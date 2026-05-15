"""
P6 XER Project Manager Planning Tool
=====================================
A Streamlit app for interrogating Primavera P6 XER schedules
without needing to open P6. Designed for Project Managers.
"""

import io
import re
import math
import warnings
from collections import defaultdict, deque
from datetime import datetime, timedelta

import networkx as nx
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# -----------------------------------------------------------------------------
# PAGE CONFIG
# -----------------------------------------------------------------------------
st.set_page_config(
    page_title="PlanTrace",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded",
)

# -----------------------------------------------------------------------------
# CUSTOM CSS  --  PlanTrace v3 Professional Control Centre
# Brand: Navy=#071827  Amber=#F5A623  Background=#F3F5F7
# -----------------------------------------------------------------------------
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');

/* ================================================================
   PLANTRACE v4 — ENTERPRISE CONTROL CENTRE
   Primary: #071827 | Panel: #0C2132 | Amber: #F5A623 | BG: #E8EAED
   ================================================================ */

/* ---------- ELIMINATE ALL STREAMLIT CHROME ---------- */
#MainMenu { visibility: hidden !important; }
footer { visibility: hidden !important; }
header { visibility: hidden !important; height: 0 !important; }
[data-testid="stToolbar"] { display: none !important; }
[data-testid="stDecoration"] { display: none !important; }
[data-testid="stStatusWidget"] { display: none !important; }
[data-testid="stHeader"] { display: none !important; height: 0 !important; }
[data-testid="collapsedControl"] { display: none !important; }
.stDeployButton { display: none !important; }
section[data-testid="stSidebarNav"] { display: none !important; }
[data-testid="stSidebarNavItems"] { display: none !important; }
button[kind="header"] { display: none !important; }
[data-testid="stSidebarNavSeparator"] { display: none !important; }

/* ---------- ROOT ---------- */
*, *::before, *::after { box-sizing: border-box; }
html, body { margin: 0; padding: 0; }
.stApp {
    background: #E8EAED !important;
    font-family: 'Inter', 'Segoe UI', -apple-system, sans-serif !important;
}
[data-testid="stAppViewContainer"] { background: #E8EAED !important; }
[data-testid="block-container"] {
    padding: 0 !important;
    max-width: 100% !important;
}

/* ---------- SIDEBAR ---------- */
[data-testid="stSidebar"] {
    background: #071827 !important;
    border-right: 1px solid #0c2030 !important;
    min-width: 192px !important;
    max-width: 192px !important;
    padding: 0 !important;
    overflow-x: hidden !important;
}
/* Kill ALL internal sidebar spacing */
[data-testid="stSidebar"] > div { padding: 0 !important; }
[data-testid="stSidebar"] section { padding: 0 !important; }
[data-testid="stSidebar"] .block-container { padding: 0 !important; }
[data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
    gap: 0 !important;
    padding: 0 !important;
}
[data-testid="stSidebar"] .element-container {
    margin: 0 !important;
    padding: 0 !important;
}
[data-testid="stSidebar"] .stMarkdown { margin: 0 !important; padding: 0 !important; }
[data-testid="stSidebar"] p { margin: 0 !important; padding: 0 !important; }
[data-testid="stSidebar"] * {
    font-family: 'Inter', 'Segoe UI', -apple-system, sans-serif !important;
    color: #2d4455 !important;
    box-sizing: border-box !important;
}

/* Sidebar file uploader */
[data-testid="stSidebar"] [data-testid="stFileUploader"] { padding: 0 10px !important; }
[data-testid="stSidebar"] [data-testid="stFileUploader"] label { display: none !important; }
[data-testid="stSidebar"] [data-testid="stFileUploadDropzone"] {
    background: #0C2132 !important;
    border: 1px dashed #173248 !important;
    border-radius: 3px !important;
    padding: 6px !important;
    min-height: 0 !important;
}
[data-testid="stSidebar"] [data-testid="stFileUploadDropzone"] span,
[data-testid="stSidebar"] [data-testid="stFileUploadDropzone"] small {
    font-size: 10px !important;
    color: #1d3045 !important;
}
[data-testid="stSidebar"] [data-testid="stFileUploadDropzone"] button {
    font-size: 10px !important;
    padding: 3px 8px !important;
    background: #0f2840 !important;
    color: #4a6880 !important;
    border: 1px solid #173248 !important;
    border-radius: 3px !important;
    box-shadow: none !important;
}

/* Sidebar slider */
[data-testid="stSidebar"] [data-testid="stSlider"] { padding: 2px 10px 6px 10px !important; }
[data-testid="stSidebar"] [data-testid="stSlider"] label {
    font-size: 9px !important;
    color: #1d3045 !important;
    text-transform: uppercase !important;
    letter-spacing: 1px !important;
    font-weight: 700 !important;
}
[data-testid="stSidebar"] [data-testid="stSliderThumb"] { background: #F5A623 !important; }
[data-testid="stSidebar"] [data-testid="stSliderTrackFill"] { background: #F5A623 !important; }
[data-testid="stSidebar"] [data-testid="stSlider"] [role="slider"] {
    background: #F5A623 !important;
    border-color: #F5A623 !important;
}

/* Sidebar NAV BUTTONS - tight, professional */
[data-testid="stSidebar"] .stButton { margin: 0 !important; padding: 0 !important; }
[data-testid="stSidebar"] .stButton > button {
    background: transparent !important;
    color: #2a3f52 !important;
    border: none !important;
    border-radius: 0 !important;
    border-left: 2px solid transparent !important;
    text-align: left !important;
    font-size: 11.5px !important;
    font-weight: 500 !important;
    padding: 6px 14px 6px 11px !important;
    width: 100% !important;
    height: auto !important;
    min-height: 0 !important;
    box-shadow: none !important;
    letter-spacing: 0.1px !important;
    line-height: 1.2 !important;
    transition: background 0.1s, color 0.1s !important;
    white-space: nowrap !important;
    overflow: hidden !important;
    text-overflow: ellipsis !important;
}
[data-testid="stSidebar"] .stButton > button:hover {
    background: #0C2132 !important;
    color: #6a8fa8 !important;
    border-left-color: transparent !important;
    transform: none !important;
    box-shadow: none !important;
}
[data-testid="stSidebar"] .stButton > button:focus {
    box-shadow: none !important;
    outline: none !important;
}
[data-testid="stSidebar"] .stButton > button:active {
    box-shadow: none !important;
    transform: none !important;
}

/* ================================================================
   MAIN CONTENT
   ================================================================ */

/* Page header bar - full bleed */
.pt-page-header {
    background: #071827;
    padding: 14px 24px 12px 24px;
    border-bottom: 2px solid #F5A623;
    margin: 0;
}

/* Content wrapper */
.pt-content {
    padding: 20px 24px;
}

/* Tabs */
.stTabs [data-baseweb="tab-list"] {
    background: #FFFFFF;
    gap: 0;
    padding: 0 20px;
    border-bottom: 1px solid #d5d9e0;
    box-shadow: none;
}
.stTabs [data-baseweb="tab"] {
    font-size: 11px;
    font-weight: 600;
    color: #6B7280;
    padding: 8px 14px;
    border-bottom: 2px solid transparent;
    margin-bottom: -1px;
    background: transparent;
    border-radius: 0;
    letter-spacing: 0.5px;
    text-transform: uppercase;
}
.stTabs [aria-selected="true"] {
    color: #071827 !important;
    border-bottom: 2px solid #F5A623 !important;
    background: transparent !important;
}
.stTabs [data-baseweb="tab"]:hover {
    color: #071827 !important;
    background: #f8fafc !important;
}
.stTabs [data-baseweb="tab-panel"] { padding: 20px 0 0 0; }

/* Metric containers */
div[data-testid="metric-container"] {
    background: #FFFFFF;
    border-radius: 5px;
    padding: 12px 14px;
    border: 1px solid #d5d9e0;
    box-shadow: 0 1px 2px rgba(7,24,39,0.04);
}

/* Dataframes */
.stDataFrame {
    border-radius: 5px;
    border: 1px solid #d5d9e0;
    overflow: hidden;
    box-shadow: 0 1px 3px rgba(7,24,39,0.04);
}
.stDataFrame thead tr th {
    background: #071827 !important;
    color: #c5d3de !important;
    font-size: 9px !important;
    font-weight: 700 !important;
    padding: 8px 12px !important;
    letter-spacing: 0.8px !important;
    text-transform: uppercase !important;
}
.stDataFrame tbody tr:nth-child(even) { background: #f6f8fa; }
.stDataFrame tbody tr:hover { background: #edf1f5 !important; }
.stDataFrame tbody tr td {
    font-size: 12px !important;
    padding: 6px 12px !important;
    color: #1a2a38 !important;
}

/* Main area buttons */
[data-testid="stAppViewContainer"] .stButton > button {
    background: #071827;
    color: white;
    border: none;
    border-radius: 4px;
    font-weight: 600;
    font-size: 12px;
    padding: 6px 14px;
    transition: all 0.12s;
    letter-spacing: 0.3px;
    box-shadow: none;
}
[data-testid="stAppViewContainer"] .stButton > button:hover {
    background: #0C2132;
    box-shadow: 0 3px 8px rgba(7,24,39,0.18);
    transform: translateY(-1px);
}
[data-testid="stAppViewContainer"] .stButton > button[kind="primary"] {
    background: #F5A623;
    color: #071827;
}
[data-testid="stAppViewContainer"] .stButton > button[kind="primary"]:hover {
    background: #d9911e;
}

/* Download buttons */
.stDownloadButton > button {
    background: #FFFFFF !important;
    color: #071827 !important;
    border: 1px solid #071827 !important;
    border-radius: 4px !important;
    font-weight: 600 !important;
    font-size: 11px !important;
    padding: 5px 12px !important;
    letter-spacing: 0.3px !important;
}
.stDownloadButton > button:hover {
    background: #071827 !important;
    color: white !important;
}

/* Inputs */
.stSelectbox > div > div,
.stMultiSelect > div > div {
    border-radius: 4px;
    border: 1px solid #d5d9e0;
    font-size: 12px;
    background: #FFFFFF;
}
.stTextInput > div > div > input {
    border-radius: 4px;
    border: 1px solid #d5d9e0;
    font-size: 12px;
    padding: 7px 10px;
    background: #FFFFFF;
    color: #111827;
}
.stTextInput > div > div > input:focus {
    border-color: #F5A623;
    box-shadow: 0 0 0 2px rgba(245,166,35,0.2);
    outline: none;
}
.stNumberInput > div > div > input {
    border-radius: 4px;
    border: 1px solid #d5d9e0;
    font-size: 12px;
    padding: 7px 10px;
}

/* Expanders */
.streamlit-expanderHeader {
    background: #FFFFFF;
    border: 1px solid #d5d9e0;
    border-radius: 4px;
    font-weight: 600;
    font-size: 12px;
    color: #111827;
    padding: 8px 12px;
}
.streamlit-expanderContent {
    border: 1px solid #d5d9e0;
    border-top: none;
    border-radius: 0 0 4px 4px;
    background: #FFFFFF;
}

/* Alerts */
div[data-testid="stAlert"] { border-radius: 4px; font-size: 12px; }

/* Headings */
h1 { display: none !important; }
h2 {
    font-size: 15px !important;
    font-weight: 700 !important;
    color: #071827 !important;
    margin: 14px 0 7px 0 !important;
    letter-spacing: -0.2px;
}
h3 {
    font-size: 13px !important;
    font-weight: 600 !important;
    color: #0C2132 !important;
    margin: 10px 0 5px 0 !important;
}
h4, h5 { font-size: 12px !important; font-weight: 600 !important; color: #374151 !important; }
p, li { font-size: 12px; color: #374151; line-height: 1.6; }
.stCaption, .stCaption p { font-size: 11px !important; color: #9CA3AF !important; }
hr { border: none; border-top: 1px solid #d5d9e0; margin: 12px 0; }

/* ================================================================
   CUSTOM COMPONENTS
   ================================================================ */

/* Status chips */
.chip {
    display: inline-flex;
    align-items: center;
    gap: 3px;
    padding: 2px 7px;
    border-radius: 3px;
    font-size: 9px;
    font-weight: 700;
    letter-spacing: 0.5px;
    text-transform: uppercase;
    font-family: 'Inter', 'Segoe UI', sans-serif;
}
.chip-red    { background: #fde8e8; color: #9b1c1c; }
.chip-amber  { background: #fef3c7; color: #92400e; }
.chip-green  { background: #def7ec; color: #03543f; }
.chip-blue   { background: #e1effe; color: #1e40af; }
.chip-grey   { background: #f3f4f6; color: #4b5563; }
.chip-teal   { background: #d5f5f6; color: #05505c; }

/* KPI cards */
.kpi-card {
    background: #FFFFFF;
    border-radius: 5px;
    padding: 14px 16px;
    border: 1px solid #d5d9e0;
    border-top: 3px solid #F5A623;
    box-shadow: 0 1px 3px rgba(7,24,39,0.05);
    height: 100%;
}
.kpi-card-red   { border-top-color: #DC2626; }
.kpi-card-amber { border-top-color: #F59E0B; }
.kpi-card-green { border-top-color: #16A34A; }
.kpi-label {
    font-size: 9px;
    font-weight: 700;
    color: #6B7280;
    letter-spacing: 1.2px;
    text-transform: uppercase;
    margin-bottom: 6px;
}
.kpi-value {
    font-size: 26px;
    font-weight: 800;
    color: #071827;
    line-height: 1;
    margin-bottom: 3px;
    font-variant-numeric: tabular-nums;
}
.kpi-sub { font-size: 10px; color: #6B7280; margin-top: 3px; }

/* Section label */
.section-label {
    font-size: 9px;
    font-weight: 700;
    color: #6B7280;
    letter-spacing: 1.5px;
    text-transform: uppercase;
    margin-bottom: 8px;
    padding-bottom: 5px;
    border-bottom: 1px solid #d5d9e0;
}

/* Attention panel */
.attention-item {
    background: #FFFFFF;
    border-radius: 4px;
    padding: 9px 12px;
    border-left: 3px solid #DC2626;
    margin-bottom: 4px;
    box-shadow: 0 1px 2px rgba(7,24,39,0.04);
}
.attention-item-amber { border-left-color: #F59E0B; }
.attention-item-blue  { border-left-color: #3B82F6; }

/* Empty state */
.empty-state {
    background: #FFFFFF;
    border-radius: 6px;
    padding: 36px 28px;
    text-align: center;
    border: 1px dashed #CBD5E0;
}

/* Data quality card */
.dq-row {
    display: flex;
    align-items: center;
    justify-content: space-between;
    padding: 5px 0;
    border-bottom: 1px solid #f1f5f9;
    font-size: 11px;
}
.dq-row:last-child { border-bottom: none; }

/* Module cards (landing page) */
.module-card {
    background: #FFFFFF;
    border-radius: 5px;
    padding: 16px;
    border: 1px solid #d5d9e0;
    border-top: 3px solid #F5A623;
    box-shadow: 0 1px 3px rgba(7,24,39,0.05);
    height: 100%;
}
.module-label {
    font-size: 8px;
    font-weight: 800;
    color: #F5A623;
    letter-spacing: 2px;
    text-transform: uppercase;
    margin-bottom: 4px;
}
.module-title {
    font-size: 13px;
    font-weight: 700;
    color: #071827;
    margin-bottom: 4px;
    line-height: 1.2;
}
.module-desc {
    font-size: 11px;
    color: #6B7280;
    line-height: 1.5;
    margin-bottom: 8px;
}
.module-outputs {
    font-size: 10px;
    color: #374151;
    padding-left: 0;
    list-style: none;
    margin: 0;
}
.module-outputs li {
    padding: 2px 0;
    display: flex;
    align-items: center;
    gap: 4px;
}
.module-outputs li::before {
    content: "";
    display: inline-block;
    width: 3px;
    height: 3px;
    border-radius: 50%;
    background: #F5A623;
    flex-shrink: 0;
}
</style>
""", unsafe_allow_html=True)

# -----------------------------------------------------------------------------
# XER PARSING  (xerparser + manual fallback)
# -----------------------------------------------------------------------------

def parse_xer_fallback(raw_text: str) -> dict:
    """
    Manual fallback parser that reads XER table format:
    %T TABLE_NAME  /  %F col1 col2 ...  /  %R val1 val2 ...
    Returns dict of {table_name: list_of_dicts}
    """
    tables = {}
    current_table = None
    current_fields = []

    for line in raw_text.splitlines():
        line = line.rstrip("\r")
        if line.startswith("%T\t"):
            current_table = line[3:].strip()
            current_fields = []
            tables[current_table] = []
        elif line.startswith("%F\t") and current_table:
            current_fields = line[3:].split("\t")
        elif line.startswith("%R\t") and current_table and current_fields:
            values = line[3:].split("\t")
            # Pad values if shorter than fields
            while len(values) < len(current_fields):
                values.append("")
            row = {current_fields[i]: values[i] for i in range(len(current_fields))}
            tables[current_table].append(row)

    return tables


def hours_to_days(hours, hours_per_day=8.0):
    """Convert hours to working days."""
    if hours is None:
        return None
    try:
        return round(float(hours) / hours_per_day, 1)
    except (TypeError, ValueError):
        return None


def safe_float(val, default=None):
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def safe_date(val):
    if val is None or str(val).strip() in ("", "None"):
        return None
    if isinstance(val, datetime):
        return val
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(val).strip(), fmt)
        except ValueError:
            pass
    return None


def parse_xer(file_bytes: bytes):
    """
    Parse an XER file. Uses xerparser library first; falls back to manual parsing.
    Returns a dict with keys: tasks_df, relationships_df, wbs_df, resources_df,
    task_resources_df, project_info, calendars_df, parse_method
    """
    # Try to decode the file
    for codec in ("cp1252", "utf-8", "latin-1"):
        try:
            raw_text = file_bytes.decode(codec)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Cannot decode XER file. Please check the file encoding.")

    result = {
        "tasks_df": pd.DataFrame(),
        "relationships_df": pd.DataFrame(),
        "wbs_df": pd.DataFrame(),
        "resources_df": pd.DataFrame(),
        "task_resources_df": pd.DataFrame(),
        "project_info": {},
        "calendars_df": pd.DataFrame(),
        "parse_method": "unknown",
    }

    # -- Try xerparser library -------------------------------------------------
    try:
        from xerparser.src.xer import Xer
        xer = Xer(raw_text)

        # Project info
        proj = None
        if xer.projects:
            proj_id = next(iter(xer.projects))
            proj = xer.projects[proj_id]
            result["project_info"] = {
                "name": getattr(proj, "name", ""),
                "data_date": getattr(proj, "last_recalc_date", None),
                "project_id": proj_id,
                "plan_start": getattr(proj, "plan_start_date", None),
                "scd_end": getattr(proj, "scd_end_date", None),
            }

        # Tasks DataFrame
        rows = []
        for uid, task in xer.tasks.items():
            tf = task.total_float_hr_cnt
            ff = task.free_float_hr_cnt
            # Effective start/finish (actual if done, early if not)
            eff_start = task.act_start_date or task.early_start_date or task.target_start_date
            eff_finish = task.act_end_date or task.early_end_date or task.target_end_date

            # WBS path
            wbs_node = xer.wbs_nodes.get(task.wbs_id)
            wbs_path = ""
            if wbs_node:
                parts = []
                n = wbs_node
                while n:
                    parts.append(getattr(n, "name", ""))
                    n = getattr(n, "parent", None)
                wbs_path = " > ".join(reversed(parts))

            # Calendar name
            cal = xer.calendars.get(task.clndr_id)
            cal_name = getattr(cal, "name", "") if cal else ""

            rows.append({
                "task_id": uid,
                "task_code": task.task_code,
                "task_name": task.name,
                "wbs_id": task.wbs_id,
                "wbs_path": wbs_path,
                "status": task.status.value if task.status else "",
                "task_type": task.type.value if task.type else "",
                "calendar": cal_name,
                "early_start": task.early_start_date,
                "early_finish": task.early_end_date,
                "late_start": task.late_start_date,
                "late_finish": task.late_end_date,
                "act_start": task.act_start_date,
                "act_finish": task.act_end_date,
                "target_start": task.target_start_date,
                "target_finish": task.target_end_date,
                "eff_start": eff_start,
                "eff_finish": eff_finish,
                "orig_dur_days": hours_to_days(task.target_drtn_hr_cnt),
                "rem_dur_days": hours_to_days(task.remain_drtn_hr_cnt),
                "total_float_days": hours_to_days(tf),
                "free_float_days": hours_to_days(ff),
                "total_float_hrs": tf,
                "is_longest_path": task.is_longest_path,
                "cstr_type": task.cstr_type,
                "cstr_date": task.cstr_date,
                "cstr_type2": task.cstr_type2,
                "cstr_date2": task.cstr_date2,
                "phys_pct": round(task.phys_complete_pct * 100, 1),
                "float_path": task.float_path,
            })

        result["tasks_df"] = pd.DataFrame(rows)

        # Relationships DataFrame
        rel_rows = []
        for uid, rel in xer.relationships.items():
            rel_rows.append({
                "pred_id": uid,
                "pred_task_id": rel.predecessor.uid if rel.predecessor else "",
                "pred_task_code": rel.predecessor.task_code if rel.predecessor else "",
                "pred_task_name": rel.predecessor.name if rel.predecessor else "",
                "succ_task_id": rel.successor.uid if rel.successor else "",
                "succ_task_code": rel.successor.task_code if rel.successor else "",
                "succ_task_name": rel.successor.name if rel.successor else "",
                "rel_type": rel.link,
                "lag_days": rel.lag,
                "lag_hrs": rel.lag_hr_cnt,
            })
        result["relationships_df"] = pd.DataFrame(rel_rows)

        # WBS DataFrame
        wbs_rows = []
        for uid, wbs in xer.wbs_nodes.items():
            wbs_rows.append({
                "wbs_id": uid,
                "wbs_code": getattr(wbs, "short_name", ""),
                "wbs_name": getattr(wbs, "name", ""),
                "parent_wbs_id": getattr(wbs, "parent_wbs_id", ""),
                "proj_id": getattr(wbs, "proj_id", ""),
            })
        result["wbs_df"] = pd.DataFrame(wbs_rows)

        # Resources & task resources
        if xer.resources:
            res_rows = []
            for uid, r in xer.resources.items():
                res_rows.append({
                    "rsrc_id": uid,
                    "rsrc_name": getattr(r, "name", ""),
                    "rsrc_short": getattr(r, "rsrc_short_name", ""),
                    "rsrc_type": getattr(r, "rsrc_type", ""),
                })
            result["resources_df"] = pd.DataFrame(res_rows)

        # Task resources (loading)
        taskrsrc_rows = []
        for uid, task in xer.tasks.items():
            for tr in getattr(task, "resources", []):
                taskrsrc_rows.append({
                    "task_id": uid,
                    "task_code": task.task_code,
                    "rsrc_id": getattr(tr, "rsrc_id", ""),
                    "target_qty": safe_float(getattr(tr, "target_qty", 0), 0),
                    "remain_qty": safe_float(getattr(tr, "remain_qty", 0), 0),
                    "act_reg_qty": safe_float(getattr(tr, "act_reg_qty", 0), 0),
                    "target_start": safe_date(getattr(tr, "target_start_date", None)),
                    "target_finish": safe_date(getattr(tr, "target_end_date", None)),
                })
        result["task_resources_df"] = pd.DataFrame(taskrsrc_rows)

        result["parse_method"] = "xerparser"
        return result

    except Exception as e:
        st.warning(f"xerparser failed ({e}), using fallback parser...")

    # -- Manual fallback -------------------------------------------------------
    try:
        tables = parse_xer_fallback(raw_text)
        return _build_from_raw_tables(tables)
    except Exception as e2:
        raise ValueError(f"Both parsers failed. Last error: {e2}")


def _build_from_raw_tables(tables: dict) -> dict:
    """Build result dict from raw parsed tables (fallback)."""
    result = {
        "tasks_df": pd.DataFrame(),
        "relationships_df": pd.DataFrame(),
        "wbs_df": pd.DataFrame(),
        "resources_df": pd.DataFrame(),
        "task_resources_df": pd.DataFrame(),
        "project_info": {},
        "calendars_df": pd.DataFrame(),
        "parse_method": "manual_fallback",
    }

    # Project info
    if "PROJECT" in tables and tables["PROJECT"]:
        proj = tables["PROJECT"][0]
        result["project_info"] = {
            "name": proj.get("proj_short_name", proj.get("proj_id", "")),
            "data_date": safe_date(proj.get("last_recalc_date")),
            "plan_start": safe_date(proj.get("plan_start_date")),
            "scd_end": safe_date(proj.get("scd_end_date")),
        }

    # Tasks
    if "TASK" in tables:
        df = pd.DataFrame(tables["TASK"])
        # Normalise date columns
        for col in ["early_start_date", "early_end_date", "late_start_date",
                    "late_end_date", "act_start_date", "act_end_date",
                    "target_start_date", "target_end_date", "cstr_date", "cstr_date2"]:
            if col in df.columns:
                df[col] = df[col].apply(safe_date)
        # Float
        for col in ["total_float_hr_cnt", "free_float_hr_cnt",
                    "target_drtn_hr_cnt", "remain_drtn_hr_cnt"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        # Build normalised columns
        df["eff_start"] = df.get("act_start_date", df.get("early_start_date"))
        df["eff_finish"] = df.get("act_end_date", df.get("early_end_date"))
        df["total_float_days"] = df.get("total_float_hr_cnt", pd.Series(dtype=float)).apply(hours_to_days)
        df["free_float_days"] = df.get("free_float_hr_cnt", pd.Series(dtype=float)).apply(hours_to_days)
        df["orig_dur_days"] = df.get("target_drtn_hr_cnt", pd.Series(dtype=float)).apply(hours_to_days)
        df["rem_dur_days"] = df.get("remain_drtn_hr_cnt", pd.Series(dtype=float)).apply(hours_to_days)

        # Rename for consistency
        rename = {
            "task_id": "task_id", "task_code": "task_code",
            "task_name": "task_name", "wbs_id": "wbs_id",
            "status_code": "status", "task_type": "task_type",
            "early_start_date": "early_start", "early_end_date": "early_finish",
            "late_start_date": "late_start", "late_end_date": "late_finish",
            "act_start_date": "act_start", "act_end_date": "act_finish",
            "target_start_date": "target_start", "target_end_date": "target_finish",
            "cstr_type": "cstr_type", "cstr_date": "cstr_date",
            "cstr_type2": "cstr_type2", "cstr_date2": "cstr_date2",
            "driving_path_flag": "is_longest_path_flag",
            "phys_complete_pct": "phys_pct",
        }
        df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
        if "is_longest_path_flag" in df.columns:
            df["is_longest_path"] = df["is_longest_path_flag"] == "Y"
        df["wbs_path"] = df.get("wbs_id", "")
        result["tasks_df"] = df

    # Relationships
    if "TASKPRED" in tables:
        df = pd.DataFrame(tables["TASKPRED"])
        if "lag_hr_cnt" in df.columns:
            df["lag_days"] = pd.to_numeric(df["lag_hr_cnt"], errors="coerce").apply(hours_to_days)
        rename_r = {"pred_type": "rel_type", "task_id": "succ_task_id",
                    "pred_task_id": "pred_task_id"}
        df = df.rename(columns={k: v for k, v in rename_r.items() if k in df.columns})
        result["relationships_df"] = df

    # WBS
    if "PROJWBS" in tables:
        df = pd.DataFrame(tables["PROJWBS"])
        rename_w = {"wbs_id": "wbs_id", "wbs_short_name": "wbs_code",
                    "wbs_name": "wbs_name", "parent_wbs_id": "parent_wbs_id"}
        df = df.rename(columns={k: v for k, v in rename_w.items() if k in df.columns})
        result["wbs_df"] = df

    # Resources
    if "RSRC" in tables:
        df = pd.DataFrame(tables["RSRC"])
        rename_rs = {"rsrc_id": "rsrc_id", "rsrc_name": "rsrc_name",
                     "rsrc_short_name": "rsrc_short"}
        df = df.rename(columns={k: v for k, v in rename_rs.items() if k in df.columns})
        result["resources_df"] = df

    # Task resources
    if "TASKRSRC" in tables:
        df = pd.DataFrame(tables["TASKRSRC"])
        for col in ["target_qty", "remain_qty", "act_reg_qty"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        for col in ["target_start_date", "target_end_date"]:
            if col in df.columns:
                df[col] = df[col].apply(safe_date)
                df = df.rename(columns={col: col.replace("_date", "")})
        result["task_resources_df"] = df

    return result


# -----------------------------------------------------------------------------
# GRAPH BUILDING
# -----------------------------------------------------------------------------

def build_graph(tasks_df: pd.DataFrame, rels_df: pd.DataFrame) -> nx.DiGraph:
    """Build a networkx directed graph from tasks and relationships."""
    G = nx.DiGraph()
    for _, row in tasks_df.iterrows():
        G.add_node(row["task_id"], **row.to_dict())
    for _, row in rels_df.iterrows():
        if row.get("pred_task_id") and row.get("succ_task_id"):
            G.add_edge(
                row["pred_task_id"],
                row["succ_task_id"],
                rel_type=row.get("rel_type", "FS"),
                lag_days=row.get("lag_days", 0),
            )
    return G


# -----------------------------------------------------------------------------
# CRITICAL PATH HELPERS
# -----------------------------------------------------------------------------

def get_critical_threshold(tasks_df: pd.DataFrame, near_crit_days: float = 10.0):
    """Classify activities as critical / near-critical / float."""
    df = tasks_df.copy()
    df["is_critical"] = df["total_float_days"].apply(
        lambda f: f is not None and f <= 0
    )
    df["is_near_critical"] = df["total_float_days"].apply(
        lambda f: f is not None and 0 < f <= near_crit_days
    )
    return df


def float_status_badge(f):
    if f is None:
        return "-"
    elif f <= 0:
        return "🔴 Critical"
    elif f <= 10:
        return "🟡 Near-Critical"
    else:
        return "🟢 Float"


# -----------------------------------------------------------------------------
# LOGIC TRACE HELPERS
# -----------------------------------------------------------------------------

def trace_predecessors(G: nx.DiGraph, task_id: str, max_depth=100) -> list:
    """BFS backwards through predecessors. Returns list of (task_id, depth)."""
    visited = {}
    queue = deque([(task_id, 0)])
    result = []
    while queue:
        node, depth = queue.popleft()
        if node in visited or depth > max_depth:
            continue
        visited[node] = depth
        if node != task_id:
            result.append((node, depth))
        for pred in G.predecessors(node):
            if pred not in visited:
                queue.append((pred, depth + 1))
    return result


def trace_successors(G: nx.DiGraph, task_id: str, max_depth=100) -> list:
    """BFS forwards through successors."""
    visited = {}
    queue = deque([(task_id, 0)])
    result = []
    while queue:
        node, depth = queue.popleft()
        if node in visited or depth > max_depth:
            continue
        visited[node] = depth
        if node != task_id:
            result.append((node, depth))
        for succ in G.successors(node):
            if succ not in visited:
                queue.append((succ, depth + 1))
    return result


def driving_path_to_activity(
    G: nx.DiGraph,
    tasks_df: pd.DataFrame,
    rels_df: pd.DataFrame,
    target_id: str,
) -> list:
    """
    Identify the most likely driving predecessor chain into a target activity.

    Driving predecessor selection priority (in order):
      1. Lowest total float  (most constrained activity wins)
      2. On P6 longest-path / driving flag where available
      3. Latest early-finish date  (latest predecessor is usually the driver)
      4. Highest lag on the connecting relationship (more constraining)

    Returns ordered list of task_ids, from chain start -> target.
    """
    task_lookup = tasks_df.set_index("task_id").to_dict("index") if not tasks_df.empty else {}

    def _score(pred_id, succ_id):
        t  = task_lookup.get(pred_id, {})
        tf = safe_float(t.get("total_float_days"), 9999)
        finish = t.get("eff_finish")
        if finish is not None:
            try:
                finish_score = -(finish.timestamp() / 86400)
            except Exception:
                finish_score = 0
        else:
            finish_score = 0
        driving_bonus = 0 if t.get("is_longest_path", False) else 1
        lag_score = 0
        if not rels_df.empty:
            rel = rels_df[
                (rels_df.get("pred_task_id", pd.Series(dtype=str)) == pred_id) &
                (rels_df.get("succ_task_id", pd.Series(dtype=str)) == succ_id)
            ]
            if not rel.empty and "lag_days" in rel.columns:
                lag_val = safe_float(rel["lag_days"].iloc[0], 0)
                lag_score = -lag_val
        return (tf, driving_bonus, finish_score, lag_score)

    path    = [target_id]
    visited = {target_id}
    current = target_id

    for _ in range(500):
        preds = list(G.predecessors(current))
        if not preds:
            break
        unvisited = [p for p in preds if p not in visited]
        if not unvisited:
            break
        best = min(unvisited, key=lambda p: _score(p, current))
        path.insert(0, best)
        visited.add(best)
        current = best

    return path


def _all_pred_paths(
    G: nx.DiGraph,
    tasks_df: pd.DataFrame,
    target_id: str,
    max_paths: int = 8,
) -> list:
    """
    Find up to max_paths predecessor chains into target_id.
    Each chain is a list of task_ids ordered start -> target.
    Only returns chains that begin at an activity with no predecessors.
    """
    task_lookup = tasks_df.set_index("task_id").to_dict("index") if not tasks_df.empty else {}

    def _float(tid):
        return safe_float(task_lookup.get(tid, {}).get("total_float_days"), 9999)

    found_paths = []

    def dfs(node, current_path, visited_set):
        if len(found_paths) >= max_paths:
            return
        preds = [p for p in G.predecessors(node) if p not in visited_set]
        if not preds:
            found_paths.append(list(reversed(current_path)))
            return
        for pred in sorted(preds, key=_float)[:4]:
            dfs(pred, current_path + [pred], visited_set | {pred})

    dfs(target_id, [target_id], {target_id})
    return found_paths


# -----------------------------------------------------------------------------
# EXPORT HELPERS
# -----------------------------------------------------------------------------

def style_header_row(ws, row_idx, fill_color="1e3a5f", font_color="FFFFFF"):
    fill = PatternFill("solid", start_color=fill_color, fgColor=fill_color)
    font = Font(bold=True, color=font_color)
    for cell in ws[row_idx]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def df_to_sheet(ws, df, sheet_title=None):
    """Write a DataFrame to an openpyxl worksheet with formatting."""
    if sheet_title:
        ws.title = sheet_title[:31]
    ws.append(list(df.columns))
    style_header_row(ws, 1)
    for r in df.itertuples(index=False):
        ws.append(list(r))
    # Auto-width
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 50)


def export_df_to_excel(sheets: dict) -> bytes:
    """sheets = {sheet_name: dataframe}. Returns Excel bytes."""
    wb = Workbook()
    first = True
    for name, df in sheets.items():
        if first:
            ws = wb.active
            first = False
        else:
            ws = wb.create_sheet()
        df_to_sheet(ws, df, name)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def format_date(d):
    if d is None:
        return "-"
    try:
        return d.strftime("%d %b %Y")
    except Exception:
        return str(d)


# -----------------------------------------------------------------------------
# PAGE: PROJECT SUMMARY
# -----------------------------------------------------------------------------

def page_project_summary(data: dict, near_crit_days: float):
    st.title("📊 Project Summary")

    proj = data["project_info"]
    tasks = data["tasks_df"]
    rels = data["relationships_df"]

    if tasks.empty:
        st.warning("No activities found in this file.")
        return

    tasks = get_critical_threshold(tasks, near_crit_days)

    # Header metrics
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Project Name", proj.get("name", "Unknown"))
    c2.metric("Data Date", format_date(proj.get("data_date")))
    c3.metric("Parse Method", data.get("parse_method", "-"))
    c4.metric("Activities", len(tasks))

    st.divider()
    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("🔴 Critical", int(tasks["is_critical"].sum()))
    c2.metric("🟡 Near-Critical", int(tasks["is_near_critical"].sum()))
    neg_float = tasks["total_float_days"].apply(lambda f: f is not None and f < 0).sum()
    c3.metric("⚠️ Negative Float", int(neg_float))
    c4.metric("🔗 Relationships", len(rels))

    # Open-ended
    if not rels.empty and "pred_task_id" in rels.columns:
        tasks_with_pred = set(rels["succ_task_id"].dropna())
        tasks_with_succ = set(rels["pred_task_id"].dropna())
        task_ids = set(tasks["task_id"])
        no_pred = len(task_ids - tasks_with_pred)
        no_succ = len(task_ids - tasks_with_succ)
        c5.metric("Open-Ended Activities", no_pred + no_succ)
    else:
        c5.metric("Open-Ended Activities", "-")

    # Date range
    valid_starts = tasks["eff_start"].dropna()
    valid_finishes = tasks["eff_finish"].dropna()
    if not valid_starts.empty and not valid_finishes.empty:
        earliest = min(valid_starts)
        latest = max(valid_finishes)
        st.info(f"**Schedule Span:** {format_date(earliest)} -> {format_date(latest)}")

    # Constraint count
    constrained = tasks["cstr_type"].apply(lambda x: bool(x) and str(x).strip() not in ("", "None")).sum() if "cstr_type" in tasks.columns else 0
    st.info(f"**Constrained Activities:** {int(constrained)}")

    st.divider()

    # Charts
    tab1, tab2, tab3 = st.tabs(["Float Distribution", "Activities by WBS", "Status Breakdown"])

    with tab1:
        float_vals = tasks["total_float_days"].dropna()
        if not float_vals.empty:
            fig = px.histogram(
                float_vals, nbins=40, title="Total Float Distribution (Days)",
                labels={"value": "Float (days)", "count": "Activities"},
                color_discrete_sequence=["#2563eb"],
            )
            fig.add_vline(x=0, line_dash="dash", line_color="red", annotation_text="Critical")
            fig.add_vline(x=near_crit_days, line_dash="dot", line_color="orange",
                          annotation_text=f"Near-Critical ({near_crit_days}d)")
            st.plotly_chart(fig, use_container_width=True)

    with tab2:
        if "wbs_path" in tasks.columns:
            # Show top-level WBS only
            tasks["wbs_top"] = tasks["wbs_path"].apply(
                lambda x: str(x).split(" > ")[0] if pd.notna(x) and x else "Unknown"
            )
            wbs_counts = tasks.groupby("wbs_top").size().reset_index(name="count")
            wbs_counts = wbs_counts.sort_values("count", ascending=False).head(20)
            fig = px.bar(wbs_counts, x="count", y="wbs_top", orientation="h",
                         title="Activities by Top-Level WBS",
                         color_discrete_sequence=["#1e3a5f"])
            fig.update_layout(yaxis_title="", xaxis_title="Activity Count")
            st.plotly_chart(fig, use_container_width=True)

    with tab3:
        if "status" in tasks.columns:
            status_counts = tasks["status"].value_counts().reset_index()
            status_counts.columns = ["Status", "Count"]
            fig = px.pie(status_counts, values="Count", names="Status",
                         title="Activity Status Breakdown",
                         color_discrete_sequence=px.colors.qualitative.Set2)
            st.plotly_chart(fig, use_container_width=True)

    # Summary table
    st.subheader("Activity Summary Table")
    display_cols = ["task_code", "task_name", "wbs_path", "eff_start", "eff_finish",
                    "total_float_days", "status", "is_critical"]
    avail = [c for c in display_cols if c in tasks.columns]
    st.dataframe(tasks[avail].head(100), use_container_width=True)


# -----------------------------------------------------------------------------
# PAGE: ACTIVITY SEARCH
# -----------------------------------------------------------------------------

def _col(df: pd.DataFrame, col: str, default="-"):
    """Safely get a column value; return default if column missing or null."""
    if col not in df.index:
        return default
    val = df.get(col, default)
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return default
    return val


def _float_color(f) -> str:
    """Return a hex colour string for a float value."""
    if f is None:
        return "#6b7280"
    try:
        f = float(f)
    except (TypeError, ValueError):
        return "#6b7280"
    if f < 0:
        return "#991b1b"   # dark red  - negative float
    if f == 0:
        return "#dc2626"   # red       - critical
    if f <= 10:
        return "#d97706"   # amber     - near-critical
    return "#15803d"       # green     - has float


def _status_label(status: str) -> str:
    """Convert P6 status code to a readable label."""
    mapping = {
        "TK_NotStart": "Not Started",
        "TK_Active":   "In Progress",
        "TK_Complete": "Complete",
        "Not Started": "Not Started",
        "In Progress": "In Progress",
        "Complete":    "Complete",
    }
    return mapping.get(str(status).strip(), str(status).strip() or "-")


def _status_colour(status: str) -> str:
    s = _status_label(status)
    if s == "Complete":    return "#15803d"
    if s == "In Progress": return "#2563eb"
    return "#6b7280"


def page_activity_search(data: dict, near_crit_days: float):
    """
    Activity Search page.
    Lets the user filter activities by ID, name, WBS, date range and
    critical status, then shows a full detail panel for the selected activity.
    """
    st.title("🔍 Activity Search")
    st.caption("Search and filter activities, then click any row to view its full detail.")

    tasks = data["tasks_df"]
    rels  = data["relationships_df"]

    if tasks.empty:
        st.warning("No activities found. Please upload an XER file first.")
        return

    tasks = get_critical_threshold(tasks, near_crit_days)

    # -------------------------------------------------------------------------
    # SEARCH / FILTER PANEL
    # -------------------------------------------------------------------------
    with st.expander("🔎  Search & Filter", expanded=True):
        r1c1, r1c2, r1c3 = st.columns(3)
        search_code = r1c1.text_input("Activity ID", placeholder="e.g. A1000")
        search_name = r1c2.text_input("Activity Name", placeholder="partial match")
        search_wbs  = r1c3.text_input("WBS", placeholder="partial match")

        r2c1, r2c2, r2c3 = st.columns(3)

        # Critical / float filter
        crit_filter = r2c1.selectbox(
            "Float Status",
            ["All", "Critical (float <= 0)", "Near-Critical", "Positive Float", "Negative Float"],
        )

        # Status filter - only show if column exists
        if "status" in tasks.columns:
            status_opts = ["All"] + sorted(
                [s for s in tasks["status"].dropna().unique() if str(s).strip()]
            )
        else:
            status_opts = ["All"]
        status_filter = r2c2.selectbox("Activity Status", status_opts)

        # Activity type filter
        if "task_type" in tasks.columns:
            type_opts = ["All"] + sorted(
                [t for t in tasks["task_type"].dropna().unique() if str(t).strip()]
            )
        else:
            type_opts = ["All"]
        type_filter = r2c3.selectbox("Activity Type", type_opts)

        # Date range - only show if date columns exist and have data
        date_from = date_to = None
        valid_starts  = tasks["eff_start"].dropna()  if "eff_start"  in tasks.columns else pd.Series(dtype=object)
        valid_finishes = tasks["eff_finish"].dropna() if "eff_finish" in tasks.columns else pd.Series(dtype=object)

        if not valid_starts.empty and not valid_finishes.empty:
            try:
                min_d = min(valid_starts).date()
                max_d = max(valid_finishes).date()
                dc1, dc2 = st.columns(2)
                date_from = dc1.date_input("Finish on or After", value=min_d,
                                           min_value=min_d, max_value=max_d)
                date_to   = dc2.date_input("Finish on or Before", value=max_d,
                                           min_value=min_d, max_value=max_d)
            except Exception:
                pass  # silently skip date filter if dates are unusable

    # -------------------------------------------------------------------------
    # APPLY FILTERS
    # -------------------------------------------------------------------------
    filtered = tasks.copy()

    if search_code.strip():
        if "task_code" in filtered.columns:
            filtered = filtered[
                filtered["task_code"].astype(str).str.contains(
                    search_code.strip(), case=False, na=False
                )
            ]
    if search_name.strip():
        if "task_name" in filtered.columns:
            filtered = filtered[
                filtered["task_name"].astype(str).str.contains(
                    search_name.strip(), case=False, na=False
                )
            ]
    if search_wbs.strip():
        if "wbs_path" in filtered.columns:
            filtered = filtered[
                filtered["wbs_path"].astype(str).str.contains(
                    search_wbs.strip(), case=False, na=False
                )
            ]

    if crit_filter == "Critical (float <= 0)" and "is_critical" in filtered.columns:
        filtered = filtered[filtered["is_critical"] == True]
    elif crit_filter == "Near-Critical" and "is_near_critical" in filtered.columns:
        filtered = filtered[filtered["is_near_critical"] == True]
    elif crit_filter == "Positive Float" and "total_float_days" in filtered.columns:
        filtered = filtered[filtered["total_float_days"].apply(
            lambda f: f is not None and safe_float(f, 1) > 0
        )]
    elif crit_filter == "Negative Float" and "total_float_days" in filtered.columns:
        filtered = filtered[filtered["total_float_days"].apply(
            lambda f: f is not None and safe_float(f, 0) < 0
        )]

    if status_filter != "All" and "status" in filtered.columns:
        filtered = filtered[filtered["status"] == status_filter]

    if type_filter != "All" and "task_type" in filtered.columns:
        filtered = filtered[filtered["task_type"] == type_filter]

    if date_from is not None and "eff_finish" in filtered.columns:
        filtered = filtered[
            filtered["eff_finish"].apply(
                lambda d: d is not None and hasattr(d, "date") and d.date() >= date_from
            )
        ]
    if date_to is not None and "eff_finish" in filtered.columns:
        filtered = filtered[
            filtered["eff_finish"].apply(
                lambda d: d is not None and hasattr(d, "date") and d.date() <= date_to
            )
        ]

    # -------------------------------------------------------------------------
    # RESULTS TABLE
    # -------------------------------------------------------------------------
    n_found = len(filtered)
    n_total = len(tasks)

    if n_found == 0:
        st.warning("No activities match your filters. Try broadening your search.")
        return

    # Build a clean display version of the table
    TABLE_COLS = {
        "task_code":        "Activity ID",
        "task_name":        "Activity Name",
        "wbs_path":         "WBS",
        "eff_start":        "Start",
        "eff_finish":       "Finish",
        "orig_dur_days":    "Orig Dur (d)",
        "rem_dur_days":     "Rem Dur (d)",
        "total_float_days": "Float (d)",
        "status":           "Status",
        "task_type":        "Type",
        "is_critical":      "Critical",
    }

    present_cols = {k: v for k, v in TABLE_COLS.items() if k in filtered.columns}
    display_df = filtered[list(present_cols.keys())].copy()
    display_df = display_df.rename(columns=present_cols)

    # Format date columns for readability
    for col in ["Start", "Finish"]:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(format_date)

    # Format critical flag
    if "Critical" in display_df.columns:
        display_df["Critical"] = display_df["Critical"].apply(
            lambda x: "Yes" if x else ""
        )

    # Friendly status labels
    if "Status" in display_df.columns:
        display_df["Status"] = display_df["Status"].apply(_status_label)

    st.markdown(
        f"<p style='color:#6b7280;font-size:13px;'>"
        f"Showing <strong>{n_found}</strong> of <strong>{n_total}</strong> activities"
        f"</p>",
        unsafe_allow_html=True,
    )

    st.dataframe(
        display_df,
        use_container_width=True,
        height=min(400, 45 + n_found * 35),
        hide_index=True,
    )

    # -------------------------------------------------------------------------
    # ACTIVITY SELECTOR  -  pick from the filtered results
    # -------------------------------------------------------------------------
    st.divider()

    # Build selector labels
    def make_label(r):
        code = str(r.get("task_code", "?"))
        name = str(r.get("task_name", "?"))
        return f"{code}  --  {name}"

    act_labels = filtered.apply(make_label, axis=1).tolist()

    selected_label = st.selectbox(
        "Select an activity to view full details",
        options=act_labels,
        index=0,
    )

    sel_idx = act_labels.index(selected_label)
    row = filtered.iloc[sel_idx]

    # -------------------------------------------------------------------------
    # SELECTED ACTIVITY BANNER
    # -------------------------------------------------------------------------
    tf_val  = row.get("total_float_days") if "total_float_days" in row.index else None
    tf_num  = safe_float(tf_val, None)
    f_color = _float_color(tf_num)

    status_raw = row.get("status", "") if "status" in row.index else ""
    s_label    = _status_label(str(status_raw))
    s_color    = _status_colour(str(status_raw))

    is_crit = bool(row.get("is_critical", False)) if "is_critical" in row.index else False
    crit_banner = (
        '<span style="background:#dc2626;color:white;padding:3px 10px;'
        'border-radius:12px;font-size:12px;font-weight:700;margin-left:10px;">CRITICAL</span>'
        if is_crit else ""
    )

    task_code = str(row.get("task_code", "-")) if "task_code" in row.index else "-"
    task_name = str(row.get("task_name", "-")) if "task_name" in row.index else "-"

    st.markdown(
        f"""
        <div style="background:#1e3a5f;color:white;border-radius:10px;
                    padding:18px 24px;margin-bottom:16px;">
            <div style="font-size:13px;color:#93c5fd;font-weight:600;
                        letter-spacing:1px;text-transform:uppercase;">
                Selected Activity
            </div>
            <div style="font-size:22px;font-weight:700;margin-top:4px;">
                {task_code}{crit_banner}
            </div>
            <div style="font-size:16px;color:#bfdbfe;margin-top:4px;">
                {task_name}
            </div>
            <div style="margin-top:10px;">
                <span style="background:{s_color};color:white;padding:3px 10px;
                             border-radius:12px;font-size:12px;font-weight:600;">
                    {s_label}
                </span>
                <span style="background:{f_color};color:white;padding:3px 10px;
                             border-radius:12px;font-size:12px;font-weight:600;
                             margin-left:8px;">
                    Float: {tf_num if tf_num is not None else "-"} days
                </span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # -------------------------------------------------------------------------
    # DETAIL PANEL  -  two column layout
    # -------------------------------------------------------------------------
    def field(label: str, value, suffix: str = "") -> str:
        """Render a labelled field as HTML. Handles missing/None gracefully."""
        if value is None or str(value).strip() in ("", "None", "nan", "-"):
            disp = '<span style="color:#9ca3af;">Not available</span>'
        else:
            disp = f'<span style="font-weight:600;color:#111827;">{value}{suffix}</span>'
        return (
            f'<div style="padding:6px 0;border-bottom:1px solid #f3f4f6;">'
            f'<span style="color:#6b7280;font-size:12px;text-transform:uppercase;'
            f'letter-spacing:0.5px;">{label}</span><br>{disp}</div>'
        )

    col_left, col_right = st.columns(2)

    with col_left:
        st.markdown("##### Identity & Classification")
        st.markdown(
            field("Activity ID",    _col(row, "task_code")) +
            field("Activity Name",  _col(row, "task_name")) +
            field("WBS",            _col(row, "wbs_path")) +
            field("Activity Type",  _col(row, "task_type")) +
            field("Calendar",       _col(row, "calendar")) +
            field("% Complete",     _col(row, "phys_pct"), suffix="%"),
            unsafe_allow_html=True,
        )

    with col_right:
        st.markdown("##### Schedule & Float")
        early_start  = format_date(row.get("early_start")  if "early_start"  in row.index else None)
        early_finish = format_date(row.get("early_finish") if "early_finish" in row.index else None)
        late_start   = format_date(row.get("late_start")   if "late_start"   in row.index else None)
        late_finish  = format_date(row.get("late_finish")  if "late_finish"  in row.index else None)
        act_start    = format_date(row.get("act_start")    if "act_start"    in row.index else None)
        act_finish   = format_date(row.get("act_finish")   if "act_finish"   in row.index else None)
        orig_dur     = _col(row, "orig_dur_days")
        rem_dur      = _col(row, "rem_dur_days")
        ff_val       = _col(row, "free_float_days")

        st.markdown(
            field("Early Start",         early_start) +
            field("Early Finish",        early_finish) +
            field("Late Start",          late_start) +
            field("Late Finish",         late_finish) +
            field("Actual Start",        act_start) +
            field("Actual Finish",       act_finish) +
            field("Original Duration",   orig_dur,  suffix=" days") +
            field("Remaining Duration",  rem_dur,   suffix=" days") +
            field("Total Float",         tf_num,    suffix=" days") +
            field("Free Float",          ff_val,    suffix=" days"),
            unsafe_allow_html=True,
        )

    # Constraint row - only show when a constraint is set
    if "cstr_type" in row.index:
        cstr = row.get("cstr_type", "")
        if cstr and str(cstr).strip() not in ("", "None", "nan"):
            cstr_date = format_date(row.get("cstr_date") if "cstr_date" in row.index else None)
            st.markdown(
                f'<div style="background:#fef3c7;border-left:4px solid #f59e0b;'
                f'border-radius:6px;padding:10px 16px;margin-top:12px;">'
                f'<strong>Constraint:</strong> {cstr} &nbsp;|&nbsp; '
                f'<strong>Constraint Date:</strong> {cstr_date}</div>',
                unsafe_allow_html=True,
            )

    # -------------------------------------------------------------------------
    # PREDECESSORS & SUCCESSORS
    # -------------------------------------------------------------------------
    st.markdown("---")
    st.markdown("##### Logic")

    pred_col, succ_col = st.columns(2)

    if not rels.empty and "task_id" in row.index:
        task_id = row["task_id"]

        # --- Predecessors ---
        with pred_col:
            st.markdown("**Predecessors**  *(what drives this activity)*")
            if "succ_task_id" in rels.columns:
                preds = rels[rels["succ_task_id"] == task_id].copy()
            else:
                preds = pd.DataFrame()

            if preds.empty:
                st.info("No predecessors -- this is an open start.")
            else:
                pred_display_cols = {
                    "pred_task_code": "Activity ID",
                    "pred_task_name": "Activity Name",
                    "rel_type":       "Link",
                    "lag_days":       "Lag (d)",
                }
                pred_show = {k: v for k, v in pred_display_cols.items()
                             if k in preds.columns}
                pred_df = preds[list(pred_show.keys())].rename(columns=pred_show)
                st.dataframe(pred_df, use_container_width=True, hide_index=True)

        # --- Successors ---
        with succ_col:
            st.markdown("**Successors**  *(what this activity drives)*")
            if "pred_task_id" in rels.columns:
                succs = rels[rels["pred_task_id"] == task_id].copy()
            else:
                succs = pd.DataFrame()

            if succs.empty:
                st.info("No successors -- this is an open finish.")
            else:
                succ_display_cols = {
                    "succ_task_code": "Activity ID",
                    "succ_task_name": "Activity Name",
                    "rel_type":       "Link",
                    "lag_days":       "Lag (d)",
                }
                succ_show = {k: v for k, v in succ_display_cols.items()
                             if k in succs.columns}
                succ_df = succs[list(succ_show.keys())].rename(columns=succ_show)
                st.dataframe(succ_df, use_container_width=True, hide_index=True)
    else:
        with pred_col:
            st.info("Relationship data not available.")
        with succ_col:
            st.info("Relationship data not available.")

    # -------------------------------------------------------------------------
    # EXPORT THIS ACTIVITY
    # -------------------------------------------------------------------------
    st.markdown("---")

    # Build a single-row detail export
    export_fields = {
        "Activity ID":        _col(row, "task_code"),
        "Activity Name":      _col(row, "task_name"),
        "WBS":                _col(row, "wbs_path"),
        "Activity Type":      _col(row, "task_type"),
        "Calendar":           _col(row, "calendar"),
        "Status":             _status_label(str(_col(row, "status", ""))),
        "% Complete":         _col(row, "phys_pct"),
        "Early Start":        early_start,
        "Early Finish":       early_finish,
        "Late Start":         late_start,
        "Late Finish":        late_finish,
        "Actual Start":       act_start,
        "Actual Finish":      act_finish,
        "Original Duration":  _col(row, "orig_dur_days"),
        "Remaining Duration": _col(row, "rem_dur_days"),
        "Total Float (days)": tf_num,
        "Free Float (days)":  _col(row, "free_float_days"),
        "Critical":           "Yes" if is_crit else "No",
        "Constraint":         _col(row, "cstr_type"),
        "Constraint Date":    format_date(row.get("cstr_date") if "cstr_date" in row.index else None),
    }
    detail_df = pd.DataFrame([export_fields])

    # Predecessors / successors for export
    export_sheets = {"Activity Detail": detail_df}

    if not rels.empty and "task_id" in row.index:
        task_id = row["task_id"]
        if not preds.empty:
            pred_exp_cols = [c for c in ["pred_task_code","pred_task_name","rel_type","lag_days"]
                             if c in preds.columns]
            export_sheets["Predecessors"] = preds[pred_exp_cols].rename(columns={
                "pred_task_code": "Activity ID", "pred_task_name": "Activity Name",
                "rel_type": "Link", "lag_days": "Lag (days)"
            })
        if not succs.empty:
            succ_exp_cols = [c for c in ["succ_task_code","succ_task_name","rel_type","lag_days"]
                             if c in succs.columns]
            export_sheets["Successors"] = succs[succ_exp_cols].rename(columns={
                "succ_task_code": "Activity ID", "succ_task_name": "Activity Name",
                "rel_type": "Link", "lag_days": "Lag (days)"
            })

    xls_bytes = export_df_to_excel(export_sheets)
    st.download_button(
        label="📥  Export Activity Detail to Excel",
        data=xls_bytes,
        file_name=f"activity_{_col(row, 'task_code', 'detail')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# -----------------------------------------------------------------------------
# PAGE: LOGIC TRACE
# -----------------------------------------------------------------------------

# Relationship type codes -> readable labels
REL_TYPE_LABELS = {
    "FS": "Finish-to-Start (FS)",
    "FF": "Finish-to-Finish (FF)",
    "SS": "Start-to-Start (SS)",
    "SF": "Start-to-Finish (SF)",
    "PR_FS": "Finish-to-Start (FS)",
    "PR_FF": "Finish-to-Finish (FF)",
    "PR_SS": "Start-to-Start (SS)",
    "PR_SF": "Start-to-Finish (SF)",
}


def _rel_label(code: str) -> str:
    return REL_TYPE_LABELS.get(str(code).strip(), str(code).strip() or "-")


def _crit_flag(tf) -> str:
    """Return a text critical flag suitable for display and export."""
    if tf is None:
        return "-"
    try:
        f = float(tf)
    except (TypeError, ValueError):
        return "-"
    if f < 0:
        return "Negative Float"
    if f == 0:
        return "Critical"
    if f <= 10:
        return "Near-Critical"
    return "Float"


def _build_full_trace_df(
    G: nx.DiGraph,
    rels_df: pd.DataFrame,
    task_lookup: dict,
    selected_id: str,
    trace_list: list,     # [(task_id, depth), ...]
    direction: str,       # "pred" | "succ" | "both"
) -> pd.DataFrame:
    """
    Build the trace results DataFrame.

    For each activity in trace_list, look up the relationship that
    connects it to the selected activity (or to the activity one step
    closer in the chain) so we can display the correct link type and lag.

    direction:
        "pred"  - activities are predecessors; depth counts backwards  (1 = direct pred)
        "succ"  - activities are successors;   depth counts forwards   (1 = direct succ)
        "both"  - mixed: negative depth = pred side, positive = succ side
    """
    rows = []

    for tid, depth in trace_list:
        t = task_lookup.get(tid, {})
        tf = t.get("total_float_days")

        # ---- find the relationship between this activity and the one at
        #      depth-1 in the chain (i.e. the step that led here) ----------
        rel_type_str = "-"
        lag_val = 0

        if not rels_df.empty:
            if direction in ("pred", "both") and depth > 0:
                # this activity is a predecessor of something; find its
                # outgoing relationship toward the selected direction
                mask = (
                    (rels_df.get("pred_task_id", pd.Series(dtype=str)) == tid) |
                    (rels_df.get("succ_task_id", pd.Series(dtype=str)) == tid)
                )
                rel_candidates = rels_df[mask]
            elif direction == "succ":
                mask = (
                    (rels_df.get("pred_task_id", pd.Series(dtype=str)) == tid) |
                    (rels_df.get("succ_task_id", pd.Series(dtype=str)) == tid)
                )
                rel_candidates = rels_df[mask]
            else:
                rel_candidates = pd.DataFrame()

            if not rel_candidates.empty and "rel_type" in rel_candidates.columns:
                rel_type_str = _rel_label(rel_candidates["rel_type"].iloc[0])
                lag_val = rel_candidates["lag_days"].iloc[0] if "lag_days" in rel_candidates.columns else 0
                try:
                    lag_val = float(lag_val) if lag_val is not None else 0
                except (TypeError, ValueError):
                    lag_val = 0

        rows.append({
            "Depth": depth,
            "Direction": "Predecessor" if depth < 0 or direction == "pred"
                         else ("Successor" if direction == "succ" else "Both"),
            "Activity ID":    t.get("task_code", tid),
            "Activity Name":  t.get("task_name", ""),
            "Link Type":      rel_type_str,
            "Lag (days)":     lag_val,
            "Start":          format_date(t.get("eff_start")),
            "Finish":         format_date(t.get("eff_finish")),
            "Total Float (d)": tf if tf is not None else "-",
            "Status":         _status_label(str(t.get("status", ""))),
            "Critical Flag":  _crit_flag(tf),
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values("Depth").reset_index(drop=True)
    return df


def _summary_bar(label: str, value: int, colour: str) -> str:
    return (
        f'<div style="display:inline-block;background:{colour};color:white;'
        f'border-radius:8px;padding:8px 18px;margin:4px 6px 4px 0;font-size:13px;">'
        f'<strong>{value}</strong> {label}</div>'
    )


def page_logic_trace(data: dict, near_crit_days: float):
    """
    Logic Trace page.
    Search and select any activity, then trace its predecessor and successor
    chains through the schedule network. Results show depth, link type, lag,
    dates, float, WBS and critical status with filters and Excel export.
    """
    st.title("🔗 Logic Trace")
    st.caption(
        "Select an activity then trace its logic through the schedule network. "
        "Predecessors flow INTO the activity. Successors flow OUT of it."
    )

    tasks = data["tasks_df"]
    rels  = data["relationships_df"]

    # -------------------------------------------------------------------------
    # GUARD RAILS
    # -------------------------------------------------------------------------
    if tasks.empty:
        st.warning("No activities found in this programme.")
        return

    tasks = get_critical_threshold(tasks, near_crit_days)

    if rels.empty:
        st.warning(
            "**No relationship data found in this XER file.** "
            "Logic tracing requires both activities and relationships. "
            "Check the XER was exported with relationships included."
        )
        st.dataframe(
            tasks[["task_code", "task_name", "total_float_days", "status"]].head(50),
            use_container_width=True, hide_index=True,
        )
        return

    # -------------------------------------------------------------------------
    # BUILD NETWORK GRAPH
    # -------------------------------------------------------------------------
    G           = build_graph(tasks, rels)
    task_lookup = tasks.set_index("task_id").to_dict("index")

    # -------------------------------------------------------------------------
    # ACTIVITY SEARCH & SELECTOR
    # -------------------------------------------------------------------------
    st.markdown(
        '<div style="font-size:12px;font-weight:700;color:#94A3B8;'
        'letter-spacing:1px;text-transform:uppercase;margin-bottom:6px;">'
        'Find Activity</div>',
        unsafe_allow_html=True,
    )

    search_col, select_col = st.columns([1, 2])

    with search_col:
        search_text = st.text_input(
            "Search by ID or name",
            placeholder="e.g. A1000 or Excavation",
            key="lt_search",
            label_visibility="collapsed",
        )

    # Filter task list by search text
    if search_text.strip():
        mask = (
            tasks["task_code"].astype(str).str.contains(search_text.strip(), case=False, na=False) |
            tasks["task_name"].astype(str).str.contains(search_text.strip(), case=False, na=False)
        )
        filtered_tasks = tasks[mask]
        if filtered_tasks.empty:
            st.warning(f"No activities match '{search_text}'. Showing all activities.")
            filtered_tasks = tasks
    else:
        filtered_tasks = tasks

    with select_col:
        def _act_label(r):
            tf   = r.get("total_float_days")
            flag = " [CRITICAL]" if (tf is not None and safe_float(tf, 1) <= 0) else ""
            return f"{r.get('task_code','?')}  --  {r.get('task_name','?')}{flag}"

        act_labels = filtered_tasks.apply(_act_label, axis=1).tolist()

        if not act_labels:
            st.warning("No activities to select.")
            return

        selected_label = st.selectbox(
            "Select activity",
            options=act_labels,
            key="logic_trace_selector",
            label_visibility="collapsed",
        )

    sel_idx      = act_labels.index(selected_label)
    selected_row = filtered_tasks.iloc[sel_idx]
    selected_id  = selected_row["task_id"]
    sel_code     = str(selected_row.get("task_code", "-"))
    sel_name     = str(selected_row.get("task_name", "-"))

    # Clear results when activity changes
    if st.session_state.get("_trace_last_id") != selected_id:
        for k in ("trace_df", "trace_label", "trace_direction"):
            st.session_state.pop(k, None)
        st.session_state["_trace_last_id"] = selected_id

    # -------------------------------------------------------------------------
    # SELECTED ACTIVITY SUMMARY CARD
    # -------------------------------------------------------------------------
    sel_tf   = safe_float(selected_row.get("total_float_days"), None) if "total_float_days" in selected_row.index else None
    sel_ff   = safe_float(selected_row.get("free_float_days"),  None) if "free_float_days"  in selected_row.index else None
    sel_fcol = _float_color(sel_tf)
    sel_crit = bool(selected_row.get("is_critical", False)) if "is_critical" in selected_row.index else False
    sel_stat = _status_label(str(selected_row.get("status", "")))
    sel_scol = _status_colour(str(selected_row.get("status", "")))
    sel_wbs  = str(selected_row.get("wbs_path", "-")) if "wbs_path" in selected_row.index else "-"
    sel_start  = format_date(selected_row.get("eff_start")  if "eff_start"  in selected_row.index else None)
    sel_finish = format_date(selected_row.get("eff_finish") if "eff_finish" in selected_row.index else None)

    crit_pill = (
        '<span style="background:#dc2626;color:white;padding:2px 10px;'
        'border-radius:12px;font-size:11px;font-weight:700;margin-left:8px;">CRITICAL</span>'
        if sel_crit else ""
    )

    # Card layout: left = identity, right = schedule metrics
    card_left = f"""
        <div style="flex:1;min-width:0;">
            <div style="font-size:11px;color:#93c5fd;font-weight:600;
                        letter-spacing:1px;text-transform:uppercase;
                        margin-bottom:6px;">Selected Activity</div>
            <div style="font-size:20px;font-weight:800;white-space:nowrap;
                        overflow:hidden;text-overflow:ellipsis;">
                {sel_code}{crit_pill}
            </div>
            <div style="font-size:14px;color:#bfdbfe;margin-top:4px;
                        white-space:nowrap;overflow:hidden;text-overflow:ellipsis;"
                 title="{sel_name}">{sel_name}</div>
            <div style="font-size:12px;color:#64748B;margin-top:4px;
                        white-space:nowrap;overflow:hidden;text-overflow:ellipsis;"
                 title="{sel_wbs}">{sel_wbs}</div>
            <div style="margin-top:10px;display:flex;gap:6px;flex-wrap:wrap;">
                <span style="background:{sel_scol};color:white;padding:3px 10px;
                             border-radius:12px;font-size:11px;">{sel_stat}</span>
                <span style="background:{sel_fcol};color:white;padding:3px 10px;
                             border-radius:12px;font-size:11px;">
                    Float: {sel_tf if sel_tf is not None else "-"} days
                </span>
            </div>
        </div>"""

    card_right = f"""
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;
                    min-width:200px;align-self:center;">
            <div style="background:#0d2640;border-radius:6px;padding:8px 12px;">
                <div style="font-size:9px;color:#64748B;text-transform:uppercase;
                            letter-spacing:0.8px;">Start</div>
                <div style="font-size:13px;font-weight:600;color:#e2e8f0;
                            margin-top:2px;">{sel_start}</div>
            </div>
            <div style="background:#0d2640;border-radius:6px;padding:8px 12px;">
                <div style="font-size:9px;color:#64748B;text-transform:uppercase;
                            letter-spacing:0.8px;">Finish</div>
                <div style="font-size:13px;font-weight:600;color:#e2e8f0;
                            margin-top:2px;">{sel_finish}</div>
            </div>
            <div style="background:#0d2640;border-radius:6px;padding:8px 12px;">
                <div style="font-size:9px;color:#64748B;text-transform:uppercase;
                            letter-spacing:0.8px;">Total Float</div>
                <div style="font-size:15px;font-weight:700;color:#F5A623;
                            margin-top:2px;">{(str(sel_tf) + "d") if sel_tf is not None else "-"}</div>
            </div>
            <div style="background:#0d2640;border-radius:6px;padding:8px 12px;">
                <div style="font-size:9px;color:#64748B;text-transform:uppercase;
                            letter-spacing:0.8px;">Free Float</div>
                <div style="font-size:13px;font-weight:600;color:#e2e8f0;
                            margin-top:2px;">{(str(sel_ff) + "d") if sel_ff is not None else "-"}</div>
            </div>
        </div>"""

    st.markdown(
        f"""
        <div style="background:#1e3a5f;color:white;border-radius:12px;
                    padding:20px 24px;margin:10px 0 20px 0;
                    display:flex;gap:24px;align-items:flex-start;flex-wrap:wrap;">
            {card_left}
            {card_right}
        </div>
        """,
        unsafe_allow_html=True,
    )

    # -------------------------------------------------------------------------
    # OPEN-END WARNINGS
    # -------------------------------------------------------------------------
    direct_preds = list(G.predecessors(selected_id))
    direct_succs = list(G.successors(selected_id))
    has_preds    = len(direct_preds) > 0
    has_succs    = len(direct_succs) > 0

    if not has_preds and not has_succs:
        st.error(
            f"**{sel_code} has no logic connections.** "
            "This activity is completely isolated in the programme network - "
            "it has neither predecessors nor successors."
        )
        return

    if not has_preds:
        st.warning(
            f"**Open Start:** {sel_code} has no predecessors. "
            "It is not driven by any logic and may have artificially high float."
        )
    if not has_succs:
        st.warning(
            f"**Open Finish:** {sel_code} has no successors. "
            "Nothing in the programme is dependent on it completing."
        )

    # -------------------------------------------------------------------------
    # QUICK STATS ROW
    # -------------------------------------------------------------------------
    all_pred_ids = [tid for tid, _ in trace_predecessors(G, selected_id)]
    all_succ_ids = [tid for tid, _ in trace_successors(G,  selected_id)]

    def count_crit_in(id_list):
        return sum(
            1 for tid in id_list
            if safe_float(task_lookup.get(tid, {}).get("total_float_days"), 1) <= 0
        )

    st.markdown(
        _summary_bar(f"direct predecessors",   len(direct_preds), "#1d4ed8") +
        _summary_bar(f"direct successors",     len(direct_succs), "#1d4ed8") +
        _summary_bar(f"total predecessors",    len(all_pred_ids), "#4338ca") +
        _summary_bar(f"total successors",      len(all_succ_ids), "#4338ca") +
        (_summary_bar("critical in pred network", count_crit_in(all_pred_ids), "#dc2626") if all_pred_ids else "") +
        (_summary_bar("critical in succ network", count_crit_in(all_succ_ids), "#dc2626") if all_succ_ids else ""),
        unsafe_allow_html=True,
    )
    st.markdown("<br>", unsafe_allow_html=True)

    # -------------------------------------------------------------------------
    # TRACE BUTTONS
    # -------------------------------------------------------------------------
    st.markdown(
        '<div style="font-size:12px;font-weight:700;color:#94A3B8;'
        'letter-spacing:1px;text-transform:uppercase;margin-bottom:8px;">'
        'Choose what to trace</div>',
        unsafe_allow_html=True,
    )

    b1, b2, b3, b4 = st.columns(4)
    btn_dir_pred = b1.button(
        "◀  Direct Predecessors",
        key="btn_dp", use_container_width=True, disabled=not has_preds,
        help="Show only the activities that directly link into this activity (depth 1).",
    )
    btn_dir_succ = b2.button(
        "▶  Direct Successors",
        key="btn_ds", use_container_width=True, disabled=not has_succs,
        help="Show only the activities this activity directly links into (depth 1).",
    )
    btn_all_pred = b3.button(
        "◀◀  All Predecessors",
        key="btn_ap", use_container_width=True, disabled=not has_preds,
        help="Trace the full predecessor network back through all levels.",
    )
    btn_all_succ = b4.button(
        "▶▶  All Successors",
        key="btn_as", use_container_width=True, disabled=not has_succs,
        help="Trace the full successor network forward through all levels.",
    )

    # -------------------------------------------------------------------------
    # HANDLE BUTTONS
    # -------------------------------------------------------------------------
    new_trace = None
    new_label = None
    new_dir   = None

    if btn_dir_pred:
        new_trace = [(p, 1) for p in direct_preds]
        new_label = f"Direct Predecessors of {sel_code}"
        new_dir   = "pred"
    elif btn_dir_succ:
        new_trace = [(s, 1) for s in direct_succs]
        new_label = f"Direct Successors of {sel_code}"
        new_dir   = "succ"
    elif btn_all_pred:
        new_trace = trace_predecessors(G, selected_id)
        new_label = f"All Predecessors of {sel_code}"
        new_dir   = "pred"
    elif btn_all_succ:
        new_trace = trace_successors(G, selected_id)
        new_label = f"All Successors of {sel_code}"
        new_dir   = "succ"

    if new_trace is not None:
        raw_df = _build_full_trace_df(G, rels, task_lookup, selected_id, new_trace, new_dir)

        # Enrich with WBS column
        if not raw_df.empty and "Activity ID" in raw_df.columns:
            code_to_wbs = tasks.set_index("task_code")["wbs_path"].to_dict() if "wbs_path" in tasks.columns else {}
            raw_df.insert(
                raw_df.columns.get_loc("Activity Name") + 1,
                "WBS",
                raw_df["Activity ID"].map(code_to_wbs).fillna("-"),
            )

        st.session_state["trace_df"]        = raw_df
        st.session_state["trace_label"]     = new_label
        st.session_state["trace_direction"] = new_dir

    # -------------------------------------------------------------------------
    # RESULTS
    # -------------------------------------------------------------------------
    if "trace_df" not in st.session_state or st.session_state["trace_df"].empty:
        st.markdown(
            '<div style="background:#f0f9ff;border:2px dashed #93c5fd;border-radius:10px;'
            'padding:32px;text-align:center;color:#1e40af;margin-top:20px;">'
            '<div style="font-size:28px;margin-bottom:10px;">🔗</div>'
            '<strong style="font-size:15px;">Press a button above to trace the logic</strong><br>'
            '<span style="font-size:13px;color:#64748B;margin-top:6px;display:block;">'
            'Results will appear here.</span>'
            '</div>',
            unsafe_allow_html=True,
        )
        return

    trace_df  = st.session_state["trace_df"]
    trace_lbl = st.session_state.get("trace_label", "Trace Results")
    trace_dir = st.session_state.get("trace_direction", "pred")

    st.divider()

    # Counts
    n_res  = len(trace_df)
    n_crit = int((trace_df["Critical Flag"] == "Critical").sum())
    n_neg  = int((trace_df["Critical Flag"] == "Negative Float").sum())
    n_near = int((trace_df["Critical Flag"] == "Near-Critical").sum())

    st.markdown(
        f"<h4 style='color:#0B1F33;margin-bottom:6px;'>{trace_lbl}</h4>",
        unsafe_allow_html=True,
    )
    st.markdown(
        _summary_bar(f"activities",      n_res,  "#374151") +
        (_summary_bar("critical",        n_crit, "#dc2626") if n_crit else "") +
        (_summary_bar("negative float",  n_neg,  "#7f1d1d") if n_neg  else "") +
        (_summary_bar("near-critical",   n_near, "#d97706") if n_near else ""),
        unsafe_allow_html=True,
    )
    st.markdown("<br>", unsafe_allow_html=True)

    if trace_dir == "pred":
        st.caption("Depth = steps back from the selected activity. Depth 1 = directly connected.")
    elif trace_dir == "succ":
        st.caption("Depth = steps forward from the selected activity. Depth 1 = directly connected.")

    # -------------------------------------------------------------------------
    # RESULT FILTERS
    # -------------------------------------------------------------------------
    with st.expander("Filter results", expanded=False):
        fc1, fc2, fc3 = st.columns(3)

        filt_flag = fc1.selectbox(
            "Float status",
            ["All", "Critical & Negative Float", "Near-Critical", "Has Float"],
            key="lt_filt_flag",
        )

        filt_wbs = fc2.text_input(
            "WBS contains",
            placeholder="e.g. Civil",
            key="lt_filt_wbs",
        )

        filt_float_max = fc3.number_input(
            "Max float (days)",
            value=9999,
            step=1,
            min_value=-9999,
            key="lt_filt_float",
            help="Show only activities with total float at or below this value.",
        )

        # Date range filter
        fd1, fd2 = st.columns(2)
        filt_finish_from = fd1.text_input(
            "Finish on or after (dd/mm/yyyy)",
            placeholder="01/01/2024",
            key="lt_filt_df",
        )
        filt_finish_to = fd2.text_input(
            "Finish on or before (dd/mm/yyyy)",
            placeholder="31/12/2025",
            key="lt_filt_dt",
        )

    # Apply filters to a copy of trace_df
    display_df = trace_df.copy()

    if filt_flag == "Critical & Negative Float":
        display_df = display_df[display_df["Critical Flag"].isin(["Critical", "Negative Float"])]
    elif filt_flag == "Near-Critical":
        display_df = display_df[display_df["Critical Flag"] == "Near-Critical"]
    elif filt_flag == "Has Float":
        display_df = display_df[display_df["Critical Flag"] == "Float"]

    if "WBS" in display_df.columns and filt_wbs.strip():
        display_df = display_df[
            display_df["WBS"].astype(str).str.contains(filt_wbs.strip(), case=False, na=False)
        ]

    if filt_float_max < 9999 and "Total Float (d)" in display_df.columns:
        display_df = display_df[
            display_df["Total Float (d)"].apply(
                lambda v: safe_float(v, 9999) <= filt_float_max
            )
        ]

    def _parse_date_filter(s):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(s.strip(), fmt)
            except Exception:
                pass
        return None

    if filt_finish_from.strip() and "Finish" in display_df.columns:
        d_from = _parse_date_filter(filt_finish_from)
        if d_from:
            # Finish column is a formatted string - compare via task lookup
            finish_dates = {}
            for _, row in tasks.iterrows():
                finish_dates[str(row.get("task_code",""))] = row.get("eff_finish")
            display_df = display_df[
                display_df["Activity ID"].apply(
                    lambda code: (lambda fd: fd is not None and fd >= d_from)(finish_dates.get(code))
                )
            ]

    if filt_finish_to.strip() and "Finish" in display_df.columns:
        d_to = _parse_date_filter(filt_finish_to)
        if d_to:
            finish_dates = {str(r.get("task_code","")): r.get("eff_finish") for _, r in tasks.iterrows()}
            display_df = display_df[
                display_df["Activity ID"].apply(
                    lambda code: (lambda fd: fd is not None and fd <= d_to)(finish_dates.get(code))
                )
            ]

    n_display = len(display_df)
    if n_display < n_res:
        st.caption(f"Showing {n_display} of {n_res} activities after filters.")

    if display_df.empty:
        st.info("No activities match the current filters.")
    else:
        # Styled table
        def _colour_flag(val):
            return {
                "Critical":       "background-color:#fee2e2;color:#991b1b;font-weight:600;",
                "Negative Float": "background-color:#fecaca;color:#7f1d1d;font-weight:700;",
                "Near-Critical":  "background-color:#fef3c7;color:#92400e;font-weight:600;",
                "Float":          "background-color:#dcfce7;color:#166534;",
            }.get(val, "")

        styled = display_df.style.applymap(_colour_flag, subset=["Critical Flag"])
        st.dataframe(styled, use_container_width=True, hide_index=True,
                     height=min(600, 45 + n_display * 35))

    # -------------------------------------------------------------------------
    # TABS: FILTERED VIEWS
    # -------------------------------------------------------------------------
    if n_res > 5:
        st.markdown("<br>", unsafe_allow_html=True)
        tab_all, tab_crit, tab_near, tab_open = st.tabs(
            ["All Results", "Critical & Neg Float", "Near-Critical", "Open Ends in Chain"]
        )

        with tab_all:
            st.dataframe(trace_df, use_container_width=True, hide_index=True)

        with tab_crit:
            crit_df = trace_df[trace_df["Critical Flag"].isin(["Critical", "Negative Float"])]
            if crit_df.empty:
                st.success("No critical or negative float activities in this trace.")
            else:
                st.dataframe(crit_df, use_container_width=True, hide_index=True)

        with tab_near:
            near_df = trace_df[trace_df["Critical Flag"] == "Near-Critical"]
            if near_df.empty:
                st.success("No near-critical activities in this trace.")
            else:
                st.dataframe(near_df, use_container_width=True, hide_index=True)

        with tab_open:
            open_rows = []
            for act_code in trace_df["Activity ID"].unique():
                match = tasks[tasks["task_code"] == act_code]
                if match.empty:
                    continue
                mid  = match.iloc[0]["task_id"]
                no_p = len(list(G.predecessors(mid))) == 0
                no_s = len(list(G.successors(mid))) == 0
                if no_p or no_s:
                    t = task_lookup.get(mid, {})
                    open_rows.append({
                        "Activity ID":   t.get("task_code", mid),
                        "Activity Name": t.get("task_name", ""),
                        "Issue": ("No predecessors" if no_p else "") +
                                 (" | No successors" if no_s else ""),
                        "Float (d)":     t.get("total_float_days", "-"),
                    })
            if open_rows:
                st.dataframe(pd.DataFrame(open_rows), use_container_width=True, hide_index=True)
            else:
                st.success("No open-ended activities found in this trace.")

    # -------------------------------------------------------------------------
    # GANTT CHART
    # -------------------------------------------------------------------------
    st.divider()
    st.markdown(
        '<div style="font-size:12px;font-weight:700;color:#94A3B8;'
        'letter-spacing:1px;text-transform:uppercase;margin-bottom:8px;">'
        'Trace Timeline</div>',
        unsafe_allow_html=True,
    )

    gantt_data = trace_df.copy().merge(
        tasks[["task_code","eff_start","eff_finish"]].rename(columns={"task_code":"Activity ID"}),
        on="Activity ID", how="left",
    ).dropna(subset=["eff_start","eff_finish"])

    if not gantt_data.empty:
        gantt_data["Label"] = (
            gantt_data["Activity ID"].astype(str) + "  " +
            gantt_data["Activity Name"].astype(str).str[:38]
        )
        gantt_data["Bar Colour"] = gantt_data["Critical Flag"].map({
            "Critical":       "Critical",
            "Negative Float": "Critical",
            "Near-Critical":  "Near-Critical",
            "Float":          "Has Float",
        }).fillna("Has Float")

        fig = px.timeline(
            gantt_data.head(80),
            x_start="eff_start", x_end="eff_finish", y="Label",
            color="Bar Colour",
            color_discrete_map={
                "Critical":     "#dc2626",
                "Near-Critical":"#d97706",
                "Has Float":    "#2563eb",
            },
            title=f"Timeline: {trace_lbl}",
            labels={"Label": ""},
        )
        fig.update_yaxes(autorange="reversed")
        fig.add_vline(
            x=datetime.now(), line_dash="dot", line_color="#94A3B8",
            annotation_text="Today", annotation_position="top left",
        )
        fig.update_layout(
            legend_title_text="Float Status",
            height=max(300, min(700, 60 + len(gantt_data.head(80)) * 28)),
            margin=dict(l=10, r=10, t=40, b=10),
            plot_bgcolor="#ffffff",
            paper_bgcolor="#ffffff",
        )
        st.plotly_chart(fig, use_container_width=True)
        if len(gantt_data) > 80:
            st.caption(f"Showing first 80 of {len(gantt_data)} activities on the timeline.")
    else:
        st.info("No date data available to build the timeline.")

    # -------------------------------------------------------------------------
    # EXCEL EXPORT
    # -------------------------------------------------------------------------
    st.divider()

    summary_data = {
        "Item": [
            "Selected Activity ID", "Activity Name", "WBS",
            "Trace Type", "Total in chain",
            "Critical in chain", "Near-Critical in chain", "Negative Float in chain",
        ],
        "Value": [
            sel_code, sel_name, sel_wbs,
            trace_lbl, n_res, n_crit, n_near, n_neg,
        ],
    }

    export_sheets = {"Summary": pd.DataFrame(summary_data), "Full Trace": trace_df}

    crit_exp = trace_df[trace_df["Critical Flag"].isin(["Critical","Negative Float"])]
    near_exp = trace_df[trace_df["Critical Flag"] == "Near-Critical"]
    if not crit_exp.empty:
        export_sheets["Critical Activities"] = crit_exp
    if not near_exp.empty:
        export_sheets["Near-Critical"]       = near_exp
    if n_display < n_res and not display_df.empty:
        export_sheets["Filtered View"] = display_df

    xls_bytes = export_df_to_excel(export_sheets)

    st.download_button(
        label="📥  Export Logic Trace to Excel",
        data=xls_bytes,
        file_name=f"logic_trace_{sel_code}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Exports Summary, Full Trace, Critical Activities, Near-Critical and any filtered view.",
    )


# -----------------------------------------------------------------------------
# PAGE: CRITICAL PATH ANALYSIS
# -----------------------------------------------------------------------------

def page_critical_path(data: dict, near_crit_days: float):
    st.title("🚨 Critical Path Analysis")

    tasks = data["tasks_df"]
    rels = data["relationships_df"]

    if tasks.empty:
        st.warning("No activities loaded.")
        return

    tasks = get_critical_threshold(tasks, near_crit_days)

    tab1, tab2, tab3, tab4 = st.tabs(
        ["Critical Activities", "Near-Critical", "Negative Float", "By WBS / Package"]
    )

    with tab1:
        critical = tasks[tasks["is_critical"]].sort_values("total_float_days")
        st.metric("Critical Activities", len(critical))
        disp = ["task_code", "task_name", "wbs_path", "eff_start", "eff_finish",
                "total_float_days", "status"]
        avail = [c for c in disp if c in critical.columns]
        st.dataframe(critical[avail], use_container_width=True)

        if not critical.empty and "eff_start" in critical.columns:
            st.subheader("Critical Path Gantt")
            gantt_df = critical.dropna(subset=["eff_start", "eff_finish"]).copy()
            gantt_df["Start"] = gantt_df["eff_start"]
            gantt_df["Finish"] = gantt_df["eff_finish"]
            gantt_df["Task"] = gantt_df["task_code"] + " - " + gantt_df["task_name"]
            if len(gantt_df) > 0:
                fig = px.timeline(
                    gantt_df.head(50),
                    x_start="Start", x_end="Finish", y="Task",
                    title="Critical Path Activities (top 50)",
                    color_discrete_sequence=["#dc2626"],
                )
                fig.update_yaxes(autorange="reversed")
                st.plotly_chart(fig, use_container_width=True)

        xls = export_df_to_excel({"Critical Path": critical[avail]})
        st.download_button("📥 Export Critical Path", xls,
                           "critical_path.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with tab2:
        near_crit = tasks[tasks["is_near_critical"]].sort_values("total_float_days")
        st.metric(f"Near-Critical (0 < float <= {near_crit_days}d)", len(near_crit))
        avail = [c for c in ["task_code","task_name","wbs_path","eff_start","eff_finish",
                              "total_float_days","status"] if c in near_crit.columns]
        st.dataframe(near_crit[avail], use_container_width=True)

    with tab3:
        neg = tasks[tasks["total_float_days"].apply(lambda f: f is not None and f < 0)].sort_values("total_float_days")
        st.metric("Negative Float Activities", len(neg))
        if not neg.empty:
            st.warning("⚠️ Activities with negative float indicate the schedule cannot be met -- investigate immediately.")
            avail = [c for c in ["task_code","task_name","total_float_days","eff_start",
                                  "eff_finish","status"] if c in neg.columns]
            st.dataframe(neg[avail], use_container_width=True)

    with tab4:
        if "wbs_path" not in tasks.columns:
            st.info("WBS data not available.")
            return
        tasks["wbs_top"] = tasks["wbs_path"].apply(
            lambda x: str(x).split(" > ")[0] if pd.notna(x) else "Unknown"
        )
        wbs_crit = tasks.groupby("wbs_top").agg(
            total=("task_id", "count"),
            critical=("is_critical", "sum"),
            near_critical=("is_near_critical", "sum"),
        ).reset_index()
        wbs_crit["crit_%"] = (wbs_crit["critical"] / wbs_crit["total"] * 100).round(1)
        fig = px.bar(
            wbs_crit, x="wbs_top", y=["critical", "near_critical"],
            title="Critical & Near-Critical by WBS",
            labels={"value": "Activities", "wbs_top": "WBS"},
            color_discrete_map={"critical": "#dc2626", "near_critical": "#f59e0b"},
            barmode="group",
        )
        st.plotly_chart(fig, use_container_width=True)
        st.dataframe(wbs_crit, use_container_width=True)


# -----------------------------------------------------------------------------
# PAGE: CRITICAL PATH TO SELECTED ACTIVITY
# -----------------------------------------------------------------------------

def _network_diagram_html(
    path_ids: list,
    all_pred_ids: list,
    task_lookup: dict,
    rels_df: pd.DataFrame,
) -> str:
    """
    Build a lightweight SVG-based network diagram showing the driving path
    as a horizontal chain, with branch predecessors shown above/below.
    Returns an HTML string ready for st.components.v1.html().
    """
    if not path_ids:
        return ""

    # Colour helpers
    def node_colour(tid):
        t  = task_lookup.get(tid, {})
        tf = safe_float(t.get("total_float_days"), 9999)
        if tf < 0:  return "#7f1d1d", "#fecaca"   # bg, border
        if tf == 0: return "#dc2626", "#fee2e2"
        if tf <= 10:return "#d97706", "#fef3c7"
        return       "#1d4ed8", "#dbeafe"

    BOX_W, BOX_H = 160, 54
    H_GAP, V_GAP = 40, 70
    n = len(path_ids)
    canvas_w = n * (BOX_W + H_GAP) + H_GAP
    canvas_h = 200

    svg_parts = [
        f'<svg xmlns="http://www.w3.org/2000/svg" '
        f'width="{canvas_w}" height="{canvas_h}" '
        f'style="font-family:Arial,sans-serif;font-size:11px;">'
    ]

    # Draw boxes for driving path (middle row y=70)
    mid_y = 80
    box_centres = {}
    for i, tid in enumerate(path_ids):
        t    = task_lookup.get(tid, {})
        x    = H_GAP + i * (BOX_W + H_GAP)
        y    = mid_y
        cx   = x + BOX_W // 2
        cy   = y + BOX_H // 2
        box_centres[tid] = (cx, cy)
        fc, bc = node_colour(tid)
        code = str(t.get("task_code", tid))[:14]
        name = str(t.get("task_name", ""))[:20]
        tf   = t.get("total_float_days")
        tf_s = f"Float: {tf}d" if tf is not None else ""
        svg_parts.append(
            f'<rect x="{x}" y="{y}" width="{BOX_W}" height="{BOX_H}" '
            f'rx="6" fill="{bc}" stroke="{fc}" stroke-width="2"/>'
            f'<text x="{cx}" y="{y+16}" text-anchor="middle" '
            f'font-weight="bold" fill="{fc}">{code}</text>'
            f'<text x="{cx}" y="{y+30}" text-anchor="middle" fill="#374151">{name}</text>'
            f'<text x="{cx}" y="{y+44}" text-anchor="middle" fill="{fc}" '
            f'font-size="10">{tf_s}</text>'
        )

        # Arrow from previous box
        if i > 0:
            prev_tid = path_ids[i - 1]
            px2, py2 = box_centres[prev_tid]
            # Get rel type
            rel_label = "FS"
            lag_label = ""
            if not rels_df.empty:
                rel = rels_df[
                    (rels_df.get("pred_task_id", pd.Series(dtype=str)) == prev_tid) &
                    (rels_df.get("succ_task_id", pd.Series(dtype=str)) == tid)
                ]
                if not rel.empty:
                    rel_label = str(rel["rel_type"].iloc[0])[-2:] if "rel_type" in rel.columns else "FS"
                    lag = safe_float(rel["lag_days"].iloc[0] if "lag_days" in rel.columns else 0, 0)
                    lag_label = f" +{int(lag)}d" if lag > 0 else (f" {int(lag)}d" if lag < 0 else "")
            ax1 = px2 + BOX_W // 2
            ax2 = x
            ay  = cy
            svg_parts.append(
                f'<line x1="{ax1}" y1="{ay}" x2="{ax2}" y2="{ay}" '
                f'stroke="#6b7280" stroke-width="2" marker-end="url(#arr)"/>'
                f'<text x="{(ax1+ax2)//2}" y="{ay-5}" text-anchor="middle" '
                f'fill="#6b7280" font-size="10">{rel_label}{lag_label}</text>'
            )

    # Arrow marker def
    svg_parts.insert(1,
        '<defs><marker id="arr" markerWidth="8" markerHeight="8" '
        'refX="6" refY="3" orient="auto">'
        '<path d="M0,0 L0,6 L8,3 z" fill="#6b7280"/>'
        '</marker></defs>'
    )

    svg_parts.append('</svg>')
    html = (
        '<div style="overflow-x:auto;background:#f8fafc;border:1px solid #e2e8f0;'
        'border-radius:8px;padding:12px;">'
        + "".join(svg_parts) +
        '</div>'
    )
    return html


def page_critical_path_to_activity(data: dict, near_crit_days: float):
    """
    Critical Path to Selected Activity page.

    Traces backwards through predecessor logic to identify the most likely
    driving chain into any selected activity or milestone.
    Uses float, finish dates, relationship type and lag to determine the driver.
    """
    st.title("🎯 Critical Path to Selected Activity")

    # -------------------------------------------------------------------------
    # EXPLANATION BANNER
    # -------------------------------------------------------------------------
    st.markdown(
        '<div style="background:#eff6ff;border-left:4px solid #3b82f6;'
        'border-radius:6px;padding:14px 18px;margin-bottom:18px;">'
        '<strong>How this works</strong><br>'
        'This shows the likely chain of activities driving the selected activity. '
        'It traces backwards through predecessor logic and identifies the most critical '
        'path based on total float, latest finish dates, relationship types and lag. '
        '<br><br>'
        '<em>This is based on available XER logic and float values and should be '
        'reviewed with the planner before being used for decision-making.</em>'
        '</div>',
        unsafe_allow_html=True,
    )

    tasks = data["tasks_df"]
    rels  = data["relationships_df"]

    # Guard rails
    if tasks.empty:
        st.warning("No activities found. Please upload an XER file first.")
        return
    if rels.empty:
        st.warning(
            "No relationship data found in this XER file. "
            "This page requires both activities and relationships."
        )
        return

    tasks = get_critical_threshold(tasks, near_crit_days)
    G           = build_graph(tasks, rels)
    task_lookup = tasks.set_index("task_id").to_dict("index")

    # -------------------------------------------------------------------------
    # ACTIVITY SELECTOR
    # -------------------------------------------------------------------------
    def _label(r):
        code = str(r.get("task_code", "?"))
        name = str(r.get("task_name", "?"))
        tf   = r.get("total_float_days")
        flag = " [CRITICAL]" if (tf is not None and safe_float(tf, 1) <= 0) else ""
        return f"{code}  --  {name}{flag}"

    act_labels = tasks.apply(_label, axis=1).tolist()
    selected_label = st.selectbox(
        "Select target activity or milestone",
        options=act_labels,
        key="cpta_selector",
        help="Choose the activity you want to understand the driving path for.",
    )
    sel_idx    = act_labels.index(selected_label)
    target_row = tasks.iloc[sel_idx]
    target_id  = target_row["task_id"]
    tgt_code   = str(target_row.get("task_code", "-"))
    tgt_name   = str(target_row.get("task_name", "-"))
    tgt_tf     = safe_float(target_row.get("total_float_days"), None) if "total_float_days" in target_row.index else None
    tgt_fcol   = _float_color(tgt_tf)
    tgt_crit   = bool(target_row.get("is_critical", False)) if "is_critical" in target_row.index else False
    tgt_stat   = _status_label(str(target_row.get("status", "")))
    tgt_scol   = _status_colour(str(target_row.get("status", "")))

    # Clear cached results when target changes
    if st.session_state.get("_cpta_last_id") != target_id:
        for k in ("cpta_path", "cpta_all_preds"):
            st.session_state.pop(k, None)
        st.session_state["_cpta_last_id"] = target_id

    # -------------------------------------------------------------------------
    # TARGET ACTIVITY BANNER
    # -------------------------------------------------------------------------
    crit_pill = (
        '<span style="background:#dc2626;color:white;padding:2px 10px;'
        'border-radius:12px;font-size:11px;font-weight:700;margin-left:8px;">CRITICAL</span>'
        if tgt_crit else ""
    )
    st.markdown(
        f"""
        <div style="background:#1e3a5f;color:white;border-radius:10px;
                    padding:16px 22px;margin:8px 0 18px 0;">
            <div style="font-size:12px;color:#93c5fd;font-weight:600;
                        letter-spacing:1px;text-transform:uppercase;">Target Activity</div>
            <div style="font-size:20px;font-weight:700;margin-top:4px;">
                {tgt_code}{crit_pill}
            </div>
            <div style="font-size:14px;color:#bfdbfe;margin-top:2px;">{tgt_name}</div>
            <div style="margin-top:10px;">
                <span style="background:{tgt_scol};color:white;padding:3px 10px;
                             border-radius:12px;font-size:12px;">{tgt_stat}</span>
                <span style="background:{tgt_fcol};color:white;padding:3px 10px;
                             border-radius:12px;font-size:12px;margin-left:6px;">
                    Float: {tgt_tf if tgt_tf is not None else "-"} days
                </span>
                <span style="color:#93c5fd;font-size:12px;margin-left:12px;">
                    Finish: {format_date(target_row.get("eff_finish") if "eff_finish" in target_row.index else None)}
                </span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Check if target has any predecessors at all
    direct_preds = list(G.predecessors(target_id))
    if not direct_preds:
        st.warning(
            f"**{tgt_code} has no predecessors.** This activity has an open start "
            "and is not driven by any logic in the programme. "
            "Nothing can be identified as the driving path."
        )
        return

    # -------------------------------------------------------------------------
    # RUN BUTTON
    # -------------------------------------------------------------------------
    run_col, _ = st.columns([1, 3])
    run_btn = run_col.button(
        "🔍  Find Driving Path",
        key="cpta_run",
        use_container_width=True,
        type="primary",
    )

    if run_btn:
        with st.spinner("Tracing predecessor network..."):
            driving_path   = driving_path_to_activity(G, tasks, rels, target_id)
            all_pred_pairs = trace_predecessors(G, target_id)
            all_pred_ids   = [p for p, _ in all_pred_pairs]
        st.session_state["cpta_path"]      = driving_path
        st.session_state["cpta_all_preds"] = all_pred_ids

    # -------------------------------------------------------------------------
    # RESULTS
    # -------------------------------------------------------------------------
    if "cpta_path" not in st.session_state:
        st.markdown(
            '<div style="background:#f0f9ff;border:1px dashed #93c5fd;border-radius:8px;'
            'padding:24px;text-align:center;color:#1e40af;margin-top:16px;">'
            '<strong>Press "Find Driving Path" above to run the analysis.</strong>'
            '</div>',
            unsafe_allow_html=True,
        )
        return

    driving_path = st.session_state["cpta_path"]
    all_pred_ids = st.session_state["cpta_all_preds"]

    # ---- KEY METRICS --------------------------------------------------------
    chain_tasks = tasks[tasks["task_id"].isin(driving_path)]
    n_chain     = len(driving_path)
    min_float   = chain_tasks["total_float_days"].min() if "total_float_days" in chain_tasks.columns else None
    n_crit_chain = int((chain_tasks["total_float_days"].apply(
        lambda f: safe_float(f, 1) <= 0
    )).sum()) if "total_float_days" in chain_tasks.columns else 0
    n_neg_chain  = int((chain_tasks["total_float_days"].apply(
        lambda f: safe_float(f, 0) < 0
    )).sum()) if "total_float_days" in chain_tasks.columns else 0

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Activities in Driving Chain",  n_chain)
    m2.metric("Lowest Float in Chain",         f"{min_float:.1f} days" if min_float is not None else "-")
    m3.metric("Critical in Chain",             n_crit_chain)
    m4.metric("Total Predecessor Network",     len(all_pred_ids))

    if n_neg_chain > 0:
        st.error(
            f"⚠️ **{n_neg_chain} activit{'y' if n_neg_chain == 1 else 'ies'} with negative float** "
            "in the driving chain. The current schedule cannot meet its target dates for this path."
        )

    st.divider()

    # ---- TABS ---------------------------------------------------------------
    tab_path, tab_network, tab_all_preds, tab_constraints = st.tabs([
        "Driving Path", "Network Diagram", "All Predecessors", "Constraints & Issues"
    ])

    # =========================================================================
    # TAB 1: DRIVING PATH TABLE
    # =========================================================================
    with tab_path:
        st.markdown(
            "The table below shows the most likely chain of activities driving "
            f"**{tgt_code}**, ordered from the earliest activity to the target. "
            "Activities are selected based on lowest float, latest finish date "
            "and relationship constraints."
        )

        path_rows = []
        for i, tid in enumerate(driving_path):
            t         = task_lookup.get(tid, {})
            tf        = t.get("total_float_days")
            is_target = (tid == target_id)

            # Relationship to next activity in chain
            rel_label = "-"
            lag_val   = 0
            if i < len(driving_path) - 1:
                next_tid = driving_path[i + 1]
                if not rels.empty:
                    rel = rels[
                        (rels.get("pred_task_id", pd.Series(dtype=str)) == tid) &
                        (rels.get("succ_task_id", pd.Series(dtype=str)) == next_tid)
                    ]
                    if not rel.empty:
                        rel_label = _rel_label(rel["rel_type"].iloc[0] if "rel_type" in rel.columns else "FS")
                        lag_val   = safe_float(rel["lag_days"].iloc[0] if "lag_days" in rel.columns else 0, 0)

            cstr = str(t.get("cstr_type", "")) if "cstr_type" in t else ""
            has_cstr = cstr.strip() not in ("", "None", "nan")

            path_rows.append({
                "Step":            i + 1,
                "Activity ID":     t.get("task_code", tid),
                "Activity Name":   t.get("task_name", ""),
                "Start":           format_date(t.get("eff_start")),
                "Finish":          format_date(t.get("eff_finish")),
                "Orig Dur (d)":    t.get("orig_dur_days", "-"),
                "Total Float (d)": tf if tf is not None else "-",
                "Link to Next":    rel_label if not is_target else "-",
                "Lag (d)":         lag_val if not is_target else "-",
                "Critical Flag":   _crit_flag(tf),
                "Constraint":      cstr if has_cstr else "",
                "Status":          _status_label(str(t.get("status", ""))),
                "Target":          "TARGET" if is_target else "",
            })

        path_df = pd.DataFrame(path_rows)

        # Colour code
        def _style_path_row(row):
            flag = row.get("Critical Flag", "")
            is_tgt = row.get("Target", "") == "TARGET"
            if is_tgt:
                return ["background-color:#1e3a5f;color:white;font-weight:700;"] * len(row)
            colour_map = {
                "Negative Float": "background-color:#fecaca;",
                "Critical":       "background-color:#fee2e2;",
                "Near-Critical":  "background-color:#fef3c7;",
            }
            style = colour_map.get(flag, "")
            return [style] * len(row)

        styled_path = path_df.style.apply(_style_path_row, axis=1)
        st.dataframe(styled_path, use_container_width=True, hide_index=True)

        # Gantt for driving path
        st.markdown("**Driving Path Timeline**")
        gantt_src = chain_tasks.dropna(subset=["eff_start","eff_finish"]).copy() if "eff_start" in chain_tasks.columns else pd.DataFrame()
        if not gantt_src.empty:
            gantt_src = gantt_src.merge(
                tasks[["task_id","task_code","task_name"]],
                on="task_id", how="left", suffixes=("","_t")
            )
            gantt_src["Label"]   = gantt_src["task_code"].astype(str) + "  " + gantt_src["task_name"].astype(str).str[:35]
            gantt_src["Colour"]  = gantt_src["task_id"].apply(
                lambda t: "Target" if t == target_id else (
                    "Critical" if safe_float(task_lookup.get(t,{}).get("total_float_days"), 1) <= 0
                    else "Near-Critical" if safe_float(task_lookup.get(t,{}).get("total_float_days"), 11) <= 10
                    else "Has Float"
                )
            )
            fig = px.timeline(
                gantt_src,
                x_start="eff_start", x_end="eff_finish", y="Label",
                color="Colour",
                color_discrete_map={
                    "Target":       "#1e3a5f",
                    "Critical":     "#dc2626",
                    "Near-Critical":"#d97706",
                    "Has Float":    "#2563eb",
                },
                title=f"Driving Path to {tgt_code}",
            )
            fig.update_yaxes(autorange="reversed")
            fig.add_vline(
                x=datetime.now(), line_dash="dot", line_color="#6b7280",
                annotation_text="Today", annotation_position="top left",
            )
            fig.update_layout(
                height=max(280, 50 + len(gantt_src) * 30),
                margin=dict(l=10, r=10, t=40, b=10),
                legend_title_text="Float Status",
            )
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No date data available for the Gantt chart.")

    # =========================================================================
    # TAB 2: NETWORK DIAGRAM
    # =========================================================================
    with tab_network:
        st.markdown(
            "A simple left-to-right network diagram of the driving path. "
            "Each box shows the activity ID, name and float. "
            "Colours: **red** = critical/negative float, **amber** = near-critical, **blue** = has float, "
            "**navy** = target activity."
        )

        if len(driving_path) > 0:
            diagram_html = _network_diagram_html(
                driving_path, all_pred_ids, task_lookup, rels
            )
            if diagram_html:
                import streamlit.components.v1 as components
                n_boxes = len(driving_path)
                diagram_w = n_boxes * 200 + 80
                components.html(diagram_html, height=220, scrolling=True)
            else:
                st.info("Could not generate network diagram.")

            st.caption(
                "Note: The diagram shows the identified driving path only. "
                "Use the All Predecessors tab to see the full predecessor network."
            )
        else:
            st.info("No path data to display.")

    # =========================================================================
    # TAB 3: ALL PREDECESSORS
    # =========================================================================
    with tab_all_preds:
        all_pred_tasks = tasks[tasks["task_id"].isin(all_pred_ids)].copy()

        if all_pred_tasks.empty:
            st.info("No predecessor activities found.")
        else:
            n_ap     = len(all_pred_tasks)
            n_ap_crit = int((all_pred_tasks["total_float_days"].apply(
                lambda f: safe_float(f, 1) <= 0
            )).sum()) if "total_float_days" in all_pred_tasks.columns else 0

            st.markdown(
                _summary_bar(f"total predecessors", n_ap, "#374151") +
                (_summary_bar("critical", n_ap_crit, "#dc2626") if n_ap_crit else ""),
                unsafe_allow_html=True,
            )
            st.markdown("<br>", unsafe_allow_html=True)
            st.caption(
                "All activities in the predecessor network of the target activity, "
                "sorted by float (most critical first)."
            )

            all_pred_tasks = all_pred_tasks.sort_values("total_float_days")

            AP_COLS = {
                "task_code":        "Activity ID",
                "task_name":        "Activity Name",
                "wbs_path":         "WBS",
                "eff_start":        "Start",
                "eff_finish":       "Finish",
                "total_float_days": "Float (d)",
                "status":           "Status",
                "is_critical":      "Critical",
            }
            ap_show = {k: v for k, v in AP_COLS.items() if k in all_pred_tasks.columns}
            ap_df   = all_pred_tasks[list(ap_show.keys())].rename(columns=ap_show).copy()
            for col in ["Start","Finish"]:
                if col in ap_df.columns:
                    ap_df[col] = ap_df[col].apply(format_date)
            if "Critical" in ap_df.columns:
                ap_df["Critical"] = ap_df["Critical"].apply(lambda x: "Yes" if x else "")
            if "Status" in ap_df.columns:
                ap_df["Status"] = ap_df["Status"].apply(_status_label)

            st.dataframe(ap_df, use_container_width=True, hide_index=True, height=400)

    # =========================================================================
    # TAB 4: CONSTRAINTS & ISSUES
    # =========================================================================
    with tab_constraints:
        st.markdown(
            "Activities in the driving path or predecessor network that have "
            "constraints, negative float, or other schedule quality issues."
        )

        issues_found = False

        # --- Negative float in driving path ---
        neg_in_path = chain_tasks[
            chain_tasks["total_float_days"].apply(lambda f: safe_float(f, 0) < 0)
        ] if "total_float_days" in chain_tasks.columns else pd.DataFrame()

        if not neg_in_path.empty:
            issues_found = True
            st.markdown(
                '<div style="background:#fef2f2;border-left:4px solid #dc2626;'
                'border-radius:6px;padding:10px 14px;margin-bottom:12px;">'
                f'<strong>⚠️ Negative Float in Driving Chain ({len(neg_in_path)} activities)</strong><br>'
                'These activities are beyond their target dates. '
                'The driving chain cannot currently meet its schedule.'
                '</div>',
                unsafe_allow_html=True,
            )
            neg_cols = {k: v for k, v in {
                "task_code":"Activity ID","task_name":"Activity Name",
                "total_float_days":"Float (d)","eff_finish":"Finish","status":"Status"
            }.items() if k in neg_in_path.columns}
            neg_disp = neg_in_path[list(neg_cols.keys())].rename(columns=neg_cols).copy()
            if "Finish" in neg_disp.columns:
                neg_disp["Finish"] = neg_disp["Finish"].apply(format_date)
            st.dataframe(neg_disp, use_container_width=True, hide_index=True)

        # --- Constraints in driving path ---
        if "cstr_type" in chain_tasks.columns:
            constrained = chain_tasks[
                chain_tasks["cstr_type"].apply(
                    lambda x: bool(x) and str(x).strip() not in ("","None","nan")
                )
            ]
            if not constrained.empty:
                issues_found = True
                st.markdown(
                    '<div style="background:#fffbeb;border-left:4px solid #f59e0b;'
                    'border-radius:6px;padding:10px 14px;margin-bottom:12px;">'
                    f'<strong>Constraints in Driving Chain ({len(constrained)} activities)</strong><br>'
                    'Constraints override schedule logic and can cause artificial float or '
                    'negative float. Each one should be reviewed with the planner.'
                    '</div>',
                    unsafe_allow_html=True,
                )
                cstr_cols = {k: v for k, v in {
                    "task_code":"Activity ID","task_name":"Activity Name",
                    "cstr_type":"Constraint Type","cstr_date":"Constraint Date",
                    "total_float_days":"Float (d)"
                }.items() if k in constrained.columns}
                cstr_disp = constrained[list(cstr_cols.keys())].rename(columns=cstr_cols).copy()
                if "Constraint Date" in cstr_disp.columns:
                    cstr_disp["Constraint Date"] = cstr_disp["Constraint Date"].apply(format_date)
                st.dataframe(cstr_disp, use_container_width=True, hide_index=True)

        # --- Constraints in full predecessor network ---
        all_pred_tasks_full = tasks[tasks["task_id"].isin(all_pred_ids)].copy()
        if "cstr_type" in all_pred_tasks_full.columns:
            all_constrained = all_pred_tasks_full[
                all_pred_tasks_full["cstr_type"].apply(
                    lambda x: bool(x) and str(x).strip() not in ("","None","nan")
                )
            ]
            if not all_constrained.empty:
                issues_found = True
                with st.expander(f"Constraints in Full Predecessor Network ({len(all_constrained)})"):
                    cstr_cols2 = {k: v for k, v in {
                        "task_code":"Activity ID","task_name":"Activity Name",
                        "cstr_type":"Constraint Type","cstr_date":"Constraint Date",
                        "total_float_days":"Float (d)"
                    }.items() if k in all_constrained.columns}
                    cstr_disp2 = all_constrained[list(cstr_cols2.keys())].rename(columns=cstr_cols2).copy()
                    if "Constraint Date" in cstr_disp2.columns:
                        cstr_disp2["Constraint Date"] = cstr_disp2["Constraint Date"].apply(format_date)
                    st.dataframe(cstr_disp2, use_container_width=True, hide_index=True)

        # --- High lag in driving path relationships ---
        if not rels.empty and "lag_days" in rels.columns:
            path_set  = set(driving_path)
            path_rels = rels[
                rels.get("pred_task_id", pd.Series(dtype=str)).isin(path_set) &
                rels.get("succ_task_id", pd.Series(dtype=str)).isin(path_set)
            ]
            high_lag = path_rels[
                path_rels["lag_days"].apply(lambda l: abs(safe_float(l, 0)) > 5)
            ] if not path_rels.empty else pd.DataFrame()

            if not high_lag.empty:
                issues_found = True
                st.markdown(
                    '<div style="background:#eff6ff;border-left:4px solid #3b82f6;'
                    'border-radius:6px;padding:10px 14px;margin-bottom:12px;">'
                    f'<strong>Significant Lag in Driving Path ({len(high_lag)} relationships)</strong><br>'
                    'Lag of more than 5 days can hide logic issues and affect float calculations.'
                    '</div>',
                    unsafe_allow_html=True,
                )
                lag_cols = {k: v for k, v in {
                    "pred_task_code":"From","pred_task_name":"From Name",
                    "succ_task_code":"To","succ_task_name":"To Name",
                    "rel_type":"Link","lag_days":"Lag (d)"
                }.items() if k in high_lag.columns}
                if lag_cols:
                    st.dataframe(
                        high_lag[list(lag_cols.keys())].rename(columns=lag_cols),
                        use_container_width=True, hide_index=True,
                    )

        if not issues_found:
            st.success(
                "No constraints, negative float or significant lag found in the driving path. "
                "The chain appears logically sound."
            )

    # =========================================================================
    # EXCEL EXPORT
    # =========================================================================
    st.divider()

    # Rebuild path_df for export (already built above in tab_path)
    export_path_df = pd.DataFrame(path_rows) if path_rows else pd.DataFrame()

    summary_rows = {
        "Item":  [
            "Target Activity ID", "Target Activity Name", "Target Finish",
            "Target Float (days)", "Activities in Driving Chain",
            "Lowest Float in Chain", "Critical in Chain", "Negative Float in Chain",
            "Total Predecessor Network",
        ],
        "Value": [
            tgt_code, tgt_name,
            format_date(target_row.get("eff_finish") if "eff_finish" in target_row.index else None),
            tgt_tf,
            n_chain, min_float, n_crit_chain, n_neg_chain, len(all_pred_ids),
        ],
    }

    export_sheets = {
        "Summary":        pd.DataFrame(summary_rows),
        "Driving Path":   export_path_df,
        "All Predecessors": ap_df if not all_pred_tasks.empty else pd.DataFrame(columns=["No data"]),
    }

    if not neg_in_path.empty:
        export_sheets["Negative Float"] = neg_disp
    if "cstr_type" in chain_tasks.columns and not constrained.empty:
        export_sheets["Constraints"] = cstr_disp

    xls_bytes = export_df_to_excel(export_sheets)

    dl_col, _ = st.columns([1, 3])
    dl_col.download_button(
        label="📥  Export Driving Path Report to Excel",
        data=xls_bytes,
        file_name=f"driving_path_{tgt_code}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Exports Summary, Driving Path, All Predecessors, Negative Float and Constraints sheets.",
        use_container_width=True,
    )

# -----------------------------------------------------------------------------
# PAGE: LABOUR HISTOGRAM
# -----------------------------------------------------------------------------

def page_labour_histogram(data: dict):
    st.title("👷 Labour Histogram")

    task_res = data["task_resources_df"]
    tasks = data["tasks_df"]
    resources = data["resources_df"]

    if task_res.empty:
        st.markdown("""
        <div class="warn-box">
        ⚠️ <strong>No resource loading found in this XER file.</strong><br>
        This usually means the programme was not resourced in P6, or resource data was not exported.
        <br><br>
        You can upload a separate resource CSV or Excel file below.
        </div>
        """, unsafe_allow_html=True)

        st.subheader("Upload Resource Loading File")
        res_file = st.file_uploader("Upload CSV or Excel (columns: task_code, rsrc_name, target_qty, target_start, target_finish)", type=["csv","xlsx"])
        if res_file:
            try:
                if res_file.name.endswith(".csv"):
                    task_res = pd.read_csv(res_file)
                else:
                    task_res = pd.read_excel(res_file)
                for col in ["target_start","target_finish"]:
                    if col in task_res.columns:
                        task_res[col] = pd.to_datetime(task_res[col], errors="coerce")
                st.success(f"Loaded {len(task_res)} resource rows.")
            except Exception as e:
                st.error(f"Could not read resource file: {e}")
                return
        else:
            return

    # Merge with task info and resource names
    if not tasks.empty and "task_id" in task_res.columns:
        task_res = task_res.merge(
            tasks[["task_id","task_code","task_name","wbs_path","is_critical" if "is_critical" in tasks.columns else "task_id"]].drop_duplicates(),
            on="task_id", how="left", suffixes=("","_task")
        )
    if not resources.empty and "rsrc_id" in task_res.columns:
        task_res = task_res.merge(resources[["rsrc_id","rsrc_name"]], on="rsrc_id", how="left", suffixes=("","_res"))
        if "rsrc_name_res" in task_res.columns:
            task_res["rsrc_name"] = task_res["rsrc_name_res"].fillna(task_res.get("rsrc_name",""))

    # Expand resource loading to weekly intervals
    def expand_to_weeks(df):
        rows = []
        for _, r in df.iterrows():
            s = pd.to_datetime(r.get("target_start") or r.get("target_start_date"))
            e = pd.to_datetime(r.get("target_finish") or r.get("target_end_date"))
            if pd.isna(s) or pd.isna(e) or s > e:
                continue
            qty = safe_float(r.get("target_qty", 0), 0)
            if qty == 0:
                continue
            weeks = max(1, math.ceil((e - s).days / 7))
            qty_per_week = qty / weeks
            current = s
            for _ in range(weeks):
                rows.append({
                    "week": current.to_period("W").start_time,
                    "month": current.to_period("M").start_time,
                    "qty": qty_per_week,
                    "rsrc_name": r.get("rsrc_name","Unknown"),
                    "task_code": r.get("task_code",""),
                    "task_name": r.get("task_name",""),
                    "wbs_path": r.get("wbs_path",""),
                })
                current += timedelta(weeks=1)
        return pd.DataFrame(rows)

    weekly = expand_to_weeks(task_res)

    if weekly.empty:
        st.warning("Could not generate histogram -- resource dates or quantities may be missing.")
        return

    # Filters
    st.sidebar.divider()
    st.sidebar.subheader("Labour Filters")
    all_resources = sorted(weekly["rsrc_name"].unique().tolist())
    sel_res = st.sidebar.multiselect("Resource / Trade", all_resources, default=all_resources)
    if sel_res:
        weekly = weekly[weekly["rsrc_name"].isin(sel_res)]

    # Metrics
    c1, c2, c3 = st.columns(3)
    c1.metric("Total Planned Hours", f"{weekly['qty'].sum():,.0f}")
    weekly_totals = weekly.groupby("week")["qty"].sum()
    c2.metric("Peak Week (hrs)", f"{weekly_totals.max():,.0f}" if not weekly_totals.empty else "-")
    c3.metric("Average Week (hrs)", f"{weekly_totals.mean():,.0f}" if not weekly_totals.empty else "-")

    tab1, tab2, tab3, tab4 = st.tabs(["By Week", "By Month", "By Resource", "By WBS"])

    with tab1:
        weekly_sum = weekly.groupby("week")["qty"].sum().reset_index()
        fig = px.bar(weekly_sum, x="week", y="qty",
                     title="Labour Loading by Week (Hours)",
                     labels={"week":"Week","qty":"Hours"},
                     color_discrete_sequence=["#2563eb"])
        st.plotly_chart(fig, use_container_width=True)

    with tab2:
        monthly_sum = weekly.groupby("month")["qty"].sum().reset_index()
        fig = px.bar(monthly_sum, x="month", y="qty",
                     title="Labour Loading by Month (Hours)",
                     labels={"month":"Month","qty":"Hours"},
                     color_discrete_sequence=["#1e3a5f"])
        st.plotly_chart(fig, use_container_width=True)

    with tab3:
        res_sum = weekly.groupby("rsrc_name")["qty"].sum().reset_index().sort_values("qty", ascending=False)
        fig = px.bar(res_sum, x="rsrc_name", y="qty",
                     title="Total Hours by Resource / Trade",
                     labels={"rsrc_name":"Resource","qty":"Hours"},
                     color_discrete_sequence=["#7c3aed"])
        st.plotly_chart(fig, use_container_width=True)

        # By week and resource stacked
        if len(sel_res) <= 10:
            by_res_week = weekly.groupby(["week","rsrc_name"])["qty"].sum().reset_index()
            fig2 = px.bar(by_res_week, x="week", y="qty", color="rsrc_name",
                          title="Weekly Labour by Resource",
                          labels={"week":"Week","qty":"Hours","rsrc_name":"Resource"})
            st.plotly_chart(fig2, use_container_width=True)

    with tab4:
        if "wbs_path" in weekly.columns:
            weekly["wbs_top"] = weekly["wbs_path"].apply(
                lambda x: str(x).split(" > ")[0] if pd.notna(x) and x else "Unknown"
            )
            wbs_sum = weekly.groupby("wbs_top")["qty"].sum().reset_index().sort_values("qty", ascending=False)
            fig = px.bar(wbs_sum, x="qty", y="wbs_top", orientation="h",
                         title="Total Hours by WBS",
                         color_discrete_sequence=["#059669"])
            st.plotly_chart(fig, use_container_width=True)

    # Export
    xls = export_df_to_excel({
        "Weekly Labour": weekly.groupby(["week","rsrc_name"])["qty"].sum().reset_index(),
        "Monthly Labour": weekly.groupby(["month","rsrc_name"])["qty"].sum().reset_index(),
        "By Resource": res_sum,
    })
    st.download_button("📥 Export Labour Data", xls, "labour_histogram.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# -----------------------------------------------------------------------------
# PAGE: SCHEDULE HEALTH CHECK
# -----------------------------------------------------------------------------

def page_health_check(data: dict, near_crit_days: float):
    st.title("🩺 Schedule Health Check")
    st.markdown("> Automated quality checks to identify common schedule issues.")

    tasks = data["tasks_df"]
    rels = data["relationships_df"]

    if tasks.empty:
        st.warning("No activities loaded.")
        return

    tasks = get_critical_threshold(tasks, near_crit_days)

    # Build predecessor/successor sets
    tasks_with_pred = set()
    tasks_with_succ = set()
    if not rels.empty:
        tasks_with_pred = set(rels["succ_task_id"].dropna()) if "succ_task_id" in rels.columns else set()
        tasks_with_succ = set(rels["pred_task_id"].dropna()) if "pred_task_id" in rels.columns else set()

    # Define checks
    checks = []

    # 1. No predecessors (excl. milestones at start)
    no_pred = tasks[~tasks["task_id"].isin(tasks_with_pred)]
    checks.append({
        "Check": "No Predecessors",
        "Count": len(no_pred),
        "Severity": "⚠️ Warning",
        "Why It Matters": "Activities with no predecessors are open-ended. They cannot be driven by logic and may cause float calculation issues.",
        "df": no_pred,
    })

    # 2. No successors
    no_succ = tasks[~tasks["task_id"].isin(tasks_with_succ)]
    checks.append({
        "Check": "No Successors",
        "Count": len(no_succ),
        "Severity": "⚠️ Warning",
        "Why It Matters": "Activities with no successors are open-ended and may have artificially high float.",
        "df": no_succ,
    })

    # 3. Negative float
    neg_float = tasks[tasks["total_float_days"].apply(lambda f: f is not None and f < 0)]
    checks.append({
        "Check": "Negative Float",
        "Count": len(neg_float),
        "Severity": "🔴 Critical",
        "Why It Matters": "Negative float means the current schedule cannot meet its target dates. Immediate attention required.",
        "df": neg_float,
    })

    # 4. High float (> 60 days)
    high_float = tasks[tasks["total_float_days"].apply(lambda f: f is not None and f > 60)]
    checks.append({
        "Check": "Very High Float (>60 days)",
        "Count": len(high_float),
        "Severity": "ℹ️ Info",
        "Why It Matters": "Activities with very high float may have missing logic or may not be properly constrained.",
        "df": high_float,
    })

    # 5. Excessive duration (> 60 working days)
    excess_dur = tasks[tasks["orig_dur_days"].apply(lambda d: d is not None and d > 60)]
    checks.append({
        "Check": "Excessive Duration (>60 days)",
        "Count": len(excess_dur),
        "Severity": "⚠️ Warning",
        "Why It Matters": "Very long activities are difficult to control and should usually be broken down into smaller work packages.",
        "df": excess_dur,
    })

    # 6. Constraints
    constrained = tasks[tasks["cstr_type"].apply(
        lambda x: bool(x) and str(x).strip() not in ("", "None")
    )] if "cstr_type" in tasks.columns else pd.DataFrame()
    checks.append({
        "Check": "Constrained Activities",
        "Count": len(constrained),
        "Severity": "⚠️ Warning",
        "Why It Matters": "Constraints override schedule logic and can create artificial float or negative float. Each constraint should be justified.",
        "df": constrained,
    })

    # 7. Excessive lag (> 10 days)
    if not rels.empty and "lag_days" in rels.columns:
        high_lag = rels[rels["lag_days"].apply(lambda l: l is not None and abs(safe_float(l,0)) > 10)]
        checks.append({
            "Check": "Excessive Lag (|lag| > 10 days)",
            "Count": len(high_lag),
            "Severity": "⚠️ Warning",
            "Why It Matters": "Excessive lag can hide critical path issues. Lag should be replaced with properly sequenced activities.",
            "df": high_lag,
        })

    # 8. Missing dates
    missing_dates = tasks[tasks["eff_start"].isna() | tasks["eff_finish"].isna()]
    checks.append({
        "Check": "Missing Start or Finish Dates",
        "Count": len(missing_dates),
        "Severity": "🔴 Critical",
        "Why It Matters": "Activities with no dates cannot be scheduled or reported on.",
        "df": missing_dates,
    })

    # 9. Actual dates in future
    now = datetime.now()
    future_actuals = tasks[
        tasks["act_start"].apply(lambda d: d is not None and d > now) |
        tasks["act_finish"].apply(lambda d: d is not None and d > now)
    ] if "act_start" in tasks.columns else pd.DataFrame()
    checks.append({
        "Check": "Future Actual Dates",
        "Count": len(future_actuals),
        "Severity": "🔴 Critical",
        "Why It Matters": "Actual start/finish dates should not be in the future. This indicates data entry errors.",
        "df": future_actuals,
    })

    # 10. Critical not started
    crit_not_started = tasks[
        tasks["is_critical"] &
        tasks["status"].apply(lambda s: str(s) in ("TK_NotStart", "Not Started") if pd.notna(s) else False)
    ] if "status" in tasks.columns else pd.DataFrame()
    checks.append({
        "Check": "Critical Activities Not Started",
        "Count": len(crit_not_started),
        "Severity": "🔴 Critical",
        "Why It Matters": "Critical activities that haven't started need immediate attention to avoid slippage.",
        "df": crit_not_started,
    })

    # 11. Near-critical due in 8 weeks
    eight_weeks = now + timedelta(weeks=8)
    near_due = tasks[
        tasks["is_near_critical"] &
        tasks["eff_finish"].apply(lambda d: d is not None and d <= eight_weeks)
    ] if "eff_finish" in tasks.columns else pd.DataFrame()
    checks.append({
        "Check": "Near-Critical Due in 8 Weeks",
        "Count": len(near_due),
        "Severity": "⚠️ Warning",
        "Why It Matters": "Near-critical activities finishing soon may become critical if not progressed.",
        "df": near_due,
    })

    # Scorecard
    st.subheader("Health Check Scorecard")
    score_data = [
        {"Check": c["Check"], "Count": c["Count"], "Severity": c["Severity"]}
        for c in checks
    ]
    score_df = pd.DataFrame(score_data)
    st.dataframe(score_df, use_container_width=True)

    # Detail per check
    st.divider()
    for chk in checks:
        with st.expander(f"{chk['Severity']} -- {chk['Check']} ({chk['Count']})"):
            st.markdown(f"**Why it matters:** {chk['Why It Matters']}")
            df = chk["df"]
            if not df.empty:
                disp = [c for c in ["task_code","task_name","wbs_path","eff_start",
                                     "eff_finish","total_float_days","status",
                                     "cstr_type","lag_days"] if c in df.columns]
                st.dataframe(df[disp].head(100), use_container_width=True)
                # Export individual check
                xls = export_df_to_excel({chk["Check"][:31]: df[disp]})
                st.download_button(
                    f"📥 Export: {chk['Check']}", xls,
                    f"health_{chk['Check'][:20].replace(' ','_')}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            else:
                st.success("✅ No issues found for this check.")

    # Full export
    all_export = {chk["Check"][:31]: chk["df"][[c for c in ["task_code","task_name","total_float_days","status"] if c in chk["df"].columns]] if not chk["df"].empty else pd.DataFrame(columns=["No issues"]) for chk in checks}
    xls_all = export_df_to_excel(all_export)
    st.download_button("📥 Export Full Health Check Report", xls_all, "schedule_health_check.xlsx",
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# -----------------------------------------------------------------------------
# PAGE: PLANNING NOTES
# -----------------------------------------------------------------------------

HIGHLIGHT_WORDS = [
    "risk", "delay", "delayed", "blocked", "constraint", "access",
    "design", "procurement", "client", "instruction", "CE", "EWN",
    "change", "issue", "hold", "pending", "late", "overrun",
]

def highlight_text(text: str) -> str:
    """Wrap highlight words in HTML span."""
    for word in HIGHLIGHT_WORDS:
        pattern = re.compile(r"\b(" + re.escape(word) + r")\b", re.IGNORECASE)
        text = pattern.sub(r'<span style="background:#fef08a;font-weight:bold;">\1</span>', text)
    return text


def page_planning_notes(data: dict):
    st.title("📝 Planning Notes")
    st.markdown("> Upload planning notes and link them to activities in the programme.")

    tasks = data["tasks_df"]
    notes_file = st.file_uploader("Upload Planning Notes (CSV, Excel, TXT, or DOCX)",
                                   type=["csv","xlsx","txt","docx"])

    if notes_file is None:
        st.info("Upload a notes file to get started. The file should contain free-text notes referencing activity IDs.")
        return

    # Read notes
    notes_text = ""
    notes_rows = []

    try:
        if notes_file.name.endswith(".csv"):
            df = pd.read_csv(notes_file)
            notes_text = " ".join(df.astype(str).values.flatten())
            notes_rows = df.to_dict("records")
        elif notes_file.name.endswith(".xlsx"):
            df = pd.read_excel(notes_file)
            notes_text = " ".join(df.astype(str).values.flatten())
            notes_rows = df.to_dict("records")
        elif notes_file.name.endswith(".txt"):
            notes_text = notes_file.read().decode("utf-8", errors="replace")
            notes_rows = [{"line": i+1, "text": line} for i, line in enumerate(notes_text.splitlines()) if line.strip()]
        elif notes_file.name.endswith(".docx"):
            from docx import Document
            doc = Document(io.BytesIO(notes_file.read()))
            lines = [p.text for p in doc.paragraphs if p.text.strip()]
            notes_text = "\n".join(lines)
            notes_rows = [{"paragraph": i+1, "text": line} for i, line in enumerate(lines)]
        else:
            st.error("Unsupported file format.")
            return
        st.success(f"Loaded notes file: {notes_file.name}")
    except Exception as e:
        st.error(f"Could not read notes file: {e}")
        return

    # Find activity IDs mentioned in notes
    if not tasks.empty and "task_code" in tasks.columns:
        task_codes = tasks["task_code"].dropna().tolist()
        found_codes = [code for code in task_codes if code in notes_text]

        st.subheader(f"Activity IDs Found in Notes: {len(found_codes)}")
        if found_codes:
            matched_tasks = tasks[tasks["task_code"].isin(found_codes)][
                ["task_code","task_name","eff_start","eff_finish","total_float_days","status"]
            ]
            st.dataframe(matched_tasks, use_container_width=True)
        else:
            st.info("No activity IDs from the programme were found in the notes.")

        # Not found
        not_found = [code for code in task_codes if code not in notes_text]
        st.caption(f"{len(not_found)} activities not mentioned in notes.")

    # Keyword search
    st.divider()
    st.subheader("Keyword Search")
    keyword = st.text_input("Search notes for keyword")

    display_rows = notes_rows
    if keyword:
        display_rows = [r for r in notes_rows if keyword.lower() in str(r).lower()]
        st.caption(f"{len(display_rows)} matching entries")

    # Display with highlights
    for row in display_rows[:100]:
        text = str(row.get("text","") or list(row.values())[-1])
        highlighted = highlight_text(text)
        st.markdown(f"<div style='background:#f8fafc;border-left:3px solid #2563eb;padding:8px;margin:4px 0;font-size:13px;'>{highlighted}</div>", unsafe_allow_html=True)

    # Full highlighted dump
    st.divider()
    st.subheader("Full Notes (with keyword highlighting)")
    highlighted_full = highlight_text(notes_text.replace("\n","<br>"))
    st.markdown(f"<div style='background:white;border:1px solid #e2e8f0;padding:16px;border-radius:8px;max-height:400px;overflow-y:auto;font-size:12px;'>{highlighted_full}</div>", unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# PAGE: PROGRAMME COMPARISON / MOVEMENT INTELLIGENCE
# -----------------------------------------------------------------------------

def _mi_card(title: str, value, subtitle: str = "", colour: str = "#0B1F33",
             bg: str = "#ffffff", border: str = "#E2E8F0") -> str:
    """Render a compact metric card as HTML."""
    return (
        f'<div style="background:{bg};border:1px solid {border};border-radius:10px;'
        f'padding:16px 18px;box-shadow:0 1px 4px rgba(11,31,51,0.07);">'
        f'<div style="font-size:10px;font-weight:700;color:#94A3B8;letter-spacing:1px;'
        f'text-transform:uppercase;margin-bottom:6px;">{title}</div>'
        f'<div style="font-size:26px;font-weight:800;color:{colour};line-height:1;">{value}</div>'
        f'{"" if not subtitle else f"<div style=font-size:11px;color:#64748B;margin-top:4px;>{subtitle}</div>"}'
        f'</div>'
    )


def _commentary_box(heading: str, body: str, colour: str = "#0B1F33",
                    bg: str = "#eff6ff", border: str = "#3b82f6") -> str:
    return (
        f'<div style="background:{bg};border-left:4px solid {border};border-radius:6px;'
        f'padding:14px 18px;margin-bottom:10px;">'
        f'<div style="font-weight:700;color:{colour};margin-bottom:4px;">{heading}</div>'
        f'<div style="font-size:14px;color:#334155;line-height:1.6;">{body}</div>'
        f'</div>'
    )


def _compute_movement(prev_tasks: pd.DataFrame, curr_tasks: pd.DataFrame,
                      near_crit: float = 10.0) -> dict:
    """
    Merge previous and current task sets and compute all movement metrics.
    Returns a dict of DataFrames and scalar stats.
    """
    prev = get_critical_threshold(prev_tasks.copy(), near_crit)
    curr = get_critical_threshold(curr_tasks.copy(), near_crit)

    # Common activities (matched on task_code)
    merged = prev.merge(curr, on="task_code", how="outer", suffixes=("_p", "_c"))

    added   = curr[~curr["task_code"].isin(prev["task_code"])].copy()
    removed = prev[~prev["task_code"].isin(curr["task_code"])].copy()
    common  = merged.dropna(subset=["task_code"]).copy()

    # ---- Date movement -------------------------------------------------------
    def _days_diff(a, b):
        """b minus a in calendar days. Positive = slipped."""
        try:
            if pd.isna(a) or pd.isna(b):
                return None
            return int((pd.Timestamp(b) - pd.Timestamp(a)).days)
        except Exception:
            return None

    common["finish_move"] = common.apply(
        lambda r: _days_diff(r.get("eff_finish_p"), r.get("eff_finish_c")), axis=1
    )
    common["start_move"] = common.apply(
        lambda r: _days_diff(r.get("eff_start_p"), r.get("eff_start_c")), axis=1
    )

    # ---- Float movement ------------------------------------------------------
    common["float_move"] = common.apply(
        lambda r: (
            safe_float(r.get("total_float_days_c"), None) is not None and
            safe_float(r.get("total_float_days_p"), None) is not None
        ) and safe_float(r.get("total_float_days_c"), 0) - safe_float(r.get("total_float_days_p"), 0)
        if (
            safe_float(r.get("total_float_days_c"), None) is not None and
            safe_float(r.get("total_float_days_p"), None) is not None
        ) else None,
        axis=1,
    )

    # ---- Critical transitions ------------------------------------------------
    def _is_crit(row, suffix):
        tf = safe_float(row.get(f"total_float_days_{suffix}"), 9999)
        return tf <= 0

    common["was_crit"]  = common.apply(lambda r: _is_crit(r, "p"), axis=1)
    common["now_crit"]  = common.apply(lambda r: _is_crit(r, "c"), axis=1)
    became_crit  = common[~common["was_crit"] & common["now_crit"]]
    no_longer_crit = common[common["was_crit"] & ~common["now_crit"]]

    # ---- Slipped / improved --------------------------------------------------
    delayed  = common[common["finish_move"].apply(lambda x: x is not None and x > 0)]
    improved = common[common["finish_move"].apply(lambda x: x is not None and x < 0)]
    float_lost   = common[common["float_move"].apply(lambda x: x is not None and x < 0)]
    float_gained = common[common["float_move"].apply(lambda x: x is not None and x > 0)]

    # ---- Project finish movement --------------------------------------------
    prev_end = prev["eff_finish"].dropna().max() if "eff_finish" in prev.columns else None
    curr_end = curr["eff_finish"].dropna().max() if "eff_finish" in curr.columns else None
    proj_finish_move = _days_diff(prev_end, curr_end)

    # ---- WBS-level summary --------------------------------------------------
    wbs_col_p = "wbs_path_p" if "wbs_path_p" in common.columns else None
    wbs_col_c = "wbs_path_c" if "wbs_path_c" in common.columns else None
    wbs_col   = wbs_col_c or wbs_col_p

    if wbs_col and not common.empty:
        common["wbs_top"] = common[wbs_col].apply(
            lambda x: str(x).split(" > ")[0] if pd.notna(x) and str(x).strip() not in ("","nan") else "Unknown"
        )
        wbs_summary = common.groupby("wbs_top").agg(
            total_activities=("task_code", "count"),
            delayed_count=("finish_move", lambda x: (x > 0).sum()),
            improved_count=("finish_move", lambda x: (x < 0).sum()),
            avg_finish_move=("finish_move", lambda x: round(x.dropna().mean(), 1) if x.dropna().any() else 0),
            max_finish_slip=("finish_move", lambda x: x.dropna().max() if not x.dropna().empty else 0),
            crit_gained=("now_crit", lambda x: (x & ~common.loc[x.index, "was_crit"]).sum()),
            crit_lost=("now_crit", lambda x: (~x & common.loc[x.index, "was_crit"]).sum()),
        ).reset_index()
    else:
        wbs_summary = pd.DataFrame()

    return {
        "merged":        common,
        "added":         added,
        "removed":       removed,
        "delayed":       delayed.sort_values("finish_move", ascending=False),
        "improved":      improved.sort_values("finish_move"),
        "float_lost":    float_lost.sort_values("float_move"),
        "float_gained":  float_gained.sort_values("float_move", ascending=False),
        "became_crit":   became_crit,
        "no_longer_crit":no_longer_crit,
        "wbs_summary":   wbs_summary,
        "proj_finish_move": proj_finish_move,
        "prev_end":      prev_end,
        "curr_end":      curr_end,
        "n_added":       len(added),
        "n_removed":     len(removed),
        "n_delayed":     len(delayed),
        "n_improved":    len(improved),
        "n_became_crit": len(became_crit),
        "n_no_longer_crit": len(no_longer_crit),
        "n_float_lost":  len(float_lost),
        "n_float_gained":len(float_gained),
        "n_common":      len(common),
    }


def _build_commentary(m: dict, prev_name: str, curr_name: str) -> list:
    """
    Generate plain-English commentary bullets from movement data.
    Returns list of (heading, body, colour, bg, border) tuples.
    """
    items = []
    pfm = m["proj_finish_move"]

    # -- What changed? ---------------------------------------------------------
    changes = []
    if pfm is not None and pfm > 0:
        changes.append(f"The project finish date has slipped by <strong>{pfm} days</strong>.")
    elif pfm is not None and pfm < 0:
        changes.append(f"The project finish date has improved by <strong>{abs(pfm)} days</strong>.")
    elif pfm == 0:
        changes.append("The project finish date is unchanged between revisions.")

    if m["n_delayed"] > 0:
        changes.append(
            f"<strong>{m['n_delayed']}</strong> activities have a later finish date in the current programme."
        )
    if m["n_improved"] > 0:
        changes.append(f"<strong>{m['n_improved']}</strong> activities have an earlier finish date.")
    if m["n_added"] > 0:
        changes.append(f"<strong>{m['n_added']}</strong> new activities have been added.")
    if m["n_removed"] > 0:
        changes.append(f"<strong>{m['n_removed']}</strong> activities have been removed.")

    if changes:
        items.append((
            "What changed?",
            " ".join(changes),
            "#1e3a5f", "#eff6ff", "#3b82f6"
        ))

    # -- Where is the risk? ----------------------------------------------------
    risks = []
    if m["n_became_crit"] > 0:
        risks.append(
            f"<strong>{m['n_became_crit']} activities became critical</strong> in the current revision. "
            "These activities now have zero or negative float and must be monitored closely."
        )
    neg_float_curr = m["merged"]["total_float_days_c"].apply(
        lambda f: safe_float(f, 0) < 0
    ).sum() if "total_float_days_c" in m["merged"].columns else 0
    if neg_float_curr > 0:
        risks.append(
            f"<strong>{int(neg_float_curr)} activities have negative float</strong> in the current programme. "
            "This means the schedule cannot currently achieve its target dates on these paths."
        )
    if m["n_float_lost"] > 0:
        worst_loss = m["float_lost"]["float_move"].min()
        risks.append(
            f"<strong>{m['n_float_lost']} activities lost float</strong>. "
            f"The worst case is a loss of {abs(worst_loss):.1f} days."
        )
    if not m["delayed"].empty:
        worst_slip = m["delayed"]["finish_move"].max()
        worst_act  = m["delayed"].iloc[0].get("task_name_c", m["delayed"].iloc[0].get("task_code",""))
        risks.append(
            f"The biggest finish date slip is <strong>{int(worst_slip)} days</strong> "
            f"({str(worst_act)[:60]})."
        )

    if risks:
        items.append((
            "Where is the risk?",
            " ".join(risks),
            "#7f1d1d", "#fef2f2", "#dc2626"
        ))
    else:
        items.append((
            "Where is the risk?",
            "No significant new risks identified between these two revisions.",
            "#166534", "#f0fdf4", "#16a34a"
        ))

    # -- What needs PM attention? ----------------------------------------------
    attention = []
    if m["n_became_crit"] > 0:
        codes = m["became_crit"]["task_code"].head(3).tolist()
        attention.append(
            f"Review the {m['n_became_crit']} newly critical activities, "
            f"including: {', '.join(str(c) for c in codes)}."
        )
    if neg_float_curr > 0:
        attention.append(
            "Investigate all activities with negative float. "
            "These require a recovery plan or a revised target date."
        )
    if not m["delayed"].empty:
        attention.append(
            "Review the biggest finish date slips in the Biggest Slips section below "
            "and confirm whether recovery actions are in place."
        )
    if m["n_removed"] > 0:
        attention.append(
            f"{m['n_removed']} activities were removed. "
            "Confirm these were intentional scope reductions and not accidental deletions."
        )

    if attention:
        items.append((
            "What needs PM attention?",
            " ".join(attention),
            "#92400e", "#fffbeb", "#F5A623"
        ))

    # -- What improved? --------------------------------------------------------
    good = []
    if m["n_no_longer_crit"] > 0:
        good.append(
            f"<strong>{m['n_no_longer_crit']} activities are no longer critical</strong>, "
            "indicating improved schedule recovery on those paths."
        )
    if m["n_improved"] > 0:
        best = abs(m["improved"]["finish_move"].min())
        good.append(
            f"<strong>{m['n_improved']} activities finished earlier</strong> than the previous revision. "
            f"The best improvement is {int(best)} days."
        )
    if m["n_float_gained"] > 0:
        good.append(f"<strong>{m['n_float_gained']} activities gained float</strong>.")
    if pfm is not None and pfm < 0:
        good.append(
            f"The overall project finish has <strong>improved by {abs(pfm)} days</strong>."
        )

    if good:
        items.append((
            "What improved?",
            " ".join(good),
            "#166534", "#f0fdf4", "#16a34a"
        ))

    return items


def _display_cols(df: pd.DataFrame, col_map: dict) -> pd.DataFrame:
    """Return a display-ready copy with renamed columns and formatted dates."""
    avail = {k: v for k, v in col_map.items() if k in df.columns}
    out = df[list(avail.keys())].copy().rename(columns=avail)
    for col in out.columns:
        if any(kw in col.lower() for kw in ("start","finish","date")):
            try:
                out[col] = out[col].apply(format_date)
            except Exception:
                pass
    return out


def page_programme_comparison():
    """
    Programme Movement Intelligence page.
    Upload two XER files (previous and current) and get a plain-English
    analysis of what changed, where the risk is, and what improved.
    """
    st.title("📅 Movement Intelligence")
    st.caption(
        "Upload the previous and current version of your XER programme to see "
        "a plain-English analysis of what has changed, where the risk is, "
        "and what needs your attention."
    )

    # ---- File uploads --------------------------------------------------------
    up_col1, up_col2 = st.columns(2)
    with up_col1:
        st.markdown(
            '<div style="font-size:12px;font-weight:700;color:#94A3B8;'
            'letter-spacing:1px;text-transform:uppercase;margin-bottom:6px;">'
            'Previous Programme</div>',
            unsafe_allow_html=True,
        )
        prev_file = st.file_uploader("Previous XER", type=["xer"], key="mi_prev_xer",
                                     label_visibility="collapsed")
    with up_col2:
        st.markdown(
            '<div style="font-size:12px;font-weight:700;color:#94A3B8;'
            'letter-spacing:1px;text-transform:uppercase;margin-bottom:6px;">'
            'Current Programme</div>',
            unsafe_allow_html=True,
        )
        curr_file = st.file_uploader("Current XER", type=["xer"], key="mi_curr_xer",
                                     label_visibility="collapsed")

    if not prev_file or not curr_file:
        st.markdown(
            '<div style="background:#f8fafc;border:2px dashed #CBD5E1;border-radius:10px;'
            'padding:32px;text-align:center;color:#64748B;margin-top:20px;">'
            '<div style="font-size:28px;margin-bottom:10px;">📅</div>'
            '<strong style="font-size:15px;">Upload both programmes above to begin</strong><br>'
            '<span style="font-size:13px;margin-top:6px;display:block;">'
            'Upload the previous revision on the left and the current revision on the right.</span>'
            '</div>',
            unsafe_allow_html=True,
        )
        return

    # ---- Parse & cache -------------------------------------------------------
    cache_key = f"mi_{prev_file.name}_{prev_file.size}_{curr_file.name}_{curr_file.size}"
    if st.session_state.get("_mi_cache_key") != cache_key:
        with st.spinner("Parsing both programmes..."):
            try:
                prev_data = parse_xer(prev_file.read())
                curr_data = parse_xer(curr_file.read())
            except Exception as e:
                st.error(f"Could not parse XER files: {e}")
                return
        st.session_state["_mi_prev"] = prev_data
        st.session_state["_mi_curr"] = curr_data
        st.session_state["_mi_cache_key"] = cache_key

    prev_data = st.session_state["_mi_prev"]
    curr_data = st.session_state["_mi_curr"]

    prev_tasks = prev_data["tasks_df"]
    curr_tasks = curr_data["tasks_df"]

    if prev_tasks.empty or curr_tasks.empty:
        st.error("Could not extract activities from one or both files. Check the XER exports.")
        return

    near_crit_days = 10.0

    # ---- Run movement analysis -----------------------------------------------
    m = _compute_movement(prev_tasks, curr_tasks, near_crit_days)

    prev_name = prev_data.get("project_info", {}).get("name", prev_file.name)
    curr_name = curr_data.get("project_info", {}).get("name", curr_file.name)

    # ---- File header strip ---------------------------------------------------
    pfm = m["proj_finish_move"]
    pfm_str = (
        f"+{pfm}d (slipped)" if pfm and pfm > 0 else
        f"{pfm}d (improved)" if pfm and pfm < 0 else
        "No change" if pfm == 0 else "N/A"
    )
    pfm_colour = "#dc2626" if pfm and pfm > 0 else "#16a34a" if pfm and pfm < 0 else "#6b7280"

    st.markdown(
        f"""
        <div style="background:#0B1F33;border-radius:12px;padding:20px 24px;
                    margin-bottom:24px;display:flex;gap:32px;flex-wrap:wrap;
                    align-items:center;">
            <div>
                <div style="font-size:10px;color:#64748B;text-transform:uppercase;
                            letter-spacing:1px;">Previous</div>
                <div style="font-size:14px;font-weight:600;color:#CBD5E1;margin-top:2px;">
                    {prev_name}</div>
                <div style="font-size:11px;color:#475569;">
                    {format_date(m["prev_end"])}</div>
            </div>
            <div style="font-size:24px;color:#F5A623;">&#8594;</div>
            <div>
                <div style="font-size:10px;color:#64748B;text-transform:uppercase;
                            letter-spacing:1px;">Current</div>
                <div style="font-size:14px;font-weight:600;color:#CBD5E1;margin-top:2px;">
                    {curr_name}</div>
                <div style="font-size:11px;color:#475569;">
                    {format_date(m["curr_end"])}</div>
            </div>
            <div style="margin-left:auto;text-align:right;">
                <div style="font-size:10px;color:#64748B;text-transform:uppercase;
                            letter-spacing:1px;">Net Project Finish Movement</div>
                <div style="font-size:28px;font-weight:800;color:{pfm_colour};margin-top:2px;">
                    {pfm_str}</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ---- Summary metric cards ------------------------------------------------
    st.markdown(
        '<div style="font-size:12px;font-weight:700;color:#94A3B8;'
        'letter-spacing:1px;text-transform:uppercase;margin-bottom:10px;">'
        'Summary</div>',
        unsafe_allow_html=True,
    )

    r1 = st.columns(5)
    r1[0].markdown(_mi_card("Activities Added",    m["n_added"],    colour="#16a34a"), unsafe_allow_html=True)
    r1[1].markdown(_mi_card("Activities Removed",  m["n_removed"],  colour="#6b7280"), unsafe_allow_html=True)
    r1[2].markdown(_mi_card("Finish Date Moved",   m["n_delayed"] + m["n_improved"], colour="#0B1F33"), unsafe_allow_html=True)
    r1[3].markdown(_mi_card("Delayed",             m["n_delayed"],  colour="#dc2626"), unsafe_allow_html=True)
    r1[4].markdown(_mi_card("Improved",            m["n_improved"], colour="#16a34a"), unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)
    r2 = st.columns(5)
    r2[0].markdown(_mi_card("Became Critical",     m["n_became_crit"],    colour="#dc2626", bg="#fef2f2", border="#fca5a5"), unsafe_allow_html=True)
    r2[1].markdown(_mi_card("No Longer Critical",  m["n_no_longer_crit"], colour="#16a34a", bg="#f0fdf4", border="#86efac"), unsafe_allow_html=True)
    r2[2].markdown(_mi_card("Lost Float",          m["n_float_lost"],     colour="#d97706", bg="#fffbeb", border="#fcd34d"), unsafe_allow_html=True)
    r2[3].markdown(_mi_card("Gained Float",        m["n_float_gained"],   colour="#16a34a", bg="#f0fdf4", border="#86efac"), unsafe_allow_html=True)
    r2[4].markdown(_mi_card("Net Project Movement", pfm_str,              colour=pfm_colour), unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ---- Written commentary --------------------------------------------------
    st.markdown(
        '<div style="font-size:12px;font-weight:700;color:#94A3B8;'
        'letter-spacing:1px;text-transform:uppercase;margin-bottom:10px;">'
        'Intelligence Summary</div>',
        unsafe_allow_html=True,
    )

    commentary = _build_commentary(m, prev_name, curr_name)
    for heading, body, col, bg, border in commentary:
        st.markdown(_commentary_box(heading, body, col, bg, border), unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # ---- Detailed sections (tabs) --------------------------------------------
    (tab_slips, tab_gains, tab_float_loss, tab_float_gain,
     tab_new_crit, tab_rem_crit, tab_added, tab_removed,
     tab_wbs, tab_export) = st.tabs([
        "Biggest Slips",
        "Biggest Gains",
        "Float Losses",
        "Float Gains",
        "New Critical",
        "No Longer Critical",
        "Added",
        "Removed",
        "WBS Movement",
        "Export",
    ])

    # Helper: common columns for merged common activities
    MOVE_COLS = {
        "task_code":           "Activity ID",
        "task_name_c":         "Activity Name",
        "wbs_path_c":          "WBS",
        "eff_finish_p":        "Previous Finish",
        "eff_finish_c":        "Current Finish",
        "finish_move":         "Finish Movement (d)",
        "total_float_days_p":  "Prev Float (d)",
        "total_float_days_c":  "Curr Float (d)",
        "float_move":          "Float Movement (d)",
    }
    ALT_NAME = {k.replace("_c", "_p"): v for k, v in MOVE_COLS.items()}

    def _safe_display(df, col_map=None):
        if df is None or df.empty:
            return pd.DataFrame(columns=["No data"])
        cols = col_map or MOVE_COLS
        avail = {k: v for k, v in cols.items() if k in df.columns}
        # Fallbacks for name column
        if "task_name_c" not in df.columns and "task_name_p" in df.columns:
            avail["task_name_p"] = "Activity Name"
        out = df[list(avail.keys())].copy().rename(columns=avail)
        for col in ["Previous Finish","Current Finish","Previous Start","Current Start"]:
            if col in out.columns:
                out[col] = out[col].apply(format_date)
        for col in ["Prev Float (d)","Curr Float (d)","Float Movement (d)","Finish Movement (d)"]:
            if col in out.columns:
                out[col] = out[col].apply(lambda x: round(float(x),1) if x is not None and str(x) not in ("","nan") else "-")
        return out

    # -- Biggest slips ---------------------------------------------------------
    with tab_slips:
        st.markdown("**Activities with the largest finish date slips** (most slipped first).")
        df_slips = _safe_display(m["delayed"].head(50))
        if df_slips.empty or "No data" in df_slips.columns:
            st.success("No activities slipped between these two revisions.")
        else:
            # Colour the movement column
            def _red_pos(val):
                try:
                    return "color:#dc2626;font-weight:600;" if float(val) > 0 else ""
                except Exception:
                    return ""
            if "Finish Movement (d)" in df_slips.columns:
                st.dataframe(
                    df_slips.style.applymap(_red_pos, subset=["Finish Movement (d)"]),
                    use_container_width=True, hide_index=True
                )
            else:
                st.dataframe(df_slips, use_container_width=True, hide_index=True)

            if not m["delayed"].empty:
                top20 = m["delayed"].head(20).copy()
                top20 = top20.dropna(subset=["finish_move"])
                label_col = "task_name_c" if "task_name_c" in top20.columns else "task_code"
                top20["Label"] = top20["task_code"].astype(str) + "  " + top20[label_col].astype(str).str[:35]
                fig = px.bar(
                    top20, x="finish_move", y="Label", orientation="h",
                    title="Top 20 Biggest Finish Date Slips (days)",
                    labels={"finish_move":"Slip (days)","Label":""},
                    color_discrete_sequence=["#dc2626"],
                )
                fig.update_yaxes(autorange="reversed")
                fig.update_layout(height=500, margin=dict(l=10,r=10,t=40,b=10),
                                  plot_bgcolor="#ffffff", paper_bgcolor="#ffffff")
                st.plotly_chart(fig, use_container_width=True)

    # -- Biggest gains ---------------------------------------------------------
    with tab_gains:
        st.markdown("**Activities with the largest finish date improvements** (most improved first).")
        df_gains = _safe_display(m["improved"].head(50))
        if df_gains.empty or "No data" in df_gains.columns:
            st.success("No activities improved their finish date between these two revisions.")
        else:
            st.dataframe(df_gains, use_container_width=True, hide_index=True)
            if not m["improved"].empty:
                top20g = m["improved"].head(20).copy().dropna(subset=["finish_move"])
                label_col = "task_name_c" if "task_name_c" in top20g.columns else "task_code"
                top20g["Label"] = top20g["task_code"].astype(str) + "  " + top20g[label_col].astype(str).str[:35]
                top20g["improvement"] = top20g["finish_move"].abs()
                fig2 = px.bar(
                    top20g, x="improvement", y="Label", orientation="h",
                    title="Top 20 Biggest Finish Date Improvements (days)",
                    labels={"improvement":"Improvement (days)","Label":""},
                    color_discrete_sequence=["#16a34a"],
                )
                fig2.update_yaxes(autorange="reversed")
                fig2.update_layout(height=500, margin=dict(l=10,r=10,t=40,b=10),
                                   plot_bgcolor="#ffffff", paper_bgcolor="#ffffff")
                st.plotly_chart(fig2, use_container_width=True)

    # -- Float losses ----------------------------------------------------------
    with tab_float_loss:
        st.markdown("**Activities that lost the most float** -- these are moving towards the critical path.")
        df_fl = _safe_display(m["float_lost"].head(50))
        if df_fl.empty or "No data" in df_fl.columns:
            st.success("No activities lost float between these two revisions.")
        else:
            st.dataframe(df_fl, use_container_width=True, hide_index=True)

    # -- Float gains -----------------------------------------------------------
    with tab_float_gain:
        st.markdown("**Activities that gained the most float** -- these have moved away from the critical path.")
        df_fg = _safe_display(m["float_gained"].head(50))
        if df_fg.empty or "No data" in df_fg.columns:
            st.success("No activities gained float between these two revisions.")
        else:
            st.dataframe(df_fg, use_container_width=True, hide_index=True)

    # -- New critical activities ------------------------------------------------
    with tab_new_crit:
        if m["became_crit"].empty:
            st.success("No activities became critical in the current revision. Good news.")
        else:
            st.warning(
                f"**{len(m['became_crit'])} activities became critical** in the current revision. "
                "These activities now have zero or negative float and require immediate attention."
            )
            df_nc = _safe_display(m["became_crit"])
            st.dataframe(df_nc, use_container_width=True, hide_index=True)

    # -- No longer critical ----------------------------------------------------
    with tab_rem_crit:
        if m["no_longer_crit"].empty:
            st.info("No activities moved off the critical path in the current revision.")
        else:
            st.success(
                f"**{len(m['no_longer_crit'])} activities are no longer critical.** "
                "This indicates recovery on those paths."
            )
            df_rc = _safe_display(m["no_longer_crit"])
            st.dataframe(df_rc, use_container_width=True, hide_index=True)

    # -- Added activities ------------------------------------------------------
    with tab_added:
        if m["added"].empty:
            st.info("No new activities were added in the current revision.")
        else:
            st.info(f"**{len(m['added'])} activities added** to the current programme.")
            added_cols = {
                "task_code":"Activity ID","task_name":"Activity Name",
                "wbs_path":"WBS","eff_start":"Start","eff_finish":"Finish",
                "total_float_days":"Float (d)","status":"Status",
            }
            st.dataframe(_safe_display(m["added"], added_cols), use_container_width=True, hide_index=True)

    # -- Removed activities ----------------------------------------------------
    with tab_removed:
        if m["removed"].empty:
            st.info("No activities were removed in the current revision.")
        else:
            st.warning(
                f"**{len(m['removed'])} activities removed.** "
                "Confirm these are intentional and not accidental deletions."
            )
            rem_cols = {
                "task_code":"Activity ID","task_name":"Activity Name",
                "wbs_path":"WBS","eff_start":"Start","eff_finish":"Finish",
                "total_float_days":"Float (d)","status":"Status",
            }
            st.dataframe(_safe_display(m["removed"], rem_cols), use_container_width=True, hide_index=True)

    # -- WBS-level movement ----------------------------------------------------
    with tab_wbs:
        st.markdown("**Finish date and critical path movement broken down by WBS area.**")
        if m["wbs_summary"].empty:
            st.info("WBS data not available for movement analysis.")
        else:
            wbs_disp = m["wbs_summary"].copy()
            wbs_disp = wbs_disp.rename(columns={
                "wbs_top":           "WBS Area",
                "total_activities":  "Total Activities",
                "delayed_count":     "Delayed",
                "improved_count":    "Improved",
                "avg_finish_move":   "Avg Finish Move (d)",
                "max_finish_slip":   "Max Slip (d)",
                "crit_gained":       "Became Critical",
                "crit_lost":         "No Longer Critical",
            })
            wbs_disp = wbs_disp.sort_values("Max Slip (d)", ascending=False)
            st.dataframe(wbs_disp, use_container_width=True, hide_index=True)

            # Chart: delayed count by WBS
            if not wbs_disp.empty and "Delayed" in wbs_disp.columns:
                fig_wbs = px.bar(
                    wbs_disp[wbs_disp["Delayed"] > 0].head(20),
                    x="Delayed", y="WBS Area", orientation="h",
                    title="Delayed Activities by WBS Area",
                    labels={"Delayed":"Activities Delayed","WBS Area":""},
                    color_discrete_sequence=["#dc2626"],
                )
                fig_wbs.update_yaxes(autorange="reversed")
                fig_wbs.update_layout(height=400, margin=dict(l=10,r=10,t=40,b=10),
                                      plot_bgcolor="#ffffff", paper_bgcolor="#ffffff")
                st.plotly_chart(fig_wbs, use_container_width=True)

            # Chart: average movement by WBS
            wbs_nonzero = wbs_disp[wbs_disp["Avg Finish Move (d)"] != 0].head(20)
            if not wbs_nonzero.empty:
                fig_avg = px.bar(
                    wbs_nonzero,
                    x="Avg Finish Move (d)", y="WBS Area", orientation="h",
                    title="Average Finish Movement by WBS (positive = slipped)",
                    labels={"Avg Finish Move (d)":"Average (days)","WBS Area":""},
                    color="Avg Finish Move (d)",
                    color_continuous_scale=["#16a34a","#f5f5f5","#dc2626"],
                    color_continuous_midpoint=0,
                )
                fig_avg.update_yaxes(autorange="reversed")
                fig_avg.update_layout(height=400, margin=dict(l=10,r=10,t=40,b=10),
                                      plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                                      coloraxis_showscale=False)
                st.plotly_chart(fig_avg, use_container_width=True)

    # -- Export ----------------------------------------------------------------
    with tab_export:
        st.markdown("**Download the full movement analysis as a formatted Excel workbook.**")

        summary_sheet = pd.DataFrame({
            "Metric": [
                "Previous Programme", "Current Programme",
                "Previous Finish", "Current Finish",
                "Net Project Finish Movement",
                "Activities Added", "Activities Removed",
                "Activities Delayed", "Activities Improved",
                "Became Critical", "No Longer Critical",
                "Lost Float", "Gained Float",
            ],
            "Value": [
                prev_name, curr_name,
                format_date(m["prev_end"]), format_date(m["curr_end"]),
                pfm_str,
                m["n_added"], m["n_removed"],
                m["n_delayed"], m["n_improved"],
                m["n_became_crit"], m["n_no_longer_crit"],
                m["n_float_lost"], m["n_float_gained"],
            ],
        })

        commentary_sheet = pd.DataFrame({
            "Section": [h for h, *_ in commentary],
            "Commentary": [b.replace("<strong>","").replace("</strong>","") for _,b,*_ in commentary],
        })

        export_sheets = {
            "Summary":          summary_sheet,
            "Commentary":       commentary_sheet,
            "Biggest Slips":    _safe_display(m["delayed"].head(100)),
            "Biggest Gains":    _safe_display(m["improved"].head(100)),
            "Float Losses":     _safe_display(m["float_lost"].head(100)),
            "Float Gains":      _safe_display(m["float_gained"].head(100)),
            "Became Critical":  _safe_display(m["became_crit"]),
            "No Longer Critical": _safe_display(m["no_longer_crit"]),
            "Added Activities": _safe_display(
                m["added"],
                {"task_code":"Activity ID","task_name":"Activity Name",
                 "wbs_path":"WBS","eff_start":"Start","eff_finish":"Finish",
                 "total_float_days":"Float (d)","status":"Status"},
            ),
            "Removed Activities": _safe_display(
                m["removed"],
                {"task_code":"Activity ID","task_name":"Activity Name",
                 "wbs_path":"WBS","eff_start":"Start","eff_finish":"Finish",
                 "total_float_days":"Float (d)","status":"Status"},
            ),
        }
        if not m["wbs_summary"].empty:
            export_sheets["WBS Movement"] = m["wbs_summary"].rename(columns={
                "wbs_top":"WBS Area","total_activities":"Total Activities",
                "delayed_count":"Delayed","improved_count":"Improved",
                "avg_finish_move":"Avg Finish Move (d)","max_finish_slip":"Max Slip (d)",
                "crit_gained":"Became Critical","crit_lost":"No Longer Critical",
            })

        xls_bytes = export_df_to_excel(export_sheets)

        st.download_button(
            label="📥  Download Movement Intelligence Report",
            data=xls_bytes,
            file_name=f"movement_intelligence_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Exports all movement data across multiple sheets including commentary.",
        )

        st.markdown(
            f"""
            <div style="background:#f8fafc;border:1px solid #E2E8F0;border-radius:8px;
                        padding:14px 18px;margin-top:12px;">
                <div style="font-size:12px;font-weight:700;color:#0B1F33;
                            margin-bottom:8px;">Workbook contents</div>
                <div style="font-size:12px;color:#64748B;line-height:2;">
                    {"".join(f"<strong>{k}</strong> &nbsp;|&nbsp; " for k in export_sheets.keys())}
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )




# -----------------------------------------------------------------------------
# PAGE: PM ACTIONS
# -----------------------------------------------------------------------------

# Risk keywords for notes scanning
_RISK_WORDS = [
    "risk","delay","delayed","delays","blocked","block","constraint","access",
    "design","procurement","client","instruction","CE","EWN","change",
    "hold","pending","late","overrun","issue","dispute","claim",
]

def _generate_actions(
    tasks: pd.DataFrame,
    rels: pd.DataFrame,
    near_crit_days: float,
    notes_text: str = "",
) -> pd.DataFrame:
    """
    Analyse the programme and generate a prioritised PM action list.
    Returns a DataFrame with one row per action.
    """
    rows = []
    now  = datetime.now()
    eight_weeks = now + timedelta(weeks=8)
    four_weeks  = now + timedelta(weeks=4)

    tasks = get_critical_threshold(tasks, near_crit_days)

    # Build predecessor/successor sets from rels
    tasks_with_pred = set()
    tasks_with_succ = set()
    if not rels.empty:
        if "succ_task_id" in rels.columns:
            tasks_with_pred = set(rels["succ_task_id"].dropna())
        if "pred_task_id" in rels.columns:
            tasks_with_succ = set(rels["pred_task_id"].dropna())

    def _wbs(row):
        wbs = row.get("wbs_path", "") if "wbs_path" in row.index else ""
        return str(wbs).split(" > ")[0] if wbs and str(wbs).strip() not in ("","nan") else "-"

    def _row(priority, category, task_code, task_name, wbs, issue, why, action):
        return {
            "Priority":         priority,
            "Category":         category,
            "Activity ID":      str(task_code),
            "Activity Name":    str(task_name),
            "WBS":              str(wbs),
            "Issue":            str(issue),
            "Why It Matters":   str(why),
            "Suggested Action": str(action),
            "Owner":            "",
            "Due Date":         "",
            "Status":           "Open",
        }

    # -- 1. Negative float -----------------------------------------------------
    neg = tasks[tasks["total_float_days"].apply(lambda f: safe_float(f,0) < 0)] \
        if "total_float_days" in tasks.columns else pd.DataFrame()
    for _, t in neg.iterrows():
        tf = round(safe_float(t.get("total_float_days"), 0), 1)
        rows.append(_row(
            "High", "Negative Float",
            t.get("task_code",""), t.get("task_name",""), _wbs(t),
            f"Activity has {tf} days negative float.",
            "Negative float means the current schedule cannot meet its target date on this path. "
            "Every day of inaction increases the delay.",
            f"Investigate the cause of the {abs(tf)}-day overrun. "
            "Agree a recovery plan with the planner and update the programme.",
        ))

    # -- 2. Critical activities not started ------------------------------------
    crit_ns = tasks[
        tasks["is_critical"] &
        tasks["status"].apply(lambda s: str(s) in ("TK_NotStart","Not Started"))
    ] if "status" in tasks.columns and "is_critical" in tasks.columns else pd.DataFrame()
    for _, t in crit_ns.iterrows():
        finish = t.get("eff_finish")
        finish_str = format_date(finish) if finish else "unknown"
        rows.append(_row(
            "High", "Critical Not Started",
            t.get("task_code",""), t.get("task_name",""), _wbs(t),
            f"Critical activity not yet started. Target finish: {finish_str}.",
            "Any delay to a critical activity directly delays the project finish date. "
            "There is no float to absorb slippage.",
            "Confirm start date with the responsible party. "
            "If start is at risk, escalate immediately and review recovery options.",
        ))

    # -- 3. Near-critical due within 4 weeks -----------------------------------
    nc_due = tasks[
        tasks["is_near_critical"] &
        tasks["eff_finish"].apply(
            lambda d: d is not None and hasattr(d,"date") and d <= four_weeks
        )
    ] if "eff_finish" in tasks.columns else pd.DataFrame()
    for _, t in nc_due.iterrows():
        tf = round(safe_float(t.get("total_float_days"),0), 1)
        rows.append(_row(
            "High", "Near-Critical Due Soon",
            t.get("task_code",""), t.get("task_name",""), _wbs(t),
            f"Near-critical activity due within 4 weeks. Float: {tf} days.",
            "With only a small float buffer and a near-term finish, any disruption "
            "will push this activity onto the critical path.",
            "Confirm the activity is progressing on schedule. "
            "If at risk, treat as critical and raise with the team now.",
        ))

    # -- 4. Near-critical due within 4-8 weeks --------------------------------
    nc_due_med = tasks[
        tasks["is_near_critical"] &
        tasks["eff_finish"].apply(
            lambda d: d is not None and hasattr(d,"date") and four_weeks < d <= eight_weeks
        )
    ] if "eff_finish" in tasks.columns else pd.DataFrame()
    for _, t in nc_due_med.iterrows():
        tf = round(safe_float(t.get("total_float_days"),0), 1)
        rows.append(_row(
            "Medium", "Near-Critical Due Soon",
            t.get("task_code",""), t.get("task_name",""), _wbs(t),
            f"Near-critical activity due in 4-8 weeks. Float: {tf} days.",
            "Limited float with an upcoming finish date. Monitor closely.",
            "Review progress and confirm no blockers. Add to weekly look-ahead.",
        ))

    # -- 5. No predecessor (open start) ----------------------------------------
    no_pred = tasks[~tasks["task_id"].isin(tasks_with_pred)] \
        if "task_id" in tasks.columns else pd.DataFrame()
    # Exclude milestones with no predecessors intentionally (start milestones)
    no_pred = no_pred[no_pred.get("task_type","").apply(
        lambda t: "Milestone" not in str(t) and "LOE" not in str(t)
    )] if "task_type" in no_pred.columns else no_pred
    for _, t in no_pred.head(20).iterrows():
        rows.append(_row(
            "Medium", "Open Logic - No Predecessor",
            t.get("task_code",""), t.get("task_name",""), _wbs(t),
            "Activity has no predecessor. Logic is open at the start.",
            "Open-start activities are not driven by schedule logic. "
            "Float calculations may be unreliable for this activity.",
            "Review with the planner. Add a predecessor or confirm the open start is intentional. "
            "If intentional, add a Start On or After constraint.",
        ))

    # -- 6. No successor (open finish) -----------------------------------------
    no_succ = tasks[~tasks["task_id"].isin(tasks_with_succ)] \
        if "task_id" in tasks.columns else pd.DataFrame()
    no_succ = no_succ[no_succ.get("task_type","").apply(
        lambda t: "Finish Milestone" not in str(t) and "LOE" not in str(t)
    )] if "task_type" in no_succ.columns else no_succ
    for _, t in no_succ.head(20).iterrows():
        rows.append(_row(
            "Medium", "Open Logic - No Successor",
            t.get("task_code",""), t.get("task_name",""), _wbs(t),
            "Activity has no successor. Logic is open at the finish.",
            "Open-finish activities may have artificially high float and can mask "
            "schedule risk downstream.",
            "Review with the planner. Add a successor or confirm the open finish "
            "is a deliberate programme end point.",
        ))

    # -- 7. Excessive lag (> 10 days) ------------------------------------------
    if not rels.empty and "lag_days" in rels.columns:
        big_lag = rels[rels["lag_days"].apply(lambda l: safe_float(l,0) > 10)]
        for _, r in big_lag.head(15).iterrows():
            code = r.get("succ_task_code", r.get("succ_task_id",""))
            name = r.get("succ_task_name","")
            lag  = safe_float(r.get("lag_days",0),0)
            pred = r.get("pred_task_code", r.get("pred_task_id",""))
            # Get WBS from tasks
            match = tasks[tasks["task_code"] == code]
            wbs   = _wbs(match.iloc[0]) if not match.empty else "-"
            rows.append(_row(
                "Medium", "Excessive Lag",
                code, name, wbs,
                f"Relationship from {pred} has {int(lag)} days lag.",
                "Excessive lag hides schedule risk. It should be replaced with properly "
                "sequenced activities so the plan reflects real work.",
                f"Challenge the {int(lag)}-day lag with the planner. "
                "Replace with an intermediate activity or reduce the lag to the minimum justified.",
            ))

    # -- 8. Long duration activities (> 60 working days) -----------------------
    long_dur = tasks[tasks["orig_dur_days"].apply(lambda d: safe_float(d,0) > 60)] \
        if "orig_dur_days" in tasks.columns else pd.DataFrame()
    for _, t in long_dur.head(15).iterrows():
        dur = int(safe_float(t.get("orig_dur_days",0),0))
        rows.append(_row(
            "Low", "Long Duration",
            t.get("task_code",""), t.get("task_name",""), _wbs(t),
            f"Activity duration is {dur} working days.",
            "Activities longer than 60 days are difficult to manage and monitor. "
            "Progress is hard to measure and problems can go undetected for weeks.",
            f"Review whether the {dur}-day activity can be broken into smaller "
            "work packages of 20-30 days. Discuss with the planner.",
        ))

    # -- 9. Activities flagged in planning notes --------------------------------
    if notes_text and "task_code" in tasks.columns:
        for word in _RISK_WORDS:
            pattern = re.compile(r"\b" + re.escape(word) + r"\b", re.IGNORECASE)
            if not pattern.search(notes_text):
                continue
            # Find activity IDs mentioned near this risk word in the notes
            for _, t in tasks.iterrows():
                code = str(t.get("task_code",""))
                if code and code in notes_text:
                    # Check the risk word appears within 200 chars of the activity ID
                    idx = notes_text.find(code)
                    snippet = notes_text[max(0,idx-200):idx+200]
                    if pattern.search(snippet):
                        rows.append(_row(
                            "High", "Planning Notes Risk",
                            code, t.get("task_name",""), _wbs(t),
                            f"Activity mentioned in planning notes alongside the word '{word}'.",
                            "Planning notes flag a potential issue against this activity that "
                            "may not be visible in the programme alone.",
                            f"Review the planning notes for {code} and confirm the '{word}' "
                            "item has been actioned or is being tracked.",
                        ))
                        break   # one action per risk word per activity

    # -- 10. Activities with constraints ---------------------------------------
    if "cstr_type" in tasks.columns:
        constrained = tasks[tasks["cstr_type"].apply(
            lambda x: bool(x) and str(x).strip() not in ("","None","nan")
        )]
        for _, t in constrained.head(20).iterrows():
            cstr = str(t.get("cstr_type",""))
            rows.append(_row(
                "Low", "Constraint",
                t.get("task_code",""), t.get("task_name",""), _wbs(t),
                f"Activity has a constraint: {cstr}.",
                "Constraints override schedule logic and can cause artificial float "
                "or negative float. Each constraint should be justified.",
                f"Confirm the {cstr} constraint is still valid. "
                "If the constraint is no longer required, ask the planner to remove it.",
            ))

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)

    # Deduplicate: same Activity ID + Category
    df = df.drop_duplicates(subset=["Activity ID","Category"]).reset_index(drop=True)

    # Sort: High first, then Medium, then Low; within each by Category
    priority_order = {"High": 0, "Medium": 1, "Low": 2}
    df["_sort"] = df["Priority"].map(priority_order)
    df = df.sort_values(["_sort","Category","Activity ID"]).drop(columns=["_sort"]).reset_index(drop=True)

    return df


def page_pm_actions(data: dict, near_crit_days: float):
    """
    PM Action Dashboard.
    Automatically generates a prioritised action list from the uploaded programme.
    Each action has priority, category, issue, why it matters, suggested action,
    and editable owner / due date / status fields.
    """
    st.title("📋 PM Actions")
    st.caption(
        "Automatically generated from your programme. "
        "Each action is prioritised and includes a suggested next step. "
        "Update the Owner, Due Date and Status columns to track progress."
    )

    tasks = data["tasks_df"]
    rels  = data["relationships_df"]

    if tasks.empty:
        st.warning("No activities found. Please upload a programme first.")
        return

    # Retrieve any planning notes text from session state (set by Planning Notes page)
    notes_text = st.session_state.get("_notes_text", "")

    # ---- Generate / cache actions -------------------------------------------
    prog_key = st.session_state.get("_xer_cache_key", "")
    cache_key = f"_pm_actions_{prog_key}_{near_crit_days}"

    if st.session_state.get("_pm_actions_key") != cache_key:
        with st.spinner("Analysing programme..."):
            actions_df = _generate_actions(tasks, rels, near_crit_days, notes_text)
        st.session_state["_pm_actions_df"]  = actions_df
        st.session_state["_pm_actions_key"] = cache_key
    else:
        actions_df = st.session_state["_pm_actions_df"]

    if actions_df.empty:
        st.success(
            "No actions generated. The programme has no obvious issues detected by "
            "the automated checks. Review the Health Check page for more detail."
        )
        return

    # ---- Summary banner ------------------------------------------------------
    n_high   = int((actions_df["Priority"] == "High").sum())
    n_med    = int((actions_df["Priority"] == "Medium").sum())
    n_low    = int((actions_df["Priority"] == "Low").sum())
    n_total  = len(actions_df)

    st.markdown(
        f"""
        <div style="background:#0B1F33;border-radius:12px;padding:20px 24px;
                    margin-bottom:20px;display:flex;gap:24px;flex-wrap:wrap;
                    align-items:center;">
            <div>
                <div style="font-size:11px;color:#64748B;text-transform:uppercase;
                            letter-spacing:1px;">Total Actions</div>
                <div style="font-size:32px;font-weight:800;color:#F5A623;
                            line-height:1;margin-top:4px;">{n_total}</div>
            </div>
            <div style="width:1px;background:#1e3a5f;align-self:stretch;"></div>
            <div>
                <div style="font-size:11px;color:#64748B;text-transform:uppercase;
                            letter-spacing:1px;">High Priority</div>
                <div style="font-size:28px;font-weight:800;color:#dc2626;
                            line-height:1;margin-top:4px;">{n_high}</div>
            </div>
            <div>
                <div style="font-size:11px;color:#64748B;text-transform:uppercase;
                            letter-spacing:1px;">Medium</div>
                <div style="font-size:28px;font-weight:800;color:#d97706;
                            line-height:1;margin-top:4px;">{n_med}</div>
            </div>
            <div>
                <div style="font-size:11px;color:#64748B;text-transform:uppercase;
                            letter-spacing:1px;">Low</div>
                <div style="font-size:28px;font-weight:800;color:#64748B;
                            line-height:1;margin-top:4px;">{n_low}</div>
            </div>
            <div style="margin-left:auto;">
                <div style="font-size:11px;color:#475569;line-height:1.6;">
                    Generated from negative float, critical path,<br>
                    open logic, lag, duration and planning notes.
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ---- Category breakdown cards -------------------------------------------
    cats = actions_df.groupby("Category").size().reset_index(name="n")
    cols = st.columns(min(len(cats), 5))
    cat_colours = {
        "Negative Float":           "#dc2626",
        "Critical Not Started":     "#b91c1c",
        "Near-Critical Due Soon":   "#d97706",
        "Open Logic - No Predecessor": "#7c3aed",
        "Open Logic - No Successor":   "#6d28d9",
        "Excessive Lag":            "#0369a1",
        "Long Duration":            "#0891b2",
        "Planning Notes Risk":      "#dc2626",
        "Constraint":               "#475569",
    }
    for i, (_, cat_row) in enumerate(cats.iterrows()):
        col = cols[i % len(cols)]
        colour = cat_colours.get(cat_row["Category"], "#374151")
        col.markdown(
            f'<div style="background:#ffffff;border:1px solid #E2E8F0;border-radius:8px;'
            f'padding:12px 14px;border-top:3px solid {colour};margin-bottom:6px;">'
            f'<div style="font-size:11px;color:#94A3B8;text-transform:uppercase;'
            f'letter-spacing:0.8px;margin-bottom:4px;">{cat_row["Category"]}</div>'
            f'<div style="font-size:22px;font-weight:800;color:{colour};">{cat_row["n"]}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # ---- Filters ------------------------------------------------------------
    with st.expander("Filter actions", expanded=False):
        fc1, fc2, fc3, fc4, fc5 = st.columns(5)

        all_priorities = ["All"] + sorted(actions_df["Priority"].unique().tolist(),
                          key=lambda p: {"High":0,"Medium":1,"Low":2}.get(p,9))
        f_priority = fc1.selectbox("Priority",  all_priorities, key="pma_f_pri")

        all_cats = ["All"] + sorted(actions_df["Category"].unique().tolist())
        f_cat    = fc2.selectbox("Category",  all_cats, key="pma_f_cat")

        all_wbs = ["All"] + sorted(actions_df["WBS"].unique().tolist())
        f_wbs   = fc3.selectbox("WBS",        all_wbs, key="pma_f_wbs")

        all_owners = ["All"] + sorted([o for o in actions_df["Owner"].unique() if o])
        f_owner  = fc4.selectbox("Owner",      all_owners, key="pma_f_own")

        all_status = ["All", "Open", "In Progress", "Closed"]
        f_status = fc5.selectbox("Status",     all_status, key="pma_f_sta")

    # Apply filters
    display_df = actions_df.copy()
    if f_priority != "All":
        display_df = display_df[display_df["Priority"] == f_priority]
    if f_cat != "All":
        display_df = display_df[display_df["Category"] == f_cat]
    if f_wbs != "All":
        display_df = display_df[display_df["WBS"] == f_wbs]
    if f_owner != "All":
        display_df = display_df[display_df["Owner"] == f_owner]
    if f_status != "All":
        display_df = display_df[display_df["Status"] == f_status]

    n_shown = len(display_df)
    st.caption(f"Showing {n_shown} of {n_total} actions.")

    if display_df.empty:
        st.info("No actions match the current filters.")
        return

    # ---- Editable action table ----------------------------------------------
    st.markdown(
        '<div style="font-size:12px;font-weight:700;color:#94A3B8;letter-spacing:1px;'
        'text-transform:uppercase;margin-bottom:8px;">Action List</div>',
        unsafe_allow_html=True,
    )
    st.caption(
        "Edit the Owner, Due Date and Status columns directly in the table below. "
        "Changes are saved while you are on this page."
    )

    # Colour-code Priority column
    def _style_priority(row):
        p = row.get("Priority","")
        base = [""] * len(row)
        idx  = list(row.index)
        try:
            pi = idx.index("Priority")
        except ValueError:
            return base
        colours = {"High":"background-color:#fef2f2;color:#991b1b;font-weight:700;",
                   "Medium":"background-color:#fffbeb;color:#92400e;font-weight:600;",
                   "Low":"background-color:#f8fafc;color:#475569;"}
        base[pi] = colours.get(p,"")
        return base

    # Use st.data_editor for the editable fields
    edit_cols = [
        "Priority","Category","Activity ID","Activity Name","WBS",
        "Issue","Why It Matters","Suggested Action",
        "Owner","Due Date","Status",
    ]
    avail_edit = [c for c in edit_cols if c in display_df.columns]

    edited_df = st.data_editor(
        display_df[avail_edit],
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        column_config={
            "Priority": st.column_config.SelectboxColumn(
                "Priority",
                options=["High","Medium","Low"],
                width="small",
            ),
            "Status": st.column_config.SelectboxColumn(
                "Status",
                options=["Open","In Progress","Closed"],
                width="small",
            ),
            "Owner": st.column_config.TextColumn("Owner", width="small"),
            "Due Date": st.column_config.TextColumn("Due Date", width="small",
                                                     help="e.g. 01/06/2025"),
            "Issue":            st.column_config.TextColumn("Issue",            width="medium"),
            "Why It Matters":   st.column_config.TextColumn("Why It Matters",   width="large"),
            "Suggested Action": st.column_config.TextColumn("Suggested Action", width="large"),
            "Activity ID":      st.column_config.TextColumn("Activity ID",      width="small"),
            "Activity Name":    st.column_config.TextColumn("Activity Name",    width="medium"),
            "WBS":              st.column_config.TextColumn("WBS",              width="medium"),
            "Category":         st.column_config.TextColumn("Category",         width="medium"),
        },
        key="pma_editor",
    )

    # Persist edits back to session state
    if edited_df is not None and not edited_df.empty:
        for col in ["Owner","Due Date","Status"]:
            if col in edited_df.columns:
                actions_df.loc[display_df.index, col] = edited_df[col].values
        st.session_state["_pm_actions_df"] = actions_df

    # ---- Highlighted high-priority section ----------------------------------
    high_df = display_df[display_df["Priority"] == "High"]
    if not high_df.empty:
        st.divider()
        st.markdown(
            f'<div style="font-size:12px;font-weight:700;color:#dc2626;letter-spacing:1px;'
            f'text-transform:uppercase;margin-bottom:12px;">'
            f'High Priority Actions ({len(high_df)})</div>',
            unsafe_allow_html=True,
        )
        for _, act in high_df.iterrows():
            st.markdown(
                f"""
                <div style="background:#ffffff;border:1px solid #fca5a5;border-left:4px solid #dc2626;
                            border-radius:8px;padding:14px 18px;margin-bottom:8px;">
                    <div style="display:flex;justify-content:space-between;align-items:flex-start;
                                flex-wrap:wrap;gap:8px;">
                        <div style="flex:1;min-width:200px;">
                            <span style="background:#dc2626;color:white;padding:2px 8px;
                                         border-radius:4px;font-size:10px;font-weight:700;
                                         letter-spacing:0.5px;text-transform:uppercase;">
                                {act.get("Category","")}
                            </span>
                            <div style="font-weight:700;color:#0B1F33;font-size:14px;
                                        margin-top:6px;">
                                {act.get("Activity ID","")} - {act.get("Activity Name","")}
                            </div>
                            <div style="font-size:12px;color:#64748B;margin-top:2px;">
                                {act.get("WBS","")}
                            </div>
                        </div>
                    </div>
                    <div style="margin-top:10px;display:grid;grid-template-columns:1fr 1fr;
                                gap:10px;">
                        <div>
                            <div style="font-size:10px;font-weight:700;color:#94A3B8;
                                        text-transform:uppercase;letter-spacing:0.8px;">Issue</div>
                            <div style="font-size:13px;color:#334155;margin-top:3px;">
                                {act.get("Issue","")}</div>
                        </div>
                        <div>
                            <div style="font-size:10px;font-weight:700;color:#94A3B8;
                                        text-transform:uppercase;letter-spacing:0.8px;">
                                Suggested Action</div>
                            <div style="font-size:13px;color:#334155;margin-top:3px;">
                                {act.get("Suggested Action","")}</div>
                        </div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    # ---- Export -------------------------------------------------------------
    st.divider()

    # Merge any edits back for export
    export_df = st.session_state.get("_pm_actions_df", actions_df).copy()

    # Add summary sheet
    summary_rows = {
        "Metric": ["Total Actions","High Priority","Medium Priority","Low Priority",
                   "Open","In Progress","Closed"],
        "Count":  [
            len(export_df),
            int((export_df["Priority"]=="High").sum()),
            int((export_df["Priority"]=="Medium").sum()),
            int((export_df["Priority"]=="Low").sum()),
            int((export_df["Status"]=="Open").sum()),
            int((export_df["Status"]=="In Progress").sum()),
            int((export_df["Status"]=="Closed").sum()),
        ],
    }

    export_sheets = {
        "Summary":           pd.DataFrame(summary_rows),
        "All Actions":       export_df,
        "High Priority":     export_df[export_df["Priority"]=="High"],
        "Medium Priority":   export_df[export_df["Priority"]=="Medium"],
        "Low Priority":      export_df[export_df["Priority"]=="Low"],
    }
    # Category sheets
    for cat in export_df["Category"].unique():
        safe_name = cat[:31]
        export_sheets[safe_name] = export_df[export_df["Category"]==cat]

    xls_bytes = export_df_to_excel(export_sheets)

    dl_col, _ = st.columns([1,3])
    dl_col.download_button(
        label="📥  Export Action List to Excel",
        data=xls_bytes,
        file_name=f"pm_actions_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        help="Exports all actions with Summary, High/Medium/Low priority sheets and one sheet per category.",
    )

    st.caption(
        "Tip: Export the action list and share it with your delivery team. "
        "Update Owner and Status as actions are completed."
    )



# -----------------------------------------------------------------------------
# PAGE: LOOKAHEAD PLANNER
# -----------------------------------------------------------------------------

def _lookahead_card(label: str, value, sublabel: str = "",
                    colour: str = "#0B1F33", bg: str = "#ffffff",
                    border_top: str = "#0B1F33") -> str:
    """Compact metric card matching PlanTrace brand."""
    return (
        f'<div style="background:{bg};border:1px solid #E2E8F0;border-radius:10px;'
        f'padding:16px 18px;border-top:3px solid {border_top};'
        f'box-shadow:0 1px 4px rgba(11,31,51,0.06);">'
        f'<div style="font-size:10px;font-weight:700;color:#94A3B8;letter-spacing:1px;'
        f'text-transform:uppercase;margin-bottom:6px;">{label}</div>'
        f'<div style="font-size:26px;font-weight:800;color:{colour};line-height:1;">{value}</div>'
        f'{"" if not sublabel else f"<div style=font-size:11px;color:#64748B;margin-top:4px;>{sublabel}</div>"}'
        f'</div>'
    )


def _lookahead_section_header(title: str, count: int, colour: str = "#0B1F33") -> str:
    return (
        f'<div style="display:flex;align-items:center;gap:10px;margin:16px 0 8px 0;">'
        f'<div style="font-size:14px;font-weight:700;color:{colour};">{title}</div>'
        f'<div style="background:{colour};color:white;border-radius:12px;'
        f'padding:1px 9px;font-size:11px;font-weight:700;">{count}</div>'
        f'</div>'
    )


def _week_labour(task_res_df: pd.DataFrame, tasks_df: pd.DataFrame,
                 window_start: datetime, window_end: datetime) -> pd.DataFrame:
    """
    Expand resource loading into a weekly series within the lookahead window.
    Returns a DataFrame with columns: week, rsrc_name, qty.
    """
    if task_res_df.empty:
        return pd.DataFrame()

    rows = []
    for _, r in task_res_df.iterrows():
        s = pd.to_datetime(r.get("target_start") or r.get("target_start_date"), errors="coerce")
        e = pd.to_datetime(r.get("target_finish") or r.get("target_end_date"), errors="coerce")
        if pd.isna(s) or pd.isna(e) or s > e:
            continue

        # Clip to window
        s_clip = max(s, pd.Timestamp(window_start))
        e_clip = min(e, pd.Timestamp(window_end))
        if s_clip > e_clip:
            continue

        qty = safe_float(r.get("target_qty", 0), 0)
        if qty == 0:
            continue

        total_days = max(1, (e - s).days)
        window_days = max(1, (e_clip - s_clip).days)
        qty_in_window = qty * (window_days / total_days)

        weeks = max(1, math.ceil(window_days / 7))
        qty_per_week = qty_in_window / weeks

        current = s_clip
        for _ in range(weeks):
            rows.append({
                "week":      current.to_period("W").start_time,
                "rsrc_name": str(r.get("rsrc_name", r.get("rsrc_id", "Unknown"))),
                "qty":       qty_per_week,
            })
            current += timedelta(weeks=1)

    return pd.DataFrame(rows) if rows else pd.DataFrame()


def page_lookahead(data: dict, near_crit_days: float):
    """
    Lookahead Planner page.
    Shows all activities starting, finishing, or at risk within a user-defined
    lookahead window, with labour demand and export.
    """
    st.title("📅 Lookahead Planner")
    st.caption(
        "A short-term delivery view from your programme. "
        "Select a lookahead window to see what is starting, finishing and at risk."
    )

    tasks    = data["tasks_df"]
    rels     = data["relationships_df"]
    task_res = data.get("task_resources_df", pd.DataFrame())

    if tasks.empty:
        st.warning("No activities found. Please upload a programme first.")
        return

    tasks = get_critical_threshold(tasks, near_crit_days)

    # ---- Lookahead window selector ------------------------------------------
    st.markdown(
        '<div style="font-size:12px;font-weight:700;color:#94A3B8;letter-spacing:1px;'
        'text-transform:uppercase;margin-bottom:8px;">Lookahead Window</div>',
        unsafe_allow_html=True,
    )

    wcol1, wcol2 = st.columns([2, 3])

    with wcol1:
        window_opt = st.selectbox(
            "Window",
            ["2 Weeks", "4 Weeks", "6 Weeks", "12 Weeks", "Custom Date Range"],
            index=1,
            label_visibility="collapsed",
            key="la_window_opt",
        )

    now = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)

    window_map = {
        "2 Weeks":  14,
        "4 Weeks":  28,
        "6 Weeks":  42,
        "12 Weeks": 84,
    }

    with wcol2:
        if window_opt == "Custom Date Range":
            dc1, dc2 = st.columns(2)
            window_start = dc1.date_input(
                "From", value=now.date(), key="la_custom_start"
            )
            window_end = dc2.date_input(
                "To", value=(now + timedelta(weeks=4)).date(), key="la_custom_end"
            )
            window_start = datetime.combine(window_start, datetime.min.time())
            window_end   = datetime.combine(window_end,   datetime.max.time().replace(microsecond=0))
        else:
            days = window_map[window_opt]
            window_start = now
            window_end   = now + timedelta(days=days)
            st.markdown(
                f'<div style="padding:8px 14px;background:#0B1F33;border-radius:8px;'
                f'font-size:13px;color:#CBD5E1;display:inline-block;">'
                f'{window_start.strftime("%d %b %Y")} &rarr; '
                f'{window_end.strftime("%d %b %Y")}'
                f'</div>',
                unsafe_allow_html=True,
            )

    st.markdown("<br>", unsafe_allow_html=True)

    # ---- Filters ------------------------------------------------------------
    with st.expander("Filters", expanded=False):
        fl1, fl2, fl3, fl4 = st.columns(4)

        wbs_opts = ["All"]
        if "wbs_path" in tasks.columns:
            tops = tasks["wbs_path"].dropna().apply(
                lambda x: str(x).split(" > ")[0] if x and str(x).strip() not in ("","nan") else "Unknown"
            ).unique().tolist()
            wbs_opts += sorted(tops)
        f_wbs = fl1.selectbox("WBS / Area", wbs_opts, key="la_f_wbs")

        status_opts = ["All"] + sorted([
            s for s in tasks["status"].dropna().unique() if str(s).strip()
        ]) if "status" in tasks.columns else ["All"]
        f_status = fl2.selectbox("Activity Status", status_opts, key="la_f_stat")

        f_crit = fl3.selectbox(
            "Float Filter",
            ["All", "Critical only", "Near-critical only", "Negative float only"],
            key="la_f_crit",
        )

        f_type = fl4.selectbox(
            "Activity Type",
            ["All"] + (sorted([t for t in tasks["task_type"].dropna().unique() if str(t).strip()])
                       if "task_type" in tasks.columns else []),
            key="la_f_type",
        )

    # ---- Build lookahead activity sets --------------------------------------
    def _in_window(d, strict_start=False):
        """True if datetime d falls within the lookahead window."""
        if d is None or not hasattr(d, "date"):
            return False
        if strict_start:
            return window_start <= d <= window_end
        return d <= window_end

    # Activities starting in window
    starting = tasks[
        tasks["eff_start"].apply(lambda d: _in_window(d, strict_start=True))
    ].copy() if "eff_start" in tasks.columns else pd.DataFrame()

    # Activities finishing in window
    finishing = tasks[
        tasks["eff_finish"].apply(lambda d: _in_window(d, strict_start=True))
    ].copy() if "eff_finish" in tasks.columns else pd.DataFrame()

    # Activities spanning the window (started before, finish after window start)
    spanning = tasks[
        tasks["eff_start"].apply(lambda d: d is not None and hasattr(d,"date") and d < window_start) &
        tasks["eff_finish"].apply(lambda d: d is not None and hasattr(d,"date") and d >= window_start)
    ].copy() if ("eff_start" in tasks.columns and "eff_finish" in tasks.columns) else pd.DataFrame()

    # All activities in window = union
    in_window_ids = (
        set(starting["task_id"].tolist() if not starting.empty else []) |
        set(finishing["task_id"].tolist() if not finishing.empty else []) |
        set(spanning["task_id"].tolist() if not spanning.empty else [])
    )
    all_window = tasks[tasks["task_id"].isin(in_window_ids)].copy()

    # Apply filters to all_window
    def _apply_filters(df):
        if df.empty:
            return df
        if f_wbs != "All" and "wbs_path" in df.columns:
            df = df[df["wbs_path"].astype(str).str.startswith(f_wbs)]
        if f_status != "All" and "status" in df.columns:
            df = df[df["status"] == f_status]
        if f_crit == "Critical only" and "is_critical" in df.columns:
            df = df[df["is_critical"] == True]
        elif f_crit == "Near-critical only" and "is_near_critical" in df.columns:
            df = df[df["is_near_critical"] == True]
        elif f_crit == "Negative float only" and "total_float_days" in df.columns:
            df = df[df["total_float_days"].apply(lambda f: safe_float(f, 0) < 0)]
        if f_type != "All" and "task_type" in df.columns:
            df = df[df["task_type"] == f_type]
        return df

    starting  = _apply_filters(starting)
    finishing = _apply_filters(finishing)
    spanning  = _apply_filters(spanning)
    all_window = _apply_filters(all_window)

    # Open logic in window
    tasks_with_pred = set(rels["succ_task_id"].dropna()) if not rels.empty and "succ_task_id" in rels.columns else set()
    tasks_with_succ = set(rels["pred_task_id"].dropna()) if not rels.empty and "pred_task_id" in rels.columns else set()

    open_start_win  = all_window[~all_window["task_id"].isin(tasks_with_pred)] if "task_id" in all_window.columns else pd.DataFrame()
    open_finish_win = all_window[~all_window["task_id"].isin(tasks_with_succ)] if "task_id" in all_window.columns else pd.DataFrame()

    # Critical & near-critical in window
    crit_win  = all_window[all_window["is_critical"] == True] if "is_critical" in all_window.columns else pd.DataFrame()
    nc_win    = all_window[all_window["is_near_critical"] == True] if "is_near_critical" in all_window.columns else pd.DataFrame()
    neg_win   = all_window[all_window["total_float_days"].apply(lambda f: safe_float(f,0) < 0)] if "total_float_days" in all_window.columns else pd.DataFrame()

    # Notes matching
    notes_text = st.session_state.get("_notes_text", "")
    notes_ids  = set()
    if notes_text and "task_code" in all_window.columns:
        for code in all_window["task_code"].dropna():
            if str(code) in notes_text:
                notes_ids.add(str(code))
    notes_win = all_window[all_window["task_code"].astype(str).isin(notes_ids)] if notes_ids else pd.DataFrame()

    risk_ids = set()
    if notes_text and "task_code" in all_window.columns:
        for _, t in all_window.iterrows():
            code = str(t.get("task_code",""))
            if not code:
                continue
            idx = notes_text.find(code)
            if idx < 0:
                continue
            snippet = notes_text[max(0,idx-300):idx+300]
            for word in _RISK_WORDS:
                if re.search(r'\b' + re.escape(word) + r'\b', snippet, re.IGNORECASE):
                    risk_ids.add(code)
                    break
    risk_win = all_window[all_window["task_code"].astype(str).isin(risk_ids)] if risk_ids else pd.DataFrame()

    # Labour in window
    labour_weekly = _week_labour(task_res, tasks, window_start, window_end) if not task_res.empty else pd.DataFrame()
    peak_labour   = int(labour_weekly.groupby("week")["qty"].sum().max()) if not labour_weekly.empty else 0

    # ---- Summary metric cards -----------------------------------------------
    open_logic_count = len(open_start_win) + len(open_finish_win)

    st.markdown(
        '<div style="font-size:12px;font-weight:700;color:#94A3B8;letter-spacing:1px;'
        'text-transform:uppercase;margin-bottom:10px;">Window Summary</div>',
        unsafe_allow_html=True,
    )

    r1 = st.columns(6)
    r1[0].markdown(_lookahead_card(
        "Starting",       len(starting),
        f"in {window_opt.lower()}",
        "#2563eb", "#eff6ff", "#2563eb"
    ), unsafe_allow_html=True)
    r1[1].markdown(_lookahead_card(
        "Finishing",      len(finishing),
        f"in {window_opt.lower()}",
        "#0B1F33", "#f0f9ff", "#0B1F33"
    ), unsafe_allow_html=True)
    r1[2].markdown(_lookahead_card(
        "Critical",       len(crit_win),
        "in window",
        "#dc2626", "#fef2f2" if crit_win.empty else "#fef2f2", "#dc2626"
    ), unsafe_allow_html=True)
    r1[3].markdown(_lookahead_card(
        "Negative Float", len(neg_win),
        "in window",
        "#dc2626" if not neg_win.empty else "#64748B",
        "#fef2f2" if not neg_win.empty else "#f8fafc",
        "#dc2626" if not neg_win.empty else "#E2E8F0"
    ), unsafe_allow_html=True)
    r1[4].markdown(_lookahead_card(
        "Labour Peak",
        f"{peak_labour:,}" if peak_labour else "N/A",
        "hrs/week in window",
        "#7c3aed", "#faf5ff", "#7c3aed"
    ), unsafe_allow_html=True)
    r1[5].markdown(_lookahead_card(
        "Open Logic",     open_logic_count,
        "in window",
        "#d97706" if open_logic_count else "#64748B",
        "#fffbeb" if open_logic_count else "#f8fafc",
        "#d97706" if open_logic_count else "#E2E8F0"
    ), unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    if all_window.empty:
        st.info(
            f"No activities found in the selected window "
            f"({window_start.strftime('%d %b %Y')} to {window_end.strftime('%d %b %Y')}). "
            "Try extending the lookahead window or removing filters."
        )
        return

    # ---- Display columns helper ---------------------------------------------
    BASE_COLS = {
        "task_code":        "Activity ID",
        "task_name":        "Activity Name",
        "wbs_path":         "WBS",
        "eff_start":        "Start",
        "eff_finish":       "Finish",
        "orig_dur_days":    "Orig Dur (d)",
        "rem_dur_days":     "Rem Dur (d)",
        "total_float_days": "Float (d)",
        "status":           "Status",
        "is_critical":      "Critical",
    }

    def _fmt_table(df, extra_cols=None):
        cols = dict(BASE_COLS)
        if extra_cols:
            cols.update(extra_cols)
        avail = {k: v for k, v in cols.items() if k in df.columns}
        out = df[list(avail.keys())].copy().rename(columns=avail)
        for col in ["Start","Finish"]:
            if col in out.columns:
                out[col] = out[col].apply(format_date)
        if "Status" in out.columns:
            out["Status"] = out["Status"].apply(_status_label)
        if "Critical" in out.columns:
            out["Critical"] = out["Critical"].apply(lambda x: "Yes" if x else "")
        if "Float (d)" in out.columns:
            out["Float (d)"] = out["Float (d)"].apply(
                lambda x: round(float(x),1) if x is not None and str(x) not in ("","nan") else "-"
            )
        return out

    def _crit_style(df):
        """Apply row colour based on float."""
        def _row_style(row):
            flag = _crit_flag(row.get("Float (d)", None))
            colour_map = {
                "Negative Float": "background-color:#fecaca;",
                "Critical":       "background-color:#fee2e2;",
                "Near-Critical":  "background-color:#fef3c7;",
            }
            style = colour_map.get(flag, "")
            return [style] * len(row)
        try:
            return df.style.apply(_row_style, axis=1)
        except Exception:
            return df

    # ---- Main tabs ----------------------------------------------------------
    (tab_start, tab_finish, tab_crit, tab_nc, tab_neg,
     tab_open, tab_notes, tab_labour, tab_gantt, tab_export) = st.tabs([
        f"Starting ({len(starting)})",
        f"Finishing ({len(finishing)})",
        f"Critical ({len(crit_win)})",
        f"Near-Critical ({len(nc_win)})",
        f"Negative Float ({len(neg_win)})",
        f"Open Logic ({open_logic_count})",
        f"Planning Notes ({len(notes_win) + len(risk_win)})",
        "Labour Demand",
        "Gantt",
        "Export",
    ])

    # -- Starting -------------------------------------------------------------
    with tab_start:
        st.markdown(
            f"Activities with a scheduled start date between "
            f"**{window_start.strftime('%d %b %Y')}** and "
            f"**{window_end.strftime('%d %b %Y')}**."
        )
        if starting.empty:
            st.info("No activities are scheduled to start in this window.")
        else:
            st.dataframe(_crit_style(_fmt_table(starting.sort_values("eff_start"))),
                         use_container_width=True, hide_index=True)

    # -- Finishing -------------------------------------------------------------
    with tab_finish:
        st.markdown(
            f"Activities with a scheduled finish date between "
            f"**{window_start.strftime('%d %b %Y')}** and "
            f"**{window_end.strftime('%d %b %Y')}**."
        )
        if finishing.empty:
            st.info("No activities are scheduled to finish in this window.")
        else:
            st.dataframe(_crit_style(_fmt_table(finishing.sort_values("eff_finish"))),
                         use_container_width=True, hide_index=True)

    # -- Critical --------------------------------------------------------------
    with tab_crit:
        if crit_win.empty:
            st.success("No critical activities fall within this window.")
        else:
            st.warning(
                f"**{len(crit_win)} critical activities** are active or due in this window. "
                "Any delay to these directly impacts the project finish date."
            )
            st.dataframe(_fmt_table(crit_win.sort_values("eff_finish")),
                         use_container_width=True, hide_index=True)

    # -- Near-Critical ---------------------------------------------------------
    with tab_nc:
        if nc_win.empty:
            st.success("No near-critical activities fall within this window.")
        else:
            st.info(
                f"**{len(nc_win)} near-critical activities** are active or due in this window. "
                f"Each has between 0 and {near_crit_days} days float."
            )
            st.dataframe(_fmt_table(nc_win.sort_values("total_float_days")),
                         use_container_width=True, hide_index=True)

    # -- Negative Float --------------------------------------------------------
    with tab_neg:
        if neg_win.empty:
            st.success("No activities with negative float are active in this window.")
        else:
            st.error(
                f"**{len(neg_win)} activities have negative float** in this window. "
                "These activities are already beyond their target dates and need immediate attention."
            )
            st.dataframe(_fmt_table(neg_win.sort_values("total_float_days")),
                         use_container_width=True, hide_index=True)

    # -- Open Logic ------------------------------------------------------------
    with tab_open:
        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown("**Open Start** -- no predecessors")
            if open_start_win.empty:
                st.success("No open-start activities in this window.")
            else:
                st.warning(f"{len(open_start_win)} activities have no predecessor in this window.")
                st.dataframe(_fmt_table(open_start_win), use_container_width=True, hide_index=True)
        with col_b:
            st.markdown("**Open Finish** -- no successors")
            if open_finish_win.empty:
                st.success("No open-finish activities in this window.")
            else:
                st.warning(f"{len(open_finish_win)} activities have no successor in this window.")
                st.dataframe(_fmt_table(open_finish_win), use_container_width=True, hide_index=True)

    # -- Planning Notes --------------------------------------------------------
    with tab_notes:
        if not notes_text:
            st.info(
                "No planning notes loaded. Upload notes on the Planning Notes page "
                "and return here to see which activities in this window are mentioned."
            )
        else:
            col_n1, col_n2 = st.columns(2)
            with col_n1:
                st.markdown("**Activities mentioned in planning notes**")
                if notes_win.empty:
                    st.info("No activities in this window are mentioned in the planning notes.")
                else:
                    st.dataframe(_fmt_table(notes_win), use_container_width=True, hide_index=True)
            with col_n2:
                st.markdown("**Activities with risk keywords in notes**")
                if risk_win.empty:
                    st.info("No risk keywords found against activities in this window.")
                else:
                    st.warning(
                        f"{len(risk_win)} activities in this window appear alongside "
                        "risk keywords (delay, blocked, CE, EWN, constraint etc.) in the planning notes."
                    )
                    st.dataframe(_fmt_table(risk_win), use_container_width=True, hide_index=True)

    # -- Labour Demand ---------------------------------------------------------
    with tab_labour:
        if labour_weekly.empty:
            st.info(
                "No resource loading data found in this XER. "
                "If the programme was resourced in P6, check the export included resources. "
                "You can also upload a resource CSV on the Labour Histogram page."
            )
        else:
            weekly_total = labour_weekly.groupby("week")["qty"].sum().reset_index()
            weekly_total["week_str"] = weekly_total["week"].dt.strftime("%d %b")

            peak_wk = weekly_total.loc[weekly_total["qty"].idxmax()]
            avg_hrs  = weekly_total["qty"].mean()
            total_hrs = weekly_total["qty"].sum()

            lm1, lm2, lm3 = st.columns(3)
            lm1.metric("Total Hours in Window", f"{total_hrs:,.0f}")
            lm2.metric("Peak Week",             f"{peak_wk['qty']:,.0f} hrs ({peak_wk['week_str']})")
            lm3.metric("Average per Week",      f"{avg_hrs:,.0f} hrs")

            fig_lab = px.bar(
                weekly_total, x="week_str", y="qty",
                title=f"Labour Demand by Week ({window_opt})",
                labels={"week_str": "Week", "qty": "Hours"},
                color_discrete_sequence=["#0B1F33"],
            )
            fig_lab.update_layout(
                height=320, margin=dict(l=10,r=10,t=40,b=10),
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            )
            st.plotly_chart(fig_lab, use_container_width=True)

            # By resource
            if "rsrc_name" in labour_weekly.columns:
                by_res = labour_weekly.groupby("rsrc_name")["qty"].sum().reset_index() \
                             .sort_values("qty", ascending=False).head(15)
                fig_res = px.bar(
                    by_res, x="qty", y="rsrc_name", orientation="h",
                    title="Hours by Resource in Window",
                    labels={"qty":"Hours","rsrc_name":""},
                    color_discrete_sequence=["#7c3aed"],
                )
                fig_res.update_yaxes(autorange="reversed")
                fig_res.update_layout(
                    height=350, margin=dict(l=10,r=10,t=40,b=10),
                    plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                )
                st.plotly_chart(fig_res, use_container_width=True)

    # -- Gantt -----------------------------------------------------------------
    with tab_gantt:
        st.markdown(
            f"All activities active in the **{window_opt.lower()}** window, "
            "colour-coded by float status."
        )
        gantt_src = all_window.dropna(subset=["eff_start","eff_finish"]).copy() \
            if "eff_start" in all_window.columns and "eff_finish" in all_window.columns \
            else pd.DataFrame()

        if gantt_src.empty:
            st.info("No date data available for the Gantt chart.")
        else:
            gantt_src["Label"] = (
                gantt_src["task_code"].astype(str) + "  " +
                gantt_src["task_name"].astype(str).str[:40]
            )
            gantt_src["Float Status"] = gantt_src["total_float_days"].apply(
                lambda f: (
                    "Negative Float" if safe_float(f,0) < 0 else
                    "Critical"       if safe_float(f,0) == 0 else
                    "Near-Critical"  if safe_float(f,0) <= near_crit_days else
                    "Has Float"
                )
            )
            gantt_src = gantt_src.sort_values("eff_start")
            max_gantt = 100

            fig_g = px.timeline(
                gantt_src.head(max_gantt),
                x_start="eff_start", x_end="eff_finish", y="Label",
                color="Float Status",
                color_discrete_map={
                    "Negative Float": "#7f1d1d",
                    "Critical":       "#dc2626",
                    "Near-Critical":  "#d97706",
                    "Has Float":      "#2563eb",
                },
                title=f"Lookahead Gantt: {window_opt} ({window_start.strftime('%d %b')} - {window_end.strftime('%d %b %Y')})",
                labels={"Label":""},
            )
            fig_g.update_yaxes(autorange="reversed")

            # Shade the window
            fig_g.add_vrect(
                x0=window_start, x1=window_end,
                fillcolor="#F5A623", opacity=0.06,
                line_width=0, annotation_text="Lookahead Window",
                annotation_position="top left",
                annotation=dict(font_color="#F5A623", font_size=11),
            )
            fig_g.add_vline(
                x=now, line_dash="dash", line_color="#0B1F33",
                annotation_text="Today", annotation_position="top right",
                annotation=dict(font_color="#0B1F33", font_size=11),
            )
            fig_g.update_layout(
                height=max(350, min(900, 55 + len(gantt_src.head(max_gantt)) * 26)),
                margin=dict(l=10,r=10,t=50,b=10),
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                legend_title_text="Float Status",
            )
            st.plotly_chart(fig_g, use_container_width=True)
            if len(gantt_src) > max_gantt:
                st.caption(f"Showing first {max_gantt} of {len(gantt_src)} activities. Apply filters to narrow the view.")

    # -- Export ----------------------------------------------------------------
    with tab_export:
        st.markdown("**Download the full lookahead as a formatted Excel workbook.**")

        window_label = (
            f"{window_opt}: {window_start.strftime('%d %b %Y')} to {window_end.strftime('%d %b %Y')}"
        )

        summary_sheet = pd.DataFrame({
            "Metric": [
                "Lookahead Window", "From", "To",
                "Activities Starting", "Activities Finishing",
                "Activities in Window (total)",
                "Critical", "Near-Critical", "Negative Float",
                "Open Start", "Open Finish",
                "In Planning Notes", "Risk Keywords",
                "Labour Peak (hrs/week)",
            ],
            "Value": [
                window_label,
                window_start.strftime("%d %b %Y"),
                window_end.strftime("%d %b %Y"),
                len(starting), len(finishing), len(all_window),
                len(crit_win), len(nc_win), len(neg_win),
                len(open_start_win), len(open_finish_win),
                len(notes_win), len(risk_win),
                peak_labour if peak_labour else "N/A",
            ],
        })

        def _exp(df):
            if df is None or df.empty:
                return pd.DataFrame(columns=["No data"])
            return _fmt_table(df)

        export_sheets = {
            "Summary":          summary_sheet,
            "Starting":         _exp(starting.sort_values("eff_start")    if not starting.empty  else starting),
            "Finishing":        _exp(finishing.sort_values("eff_finish")   if not finishing.empty else finishing),
            "All in Window":    _exp(all_window.sort_values("eff_start")   if not all_window.empty else all_window),
            "Critical":         _exp(crit_win),
            "Near-Critical":    _exp(nc_win),
            "Negative Float":   _exp(neg_win),
            "Open Start":       _exp(open_start_win),
            "Open Finish":      _exp(open_finish_win),
        }
        if not notes_win.empty:
            export_sheets["In Planning Notes"] = _exp(notes_win)
        if not risk_win.empty:
            export_sheets["Risk Keywords"]     = _exp(risk_win)
        if not labour_weekly.empty:
            lw_exp = labour_weekly.copy()
            lw_exp["week"] = lw_exp["week"].dt.strftime("%d %b %Y")
            export_sheets["Labour by Week"] = lw_exp.groupby(["week","rsrc_name"])["qty"] \
                .sum().reset_index().rename(columns={"week":"Week","rsrc_name":"Resource","qty":"Hours"})

        xls_bytes = export_df_to_excel(export_sheets)

        st.download_button(
            label="📥  Download Lookahead Report",
            data=xls_bytes,
            file_name=f"lookahead_{window_opt.replace(' ','_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Exports summary, starting, finishing, critical, open logic and labour sheets.",
        )

        st.markdown(
            '<div style="background:#f8fafc;border:1px solid #E2E8F0;border-radius:8px;'
            'padding:14px 18px;margin-top:12px;">'
            '<div style="font-size:12px;font-weight:700;color:#0B1F33;margin-bottom:6px;">'
            'Workbook sheets</div>'
            '<div style="font-size:12px;color:#64748B;line-height:2;">'
            + "".join(f"<strong>{k}</strong> &nbsp;|&nbsp; " for k in export_sheets.keys()) +
            '</div></div>',
            unsafe_allow_html=True,
        )




# -----------------------------------------------------------------------------
# PAGE: MILESTONE TRACKER
# -----------------------------------------------------------------------------

# Keywords that indicate a milestone activity by name
_MILESTONE_KEYWORDS = [
    "milestone", "complete", "completion", "handover", "hand over",
    "energisation", "energization", "energise", "energize",
    "access", "installation start", "install start",
    "commissioning", "commission", "delivery", "deliver",
    "approval", "approve", "sign off", "sign-off", "signoff",
    "ready for", "available", "issued", "award", "award of contract",
    "start on site", "mobilisation", "mobilization", "end", "finish",
    "practical completion", "pc", "substantial completion",
    "first fix", "second fix", "inspection", "witness",
    "notice to proceed", "ntp", "key date",
]


def _detect_milestones(tasks: pd.DataFrame) -> pd.DataFrame:
    """
    Auto-detect milestone activities from the tasks DataFrame.
    Returns a copy with a boolean 'is_milestone_detected' column.
    """
    df = tasks.copy()

    flags = pd.Series(False, index=df.index)

    # 1. Activity type contains "milestone" (covers TT_Mile and TT_FinMile)
    if "task_type" in df.columns:
        flags |= df["task_type"].astype(str).str.lower().str.contains(
            "milestone", na=False
        )

    # 2. Zero original duration
    if "orig_dur_days" in df.columns:
        flags |= df["orig_dur_days"].apply(lambda d: safe_float(d, 1) == 0)

    # 3. Name contains a milestone keyword
    if "task_name" in df.columns:
        kw_pattern = "|".join(re.escape(k) for k in _MILESTONE_KEYWORDS)
        flags |= df["task_name"].astype(str).str.lower().str.contains(
            kw_pattern, na=False
        )

    df["is_milestone_detected"] = flags
    return df


def _risk_rating(tf, has_constraint: bool = False, in_notes_risk: bool = False) -> str:
    """Return a simple risk rating string for a milestone."""
    f = safe_float(tf, 9999)
    if f < 0 or in_notes_risk:
        return "High"
    if f == 0 or has_constraint:
        return "High"
    if f <= 10:
        return "Medium"
    return "Low"


def _risk_colour(rating: str) -> str:
    return {"High": "#dc2626", "Medium": "#d97706", "Low": "#16a34a"}.get(rating, "#6b7280")


def _risk_bg(rating: str) -> str:
    return {"High": "#fef2f2", "Medium": "#fffbeb", "Low": "#f0fdf4"}.get(rating, "#f8fafc")


def _milestone_header_card(row: pd.Series, movement_days=None) -> str:
    """Render a single milestone summary card as HTML."""
    tf        = safe_float(row.get("total_float_days"), None)
    f_col     = _float_color(tf)
    is_crit   = bool(row.get("is_critical", False))
    status    = _status_label(str(row.get("status", "")))
    s_col     = _status_colour(str(row.get("status", "")))
    code      = str(row.get("task_code", "-"))
    name      = str(row.get("task_name", "-"))
    wbs       = str(row.get("wbs_path", "-"))
    finish    = format_date(row.get("eff_finish"))
    start     = format_date(row.get("eff_start"))
    rating    = _risk_rating(
        tf,
        has_constraint=bool(row.get("cstr_type","")) and str(row.get("cstr_type","")).strip() not in ("","None","nan"),
    )
    r_col     = _risk_colour(rating)

    crit_pill = (
        '<span style="background:#dc2626;color:white;padding:1px 8px;'
        'border-radius:10px;font-size:10px;font-weight:700;margin-left:6px;">CRITICAL</span>'
        if is_crit else ""
    )

    move_html = ""
    if movement_days is not None:
        m_col  = "#dc2626" if movement_days > 0 else "#16a34a" if movement_days < 0 else "#6b7280"
        m_sign = "+" if movement_days > 0 else ""
        move_html = (
            f'<span style="background:{m_col};color:white;padding:2px 8px;'
            f'border-radius:10px;font-size:10px;font-weight:700;margin-left:6px;">'
            f'{m_sign}{movement_days}d movement</span>'
        )

    return f"""
    <div style="background:#ffffff;border:1px solid #E2E8F0;border-radius:12px;
                padding:18px 22px;margin-bottom:10px;
                border-left:5px solid {r_col};
                box-shadow:0 2px 6px rgba(11,31,51,0.07);">
        <div style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:8px;">
            <div style="flex:1;min-width:0;">
                <div style="font-size:10px;font-weight:700;color:#94A3B8;letter-spacing:1px;
                            text-transform:uppercase;margin-bottom:4px;">Milestone</div>
                <div style="font-size:16px;font-weight:800;color:#0B1F33;">
                    {code}{crit_pill}{move_html}
                </div>
                <div style="font-size:14px;color:#334155;margin-top:3px;
                            white-space:nowrap;overflow:hidden;text-overflow:ellipsis;"
                     title="{name}">{name}</div>
                <div style="font-size:11px;color:#94A3B8;margin-top:2px;">{wbs}</div>
            </div>
            <div style="display:grid;grid-template-columns:repeat(4,auto);gap:10px;align-items:start;">
                <div style="text-align:center;">
                    <div style="font-size:9px;color:#94A3B8;text-transform:uppercase;
                                letter-spacing:0.8px;">Start</div>
                    <div style="font-size:12px;font-weight:600;color:#0B1F33;margin-top:2px;">{start}</div>
                </div>
                <div style="text-align:center;">
                    <div style="font-size:9px;color:#94A3B8;text-transform:uppercase;
                                letter-spacing:0.8px;">Finish</div>
                    <div style="font-size:12px;font-weight:700;color:#0B1F33;margin-top:2px;">{finish}</div>
                </div>
                <div style="text-align:center;">
                    <div style="font-size:9px;color:#94A3B8;text-transform:uppercase;
                                letter-spacing:0.8px;">Float</div>
                    <div style="font-size:14px;font-weight:800;color:{f_col};margin-top:2px;">
                        {(str(tf) + "d") if tf is not None else "-"}</div>
                </div>
                <div style="text-align:center;">
                    <div style="font-size:9px;color:#94A3B8;text-transform:uppercase;
                                letter-spacing:0.8px;">Risk</div>
                    <div style="background:{r_col};color:white;border-radius:8px;
                                padding:2px 10px;font-size:11px;font-weight:700;margin-top:2px;">
                        {rating}</div>
                </div>
            </div>
        </div>
        <div style="margin-top:10px;">
            <span style="background:{s_col};color:white;padding:2px 9px;
                         border-radius:10px;font-size:10px;">{status}</span>
        </div>
    </div>"""


def page_milestone_tracker(data: dict, near_crit_days: float):
    """
    Milestone Tracker page.
    Auto-detects milestones from the XER, allows manual additions,
    and shows driving path, successors, movement, notes and risk per milestone.
    """
    st.title("🏁 Milestone Tracker")
    st.caption(
        "Key programme milestones auto-detected from your XER. "
        "Select any milestone to see what is driving it, what it drives, and its current risk."
    )

    tasks = data["tasks_df"]
    rels  = data["relationships_df"]

    if tasks.empty:
        st.warning("No activities found. Please upload a programme first.")
        return

    tasks = get_critical_threshold(tasks, near_crit_days)
    tasks_with_milestones = _detect_milestones(tasks)

    # Build graph
    G           = build_graph(tasks, rels)
    task_lookup = tasks.set_index("task_id").to_dict("index")

    # ---- Comparison data (finish movement) ----------------------------------
    movement_map = {}
    if "_mi_prev" in st.session_state and "_mi_curr" in st.session_state:
        try:
            prev_t = st.session_state["_mi_prev"]["tasks_df"]
            curr_t = st.session_state["_mi_curr"]["tasks_df"]
            if not prev_t.empty and not curr_t.empty:
                merged = prev_t[["task_code","eff_finish"]].merge(
                    curr_t[["task_code","eff_finish"]], on="task_code",
                    suffixes=("_p","_c"), how="inner"
                )
                for _, r in merged.iterrows():
                    try:
                        p = pd.Timestamp(r["eff_finish_p"])
                        c = pd.Timestamp(r["eff_finish_c"])
                        movement_map[str(r["task_code"])] = int((c - p).days)
                    except Exception:
                        pass
        except Exception:
            pass

    # ---- Notes data ---------------------------------------------------------
    notes_text = st.session_state.get("_notes_text", "")
    notes_ids  = set()
    risk_ids   = set()
    if notes_text and "task_code" in tasks.columns:
        for _, t in tasks.iterrows():
            code = str(t.get("task_code",""))
            if not code or code not in notes_text:
                continue
            notes_ids.add(code)
            idx = notes_text.find(code)
            snippet = notes_text[max(0,idx-300):idx+300]
            for word in _RISK_WORDS:
                if re.search(r'\b' + re.escape(word) + r'\b', snippet, re.IGNORECASE):
                    risk_ids.add(code)
                    break

    # ---- Milestone selection -------------------------------------------------
    auto_milestones = tasks_with_milestones[
        tasks_with_milestones["is_milestone_detected"]
    ].copy()

    # Session state for manually added milestones
    if "ms_manual_ids" not in st.session_state:
        st.session_state["ms_manual_ids"] = set()

    # ---- Sidebar-style controls panel ---------------------------------------
    with st.expander("Configure Milestones", expanded=True):
        cfg1, cfg2 = st.columns([1, 2])

        with cfg1:
            st.markdown(
                '<div style="font-size:11px;font-weight:700;color:#94A3B8;'
                'letter-spacing:1px;text-transform:uppercase;margin-bottom:6px;">'
                'Auto-Detected</div>',
                unsafe_allow_html=True,
            )
            st.caption(
                f"{len(auto_milestones)} milestones detected from activity type, "
                "zero duration, or name keywords."
            )

            show_auto = st.checkbox("Include auto-detected milestones", value=True, key="ms_show_auto")

        with cfg2:
            st.markdown(
                '<div style="font-size:11px;font-weight:700;color:#94A3B8;'
                'letter-spacing:1px;text-transform:uppercase;margin-bottom:6px;">'
                'Add Key Activities Manually</div>',
                unsafe_allow_html=True,
            )
            manual_search = st.text_input(
                "Search to add",
                placeholder="Type Activity ID or name",
                key="ms_manual_search",
                label_visibility="collapsed",
            )
            if manual_search.strip():
                mask = (
                    tasks["task_code"].astype(str).str.contains(manual_search.strip(), case=False, na=False) |
                    tasks["task_name"].astype(str).str.contains(manual_search.strip(), case=False, na=False)
                )
                search_results = tasks[mask]
                if not search_results.empty:
                    add_labels = search_results.apply(
                        lambda r: f"{r.get('task_code','')}  --  {r.get('task_name','')}", axis=1
                    ).tolist()
                    add_sel = st.selectbox("Select to add", add_labels, key="ms_add_sel", label_visibility="collapsed")
                    if st.button("Add as milestone", key="ms_add_btn"):
                        sel_code = add_sel.split("  --  ")[0].strip()
                        match = tasks[tasks["task_code"] == sel_code]
                        if not match.empty:
                            st.session_state["ms_manual_ids"].add(match.iloc[0]["task_id"])
                            st.success(f"Added {sel_code}")
                else:
                    st.caption("No activities match.")

            if st.session_state["ms_manual_ids"]:
                if st.button("Clear manual selections", key="ms_clear"):
                    st.session_state["ms_manual_ids"] = set()

        # Filters
        st.markdown("<hr style='border:none;border-top:1px solid #E2E8F0;margin:10px 0;'>", unsafe_allow_html=True)
        fa, fb, fc = st.columns(3)
        f_risk = fa.selectbox("Risk filter", ["All","High","Medium","Low"], key="ms_f_risk")
        f_crit = fb.selectbox("Float filter", ["All","Critical","Near-Critical","Negative Float"], key="ms_f_crit")
        f_wbs  = fc.text_input("WBS contains", placeholder="e.g. Civil", key="ms_f_wbs")

    # ---- Build final milestone list -----------------------------------------
    milestone_ids = set()
    if show_auto:
        milestone_ids |= set(auto_milestones["task_id"].tolist())
    milestone_ids |= st.session_state["ms_manual_ids"]

    if not milestone_ids:
        st.info(
            "No milestones detected or selected. "
            "Either enable auto-detection above or search for and add activities manually."
        )
        return

    milestones = tasks[tasks["task_id"].isin(milestone_ids)].copy()

    # Apply risk rating
    def _rate(row):
        tf   = safe_float(row.get("total_float_days"), 9999)
        cstr = str(row.get("cstr_type","")).strip() not in ("","None","nan")
        risk = str(row.get("task_code","")) in risk_ids
        return _risk_rating(tf, cstr, risk)

    milestones["risk_rating"] = milestones.apply(_rate, axis=1)

    # Apply filters
    if f_risk != "All":
        milestones = milestones[milestones["risk_rating"] == f_risk]
    if f_crit == "Critical":
        milestones = milestones[milestones["is_critical"] == True]
    elif f_crit == "Near-Critical":
        milestones = milestones[milestones["is_near_critical"] == True]
    elif f_crit == "Negative Float":
        milestones = milestones[milestones["total_float_days"].apply(lambda f: safe_float(f,0) < 0)]
    if f_wbs.strip() and "wbs_path" in milestones.columns:
        milestones = milestones[milestones["wbs_path"].astype(str).str.contains(f_wbs.strip(), case=False, na=False)]

    if milestones.empty:
        st.info("No milestones match the current filters.")
        return

    milestones = milestones.sort_values("eff_finish" if "eff_finish" in milestones.columns else "task_code")

    # ---- Summary metrics ----------------------------------------------------
    n_total  = len(milestones)
    n_high   = int((milestones["risk_rating"] == "High").sum())
    n_med    = int((milestones["risk_rating"] == "Medium").sum())
    n_low    = int((milestones["risk_rating"] == "Low").sum())
    n_crit   = int(milestones["is_critical"].sum()) if "is_critical" in milestones.columns else 0
    n_neg    = int(milestones["total_float_days"].apply(lambda f: safe_float(f,0) < 0).sum()) if "total_float_days" in milestones.columns else 0

    st.markdown(
        f"""
        <div style="background:#0B1F33;border-radius:12px;padding:16px 22px;
                    margin-bottom:18px;display:flex;gap:20px;flex-wrap:wrap;align-items:center;">
            <div>
                <div style="font-size:10px;color:#64748B;text-transform:uppercase;
                            letter-spacing:1px;">Milestones</div>
                <div style="font-size:28px;font-weight:800;color:#F5A623;line-height:1;margin-top:3px;">{n_total}</div>
            </div>
            <div style="width:1px;background:#1e3a5f;align-self:stretch;"></div>
            <div>
                <div style="font-size:10px;color:#64748B;text-transform:uppercase;letter-spacing:1px;">High Risk</div>
                <div style="font-size:24px;font-weight:800;color:#dc2626;line-height:1;margin-top:3px;">{n_high}</div>
            </div>
            <div>
                <div style="font-size:10px;color:#64748B;text-transform:uppercase;letter-spacing:1px;">Medium</div>
                <div style="font-size:24px;font-weight:800;color:#d97706;line-height:1;margin-top:3px;">{n_med}</div>
            </div>
            <div>
                <div style="font-size:10px;color:#64748B;text-transform:uppercase;letter-spacing:1px;">Low</div>
                <div style="font-size:24px;font-weight:800;color:#16a34a;line-height:1;margin-top:3px;">{n_low}</div>
            </div>
            <div style="width:1px;background:#1e3a5f;align-self:stretch;"></div>
            <div>
                <div style="font-size:10px;color:#64748B;text-transform:uppercase;letter-spacing:1px;">Critical</div>
                <div style="font-size:24px;font-weight:800;color:#dc2626;line-height:1;margin-top:3px;">{n_crit}</div>
            </div>
            <div>
                <div style="font-size:10px;color:#64748B;text-transform:uppercase;letter-spacing:1px;">Neg Float</div>
                <div style="font-size:24px;font-weight:800;color:#7f1d1d;line-height:1;margin-top:3px;">{n_neg}</div>
            </div>
            {"" if not movement_map else '<div style="margin-left:auto;font-size:11px;color:#F5A623;">Comparison data loaded</div>'}
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ---- Summary table tab + individual milestone tabs ----------------------
    tab_names = ["Summary Table"] + [
        f"{row.get('task_code','?')}" for _, row in milestones.head(10).iterrows()
    ]
    if len(milestones) > 10:
        tab_names[-1] = f"{tab_names[-1]}..."
    tabs = st.tabs(tab_names)

    # =========================================================================
    # TAB 0: Summary Table
    # =========================================================================
    with tabs[0]:
        st.markdown(
            '<div style="font-size:12px;font-weight:700;color:#94A3B8;letter-spacing:1px;'
            'text-transform:uppercase;margin-bottom:8px;">All Milestones</div>',
            unsafe_allow_html=True,
        )

        table_rows = []
        for _, ms in milestones.iterrows():
            code = str(ms.get("task_code",""))
            tf   = safe_float(ms.get("total_float_days"), None)
            move = movement_map.get(code)
            table_rows.append({
                "Activity ID":      code,
                "Activity Name":    str(ms.get("task_name","")),
                "WBS":              str(ms.get("wbs_path","")).split(" > ")[0],
                "Forecast Start":   format_date(ms.get("eff_start")),
                "Forecast Finish":  format_date(ms.get("eff_finish")),
                "Total Float (d)":  round(tf, 1) if tf is not None else "-",
                "Critical":         "Yes" if ms.get("is_critical") else "",
                "Movement (d)":     move if move is not None else "-",
                "Status":           _status_label(str(ms.get("status",""))),
                "Risk":             ms.get("risk_rating",""),
                "In Notes":         "Yes" if code in notes_ids else "",
                "Risk Keywords":    "Yes" if code in risk_ids else "",
            })

        summary_df = pd.DataFrame(table_rows)

        def _style_summary(row):
            risk = row.get("Risk","")
            colour_map = {
                "High":   "background-color:#fef2f2;",
                "Medium": "background-color:#fffbeb;",
                "Low":    "",
            }
            return [colour_map.get(risk,"")] * len(row)

        st.dataframe(
            summary_df.style.apply(_style_summary, axis=1),
            use_container_width=True, hide_index=True,
        )

        # Timeline chart of all milestones
        gantt_src = milestones.dropna(subset=["eff_finish"]).copy() \
            if "eff_finish" in milestones.columns else pd.DataFrame()
        if not gantt_src.empty:
            gantt_src["Label"] = gantt_src["task_code"].astype(str) + "  " + gantt_src["task_name"].astype(str).str[:40]
            gantt_src["Risk"]  = gantt_src["risk_rating"]
            start_col  = "eff_start"  if "eff_start"  in gantt_src.columns else "eff_finish"
            finish_col = "eff_finish"

            # For zero-duration milestones, give a 1-day bar for visibility
            gantt_src["_plot_start"]  = gantt_src[start_col]
            gantt_src["_plot_finish"] = gantt_src.apply(
                lambda r: r[finish_col] + timedelta(days=1)
                if r.get(start_col) == r.get(finish_col) or pd.isna(r.get(start_col))
                else r[finish_col],
                axis=1,
            )

            fig_ms = px.timeline(
                gantt_src,
                x_start="_plot_start", x_end="_plot_finish", y="Label",
                color="Risk",
                color_discrete_map={
                    "High":   "#dc2626",
                    "Medium": "#d97706",
                    "Low":    "#16a34a",
                },
                title="Milestone Timeline",
                labels={"Label":""},
            )
            fig_ms.update_yaxes(autorange="reversed")
            fig_ms.add_vline(
                x=datetime.now(), line_dash="dash", line_color="#0B1F33",
                annotation_text="Today", annotation_position="top right",
                annotation=dict(font_color="#0B1F33"),
            )
            fig_ms.update_layout(
                height=max(300, min(800, 50 + len(gantt_src) * 28)),
                margin=dict(l=10,r=10,t=50,b=10),
                plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
                legend_title_text="Risk",
            )
            st.plotly_chart(fig_ms, use_container_width=True)

    # =========================================================================
    # TABS 1-N: Individual milestone detail
    # =========================================================================
    for tab_idx, (_, ms) in enumerate(milestones.head(10).iterrows(), start=1):
        if tab_idx >= len(tabs):
            break
        with tabs[tab_idx]:
            ms_code = str(ms.get("task_code",""))
            ms_id   = ms["task_id"]
            move    = movement_map.get(ms_code)

            # Header card
            st.markdown(
                _milestone_header_card(ms, movement_days=move),
                unsafe_allow_html=True,
            )

            # Constraint warning
            if "cstr_type" in ms.index:
                cstr = str(ms.get("cstr_type","")).strip()
                if cstr not in ("","None","nan"):
                    cdate = format_date(ms.get("cstr_date") if "cstr_date" in ms.index else None)
                    st.markdown(
                        f'<div style="background:#fffbeb;border-left:4px solid #F5A623;'
                        f'border-radius:6px;padding:10px 16px;margin-bottom:10px;">'
                        f'<strong>Constraint:</strong> {cstr} &nbsp;|&nbsp; '
                        f'<strong>Date:</strong> {cdate}</div>',
                        unsafe_allow_html=True,
                    )

            # Notes against this milestone
            if ms_code in notes_ids:
                note_flag = "Risk keywords found in notes." if ms_code in risk_ids else "Mentioned in planning notes."
                note_col  = "#dc2626" if ms_code in risk_ids else "#2563eb"
                st.markdown(
                    f'<div style="background:#eff6ff;border-left:4px solid {note_col};'
                    f'border-radius:6px;padding:10px 16px;margin-bottom:10px;">'
                    f'<strong>Planning Notes:</strong> {note_flag}</div>',
                    unsafe_allow_html=True,
                )

            # ---- Action buttons ---------------------------------------------
            ba, bb = st.columns(2)
            btn_drive = ba.button("Find Driving Path",  key=f"ms_drive_{ms_id}", use_container_width=True)
            btn_succ  = bb.button("Show Successors",    key=f"ms_succ_{ms_id}",  use_container_width=True)

            # ---- Driving Path ------------------------------------------------
            if btn_drive:
                with st.spinner("Tracing driving path..."):
                    direct_preds = list(G.predecessors(ms_id))
                    if not direct_preds:
                        st.warning(f"{ms_code} has no predecessors. No driving path can be identified.")
                    else:
                        path = driving_path_to_activity(G, tasks, rels, ms_id)
                        path_rows = []
                        for i, tid in enumerate(path):
                            t  = task_lookup.get(tid, {})
                            tf = t.get("total_float_days")
                            is_tgt = (tid == ms_id)
                            # Relationship to next step
                            rl, lg = "-", 0
                            if i < len(path) - 1:
                                next_id = path[i+1]
                                if not rels.empty:
                                    rel = rels[
                                        (rels.get("pred_task_id", pd.Series(dtype=str)) == tid) &
                                        (rels.get("succ_task_id", pd.Series(dtype=str)) == next_id)
                                    ]
                                    if not rel.empty:
                                        rl = _rel_label(rel["rel_type"].iloc[0] if "rel_type" in rel.columns else "FS")
                                        lg = safe_float(rel["lag_days"].iloc[0] if "lag_days" in rel.columns else 0, 0)
                            path_rows.append({
                                "Step":            i + 1,
                                "Activity ID":     t.get("task_code", tid),
                                "Activity Name":   t.get("task_name",""),
                                "Start":           format_date(t.get("eff_start")),
                                "Finish":          format_date(t.get("eff_finish")),
                                "Float (d)":       round(float(tf),1) if tf is not None else "-",
                                "Link":            rl if not is_tgt else "-",
                                "Lag (d)":         lg if not is_tgt else "-",
                                "Flag":            _crit_flag(tf),
                                "Milestone":       "TARGET" if is_tgt else "",
                            })

                        path_df = pd.DataFrame(path_rows)
                        st.markdown("**Driving Path**")
                        st.caption("Activities ordered from chain start to the milestone target.")

                        def _path_row_style(row):
                            flag = row.get("Flag","")
                            is_t = row.get("Milestone","") == "TARGET"
                            if is_t:
                                return ["background-color:#0B1F33;color:white;font-weight:700;"] * len(row)
                            cm = {"Critical":"background-color:#fee2e2;",
                                  "Negative Float":"background-color:#fecaca;",
                                  "Near-Critical":"background-color:#fef3c7;"}
                            return [cm.get(flag,"")] * len(row)

                        st.dataframe(path_df.style.apply(_path_row_style, axis=1),
                                     use_container_width=True, hide_index=True)

                        # Mini Gantt for path
                        path_task_ids = [p for p in path if p != ms_id]
                        path_tasks    = tasks[tasks["task_id"].isin(path)].copy()
                        gantt_p = path_tasks.dropna(subset=["eff_start","eff_finish"]).copy() \
                            if "eff_start" in path_tasks.columns else pd.DataFrame()
                        if not gantt_p.empty:
                            gantt_p["Label"] = gantt_p["task_code"].astype(str) + "  " + gantt_p["task_name"].astype(str).str[:35]
                            gantt_p["Type"]  = gantt_p["task_id"].apply(lambda t: "Milestone" if t == ms_id else "Driving Path")
                            fig_dp = px.timeline(
                                gantt_p, x_start="eff_start", x_end="eff_finish", y="Label",
                                color="Type",
                                color_discrete_map={"Milestone":"#0B1F33","Driving Path":"#dc2626"},
                                title=f"Driving Path to {ms_code}",
                            )
                            fig_dp.update_yaxes(autorange="reversed")
                            fig_dp.add_vline(x=datetime.now(), line_dash="dot", line_color="#94A3B8",
                                             annotation_text="Today")
                            fig_dp.update_layout(height=max(250, 50+len(gantt_p)*28),
                                                 margin=dict(l=10,r=10,t=40,b=10),
                                                 plot_bgcolor="#ffffff", paper_bgcolor="#ffffff")
                            st.plotly_chart(fig_dp, use_container_width=True)

            # ---- Successors --------------------------------------------------
            if btn_succ:
                direct_succs = list(G.successors(ms_id))
                if not direct_succs:
                    st.warning(f"{ms_code} has no successors.")
                else:
                    all_succs = trace_successors(G, ms_id)
                    succ_df   = _build_full_trace_df(G, rels, task_lookup, ms_id, all_succs, "succ")
                    # Add WBS
                    if not succ_df.empty and "Activity ID" in succ_df.columns:
                        code_to_wbs = tasks.set_index("task_code")["wbs_path"].to_dict() if "wbs_path" in tasks.columns else {}
                        succ_df.insert(succ_df.columns.get_loc("Activity Name")+1, "WBS",
                                       succ_df["Activity ID"].map(code_to_wbs).fillna("-"))

                    st.markdown(f"**All Successors of {ms_code}** ({len(succ_df)} activities)")

                    def _succ_style(val):
                        return {"Critical":"background-color:#fee2e2;color:#991b1b;font-weight:600;",
                                "Negative Float":"background-color:#fecaca;color:#7f1d1d;font-weight:700;",
                                "Near-Critical":"background-color:#fef3c7;color:#92400e;font-weight:600;",
                                "Float":"background-color:#dcfce7;color:#166534;"}.get(val,"")

                    st.dataframe(
                        succ_df.style.applymap(_succ_style, subset=["Critical Flag"]),
                        use_container_width=True, hide_index=True, height=min(400, 45+len(succ_df)*35),
                    )

            # ---- Individual export -------------------------------------------
            st.divider()
            exp_col, _ = st.columns([1,3])

            exp_detail = pd.DataFrame([{
                "Activity ID":    ms_code,
                "Activity Name":  str(ms.get("task_name","")),
                "WBS":            str(ms.get("wbs_path","")),
                "Forecast Start": format_date(ms.get("eff_start")),
                "Forecast Finish":format_date(ms.get("eff_finish")),
                "Total Float (d)":safe_float(ms.get("total_float_days"), None),
                "Critical":       "Yes" if ms.get("is_critical") else "No",
                "Movement (d)":   move if move is not None else "N/A",
                "Status":         _status_label(str(ms.get("status",""))),
                "Risk Rating":    ms.get("risk_rating",""),
                "In Notes":       "Yes" if ms_code in notes_ids else "No",
                "Risk Keywords":  "Yes" if ms_code in risk_ids else "No",
                "Constraint":     str(ms.get("cstr_type","")).strip() or "None",
            }])

            xls_ms = export_df_to_excel({"Milestone Detail": exp_detail})
            exp_col.download_button(
                label=f"📥 Export {ms_code} Report",
                data=xls_ms,
                file_name=f"milestone_{ms_code}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key=f"ms_exp_{ms_id}",
            )

    # =========================================================================
    # FULL EXPORT
    # =========================================================================
    st.divider()

    full_exp_rows = []
    for _, ms in milestones.iterrows():
        code = str(ms.get("task_code",""))
        tf   = safe_float(ms.get("total_float_days"), None)
        full_exp_rows.append({
            "Activity ID":      code,
            "Activity Name":    str(ms.get("task_name","")),
            "WBS":              str(ms.get("wbs_path","")),
            "Forecast Start":   format_date(ms.get("eff_start")),
            "Forecast Finish":  format_date(ms.get("eff_finish")),
            "Total Float (d)":  round(tf,1) if tf is not None else "-",
            "Critical":         "Yes" if ms.get("is_critical") else "",
            "Risk Rating":      ms.get("risk_rating",""),
            "Movement (d)":     movement_map.get(code, "-"),
            "Status":           _status_label(str(ms.get("status",""))),
            "In Notes":         "Yes" if code in notes_ids else "",
            "Risk Keywords":    "Yes" if code in risk_ids else "",
            "Constraint":       str(ms.get("cstr_type","")).strip() or "",
        })

    full_df = pd.DataFrame(full_exp_rows)
    high_df = full_df[full_df["Risk Rating"] == "High"]
    med_df  = full_df[full_df["Risk Rating"] == "Medium"]

    xls_full = export_df_to_excel({
        "All Milestones": full_df,
        "High Risk":      high_df if not high_df.empty else pd.DataFrame(columns=["No data"]),
        "Medium Risk":    med_df  if not med_df.empty  else pd.DataFrame(columns=["No data"]),
    })

    dl_col, _ = st.columns([1,3])
    dl_col.download_button(
        label="📥  Export All Milestones to Excel",
        data=xls_full,
        file_name=f"milestones_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        help="Exports All Milestones, High Risk and Medium Risk sheets.",
    )




# -----------------------------------------------------------------------------
# PAGE: RISK & OPPORTUNITY REGISTER
# -----------------------------------------------------------------------------

# Probability and impact scoring for auto-generated items
_PROB_MAP   = {"High": "High", "Medium": "Medium", "Low": "Low"}
_IMPACT_MAP = {"High": "High", "Medium": "Medium", "Low": "Low"}

# RAG colour helpers reused from rest of app
def _rag_colour(priority: str) -> tuple:
    """Return (text_colour, bg_colour, border_colour) for a priority level."""
    return {
        "High":   ("#991b1b", "#fef2f2", "#fca5a5"),
        "Medium": ("#92400e", "#fffbeb", "#fcd34d"),
        "Low":    ("#166534", "#f0fdf4", "#86efac"),
    }.get(priority, ("#374151", "#f9fafb", "#E2E8F0"))


def _ro_card_pill(item_type: str) -> str:
    """Coloured pill for Risk vs Opportunity."""
    if item_type == "Risk":
        return (
            '<span style="background:#dc2626;color:white;padding:2px 10px;'
            'border-radius:10px;font-size:10px;font-weight:700;letter-spacing:0.5px;">'
            'RISK</span>'
        )
    return (
        '<span style="background:#16a34a;color:white;padding:2px 10px;'
        'border-radius:10px;font-size:10px;font-weight:700;letter-spacing:0.5px;">'
        'OPPORTUNITY</span>'
    )


def _ro_row(item_type, priority, act_code, act_name, wbs,
            description, cause, effect, mitigation):
    """Build a single register row dict."""
    return {
        "Type":            item_type,
        "Priority":        priority,
        "Activity ID":     str(act_code),
        "Activity Name":   str(act_name),
        "WBS":             str(wbs),
        "Description":     str(description),
        "Cause":           str(cause),
        "Effect":          str(effect),
        "Mitigation / Action": str(mitigation),
        "Owner":           "",
        "Due Date":        "",
        "Status":          "Open",
    }


def _generate_register(
    tasks: pd.DataFrame,
    rels:  pd.DataFrame,
    near_crit_days: float,
    notes_text: str = "",
) -> pd.DataFrame:
    """
    Generate a draft Risk & Opportunity register from programme data.
    Returns a DataFrame with one row per item.
    """
    rows = []
    now  = datetime.now()
    eight_weeks = now + timedelta(weeks=8)
    tasks = get_critical_threshold(tasks, near_crit_days)

    def _wbs(row):
        w = row.get("wbs_path","") if "wbs_path" in row.index else ""
        return str(w).split(" > ")[0] if w and str(w).strip() not in ("","nan") else "-"

    # -- Predecessor / successor lookup ----------------------------------------
    tasks_with_pred = set()
    tasks_with_succ = set()
    if not rels.empty:
        if "succ_task_id" in rels.columns:
            tasks_with_pred = set(rels["succ_task_id"].dropna())
        if "pred_task_id" in rels.columns:
            tasks_with_succ = set(rels["pred_task_id"].dropna())

    # =========================================================================
    # RISKS
    # =========================================================================

    # R1: Negative float
    if "total_float_days" in tasks.columns:
        neg = tasks[tasks["total_float_days"].apply(lambda f: safe_float(f,0) < 0)]
        for _, t in neg.iterrows():
            tf = round(safe_float(t.get("total_float_days"),0), 1)
            rows.append(_ro_row(
                "Risk", "High",
                t.get("task_code",""), t.get("task_name",""), _wbs(t),
                f"Activity has {abs(tf)} days negative float.",
                f"The current schedule logic and constraints result in a {abs(tf)}-day overrun on this activity.",
                "The project completion date cannot be achieved on this path without intervention. "
                "Every day of further delay worsens the position.",
                f"Raise with the planner immediately. Develop a recovery plan to recover the {abs(tf)} days. "
                "Consider acceleration, scope reduction, or parallel working.",
            ))

    # R2: Critical activities not started
    if "status" in tasks.columns and "is_critical" in tasks.columns:
        crit_ns = tasks[
            tasks["is_critical"] &
            tasks["status"].apply(lambda s: str(s) in ("TK_NotStart","Not Started"))
        ]
        for _, t in crit_ns.iterrows():
            finish = format_date(t.get("eff_finish"))
            rows.append(_ro_row(
                "Risk", "High",
                t.get("task_code",""), t.get("task_name",""), _wbs(t),
                f"Critical activity not yet started. Target finish: {finish}.",
                "The activity is on the critical path but mobilisation or enabling works have not commenced.",
                "Any further delay to this activity will directly delay the project finish date.",
                "Confirm the start date and mobilisation plan. If at risk, escalate to the delivery team and PM.",
            ))

    # R3: Near-critical due within 8 weeks
    if "eff_finish" in tasks.columns and "is_near_critical" in tasks.columns:
        nc_soon = tasks[
            tasks["is_near_critical"] &
            tasks["eff_finish"].apply(
                lambda d: d is not None and hasattr(d,"date") and d <= eight_weeks
            )
        ]
        for _, t in nc_soon.iterrows():
            tf = round(safe_float(t.get("total_float_days"),0), 1)
            rows.append(_ro_row(
                "Risk", "Medium",
                t.get("task_code",""), t.get("task_name",""), _wbs(t),
                f"Near-critical activity due within 8 weeks with only {tf} days float.",
                "Limited schedule buffer combined with imminent completion requirements.",
                "If this activity is delayed, it will move onto the critical path and threaten the project finish.",
                f"Monitor weekly. If float drops below 5 days, treat as critical and escalate.",
            ))

    # R4: No predecessor (open start)
    if "task_id" in tasks.columns and "task_type" in tasks.columns:
        no_pred = tasks[
            ~tasks["task_id"].isin(tasks_with_pred) &
            ~tasks["task_type"].astype(str).str.contains("Milestone|LOE|WBS", na=False)
        ]
        for _, t in no_pred.head(15).iterrows():
            rows.append(_ro_row(
                "Risk", "Medium",
                t.get("task_code",""), t.get("task_name",""), _wbs(t),
                "Activity has no predecessor. Programme logic is open at the start.",
                "Missing or incomplete logic in the schedule. The activity may start earlier or later than intended.",
                "Float calculations for this activity may be unreliable. "
                "The activity could be delayed without any schedule warning.",
                "Review with the planner. Add a logical predecessor or add a constraint if the start date is fixed.",
            ))

    # R5: No successor (open finish)
    if "task_id" in tasks.columns and "task_type" in tasks.columns:
        no_succ = tasks[
            ~tasks["task_id"].isin(tasks_with_succ) &
            ~tasks["task_type"].astype(str).str.contains("Finish Milestone|LOE|WBS", na=False)
        ]
        for _, t in no_succ.head(15).iterrows():
            rows.append(_ro_row(
                "Risk", "Medium",
                t.get("task_code",""), t.get("task_name",""), _wbs(t),
                "Activity has no successor. Programme logic is open at the finish.",
                "Missing logic means this activity does not drive any subsequent work in the schedule.",
                "The activity may show artificially high float. "
                "Delays may not cascade correctly through the programme.",
                "Review with the planner. Add a logical successor or confirm the activity is a deliberate end point.",
            ))

    # R6: Excessive lag (> 10 days)
    if not rels.empty and "lag_days" in rels.columns:
        big_lag = rels[rels["lag_days"].apply(lambda l: safe_float(l,0) > 10)]
        for _, r in big_lag.head(10).iterrows():
            code = r.get("succ_task_code", r.get("succ_task_id",""))
            name = r.get("succ_task_name","")
            pred = r.get("pred_task_code", r.get("pred_task_id",""))
            lag  = int(safe_float(r.get("lag_days",0),0))
            match = tasks[tasks["task_code"] == str(code)]
            wbs   = _wbs(match.iloc[0]) if not match.empty else "-"
            rows.append(_ro_row(
                "Risk", "Low",
                code, name, wbs,
                f"Relationship from {pred} has {lag} days lag.",
                f"A {lag}-day lag has been applied instead of modelling the actual work sequence.",
                "Excessive lag disguises schedule risk and inflates float on successor activities. "
                "It may also mask negative float.",
                f"Challenge the {lag}-day lag with the planner. Replace with a properly sequenced "
                "activity or reduce the lag to the minimum justified by contract or site conditions.",
            ))

    # R7: Long duration activities (> 60 days)
    if "orig_dur_days" in tasks.columns:
        long_dur = tasks[tasks["orig_dur_days"].apply(lambda d: safe_float(d,0) > 60)]
        for _, t in long_dur.head(10).iterrows():
            dur = int(safe_float(t.get("orig_dur_days",0),0))
            rows.append(_ro_row(
                "Risk", "Low",
                t.get("task_code",""), t.get("task_name",""), _wbs(t),
                f"Activity duration is {dur} working days.",
                "Activity is too long to manage or monitor effectively as a single work package.",
                "Problems can go undetected for weeks. Float burn and delays may not surface until it is too late.",
                f"Break the {dur}-day activity into smaller work packages of 20-30 days. "
                "Discuss with the planner to improve schedule resolution.",
            ))

    # R8: Planning notes risk words
    if notes_text and "task_code" in tasks.columns:
        for word in _RISK_WORDS:
            pattern = re.compile(r'\b' + re.escape(word) + r'\b', re.IGNORECASE)
            if not pattern.search(notes_text):
                continue
            for _, t in tasks.iterrows():
                code = str(t.get("task_code",""))
                if not code or code not in notes_text:
                    continue
                idx = notes_text.find(code)
                snippet = notes_text[max(0,idx-300):idx+300]
                if pattern.search(snippet):
                    rows.append(_ro_row(
                        "Risk", "High",
                        code, t.get("task_name",""), _wbs(t),
                        f"Planning notes reference '{word}' against this activity.",
                        f"The planning note indicates a potential {word} issue that is not fully visible in the programme.",
                        "This risk may not be reflected in the current schedule float or critical path.",
                        f"Review the planning note for {code}. Confirm whether the '{word}' item has been "
                        "resolved, raise a formal risk if not, and update the programme if dates are affected.",
                    ))
                    break

    # R9: Activities delayed in comparison (> 10 days slip)
    if "_mi_prev" in st.session_state and "_mi_curr" in st.session_state:
        try:
            prev_t = st.session_state["_mi_prev"]["tasks_df"]
            curr_t = st.session_state["_mi_curr"]["tasks_df"]
            if not prev_t.empty and not curr_t.empty:
                merged_comp = prev_t[["task_code","eff_finish","task_name"]].merge(
                    curr_t[["task_code","eff_finish"]], on="task_code", suffixes=("_p","_c"), how="inner"
                )
                for _, r in merged_comp.iterrows():
                    try:
                        slip = int((pd.Timestamp(r["eff_finish_c"]) - pd.Timestamp(r["eff_finish_p"])).days)
                        if slip > 10:
                            match = tasks[tasks["task_code"] == str(r["task_code"])]
                            wbs   = _wbs(match.iloc[0]) if not match.empty else "-"
                            rows.append(_ro_row(
                                "Risk", "High" if slip > 30 else "Medium",
                                str(r["task_code"]), str(r.get("task_name","")), wbs,
                                f"Activity finish date has slipped {slip} days since the previous programme revision.",
                                f"The activity's forecast finish moved {slip} days later between the two programme versions.",
                                f"A {slip}-day slip on this activity may delay downstream work and impact the project finish date.",
                                f"Investigate the reason for the {slip}-day slip. Agree a recovery programme with the delivery team. "
                                "Update the programme and issue a revised schedule.",
                            ))
                    except Exception:
                        pass
        except Exception:
            pass

    # =========================================================================
    # OPPORTUNITIES
    # =========================================================================

    # O1: Activities pulled earlier in comparison (> 5 days improvement)
    if "_mi_prev" in st.session_state and "_mi_curr" in st.session_state:
        try:
            prev_t = st.session_state["_mi_prev"]["tasks_df"]
            curr_t = st.session_state["_mi_curr"]["tasks_df"]
            if not prev_t.empty and not curr_t.empty:
                merged_opp = prev_t[["task_code","eff_finish","task_name"]].merge(
                    curr_t[["task_code","eff_finish"]], on="task_code", suffixes=("_p","_c"), how="inner"
                )
                for _, r in merged_opp.iterrows():
                    try:
                        gain = int((pd.Timestamp(r["eff_finish_p"]) - pd.Timestamp(r["eff_finish_c"])).days)
                        if gain > 5:
                            match = tasks[tasks["task_code"] == str(r["task_code"])]
                            wbs   = _wbs(match.iloc[0]) if not match.empty else "-"
                            rows.append(_ro_row(
                                "Opportunity", "Medium",
                                str(r["task_code"]), str(r.get("task_name","")), wbs,
                                f"Activity finish date has improved by {gain} days since the previous revision.",
                                f"The activity's forecast finish moved {gain} days earlier between programme versions.",
                                f"This improvement may allow successor activities to start earlier and could reduce the overall project duration.",
                                f"Review whether the {gain}-day improvement can be formally recognised in the programme. "
                                "Check if successor activities can be brought forward accordingly.",
                            ))
                    except Exception:
                        pass
        except Exception:
            pass

    # O2: High float activities that could be resequenced (float > 30 days)
    if "total_float_days" in tasks.columns:
        high_float = tasks[tasks["total_float_days"].apply(lambda f: safe_float(f,0) > 30)]
        for _, t in high_float.head(10).iterrows():
            tf = int(safe_float(t.get("total_float_days"),0))
            rows.append(_ro_row(
                "Opportunity", "Low",
                t.get("task_code",""), t.get("task_name",""), _wbs(t),
                f"Activity has {tf} days total float. There may be scope to resequence.",
                f"Significant float of {tf} days suggests this activity's resources or timing could be optimised.",
                "The float window could be used to smooth resource demand, resolve clashes, or accelerate predecessor activities.",
                f"Review with the planner whether the {tf} days of float can be used to resource-level, "
                "front-load critical activities, or release resource to other packages.",
            ))

    # O3: Float gained in comparison
    if "_mi_prev" in st.session_state and "_mi_curr" in st.session_state:
        try:
            prev_t = st.session_state["_mi_prev"]["tasks_df"]
            curr_t = st.session_state["_mi_curr"]["tasks_df"]
            if not prev_t.empty and not curr_t.empty and \
               "total_float_days" in prev_t.columns and "total_float_days" in curr_t.columns:
                mf = prev_t[["task_code","total_float_days","task_name"]].merge(
                    curr_t[["task_code","total_float_days"]], on="task_code", suffixes=("_p","_c"), how="inner"
                )
                for _, r in mf.iterrows():
                    fp = safe_float(r.get("total_float_days_p"), None)
                    fc = safe_float(r.get("total_float_days_c"), None)
                    if fp is not None and fc is not None and fc - fp > 10:
                        match = tasks[tasks["task_code"] == str(r["task_code"])]
                        wbs   = _wbs(match.iloc[0]) if not match.empty else "-"
                        gain  = round(fc - fp, 1)
                        rows.append(_ro_row(
                            "Opportunity", "Low",
                            str(r["task_code"]), str(r.get("task_name","")), wbs,
                            f"Activity gained {gain} days float since the previous revision.",
                            "Schedule logic or date changes in the current revision have created additional float.",
                            "This float could be used for resource smoothing, risk buffer, or to accommodate other priorities.",
                            f"Review whether the additional {gain} days of float has been deliberately created or is an unintended consequence. "
                            "Confirm with the planner that this aligns with programme strategy.",
                        ))
        except Exception:
            pass

    # O4: Near-critical activities where predecessors have high float
    if "total_float_days" in tasks.columns and not rels.empty:
        nc_acts = tasks[tasks["is_near_critical"]] if "is_near_critical" in tasks.columns else pd.DataFrame()
        if not nc_acts.empty and "pred_task_id" in rels.columns:
            task_float_map = tasks.set_index("task_id")["total_float_days"].to_dict() if "task_id" in tasks.columns else {}
            for _, t in nc_acts.head(10).iterrows():
                tid = t.get("task_id","")
                pred_ids = rels[rels["succ_task_id"] == tid]["pred_task_id"].tolist() if "succ_task_id" in rels.columns else []
                high_float_preds = [
                    p for p in pred_ids
                    if safe_float(task_float_map.get(p), 0) > 20
                ]
                if high_float_preds:
                    tf = round(safe_float(t.get("total_float_days"),0), 1)
                    rows.append(_ro_row(
                        "Opportunity", "Medium",
                        t.get("task_code",""), t.get("task_name",""), _wbs(t),
                        f"Near-critical activity ({tf}d float) has predecessors with high float.",
                        f"One or more predecessor activities have significant float, meaning they could potentially "
                        "be accelerated without impacting the overall programme.",
                        f"Accelerating high-float predecessors could create additional buffer on this near-critical path.",
                        f"Review whether the high-float predecessors can be started earlier or completed faster "
                        "to increase the float buffer on this activity.",
                    ))

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    df = df.drop_duplicates(subset=["Type","Activity ID","Description"]).reset_index(drop=True)

    # Sort: Risks first, then Opportunities; within each by Priority
    priority_order = {"High": 0, "Medium": 1, "Low": 2}
    type_order     = {"Risk": 0, "Opportunity": 1}
    df["_ts"] = df["Type"].map(type_order)
    df["_ps"] = df["Priority"].map(priority_order)
    df = df.sort_values(["_ts","_ps","Activity ID"]).drop(columns=["_ts","_ps"]).reset_index(drop=True)

    return df


def page_risk_register(data: dict, near_crit_days: float):
    """
    Risk & Opportunity Register page.
    Auto-generates a draft register from programme data with
    editable owner, due date and status fields.
    """
    st.title("⚠️ Risk & Opportunity Register")
    st.caption(
        "Auto-generated draft register based on the uploaded programme. "
        "Edit Owner, Due Date and Status inline. Export when complete."
    )

    tasks = data["tasks_df"]
    rels  = data["relationships_df"]

    if tasks.empty:
        st.warning("No activities found. Please upload a programme first.")
        return

    notes_text = st.session_state.get("_notes_text", "")

    # ---- Generate / cache register ------------------------------------------
    prog_key  = st.session_state.get("_xer_cache_key", "")
    cache_key = f"_ro_register_{prog_key}_{near_crit_days}"

    if st.session_state.get("_ro_register_key") != cache_key:
        with st.spinner("Analysing programme for risks and opportunities..."):
            register_df = _generate_register(tasks, rels, near_crit_days, notes_text)
        st.session_state["_ro_register_df"]  = register_df
        st.session_state["_ro_register_key"] = cache_key
    else:
        register_df = st.session_state["_ro_register_df"]

    if register_df.empty:
        st.success("No risks or opportunities generated from the current programme data.")
        return

    # ---- Summary banner -----------------------------------------------------
    n_total = len(register_df)
    n_risk  = int((register_df["Type"] == "Risk").sum())
    n_opp   = int((register_df["Type"] == "Opportunity").sum())
    n_high  = int((register_df["Priority"] == "High").sum())
    n_med   = int((register_df["Priority"] == "Medium").sum())
    n_low   = int((register_df["Priority"] == "Low").sum())

    # Risk counts by priority
    r_high = int(((register_df["Type"]=="Risk") & (register_df["Priority"]=="High")).sum())
    r_med  = int(((register_df["Type"]=="Risk") & (register_df["Priority"]=="Medium")).sum())
    r_low  = int(((register_df["Type"]=="Risk") & (register_df["Priority"]=="Low")).sum())
    o_high = int(((register_df["Type"]=="Opportunity") & (register_df["Priority"]=="High")).sum())
    o_med  = int(((register_df["Type"]=="Opportunity") & (register_df["Priority"]=="Medium")).sum())
    o_low  = int(((register_df["Type"]=="Opportunity") & (register_df["Priority"]=="Low")).sum())

    st.markdown(
        f"""
        <div style="display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:20px;">
            <div style="background:#0B1F33;border-radius:12px;padding:16px 20px;">
                <div style="font-size:11px;font-weight:700;color:#64748B;text-transform:uppercase;
                            letter-spacing:1px;margin-bottom:8px;">Risks  ({n_risk})</div>
                <div style="display:flex;gap:12px;">
                    <div style="text-align:center;">
                        <div style="font-size:9px;color:#64748B;text-transform:uppercase;letter-spacing:0.8px;">High</div>
                        <div style="font-size:24px;font-weight:800;color:#dc2626;line-height:1.1;">{r_high}</div>
                    </div>
                    <div style="text-align:center;">
                        <div style="font-size:9px;color:#64748B;text-transform:uppercase;letter-spacing:0.8px;">Medium</div>
                        <div style="font-size:24px;font-weight:800;color:#d97706;line-height:1.1;">{r_med}</div>
                    </div>
                    <div style="text-align:center;">
                        <div style="font-size:9px;color:#64748B;text-transform:uppercase;letter-spacing:0.8px;">Low</div>
                        <div style="font-size:24px;font-weight:800;color:#94A3B8;line-height:1.1;">{r_low}</div>
                    </div>
                </div>
            </div>
            <div style="background:#0B1F33;border-radius:12px;padding:16px 20px;">
                <div style="font-size:11px;font-weight:700;color:#64748B;text-transform:uppercase;
                            letter-spacing:1px;margin-bottom:8px;">Opportunities  ({n_opp})</div>
                <div style="display:flex;gap:12px;">
                    <div style="text-align:center;">
                        <div style="font-size:9px;color:#64748B;text-transform:uppercase;letter-spacing:0.8px;">High</div>
                        <div style="font-size:24px;font-weight:800;color:#16a34a;line-height:1.1;">{o_high}</div>
                    </div>
                    <div style="text-align:center;">
                        <div style="font-size:9px;color:#64748B;text-transform:uppercase;letter-spacing:0.8px;">Medium</div>
                        <div style="font-size:24px;font-weight:800;color:#16a34a;line-height:1.1;">{o_med}</div>
                    </div>
                    <div style="text-align:center;">
                        <div style="font-size:9px;color:#64748B;text-transform:uppercase;letter-spacing:0.8px;">Low</div>
                        <div style="font-size:24px;font-weight:800;color:#94A3B8;line-height:1.1;">{o_low}</div>
                    </div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ---- Filters ------------------------------------------------------------
    with st.expander("Filter register", expanded=False):
        f1, f2, f3, f4 = st.columns(4)

        f_type   = f1.selectbox("Type",     ["All","Risk","Opportunity"], key="ro_f_type")
        f_pri    = f2.selectbox("Priority", ["All","High","Medium","Low"],  key="ro_f_pri")
        f_status = f3.selectbox("Status",   ["All","Open","In Progress","Closed"], key="ro_f_status")

        all_wbs  = ["All"] + sorted(register_df["WBS"].unique().tolist())
        f_wbs    = f4.selectbox("WBS",      all_wbs, key="ro_f_wbs")

    filtered = register_df.copy()
    if f_type   != "All": filtered = filtered[filtered["Type"]     == f_type]
    if f_pri    != "All": filtered = filtered[filtered["Priority"] == f_pri]
    if f_status != "All": filtered = filtered[filtered["Status"]   == f_status]
    if f_wbs    != "All": filtered = filtered[filtered["WBS"]      == f_wbs]

    st.caption(f"Showing {len(filtered)} of {n_total} items.")

    # ---- Tabs: Risks | Opportunities | Full Register | Export ----------------
    tab_risks, tab_opps, tab_full, tab_export = st.tabs([
        f"Risks ({n_risk})",
        f"Opportunities ({n_opp})",
        "Full Register",
        "Export",
    ])

    EDIT_COLS = [
        "Type","Priority","Activity ID","Activity Name","WBS",
        "Description","Cause","Effect","Mitigation / Action",
        "Owner","Due Date","Status",
    ]

    COL_CONFIG = {
        "Type": st.column_config.SelectboxColumn(
            "Type", options=["Risk","Opportunity"], width="small"
        ),
        "Priority": st.column_config.SelectboxColumn(
            "Priority", options=["High","Medium","Low"], width="small"
        ),
        "Status": st.column_config.SelectboxColumn(
            "Status", options=["Open","In Progress","Closed"], width="small"
        ),
        "Owner":              st.column_config.TextColumn("Owner",    width="small"),
        "Due Date":           st.column_config.TextColumn("Due Date", width="small"),
        "Activity ID":        st.column_config.TextColumn("Activity ID",   width="small"),
        "Activity Name":      st.column_config.TextColumn("Activity Name", width="medium"),
        "WBS":                st.column_config.TextColumn("WBS",           width="medium"),
        "Description":        st.column_config.TextColumn("Description",   width="large"),
        "Cause":              st.column_config.TextColumn("Cause",         width="large"),
        "Effect":             st.column_config.TextColumn("Effect",        width="large"),
        "Mitigation / Action":st.column_config.TextColumn("Mitigation / Action", width="large"),
    }

    def _style_row(row):
        """Row background by Type and Priority."""
        if row.get("Type","") == "Opportunity":
            return ["background-color:#f0fdf4;"] * len(row)
        priority = row.get("Priority","")
        cm = {
            "High":   "background-color:#fef2f2;",
            "Medium": "background-color:#fffbeb;",
        }
        return [cm.get(priority,"")] * len(row)

    def _show_editable(df, key_suffix):
        avail = [c for c in EDIT_COLS if c in df.columns]
        edited = st.data_editor(
            df[avail].style.apply(_style_row, axis=1),
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            column_config=COL_CONFIG,
            key=f"ro_editor_{key_suffix}",
        )
        return edited

    # -- Risks tab --------------------------------------------------------------
    with tab_risks:
        risks_df = filtered[filtered["Type"] == "Risk"].copy()
        if risks_df.empty:
            st.success("No risks match the current filters.")
        else:
            st.markdown(
                '<div style="font-size:12px;font-weight:700;color:#94A3B8;letter-spacing:1px;'
                'text-transform:uppercase;margin-bottom:8px;">Risks</div>',
                unsafe_allow_html=True,
            )
            st.caption("Edit Owner, Due Date and Status directly in the table.")
            edited_risks = _show_editable(risks_df, "risks")

            # Persist edits
            if edited_risks is not None and not edited_risks.empty:
                for col in ["Owner","Due Date","Status","Priority"]:
                    if col in edited_risks.columns:
                        register_df.loc[risks_df.index, col] = edited_risks[col].values
                st.session_state["_ro_register_df"] = register_df

            # High priority cards
            high_risks = risks_df[risks_df["Priority"] == "High"]
            if not high_risks.empty:
                st.markdown("---")
                st.markdown(
                    f'<div style="font-size:12px;font-weight:700;color:#dc2626;letter-spacing:1px;'
                    f'text-transform:uppercase;margin-bottom:10px;">'
                    f'High Priority Risks ({len(high_risks)})</div>',
                    unsafe_allow_html=True,
                )
                for _, item in high_risks.iterrows():
                    tc, bc, brc = _rag_colour("High")
                    st.markdown(
                        f"""
                        <div style="background:{bc};border:1px solid {brc};border-left:5px solid #dc2626;
                                    border-radius:8px;padding:14px 18px;margin-bottom:8px;">
                            <div style="display:flex;align-items:flex-start;justify-content:space-between;
                                        gap:12px;flex-wrap:wrap;">
                                <div style="flex:1;min-width:0;">
                                    {_ro_card_pill("Risk")}
                                    <div style="font-weight:700;color:#0B1F33;font-size:14px;margin-top:6px;">
                                        {item.get("Activity ID","")} - {item.get("Activity Name","")}
                                    </div>
                                    <div style="font-size:12px;color:#64748B;margin-top:2px;">
                                        {item.get("WBS","")}
                                    </div>
                                </div>
                            </div>
                            <div style="margin-top:10px;display:grid;grid-template-columns:1fr 1fr 1fr;gap:10px;">
                                <div>
                                    <div style="font-size:10px;font-weight:700;color:#94A3B8;
                                                text-transform:uppercase;letter-spacing:0.8px;">Description</div>
                                    <div style="font-size:13px;color:#334155;margin-top:3px;">{item.get("Description","")}</div>
                                </div>
                                <div>
                                    <div style="font-size:10px;font-weight:700;color:#94A3B8;
                                                text-transform:uppercase;letter-spacing:0.8px;">Effect</div>
                                    <div style="font-size:13px;color:#334155;margin-top:3px;">{item.get("Effect","")}</div>
                                </div>
                                <div>
                                    <div style="font-size:10px;font-weight:700;color:#94A3B8;
                                                text-transform:uppercase;letter-spacing:0.8px;">Mitigation</div>
                                    <div style="font-size:13px;color:#334155;margin-top:3px;">{item.get("Mitigation / Action","")}</div>
                                </div>
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

    # -- Opportunities tab ------------------------------------------------------
    with tab_opps:
        opps_df = filtered[filtered["Type"] == "Opportunity"].copy()
        if opps_df.empty:
            st.info("No opportunities detected from the current programme data.")
            if "_mi_prev" not in st.session_state:
                st.caption(
                    "Note: Comparison-based opportunities (pulled-forward activities, float gained) "
                    "require both previous and current XER files to be uploaded on the "
                    "Programme Comparison page."
                )
        else:
            st.markdown(
                '<div style="font-size:12px;font-weight:700;color:#16a34a;letter-spacing:1px;'
                'text-transform:uppercase;margin-bottom:8px;">Opportunities</div>',
                unsafe_allow_html=True,
            )
            st.caption("Review these opportunities with the planner to see if they can be captured.")
            edited_opps = _show_editable(opps_df, "opps")

            if edited_opps is not None and not edited_opps.empty:
                for col in ["Owner","Due Date","Status","Priority"]:
                    if col in edited_opps.columns:
                        register_df.loc[opps_df.index, col] = edited_opps[col].values
                st.session_state["_ro_register_df"] = register_df

    # -- Full register tab ------------------------------------------------------
    with tab_full:
        st.caption("Complete register -- all risks and opportunities combined.")
        edited_full = _show_editable(filtered, "full")
        if edited_full is not None and not edited_full.empty:
            for col in ["Owner","Due Date","Status","Priority"]:
                if col in edited_full.columns:
                    register_df.loc[filtered.index, col] = edited_full[col].values
            st.session_state["_ro_register_df"] = register_df

    # -- Export tab -------------------------------------------------------------
    with tab_export:
        st.markdown("Download the full Risk & Opportunity Register as a formatted Excel workbook.")

        final_df = st.session_state.get("_ro_register_df", register_df)

        # Summary sheet
        status_counts = final_df.groupby(["Type","Priority"]).size().reset_index(name="Count")
        owner_counts  = final_df[final_df["Owner"] != ""].groupby("Owner").size().reset_index(name="Assigned Items")

        summary_data  = pd.DataFrame({
            "Metric": [
                "Total Items","Total Risks","Total Opportunities",
                "High Priority Risks","Medium Priority Risks","Low Priority Risks",
                "High Priority Opportunities","Medium Priority Opportunities","Low Priority Opportunities",
                "Open Items","In Progress","Closed",
            ],
            "Count": [
                len(final_df),
                int((final_df["Type"]=="Risk").sum()),
                int((final_df["Type"]=="Opportunity").sum()),
                int(((final_df["Type"]=="Risk")&(final_df["Priority"]=="High")).sum()),
                int(((final_df["Type"]=="Risk")&(final_df["Priority"]=="Medium")).sum()),
                int(((final_df["Type"]=="Risk")&(final_df["Priority"]=="Low")).sum()),
                int(((final_df["Type"]=="Opportunity")&(final_df["Priority"]=="High")).sum()),
                int(((final_df["Type"]=="Opportunity")&(final_df["Priority"]=="Medium")).sum()),
                int(((final_df["Type"]=="Opportunity")&(final_df["Priority"]=="Low")).sum()),
                int((final_df["Status"]=="Open").sum()),
                int((final_df["Status"]=="In Progress").sum()),
                int((final_df["Status"]=="Closed").sum()),
            ],
        })

        export_sheets = {
            "Summary":          summary_data,
            "Full Register":    final_df,
            "Risks":            final_df[final_df["Type"]=="Risk"],
            "Opportunities":    final_df[final_df["Type"]=="Opportunity"],
            "High Priority":    final_df[final_df["Priority"]=="High"],
            "Open Items":       final_df[final_df["Status"]=="Open"],
        }
        if not owner_counts.empty:
            export_sheets["By Owner"] = owner_counts

        xls_bytes = export_df_to_excel(export_sheets)

        dl_col, _ = st.columns([1,3])
        dl_col.download_button(
            label="📥  Download Risk & Opportunity Register",
            data=xls_bytes,
            file_name=f"risk_register_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            help="Exports Summary, Full Register, Risks, Opportunities, High Priority and Open Items sheets.",
        )

        st.markdown(
            '<div style="background:#f8fafc;border:1px solid #E2E8F0;border-radius:8px;'
            'padding:14px 18px;margin-top:12px;">'
            '<div style="font-size:12px;font-weight:700;color:#0B1F33;margin-bottom:6px;">'
            'Workbook sheets</div>'
            '<div style="font-size:12px;color:#64748B;line-height:2;">'
            + "".join(f"<strong>{k}</strong> &nbsp;|&nbsp; " for k in export_sheets.keys()) +
            '</div></div>',
            unsafe_allow_html=True,
        )
        st.caption(
            "Note: This is an auto-generated draft register. "
            "Review all items with the project team before issuing formally."
        )



# -----------------------------------------------------------------------------
# PAGE: EXPORT REPORTS
# -----------------------------------------------------------------------------

def page_export_reports(data: dict, near_crit_days: float):
    st.title("📥 Export Reports")
    st.markdown("> Download all schedule data as formatted Excel reports.")

    tasks = data["tasks_df"]
    rels = data["relationships_df"]
    wbs = data["wbs_df"]
    resources = data["resources_df"]

    if tasks.empty:
        st.warning("No data loaded to export.")
        return

    tasks = get_critical_threshold(tasks, near_crit_days)
    critical = tasks[tasks["is_critical"]]
    neg_float = tasks[tasks["total_float_days"].apply(lambda f: f is not None and f < 0)]

    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Single-Sheet Exports")

        # All activities
        avail = [c for c in ["task_code","task_name","wbs_path","eff_start","eff_finish",
                              "orig_dur_days","rem_dur_days","total_float_days","free_float_days",
                              "status","task_type","is_critical","cstr_type"] if c in tasks.columns]
        xls = export_df_to_excel({"All Activities": tasks[avail]})
        st.download_button("📄 All Activities", xls, "all_activities.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Critical path
        avail_c = [c for c in avail if c in critical.columns]
        xls2 = export_df_to_excel({"Critical Path": critical[avail_c]})
        st.download_button("🔴 Critical Path Activities", xls2, "critical_path.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Relationships
        if not rels.empty:
            xls3 = export_df_to_excel({"Relationships": rels})
            st.download_button("🔗 All Relationships", xls3, "relationships.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with col2:
        st.subheader("Multi-Sheet Reports")

        # Full schedule pack
        sheets = {"All Activities": tasks[avail]}
        if not critical.empty:
            sheets["Critical Path"] = critical[avail_c]
        if not neg_float.empty:
            sheets["Negative Float"] = neg_float[[c for c in avail if c in neg_float.columns]]
        if not rels.empty:
            sheets["Relationships"] = rels
        if not wbs.empty:
            sheets["WBS"] = wbs
        if not resources.empty:
            sheets["Resources"] = resources

        xls_full = export_df_to_excel(sheets)
        st.download_button("📦 Full Schedule Data Pack", xls_full, "schedule_data_pack.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # WBS summary
        if "wbs_path" in tasks.columns:
            tasks["wbs_top"] = tasks["wbs_path"].apply(
                lambda x: str(x).split(" > ")[0] if pd.notna(x) and x else "Unknown"
            )
            wbs_summary = tasks.groupby("wbs_top").agg(
                total=("task_id","count"),
                critical=("is_critical","sum"),
                near_critical=("is_near_critical","sum"),
            ).reset_index()
            xls_wbs = export_df_to_excel({"WBS Summary": wbs_summary})
            st.download_button("🌲 WBS Summary", xls_wbs, "wbs_summary.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# -----------------------------------------------------------------------------
# PAGE: HOME  (PlanTrace branded landing page)
# -----------------------------------------------------------------------------

def _page_home():
    """
    PlanTrace branded homepage.
    Shown when no XER is loaded, or when the user navigates to Home.
    """

    # ---- Hero section -------------------------------------------------------
    st.markdown(
        """
        <div style="padding: 48px 0 24px 0;">
            <div style="font-size: 13px; font-weight: 700; color: #F5A623;
                        letter-spacing: 2px; text-transform: uppercase;
                        margin-bottom: 10px;">
                Project Programme Intelligence
            </div>
            <div style="font-size: 52px; font-weight: 900; color: #0B1F33;
                        line-height: 1.1; letter-spacing: -1px;">
                PlanTrace
            </div>
            <div style="width: 56px; height: 4px; background: #F5A623;
                        border-radius: 2px; margin: 14px 0 18px 0;"></div>
            <div style="font-size: 20px; font-weight: 400; color: #334155;
                        margin-bottom: 10px;">
                Trace logic. Expose risk. Drive delivery.
            </div>
            <div style="font-size: 15px; color: #64748B; max-width: 680px;
                        line-height: 1.7; margin-bottom: 32px;">
                Project planning intelligence for delivery teams. Upload an XER programme,
                trace predecessors and successors, review critical paths, check programme
                health and understand labour demand &mdash; without opening P6.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ---- Upload prompt -------------------------------------------------------
    st.markdown(
        """
        <div style="background:#0B1F33; border-radius:12px; padding:22px 28px;
                    display:flex; align-items:center; gap:20px; margin-bottom:36px;
                    max-width:560px;">
            <div style="font-size:28px;">📂</div>
            <div>
                <div style="color:#F5A623;font-weight:700;font-size:15px;
                            margin-bottom:4px;">Ready to start</div>
                <div style="color:#CBD5E1;font-size:13px;line-height:1.5;">
                    Upload your <strong style="color:#fff;">.xer file</strong>
                    using the panel on the left to begin analysis.
                    <br>Export from P6 via
                    <strong style="color:#F5A623;">File &rarr; Export &rarr; Primavera P6 XER</strong>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ---- Feature cards -------------------------------------------------------
    st.markdown(
        '<div style="font-size:13px;font-weight:700;color:#94A3B8;'
        'letter-spacing:1.5px;text-transform:uppercase;margin-bottom:16px;">'
        'What PlanTrace does</div>',
        unsafe_allow_html=True,
    )

    CARDS = [
        {
            "icon": "🔗",
            "title": "Logic Trace",
            "body": (
                "See what drives an activity and what it impacts. "
                "Trace full predecessor and successor chains across the network, "
                "with depth levels and relationship types shown at every step."
            ),
        },
        {
            "icon": "🚨",
            "title": "Critical Path",
            "body": (
                "Review the full critical path, near-critical work and negative float. "
                "Identify which activity or milestone is at risk and understand "
                "exactly what is driving it."
            ),
        },
        {
            "icon": "👷",
            "title": "Labour Demand",
            "body": (
                "View labour histograms by week, month, WBS and resource. "
                "Identify peak demand periods and understand resource loading "
                "across the programme."
            ),
        },
        {
            "icon": "🩺",
            "title": "Programme Health",
            "body": (
                "Find missing logic, open ends, constraints, excessive lag and "
                "planning risk before they cause problems. "
                "Eleven automated quality checks with export."
            ),
        },
    ]

    cols = st.columns(4, gap="medium")
    for col, card in zip(cols, CARDS):
        with col:
            st.markdown(
                f"""
                <div class="pt-card">
                    <div class="pt-card-icon">{card["icon"]}</div>
                    <div class="pt-card-accent"></div>
                    <div class="pt-card-title">{card["title"]}</div>
                    <div class="pt-card-body">{card["body"]}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    # ---- What's in the tool -------------------------------------------------
    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown(
        '<div style="font-size:13px;font-weight:700;color:#94A3B8;'
        'letter-spacing:1.5px;text-transform:uppercase;margin-bottom:16px;">'
        'All pages</div>',
        unsafe_allow_html=True,
    )

    PAGE_LIST = [
        ("📊", "Project Summary",          "Activity counts, float distribution, WBS breakdown and schedule span."),
        ("🔍", "Activity Search",           "Search and filter activities. View full detail, dates, float and logic."),
        ("🔗", "Logic Trace",               "Trace predecessors and successors through the network with depth levels."),
        ("🚨", "Critical Path Analysis",    "Full critical path, near-critical and negative float by WBS."),
        ("🎯", "Critical Path to Activity", "Identify the driving chain into any selected activity or milestone."),
        ("👷", "Labour Histogram",          "Weekly and monthly labour demand by resource, WBS and package."),
        ("🩺", "Schedule Health Check",     "Eleven automated quality checks with counts, tables and export."),
        ("📝", "Planning Notes",            "Upload notes, link to activities, keyword search and highlighting."),
        ("📅", "Programme Comparison",      "Compare two XER revisions. See what moved, changed or became critical."),
        ("📥", "Export Reports",            "Download all data as formatted Excel workbooks."),
    ]

    left, right = st.columns(2, gap="large")
    for i, (icon, title, desc) in enumerate(PAGE_LIST):
        col = left if i % 2 == 0 else right
        with col:
            st.markdown(
                f"""
                <div style="display:flex;gap:14px;align-items:flex-start;
                            padding:14px 0;border-bottom:1px solid #E2E8F0;">
                    <div style="font-size:22px;min-width:30px;margin-top:2px;">{icon}</div>
                    <div>
                        <div style="font-weight:700;color:#0B1F33;
                                    font-size:14px;margin-bottom:3px;">{title}</div>
                        <div style="color:#64748B;font-size:13px;
                                    line-height:1.5;">{desc}</div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    # ---- Footer -------------------------------------------------------------
    st.markdown("<br><br>", unsafe_allow_html=True)
    st.markdown(
        """
        <div style="border-top:1px solid #E2E8F0;padding-top:18px;
                    display:flex;justify-content:space-between;align-items:center;">
            <div style="font-size:13px;color:#94A3B8;">
                <strong style="color:#0B1F33;">PlanTrace</strong>
                &nbsp;&nbsp;|&nbsp;&nbsp;
                Built for Primavera P6 XER programmes
                &nbsp;&nbsp;|&nbsp;&nbsp;
                No P6 licence required
            </div>
            <div style="font-size:12px;color:#CBD5E1;">
                Upload a .xer file to begin
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


# -----------------------------------------------------------------------------
# UI COMPONENTS  --  PlanTrace v2 Premium
# Brand: Plan=#2563EB  Trace=#2DD4BF  Risk=#DC2626  OK=#16A34A  Navy=#0A1628
# -----------------------------------------------------------------------------

# -- Brand colours ------------------------------------------------------------
PT_BLUE  = "#071827"   # Primary navy (was blue, now proper navy)
PT_TEAL  = "#F5A623"   # Accent amber (replaces teal)
PT_NAVY  = "#071827"   # Dark sidebar / header
PT_NAVY2 = "#0B2438"   # Secondary navy
PT_AMBER = "#F5A623"   # Amber accent
PT_RED   = "#DC2626"   # Risk / critical
PT_GREEN = "#16A34A"   # Healthy / complete
PT_BG    = "#F3F5F7"   # Page background
PT_CARD  = "#FFFFFF"   # Card background

# -- Navigation definition ----------------------------------------------------
_NAV = [
    ("overview",   "Overview",       ""),
    ("programme",  "Programme",      ""),
    ("logic",      "Logic",          ""),
    ("critical",   "Critical Path",  ""),
    ("labour",     "Labour",         ""),
    ("health",     "Health Check",   ""),
    ("comparison", "Comparison",     ""),
    ("pm_actions", "PM Actions",     ""),
    ("risk",       "Risk Register",  ""),
    ("reports",    "Reports",        ""),
    ("settings",   "Settings",       ""),
]

_NEEDS_PROG = {
    "overview", "programme", "logic", "critical",
    "labour", "health", "pm_actions", "risk", "reports",
}


# -----------------------------------------------------------------------------
# HELPER: embed logo as base64 img tag
# -----------------------------------------------------------------------------
import base64 as _b64
import os as _os

def _logo_tag(width: int = 100) -> str:
    """Return an <img> tag with the PlanTrace logo embedded as base64."""
    _LOGO_B64 = "data:image/jpeg;base64,/9j/4QhoRXhpZgAASUkqAAgAAAADAA4BAgD4AwAAMgAAADsBAgAlAAAAKgQAAGmHBAABAAAATwQAAAAAAABTaWduYXR1cmU6IDRqckdNZUlIVUN3Nm5LR2plSFY4TnQremFTNCs2OGt4N29GV1ljUWZKRDNyMGFJbVg3UE1mbEZpdElQaUg5VGkvTUlQbW1zZWJiNkxZVzZEOSt6cy9hRVdESVg1aXJkeUdtTzZuWkhrL0owVFgzcVo4TmVMNnhRWlJMNWUwQlQ3c1Nyak9LeVo4MlNqQStvOG5qcFpZdWJxZDEzT0h3VHUyZUIxWVBhbGZYdHhrQlJBR3FCeVViWDUrR0pnbk9HSmpMZFFuSmdSdCt6dStiUzF4UGV5d1h3d0dNNlQySC9TblA3bWxOR0xZcmRQT3FaZEtSS0RtTFpJbTNTcWdzWGYyU1FldnZwQ1NKUURLUTF5aWQyUjNYUVpvQ283WWhsckgwM2JheWV6N0dwbkVMNDEyNG9HcXdDUnRYMldkdkdWQzFHZWxCZ3hmV002aGN0TzEzSmFxM045aG9xdnFEVk81T1Rhc01Dci9nVXVWMzgwQ2JSMmkvMms1TFkxcElhWkd2b1RmeGsrRThTa25BRmVLeDkrampFZk5qT2xRSk05bUpTZzI5WDhyczJmOG45Q0xPb1JJU290OGc0eTMvNXJBaHpOTmt0Z0t3VDFqN3VsUndpeSswOTQ5SWVsMXF6N2QzVTV6L1FjWmRUVlRtZjkzVHZyVnJtK29zdlM4dUk5S281QXRZWGE0Ykx4cEQ2VTlIbXBsdDVKN0lkQjRHNjZSeGk0OTlQY1dnTTdDbXh4VVIrWFpPbW9TajNsQkNnc3puL29WUWRjKy82SG5GQmRTcysvejFIdy9pTStlQm9wYWpVbXFmRDlrUUIxWmxrZVFBNzJuMWlsWmduZHVsbS9DTUs4OWl6QkZZM0daWExaaks1eVZ3OGVWZk92eEpmbStySDFybWdvdks1bTdsNUZSM1ppTGtGZWhwaGJydFcrdmZ4WGZCSTR0aEF3UCt6dGJYMUQrelYzNDVoVDJQcjFLMEJzT1Y3eDAyRzNxSEt5U3dZQnpyMUxxYURyQytRcnlnVU1pNCtHcTdUTk5vSVZjeEhsNnZzYVNDSzVLbDhycWg5dXdhZTl1d3VpdjNCS29mWHQrOUVNSXBHMVJNaFRIRi9DOHNQcFNBYkRwWWs4emt0M1ZGZk9SUVpaRVR0M01mU28yV0RBOUJjM3IyT21WQ094NVRReGtnN3pLOFBkdUVlTjM0WVNncWQrZ29pK0F6QXp4Yy9yR1NUcVMramZobU1QMlVKUkt0VVo4OUFoNU81ZTJCRlh1UWRBMUdxS09DeVVJajgvdUVVWkJBRFg2SllNTUF6cmZwdm1YMnNIRU1sTDlQU2tuRE09AGQyYWVmMWMyLTk3NjMtNDE2Yi05MTc2LTAwMjk1NmUyZDhkMgABAIaSBwD/AwAAYQQAAAAAAABBU0NJSQAAAFNpZ25hdHVyZTogNGpyR01lSUhVQ3c2bktHamVIVjhOdCt6YVM0KzY4a3g3b0ZXWWNRZkpEM3IwYUltWDdQTWZsRml0SVBpSDlUaS9NSVBtbXNlYmI2TFlXNkQ5K3pzL2FFV0RJWDVpcmR5R21PNm5aSGsvSjBUWDNxWjhOZUw2eFFaUkw1ZTBCVDdzU3JqT0t5WjgyU2pBK284bmpwWll1YnFkMTNPSHdUdTJlQjFZUGFsZlh0eGtCUkFHcUJ5VWJYNStHSmduT0dKakxkUW5KZ1J0K3p1K2JTMXhQZXl3WHd3R002VDJIL1NuUDdtbE5HTFlyZFBPcVpkS1JLRG1MWkltM1NxZ3NYZjJTUWV2dnBDU0pRREtRMXlpZDJSM1hRWm9DbzdZaGxySDAzYmF5ZXo3R3BuRUw0MTI0b0dxd0NSdFgyV2R2R1ZDMUdlbEJneGZXTTZoY3RPMTNKYXEzTjlob3F2cURWTzVPVGFzTUNyL2dVdVYzODBDYlIyaS8yazVMWTFwSWFaR3ZvVGZ4aytFOFNrbkFGZUt4OStqakVmTmpPbFFKTTltSlNnMjlYOHJzMmY4bjlDTE9vUklTb3Q4ZzR5My81ckFoek5Oa3RnS3dUMWo3dWxSd2l5KzA5NDlJZWwxcXo3ZDNVNXovUWNaZFRWVG1mOTNUdnJWcm0rb3N2Uzh1STlLbzVBdFlYYTRiTHhwRDZVOUhtcGx0NUo3SWRCNEc2NlJ4aTQ5OVBjV2dNN0NteHhVUitYWk9tb1NqM2xCQ2dzem4vb1ZRZGMrLzZIbkZCZFNzKy96MUh3L2lNK2VCb3BhalVtcWZEOWtRQjFabGtlUUE3Mm4xaWxaZ25kdWxtL0NNSzg5aXpCRlkzR1pYTFpqSzV5Vnc4ZVZmT3Z4SmZtK3JIMXJtZ292SzVtN2w1RlIzWmlMa0ZlaHBoYnJ0Vyt2ZnhYZkJJNHRoQXdQK3p0YlgxRCt6VjM0NWhUMlByMUswQnNPVjd4MDJHM3FIS3lTd1lCenIxTHFhRHJDK1FyeWdVTWk0K0dxN1ROTm9JVmN4SGw2dnNhU0NLNUtsOHJxaDl1d2FlOXV3dWl2M0JLb2ZYdCs5RU1JcEcxUk1oVEhGL0M4c1BwU0FiRHBZazh6a3QzVkZmT1JRWlpFVHQzTWZTbzJXREE5QmMzcjJPbVZDT3g1VFF4a2c3eks4UGR1RWVOMzRZU2dxZCtnb2krQXpBenhjL3JHU1RxUytqZmhtTVAyVUpSS3RVWjg5QWg1TzVlMkJGWHVRZEExR3FLT0N5VUlqOC91RVVaQkFEWDZKWU1NQXpyZnB2bVgyc0hFTWxMOVBTa25ETT3/4AAQSkZJRgABAQAAAQABAAD/2wBDAAIBAQEBAQIBAQECAgICAgQDAgICAgUEBAMEBgUGBgYFBgYGBwkIBgcJBwYGCAsICQoKCgoKBggLDAsKDAkKCgr/2wBDAQICAgICAgUDAwUKBwYHCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgr/wAARCASQAxADASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD91grDgL+OaTa27oeafg5xg0AHIwP0rQzE8s8duacFOc7TS4bgYNKAe4PWgAAYY+U80u1uSBSEEdqXkE/SgBoVtpyO1BV+mO3rSnkfjRhsjrQAhXknFIFOeh6U8Z54pBnPAoAaVI4xShTjG30obOQcGnDOOR2oAaFJwNvSmsrEY21IFbpj2prAkfd7+lAEZQ4JwfalVTuwF/Gn7WII29qArdcdqAECnup605VY44NJgjHFClsjg8UAORGzkjvSgHAO09aATjoaT5sDg0DuLg91NNIIHI5//VSjJPTtSMD6UCGMrDt9KFVsZ2mnODzx3pADx8vegAwc52nrxRtOOFOMc0g3HsaX5uymgAAJHC96Apz0NKpPpSgkn7tACgHrtNCgnPFGDjgH8qTJA4zQANngY6UENycHrQTk/wD1qTJ5xnrQAqhsZINLg91/SmqW6gd6cevQ/lQAmxuOO1NCn0NPJbjjt3pozzhaAAK2fu9qXB2gbT0oOc8ikyQOnSgAwSTwfxpcENwKRd2cgGnZPpQAYbsKbtYv070pJI4FNG4NgCgBxDKuMd6UBiBx2pMkAjHTNJuJ7dqBikMSRg8UzDDqvU8U7JJOBSjJHTofSkA0Kw6A9ad82TwfagBjnjoPSgZDHg0wAqcAY7elG1uu31pWzjgdqbliMgUCEwemDRg+nf0pCTuGB2pwJxjH8VAxFVj/AA/nS7SMDHalUMRjHP0pGDA8A0CHAfoKTDEHA5oyw7fwikJbGAKAAht2dp5pQCScDtSHdxlTQGOeR2oAUKcj5TSEN02npTlbLcU07uwoANrEfd7ClCnHK9BSnPXBoJJzx24oAiZDx8poC89OgpWycEClAbHA/hoAAM8YP5UuDt+6etIM+lHzDsetAChW4oAYvyO/FL82Bwev9KRS3p+NABhgBkGjB9COPSja2B+lGGwOKAAjtg9uabtYdjS/N2FKQfTvxQAiIeMigKSOh604ZwOO1GSR06UANZSOAp69aaqnHSnNuPQUJuwMDp/9amAYOfu04ZHVTSDOTx39KMtjgUgFAY5IpHRi2MdqAXGcA9DQWJ/KgBpVuTjpSKjbehpeSTRzjgH8qAFAOfu96cuQBlT0poJ64P5U4EkDA7UAKAT/AA0bSSePpSDPQilyewNACMpzwvFNwSxytOO7H3T0603nJNABgjjBpyg5Hymm5OQdp6UqluOKAHLkkYH50rqxzj1poY54HelLnGMHqaADa2AMdqQBjnI796CxJG0Um5j0BoADk9qQBjnCnrQM46GlXODweDQAm1ixOMUrq3OOmKUbieAelI2SCCPegAJPQc0Atnp0FLzjr3pOd1AC5PQg9aWm5OBgHNOBJA5NABkg/h60biD+AoIYnIz9aMN1GenegBNxAyBRuYkcUh3BeppDknvQA8P6igE7hj1pmWB4J6Uqs2aAFJfI+lKCe+OlN5OB7Uo3dMHpQA4ZPb9aQ5xjr+NKCxPOaQ7iOp4oAcOnT9aTk9BQCwH5UDOaAGvuHT1FIGJx9aU57ZpMHg88GgByuc9DRucgDFIM8jmkw2B1/OgB4JwKQk46elC7gBjPSkJIGcmgAfPYd/WkBI60rFv1pFLY4oAXnsKDnOaBuznnr60fNx16UAChhg/1pc88jvQueDk9qQ55NAAS3YUis3XGKG3E9/zpAWwdoP1oAUsR0Wjcxz16+lIc8ZyOO9BzyOaAHKWC8g/iaXcc/dNMUlelKSSeKAFLNgYU9KRWJ5xSc4GPSkBYdzQA4sc8DpQS3XFNy2eM/nS5PoelACoWOTinZbOSO1MXcM9elOJO7NACjPXFGDu6d6M459qQFi3OetAAS2DwetJufPQ9KXnHOfvetBDEjg/hQAKXJ6Hp6UuWA6Hr6Uirz93tRzjGKABSTn8O9Lk5/wDr0gLYxkjmly3XJoGDE5GD+tMLMB0NOJPHU01hkYoENy2QMHj1qRN2OneowDu6VIuQPx60ASKDjgduOaawPp29aeCe3tTWLHgCgBhJB59B3pMt6dx1pTuB79BTQTjg9+aAFYtnoaNx60hznIz+NKN2e9ACqTu6dqMk44/Ohd4ORnpzR82RigBSTjpSbmxwKQlgOlAyRnBoADuHb9aUZHbt60hJxSjIyPagAG7H/wBelw2Oh60iljxz+NByQKAFG/HAP+RSKT6fnSknGAO9NUn0NADvmAGBnijLYGfam/NgA+lGTwAe1AAxIHUfnSsST0700s3qelKxY9STzQAoJx0oG707etALbcf57ULu64NACMWx0pFYj/65pSGxjn86Rc5BGetCAUZ5JFL8xGAKRc9aU5oARSwBI9PWly3YGkBcA7aCW4PPSgAJfkbf0oG4DOKQ7snk/nQue2aAD5uODSgnbwP4fWkOc0oJwO3FACgt1xS85/H1pAT6UZO78aAAliOnamgnOf5mlJOfwpF3ZJBoACDkYz0pQWPVTwKDk9B2oxk8UAHIOMHrRubHTuaAp3DGaUhsY5zmgAy2Rx2pAWHb86BuzkZwBQA2eBQAEkdaAxwTjvSZbGRnr3pAzDPJHNAEikk8A0MCVPH50wH5u9OyQO/TNADycccdfpTM/NyBSse3p703Pzfy5pALu6cH86VenTv600AZHApcADPFMB24ZB4/OlJGMD0phwD94fnSnnkelACBsLwPyozmTIHemsfl6Hp6Uo27ufagB3U8Dt6+9GRu6frS4HrSfLnGR+dACEjgEHp3pc8cjsO9IQox9P71L7YP4mgA3D0/M0FhjgdvWkGfQ9PWg/T9aAHFhk/h3oyQOn600kYOB2/vUbjjjP50APJBGNp7UA9BjvUZPHI/SlUggDjrQA4H5icUHoCF7d6aPf8AlS4GB0HSgBdwOMD9aaWAHToPWlwCP/r0jHgn+tAAX4696FYYHI/P6UhwTnHf0zSpQABs9v1pS3PA7etATIwR+lIVxxg0AODcDjvSZGCcfrSDgDCngilzkdO/rQAbgT+P96mq2BjHOOxp20H/APXTCq8genrQAMy5H8qUtkHANJg8AZ9sCkwBkkDr3FAxQTtzinAgnn19aRQPb8KFz/k0CFUggfT1pONuMfrSgDA/xoABHPp60ANJ5x9O9KCPTt60uCTwD+YpNpA5B6e1AAv07Uu7JOB+tN4yeD931pcZJ4oAXdz+HrSbvn6d6Nuex/Gk8sbug6+lADywxnr83rTsjjioxkKBjv605SMj/GgBynvt7UhbPbv600deB29aXGO1AADnt6Um7BOf50gPUkd/WjdknBpIp7ClgSABn8acAcfdPSmHOen6UqcDkCmSLt5+72pSwC4P971ppI7Y6CkZhjHv60APEnGAP1oL5IG01HkdyKMD/Z9qBj2b0B6U1Tjj39aMAHoOg/z1ozx+VAhxOWBx3pC3oP1puBnOO9KfqD+NAD8gEAD9aRiMgD09aaCNwAxz70u5e/p/eoAViMf/AF6M5PT170hPGcH86CM0AIxB7dh3pdwxye1Nbtx+tA74Hb1oAerA9B+tDONvT8zTAf8AOc0ue5P5GgB24cYbvTQ3J4/KgBcdT+ftRgZx60AGcgfKe1GRgHHT3pNowBgfnShc4HHT1oATcAvI/WlLA85HUfxU0jjkfrmnEgnigBykcYHb1pM/LjB/OlHYYP5U3BxkqfyoAGPJ4/WhXXjA7+tIc/3cULgY55z60APXBGQO/rS7gf4c/jTR9O/rSg8j6etAApx+XrSHlhwPzoABz0OR60HIxwaAAnOfwxzQD8v+BpD3OB19KBgjkj8aAHZBz8vf1pRjA+U/dpMjHUdfWgFdoPt2NACKRg+w9aMjcSR39aM8H6etAznoetACkg9fT1poIycDt60vbkHp600Dn/69ADtwLZA7etKG54H603jPT+H1ox14FAEgI3Zx3obBHHPzdaaOG6jIIpSxI6E8+tABkHgjtTQ3fb29aGySDt/lTOOw7e1ADiwPIHc96bvBPTuO9I2M9B+dA7jAP40AOVuef505myOmeKYMZ7dKV8kde9AAQ2Tn1pAG3fQcU/5QDwaTK78YNAAFOQRnpS4PYE80ZBA4PSlGOBt/SgBr59D0pDuJOOOKc5HYfrSMRznPT1oAjAfBzn86UZ3fiKMgAjBpwxvyQetADjuNJg5HJ/ClyD0B6elGQW5BoAacjjJ+7S4IzgUrdBwelLlT2PIHegBu0/3e1IVIHTtT8qB0PSkcgDp6c4oAaykg+1IQetOYqc5B4x2pMj1NADWGV4FA3dv71KSMdD2pVxxz39aBgobJ5PX1oLEKCB2pylcknP50jEYGAenrQIOTjA/Sm4YjoaeMYHFNJGOhxxQNAQcEk96VVbt/OlYqRnB+93FKjA4O0/lQABW64pCjdh0FOUr2X9DQxHUKelAhoQjt3pCD+tO46kHgjtSErk8HqaAEwwGPSkAYjgHpTtwxnB6etC4wQAT8tADCjEggUbWAP1qRsbuFNJnIPyk8+tAxoDEdO9ABz+Apy4xyOho4Jzg9B60CEAb07UKGIpQRtHB4HcUikFeR+lAxdrHvSFT6dqdlfQ9aDt7A9KBEeDkj260uw5Jx1pRtyevSl4LZwfxoAQKegHagZ3fjTgRjgdvSgfeyQevYGgBrKxGR60mW4HPSnMQRxn73HFNyvA56dxQAik5IGaXa3YdvShSM5/nS5GBx2oATD54FJgkkUqlRkAdx2oByeh60DuIyNkYFADAcA05tpAGDRgAEbTQAw7ieaCHPXP3qXA7DtQ2COFOd1AIFDdjQd3QUoxtxg0rEDoO3rQIT5j/3yKQbgp+b9aczA9j90dKRWG3nNA7CNknr3FB3ZOPSnZBbJU/lSHaSeO1AhNrHmgKQByelODLuHB7+tBK8HB4HcUABVjzjn6UFTnGKcSpXJz1oJHbNADGBxjHpQUOMhccelDsCBwenpRuBHKnoOooAaFNKc4xSjb1ANKVBAGDQAi78cE0qqc5/rSrjaMqevf6UKeSeaADBKjAo2EgcdqAw2gAHp604suBkUDIypxihlbkZ/I04leRg0rYIPynr3oENCnjApNrY4WnEr1x+lAKkdO3pQA0qeuO/pSAEgYGRT2wc4B6+lIGXAyD/AJxQAmG9O/pSjdjFINuc4704sOw/WgBEBIP+7SMpJBx+lOBBBG09KRiuQcGgBjZwcinLuxwO9BOcgA9aFIxyD1oAPmJ6mkG7HJPSnEgnlTRwVHyn7vrQMaoYjqaUBsnI70qgbeQenelymScd6BCbT6dvSmqCDjFSHb0A7elNBGCSP0oAMN6du9KA3NJ8oPQ9B2pwYYAx260AJg7vxGaRgzDp3pygE5wetBwecHrQMZhuOn50wZ6AVLu6DB/KkAB/h/M0CI9hyfp6UBDzz+lPwO4PSgbOeDQA0KdxFDA4xj3p3BOf6UHHA5PFUgHmmljnNKT70053cCpAUHovX8acDxx6UzccDkfnTgeeKAAkjof1pGJzxSsOfwppGOT6d6AGsRilz83GfzpMEryf0pdp3cnuKAHKx6U4HjPtTfu8gj86A3POOmKAFYnp7UoOex6U1iOCcdKASfTigB27jgdvWmMTgYHb1pckdT27mkIJHXtQAZOKQE+nalIIGf60nPX+tACEkf8A66crH5Rk9fWkI+Xr2oVeRz36UAODZ7H86XsMcUg75xijHAoGxwyMCmtxx7UE4GfakZsjqOlAgYnOPf1pUJAHfn1pGBIJzQoB43DrQA9TgdPxpC3H4UD1BHT1oIwO3SgA3e3frmgkZOB39aQZOMkdaXk5ORQAmQOnFAb5SSe3rQQMY9BSAccHt60AOLD0pN2c5HQ0h7cikIPPPegBwY9AO9KW9FPSmhSQfrS/lwfWgYpPyggHpSLnnjtRjI6jpSL0/rQA8/T0pC3GQO1GMnr096QnnGR0waAAEjNKSC3HpTc5J5oJyc5B9KBDieenakUjdyO/rScnpjj0oAO7qOtAC8EHA703qelKQQMAjrTSORkjp60AKpPYnpS7hjp2powPy9aXk9O3egBwPBHvSLnNIuTnkdqUdeoz7UAO7DilHI6U0nGBmlB7Z/WgBp4/KlPTH+1QRkdegoJweT/F60AAzj8KQsc4/XNLn6UjAjoO1ACMfm6dqTd2B70bTnH+e1IQSOo60APzzwKXJ/So+/XuKeDg8YoAFPI4oJPH0pM4YcUHJxj0oAcWOMc0FuP/AK9NJJ4xRk9z2oAGJK0q8flTSDjAPanKD37e9ACpmlJ45ApAdo6frQwOKAFB4Ax3pC4x0/WjB9f4vWm5JJoAUNwOKcGIIqPB25x29ad0GcdBQApPelYgnIprHg8UZyCcigBew+lC/j+dGcjqOlJzjt7UABOVx/WkDc8D8zQc9MiheQOR+dAB0PA70pY9qQcg59aMgd+nvQAoYFTuHalLDHQ/nTR04/u0pB4wO3c0AJnOSR09aVSAOKaQck8fnTlGQcEdfWgBc5P3aXIwOO1Jg+negEleT2oAVSMdO1Lnvg0wE9mHajqeo/OgBzHk4HT3phOSfr60pPvTcAk80AOLZxj0pwb0BphB/T1pR7kdKAHKwzzzz60uQR07+tNUHPBHX1o/hyGGM+tACjk4wPzpoJzn26A0uefvdvWkHP5UDF3cf/Xpp4BzS8kduKTnkcdOtACA8nHpSkkik79ulGOBQIkDZBGCfxppYBun0p+QRjj603+L7w6+tACbgcALTg3t39M00HoCf1pwIx94fXNABuBPCDpQTznb2/woY4OcjpQW25zj8KAEBXbkp3pQV3DCjqO1ICSOo69xTwDuB3DrQAm7/Z/SkyNw+XvTmODjcOfekBAYcj86AGscYyvb1oDccr6c0rDOMY6etBI/vCgAVwOi/rSFsD7v60oOOpHT0prEYzkdKAFLD+7+NGR1A/WlyCDkjoO9OXHcj86AGEDH3BQpBIyo+9TyoK8MvakxjByOtAArYz8tDMu0YXrQHHZhSHkAZWgBpZQBhetN3cZx+tPxgckUxgAPvCgBxYYOF70qt/snr/hSNznOODSoR03D8/pQA5WwOn60ZGchO1APGMj86Un0x+dACIwwOO/rQSMHilDYwSejetN3DqeOaAF3jPC/jTAwx90dPSnDk9R+VIOmNwzgUAN3jIwtLkc/L3HalJ5wewpQpAPzDr60AC4xnZ3pQwzwP1oX7uSR+dKCCeKAAHgcHp60gYYzjtTlIwOn500DjqOlAAW5xtPWmlueh9etKx5PIHPrTSVOBkdKAEDcnjt60u4Fj8tNH4dPWng5JOenvQAoI67aQEFvuDr6U4fN0I6dTR0b7w60ANYjH3R970oDDj5e1OBwMlh19aBk4IP50AMGM/c7UcZztpQpzwR0pePUdKAG5HOV9KA3JAH607rkAg803ox+YfnQA8sAR8tAOBnbmkJH98fnSgrj7w6UANyDj5aCec4780DBbG4dfWndRzj73rQAA8dB19KRiOPkHTjilB9DQeOhHTsaAGMRn7g6UDAGdvp2px4Odw6DpSA5XggUAIcbuEHX0o3YzhR+ApSCWHTrQcc8j86AEyN3TtSEgAfL29acuAOo6dKHUcYIoAaWA4we3OaAw9P1pSe26nAAdCOnrQAzjA+XtT8552ntSMBjjFOzkHp09aAEBGOh/OlJUDO0daBjGNw/Ohug+YUANDDjIHWmhhu+6KXccDDDr60iOM9R+dABlccIKdkDGFHI9aQAkDkU7G3AyKAGEgrkr+tO3cHC0hIx94UrENnpQAqsMDg9KA3H3TSBgAOnT1pQwIzn9al6DEJHPy0KQMcd6GYEcMPzpVPA5H50JiEyMZ20Z4+7S5GDyPzoIyM5HTvVAIpABGwdKQkZHydqVckfeHShmAONw6elADSeo205W4+739aQ85wRSoCFzkdaAFDegNIGwAQp6etLkHnigEFR06UAIp46dPelB5PHf1pB0pRjJ5/M0AHHXb29KQEZb5adkEZ4/E0Ic55FABnn7p6etGeThTSkgnqOnrSE+4/OgBA2G6d6QuOy4+agk5+8OvrSEkDqPvdjQA4Pz07elNGD/B2pd3TBFCexHT1oAQlRn5f1pSRzgelGQTjIz7mlzgHn0oAbkbskfTihj1ASkJyxOR19aUsMYDd/WgYcenfPSkB+bNIOc8d6By2aBD+RjH5U4dOB0pg6DH6U7HbH60AIxycUhIyduOnpSsPQCjYM8jt60AAwq9O/pQWGen0o25TOP1pG5bp6UAOznoO3FA6/hSdf/wBdGAp6frQAHgYx2oHXp27CkbBxxR7EZx70AG3AJCjpSd84pce360hHqP1oAevPNOXrkDtUYPHT0704HjP9aAHEgADHSk7gAd6CcgYpAcgDbQADjPA/KkzkAACjueOh9aQjOOPSgAHQDHpSHjIwOvpS8EdO3rQyjGcH86AGtxkKO/pT04xx3GKa4znjoeKVQOOO9MBy56Y6+1DegH8NIOOw/Og4I4HakADt9aQ5BOBSgAEcd6aQDyAevrQAu4gnFCtxjjp3pp5HAoTjjHpQBJwcfKOOnFL/AAkAd/SkXjoo6UHAB4HX0oATPB4zQT83Tv2pv4UYGen60ASA8DjtTRz2/SkBGAcdvWkUjBGOlADmGPuikyQfoPSgjJxj9fpSEd8dqAEHH4Up+9nHekx1yBQRlu/50APB9AOlCkZGemaQfQ9PWlX72cdBQAucrwAeeKTcAQAPypGAwTjv60hIOOOg9aAFVhnA/Sl5I4HpTAefu0o9NvTrQA/tuPrSY+bp3pAeo9xQByeKAHDGBgdKAe2KZ6AD9acBnigA+ooI9Bzu7UgUY+72oYn0/ioAcD7du9NJBGfakHB/+vSk56DtQA7dyQB+tNHAIGKCcnp+tIMbc8H8aAHEDOMDrS46gDtTTjdjA6+lO47gUAAzn8KD2HtRkZB/rSHHBAoACOMAdvSj/gPf0oY8cj9aTHYCgAJxxgdPSnDp07elMI9v1pw6cjtQA4cdP5Uxm7d6XHp/OmtkgA0AJnAGAOvpSIcEkj9KUDgDHf1poXk4A/OgB4I2DA/h/KlB9ulNA4GAOlOxkgYoAQ8ZFI3GQPWhgMZIPT1pGGe3egBy444HX0pU+n5CkXAGeenrQB22/rQNCsOwHelQdPY+lBGR05pEUcYpBcUd+O/pQxwcAdqTg8gHr60ZzTECH27UMehxSAcHjtSEHd90dPWgBd3B+vFKvAxjv6U0DknA/KlXp0oGOPJ6fpSAcDp0pCcn7p9+aUcgey0CEHAOPTtThgH7vWkAAz7D1o4yTgfnQAp47dvSkBzkUH6U0c5GKAHkn07Uv+FM6nkdvWnKeDx+tAAMbu3X0pDznAB57CjGSfl7+lBHB47+tAACAR0HHYUKwHK0YHH09aQAk0AOyew7U3A5J9u1J+HQUoI7e1A0hCvPPr6UNkEgc80mMnFK4GDgUDFwD270AfMDj9aTcMYKnrRuw2cHGaCR2AAPlFOGOO9N3DIHPXFKGGcY6+9AA3X7vb1pSASfl5x60EqBkL29aQsMHrwPWgBflC8D9aQgFs47jvSbwV78e9DMN2MHr60AO4PUdu9A+8OBSbwM4B6etAbDfh3NACEDsvalKj+72HekOD2P50u4c5B6DvQAgAxwv45pCOnHb1pwZRzg9PWmuw9D09aADGRkjoBSr9KTd7dh3oBwM4PvQApIAA2ijPQbRQSPQ/nSAjjr19aAHAAnO2kIyMbew6UisM4GfzpdwwMA/nQAo5H3e3ekIB6DtS7gDn+tN3jG7FACsAM/KOtKoAx8tIxBznPWlVhwDnqO9ACgD+6OtBA4+XtSBwOcH86QuoPTt60AOUA449KaR3xQkijHB6+tBcHOQevNA0IRgYC0ijqSufxpSwIzg/nQHXBGD09aAFwAfu/rS8YI296YXG76+9ODLg8d/WgQgA5G39aNvP3aVSuOR39aMgnv19aADA2529qavPbPFPyCBwfu+tNGDng9KBhwD0pcAfw9qMgnof8AOKN6+h796AAAc/LQQC33aAR6H86UlWJ47etAAMA8L+tCjn7p/OlBB5wenrSAgMOD19aBCNggkr1PpSfxD5e3pSuy+nf1pMrkcHp60AIuM/dpQuRkrSAgHp+tKWGCMHr60DFGMEbaToen40BxzkenQ0Bl3dD780ABUZxtHFOUf7P6Um8ccHn3o3LjofzoAXv0HSkcAjO0daUMuRjPHvQW4x/tetADcAdqMD+729aXIAxzSHGQQD+dAgOCeg4A70ZABIFDFcd+g700MMev40AO3DPTv6044/u0zcAcc/nSlwAQB+tACjhun50hA4+WhSN3T9aGZeOP1oAUgZHyjt2oA/2RSFhjIHp3pQcdB+tACYBA+UU4Dr8vb1puRjp29acGGDkHoO9AAq98UEcYx0pVIxwp/Og7cdD+dADdoAHy9/U0gUA52/jTwy4wQevrTQw3Ec/nQAYG37vb1pcDH3fxoyAo4PT1o3AYOD+dADWXttFDKOu3vQzLtJI96GYHt39aAFAA/h7cUqqN33aQMAAcHp60qsuOhoHcXbxwtC4wBtoLKBxmkDAAdfzoEJwP4aMAdB29aQMCOlKGB6j9aAFVRtJ2jp3NIyj+4PypyuoB47etMdwMZB6etACHHPy9h3oXAGMfrSMwyTg9PWlVlxnB69zQA7C7sBe/alAAAAXt60gKk55/OjcNo69PWmwBRkEle1GPmJ2ikVhtPHb1pcrvPB6+tIAwByF7UgXk/LmnEj+6enrQpBzx+tACbcdF7Cjgfw9uaUlc/hSbl9D065oAO+cd/WgkNzt7mk3DPOeo70MygZwevrQA4Hpx2PegDn7vUetIrjI4P50qkEZwfzoGIVyM47d6No5JUdqUlemO2DzSbhzwfrmgYgAz0pXA2HC9qAw3ZOaHYYPB6UDECd80hXBzSg8/d/Sk5J6UEC7SCCCetKAeKQc4GaUemTQArZJ60Yweo7UjkjselISQSR6UALtO3qPyoZcNye9Jklenel6np3oAXafX9KUKSwPH1xRz6UdD3oAQjABGKNpPf9KGPI49+lAJPVemOaADaeec/hSMOBz2pR/SmtwM47DrQABTilC5HUUmcA54pwJxj9aAEI4zkUgBBHPc05hxkj0pMHI/3qABR/tA/hSEALwe1KMntSMCAPpQAYOMA4ox2zRjFHIGTmgBzKQDz+lABIzupGJP4GgHj+tABgjoe1NPBHzdqd7/ANKbjGMZoAFHTmgjqAe/pQucAkHqKOSe/Wga0Db1yf0oCgqef4aUAnqKXHy4x/DQJibSOjD86ACQcmlOc4pNpAPHegByjPQ96XBHfvTVzj8aUHPb86ADbhe/Smhc9D2p2cgY9D2pF4U/SgAKnIOf0oKnPB9e1HGenSg9celAAo55x0p20g8VGOM/7tODck89KAHAGkwC3UdaGJIxj603d8/40ABB7MOtAHIGabuAWlzkigAGPWlxjuKavJHHanDg4/pQAbSQST+lAHJ5oHQjntSEkE59aAHYIxz/ACpMHGM9qaWPHNKDmgBcE4+gp+3jOf4ulRAk96kB7YP3qAF2k96aV5A/rTvbmmnOc4PSgAKkdx0FJt4+8PypcnOcHoKQklen40AIy4bqOtLgnuKRsk9D2peQelAChSWGf5UhXBAB/SlHDDjHBpDng4oAACRjP6U4Lng4pueOR0pykUABXI69qAvHXoKXtxQOBj2oAMdqRjgYyKQn0HamscrzQA8Djg/xelIAeu4fgKaCff71KGJ7UIegvG3IbtSkZHUUnOB/u0vI4HpzQIQqSCMj8qVlx3pOoyQegzSt64oATBwORQq/7X6UEnGOfwFIDxzQA45JxmkUdMevrSEk8g0L2A/SgBMY496UA5pAM9AaUcjkUAC8g4P8NIVP94flSc4P0pSxGATQA1l6kkUqg4OD3pOeeKcqnnjv6UAOUMPSlKYUH2oAJ7HgdxRngEA9KAGqDzz29KApznPX2pBn3pRz1BoAcVPQ/wAqQA5I/lRkdf0pBz19KAFYEdPT0pCO4/Gjr+Ao56EUAIF3NjPcUu3j7w60gB3f8CpwBA6HrQAKhyASKUDsD2HSjaQOnY0KDngdqADBBAzSEHkg9qU89u3cU3OCRj9KAEIO480j5w3XpS4JamyA7c0DuSqxA/GkLndnHanbVHGKaQu7GD+dAgUnI4/Wl3EkfL3pdowMr+dGFx93pQA1pD2HagyMSeB92kbb6HpSgKWxjt3NAAjlV+7Tg53dO4pCqgcLRlQwOD1oAdvOen60u45Hy9/WkAB6A/jSqBn8KAELE4+XsaUvxjbQQowAP1pG2EdDQAm85+6KazkjIA4HrSkA8YNI2MdD09aAAucElaXfxgJSbVPOD69aVQCMYP50APMny9KTccZ2j7xoJUDGDSbhgYHf1oAQMwJO2gyHaDt7UfL1P9aBtIGVNADt3H3e1DNhD8vb1pcKRgA0jbT2oAY8hAOR3pVkOchRSSAA4A70Ltx09qB6Artjhe9BJBHy/rTgFJ+6fypSF4+U0CGqTgcd/WjJxkLTlCnGQaTC5Py+vegAD7QBt7UCT5en8NHygfdNA27Twfu9zQAu45Hy0Fzg8UYXrtpCVOcA0AHmEKRt7+tAfnO3v603KnPB5NKAM9O/pQAoY4Hy9vWhXO3oOlKAMDjtSADHAPT1oAVmyeEzz60hLddo6UpC5zjvR8uOR29aAGBiCTjsKPMJY8frSjBJ+XtSlV3dP1oAaHI7Um87+B3p5UAcD9KaFG7JB6+lADdxK5C0ocgjj9aAgI6HrTgoBGAaAGo+D0H3aXd82cfrShQD909PSjC9hQMBL/s+nekLZJwvp3oG3k4PGKUYYnigQ0k5GF70qsR/DTtqkYC0gCAdOgoAYGIIwvapA5wPl70z5QRx6Zpy7RjA70AODHOcUjOeBjtS4THA+lI230NACGVs4C/w00OQD8v60pA3dOwoIUDkDrQAb/mA29/WlLE/w0EKG6Hr60ALgcHp60AKWII4FNZzwNopw2k8jtTWCkjjv60ABY4OF7+tKGJ7Cl2qT909aVVHQg/nQApbodo4pS+BwO3rSNtx0PT1peDnjtQBGScfdppc4AIHUU8hT2prBQPuntQAiyMQMjv/AEpUZgxJWjaMDjue9KFUfw/qaAF35UZA6UrPkfdHSkGCudv60uFPY/nQAxnIGCo6UNIcEkUrBSPu0jBRnA70AODnjjv600MSuNlKACAcelOAB/hNAxpJ7pSo3A+UcCghcZwaF28ADt60CEyRzspA5HOBTiFJ4FBCj+H9aAGCQ4PH8NBY7hx+tLtXacKfu96NoHVf0oAQPjOR2FOWQ4J2/wCcUmBz8tClcHIPWgByuf7vek3nA47etJuA4waUBcDA7UACvkD5fSlBIJwv60i7SOQfzpwC89eKAELHptFNDkHOBxinkLjp29aaACT8v60AN3Z7dvWlDYPSnEc8DtQAvYfpQAisd3T+L1pckrwB1pBgtwD970pVwRyOlADg54+XtQGHp2FGVz0PT1pAVz07elAAW5zt/WjJ5wvYd6CVPagFeePSgaEz82SO/rTZD8pwvX/61PBBY5zSPtwflNACknHT60nOelLu7YpCQDQIXnGSKOcf/XpC4AoDD9aAEY5OR6elAJye9DY3ZwOlL3OB29aADJ29Pyozzn3o6Dp+ZobBPTv60AKCfTtQpIP4UAjJ+nrQDg0AKSe1JuPcdqGOcEUA57dvWgBvX+GkY9qeDx+HrTSOMjr2oABkjGO1KuRzj86UcDGO3rQOOlACHPpRvwAPenHGOlNIHHHf1oAbuOTQOg/xpQeSaAflHH60AOUn0+tJn1FOHbAxx60jdc4oGNcn07+tIh5HHcdKdJ83Qd6Rf6igByHHQfrQWPBx2oGQeP50MRjGO1AApPGB0NBOTkUKwHfvQTyTQITJ7LQGwOnalOD0H60gIAPHb1oAC2McUhbg8c59aVsZ4FGQQeB19aAGj6UvO7p3oH3c4+nNBOTnHegBwY4Ax2pAcDp2pQRjHHT1poIH5UAKzHuPTvQCT27UHrzjrR9Ac4oARTknjtmnHOelIp68fmaUnJ4oAPw7UDGeB3pd3t075pNwDZOOvrQAg6Ehe9OBAxkHp60m7I49aUOMgcfnQA3jPSk5PboacCCc4H50jEEY68UANBOCcfrSA/N+NOB7D2pOM5460ALk4GB/nijcQPwpc8DAo5HAB6UANBORxS84wB3o79KUnPTs1ACgnGMCkJ7gUKRt/GkJHYCgBSxPOO1IGOM8D8KUsM9O3rTS3HT9aAHFju79aQEgfhRnLdBS8dAO3rQAgJzn2ozyM0v3T+BpGHTAxQAob27UofGePXvSduD2pcj/ACaAAseP8aAx547UhPYA04HqPb1oAQfSkJz2704cc0hwO3T3oAAxGB70KQOooUjA+XofWlVhnP8AWgBOoBxShj6UE5AGO3rR24HQetADD9McUMT6d6VupwKGJboOlAApx0FCkjnj60L09cClHA/DqTQAhJx07Uit04796Vscj8qVeD0oAarYPTvSk9cCgDnp3pcZ7DpQAinrwelGTkHFKpwD9PWgHJHHp0oATqDx3pvOMYqU8ggLTMDHTt60AMJOenYUqsccjoPWlYZPA7euaDjGPagAQn9OlOBPIx39aavfinZGf6ZoACT6DpTck5yKeT6DtTe5+X6UABOaBnA47UjDJ49KUdMAUACk7s470jNx07+tPX73I79qaxBGMZ59aAEDdOD35zSBj6dqXgEYH60gGTnH60AG44HHekDnofal6jgUDuMjqO9AwD89KVmOGwBxSf560rYNADiPqaaeoIX9KCyjgY6+lNLLu6fpQIDjjilyegFNyvGB29KcNp4x+lAC8+lGDnG00ny56dqUbc4x296ADqM7T19KUjkcfpSZQD5aU43cZ/KgAw3UigDLcijK9AO3pSZXd0/SgAI7DPT1ox6g9PWkO3jjt6U7KfpQAAH+6aCMYwvpQduOKQle47elADscHj86XHotJlQM47DtQpU8j+VACEcZApOeDtpTjAx/Kk44Oc80AAGf4eaADwCKPkyc/wAqTjA5oAfjvj9KRunNJlcflSMY+h7CgAPPAXvSqM44NIxXr7+lKpUAf4UBqL36GkYHPC9qBsz/APWpGKcfT0oGCk4H1o4PY9aRSuB7H0o+Uk8d/SgY/qMgGkGccjtSZHUUAptPPagkUgn+E0YPPBpG2dv5UnynODyD6UAOAJGcUhBJ6HrQpHQDv6Up2A4x27CgYAfKMDtTRk9jS/JgEY6elIu096BC85zg0vsQelBCbe3Sg7CeAOnpQAi+4peM529qbheeO3pQSueg7dqAHE5+6KQE7s470mQTkfjxSAjI+vpQOw4E45FLkkg0z5cYxnn0pVK5H+FAh4z1xR24HamgqDx+gpQy8dOlABg84HegDnOD1oBjxxjtSDbnoOvpQA4LjHy9qMcfdNICuAfalLJjGRxQAm3kECnEcZK00MoII6UpZdvbr6UD1EIyOAaawOenSjKsM4/SkJXPSgBTkH7tB6cCg7M4A7UnyleKBC85xjvTuemD+NMJXPBpcjPbp6UAOwB27Ug69D0oUrnp29KDt+nFADgM9F9KXBx9003KY5x0FLuQg47e1ACMP9n0pQeOnQUjbe1AK+vagBwz2U/lSPnrik3KeMfpSFh2Hp2oAMk4+XvQpz0U03KcfX0pQU5PtQA8ElR8valIJHQ03KBRigMpA+npQAMB120NnH3cUjFMYPahyvagBew+XNKvrtpilOD/AEpybcf/AFqAHE8cDt6UgHTikLKeoH5UisoA9j6UAOGOfl70enFN3Lnt19KUsvUUAABIOF/Sl/A0wFdpz/KnArxx29KAHeox+dIOn3T0pAV5+vpSBlxnPan0AXGcDafypCMgEL29KGdAeGH0zRlSoAx930pACjjpThkk4Hf1pqlCP/rU75N3Qcd8UAKR7U31+XpSnZj8PSmgpknH5CgBzAZ4Wj8KR2XPTt6UnygcDtQA4cHG3vQcY+6evrTcruz70ZTkcdfSgBcHrt7cUm3k4WlynRfT0pAVzwO9AC7f9mhB147+lHyjqO3pQCmCRj8qBigEHGD0pHVselIXUMf6Ch2XHP8AKgLCkHjPr1pu3LZOfxp5789/WkAG6gQgQADFOAHpRxxijoBgUAGAOnal4yeT0pCopRgk8dqAAAAdT19KGA3UvBXBPf0pHwCf8KAD1HJ4oVfmzmkwB37Uq8NmgBCBxj0p2OfwpCB2Hb0p3UcDtQA0+2elNIAxj0p3Hv0ppAJ6UAOUDBPPFOUZ6ZpqgAEU5OmcUABCkUh2jAGeDSsBikwDjr96gaEKjJOTTSOAf0qTHJOfWmkAAYoENHtQSCTmlwOOKaFzigYpGM8nrSr0pWXPPv1pMD9RQIXjrn9aY2Mgin5GcjimsADgnuc0ANXBxx3o4z170qgDHJ6+lGOfx5oADgd6AeCQT0oOM8fnQFG0jPagAyCcg0ZyD83elKDgY9KTbjP1oAFxilbHp9KRcYOO1K3XigA6jr29aRcE/eozwOe1Ccdu1ADhgDg+lLikJ9D6dqXt+FACKM5A9OtDAbuDQNvI9qVsFvpQAm0Z6UgUBsg96djgUBRux70AN425Hr3FHAOfanFRt4PekKjI5oAaCCeDSg9MGm4GcYpR6Z6UDHKODj+VJxkmlXG0j3o/i49aAAjnigjjqaMHijHBAoAbxkdelL24z1owMj60uOBz3oATHFIQM4NOGMdM00rzwv6UCDjO329aAML1pcDOc/w8ijA25z360AIQN2M8ilA5P0pCBn8fSlA4xz0oAFXn8KCRxSrjt6UhB7enPFAC++e4o4zR1P5UAcEe1AxG6cE/hSqBjOD931pGAI4HalAxn6UCBQPekYDGKcoA4pHA29O1ADcYAx/eoUYOc/XilCrgf73rSY5/rQAuQQOfwoUdPpTcDA/wpyAADjtQAEDBznihsdBnrSsvBx60MAenrQO2gijOME05UHoenrQgx360qjAyp7UCGkYGRTBzjmnuAOB+OKaFGR9aAGqRnvSgAD8KQLg/jTsAmgBVHB5J4oxyOD07mlVVKnI7UhXBFAAMYOCaT6NSkZyaQAYPFAARk96UAAdO3rQQM4pcAKB7etADVAweO1OAwT9aABgj+lGOT9aAFwOp9KFAJpTgHn0pEwSfpQAjD054pCMYIz71IyjgD0ppUYA/KgBi4B69DS4BGcnr6UoT5sdMGl2ADA9aAEB569qQEdqdtyw9hTQOc0AKAPf86Q9DzjBpTjA5pvAB+tDGhMjPBpJMcnNL3pH4BxQBKPc9+9J8oPag54Hv6U0KN3X9KBCjGR0+lOH4cH0poHA5/SlwenJ/CgBTjPAH5UpwQRjt1xTSpB6Hpzil285x29KADnZz+opGIzzjrSdV/wDrUjfezkUAOJByRijHIxTeucN29KcqnPFA7juwG0dPSl464HQdBTCvTH8qcBnt254oATv/APWpDjHQdBTtpxyD+VMYcD/CgQ7AHYdBTxgAYApnQH6elKpx0PagBT93p+dJkAjGOvpSEZHB/wDHaQKTjOfyoGOXkkcGlIGAABTV6kHP5UpXgY/lQIMA4wBTcqF5AzSkYxgfpTSfT+VADsgLjjrSDqMY6+n0pGyRyD19KEBABGevpQAozjPH5UHrjg/hQFzxg/lQwJOSP0oAQAYHA/L6UEcEYHWlQDHU/lQw5OfX0oAQDB7flQoGMgDp6UmPY/lQgypwSOOuKAFZcsOBQQMHp19KU8NgHt6UhXg/X0oGIvTt+VKTzzj8qQA8nB/KggkknP5UAAAwPlH5ULjGQB+VKAcDn9KRV4PH6UCHA88kdR2pQR6Dp6UFeeB6dqaRx+HpQAox0wOnYUoPzYOPyqMEZ/D0p65zg59uKAHKB1wD+FKoBbAx17CgDtjoPSk24bqOvpQAMPlwPX0ppwQMgdPSlYYHB/i9KaBnBB/SgBuAPT8qXHORjr6Um05/+saULwOPTtQMcpGeg6+lKNpY/dpo4zj8eKVOT1/SgB2AMcDp6UADHQdM9KCcAAenpSEnHQ9PSgQ0deg/KlOAOg+96U0jJ9ce1LgkdD170FaCjHt9BSNtyAAKAM//AKqNo7H8qBdQwSM/7PXFKqkrzigqRzg/dpOoGO3tQIUgZHTrSkg9McU05z+PoaUHqM/pQAbQSOB0pduMcDpQoyeT29KCo6Z7elAC7cDPHal7dB0o7Z9hzikJ6nPbnigBCAOqjt1FAGcnHQUjD/OKCMg89qAFGOwpCeOg7dqByen6UmDgdvwoAUZ/UdhQFy2ePypQp2g7h19KRQCeT3oAMLgHA/KlAGBjHbtTSOBz+opyjp9KAA8ryB+VDYwf8KQgEc+npSSdTn19KAFXsOPypVPHamZyBz+lKpJGBnp6UAPbB5IB59KQAHA460HJPIPX0oUcDg8ewoAbgBTwKABnt+VKVAzijaScgfpQAJ3x6elBGSDx05oVTjv09KAMEA56elADgowTtA59KABjoPypCpGf/iaAOO/5UAKMDg4/KmrwBn0pduTkg9PSgIcDHp6UAIvQ8D8BSjqenX0pFXI5B6UpUBuh4PpQAEDrx0/woTGD8o6elBHPPp3WkQcn/CgCQYzwB0oAHHA/KmnqMDt6UoJOMA9PSgBB97t970oIGOPWgA7sc8H0o7d+vegBccA4/SmcdwOv92lPXI9Kbgseh60DEJ54x9MUgUHJwO3al2ggDml2YBwOOO1AXG4Xd2pXxjORxRtyev6UrJkcEcLQIdk9v503J3cCjd6Z4FID83XpQA7J9Oh9aXJxTN44+anFgf4qABj6AUhJzkL2oc+56U0gZPJ6etADtxAP+NIxbdmmbgVOCfwpw57nr1zQMcM8/TjmnKTnp+tNBGcEnr605ThhjvQApB44NO7dO3c0wt0xS7yeD2oACTg4X9aYzH07UoPb9c0jcjGT+dAg3n9PWlUk9scUwsQOvQCnA4H3qAHHJHSjnjgUjH5cE/rRnB6nqe9ADlB9P1oycDj9aaGB4z0oLfKBk9PWgBcnHSmknnj8qCeMfSkY8d6BgxbP3T19aVScD5f1pG755o6DAPQUCFGf7vb1oYsei/rTSwJJBoLZOM0APQk44/WhgSScd/WmxtjGD37U4N3OfegBGDdcU1WODx0FOOCMYPSmY7c9BQBIC3Uj9aPmx07+tNVugzSrwOCetAAN2CNtHOc4/WgKMepzSFs/40DHDJA+nrQhOCAM8U1XyAM0BuOvb1oAcxI6D9aaxIHCnp60pYZ5zxSNjH4etAhqlsk4NSKT/k1Hxknn86epBfNAEq59KMHrg9fWmq3HfpSg/N170ANOcE47+tNJYMAFpS3HXv6UxuSCCaAAE5+7S84Ax29aao9AelO68c8e9AwBODx+tAZg1Ju6jntTSec570CJdzccZ49aMkrjFMzxilz2FABkhs7fzNKucdB1po6inZ6DnrQA5c4/H1obOO350gYYxmgk569qAF6HoM49aYxbBwO9Ozjuenemn5h680AIS2c4707J5PtTMYPA7ilzgdTQBIjH9KCx4GP1pgYgj6UpbIHJ4oAcSeuO9IGbnj9aazcYzSAj1oAcxJx9KUFgD9PWmElvwp6k4PPagAXP+TRkjjHelwAOppHAwDjvQAoLYHHfpmhdx7U0HA6nhvWlDgnG6gAOcDjt60oLYHHakJ+UAE9PWkJ96AFLHBwO1I5PoetIxyc89PWlJDHOTQALnrjtTlDdCKah7ZPSlBBORQAHOMY/WhckDgUE9vzoRiMHJoAXnJ+vrQSSc0gII5ozntigAHcBe1Jk8YXp70LyMn+760jEdM0AKHYg5B60Anafl/WmdMkE9aVSMdO9AEgJJzinc7eg6etMJ9M0ofgAE9KAFUH07UHd1x+tIHwMgkUhYHqKABs56frTQzAnjvQx/wDrU0EAn296AJSTxx2pST129qZnJzntQGU/l6UAKCQ33e9Lk46d/Woyefxpd+P4j1oAfye3b1pATngUgbnqelG4Z6UDsOAPHHajJ5OPTvTdwHSjdnI9KBCgtu/+vQxYqR7U3ec8E9KbI529eDTQDiFAySM0ny7vvd6MDOMUh69KQCkICMGlyM9e/HFJtyfu9PekIPUD9aBisR3YdKDjOdwprdenanDB4x2FAhoCY4x+Apy7Qevf0o2fLwuKAuG/EUDH4XGc55pPlJGCOlBGe3f1pDndwP1oAMrgc9qcNmM5HX1phBHRe1PA9h1oEKAhxg0jbMZz2pMHjj9aRsjOB2oAG2dc9hQCpA/rQwJH3e3FIVz0H60AOypX73aj5Tg7u9IRwcDtRz6fxUAL8uT8wpPlwMNj6UnfgfrSAZAx6UAOwCAcnPFISvenAA4+WkK/Ljb29aB3BtmCM0AJj7wpH4yQuPqaE47fr9KBBhOoNIAvY9qcFGMAfrQY/QUACBcDnv60o29jQoGAPQ0DuMfrQMPlxwecU0BTyW9KXHGAKQf0FACgJkAntSjYBw3Q00jodp/E0Hvx3oEOGOue9NwvUGgZ6YPU0BcHgDr60AIuzAII6UBl2/e7UgHTikHTgfrQA/KZOD+FHyf3utIF5PHf1pdvTj9aAEXZjrTlKbjg0wZ5GO1OA+Y8frQA9dvZh70q4LD5h1pgB6AfrTlJ3/iKAAgbc7qaVXI+btTyMr93ofWmkZIGKAGgKOQc8U4hM9unc0ijjHtQVySf60ANO3nkdu9N+XJz607bycCkA5oAX5OMH/PFO+UZwefam7QRwtOAPPH60DEUIOM0/CgcHnNNVe1OKjP3e9AhMA96RivZu1KVJ6LTWB7DtQApK5wG7UfLtzn0pAOeR2HU0u3C4H86AAhc/e70p25OT29aQg7/AMaNpPAFAB8gI+YdKacDoe1Lgkj5TSMp4GKABihyfelATPB7etIUyfujrSqvoBQAbRjg9qkwgzg9hTSMDgdqXHXC+lADvlOOaa4BH3qAOnH60j9OlACDZ13d6TKlvvd6AOf+BetIFwc46+9AC5GB8w6UcHv6Uh6Aj09aUA8Db+dAB8oXr+NObaASGpmBjkU4ng4FAxRsIGW7Uo24+9+GKaAOP8aFGB0H50AOO317UKEGMNSY9BzihcEAH+dAgBXHBo47NTccdO/rS8noOnvQMVdpBOe1IdpIwR0pF6cf3aCP9ntQIQhSDgilAXBOe9IwJzilUEjp3NTqUKNp5z355oUoQCD2pMEHp+tIOMcVQhwK44NINgJ570iDK4x3707aFJ4H50CEwoXO40ALjOf84o2jGAO3rSeooAUFAeG7Uvy9iPemgHPT8zTgp5wP1oATAz97vTW2j+LvTyvOPcU0rlc4/WgABXPDChSOzduaaVO4cGhVOMe1AxxKA4zSBlyct6d6CDn7vejb1AWgA+UkjdSPjgbh1oCnceKGQ4Py9DxTAl9Ov1pCo3Z56UoIA+8eDSAjdncaQhSBx1oIBHQ0oxkHcaQlR/EetACYGR8ppwUf3T0HemswHRj0oEg5+c9KAHbV29DSkAsOD1pu8bT8569qN4LdT1oGKcZ78etAwMdenrSEgfxGgEZBD/rQIGAyDg9KcAPQ0xiMjD9vWlXA53HtQA9QCOn60jKMdO1CsAOGPIoLYHDdhQAmAR0xxRxngUMy9Q/bmk3Yx81ACkDHQ0YBAIzyfekLADAY9e1AcHAJbrQAgHPK0h4Awp6UqlR0J6+tISCByenegBwIGDg9KDgrz7dqRTwPmNBI253HigAcgjOO/pQpGBwaQsMcMevrSpjH3v8APFAD1AxnBpSF/umkU46Mf84oZlB4J6UAIOMZB60DGeh600MOPmPXtSAg9z1oGh5Axggmm5GDwfu0oIGOT+dNyMY3Hp2oEB4P/wBalwOflJ5poYHBy1OUjB5brQAqj5c7TQ3XGDQuNvBNGATjdQAhAwDg00YxyD/hSjbx83akBGMbjQAvtjv3NLjpkdvWjIz94/5xRuAxhj0oARQOSFPSlx8xwO3pSK3UliPwpwxuOG/CgBwGMcE8UAANnB6ikDDsT+VAcBvvH71AC5HoevrRwSCFJ+lRl+OXPWnKwGME0AKvHJU9KCM9j0pqt6N2pcjoT260DW4uOvHpTcDceO9KCACQ3cUEgtncaAAAADrSqARnFN3AY+Y05WGOGNACgDjINK2MfdPWmgjOQ3p/SjeD0b+KgBcDGcHrSFRkcE8UKw7MaUsOoJ6UCE24/hPQdaXAIyQTzQWX1P3fWk3gDhz1FAAQCw+U9aCAcgA9KTcC3U8tSgg/xHpQAoH+z260Mo44pAwBHzfhQWBxhu1ACFRzxQAOOP1oJHQt6UAjONxNACtgD7p6UvGCMdqYT0wx6U7cOfmPSgBRgHpSHaR0PWgMB0JoJ4GCaABQMD5T19abxnOKcHGB8zdaZuAP3jQAYAH3T0peMZINNLDHU9PWnAjj5vSgBGAIzg8Ghs54B696GI2n5qRiOec4oKWw8Y4zmlUjGcU0MAB85oVhjhzQIecddppF6A4oLDsx9aRWXA+Y0CDHf3pRg4OD0pCV5IPelUj1PSgBFUYJIbpSEDPenLgDqelIxHHzGgBMDJ4NKoGw8GjI5wxoB+U/MaBgQM5wabkYGQelG4FuppARgcnkUAOQDH3TT+P7p/GmR428E9KeCOfmNMQhGex6U0AZ4U9PSnEjPDHpTAwAJLfrSAcQM9O1LgccGm55wG7UqsOMsTQAoHOdp6+tBAxjB60gf5up60pYYJBbr60AJtXqFJ4pAg/umnBgCME9KarAdGPSgBdoIHBpNo5GO9GR03Hr/hTlK4OWPagBoAzyDQwAGAD69KcNucg9qGC8YbtimrABzg/NSAnd1FKT2pM4akAc8YP4Yowe5H5Ubugx2pe2Tj8aAGtknqOlGG5we3pStweAOnpQT8x4H5UAIu7aef0pp4br/F6U7PHHrSMOenegBcc8Ffw+tKcg9f0oGTnI7UDGeg9qAGncOh/h9KUbume47Uhwew6elOAUdh0pgKN3H09KRicdQPwpVx6U1utIBDnB59O1LlvUdaTPbApVHGPagBWzj/61C5457+lNJ9hS56DjrQAihiSSw60YJA6cUo65x3oGSAMDigByglevakYHbjPb0p3J7Dp6UjDnOKAGsuOMjrSoCMcihm5IA70KwwPr1oAcN2Ovb0pr7vUflQCD0ApGOTwO1ACDJxg/pRk4PPXPalHI5HfvQeSfr2oAM+uKRQdvUdKUnmlXABOB0oATYcjkce1KRweR1pxHsKbuOCPegBVzjtQMjuPwpByOcUrHnpQAYOBz29KYAcHBHT0pxPyjI7Ug5z06d6AFwwPH8qCGHf8ASgjJwAO3QUrD2HegBgBOee3pSgnd/wDWoAGTwOlKVGeBQAKGIGSKPm3dR97mlxzjHQUmfmzjv60AMYEjgjr6Uo7cj8qOPQdaASCPYetAAmc8Ht6UoDZHP6ULz2HSlxwMgcUAADYyW7jtSE88n9KcCAOg7dab/F070DAA4GGH5U5d2Ov6UL0Bx0pc8dKBCE4P3v0pDu9e/pQTzwPypTgjpjnmgdhBk9xQxJI5FKvAz796Tn0H5UAHze3SgBtvUUu4k9B0ppI29B+dAgOd3UfepSWJ4I/D601jk9KcOc8Dp6UAKd2eT69qbzkHHb0pVPPSkOOwHA7UALzjj27UvzZxkflSH/dFKOmMUANbI79vSl5weR0pGxxx2pQcg5A6dcUAIC3TIpxPTpSL3oLY6UAA3EAA9/SmAMXOSDTi3A6daaDzn+lAAQRjBH+cU4AkAgjt2pOcfhTlx39KAAglcZ/SkbIBwe/TFOxxwKR8E8KKAG8nHPf0pUDFc5/SkUng47+tOUk9qAFOexH5UcgDHp2FDMSM496Bg4470ANO7HPr6U5c/wB4flTQBnp3/rTuP1oAdHnGB6U1wSfvAUqtgEYHA7mhiSM47UANOTn5hQNxX8KNx5BGeaQYx0/SgBMHOd3ehQdoOe3cCg8nHFKFOBx0FACx52np064p2CCeR1/u0injoOlL3OBQAjbscHt6UzDc/SnHH93tSADPQUAIQwPXt0xSrux17elLgHHHYUpHfb2oAaB8+AR1owSOSOvpSg8jjv60Y4xjv60AAB4ywpoLevanbiCMdvU03qen40AA3Z69z2pyk88+lNAAxwOtKuME4HGO1ADsnccN27Ujk4yG70AnOSKHJwQPrQAgXnp39KMfN/WnEjgkUg69KAAKRjrwPWlAx60Enj5aUcjBWgBjDJ/CkIz25xTj16dvWgqMk+3rQA0A465pWA/lQSNvt9aC3OAv60AL17H86TBJGePU5oBK5+XtQpAPT9aAGsOQfanbefw9aQ5Jxjt608AnqO1ACKvHfp60jA4xjtTwMfl2NNccdO1ACEZByOgGKXb36UZPINKp9B2oAaw46UEYwOeDTiBjGP1pOOOO/rQA1R3pwA469BSKSSePxpwxtHTp60ALgN2/WmvwOh6U8Dpx2prkgZxQBG/070qD+dJIT0C96WMk4z6igBQCTnn86QqffpT15PA/Wkbk49vWgBFXOPw70BcnJHenoOmP50Y5/GgBoXC4xSgHkj09aceTwv60g4BBU9PWgBhB/SjHXPrTmG3tSepx39aABelGznp37UDgdPyNLuyen60AN24X8KFGcjHanZGOfT1pB0PHagAwc4x3Hejbx+FKcZpSR09vWgBqr1GO1PKjdikU85I4+tKTk5x2oATGPl9qaQM56c04njgfmaaDlskd6AGMM569aAMnv0pTzk4HWlGMg4HfvQAIPr05p204yAelIuOw7etO7dO3rQA3BPP0pAMtz60uTjoO3WkB57daAFUDAGP1oxkCjcQAOP8AOKN+QRxQAhHQZ7UY54z96kHJHH61J1H3e/rQMaFPWggEjGfrTh0xt/WkI5HB/OgQ0g8f7tN5xye3rT257fwjvTcgr07+tAAQAfxpcegP50hHt39acOM8dvWgBMEnPPekI6ZHanLwe350j5HQdvWgAwT19vSgD2/WgnAIx270oOBn+tADW6AD0pQp64/h9aD/AEFLnAIx29aAAD0z+dI3TA9KUZzwKCBxwPzoAZyB36+tIox+frT+Co6df71Io5PH5mgAAGOnan4yBj86aMYHHb+9T16AgHp60AJjjkHp3pHUAf8A16XdwTSMQc9OvY0ANK+g70IpPY9O5pQcf/rpV9cdvWgYY4yB2oA6fWjOSQB39aUEcZ/nQIZj6/nS4x/Tmg8k8DrS/ePT9aABenXt60u3p2oU/Kcr29aMkY+U0AIRwaABzTuOeO/rSY4yB29aAE247d6AnH3f4aUnnH9acBwOO3PNADQCRjHagqcmnDgdO1DDn/69ADMEcj09aQLnPFPOMdO3rTV4JPr70ALjBH09aUA5zQx6fKenrSE9QVoAQD5gAe9BXIz+tKrHdkDofWnHG3IHegBgXB6dKbjB6VJwccdvWmHBPA/WgBNp7Z/OlUHk9KBwACPxzSqfb070ACr83WnFMjHoKAfm5H60rMMcY5oAjXsecUg4bIzT8DB5NIAu4HcaAEA6HaeppQBnOD05pQOB160dDnJoARj6A9KC3pnpQzD1PT0o3A55PQUDsNzhTwfzpCBuGQetPwNp+Y0MFLcsaAsMKDsp6UY+cHBp5xnIJoAGQQT70CGYxjKngetPUADAB6DvQQOM5/EU5SP7x6UDEx/snp601hkcA9PWng/7X6Ujc4y36UCGkccqeAOtAGcfKevrTsAjqelKABjBP5UANbgDAPHvTAc44qQ7SvJ7+lN4yMMetA0MQYJIWng4AIB6UDAOdxpVKlQMnP0oBodgY4B6U1/u/dPT1p/HYnkelIQMfeP5UCInUAE4/ioQYxhT+dSMoOck9elIqqACCfyoARRkdD270vvg0oBz97v6UMc85P5UAIpwB8p/OlxgEYPX1oXBAOTwfSgYBzk9aAFGCOh6etNGNp4/hp/GO/T0pFIAPJ6UAMYZb7ho98HrTiFOOTRlBkFj19KAGrnb0P50Zwc4b86UMu3gn8qOCchqAEXOBwelCDK4CnpSgd8np6UIRt6nP0oAUgZ6Hr60Yx2PT1pcg85NBOcAk/lQAwd/lPSnA8n5T+dIMc9enpQcE5yelACDgZAP50cF8kd6XcB3P1xSbhuzuPWgBCoI+7mlGRjAPSjII4Y9aA3TJPT0oARD9enr9KcvrtPT1pFwDnJ6elOA4HXp6UANxgE89u9IepOD+dSdc8nr2FMIBJJz+VADCOmFPalTjgA0uOByelOTGMljx7UANUDjANOIGMhT96gEZAyfypxGRznrQPoIoGfunpSMAegPSnKcdzQdpwQW6UCGOPRe1MXC8YPUd6lJDHqfu03OByx6igBAAD07igrnkA9PWnDaT95uvpTsc8Ht6UAMwARwfxPtSN7KelSBQDnJ/KmNjg8/lQA3bnnB7UY9QaftB/iPGKXaCMEmgCMnHOD0pQOvB/P6UOABjcenpShsgkk9KAEA6fKe3el7fdPHvSryNwJ+lB28cmgBo4A4PX1pqABs7T+dPyAB8x60g453GgBM8DIPT1+lL1HAPbFGRjgnp6Uoxxye3agBDgr900pJ5AzRt4wSevpSt0OSetBXQaM8H5unrQBxjafzp4AxnJ6elIABj5jQITAIzg0ozxgH86CevzH8qVDkDr+VA3YbtIH3T+dAwMYU9PWnYB5yevpQABwSeB6UEjFPBGD09aD1B2n86eMAH5j0pCAMHJoAaWwCMHt1NOU8dD6cUmRzye3anKRtyCevpQAqr1+VvzoUEAfKelKcZ6npRkYHJ6UAIuMdD09aABk8Hk+tC8Dgn8qcvcnPWgBhGF+6fu+tNXqeD0qQ7em4/lTQBk/e/KgBCMdj0FIAP7hp5Az36UmBx16elADT97vwfWkLE8DPX1p2QSeT16gUHB5y3WgAU8Dg/nTVGOx6VICBg5PT0pq4zkE9P7tADcDOcH060KOvB7d6cRnqT19KFxyQT27UD6AucnrQRgYANKOvU/lSnp0PSgQhbA696Td8w9qAPRaQhsnj1oAUuDjBo3cYBpuDuA28UpBHG39aABjz16D1pwOSRk9KYwPPHb1pwBz93P40DQ7oOG/Sh2O7Ge/pSjJX7v60j9encUAIT6GkDYPWkbIyNvcUZOcY7dzQIN3IC/hTg3+12xTcZ7du9OAPp+tAChuPvUMTg89KPT5f1prHPb0oAdu4xn6UZxzTTnH3fTqaCDxx+tACl+OD+lG4nHzfxelIQcY29qVQ3p3PegA991AJAGDQATn5aCDgYWgBwJ45pM8cEcelJzt+729KVixXO2gBGPH40obIGTSNkg/L+tAz/dP+cUDF3e9NJxQSey96Q56AdqBArYxnilDZFNXIxgd+5pVBAztHX1oAeCfXtQHwCN3YUg5xgdqQA8jae1ACls9/0phfGck/lSkEEYXt60jIcHC0ACucYz39KN2T1/KkCNj7vegA/wB3v60AODAgfShG7UiAkD5e3rTlU46D86ADcMfeFLv9KCp3fd/Wm4JxkdT60DDd1NDHng9qFDYJ29qUgkn5aAG7jnGaQMc/e/ipcMeNtADZ6d6BAG3DOe/pSqefvdqTacDC96dtIx8vagAVuCM9qdkdcimgN/dPQU75s9O3rQOwm7qP1ppY880pJ5BHcd6acknjv60ADEEdelKre/akAJI4/M/SlCk8YH50ACscghu3NPyMDDc01VbPT9fpTgTjoPvetACg8fepCxGOenelOc9BTXHt2oEIW+bG7tSFv9r8hQcg/dPShVYp92gBc5br39KduGD83akKktnB6ihgc8CgBN/zDmkJJI5oAbcOP1owTjK0APDehoLAdCPzpDkA8frSnccZX9aAI5GOOPwoDDB57UjBiBxS4Pp2FADlb3pS3HXvTeR0HfuaVgSORQAu7KgZ78cUmQD1o6jkdz3pADk/Jx9aAELHA+lOGeOO1IVbA+XtTgvAyKAAk4IJpGYEHB70MPl+7SMCSeKADf2z+lAbIxmkAbj5f1oAOOVHX1oAUnnrSo2AATSYPYDp60KpwOO9ACh85wRRuz37UmDjO2lUE4O2gBVfgj2oJGOopFUkdO39aVuoAX8aAG560qHjr34pCp54pyq3TFAAWAPBpA+QMelABz92gA4Hynp1zQAqPxS7gBjIpmDt+7S55Pynr60AOZs9Dn8aZv5OTRnodp/OkAPcUALvyRtbtS7gcYNMAbPT+GnAHn5f1oAVTzjPelyMZB7+lNGc/dPX0pecD5T97uaADeQRg0K3cHtQAcj5f1pAD/d7CgBe30oDYzg9KME/wfnSYPI20DsOB+Y4NKTleD+NIqtk8elOK8dMcUCGjJ4pOSc0oxmgDkAUANw2Rx60EMMYBp+OBzQBnoKAI2z79Kdk889qVl6fT0oCkHr+lJjQm5sZoLnOcntSkEjA/lTWHzZJ9O1JPUALHqPWm5IYYpwGe9CAZxVCBScjn607J2/gOaQgDpTwBnigBuWHemkkd+1PHIppX+VACbmAP0FO5wMUADHU9KUDtmgBG6UAnPXuaD060DPB3DrS6gALd89O9BJxnJ6UufQ03cSvUdKEMdk4ppJ2Z9u9OGT37dqCB0pghrMcGjd6c8f0odfcde1G3jg9aAGhnP4UEtnOTTivJGRSbenNAAN3B96VQ22lUDPWlAXOBQNCAscZoDHGc9hS4A6Uh4B5xxQSJk5FIQcHNPHTkmkA6gmgBoGPzpBnP48GnBRijaCeDz9aAERmwPoaVSduR+v0o2jAx6ULzxQApLkn/CkGc/zpTjPB+tAGetAxADjA9KUk80Kue9KQCcg9qAGYJJxQv3s+4pxA5ApoGGwT3oEALbR9adnGMelC8ZGT1ozyADQMFPHTtSZ5xigdcD0peoyDQAws2Tj2pAzbjj1pwXdn60BQWPNAAu7I4pwBAA9qAB0B/Wl4PQ0CG80uW9f4qOMilIAHB6GgeoZbNNcnt6U7kikK5O2gQ0k5yPSgZ2nilAw2M9hRjjKn9KAFyc9T1FGSDkUg69aevegBMEEEcUhzxj8akA5pjAAgAfpQA0lsflSb265pSox1oH1FADSTjJ9OKMtk/SlYLjI/lSgDBB9KAEycYFDE44z2pygnoaay5HX9KAAMQOD3P8qFZskkdqMcDnvSgk87qAFyQOPSlzjkGmZ4654pRnuaAAlttIS3OT6UpAPFK2ACAe9ADQWGMelKAccUoHA+nrSgZ70AN+bjBpAx4px6de9C4IHNACDcVyaVc5oGOg9e1JkDtQAKzY/ClLEkfWmjocHHFOBJ70AAyATmgHg89KMggjdmkGcZoAAzGkBbAOaGAzkHtzS7eAPagBoJxnPPFCs+SaFA6A9KcFGBj86AGgtQM459utOIz+ApFA5oAQbj1HQUuef5UEKCOe1GAelAxATuzz1FOBbb360ig7uvenKOMZ70kAgZgAc44pqk/oKceG49O1NAOeD25pghylvSlBOTxQBkfT1oByDzQAoY5Ofags34UgILHmhyACc9+KAAYHUjr3oAXdj3pFBH505QQetAhQF4xj3pwC7ccUnAA5pQTzyaAEZVzyKTauc4J4Halbd6mgZJOSeg70ANCrjkHqO1DKAen6U8D5c7h+dNcHPDdvWgCMhQT7ClG3OD6UpU54J6UKpzQAjAfp2FOAGeAfu+lIRnAGelKEbHOelAANo//VRgcfQdqUK3Ymm4OO/SgBRtx1pQFH5UAHHU8DigZIwD2oAQhcD8KAFIGPX0NKc44zSAdBk/nQA0Fd2P6UmFCjjt6U8Jyev500rgDB/WgBRtIH+FLkbc5pqgkAgUvzepoGKwB7Hr6UBRjofyoII65605QSAQT1oENIUHoevpTTt6Anp6U/GTwT601hz36UACbRjml+X9fSkUHAwSOlAznqevrQA7AAOPT0pMLg/T0o5xxSrnaeSOKAGnaGwD+lHy8nvn0pSORzik5IOG70DQgx1x39KUbSeB29KaFOMZNKBzn0xQDsOXaVBA7elIAAD9PShcgABu3SgA+p6UCA4z09O1HGeh4HpQw7ZNGD+lACrjBx6elBx6enamruycE9KUk54zQAEjv+opuVznNHpQuS33j1oAX5QM570fKccdvSg5x97vS4ORg549aAGqMH7vb0pQBjgd/ShVOaUA4wTQNAoX9aPlzwKcAcH5vpTcHdwaAYDbxg9qMjHA/SjB4AP+eKMMBwD+FAhMrke3tS5XHvn0pMHORS4OOT3oAUYI/wDrUYXg5owcZyevFGCCME9KAEwp59vSk4xyP0pfm9T92gDgkNj8aAE+UnnPX0+lOBXB57elIQc5z3pTuGcZ6UAKCueD1pG2YoBYsOaQhuMHtQANt/lQNuaUgjnJ7UmMjBNAxG2kf/WpQFwT7U0gjGPSlUHH3j93igQq47GggdR6+lCg54zQQ3GCetAANuB9aQBc8CnDOAcn71Iq85z3oATAwMeg7GgAA9P0pdpwMHPFAU0AIduDz+lKxB6H07UhyR3obPqetADl29jQMYH+FC8AcmlUEjOe1ADTtIOKBt4w3elcHse9NGePmoADtBz/AEpcKfXp6U0A5OCadtweP50AIFX9PSj5cj6elOUNtPXpQRyOT0oGhuVyf8KFAC5/HpQQckg0Lkjhj/nFAhcDOP6UBV2jgdPSggj160c7cZPSgBAAO3alwM4A6e1A9dx6UEYPXv3oACFz+HpTV2kn6elKRzwe1IoPXnpQA4heMDt6UqhSf/rUhByOvQU4Bhgc9KAEVRnn19KUhQOP73XFKoOQcnrSHJBO49aAGlQecfpSADOSP0pxxkEGmD6mgB2FxgevpSDbg49fSgDjqetG088n60AIp+Yj+lKTxjB6+lJjDUEcDHpQO48cc+lJuw2cH86cecDOeeaYcluKBC7iMdfzpwbn/wCvTDkgEevpSrkDFADy4PGPxzS9ScjoPWmZPp2pQxyR7UAPBIX6UjNz0/Wml8Dj19aRn+bp+tACn73I6e9IDhhgfjmgsSSfakUsGoAXPA+Xt607Ge3buaYSc4Apwb+VADhjjjt601sDHy9vWl3YzzTWY9cUAHUYPp604EEcDH40wHIOBSg+lACluOFH50Z4xgfnTSxHQfWgOflGR1oAcvyk8frSMQVB/TdQD3zSFiQDigAyP7v60nbBA/OlG707UhHOaBod+A6+tOUrtHy/jmmtnGScc0Kfl4xQIXIPBX/x6mscdv1pSwz179KYznofSgBykccdx3o3ZzxSRtngGkJOePWgB4YDjb+tC9DgY+X1phc5pVZgp4PSgBx5YHnp60hPB47/AN6lye47Um7r9fWgBARjkd/Wjoef/QqAeTz3obcTnFAwUkDHt60K3BG3t6mjJ29O1IuTxigQpOD/APXoyBjA7f3qMHt6UHpk+lAAuOeO3rRn5uR+tID1pSecA0AAGOQv60g+8PY9d1PPTGKbznkd6AA8j8aVSOOO396mbuPfNKH+YCgBwPJwPrzTgeQSPrzTFYntTi2VAoGKOQeO4703d82cd/WlB460gPOMd6AF4wMD9fpRgAfdFBJOKM/KfpQIQEZHHf1pckD8f71MLfN0708H5ce9AxRkDp+tIT04/M0oJAIz2pDnIwKBCMCeozhfWkXIGfQj+KnZJPPpTWJx170AOJG78em6jgjp29aQ5z0pQcnkdqAAfeyBj8aRiMDI7etKpO6kbJ4oAD9O470Z45H15oJz1pcnnigYxjjA9vWlGcE/7PrSMe+KASePagBwwedv60E8dP1pFY9sUjNkCgQ4E8c9/WheCeP1pA54Ge9Gff8AGgBSxwBj/wAepcjHA6e9N3HAGMcetKCemO1ACHGOn40E4zkd/WlbnJ/pSOSOnrQAqngYH/j3tT1IwML29ajRiB+FPGfXtQAjHHUd/WmjJxx39aexPIHrSJnGAKAGjABGO/rS7tvGP/HqCOfxpGyeRQAq9MD09aDyR1/OhTgHHp60bgep7UAGcZ+v96jPHT9c0m4nIA70A57d6AA8Z479jQDhRn0/vUNlqO34UAICMH5e3rTgcN07+v0pq5Ix7U7nnigA4wOO3rSL0Py9PenEn8hTVPJoAceDgjPApVYDoO3rSHJx9KXNAAG+bgdD60jHI5/vUDP6+tIcnkjnNACFzwMn86aGOeh/OnEsSMDpTec9O1ACqe5HcfxU7IOTj8M01c4xjvTlJIOTQAZAbgd/WmucA4OOaXOWpJCf1oAfyBxnrTed3KmnMW9j+NN3ENnFACnHB2/hSjJGMUgcjAzShvTFAAc54H60hJBPXp607cfXtTWLZ60DE3kLyD1HpSFznHPbvQdxHPr6UZbIG6gQEnkbT0oBOfun65pQc5A9KACCOB+dACZJx8p6U4E9Np4FNOcjIFLyPTpQAu4j1/OkJz2NAJ9RSE9On5UALzgkKe1Ck4xtNLk4wcdKBntjigdxCeOR37mgNgrx39aU5x1pvPGD3oEKGPv9aTdwMA/nSbmJ+lAZsDn9aAHDt8poJIH3T0oBIAP0pSTjt0oAGJx0PX1pFLYHB60rlucnvSKWHOR+f0oAQs2c4PWmszeh6U78R1pCCSOewoAahPHB6+tKCScBT1pygjGCOtHOenegBDkAnH5mhScfgP4qdyRwaQZ9e3tQAZ9jS7uDwevrTTnjpSjIBy3egBQSR0PWkOe6GhSQME8/WlLHPb86AEwcfdPShM/3TSgHHbpSKSPTgUAKc8fKelBJxkA9KU+2OAKbk4wcdKAEBJzweaUsc/dP1pATznsKU5LdvxoAXLeh6Ugzn7p+91FKCfypOQ3UdaB3sMJOO/X1pcnI6/nSMSB97vRg5zkdOgoEKjEnkGnAnA4PSmLu/SnqTwM/rQA5TxkqeooyQcbSaEzjnjkUAnceR1oATkAYU9PWjJ28KelBOMHjpSAnGeOlA7iEknoetLk46d/WgE5HFHOOSOtIaasOVjjAU9KMng4NIDx1FBbHRhTJHMMH7p6etICcZwfzoLNnqOnrSBiRmgAbOfx/vUuevyn86CTnqOtG45OMUAAJBzg/nQ2ePlNCkhs5FKTnHIoAQgj+E9qTJx909OtOJOMcfnRyOhH50ARtuPO00oyedp6UPnA4pRnHbpQAi5H8J/OhmPp39aVSf9mkJJGc+nagAVjgZB60gY5+6fypRuIAz3FIM9Dj86AAEhQCD0pV6DCnpTRkAYI/OlBxgcUAK2f7ppGJ7g9fWlJPPSlbPqOtACLnGdvb1p65/ummgtjt09aVCR1x0oAXnAyppV6Dg/nQScduPakDYAwRQAhzuxtPWkOcY2mlLc9R1pOcYwOlAAp4yAenrQT7H8TSKWGcHtS5bPTt6UABbAOB39abuOM8/lTiWGTnv6Ug3YoAD1+6adjKjg9KTnqcfnSjJA6dPWnoAiAkfdPTrSnrjaetAJx0HAoB+lIBDu6YPSkBOTwelOJ+nIpoJBPI6UAOLHjg9BSgk/wnpTCxyPoOpp6kk9KABc+hoI4xz19aUZ45zzSEkDqOvNACYwRwenrSAH+6acCcjmkVieaAEHToetIC2CMHqO9OyQKaS2DigBMnd900SEkdPek3MSRihi2D9KAJDnpTTknPp3pxPb3pON340AByMD39aUZ7kD6mjOcYpc4/OgA57HtQc5P8qRiAenagOCSMdqAAgkf40EHdnNBbg8UMcnOPrQADPr29aUA5/wDr0m7r9KVWwwNACEe1LyB+HrSMw4wO1Lu7f7NACEMB/wDXpDn06DrTsjqPSmEj0oAcM47fnSjkf400NweO1OUgdaAAg46U3qBx+tKzD0pAw4GO9AAAeoI6+tABIGMfnQG7gd6UHIHHagBVHAOaOcf40bvajcCc0ADA/wD6zSDpx7d6Hcdff0oVhwPcUAAHPHr60FT0x29aUEZ60FgR+FAAoOBx3FJzTlIyPamswySKAAjnHH50gBx17etBYZ6UIwAI4PHpQApUg0h6HkdaeTz0/SkL9cDvQA0BsdvzpefUfnSAjGcUFuelAC84B9u5oXJ4HpS7hgfSkVsk0ALtb1HbvSEH9PWnFh1+nNIx4/CgBgU88UuDn8PWgEcmnHrQAmD6dqQg7scdacSMZx2pu7DdO9A7CEEDj19aNp44PSlLZGcfxdKAV447e9AhiqQenanjOeB2pARnIXFKDkDA7UAOXJH4igZ3Z96AwxSK3zZNAAcgD6etGDj/AOvTs5AxSEgL+FADVHtTsED8fWmkjP4+lO3Ar+NAwwT6Uhz1HpS9vxprMM5x2oAOc9O1HOOg696QuM8D+Gk8w7elAEhzu7Uc9vSmF8tyO9ODA/lQAoznOP1owemO1IrYPPpQxB/KgB2TgED070vOOP1NN3jA57DvShxzg0CEZePwoCkDGR09aGYY60oYHP0oGIAf8mmnpinZANNZsjPtQIVQcDBHX0oHpSowAHB4NJuGdwoAbyRkUAEDIPak3ZA+WlVugoAMHnmlOTn60hOaViCaBirnAx/OlQH0/WkQjHJ70qlT9aAFOSORSDOB7e9IX4zQjDIx2oEGD1P6mjGe3QUZH60pPtQA0DOeR09aCOe3SnIwAI9qCc9u1ACbeD0696TBINO3ZyQvf1pARjmgBOc//XpRkKDjt60EgnNIWGB9KAFHIPQ0oJzk88+tMVuuc0/IJJPrQAhGevp60gB646U5iDz7U0HqcUAIc5GR2705cmmvyRx2p2RgcdB6UAKpII57+tBJxnPf1pgYK3Tv6UvmZ/h70AKc+v60gz0z+NL5mT+B70gI7AdPWgAz0/xo+Ygnjr60ZBGB7ZoB6n3oAAvPWkdTtwPxo3DdnFOZhgj0FAAQMdB1zTeAwyKfu7Yppb5v/rUALxgDFHbGKTdgAZHXvQWNACMRmjgk4GeKR2PXHalDHkkD8aAFOAvNDEEj8KTfkcfyo3Hd0/ioAcOe1GADnFIHGOn60bvQdvWgBCOcAdqcowenamljnp2p2T1x0oAAARkKelIRk9KUH6UjHjpQAoHHT86MY7Um/APH5ml3cdKAEfpnFNBHy/WnM3HFJk5AIH3vSgAAGM7f1owMDApVb6flQWIUHA6elAAD0OORQemQKN2OcdqCxA6D8qABwPTvSL9O9KzE5/CgHn8fSgAXGeBQfTHajcew/nSFyDQAo5HSmtgk/WnK3sKRyc5x3oGhCwz/APXoUjB+nrSAnP09qchO3H+z6UCHd+lJkc8frQSQe1G5sH/CgYgIPalbGfu0biQeO/pSFsnt+VAg7dO1KnPbnFNVyQOKVX+XIGePWgB/4elIR7dsUhY9x0pd/HSgAUdeKcw54FNDY6jtSls54H4mgAccYx2pnBbk96fv54HSmbm39O9A0xAQeff1pQQDgDHHrTdxxn39KUOQRj0oEAx1ApwA29D0pqOegHb0pwZumBQAAdcj0oXrQCQT06CguQT/AIUAOAAAytIQCOF7Uhc8c/nS7sdulADOrdOpp3bPvTSxzj0pQxxgetA0LkbelNYjOPalyaGJyBjt6UANJBIAHQDvQcAdP1pcknkdhSZGM4H5UCA4z0704EcgDtSZO4cD73pQW9qAHLjPTtTSeelG4g/hSEkkcUAOPTp2oB4PFISSOn6UufagAY98Ht3oGDk47UhJx92lz1H60AAbHb9aDjH/ANekDUpY4xigBQeBx3pOM5xil3kgcd6TOe1A0xMjA4pRjI4pNzBenalJ9FFAhCfakc5z/U0FuOlBfI6d6AFQjjg/nTk+n60wMfQdPSnq2BnrxQAEcYxSLzjjvTyewFMDYwKAAYB6d6dxjjH51GGb2/Cnq54oAVSACT6UFR6UBjjj0oLY4wOlABkd/X1pR0pA55I/lShvlPA6+lACMAecU0rgDjtTt1JuJGD6UAIg65Han4B429qYrcdqcrfNQApwT07UmAe344pxbjIpAxJwaAGsOny9hQ2MdO2KcWz+VIWoAjON2Se9Hykceuadkl+vekIwOR39KB9ADAHGO3rRnnilGR2HSkH0HT0oEHBFAXA5FAJB5A/KjdjPA6elAB36UrHg8dqbu5IwPypJGJU4PU+lAEoZccjnNISmc7e1BApOc8elACllBAwO/ak3LkHbSlRgH+tJtzxg0ADFcdO1BKZPy9qCp9PxpMZJ+goGKpQL9307Ubk3D5f4qQrheBRtOenf1oEKWTOAv6Ublz93saNhPalCjP8AKgBCV6he1O3IB0pNo4zRt+vSgBwKE/d7elNYrjhe1A68CgjjGO1AAxXqF6Yo3L2X9KCuR09KAuOi0ABKbPujp6UhZcglf4qRunHpSEZIH+1QMehQE8fpRuTbwvY0m0dqTGEB9qAHsVxwO1ISm0/L+lAA9O3WjaPT9aBCkoQcr+lKNgbOO/pSMOwFGMUAGVIxt/SkYpu4UUu3mmkdKABWUY+X0pcrjlaaOo49KAAScDv60DsOTYABt/SlRkx93+Gk28/KKAvHA/hoEOOwsCAPyoYoQTgflRj2pp6Hj9KBi5Qqfl/SkGzJ47+lAHBA9fSgD2oENXbx8v60Kybfu9qXZwCB2pFGc+w4oAdmPcTt70blz939aRhznHegg56dqAEDryQvanhlGfl9O1RgcnjtTwOSQPxoAcrR91/SkzHv4XvRt9vxpADu5HegBDs28Dv6UDYCOKcFBGSO9Jt+YYH6Ur6gIuz+7/D60AqD07UBc5+XtSlcDkUwDcnJAHQUhZMnAHX0oxnPNIB8xOKAHZXj5f0+lClcY2/pRtyB9KMccjtQAgZMj5c8UAr12/xUhGMYHagjgcfxd6BjgU67frSsycALTQAR0+lB5xQMVinoPuimKyAcr6dqUjJ/AU1lwOB+lAhwZd/3e4oJTsBTSuGxjuO9O2+v4UALlcj5aQleML2pwBLYxTdvIHtQIdlMZ2+nalAQ8hf0pNvNOA44A6UANOwj7v8ADSlk/u/54oZcjAHakC9cjoKAAbSfu0uUCj5fSlUDHTt60jDigBVKY+6OvP5U0bNxO0flS4/9CpMeg/SgAyhUYQdDSsUxwO1Nxhc47UoHPTtQA07dudg/Kg7ecrTmUYzj9KR1HIFACKy5Hy9qchXbgL29KQLx0HSnKozx6dqAHBkIzt9O1JlCAdtBGBwKQc0DWjG/L1K/pTkKDjZ344pNo5AH60FeflFAXHKUx93+GkbYSOB09KRc7Tx2pdo9D+VAgyhBwtLuTBwtNxweKACe3egLjhsznaOtNLIQOO1BHGNpoK9OO1ADVZcD5fSnIyAnj9KaF6+wox7UAPDIBwvb0oDLu6fpTcc5x2oHXgfrQA7cpH3f4aXcnOV700AE5x2pSAeMfjQMAVznb/FSsUPO3v6e9IFOeR/FQoJHQ0CFBTgbaF8v+7/DSEcjFNzk4A7UAKQmc4HGKPk+bAHT0o7EkdqPX9KAGjaCRj0prFcAlehp2CW6UkikgnHegBf89KBw2M9+tKPp+YpQRuyfXrigBMgEc/rRxx9adngYoz/tfnQA00hxjr2HenlvU9vSjOSR7CgBi4K5BpcfNTx8o6nrQTzkevpQAzgd+3PNHGQenFSDrn2pBwR9KAGZPB9qU8ryOw5oPbin+/tQBHz/APrpOcfhUmRwQO3rTWOOKADkDnPQUEgc4pQ3B47UBsdP5UAMbp19KQ54yT17/SnM3GM9x2pA/Tp+AoGIp9T+tGQMfSnA46E0gbABOelAhQcAc9qCeOfbrSqwHAPTrSkqo+YgfU4oC4jdOB3pB/hTj1xt7+lSKjEfcP8A3zQF0RYOeh603J469KmZGXO5Mc9ximOu1sYxgU7MCMZ4we4zQCR1Jp/PGB3pCcNgnvSHcRW460qnI69qC3oTQHwOp6UCBm5GKQtwTn9aUsCRyaN+c89/SgaDIwcnv60A9welCnj/AOxp3J70Axq9Og+76UgAx36U/IwOe1IuOcZ6UCGsOe/WkOevt1qQ49/zpDg9u3rQBGCcn6U4HBPWlGOfp60Hhu9AADzxRuw4we9Lnjihm569/SgBAeM/7VOHakByOp69qcGwQOaAGr1x7UdcEHFO3d89qUYOM0ARevtijHzEEd6k9/cc0ncfWgBoJwBz2pR04NL6fhSg4BOTQBHnn8O9BPHX+L1p+eRz+dHB7/xUAR+wP5UE8jmpR9T09KaSRigBgPqe3ekGCuQacW5wCelCnCnk9aAE/i696XOD17U7POcnrQST1zwKAGg88elISRjrwKcDz17UFumKADccZOe1Ln60px36jFAIPb9aAGk5A49KUAEHjt60E4A69PWnLzzk9KAEDEHr270jNgHHpThxyPSms2OB69QKAEBzwT39aaG+bGe/rS7sYye/pSg4PJNACZAAx6etKCBjntQGyo5PSnAnAwO1ADScrnNI3P8AF3p2DjH9KUnB6nr6UD6DRn1pVOQO9OU9DzQp+vSgQ3OAO3FAPAw3WnHGM5P501W6c/rQA3dxyeRRn0P604cmgjr/AIUANU5BGf4aCwBGD2pwIAOT29KCQMc9vSgBuQAef1pQQR1/Wjfgnr19KcjHHfrQAzrxilX7oz/dpx5PelB4H09aBkYBIJOeBRwSR/WnZO05z0oJwT160BYb3/ChMA9acW9CelIrHPX9KBADn8hSjGenQ0uemM9KUNgdeaAEX7w+tJwFyMU4N82Se9IzZXPPWgBmeRg0ikn8qduAIGaRCP09KAEyc4Gfz+lN3EE1KcY5x+VMIC5xigAU5NKwyPxpFYhuo/OnbvloAMnPUfTikBO7+tO5yOP0puDnJ/OgA3nA6fnShs9x2pMHAI/WlVSQOn5UABODxjp6UuSSen5UhH8qCCSeB+NADhnb0H5UvOccdfSkAwvAHWg9elADuT0A6U3JB5AoweoA/GnKDkHjpQBGc8Zx2pQzZ59PSlYDj5aaQT1H50AOzgcAdKa3A6ClGR2HHtTT7fyoACzc9OlJuwMnH50uMjg9qTBA60AIzHHUdfSkVskA4ps0kUMbyzzJGiKWd3YKqqBkkk8AD1Nfmr/wUY/4OMvgZ+zpdX/wo/Y903T/AIkeMYC8F34gknP9haXKMg4dCGvXB/hjKx+sh6V7GSZBm/EWMWGy+k5y69orvJvRL136XOTG4/CZfR9pXkkvxfoj9GvFfjfwf4A8PXfi/wAd+KdM0TSbCMyX2qavfx21vbqP4nkkKqo+pr4Z/aa/4OOP+Cf/AMEA+k/CG91X4r6vHJsdPDUJtdOTHUm9uFAkHoYkkB9RX4a/tPftoftQftn+Kv8AhL/2lfjDqviSRZN9npkkvlafZeggtY8RR46Z27vVj1rzMucc/pX9BcO+BmX0oRq5xWc5fyQ92Po5fE/lyn59mHG+InJxwsFFd3q/u2/M/Tr42f8AB0f+2B4nnntPgn8HPBfg+1cMIZbtZdSukBHBLuUQkc/wCvmD4gf8Fpv+ConxGdzqv7Xuv6fE3SDQbW2slX2Bjj3frXzA79sdOlMO4/wMfwr9PwHAfCGWRSoYKn6uKk/vld/ifO1s8zbEu860vk7fgj1rUP8AgoF+3dqs5ub/APbO+JzuTkkeMLhefoGAra8L/wDBT3/gon4QkWTQ/wBtD4gjb90XeutcD8pA2a8GLopw0gHsTU0Y3HcuSPavXlkGRzjyvDU2v8Ef8jkeOxy19pL72fbPww/4OF/+CpHw3u431D4y6Z4pt0zvtfEvh6CUOPdowjfrX1b8Ff8Ag6+v45obL9pH9keCaEuoudU8C655cqLnlhbXQKuQMnHnL+tfjzI+0bTx+NQs5YgDNfN5l4c8F5mmqmDhF94e4/8AyW343O/C8Q5zh3eNVtdnqvxuf1Dfsof8Fff+Cfv7ZjwaR8JPj3Y2Wvz4A8J+LIzpepBsZKrHMdk5HrC8g96+liXDEFCMdciv454VBZWPJUgqe4I6EHsa+2/2Dv8Agun+2X+xhPZ+EfEniOT4j+BYWCy+GPFV48lzaxd/sl6d0kJHZH8yPttXrX5PxH4G1qVJ1slrc9v+Xc7Jv0krK/k0v8R9Vl3G0ZVFDGQsv5l+q/yP6Pdx6nH4Um7AIyOleA/sM/8ABSn9lf8A4KBeE/7W+CHjYRa5bQCTWPBmrssOqad6logSJYwePNjLJ6kHivfQcDI/z0r8GxuAxuW4qWGxVNwqR3jJWa/rvs+h93Qr0cTSVSlJSi+qFLEHqtKpyp6daQgg8k0qjOR157muU1FHSnbsHGD+dIoyB/QUMuGyaAE3nHbpSqeOMdqbggYye9C89KQDyTxwPxoGepA4HekwAAQAeKPwHSgAGRnIHSg8np3oXJ4Awa5r4z/Fbw58C/hL4m+M3i+3uptK8LaHdarqMdjGHmeGCMyOEViAWwpwCRz3FXSpzrVI04K7bSS7t7EzlGEHKWyOk3Y4wPzpdxznivzPm/4Omv2CtgktPhB8WJgy5H/EjsF6/W8qBP8Ag6h/YZ3fP8DPiwBjr/Zenf8AyZX2v/EN+OLX+oT+5f5nj/6xZMnZ1l+P+R+nAPc4696d93jjgV+Zcf8AwdQ/sGhct8F/iwPb+x9P/wDkynj/AIOo/wBgpjj/AIUv8WP/AAUad/8AJlR/xDrjb/oBqfcv8w/1hyb/AJ/L8f8AI/TAHnPH0xSk+tfmTN/wdQfsMjJtvgb8WJPTOm6cv/t5UUH/AAdQ/sVvJtm+AXxSRSfvfZdOP6C6p/8AEOOOLX+oT+5f5ifEeSr/AJfI/Ttd3cUEHPT9a/PzwD/wcv8A/BM/xZMkHifUPHfhYsQDJrHhJpY0z3LWrynH4V9Yfs9ftwfsiftYWq3H7Ov7RHhXxVK0e9tP0/U1W9jH+3aybZl/FBXi5jwzxFlMXPGYSpCK6uErffa34nbh8zy/FO1KrFvtdX+49S7Dp24zTSxwc4p7IwOCpBHUEUwgha8M7hu7nmnKxPYfhTCDnnt7V5r+2n8XPGH7Pv7HfxU+Ovw/S0fXfB3w+1XWdHXUIPNga5trZ5Y/MQEbk3KMjIyM8itsPQnia8aMN5NJerdkE5RpwcnstT1EAAd/ypsnspOK/n7tf+Doj/gpIkKNN4V+FcmVBJHhO6Hb2vKjvv8Ag6E/4KSSoTF4d+F0XHbwjO3/AKFd1+qf8QW426xp/wDgf/APlnxhk17Xl93/AAT+gN856Dpjk0qE4ycflXmv7GPxT8afHb9kP4X/ABs+I4shr/i/wFpms6uNOtTDbi4uIFlcRpubYo3YAyfqetembcDoOvrX5VXozw1edGe8W0/VOzPp4TjUgprZq/3i9+g5PpQck9vejHIxijH04x1rIoQYz26UHBxnFOAPTPakYGgBCcDJx+VCk98UjAk5pQPYUAKSTwcdu1LlvQdPSk9KX147UAGT1GPfmmMeM8HGKdjngim0AIM9AB970pFLZ7flTtpIHA696RUx1A59qAFBOAOOlOBwB+HU0m3gcdvWnDnAx3oARjxzihiTknHWgjihu+fXtQO+gAnGOOlKCeuB2700dBwPzoHHVR09KBCkkjJoB5HPSkJPoM0qnOP60AID04FO5OeKacg9cUE46igBQSQenSkO7OQB09KBuwen50hU/wB0dPSgBM8HOKcnTt1pBk5GB+dOTpQAo9v50ZOBwOnrQeDwO3PNIegz6UDGqTyOOlISTwcfWnAZyMDp3pNvPIH5UCEBPPT86EJyc4pdn0PFCrycAfWgBdxPp0FLk57dOaCDntxilx7Dp6UANDHPUcGkZjjgDr0pcc8Y69aaRkduvpQAZJ5wKF/Dp6UgU5HA6U5RjsOlAB27dKTJOenFPwSAcfnTcHByf0oAaMgnH8qcWIB57UmOefzxSOOCB6UAS4J4Pp3NIFYkfL+tKGX15pMqGHP40ABVsj5f1oAbgY7+tG5eADRkdDn8qAEYEH7valbIzhTnHrSHGeAelBwCeO1AApYpyppxyW4U9fWmZ+XOD+VKTk556+tADiCOg7etKuQRx+tJuAyAaAw3UADK2Mle1IwOOnp3pGYcEEdPShnB4J6UDEyxPIH500lum3tTgRgHmmuAQOtACgkLkjrWP8RPiJ4E+EfgTVfid8T/ABZY6F4e0Sze61bVtSmEcNtCvVmJ79goyWJAAJIFXdd8QaF4S0G98U+J9XttP03TLOS61C/vZhHDbwRqWeR2OAFVQST7V/Op/wAFkP8Agrj4w/4KD/EWX4a/Di9udL+Efh7UGOh6aGKPrsyEgajdLnnIz5UR4jU5Pzkkfa8E8F47jPM/YUvdpQs6k/5V2XeT6L5vRHi51nNDJ8Nzy1k/hXf/AICOi/4K5/8ABcr4i/tu39/8Df2dr7UvCnwnilaK5ZZDDf8AigA433O05itz1FuDyDmTP3R+fEYCgKuAB0ApJBjq34V3H7OP7N3xu/a0+K9h8FP2ffAN54i8Q6gdy29soWK2iBw09xK3yQQrnl3IHYZJAP8AXuV5RknCOVexwyVKlBXlJvfvKUnu/N/LTQ/JcTisbm+J5ptyk9l+iRxyOoYBmxjkkmvqL9jf/gkH+3X+29bQ+IPhZ8K/7H8MSuAfGXi+VrDTiuRlosqZbrA/54oy9iwr9av+Ccn/AAb5/s1fsmW9l8R/2j7bT/if8QkCSj7daFtF0iUAHbbW8g/fsD/y2mBJIyqR1+hkQjgt0t4o1RI0CxxquFVQOAAOAB6CvyDijxvhQcsPkVNSa09rNaf9uw0b8nK3+Fo+tyzgnntUxsrf3V+r/wAvvPy0/Z4/4Nbf2a/CVpban+0v8a/EnjLUfLBurHQFXSrEP3Cn95Mw99yk+g6D6r8B/wDBFj/glz8O0jTSf2NvCuoyJg+f4le41NmI7kXErL+mK+oCw4xn6CkL8Z6V+LZjxvxbms3LE42o79FJxX/gMbL8D7PDZNleFX7ujH5q7+9nl2m/sJ/sO6TbLaad+xd8JIY1UAInw603p9TCawPG3/BMf/gnX8QLdoPE37EHwwbf1ksfCNvZyfg9sI2H4GvcGZeT+VAfOMZNeLTzXM6VTnhXmpd1OSf33OyWGw0lZwVvRHwX8av+DcT/AIJq/E+KebwX4S8S+Bbp1Ply+HfEks0SHHXyrvzRjPYEV8O/tK/8Gtv7SXguO51v9l34z6D43towzxaHr8f9lX7jnCpIS8Dt2+Zox71+6hKnsfyoGwcc19blXiZxnlMly4qVRdqnvp/N+990keTiuHMoxW9JRfeOn5afgfyJfGT4D/Gr9m/xnJ8O/j18LNd8I63EMnTte09oGkX+/GT8sqf7aFlPY1yxfcMfrX9bn7RX7MvwE/ay+HE/wr/aG+GOmeKNFm+aOC/i/eWsnaWCVcSQSDsyMD26cV+IH/BUP/g32+Kf7J9lqPxv/ZVuNS8c/D2DfcajpMkXmaxoMQJJZ1Qf6XAo6yoA6j764Bev33gvxdyjP5xwmYJUK70Tv7kn5N/C32l6KTeh8Nm/CWJwSdWg+eC+9f5/I/PfwF8RfHnwn8Zaf8Rvhj4w1HQNe0e5W40zWNJu2hnt5BzlWU59iDwRkEEHFfvN/wAEf/8Agun4P/bJGn/s6ftQ3en+HviqEEOlamqrBY+KyB/yzHSC8wMmH7snJj7xr/P8GDchgRjgg8GljubuyuY7+wu5beeCRZYJ4JCjxOp3K6spBVgQCCOQRkV9dxlwTlHGOB9niFy1Yr3Ki+KPr3j3T+VnqeZlGcYrKK94O8XvHo/8mf2MhjnAx70oLHOB2r44/wCCFX7T/wC03+1v+wlpnxB/ad8K3Kahp2pSaZoni65YK3iqxiAAvWTqJFcNE8n3ZWjLDndX2SIjnYASScDHc1/FGbZfXyjMq2CrNOVOTi2ndNrs/wCmtnqfseGrwxOHjVjtJX1I5bmC1ge5upkijjQtJJIwVUUDJJJ4AA5J7V8oeKP+C6f/AASh8J+PW+HOrftkaI96l19nmvdP0fUbvToZM4+a9ht2twoPVw5UdyBzX58f8HAP/BZyP4g3ep/sG/sneL92gW8rW/xI8W6ZccanKpw2mW8inmBSMTODiRh5Y+QNv/I+MhcYPQcAdq/XuD/CP+18sWNzWc6fPrCMbJ26SlzJ79Fa9tW9bL5HNeK/quI9lhoqVt29vRW/M/sc8NeJvDXjXw9ZeLfB/iCx1bStStluNO1PTbtJ7e6hcArJHIhKupB4IJFXVzzwevc1+S//AAai/HLx34o+EPxW/Z71/U7i60Lwbqemap4cWZywsjf/AGpbmBM/dQvbJKFHAZ5D/Ea/WjABr8t4myOpw5ntfLpy5vZtWfdNKSdujs1ddz6fLsZHMMFDERVubp+DFyxHA7etBDZOB2pMjP3jxShgR+FeJY7RVBORtPWvFP8AgpLv/wCHfnxpOOnwx1rGf+vOSva1xuNeKf8ABSg4/wCCfPxr2n/mmGtf+kklejktv7Yw3/XyH/pSObG/7lU/wv8AI/lTtUU2sOAP9Sv8qV0GKLM5tYuP+WS/yp7KSeK/0Iu5H4FLSTKroT0FIowwyeK+2P2Xv+CC37b/AO2D8CNB/aL+E194Li8P+I4ZZNOGseIXgnKxyvE25BC235o2xz0xXdv/AMGw3/BSWNgv9q/DkjPUeLXH84K+Jr8d8IYbESo1MbTUotppy2admvkz2aWR5vVpqcaMrPVaH55xgbQTTgCe1fokn/BsP/wUUU8+IPhyc9ceK34/8gVLf/8ABsT/AMFFrWyNzY638PLuQLkW8Xil1Y+2WgA/Wqj4hcFPfH0//AiZcP50n/Al9x+darip7DUtR0jUYNZ0i/uLS8tXD2t5aTtFNCwOQyOhDKR6g19G/tGf8Eif+Chf7Lei3Xij4mfs1a3Po1nzc614fCalbxJzl3NszsqgDlmUAd8V81o6ugdHDA9CDkGvp8vzTKs3oOpgq0Ksdm4yUl6OzZ5tfDYvB1LVoOD800foH+wf/wAHDv7XP7MWo2Pg/wDaFvLn4qeCUKxzR6rOBrNlF03QXbf64gfwTbs4xvXrX7qfs0ftNfBT9r74PaZ8c/gD40h1vw/qYKrKg2TWsy48y2uIj80MyEgMjeoIyCCf5JmOSRX1P/wSU/4KR+MP+Cdn7S1l4hvdUuZvh54luYbP4g6EpLI1vnat9GpOBcQZLAjl03xnhhj8e8RPCzLs1wk8dlNNU8RG7cY6Rqd1ZaKXZrd6Pe6+w4d4oxGHqxoYuXNB6Xe8fn2P6bdhJ4H1FeG/8FQMj/gmz8fPQ/CDXx+dlJXuGn6lpWtaVa67omow3djfW0dxZXdu4aOeF1DJIrDhlZSCCOxrw7/gqFx/wTa+PnP/ADSLXv8A0jkr+ZMnT/tfDp/8/If+lI/S8Y19UqP+6/yP5V4QRGoPZR3ps4+U8dRT0GFGB2FJIMofpX+getz8CW5/V1/wTvtmtP8Agn98Dbcj7vwj8P8AB/68Iq9kGSOR39a8l/YKZR+wl8EsdP8AhUvh7H/gvhr1jcu3A9a/z5zR3zSu/wC/P/0pn75hf92h6L8h2CTkD9aQ5HRe3rQT82efzoyOpJ6etcBuCgg8Dt60MCAPl7Uq4yD/AFprtggBjzTQDWBx07DqaXB67T7UhYcc9T2FKCD1JoYB82QcdhS8gHj+HuaaxBUc+lKCMHntSAAG4O39aQg4Bx2pQQf4aQ4wBj9aAFTgAY7/AN6gE7vu9vWkUgAAZ6+tIHGepoAfg7QAvYZ5pcHqF/Oo9/AG7tUiEED5qAEOcHg9KGDHPHf1pWxtPPamyEc896AAbsDj9aQbiOmePWgOMDrQpXjk9KAEYt1x1PrToy2Bx355pDggDNCYG3B702ApDcnFGGPRf1pMg8j19KDg9u1IATIGCp6etBycYX680LtAz/s0ErkDnpQApVucL39aUbtudtJ8vOM9R2oBAB/OgBSWJ6H86QbmUHHb1pu4Z4pVIIHJoGKoYg4XtS4JP3e/HNIpGD9KXKg4PrQAbW9O3rSBTz8vb1pdwPIz07UgI557etAhzZJzjt60KGPQdqaW54z93ilDDt6UAAU5Py9+xpHU44Xv60uRnknrQSoBPPJpoAAOR8p6etAUnnaelA2549PQ0Aj9KGAYJ7d/WkIODx2Hel3Ck3DkY7ikA35t3CmlYNg/KenpRxuJH8qQlSPwoAeH4xSbskHjNJg56UmMNx6UDsO3dPlHSng4/hPT1pgXHY9KUjjgfpQIUt/sjp60hPOcdv71NY9R7UinJ/KgBwYbc7e/rQWG7oOvc0mOMCkIwce9Ax5bHp+dIGO7OP1pMZIOKAATjFACMwJHyjntmlB/2e3rSFenHagKCMYzxQFxQwIwB29aCFzkgUYxxXiH/BRb9sjw9+wd+yB4u/aN1dY5tQ0+1Fn4X06Q/wDH9q9xlLWHH90PmR/SOJz2rowmFr47FQw1CN5zajFd23ZGdWtCjSlUnokrv5H5o/8AByL/AMFM9S1DxE3/AATy+C/iEw6dYrFc/FG+tZebq4OHh0vI/gRSsso/iZkQ/cYH8hWkJ+Ynr7Vd8VeLfEfjvxRqXjjxlrU2o6vrN/NfatqFw26S5uZXMkkjH1ZmJ/Gm+G/DmveMvEen+D/CekT6hqurX0Vnpun2qbpbm4lcJHEg7szMoH1r+5+E+HcFwjkVPBU7e6rzltzS+1J+Xbskl0PxHNcwr5tjnVl10S7Loj0r9jH9jL40ft4/HbTPgJ8EdJV7y6/f6tq9yjfZNHslYCS7uGHRFyAF+87EIvJ4/pZ/YR/YJ/Z//wCCe/wYt/hL8EPDy/aZkSTxJ4ou4l+367dhcGedx0XOdkQ+SNTgDOSeN/4JSf8ABOPwn/wTl/ZntPAjw2134619Yr/4ga7CufPvNvy2sbdfs8AYog/iO+QgF8D6dIPHFfzF4k8f1+K8e8JhZNYSm9F/z8a+3Ly/lXRavV6fpfDmRU8sw6q1FerLfy8l+oAqSTsHtmlLcDjtSAAGgKSuc84r8uufTDsnGMdvWmnheg6CnEU18Y/CkAxyOTtHWlU5wduf/wBVDAc4HegKAM0AA7HbzgdaGPzY29PelCjtjpSMoz+dACZBA4oyQxYZH0oUYxjsKCoHbvRcND8d/wDgul/wRH0mLRta/bc/Y28HJbSWqyX/AMQ/A2mQYjeMZaXUrKNfuleWmhUYK5kQAhlb4x/4I4/8Et/EX/BR345C/wDFsFzZ/C7wpcRTeMtYhJQ3jH5o9Ot3/wCesoGWYf6uPLfeZAf6U349D7EZBrnPgh8EPhP+zx4GT4a/BHwBpnhrQkvrm8XS9Lh2RCeeVpZZDkkks7HqeAFUYVQB+r5b4u5/l3C1TKn71XSNOq3rCL3v3a2g+l9fhR8vX4VwOIzJYnaO7j0b/RPqdH4U8K+GvAvhbT/Bfg/QrTS9H0mxis9M02xhEcNpbxoESJFHCqqgAfSvy5/4OB/+CwsHwM8Laj+xF+zF4sH/AAnGr2xh8da/p0+G8P2ci82cbr0u5VI3Y5ijPZnXb6r/AMFuv+CwWj/sF+An+BXwQ1i3u/i94ksN1uU2unhizkBAvZl6ee3PkxHuPMYbVAf+dzX9a1bxJq11r+vapcXt9fXL3F7e3czSS3EzsWeR3Yks7MSSx5JJJr3PCzw7lm1SOd5pH90neEX/AMvH/M7/AGU9v5n5LXh4m4gWGTweGfvbSfZdl5/l+WVsP/6q2vhz8PvGvxb8e6R8MPhr4autZ1/XtQisdI0uyj3S3M8jbVRR9eSTgAAkkAE1lRW9zd3SWdhbSzTTSLHDBBGXeR2OAqqOWJJAAHJzX9CP/BDD/gjlZfsR+DLf9pf9oLRYZ/i14i04G2sZkDDwpZSjm3Q/8/Ui4Erj7o/drxvLfsvGvF2C4Ryx16nvVJXVOH8z8+0V1fy3aPk8pyutm2I5I6RW77f8Hse9f8EnP+Cdfh7/AIJvfsuW3wwmu7fUPGOvTJqfj3WbflLi+2BVt4ieTBAvyITjcTI+AZCB9O5Xrjv60E8DjtTQSeK/jPMMfis0x1TF4mXNUm7t+b/RbJdFofruHw9LC0I0qatGOiFDc52npS7vVe1NABPTt1pSB6Vyamw5WGfuj24rxb/gpD83/BP/AONaKoOfhdrfH/bnJXs46nPpXjf/AAUUTd+wT8Z1H/RMNb/9IpK9DJ/+Rth/8cP/AEpHNjLfVKn+F/kfynWZ22cOf+eS/wAqkDc5xTLdCLWL2iX+Qpa/0Ii7H4BJ3kz+l3/ggteif/gk/wDCeIN/qrK+Q8dD9vnP9a+u5W56Cvi3/g36u2m/4JV/DpCf9VNqSDPtey19nlyxxzX8DcXQ5OK8cv8Ap9U/9LZ+85VLmyyi/wC7H8kNBGc8Uolx0PfFMzxg0ZO78a+fO4nSYqcoccdjivyk/wCC/v8AwSH+HviX4Uax+3P+zN4EttJ8UeHwb3x/omj2wjh1qxz+9vliXCrcw/fcqB5se8tllUn9V0B5zUWraFp3irRb3wprVss1nqlnLaXcTjIeOVCjAj0IY17vDnEGYcM5rTx+Ek04vVX0lHrFrqmvueq1Rw5jl9DMsLKjVV77eT6NH8d7MSNw7+9MIJUr2NdX8cvAMnwq+Mniv4ZyAA+H/Ed7p2Ac8QTvH/7LXK4OOK/vKjWjiaMasdpJNej1PwmcXTqNPof0U/8ABud+0zqPx7/4J16d4I8Sao91q3wz1yfw3M0rEv8AYwq3FkTnqBDL5YP/AExx2r3n/gqOwT/gmr8fSi8/8Kh1/Az/ANOUlfnH/wAGnHiK5h1T45eDCB5Elr4f1BeORIsl9CT+Ksv5V+jn/BUIf8a3fj0CeP8AhUevf+kclfxxxZltPK/EupQh8LrU5L/t/ln+DlY/Y8rxMsTw5GpLfka+66/Q/ldiQ+Wu7+6P5USKAhI9KWM/KAfQUrjKke1f2ly6n41d8x/V5+wPIG/YP+CJ9fhH4e/9N8Nesgkj7vf1ryH/AIJ/5H7B3wRz/wBEk8P/APpBFXrpwBxX+emaaZnX/wAc/wD0pn9AYb/d4ei/Idu5+7nn1p+QOSuKYM54FeGftyf8FGf2X/8Agn14Ih8VfHzxe/8AaWoxudA8J6RGs+p6sV4PlRZASMHhppCsanjJOBWWCwWLzHFRw2FpudSWijFXb+X9WHWr0cPTdSpJJLqz3gMM9Mcc5qObhd23j121+BX7V3/BzH+2n8Wr+40f9m/QdJ+FuhsSsM8MSalqzr2L3Ey+VGfaOLj+8etfE/xP/bZ/bJ+M92978Uv2q/iJrRc5aK68X3axD6Ro6oB9AK/Ycq8DuJsXTVTGVYUb9NZyXrb3fukz5LFcaZbSly0oufnsvx1/A/rFN7ZqQhuo8+hkGanjdSvycjgZHNfx5t408dPL57ePvEG/Od/9vXOfz312/wAMv2yf2vPg1qEeofDL9qX4g6Q0bArFB4runi49Y3dlP4ivWreAuMVNuljot+cGvx5n+Ryw44oN+9Sf3n9bLDnAWkI9u3rX4L/spf8ABz1+1d8Krmz0X9p7wfpnxK0SPCXN5bwpp2rov99JUHkykf3XjG7H3161+wf7Fn7f/wCy9/wUA+HcnxC/Zx8eC+azKprfh/UIxb6npEjDhLm3ySoODtkUtG+DtY4Nfl3EvAnEnCr5sZSvT6Tj70fm9Gv+3kr9Ln02W53l+aL9zK0uz0f/AAfkex5wOn60jcrnA7d6cR70jDIr489YjzgjKjr1zSh0DE5FfLH/AAWZ/at+Mf7Ff7CesfHn4Da7a6f4ks/EelWdtc3mnx3Uflzz7JAY5AVOVyM9R2r8fX/4OO/+CpcnT4reG1908E2Q/wDZa++4Y8Oc+4sy94zBSpqCk4+9KSd0k+kZd+54WZ8Q4HKq6pVlJtq+iX+aP6KgBgHA49qkUgcba/Db/gnN/wAF0/8AgoX+0J+3P8L/AIHfF74maPe+GvFHimOw1e2t/C1nbvJE0Uh2iREDL8wXkHPFfsd+0D+0V8Gv2VPhNqfxu+P/AI8s/DvhvSYx9pvbs5eWU52QQxj5ppnIwsaAsT7AkeZxHwZnPDOZU8DiUp1KiTioXle7aS1ine62SOjLs4weZYaVendRi7O9l+rO8ONvCZ4qO6YRjdJwPVuK/Bv9tf8A4OYf2mvi9ql54W/ZA0hPhr4XDtHb6xcRR3Ot3iZ4kZmDRWueuxAzDP8ArDXwN8Q/2rP2ovixqUmrfEn9o7x5rU8rEub/AMXXjLz6KJAo/AV+hZL4HcRY+gquNrQoX+zZzkvVKyX/AIE/M8PGcZ5fQqOFKLnbrsv8/wAD+thZ4WHyyK3A6HNSK/X5eB1r+QvQvjX8bfDFyl74Z+NXjTTZo/uS2Pi29jZfpiWvqL9ln/guz/wUQ/Zm1e3GpfGG4+IGhpIDdaB46/0zzU7qlzxNETnqGIz1B6V2Zj4D51h6LnhMVCo10acL+j95ffZeZlQ42wU5pVYOK77n9KW4cDb+tAOMcd/X6V80/wDBOn/gqL+zp/wUd8Cyap8M7yTRvFelwK3iXwPqk6m8sMnHmxkYFxbk8CVRwcBgjcV9LrhsV+J5jl+NyrGTwuLpuFSLs09/+G7NaNao+woYijiqSqUpXi+oi/7v15pQRjO39aAMDHNBB7fyriNgB4PA+760m4ZHA/OgL1wf4acFJHT6UAAIwRt7+tAbK4wOnrSAHJFAzjjNACHBP3R19aFYYGR29aUjrx3pMYA47UAKpGOR27mlDZboOvrTVHGBTgOef5UAGQBjb29aQMBk/wBaUqMcelNxySM9aAAnvgfd9aUMoOMfrS7R2z09aQgA8d6AFVstyO/rSEj07/3qTHzflQoyOlAChgMcdvWlVuc7R931puDkYFIP6UAPLDPQceppN2c49u9GT1oHc0AGQD+HrSMx6be9KOvXtTZBxwe/NADjj9aTOTjH50BwRzjr3NKD83bNAxwYYFBI9qCScY/lQCeBj9KBDHUHnjp60cZPBpzkk9OlJkc9OlAw4IzQcE89sd6QPtHBFO35bj1oCwYH6UL1/CnEn07UgOCPpQIRgvB+lIQvb09KcxzjjtQT1wR/k0AMOMZGPyr8M/8Ag6H/AGurzx18ePC37HHh3VN2keB7FdY1+CNxtk1a6TEQYDvFbHAz0Ny9fuVc3lrY28l7fSKkMMbSTSMeFRRlj9AAa/ks/a2+OF5+01+1D8QP2gL2+e4XxZ4uvr6ykkGD9kMpW2XHbEKxjHoK/ZvBLJIZhxLUx1RXjh43X+OeifyXN87Hx/GeNlh8ujRi9ZvX0Wv52PNySOf0r9R/+DY/9hi1+LPx41v9tj4haN5+kfDthYeEI548xy61NGS9wAept4G49HnVuqCvy4uSkMDzvnailmwOw5Nf1O/8EqP2XT+x9+wJ8N/g3qOmLba1/YSat4oULhjqd7/pE6t6lC6w59IhX6t4w8QTybhf6tSlapiXyefIleb/ACj/ANvHy/COAWMzH2sl7sNfn0/z+R9DnG0cd+4phIOP8aV5OOOOab5mcDPev5DP1gTv0NGeAKAwDZBAppc7RQMeTgZxSFieoIpFcnH06UZOOcUDSFbByffrigEDt+lDtnOPX1oDY/KgQZB6Dr7UEZ5Cn8RQG54A6+lKWJxgdqBCAcg4/Sgrkn6+lKG6c9x3pV5J6YoAjaPjAFfIv/BW/wD4Kq+B/wDgmx8GUTREtNY+Jnie3kTwb4elbKQgfK1/cgciCNui8GRxtHG4r6/+3R+2x8HP2Bv2ftU+O3xf1BX8lGg0DQopgtzrV+VJjtYQfU8s/REBY9AD/L5+1N+0z8Vv2v8A4567+0H8Z9dN7ruvXJdkQkQ2UA4itYFJ+SKNcKo+pOWYk/q3hl4fy4qxv1zGxawtN6/9PJfyryX2n8urt8txJnyy2l7Gi/3kv/JV39exzHxK+JXj34x/EHWPit8UPFN3rfiLX7+S91jVb2TdLczucsx9BwAFGAoAUAAAVgscgluAB36AUkrleQenav1M/wCCAX/BHi6+P2v6f+3D+0/4PRvAGl3Hm+CPDuow5HiK8RsC7kQ9bSJ1O0HiaRe6Id39M8QZ9lXCeUSxWI92EFaMVo5O2kYr+klrsj86wGBxOaYtU4at7vt5s9j/AOCAP/BGu28D6fpX7en7VvhDdr12i3Xw18K6nb/8g2EjKapPG3SdwcwowzGv7w/My7P1yDNuLNySeaXZj5m+uaT8Ohr+LuI+Icw4nzWeOxb1ey6Rj0ivJd+r1Z+w5fgKGXYZUaS9X3fdjlYFRk80LgjihWwB0/OhG4J4/KvCR2i8f3e3XFBIxwvb0oLDdkEUbhwcdqroAgwc143/AMFFJlg/YK+M0pAwPhjrfX/rzkFeyq27g+teKf8ABSIkfsAfGgAjP/CsdZ7/APTo9d+T65th/wDHD/0pHPjP9zqf4X+R/KrbsGtoweP3S/yFKOTTYP8Aj3jH/TNf5U8HnJr/AEEjrI/AJfEz+j7/AIN8oXg/4JWeAC38d5qjDPp9tl/wr7Tycj6elfGn/BADY3/BKj4buo6vqec/9f8APX2YVUHgDpX8G8Yu/FmP/wCv1T/0tn7rk/8AyKqH+GP5DBjP/wBalBGc0L39vSgZz2r5xanpEsRGMUXGqWOh2c+uanOsVtZwPPcSscBI0UszEnsADQgK5BAGfUV+bn/BfX/gqx4H/Z2+Cut/sf8Awb8Ww3nxL8YWBstZFhOH/wCEc0yVSJnmZfuXEqZSOL7wDmQ4AXd7GQZHjuIs1pYDCRblNpPtFdZPsktX/mcePxtHAYWVao7W/F9j8Pf2h/HkfxT+PPjP4lREbPEHirUNRjweNs1zJIv6MK46mMwIAHQDAFKuWO3I+tf37Rw9PDUY0obRSS9Fofg826k3J9T9kf8Ag038JzPF8c/HUkZEQOgabExHDPm9mcfgDH/30K/RL/gqNFv/AOCa/wAfFUcn4Ra8f/JOQ141/wAG/H7Muofs1f8ABObQNR8QWTQ6v8RdTn8W38cq4ZIZ1SK0Q/8AbtDE/wBZDXtv/BS1Bcf8E6PjxARjPwi8QcH/AK8ZTX8YcWZnSzXxKq4mm7x9tCK81Bxhf58tz9kyzDSwnDkact+Rv77v9T+VOINgfQVIy4HPpTkXag6dBQ/CnntX9sKyep+LuV2f1afsA4P7B3wSb/qkvh7/ANIIa9cJXHJ6147/AME8ZWk/YE+CDtnJ+Eugdf8Arxjr2RMFtuK/zvzbTNcQv78//Smf0DhtcND0X5Hj/wC3j+2L4I/YR/Zf8Q/tF+N7Zb19PjW20HRxLsbVNSlytvbg9lLZZ2/hRHbnGK/mI/aH/aI+LX7U3xg1r45/G3xZLrHiLXLnzLq4bIjhjH3LeFM4ihjHyog4A9SST+kX/B1J+0Pf6r8cfhz+y3p2qA6f4e8NP4k1a2jkGGvLuV4IN4z1SG3kIz2nJr8nGus8Z/Gv6r8GuFcLlPDcM2nFOtiLtPrGCdkl625n3uux+W8XZlWxWYPDRfuQ6d31b/IsOFYZGOe1ej/Ar9jD9rP9qCCa7/Z3/Zx8YeMbWGUxzajo2jO1nHIOqG5fbDuHdd+R6V7T/wAEWv2EdI/b+/bHs/Avj61kl8F+GLBtb8YpG5Q3FsjqkdqGBBXzpXVSQQQgkIIIFf0q+F/DHhfwR4csfBvgrw5YaRo+l2y2+maTplokFtaQqMLHHGgCooA6ACq8Q/FNcJYpYDB0VUrtKUnJvlgnsmlq297XVlZ63Fw9wx/atJ160nGF7K27/wCAfzI33/BFL/gqvptqby4/Yi8WOoXO23urGV/++UuS2fbFeC/Fn4HfGn4DeI/+ER+N/wAIvEvhDU2JEdn4k0aazaXHUxmRQJB7qSK/r2bawwFFcf8AHn9n34O/tQfC/Ufg58d/AVj4j8PanCUmsr6IFoWwQJYX+9BKucrIhDKeQa/Pss8eM2hiY/2hhYSp9eTmjJLuuZyT9NL90fQ4jgjCOk/Y1GpdL2a/Q/kQ2GvQf2Yf2nfjH+x18a9H+P8A8CvE76brujyjfGWP2fULYnMlpcoDiSCQDDKehwwwyg10f7eX7JfiT9iX9rHxj+zf4guJLqHQdQ3aLqcqBTqGnSgSW05A4DGMhWA4Dq4HSvHWXKnNf0Q5ZdnuVxnG1SjWinqtHGS6r0Z+e3r4DFNbTg/uaP6zv2Mf2qfAP7a37NXhT9pP4cgxWPiKw33WnySBpNOvI2MdxaSEfxRyqy5/iXa3RhXpzkYIJr8b/wDg1Q+PurRyfFD9mDUtULWSra+JtItXbPlSk/ZrkqM8BlW3J46rmv2Od8j5vwr+I+MMhXDfEeIwEfhi7x/wyScfmk7PzR+0ZPjv7Sy6nXe7WvqtGfAX/Byxc+V/wS81GDI/f+PNFT64kkb+lfzwrnOa/oU/4OYdz/8ABNIhTwPiFpBb8pq/ntAr+jfBSP8AxiEv+vsvyifnnGsv+FZL+6v1Pbf+CcPxD8M/CT9u74S/FHxnrNvp2kaD46srvVdQu5NkVvbKW8yRj2AUk/hXZf8ABWD/AIKN+Ov+Ci37SV54pkv7u0+H/h64ktfAHht5CEht84N5KmcG5mwGY9VXbGOFOfmi3fy/lxx719hf8ElP+CT3jH/go/8AE+613xHf3Wh/DTwzcxr4o1+FAJrqYgMthabgQZWXDO5yIkYEgllB+6zmjkGT4t8R5g0nRp8ib1sm2/dXWcr8qtr0W7PFy+rj8VD6hh/tO/8Aw/kj45s7O71G8XTtNtJbi4fPl28EZeRvoq5JrT1T4d/EPQ7I6lrngDXrK2UZa4u9GuI4wPUsyAY/Gv6tP2ff2Q/2Y/2V/Ddr4Y+APwR8O+G4rSERi8s9Nja9mHGWluXBllY45LMf6V6XJNNNE0FxK0iMuGjkO5WHoQeDX5HiPpAQp12sPgLw7yqWb+Sg0vvZ9dS4FvC9WvZ+SuvzR/HPE8cqeZFKrqf4lYEfpTjj0zX9IP8AwUu/4Is/syftsfDXVdd+Gvw90PwT8Ure3efQ/E2iWCWkWoTgZFvfxxBUmjkPy+aR5kZIYEgMrfzl+JfDfiHwZ4l1HwZ4u0ibTtW0i/mstT0+5XEltcROUkjYeoZSPwr9V4N48yrjbBzqYeLp1IW54SabV9mmt4vXWyd1qlpf5jOsjxGTVUpPmjLZ/o/M6H4C/Hz4qfsxfF7Q/jp8FvFE2j+I/D12J7G6jJ2SL0eGVekkMi5R0PDKT3wa/qa/Yo/am8Gftpfsv+EP2lPBkK28PiPTQ1/pwkDHT76M+Xc2xPU7JFYAnqu096/k0YnOa/Z7/g1I+PWqan4S+K/7M+q3zPBot9Y+ItHidydkd0HgnCjsPMhiJ9396/P/ABp4eoY7I1mkI2q0Wk33hJ2s/RtNdte57fBuYVKWMeFb92SuvJr/AIB+v/HXH6UEZ/8A1Um7t705WBOeK/ldH6cC8dj0peO360isMHjt60M4zkDHHrTACBzn+VIDxk9vajeSSaF6dB1oGB56A/lR1Axnp6U7OegFIH6D2oAap44BpQR6Hr3oBGOR+lBbk0CFJB4x29KaDnJ/pRuPbHA9aEJOT/WgB2QWH0owDxj9KXPPHp60qtz0/WgBgGGzjvzSkD079qXd82D69jSOeOB37mgBCRkcdqaAKXeMg4/U0Kec8UAHak3AAg0udw6Dp3pCeTzQNBu+f/61I5+X7vSjdlqHOQcY4HFACbWHIz19KUKd2TmnBOMAjOaQLg9RQIUJ0+lGwkjjPFAHQjFLjI5I/GgBrKewP5UhU7unb0pxUcdKChzyB0FADAjBeh4pdpLYwetG0BeAKXB3duvegBShOMA96NpyODS4HoOlIQMjAH50ABHTg9PSgg8gA9fSmntgUvXqM0AeO/8ABQr4k3vwZ/YQ+MXxN06fyrzSPhvq76fIeMXD2zxRH/v5ItfylC3S1ijtovuxIEX6AAV/TR/wXF1J9K/4JTfGGaJipn0rT7f8JNUtFP6Zr+ZqU5PJ71/UPgPhoRyPF1+sqij8oxTX/pTPzHjqq3jqVLoo3+9tfoeofsNfCGH4+/tl/C34N3lv5trr3jvTYb+PGd1qk6yzj6eXG4/Gv6xjIZHaRjyzE8+9fzWf8ECPCyeKP+Cqvw5aWHeulw6nqHH8JjspAD+bV/SiMgAZr47x1xkqvEWHw19IUr/OUnf8Io9ngiko5dOfeX5JA+MABhz6UmP8/jS/Njv+lNweBnv61+Hn2oANnv1pNpOMUAZbt+VKFOAcL+dACqpOCc9PSgoQOnb0pwUEAcdKTA29R09aB3YFeDxQFPZTSsCcnge5oA4GADx6UCECMTwD19KXYf8AIoHBzxS7MnFACLHkVw/7SX7SPwj/AGRvgvrnx9+N/idNL0DQ4N8z9ZbmY5EdvCnWSaRvlVB1PPABI6bxx438G/DHwZqnxF+IXiSz0fQ9FsZLzVtUv5hHDbQIMs7E9gO3UnAGSa/m6/4LCf8ABT7xJ/wUa+OuzwzdXlh8MPClxJF4L0SbKG5bo+o3CZ/10g4UH/Vx4UclyfuuA+CcXxlmns9Y0IWdSfZfyr+8+nZavs/DzzOqOUYa+83sv1fkebf8FFf+CgPxa/4KJ/H+8+L/AMQ5JLHSLTfbeEPC0dwXg0ax3ZVB2aZsBpJP4m4HyqoHz83ykk/lUzhurHOTwK+lP+CXv/BNL4n/APBSj49R+BdDe40rwbojR3HjnxaIsrp9uSSsERPyvcy4IROw3O3yqc/2BUqZPwpkd5WpYejH7kui7t/e2+7PyeKxmbY6y96c3/XyR6P/AMEWf+CReu/8FE/iofib8VLO4sfg/wCE79Brt0Mo+vXS4YabbsMcYIM0g+4h2j5nBX+jvRdH0DwxoNn4a8MaRa6dpmm2kdrYWFlCscNtBGoVI0RRhVVQAAOABWB8Evgj8L/2dfhNofwQ+DHhW30Xwz4dsVtdL0+3H3VHJd2PLyOxLu5+ZmYk8mvif/gtt/wV90X9hnwDN8A/gVrtvcfF7xFYZikjKyL4Ws5AR9tmByPtDD/URN3/AHrDaoD/AMmZ5nOeeKHE8KGFg+W9qcOkY9ZS8+sn00S6X/U8HhMFw1lrnUevV932X6H0dbft+fAjWP25o/2A/DOqHU/GEPhi81jXZrSVTBpLQ+SUs5D/ABXDpKZCg/1aoN3LgV7b5WRxX8zv/BGD4vav4P8A+Crnwt8W6zrNzcT+IvEN1p2q3l1cNJLdPewTIzyOxy7NIysWYkluTzX9McfKD+Wa5vEHg+lwZmdDCU5ualSjJyfWV5KVl0Wisu3fc0yDNJ5thp1ZK1pNJeVlYYUOOM9e1NCEA5z3qUKCBkDPvTCp28Y6etfAnuiAMT3/AC+lJsYc8/dpSPm5A/E0bT046UACpyfr3rxT/gpVlf8Agn38aijDj4Zazj/wEevbVX2H5V4l/wAFKkdv+CfHxrVTz/wq/Wuf+3SSvSyfTN8P/jh/6Ujmxv8AudT/AAv8j+Va3YG2iPrEv8qVnH4VBYsWsYCT/wAsV/kKkfp1xX+giZ+CNWmz9Jv2Af8Ag4Zg/YR/ZP8ADH7MEX7HbeKz4fa6Z9cb4g/YROZrmSbiH7DLswHC/fOcZ4zivaE/4O1bWQ/vP+Ce8i/7vxXB/wDcZX42OcHB7U3cB3r89xvhhwTj8XUxVfC3nOTlJ+0qK7k7t2U0lq9kfQUOJM4oUY04VLJKy0Wy+R+zsX/B2fohHz/8E/rkZ9Pikp/9xtVNZ/4OyneykHhv9gVIrojEUuo/FAvEp9WSPTwWHsGHTrX42iT0P6UplOOtcy8J+Av+gT/ypV/+TNf9ac9/5+f+Sr/I++P2mv8Ag4w/4KGfHrSLrwx4N8QaJ8ONMu4jHJH4KsmS7Kkcj7XO8kqnrynl18HXmo6jq2oXGr6vqE93eXczTXV1czNJLNIxyzu7ElmJOSSSSetVTMO5/WlWQEbjwB1JPAr7HJciybh+m6eX0I0097LV+r3fzbPJxWNxuOlzV5uXr+iLAOThTX1D/wAEk/8Agnh4q/4KI/tR2Pgu5sLiLwJ4dljv/iDrKAqsNmGytojdPOuCDGoHIXe/RK6P/gnN/wAEWP2rf2+NSsvFV1oF14H+HEkga68ca9ZtH9qizyNPgfDXbkEgSDEIIOXJG2v6D/2Uv2SfgT+xT8GNO+BP7PvhBNL0az/e3c8h33WpXRAD3VzLjMszYGScAABVCqAB+deJHidgMkwk8BltRTxUk03F3VPu29ubsuj1dtn9Hw9wzXxlWNfExtTWtnvL/gHoFjp+naTYW+j6PYRWlnZ26QWdrAgVIYkUKiKOyhQAB2Arxr/gpMwj/wCCdvx2ZgOPhD4g/wDSCWvazkNwfyrw/wD4KZs6f8E4vj247fB/xCf/ACQlr+WcpvLNsO3/AM/If+lI/S8ZH/Y6i/uv8j+WDOAMDsKjnbEbfSkViyg+wpJSdjE+lf6EqTbP5+5bSP6uv2B7eOH9hH4JRR8AfCXw9j/wXwmvWgpA615J+wLIz/sH/BGQ9T8I/Dp/8p0Net547/ga/wA8s0v/AGnX/wAcv/Smf0Hh1/s8PRfkfG/7YH/BDX9i79tb496x+0h8Yrzxp/wkWuQWkN5/ZniUw26Jb26QRrHHsIQbUBOOrMx715Pdf8Gwf/BOyZSItd+I8R7GPxUnH/fUJr9IHYnjjp1Jrxf9qT/goP8AsYfsZQeV+0d+0DoehaiyBofD0EpvNUlBGQRZwB5gpBHzsqr719FlPFfHElTwOX4ms7JKMIOTsl0UVfRHBicvyiPNVrwiurb/AMzj/wDgnZ/wSv8A2bv+CbN94s1P4F6p4mvrnxhBZw6lN4l1KO5MUds0rIkWyNNoJlYtnOdq9MV9ODjv+Nflf8X/APg6l/ZX8L3z2HwV/Z58Z+KVTIF9q93b6ZE5z1CAyvj6gH2rwrxj/wAHXfx+v5WXwB+yZ4S0+P8AgbVtdublvxCLGK+greHniNn+JeLxVBuc7XlUnBPRJK6cr6JW1R58c/yDA0/Z05qy6JNo/cfaNoyKcucYGa/Ay4/4Omf28pj/AKF8HfhbAOwfTb6Q/wDpUKrv/wAHR3/BRF/9R8OvhMv+94cvm/8Ab4V0x8FuOZL4Kf8A4MRhLjHJoveX3f8ABLv/AAdEaFZWH7dPhjV4YVWa/wDh5b/aGAwW2XMyrn14zX5p+UD17dK90/bt/b5+NP8AwUQ+KWmfF344eHvDen6ppmiJpkMfhm0nhheJZHkDMs00p3ZcjggYA4zXiRjOORX9M8HZTjcm4YwuBxaSqU42dndbvZ+h+YZvi6WLzOrWpfDJ3R+h3/BsPfy2n/BRLWtPTO27+Gl+H/4DcWzCv39cEjtX8/8A/wAGyqhf+CkF7wOfhrqeP+/1rX9ATAjoa/mfxpio8bSt/wA+4fqfp/BknLJV/if6HwB/wcpQhv8AgmLeynqnjzRiP++3FfzzhSDkdK/oW/4OXZWh/wCCYlzHjBk8f6KDz/tSH+lfz0Fj61+veCWnB0m/+fsvyifJca/8jZf4V+pb0nTNR1zVbTQNFtvOvb+6itrOH/npNI4RF/FmFf1a/sRfsr+Dv2Mv2VPBn7Ofgy2jCaDpCHVbxUAa/wBRk/eXV05HVnmZiM9FCL0UAfy4/s4n/jIv4ehl3A+OtIBH/b5FX9c2CY1Lc55618r4847EJYLBp/u3zza7yVkvuTdvVnrcCUafLWqte9ovkQshxx+lKw5yPWhwCCSO1BXNfzkfoQeW0iEIOqnkV/Nh/wAF9fhfp/wv/wCCpnxGXSrdYYdfNjrRRAAPMubWNpG+pcMT7k1/SlDkEc1/PH/wctQ+X/wU/wBVkA+94M0U5+kBH9K/YPBKpOPGEoJ6SpSv8pQZ8jxnCLylN9JL8mfAO3PQ1+mn/Bq/NND+3X4/tEk+S4+EUrOpPUpqliQf/Hz+dfmYhzxX6Xf8GsYdv2/fGoHK/wDCnr3P/gz06v3rxJgv9Rsd/gX/AKVE+E4clJZ1Rt3P3rCknnrmlwccUpXvgUbe2RX8Sn7SNVWweD0o2ndxnpzxShSc428j+lLt6Hjp6UAG08/WlCHb07UgBGRgdRSgnb26etACFSDxnr2pApKjOTx6U4gHnI60gAKjkdKAECnGBnilKHJ4PWkA4Odv40uBkkbaAG7Djp2oReuFNO2gjovT1pFAGTx+dAD9pzkA0BD0x+OKQfe/h/OlyD6cD1oAQqc4x0NIyEjg9+1Kc5zx+dByRzjOaAG7Dx1pFU5+lO2gkYx+FIBk9unrQAgRh0B6elBiY5JB/KnBcDkDp60oXCnp2oC5Hs5Iz+tK6kjj05pSPmPT86RzkD60Bck3DgY7+lNLAtx/Kj5en9KTA3dfzoAcGBxinqwOM1H8mBgjpTgVx16UAKzDgUEqSeO1Nbb6ikJUEnP6UAKGCr0oLLv4A6+lNOCCM/pSHaWzzQBLvXnjHApGcZB9KYGTsT0pwwTkg9KBiNg4OO3pSlgRjHamkKD07Uu0YGKBHx//AMF7ZRF/wSg+KQB++dHX6/8AE0tT/Sv5p5HO7OR1r+lH/gvyWH/BKH4mFP8An50YH6f2nb1/NZL3+tf1X4Fr/jGK7/6fP/0iB+V8ba5tD/AvzkfdX/BuM6j/AIKleHCe/hXWAPxgFf0Yqe2K/nG/4NzJEi/4Kl+Ft/8AF4c1ZR9fs9f0cKR2z9K/NPG1W4xi/wDp1D85H1HBf/Ioa/vP8kKduPuj64pMrj6GieW0tbeS7u50iihjZ5pZHCqiqCWYk8AADJJ9K/G39sz/AIOifFHh74mah4I/Ym+DPhvU9C0y7aBPGHjT7RN/au04MsFrBJF5UJ52M7szLhiqZAr4Lh3hXO+K8TKhltLncVeTbSir7Xb79Fq99ND3MwzTBZXBSxErX2W7fyP2RDDcevX1pSV2gV+cf/BK3/gv94D/AG1/HVt+z9+0d4L0vwF4/wBQbb4eutNvXbSdefGfIj84l7a5wPliZnEmMK+7CH9GyBjiuPPMhzXhzHvB5hTcJrXo00+qa0a9PR6m2Cx2Gx9FVaErodkEdO1JvHYfpQAOuT+VNO3HJP5V5B1j2YHPHf0oV1OARTSFGeaF28HP6UCHrtI49KZqN9p+k2M+rarew21rawvNdXNxIEjijUFmdmY4VQASSeABT41DNtXkngADrX4j/wDBej/gsmvxY1bVf2GP2XPEwbwnZXBtviD4q0+cFdanRvm0+3dTzbIwxK4/1rKVHyKS/wBNwnwtmHF2bxwWFVlvOXSEerfn2XV/NrzM2zShlWEdapv0Xd/1ueYf8Ftf+Cv19+294vm/Z7+A2szW/wAJtAv8vdxEofFN3GcC5ccEWyn/AFUZ+9/rGGdoX89JFUccHHtxU7FW6EEe1db8BPgB8U/2ofjDovwK+CvheTV/EWvXPlWdsnCRoOXnlfpHDGuWdzwAPXAP9r5XlOTcH5GsPh7Qo01eUn17yk+73b+S0SR+NYrF4zN8Y5z1lJ6L8kjoP2Iv2KvjH+3t+0DpnwF+D1jsef8A0jXtcuIybXRLBWAkupiPTOETOZHIUdSR/Td+yF+yh8Hf2KPgTo/7P3wO0MWuk6ZHvu7uVR9p1S7YDzby4Yfflcj6KAFXCqAOH/4Jyf8ABPf4U/8ABOz4A23wn8CiLUNev9l1408VNDtm1i9C4z6rBGCVij/hXJPzMxP0CgGMqe/QV/J3iPx9X4vx/sMO3HC037q/mf8APJf+krovNs/VeHsihlOH55q9SW/l5I4r9qn4k+LPg5+y78R/i54BtbafW/C/gXVdW0iG8GYmube1klj3juoZQSO4GK/k18e+PPGnxR8bar8SviL4qvdb1/Xb6S91fWNQlLzXc8h3NI59Sew4UAAAAAD+rX9t+5jsf2IvjNdzNhIvhR4hZif+wdPX8mqIRGueyj+Vfo3gHQovDY6ryrm5oK/W1m7X7X1PneOqklVowvpZ6fNHqP7CmrT6N+3F8GtThlZWi+Keg4Knsb+FT+hr+sdGA+Udj2r+Tz9g3RJtf/bn+DWkwIS0vxQ0MgAf3b2Jz+imv6v15OfevJ8eGv7Xwff2cv8A0o7OB7/VKvqvyJlccc9qaCCOgNINuBzSpgjr+lfgZ9yLkHoPTt7U7dzx6elNwvp2Hag7R09PSq6AOUg8V4v/AMFHoxL+wF8ao1A/5JfreP8AwDkr2UFQc147/wAFD3A/YN+Mu4jB+GOtjB/685K78o/5G2H/AMcP/Skc+M/3Sp/hf5H8oenn/QYf+uK/yFTgE/nUNoAlnCvpEo/QVJvANf6CJpPU/BJq82z9EP2Dv+DfDxJ+3v8Ass6D+0zoX7XFl4UXWp7uI6Jc+BnvfJME7xE+ct5HuzsB+4MZxz1r1k/8GlPxPBwP29PD592+G9x/S/r7R/4N2JCf+CVXgYE9NV1fr/1/SmvuAsM8tjmv5M4l8TeNsv4hxeGo4q0IVJxiuSm7JSaS1hfRd9T9ay7h7J6uX0pypXbim9Xvb1PxK/4hK/iqG5/bx8N49vhxc/8AydUsX/Bpb8RiQLn9u/QguefL+HM5P63wr9rQ2T96lODXif8AEWOPH/zF/wDklP8A+ROv/VrJf+fX4v8AzPyW+Gf/AAaf/ArTL2O5+Ln7Wvi/XY1T95a6FoVrpqls/wB52nbH5H3r7F/Zq/4Irf8ABNn9l+e21vwj+zrp+v61aurw6142lOrTRuOQ6JNmJGB5DKgIPSvqQeuRTs89a8PM+N+Lc2g4YnGTcXuk+VP1UbJnXQybKsNJShRV13V/zHYQABVACqFUAYAA6AegoZ+fwphIwOv5UuQR9OK+UPUTANgjFeJf8FLwz/8ABOb49hF5Pwe8RcH/ALB8te2DGRx3rxb/AIKTOsX/AATr+PLn/oj3iLP/AIL5q9DKP+Rrh/8AHD/0pGGMf+yVP8L/ACP5VEwqLxxikuMeUxB/hoDDYoz0FNnOY2Gf4a/0KXxH8/L4j+rz9glVX9g/4IKMcfCDw3/6bYK9aRd4HfNeTfsBgf8ADB/wR4H/ACSHw4Bn/sGwU39vT9oJ/wBlP9jX4kftAWTqt74c8K3M+lh+hvHAit//ACK6H8K/z6xWFrYvO54emrynUcV6uVl+J+/wrQo4JVJbKN/uR+eH/Bbv/guN4n+EXivU/wBjn9jDxOLLXrAtbeN/HdoQ0mnS9GsbI8gTL0kn6ofkTDBmX8WNWvL7XNWutf1vUbi9v76dpr6+vJ2lnuZWOWkkkclnYnksxJNO1PVdT1vUbjWdYvpLm8vJnnu7mZyzzSuSzuxPJYsSSfU1XZsDcx6V/bPCnCOVcJZZHDYWC57Lnnb3pvq2+3ZbJH4tmea4zNMS6lSWnRdEivchYU3s20DvnitPwd4A8f8AxBultPAPgTWtblY4WLSNLmuWP4Rqa/az/gjZ/wAEHPgCvwY8O/tWfto+B4fF/iTxLaRan4f8H6uC2m6NaP8ANC80PAubh1IciTMcYZVClgWr9UfCfg3wf4D06LR/A/hPS9FtIUCxWukafFbRoo6ALGqgCvzLiPxoy7KsfUwuAoOs4NpycuWN1o7aNtJ9dL9NNT6bLuD6+Jw8atafLzK9rXf6H8rGi/8ABOr/AIKCa1bpd6R+w58XLmJxlJIfh5qJU/j5Nbll/wAEvv8Ago/c8L+wj8WVz/z08C3if+hIK/qiFzKVAMzn6yGmqQG39/Uk18p/xHrPk/dwlL5uf+aPU/1GwDWtWX4H8nHx7/ZD/aW/ZWm0m1/aL+Dmr+D5dcgkm0qDWURJJ0RgrsEVyy4JA+YDrXnjIAOlfqH/AMHT/jK1v/2tfh54QhuFaTTPh3586K3Kma8l25HbIjr8uzKAMtX9A8IZ1iuIeG8NmOJiozqptqN7LVpWu29kup+dZzgaeAzKph6TbUXbX0R+g3/Bs5lP+Ckd0Sv/ADTXVMn/ALb2tf0Au+eAO/WvwF/4Nk4xP/wUV1SQL/q/hnqRz6ZubUV+/MgAr+ZPGr/kuJf9e4fkz9P4K0yRf4mfnr/wc0y7v+CaYTP/ADUHSD/6Nr+fGv6Cv+Dm0qn/AATVQdz8Q9IH6TV/Pop4r9e8Fv8AkjH/ANfZ/lE+T4zu82/7dR237NKq37SXw6Vuh8faPn/wNir+uVDtiT6Cv5Fv2c5RB+0Z8PJSfu+PdHPT/p9ir+uqPmFD/sj+VfC+PP8Av2C/wz/NHt8CaUa3qv1GMAF6fpQQOcAce1OYIaR1UE49a/n8++FQ4x9RX883/BzCyH/gpvqIGMjwXo4P/flv8a/oZTaBX87n/ByvMX/4Kf62v/PPwlow/wDJVT/Wv1/wSX/GZv8A69T/ADifJ8Z/8idf44/kz4EDAA81+nH/AAatQl/27vHtwB9z4QXQ/PVNP/wr8wd/r61+oP8AwamyB/24PiIDjj4Rzf8Ap0sK/fvEy/8AqJjv8C/9LifDcOQ/4W6Pr+h+8YcHj+lLvUcqOM9cVHuXP4+lLxjOe1fxEfso5HHTHb0oYrkEDt6UwFcE89KCy4FADmZeRx1FAkB4A6H0+lMJyT16+lC7cd6NQJdwzz60m4YH09ab8vr+lLgYGPT0oAFYYP8AhTgwBOB39KaoHJxSjbngHrQAFhj7v8PpSArySo/KghcfhSDbknP6UAOLAtkelKJBwcdvSmnBPHp6UHbjFADt43ZpHKkcep7U35S2M9KX5ewPWgBS4457elCPzn264pMLnv0pBjtQBIGBAwB1oLDnGO3QU0FegP6UqhcH/CgBGPzfj1xTWYEY9O9KQueD+lI4Hb0oATcPXvQME5zSkAAHI/Ok+XOSRQAuV4APb0oJXrn8xSZHcjrQCOgYUAOG3Iy1KdoPB7etJwCAGH50vGSQR0HQ0AMIXbwf0pGCk43elOIVh94daQou7lh24oHcMDnk9KVQAaXYuD0/OhQu7HH50CDCnGMdKftU8DHFM445FPBU9Me1AHx9/wAF77fzf+CT/wAUwByh0dvy1W1r+aScEHr3r+mH/gvMVH/BJ34sZ/556SP/ACq2tfzPXJ+Y896/qvwLduF6/wD1+f8A6RA/LON1/wAKsP8AAvzZ9s/8G7gZv+Cp3g5vTRNWz/4CtX9IKDKjiv5wP+DdVN//AAVR8Id8aDqx/wDJZq/o/TGOa/NfG6V+MIf9eo/+lSPpuDF/wky/xv8AJHD/ALVHgrxn8Rf2X/iN8P8A4ctjxBrngbVbDRcPtLXMtrIiKDxgknAPvX8kmt6NrPhnW7zw14h0uex1DTrp7a9sbqIpLbyxna8bqeVZWBBB7iv7GPNVVzu6V+Y3/Bcr/ginB+1Vp+oftc/speHY4viZZW3m+JfDVqiqviuFF/1kY4C3yqOO04AU4cKSvCPjTA8M5jUweN92lXcff/lktFzf3Xffo/K9jirJa2ZUY1qOsoX07ry8/wAz8HLS5uLS4jvbO5lgmhkWSCeCQo8TqQVdWXlWBAII5BAIr92f+CHv/BbC2/aO0/T/ANkn9r3xhBF8RbdVt/Cfii+cIvimMDCwStwBfKB7eeOR84YN+EkkU9pM9rdwSQzQyNHNDNGUeN1JDKykAqwIIIOCCCDzRb3lzZXMd9YXcsE0MiyQTwSFHidSCrKwOVYEAgjkEZFf0bxjwflfGeV+wxGk1rCot4v9Yvquvk7Nfn+U5tismxXPDVfaj3/yZ/Y0yADpgfSmMFA6/jX5m/8ABDn/AILXQ/tT6bp37Iv7VXiBI/idZW3l+GvEdywVfFcEa52OegvkVSSP+WyjePmDA/pixGBiv4rz7Icy4azOeBx0OWcdn0kukovqn/wHZpo/Y8DjsNmOGVai7p/g+z8wJHrSrjbnt9KZkcnrXwh/wWp/4K5aL+wV8O2+Dfwc1W2u/i54msC1kqlXXw3ZuCPt0y9PNPIhiPUgu3yqA2WTZNmGf5lTwOChzVJv5JdW30S3b/UMbjMPgMPKvWdor8fJeZ5D/wAHAH/BXwfBnw9qH7DX7Mfixo/GOq25i8f+ItNnw+h2brzYxOv3bmVT85BzFGccM42/hxG4TCpwBxwOlWtd1vV/EmsXev69qtxfX19cvcX17dzGSW4mdizyO7ElnZiSSeSTUGm2N9q+p22i6NYT3d5eXCQWlpawmSWeV2CpGiLyzMxAAHJJxX9qcI8KZfwZk6wtF3k9ak3pzS6vyS2S6Lzu3+NZtmmIznF+0ntsl2X+Z0Xwt+G3j/41fEPRvhN8KfCt1rfiPxBfpZaRpVmmZLiZzwPRVABZmOFVVZiQATX9If8AwSn/AOCWvw2/4Jx/B4Q3Is9Z+JHiC2jfxp4qRM5b732K2J5S2jP0MjDew+6q+Yf8ER/+CRVn+wf4Gb44/HHTbW5+LPiawCSx4Dr4ZsnwTZRt3nfA86QcZAjUlVJb79VlwMEfnX8+eKPiJLiHEPLMvn/s0Hq1/wAvJL/2xdO712sffcL8OrAU1icQv3j2X8q/zDYvGAPalVR06UfITyf1pBjBAP8An/Ir8bSufZNHzt/wV28dxfDr/gmJ8bNdFwscl14HuNLgJIG57xktQBnv+9P5V/Lw6jcdpGPav3o/4OePj9Y+Av2JvD3wPt7wDUPHfi6FjECdxtbJfOkb6eY0I/GvwTEu7qa/rLwOy6WF4UqYma/i1G16RSj+akflHG1f2uaRpr7EUvm9f8j6f/4Iw+CU8c/8FRPg7prwmSOz8SSajKAmdot7aWQE+g3Bea/p22AKMV+Av/BsN8J5PHH/AAUC1f4kz24Nr4L8BXk5lYcLNcyR28Yz6kGUj/dNfv6MdyOK/MPG/GxxHF0KMX/DpRT9W5S/Jo+p4LoSpZRzv7Un9ysiPA6g+tC4wfmpxC44x0pqkDuOvHNfjZ9aOGOMHtSNt7HtQSAcAjp60ErgfMOlACcZJrxf/go04T9gb4zEf9Ey1rv/ANOklez5XJO4cV4p/wAFIs/8MAfGjaR/yTHWj/5KPXo5Or5th/8AHD/0pHPjbfU6n+F/kfyn24LW0Rz/AMsl/kKVhg4oszmziP8A0yX+VOI+bNf6ApXPwWTtJn9H3/BvFEsX/BKfwBj+O+1Vj9ft01fbhAJ618V/8G+UKxf8Ep/h0QR81xqhP/gdNX2oVHXcOlfwlxl/yVmO/wCv1T/0pn7nlX/Iso/4Y/kIu3PWgYx1oXGcBh+dKCDzuH5180d4ZHOD37UYXPBFLxgnNNJXcfmHX1oAXjAyRQQuMg0fKDgEdKXC46igBMAtxzzXi3/BSZVb/gnZ8eFbGP8AhT3iLI/7h81e0naeQwrxj/gpMB/w7r+PIBB/4s94i/8ATfNXo5P/AMjbD/44f+lI58Z/ulT0f5H8pkRPlqfYUT8RMc/w0kH+pT/dFLMMxMPVTX+hl7s/BPtH9ZH7CkC2/wCwx8FI1HA+Enhzr/2DYDXgP/BwZHqM3/BKP4kRaepP77SjMB2jF/CW/kK+hP2HdjfsQfBZlxj/AIVJ4cI/8FlvUv7Zn7PFn+1b+yl4/wD2e5JESfxT4ZubSweT7sd3t3wMfQCVUz7Zr+CcDjaWX8WU8VV+GFZSfop3Z+416Mq+Uypw3cLL7j+SqOQkZPpzThNDE6zTrlEYM49QDk1Prmgaz4U16+8LeItOks9Q0y8ls9Qs51w8E8TlJI2HYqysp9xVYgMCCM+vFf3U5RqQvF3T6n4g1yz16H9ePwB17QfEvwK8GeIPC80cum3nhXT5rCSI/K0LW6FCPbGK6wntivw7/wCCMn/BePwZ+zL8LNN/ZJ/bKi1FPDOikxeDvHGn2zXR062ZiwsryFcyNEhJ8uaMMVUhGXChq/UPQf8AgqZ/wTw8S6QutaT+2L4Ba3dNwabxDFEwGO6OVYH2IzX8R8R8D8RZLmtWjPDzlHmfLOMXKMlfR3Sett09Uz9ny/O8txWEjP2iTsrptKz+Z9BAg/N7VneOvG/g74ZeCdV+I3xD8R2uj6FodhLe6vqt9KEitbeNSzyMT2AHTqeAMk18hfHj/gvl/wAEzPgVotxd2vx0HjXU4lP2fQvA1k99NO/ZfN+WCP8A3mkGPevyA/4Kdf8ABZr9oD/gozIvgSPTf+EK+GttcLNaeDLG8Msl9Ipyk9/MAPOccFY1AjQ4IDMN9epwv4X8TcQ4qPtaMqFC/vTmuXT+7F2cn20t3ZzZpxNl2AoNwkpz6JfqzyD/AIKFftc6p+27+2J42/aLumuItP1fUvI8N2VwcNaaXABFaxkD7reWodh/fkavGFk53E80z7JefZG1NbWQ20cyxNOIzsWQgkIW6BiFYgdSFJ7VEJs1/XuBwmGy7CU8Jh1aFOKil2SVkfkWIqVcTWlVnvJ3fzP0q/4Nfkjf9v3xJKVBZPhjeYP1vLYV+97gN3r8Cv8Ag15uFX9v/wASQEZL/DG8x+F3bV++zDI6Cv5R8ZL/AOu8/wDr3D8j9W4PVslXqz87f+DnJCf+CbVuQ3T4i6Tn8pq/n1U4PWv6Ef8Ag5ohEn/BNQNjO34haQQf+/or+fDaAeK/YvBb/kjX/wBfZ/lE+Q4xf/Ct/wBuo639n3DftBeAOf8AmetI/wDSyKv67ocCBP8AdFfyKfs5xmX9o34dwjnd4+0cY+t7FX9dmwKirjoBXwvjz/vuB/wz/OJ7nAv8Gt6r9RGPJwxpGAbJBzzQdoyeKDtz1HWv5/PvRYyq4Ga/nX/4OUB/xtH8Rj08L6L/AOkaV/RQNmMgiv53P+DlWLH/AAVF19wPv+FNFP8A5KqP6V+w+CGvGr/69T/OB8nxnpk6/wAa/Jn5+suORX6c/wDBqjIy/t0/EJM/e+EU/wCmqWFfmWIyRX6a/wDBq3CV/br8eOAP+SR3Gf8AwZ2Ff0D4mQ/4wTH/AOBf+lRPh+HZ/wDCzR9T95kIzljS4U9D9Kbwp6jrTxjsw/Ov4dP2QRQpHWk+Xg57ClCgj7w6UYXOCR9aBpilQSSD39KUBQOtJhcEBhilXaB94daBCkL3pQFIGB2o+XOdwpx2EA57UANCqQeaUBSxGeh9KaCuOCKUYJILDr60ABAIyD2puASeacSn94U0YJJyKAFJXOc9qU7ccHtTWZeMEdPWlBBGNw/E0AAC7slh1pSBjr3pq7S2Sw6047e7dzQAbVyMH6UgCk8U7aDg7h+dIqjOdw6etAAAMYOaUbQCAaBjgkighSDgjtQAmAW4NDIMZyOOlKCucbh+dDbSCQR7U7AN2c8GkK4OPpS78/nSMecmkAmzBBOetLtwRyaAxzjjr2pQ3GMcZoARh/KkOemewpWz+lIc5zjsKAEwQuAacB8wGfTvTckqcilBbdz6igCTGec0hHIIpAxHb3oLNnpQAnQAGlwduMd6YWPBxTlLMMY6f/WoA+QP+C9e4f8ABJ74qAHqdIGP+4pbV/NK43MTnvX9Lf8AwXoy3/BKD4pL6tpH/p0tq/mnlXDn61/Vfgav+MYr/wDX5/8ApED8s42a/tSH+Bfmz7k/4NyLPzf+CpHhufH+p8M6u3/kDH9a/owH071/O3/wbeDP/BT7RuM/8Ulq2OP+mQr+iPoM571+aeNqtxhH/r1D85H03BX/ACKJf43+SHlsDrTc5IH+zSE5HWgZOMelfj59dY/Kj/gu3/wRWX4zafq/7bH7I/hEnxpbo13478IabDz4hiUZe+t0H/L4oGXQf69QSP3q/vPw7R93Xg+lf2OAuDuBII6EGvyI/wCC7v8AwRMHiqLVv24v2OvBoGrIHvPiJ4K0u3/4/lGWfUrSNf8AlsBkzRKP3gzIo3hg39A+Fvia8J7PJc2n+72p1H9ntCT/AJf5X02elrfC8TcOqspYvDL3t5Lv5rz7n40+H9a1rwzrNp4l8N6vc6fqOnXUdzYX9lO0U1tMjBkljdcFXVgCCOQRX9AP/BFX/gsxp37cWjQfs6ftDajZ6d8W9Ks82d5lYofFlugy00a9FulAzJEOGGZEGNyr/PpA6soKnORwQetafhjxN4i8F+I7Dxh4Q1680rVtKvI7vTNS0+4aKe0nRgySxupBVlIBBBr9k404My7jPK3Rq+7VjrTn1i+z7xfVfNanxeT51iMmxXNHWL+Jd/8Ag+Z/Sl/wVm/4KfeBv+Cb/wAIEGmx22rfEnxLbSDwb4clbKRgfKb+5A5W3jY8LwZXGwdGK/zf/E/4kePfjF4/1j4qfE7xXda34h16/e81jVr2TdLczOckn0A4AUYCqAAAABWt8f8A9ob40ftSfFO/+NPx68e3fiLxHqMcUdxf3QVQscaBUjREAWNFA+6oAyWPViTxZkB6HpXH4f8AA2E4Ly60rTxM178//bY/3V+L1fRK8+zytnOJutKa+Ffq/MhmIRSxOBX7j/8ABBf/AII5W3wM8P6b+21+1H4Rz471SAXHgfw5qUPPh20dflu5Ub7t5KpyARmFCOjs23yf/ggN/wAEeoPGl1pn7en7VXhBZdIhdbn4ZeFdSgyt7IrArq1xG3WJSP3CMMOw80jaqZ/aR2ZmLOSSepPUmvyrxY8RliJTyPLJ+6tKs09+8Ivt/M+vw7Xv9Zwrw97OKxmJWr+FPp5v9AX5hggdeKA3agdBmk2sTwK/nxM+/TAckn3p8OXO1eSTx70wcD+tfKn/AAWB/b90b9gX9kXVPEum6rEvjjxYkuj+BbLdmT7Q6YlvMf3LdG3k9N5jX+KvQyzLsVm+PpYLDRvOo1FL16vyW7fRJswxWJpYShKtUdoxV2fjx/wcB/tf2n7Uv7feqeFvCOrrdeFvhjaHw3pMkMm6K4vFkL39wuOOZz5II6i1B6Gvhs5AyOfarN9cy3s7XU8zSPIxaR3YlmY8kk9yTyTXQ/BL4PeNf2gvi/4b+CHw301rvXPFOsQ6dpkKjrJI2Nx9FUZdj2VSe1f3fleXYLhnI6WDhK1OjCzk9NleUn6u8mfhuKxNbMsbKq1703t+SP3A/wCDXb9nS8+HH7GXiX9oPxBYiK9+Jfik/wBmsyYY6Xp4aCJs9cPcvdsOxCqa/TQ9MAfQGuK/Z6+DXhb9nT4JeFPgT4LUDTPCOg22lWjBceYIYwrSfV33Ofdq7MNn5sdfav4b4lzief5/icwe1Sba8o7RXyikj9ty7CrBYKnQ/lSXz6/iLtwAfQU0KcflT8kgcdqYCcZHpXhnYGDn+lNZf5UuSSevSkOegH1oECjrxXi//BR8f8a//jWxAwPhdrfU/wDTnJXs6tjNYvxK+H3hD4ufD3XPhX4/sXutE8RaXPp2r2sc7RNLbTIUkQOvK5ViMjkZrpwNeOFx9KtLaMot28mmZV6bq0JwW7TX3o/kA0/BsoR/0yX+QqbyyxAx3r+kW3/4N8/+CUFvbrDD+z5qWFXAJ8ZagT/6Npyf8G/H/BKYScfALU/ofGN/j/0ZX9Sx8cuE1vRrf+Aw/wDlh+Yy4IzZybU4fe//AJET/g32Vv8Ah1R8OwRjF1qoGf8Ar/mr7RC/MAPSuL/Z5/Z7+Ef7K/wl034IfA3w0+keGtJeZrCwkvZLgxmWQyOd8hLHLMTyeK7Q9Qa/mjP8wpZrneJxlJNRqVJySe9pSbV99de5+lYGhPDYKnRlvGKT+SEA/kKMdsUo5/L0oPHH9K8g6kNwPm4pQPm/GkyeeKATn8aBDtuMc9qUKe9IGJx9OaUMfT8aAEA6D2rxj/gpHx/wTw+PGcY/4U94jzk/9Q6Y/wBK9n6nhawvih8OvCPxi+GviH4R+P7KS50LxTol1pOs20U7RNLa3ETRSqHXlCUcjI5GciurBV44bG0q0toyi38mmZV4OrRlBdU0fyB24AiRT/dFOZCwbA7V/R5b/wDBvF/wSjgiCH4I65JgYzL43vyf/RlRXX/Bu7/wSplJ2fB3xDEMciPxxe/1Y1/VS8deEb/wq3/gMP8A5M/MZcE5s3dTh97/AMj6M/YQkY/sMfBTPUfCPw5n/wAFsFerI5RsisT4d+CPDfwx8AaF8M/Blk9to3hzRrXS9Jt5JjI0VrbxLFEpduWIRFG48nGTWxkkZxX8rYutDEYupVjtKTa9G7n6dQg6dGMHukkfkN/wXd/4Il+MfiZ4z1f9uL9j3wxLqmq6iPtPxA8D2EW6e7mAw2o2aD/WSMoBlhA3MRvXLFgfxkls7izuZLS8tnimhkaOeKRCrRuDhlYHlSDwQeRX9iqkhhivnz9q/wD4JT/sIftp38viP42fA6zTxDMuJPFXh6ZtP1GQ+skkWFnI7ear4r9n4I8X55Jg4YDNYOpShpGcbc8Utk02lJLo7ppaanx2d8JLG1XXwslGT3T2fnfofy5wxjp0qYWURO9o1PuRX7Z+Pv8Ag1g+Al9eST/Db9pzxXpUROY7fVdLt7vZ143L5ZP5VneGv+DVz4YW16j+Lf2ufEN1Bn54tP8ADsELH6M7tj8q/V4eMPAns+f28r9vZzv+VvxPkf8AVDPlK3Iv/Al/mfjB9lG3Crx7Cvdv2Ef+CaX7UX/BQnx1H4c+Cfg14PD9tcqmveN9ViaPStLTPzbpP+W8oHSCPLk9dgy4/a34Ff8ABvh/wTg+Dt1BrHirwNrPjy7gKsq+L9YL2xYetvAI0cf7Lbh6g19r+G9B8N+DPD9p4T8GeHtP0fSbCIRWOl6VZpb21vGOipFGAqKPQAV8dxH47YX6vKlktBub+3UskvNRTbb7XaXdPY9zLOBqvtFPGzVv5Y9fV/5H4+/8Fw/+CdXwY/YT/wCCQvw6+HPwRtZJx4e+LFrceI9du0UXes3d1p95FJdzbeMlkiVUHCIqoOASfxsjkI5P5V/W3+1V+yn8EP20/hLJ8D/2hPDdzqvhyXVLXUHtLTU5bRzPbsWjPmREMBknIB5Br5uX/g3k/wCCTyjDfs/6wfc+PdU/+PV4HBPirl+R5RKhmqq1Ksqk5uSUXfmaerco63v0sd+d8MYjHYpTw3LGKilZtrb0TPzM/wCDYXVBa/8ABSLU9Occ3fwu1Pb9VurNv5Zr+gx84r5s/Zd/4JKfsG/sa/Fpfjd+zv8ACe/0XxImlz6cLyfxVfXaG3mKmRDHNIyHJRecZGOK+kzyOa/PvEDiXAcVcQvH4SMowcYxtJJO6v2bX4nvZFl1fLMB7Cq03dvS9vxSPgD/AIOWoEk/4JkzsV5Xx9o5HP8AtSV/PKwIYjFf1oftT/so/BT9s34Wf8KY+Pmg3epeHzqUF+1rZ6i9qzTRZ2HenOBuPHevnCL/AIN6P+CVKn978Edaf/f8aX39HFfeeHfiXkXCmQSwOMhUlNzlL3Yxas1FdZLXTseDxBw1j80x6rUXFKyWrd/yZ/Pn+y8hk/ak+GSbTg/EXRAf/A6Gv66JfSvjrwf/AMEGv+CXngPxTpfjfwz8CtRh1PRtSgv9OuX8XXz+VcQyLJG2DJg4ZQcEYOK+wnlDHp1NfOeJvGmVcZYjDVMFGcVTUk+dJbtNWtKXbyPT4aybFZRTqRrNPmtazb790hjqCDzTTxmnscrkU1wT0Br8uPpxONnXtX88/wDwctCMf8FPNVKNy3g3Ri3PfyCP5Yr+hcdNvr1r5p/ag/4JC/sKftm/F+4+On7Qnw81fVfENzZwWss9r4mubWPyoUCIvlxMBwB171994ccUYDhHiL6/jIylDklG0Um7txfVrt3PC4iyyvmuXqhSaT5k9fJPyfc/mBjTK5r9Ov8Ag1ctyf22viLKV+58JZOT76pY/wCFfonZ/wDBvv8A8EnbQYP7Ot9LjvN4y1E/ymFeq/stf8Ezv2K/2J/HGpfEf9mX4Ry+HdY1fSP7L1G5bX7y6WW181JtmyeVlB3xodwGeMdK/U+MfGDhriDhnFZdhqVVTqRSTlGCXxJ62m307M+ZynhPMMDmNPEVJxcYu+jd/wAj3EKd3405VwKBjOTQDxjFfzcfoQKvByO1LsBwSelKoJB+lBBBxigBMYB+lC/1pSOtA6c+tACgc8DrzSBcAcdBSAnOKUEkDA7UACg9Aego2kE49aFJxnHOKAck5XoaADb2zTcc49qcST1Hak7GgAK44HpSBcGnYJ7dqT3x9aAADkZ/vUDkde/rR349aQZAzjvQA4EgjntSrngewpgJyM0qntntQA9RRtzk5xSBjShyM8dqAA4B60jdPwo3HJyO9IxOMD0pgAIA6Dr6UmVLZCj8qXAHUd80g254XoKQC7lIHH6UoII6Dr3pvy8YFLkY6UABYZyFHA9KM5Y/LQSP7tHynnHb/CgBNyheFz+FKSM52jrSYUL92lyN33aAFOORtpARuzinEIQcDtTflLcCgAIHGAKUEDoO1IWXpgduaUbTxgUAfIX/AAXoYf8ADqH4pjjppOOP+onbV/NNL1/Gv6WP+C8wB/4JS/FFTjGNK/8ATlb1/NNL161/V3gWr8MV/wDr8/8A0iB+VcbL/hWh/gX5s+7/APg2+IH/AAVB0TPfwlq4/wDIIr+ifAx0Ffzqf8G4ZH/D0HQCc/8AIsat/wCiK/orypGcdK/M/G//AJLGP/XqH5yPp+C/+RQ/8T/JCHB/hH5UqKDghRx7UmQR0pQy8DbX46fXj1AByAPyo345DYx0wabuUHgUhfdjI6UrMD8Uv+C9n/BGNfAE+r/t2/sl+FcaFPK138R/B+nQf8g2Rjl9TtY1/wCWDE5mjA/dkmQfKWC/kxE+7DA9ehFf2IXEFtd20trdW6SwyxtHLFKgZXRhgqQeCCCQQeMGvxb/AOCpX/Bun8Srf4h3nxt/4J3eF7TVNE1eZp9V+G32+K2uNLnY5ZrF5mVJLckk+SzK0ZOF3LgL/Rnhl4o0aVCOU53Vty6U6knpb+Wbe1ukn00fS/57xNwxKpJ4rBx1fxRX5r9UfkyFDDnr2r9Bv+CHX/BJJ/20/iB/w0P8etBcfCrwvf7IrOdSq+J9QQg/Zl9baM4MzD7xIiHJcrV/Y3/4N7f24fjH8VLC0/aR+HF58NPBVtcq+u6rql9bG9niBBMNnDFJITI4+USPtRMlvmICn98Phd8MPAPwX+Hmi/Cj4X+GbXRfDvh/T47LSNKs1xHbwIMAepJ6sx5ZmLEkkmvW8TfEvCYXA/2fk1ZTq1F704NNQj2TWnM/J+6td7HDwxwzVqV/rGMg1GOyfV+nZfibdjZ2ljbR2VjaRQQQII4IIYwqRooAVVUcKoAAAHAAAqVgpPCjp6UmVzjHenfKcYFfzA0fp4DAAGPTtSEIGzS/Lj6Vwn7Rn7TPwM/ZJ+Fl98Z/2g/iDZeHdBsgR5tw+ZrqXBKwW8Q+eeVugRAT9ByLo0K2JrRpUouUpOySV22+iS3InUhSg5Tdkupc+PHx1+F/7NHwh1745/GTxPFpHhzw5ZNdaheS8sccJFGvWSV2wiIOWZgBX8xX/BRP9vL4kf8ABQ39pPUvjn42jksNLiU2XhDw6ZtyaRpqsSkXHBkYkvI4+87Hsqgejf8ABVP/AIKv/FX/AIKSfERbKOC58O/DTRLtn8LeD/PyzsMqL28KnElwyk4HKxKSq5JZm+Rnj9/pX9aeGPh0+FsP/aGPSeKmtFv7OL6f4n9prbZdW/yriXiH+0qnsKD/AHa/8mff07ACG4/Sv2V/4Nlv+Ce1zpdpqX/BQr4oaDtkv4JtI+GkVzFgrCSUvNRXPTeQbdGH8KzY4YV8B/8ABK3/AIJv+Ov+CjP7Rdr4LjgubLwLoE0V38QPEMalRbWhYkWsTd7ifayIB90b5DwnP9OPhLwn4U+H/hTTPAvgjQbXS9F0WwhsdJ0uzj2RWltEgSOJFHRVUAD6V4fjLxrDC4N5FhJfvKi/eNfZh0j6y6r+Xf4kd3CGTOrV+u1V7q+Hzff5fn6F0fKoX0pBtB6d/WnEgjp9OKaxU9v0r+Yj9KuPDAoAFHT0pARj7o/KkUgAcfw05SMfd/SgBDgcBR0pCVz90flTyqjHHbsKawXHTtQAwbf7v04pGI3ZAyfXFKApyNopCELdAeBQNCZGegFOSTkUmF6YpUA3fdFAD8hlGVFGeRxnikG0jp3pflGOKABWA/hFBIxkikBGeFPWlUKRyO1A7DTjn5f0oXAb7o/KnYQj7vU0m1d3TnNA7DgQQAB/nil+UDhRSfIQMD0oJTbmgljdy56dqcGBHuDTflJwBS4G0cDtTYhckdDSMecYHXjilGPQUh2/3RSAQY9B0FKMBcBR1oIXj5e1Jxjgd6AHBuckD8qMqRjaPypMjPTFLw3btQAw43ZAppIOOKkwpONtNIH93tQAmccgD64pVYDoO3pRtU446YpQq7Rhaq4ChgR0H5U8NlfwqMhQMYpy7cZwOlDegDhj0H5UjPxjA/KgbTziglcDipAQHOMqOvpQH5z7+lC7SANnQ01sZzigAMoK4wPypFbOAQKbkEDK05dpPTtQA7gcYpZNuCMD8qT5QOnQUr7RxgUDG/L/AHR7cUqEAnAH4CkXbgfL+lKuz07+lAXH78jAH6Um4EdP0pCV6baF24HFAhCFySVHX0peAcbR+VKdpYnHekO0DIFAApAHKjgUrYODgflSDaB9ztSkoTnb+lACEjnAHX0pB/ujr6UpYHPy96Btx92gAGByAPypSVIGFHT0pDtHQdqPlwAFPSgAXGOg/Klyoz8o6+lINuCNvSlBUngd6AA7R0UdPSm4Uk/KPyp7Bf7vam4GT8tAAApI47UuQAML2pG29h2pSV7DoKpAJnLdB144oO0j7tKu3dkqOtBCEZ29/wCtJ7gNBAx8v0pARnhR09KViowMU0Fc8LSAUMP7o9elOVgAeB6dKZ8u0fL29Kcu3B47+lACgjcTgflQ2MYCjGPShSuTx+lK5U4OP0oAQ7PT9KQgZ6U7GRggde9GBnkjrQAwhSOlLwMcU4qrYxRtB49/WgBvHHH60fKT07c80pVQcCl2jpjsOaAGkLt+7+tBxnJHf1pwUMucUhVd3HrQAjEdh29aTjPIH504gcqOeKTYM4x29KAGHH93sKeT2xQVGRx1HpSlQeB1A/rQB8f/APBegkf8Eo/iiVHfSc59P7Tt6/mol+9+Nf0r/wDBewhP+CTvxUI7f2QB/wCDS1r+aeU5Y81/VvgS/wDjF6//AF+f/pED8s43/wCRtD/AvzZ92/8ABuCEP/BT7Qy/bwrquPr5Ir+ipeAOK/nN/wCDc2Yx/wDBUbwymf8AWeHNWH/kDNf0ZAjb1r8z8bv+Sxj/ANeo/nI+k4K/5FMv8T/JBkHqvT3pAQQMr39aVgu3qOnrQAvBHp61+PH2ALgdFpCMqPlHanKoJIH4U7Yu0HHpQIaF/wBkflSSYxjAxUmxSMZr5z/4KYf8FIfhD/wTR+C9j8UPiRot7rmqa7qLWHhfwvpsqxz6lOqb5GMj/LFFGhBeQg43KAGLAV14DA4vMsZDC4aDnUm7JLdv+vu3Mq1anh6TqVHaK3Z9ClY0+6gH0FLlQuAB+Vfjlbf8HZ9g3/H7+wfc+3lfEJP62dZ/iT/g7K8RSWbx+C/2EbGKcj5JdY+IUkiKfUpFZqT9Nw+tffrwq47k7fVP/KlL/wCTPD/1nyVa+1/B/wCR+zgUM2NnbNUvE3ibwx4K0O48TeMvENjpGm2sZe51DVLyO3hiUDks8hCgfjX8/nxV/wCDmP8A4KT/ABCtzZ+BE8B+AIjn974c8ONdXJB7GW/kmX15WNTXx18df2q/2kv2mtVOs/H343+JfFkxbKprOrSSxR/7kWfLT6KoFfT5V4H8QYiSlj60KUey9+X4Wj/5MzysZxrgqWmHg5Pu9F+rP27/AG5P+DkL9k/4DWN74I/ZWgT4p+Lo90Q1CBnh0KxcfxPccNdEHokI2n/notfip+1b+2P+0V+2z8Tn+Kn7RnxFudcv0DJp9kB5VlpsROfKtoFOyJfUjLN1ZmPNeXoAvf8ACnjB7dK/cOE+AOHuEo8+FhzVXvUlrL0XSK9EvO58XmmfY/NXy1HaPZbf8EcWJOSa9w/YI/YC+OP/AAUL+NEXwp+ENgttYWmybxP4pvYmNnotoWx5khH35DyI4QdzkdlDMvrn/BNL/giz+0X/AMFANSsvHetw3Hgr4XmYG68Y39r+91BAeY9PhbH2hjyPNOIk5yzEbK/oD/Zh/ZR+BH7HPwnsfgr+z74Ig0XRbM+ZMc77m+uCAGubmU/NNK2OWPAGAoVQFHy/iB4pYDh2nPBZdJVMVs2tY0/8T2cl/L0+1bZ+lkHC9fHyVfELlp/jL08vP7ij+x3+yH8Gf2IPgXpPwE+COiG307Tx5t9qFwoN1qt4wHm3lw4+9I5A9lUKigKoFepllIJCjr6UFRkYowuCB6+tfybicTXxmInXrycpybbb3be7Z+q0qVOjTUIKyWyGjHdaacHnFPCqRkUFAWO0Hr2rAsYpXHTtT48Y6dqaEB5x2p0YXHbpQA8BewpCqkcDt61JgZ4oKgngjpQBAUwTlapT+IPDdtdNY3XiKwjmj4eGS9jDqfQgtkVpFAT071+Sf7ZkTD9r34gSBzn/AISFxkEj+BK6cLh/rE3G9j4HxB42lwNltLFRoe1558tublto3e9n2P1eXWtAcfLrlkfcXaf405dZ0ENk63Zj/t7T/Gvxiju71Vwt7MB7TN/jQbu+J+a/nP8A22b/ABrr/s3+9+B+Ur6Qcrf8i7/yr/8AaH7PjV9Bbhdaszz2u0/+Kp39paMB/wAhe14/6eV/xr8XDe3a8/bJf+/rf40+PV79eBfzj/ts3+NNZan9r8A/4mDn/wBC7/yr/wDaH7P/ANp6Mc/8Te16/wDP0n+NIdY0JPva3Zjjvdp/jX4zLrOqZ+XU7gfSdv8AGkl1fVSOdSuD/wBt2/xo/sx/zfgP/iYWp/0Lv/Kv/wBofsymueHzkDXbL/wMT/Gmtr3h5c7vEFj1/wCf2P8A+Kr8YJNU1A5zezf9/W/xqH7dfMeLyX/v63+NNZU/5vwE/pCzX/Mu/wDKv/2h+0n/AAkPho8f8JFYcf8AT7H/APFUHX/Dh4HiGw/8DU/+Kr8XVvL4n/j9m/7+t/jTxd3vU3s3/f5v8aayp/zfgYy+kRNf8y7/AMq//aH7QW2q6PdzCCz1W1mc9EiuUYn8Aas/LgfKeD61+Y3/AATqu7mX9rvwpDNdSMpe5JVpCR/x7ye9fp3gbRz/AJxXBiaH1eryXufsHh7xo+OsmqY50fZcs3C3NzbRi73sv5u3QAFIzig4JBCj8aVdpGCfTvQVUHGK5z7uyGkqTnb2HekAUdFqQqOoHakUKVycfWgQhADY2jrS/KP4e1KQueD39aT5RkD09aBjeMj5R9aRgp5C07C54xSEDjp0oEBA7ijC8cYx70uQO47Upxk4HX2oAjbGM7fwoBHXb2pWAPQdvSlCjHTsKAEXA6L255pWwcYX9aVE4yB2pWQAZxigBqgYGV7460mF9P1qTA2g+4poCk4oAjIBAO0fiaUADkLTiq4BB/Wjao6HtQA07R/DStt9O9I23JORSsFI4PfigYDaQM4pBjOccZ9acqjjp1pVCle3SgegwgY4FKpHBxTyoOcUgRVx9fWgkQFSc7frSgA9v1poAB49fSpAAeQP0oAaoUA4Xt60h25ztp6qCOR2oZQCMelADQBkjaOtAAxytOwvOT+ZowhTIPYd6AEwp7frSELtGAOlOIBOR7UYG0H2oAjGOTilGMkY7+tAAYdO1KAMn6igBflxjb29abxzkfrT9q/pTflGckdPWgBG2k8Z6UgxwNv605lUnqOnrSqqnjjgUAAABzt/WhgvJx345pVUbsZ/Wkk2gYGOvagBjbcjj9aYOTnb196mKgkHHrTUjBPTtQAwYA6dvWlULzx6U4Kp6enrS7VGcAdqAE4DZxSORzx0NOKjJA/CkkXg8Z70AG0jkfypOQevH0p2SR06Um478HPWgYLgbcEfTFLn3FJuPGCfzpQze/WgQEZ6Y70YAPUdBQWOevagydfYetAxCBtzQc5HQ8+tG8svH86AxLd+ooAXAPJweKTGG6dvQ0pc5/DnrSbjn8PWgQ0jgcdqUY29e9BJyMZ6Uqs2MDn61SA+Pv8AgvYN3/BJ/wCKan/qEY/8GltX81EoO81/Sp/wXyZh/wAEovieV7y6QD9P7Tt6/mskOWP1r+qfA3/kl6//AF+f/pED8r43/wCRrD/AvzZ9wf8ABulEx/4Km+FCOg8Pavn/AMBjX9GuRwM81/Oj/wAG5jqn/BUfwzu6t4b1YL9fs5r+i/dkDHrX5r43f8lhD/r1H/0qR9LwV/yKJf43+SEJOeG7jvSZyACc07ccYyabuOB1/Ovx4+wuKuM1IoIAwvp2qJZOeP5mno/AP0oCxKin+7+lfjV/wduSMbj9n+33fKI/E7gZ77tLGa/ZVZCR1H1r8Z/+Dtpx/aPwBBP/AC7eJv8A0PTf8K+/8K9ePMH/ANv/APpuZ4PFGmR1vl/6Uj8cAvTp0p6detNQE9a+5/8AgkL/AMEefDX/AAU/8G+NvFfiD9oDUfBZ8IaxZ2SW9j4djvvtazQvJvLPNHsIKEYwfWv66zrOcv4ey+WOx0nGnFpNpOW7stEm9z8nweDxGYYhUaKvJ+dj4fGAKA4JA6Z9a/dX4ef8GrP7HXh+6+0fEv8AaC+I/igDGLa1az0uI+ufLikc59nFfV37Pn/BIj/gnX+zPJBqPw8/Zh0C51C2UbNU8Ro2p3OR33XJYA/QCvzTMPG7hbDRawtOpVfoor727/8Akp9FQ4KzOb/eSjFfefgD+yR/wS5/be/bS1FD8GfgjqC6NvC3PivX1NhpcAJGT58oHnEA52wh29q/Xj9hL/g3N/Zd/Z7Nl47/AGn7uP4peKIGSVNNuLcw6JaSA54tyd1zg/8APU7Tj7lfovHtjiSCKNUjiXbFFGgVUXsABwB7CpEZhwK/IuJfFnifP4yo0ZfV6T+zB+815z3+7lXkfWZdwpluAanNc8l1e33DLW1tLC1isrG2ihgt4lit4IYwiRIoAVFUcKoAAAAwBUnzHOR29KXJIGeePSkDHHGOBX5e9z6hAcccfpSZPOPX0pWYkjnt60mcA8n8c0hAuSvPNLjJ6D8qAxAOKMknPv1oAQDABHp2FLGxHTt+NAYkD/GhCcc9cUASA88fypxI6cfjTNxycfzNBc9vT1oAMgNnHftX5MftlyQL+1t49RnUN/b7kgn1RK/WXcST9etZ114M8GajeSahqfgzSLm4lOZJ59Mhd3PHVipJrpwuI+rzcrXufA+IXBU+OMspYWFZUnCfNdx5r+61bRrufjQmwjg5pWTOQFJ/Cv2WXwZ4ITiPwTo6/wC7pcI/9lo/4RLwhnjwfpX/AILYv/ia63mX938T8lX0fMS/+Zgv/BT/APkz8YLlhGcMcfXioUnRpQquCfQGv2hk8HeCpc+Z4K0duf4tLhOf/Ha8l/bq8G+DrD9kXxte6Z4P0q1uE0+22T22nRRuv+m2+cMqgjrVQzHmko8u/mefm3gVisqyuvjXjoyVKE529m1fki5Wvzu17WufmEj4A/xpzHcvFQO+D1o80gFvQHt7V6uyPwCUnEbKwRyCCfwp0alv4D+VfrV+y94R8K3H7NfgKW98K6bNJJ4P05pHmsI2ZibdCSSV55Ndt/wg/gYNkeCNHz/2C4f/AImvPeaRjJrl/E/oTCeAeIxmDp1/r6XPGMrezel0nb4z8Y9gA5U/lQeBzX7Pjwj4NX7vgzSOfTTIf/iaZJ4O8FSr+88FaOfXOlQ//E0v7Vj/ACfiaS+jtiZbZiv/AAU//kz8yf8AgnjdiP8AbH8HIG+9Pcg4/wCvaWv1F3gqCMVlWngnwTpuopqml+DNJtruI5iurbTYY5EyMHDKoI4NaSkxrk+1ediq6xNXmSsfsXhzwVV4FyepgaldVeebndR5d4xVrXf8pNkkdO9I3XPFNDnpjtSlmA5/OuY/QBQQO4HAoBwOvpTd7ZwfQUqNuGcDr6UCsBZs4JHWgsf0oJOR/iaQseBj8/woELxuGMdPSgg8Yx27Uqk5Ge/oKDwAc0CEJPqPypd3v29TSMxwefpQWPTJ6d6YDSRx8w/OnKRtJwOgppbI5NKpIJz6DvRcB+3PIHb0pecDH8qQOcd/fmlLnA+lIBMdOB19KbnHPH5UF3wMnvSBzuwaAAEkDnoPSlzlR/U0uTtGPzxRuIGc8e9ADGPynkdPWlPse4pWYkEZ6UuTg/WgYq9vp6U5Pu9TTVJAHsKVW4xzx15oC4rHP5UzpjBHXtSs3HB7UiseCTnmgQgzzjnn3pVBwD7UBie3elDEjkdqAHICF7cDmgqM5x06cUquQD/9ekd2GCaBjWY84b060IzY+8PzpHkIyMHoKFbjJJpJ3EODH179s0gxgYHalyRjPWk3ZGCOnvTAbj5TwOlCjB7dfSgM2OKMkHmgBQcc+1AI/wAiky1CM2c57UAOJOeP7voaNzZ/+tSbjn8KNxBBxQA5WwRyOvpSMflGD1P9aQM2cAnr60Fjj159aAE28ggfpSpn07UBzx249aRWPb0oAf74pB3xj8qNxz07e9KGbn6UDG9GIGPwFI+SPwp247sk0x2IUYpoRLjPUDr3puPnzgdfSnEDPGeRTBw3T9aQxcYAHBpQO20fWk4wOKDgcgUCBwAQAB+dJgknjt60jEE9P1pOM4x2HegYoxtIA/WlYZY9OvrTSQo/+vS5Bbp39aAHMBg8CgA7u1IcHn29aAFLc/zoEBHONvanBQeNvb/CmkL6duKcAuadwPj/AP4L3Irf8EnfinwOP7Ix/wCDS2r+aVxhiK/pW/4L7yGP/glB8TlX+KbRlP0/tS2r+ayReSTX9U+BavwxX/6/P/0iB+V8bv8A4Vof4F+cj7Z/4N4JTH/wVN8HY/i0bVFP/gM1f0dqQRyR7V/OH/wbxx7v+Cpng1j20bVSP/AVq/o6QDAyBX5z44K3F9P/AK9R/wDSpn0nBP8AyKZf43+SHntx+VMI6cf0p2Pb9aawGAP/AGavxs+xQgAyTx19aVQMDp0po659/WhRleB1HrQNMmDA8V+Mv/B2uxbXfgInUCw8SH/yLp/+FfswoyeVzzX43/8AB2nCn2v4Bz4G77P4lXPtv00/1r9E8J0v9fcIn/f/APTcz57ip/8ACHVt/d/9KR+OSdPxr9s/+DT8Rp8GfjG38R8XaZn6C0k/xr8TVA4Fftj/AMGoCn/hTnxjc9P+Es0wD/wEkr+hfF1L/UPEf4qf/pcT8/4Uf/C5T+f5M/WoDPIX9aGUHjAoGMc9veggZx7V/G5+vjQvTgde9OUD0HXtSqi5Hyjr6UoQDPHfigBwQH+EfnSBMKTgfd/vU4ADt+tJgYPHbsaQCFeeg/OjaCDkfrQ2PT9aQkc/X+9QMUKMcD9aNv8Asjt3pFYDt39aXjPH6U0ABeASvb1pFHHI/WlCgAHGePWkUDpjNIBxGScgUuBjG3t60nGcAHp60uAO3b1oECqM8r0NI3XkdKX5TnIpCoLHjvQMaeTyP1o2nd0pThTxTVIJ6dqYXsABI5A69fxryP8Ab7Ij/Y58dMO2n2/T/r9tq9eAU8EdzXkX7fip/wAMcePVH/QLgI/C7t60ofxo+qPn+Ln/AMYpmH/Xit/6bkflMTk0rcxtg/wn+VRNJhs0GU7Tg9jX1FlY/wA85vRn7EfszAD9nLwEF6DwfpuP/AZK7Rzzx61wn7MUhf8AZu8Akk5Pg/Tj/wCS6V3OcnHP518pP42f6NZPrlGH/wAEP/SUITkjIH4mjqOMfnSFQR+FKADwak9JOwgjJbPHT1rC+JfxK8B/B7wfc+O/iN4ig03TbbrLKctI56RxqOXc9lGTWd8dvj18OP2cvAUvj34iap5ceTHYWEBBuL+fBIiiUnk+pPyqMkkCvzD/AGjv2lviB+0141PivxlciC0gLLpGiwSEwWER7L/fcj70h5b2GBXZhcJLESu9I/1sfmfiD4kYDgzDexpWqYqS92HSP96fl2W78lqfWeq/8FdPhXaXbxaT8K9fuoVYhJZJoY9w9cZOKpP/AMFg/AAHHwY1s/8Ab9DXwsy5Gaice1eystwttvxP5yn4zcfuTaxEV/3Dh/kfdR/4LCeBtxx8E9ZwT1+3w08f8Fhvh/HGWl+DGsjHOBfQ18IEcdKjkAGRipeWYZ9PxIfjP4gL/mJj/wCC4f5H7NfCr4gWHxY+Gmg/EvSrGW1t9d0uG9htpmBeNZF3BWI4yK3eN2T0+teX/sTy+f8AskfDplbj/hE7Uc+wI/pVv9qD9o3wt+zJ8NpPGeuQi81C5kNvomkrLta8uMZwT/DGo+Z27DgckA+A6cvbOEe9j+vcJnWHo8L0c1x81GPsoTnLZaxTf3t6JeiOs8f/ABR+Hvwl8PN4p+JPiyy0ixVtqy3cmDI391EGWkb2UE180fFX/gq/4G0a6ex+FPw+utYC8LfanP8AZYmPqFAZyPrivjL4rfGP4h/G3xhN46+JHiGS+vJMrBGMrDaR5yIoUBwiD25PUknmuaaUt9416tLLqaV56v8AA/mnijxyz3HV5UsniqNLpJpSqPz1vGN+yTfmfS2t/wDBVv8AaJu5i2i+EvCNlHnhZLa4mP5+av8AKqUH/BVj9p+3bdcaH4NnA/hOmXCZ/ETmvm+d404ZsZ7E1Fw3Qg+wNdUcLh1pyn55LxG47nPmeYVL+qt9yVj678Jf8FgPHFtOkXj74J6Zcxg/PPo2qSRtj2SVWH619E/A39v39nb433MOg2viN9B1q4YJBpWvAQmZz/DHICY3PoMgn0r8u2iyOetQz26yIVYcHtRUy7Dzjpoz6bJfGjjXLK0XiaixEOsZxSdvKUUmn5u/oftwBjgrj1FDDjkH6V8P/wDBOT9t7Xr3WbL9nf4xazJeG4xD4W1q7kLS7wOLOVjy4IH7tjzkbDnK19wEJxgnp614NajOhPlkf1ZwnxVlnGGURx+DdltKL3jJbp/fdPZqzG7QQBjvSoo3ZIH4UAjgAHr60qj5sD8awbsz6fSwojyoGw5+tea/HH9rf4Cfs/BrDxz4vWXVVQMuiaYvn3XtuUcR/wDAyK8A/bW/4KH3Wk3d38Jv2e9bRZYmaHV/FFuQ2xgSGhtj0yOhl9eF6Zr4hvrq7v7mW9vrqSaeaRpJpppCzSMeSzMSSST3NephsvdRc09EfgPHXjTh8nxE8DksVVqR0lUesE+0Un7zXe9vU+vfiD/wV612WVofhh8HrWJAfluNd1BmJ56lIsflurg7r/grD+1NLIWt9G8GwrnhF0q4b9TPXzlJt68VBIuTXowwWHgvhPwvGeKPHuLqOTx04+UeWK/BI+pNC/4K6ftCWUynXvh94S1CMfeWFbm3Y/Q73H6V7P8ACX/grT8GfFc8WmfFPwlqXhaeQgG8jYXlqCe5KAOB77TX56pEe4p3lA/eFRUwOHmtreh2ZZ4vce5dVUnivapfZqRjJP5pKX3SR+03hvxT4Z8Z6Fb+J/CGvWmp6deJvtr6xuBJFKPZl7+o6jvV5TjB96/JL9nf9qD4nfs0+JxrHgvUTPp00qnVNCuZG+zXq98j+CTHSRfmHfIyp/UD4I/GnwT8fPh1ZfEbwLds9rc5S4tpcCW0nXG+GQDoy5HsQQwyCK8fEYadB90f014f+JWW8b0XRcfZYmCvKF7pr+aD6rut11utX1oxnjFOUDjApnHpThjPeuc/Sh6AMDxyfesH4lfFH4cfB7w23i34neL7PRrAHakl3J80z/3I0XLSN/sqCa84/a//AGw/DH7LfhNY7e3j1PxTqUROj6OZCFUdPPmI5WIHt1Y8DuR+aXxL+K/xG+M/i6bxz8T/ABVcapqM2QrStiOBM5EcSD5Y0HoB9cnk9mGwcq/vS0X5n5Lx94rZfwlN4LCRVbFdVf3YduZrVv8Aur5taX+zfih/wVy8F6Tcy2Xwp+GF5qyqSI73VrkWqP6EIu5sfXBryTWv+Ctf7TFy5OkeEvBtknZXsbmcj8TOv8q+bpMFdx/KqF48aNl2A+pr2aWAwsF8N/U/nXMvFfj3HVHL646a7QUYpfhf72z6Z07/AIK0ftRwzBr3w94KukB5T+y7mIkfUXB/lXp/w5/4K+2Nw0Vv8Ufg/Jb5AEt1oeoeYB6kJIAfwzXwpbyRuwCOCfY1oW+0kcdqupgsLKPw2OXB+KfH+CqKax0peU1GS/FX+5n68fBr9pH4LfHuzaT4a+M4Lm6jj33Gl3H7q7iHqYm5I/2lyPeu6C5ySP1r8XdG1/WvDWqW+veHNXubC+tJBJa3lnO0ckLjoysOQa+/P2Ff2+v+F03MHwi+MM8Fv4qEeNN1JVCR6soHKlRwk4AzgcOMkYPFeNisFKj70dUfvvh/4x4biLEQy/NoKlXlpGS+Cb7a35Zdk20+jvofUhQAHK00LknA6U+QDkc+/NR5yTkevOa4D9yHnrnaOlGM9FHSmnGen60oI9O3rQIAvzYx1PrRjA6Dr60mefx9aUAHqO/96gBQq8cDp60KAR2/Ogcnn065pAQDj+tA7D1APUDpSheD8o6daYOgPt60oIOT7etAARliMDPemuueRjilYjdn+tI+PU8e9AC85yP5U3+L6U/aByfX0ppA38H60CEJ6DPSjg9KOOMt29aUAHnPQ0DGtTfoe1SlQCOT70hTHOT09KAIjnHXt1pf4vxFLtAU84/ClKjd17+lAC96O/4etLgZ5NKEAOAO1AhpB6j0pwyfwpCuAMfyp2Ac4zQM+O/+C+kbS/8ABKD4nEfwS6Mxz7apbV/Nc45J96/pV/4L3kJ/wSb+Kh/7A4H/AINbWv5qGcZ61/VvgT/yS9f/AK/P/wBIgflPHC/4Vof4F+bPtn/g3pA/4em+CiRn/iUaqP8AyVav6OUAx0r+cT/g3qZf+Hp3gj30rVB/5KtX9DPxd+MHwr+AHw71L4s/Gjx3p3hvw5o8XmahqupzbI488BQOrux4VFBZjwATX5343UqlXjKjCmnKTpQSS1bfNOyS6n0fBUowyacpOyUnv6I6XAxjFII5JF/dRs2D/CpP8q/Fv9tn/g6B8c6xqd34L/YN+HdrpGnxsyJ458YWYnu5/wDbgss+XEO4Mxc88ovSvgP4nf8ABTj/AIKH/F+7a78eftnfEO43Nu+z2HiOWwgB9obQxRj/AL5rnyXwV4szKkquJcKCfSTbl/4DHRejkn3ReN4xyzDTcaac2u2i+9/5H9TsiPBxNE6+7IR/Onquen41/KV4F/4KJ/t+fC7VI9Z8Cftn/EuxmiYFUl8X3V1C3IOGhuHeNxx0ZSK+3/2Nf+Dnn9o/4e6ra+Gv2zPBWn+PNEZws/iDQrOLT9VgU9XMa4t58f3dsZP96qzfwT4oy+k6mFnCul0V4y+Sen/kwsJxnl1ZpVYuH4r+vkfuzswcV+Nn/B2srfbPgE2ePs/iYdf9vTK/VP8AZo/ao+Af7YPwvtvjB+zt8RrLxFolw3lzPbkrPZT4y0FxC2HglAPKuBxgjIINflV/wdqyEX3wDjYf8u3iUj/vvTa8fwvw9bDeIeFpVouMouommrNP2c9Gnqjt4lq06uQVJQd0+XVf4kfjqp4Br9t/+DUGID4C/F+f+941sF/KzP8AjX4iq+BX7d/8Gnsiv8A/i/FnlfG1gT+Nmf8ACv33xef/ABgeI/xU/wD0uJ8Jwon/AG3D5/kz9YRz270oB/SnBBnHPWjAGMHtX8bn68LGAQKdgdfehABtye/SgEY60AL1ppzt/wCA0owOmfxoGME57elACMPwphGAeakZeR+lNKg557+lAxq98+tKOenrQABk5NKoHqevrQIUYwOO1NHT8KkVRtB5zj1pgAxxnp60DDPGMUoIzx6daNozjmlIwSc9qAQmcA5oPJJ54o6knJprY3H/ABoACcngUgPP40nGeoFKoyeT+lAhVJI614//AMFBGZf2N/HYU9dPth+d9bCvYQMrjPevHv8AgoAhP7HHjodf+Jfbf+l1tWtD+NH1R89xd/ySmYf9eK3/AKbkflC7ndkGjedp+hpD1pGGEY542nn8K+p+yf56VFa5+xP7Mox+zh4BA/6E/Tsf+AyV2xJyf89q4v8AZnTH7OngNeePB+m8f9u0ddow5Ir5Ofxs/wBG8nVspw6/uQ/9JQcZ6UoORTWI/wAmlTaTgkdfWpPSsfm1/wAFWtV1O6/awi0m81GaW1svCFi9lbPISkJleYyFR0BYqpJ77R6CvnFX53A9K+hv+CrSlf2xG54PgvS8Y/37mvnUbh3r38I2qET+DPENylxtmF3/AMvZfmezfsf/ALLbftX+K9Y8Kr48TQDpOmJeee+nG580NL5e3AdNuMg55r3a4/4I6aipPk/tE2Z9N3hmT+k9c/8A8EfCf+FweMAf+hViPP8A1+JX36/Jx71z4jG4mnVcYy09F/kfs3hp4ccH8RcH0cbmGH56spTTfPOOik0tIyS2XY+G3/4I8eIvux/tAaYf97w7MP8A2rUM/wDwR08VEEp+0FpHsP8AhHZ//j1fdXRuB2pMDsec+tc/9pYz+b8F/kfef8Qa8PHvhH/4Mq//ACZyfwD+GFx8GPgt4Y+FN7rcWpz6BpMdnLfQwGJJ2Uk7gjElRz0JPTrX54/8FC/jZcfFn9o7U9JtrotpfhR20mwRW+Uujfv3+pkyv0jFfp5lo/njUFgMqGOASOgz9a/Ni7/4Jiftjavq13rWr2fhc3N7eTXNw6+IiwMkkjO3/LL1Y1WCnTVVzqPX/M+d8YMrz2rw5hMoyXDznSv7ygnK0aaShFvV9b678p89+ZubPTFfQf7Ef7Fg/aVvbvxb41v7iz8LaZOIZfsjbZr6fGTEjY+RQCCzDnkAdSRMn/BLX9qxBuNv4YPsddP/AMar7Z/ZM+EWtfAr9n3Qfhx4ngtk1W28+XU/sk/mxtLJM7ZD4G75So6dvauvF4uMaP7uSuz8z8OPDHMcdxFF57hJwoQi5WlFpSkmkovy1u11tbYd4M/ZS/Z18A2MeneG/hDoiLGuPNuLJZ5G92eTczH3Jq9r/wCzh8B/E9u1trfwj8Pzo4wc6XGpx9VAI/Ou03cY/WnoVJwD2rxVUq3vzM/q2GSZPToexjhqahty8kbfdax8N/tof8E7vD3gPwfefFz4HRXKWunRmbWPD8khm8uAfemgY5bC9WjbPy5IIxg/G7KpGevuK/ay+0611PTp9NvIFkhuYXimjcZDKwKkEehBIr8ZfGejJ4a8Y6x4ciHyWGq3Fsg/2UkZR+gr2sBip1IuE3qj+WfGngzK+Hcdh8dl8FThW5lKK0ipRs7pdE09lpdablDT7y80vU4NV0+5aC5tZ0lt5kOGjdW3KwPqGAP4V+wXwV+Ia/Fj4SeHfiOqqraxpMNxMiHhJSMSD8HDV+PGDkmv04/4JqatNqv7I+hRzSFjaXl5brk9FWdiB+tTmcb0lLzOrwDzGrR4hxOBv7tSnzW84SST+6bPdyMgV8w/8FL/ANpm9+E3gK2+EXgy/aHXPFMDm8uYXKyWmng7X2kcq0h+QHsofvivqAxuVCouSeg981+T/wC2b8SW+LH7SnirxNFdGW0tdQbTtNYHIEFsTECPYurt/wACrhwNBVa13sj9W8YeJ63D3C3scPLlq4h8ifVRtebXysvLmPM1uGIAB4AxigksoRVJLHCqBkk/SoTlelfUX/BL39n/AE74j/FC8+LfijT0uNP8KFDYRSoCj375KMQevlqCw9GKntXuVayoUnJ9D+SuHsgxXE2d0Mtw7tKo7X7Jayk/RJvz2Nv9mz/gltrPjTS7fxn+0Fq13olpcIHtvD1jhbx1PQzuwIgz/cAL+pU8V9EaX/wTs/ZD0uzW0PwnhusIAZby9mkc8dSS/WvbGC4ySeuc96GJHJPtXz9TF16sruVvQ/tHI/DTg3I8JGjDCQqSW86kVOTff3k7eisvI+aPid/wSx+AHirS5W+HM194Y1HaTA0dy1zbM3YPHISwHurAjrg9K+GPjj8CvH/7P3jWXwR8QdKEMwXzLS5hO6G7izgSRNj5h6jgg8ECv2BVsADmvGP28fgXp/xt+AGrPDYq2teH7eTUtGnC/OGjXdJF/uugII9Qp7VvhsZVjJRm7pnxfiJ4TZHmOU1MZlFBUcRTTlaCtGaWrXKtE7bNJa6O/T8r3Pavf/8AgnV8frn4PfHO18KatqnleH/FciWWopK3yRTnIt5+ehDnYT/ckPoK+fPPVhnnnkcUsF9LZXCXdvIUkjYOjjghhyD+detOmqkGn1P5XyDN8bw/nFHMMNpOnJP1XWL8pK6fkz9tdpGRjGOCKwvih8Q9A+Evw91f4k+KJdtho1i9zMF+9IR92Nf9pmKqPdhVP4EePF+Kvwa8M/EIHLato0E8x/6aFcP/AOPA186f8FafiRJo3w58P/Cexn2nWdRN7fqG5aG3HyKfbzHB+qCvCp0lOqoH918ScTUMo4Sq5xDX3FKHnKSXJ+LV/I+Kviv8TvFXxj8e6n8SPGl4Zb/U7guyBvkhTokKeiIuFH59Sa5mQndnpipdjfWvT/2NvghZ/Hf9oLSPCeuW/m6RZhtQ1iPtJBFgiI+zuUU+xavalJUad+iP4bwWGx3EWcwoRfNWrzSu+spPVt/iz0z9kz/gnPrnxk0i0+I3xc1O40bw7dKsljYWw23moRno+WH7mI9mwWYcgAENX134N/Yw/Zg8DWiWujfBzR5Cq4M9/bC5kb3LS7iTXpqRRRRLDDGqIihURFwFAGAAB0AHGKdgDAJ/WvHqYuvVd+ax/a3DPhzwvw1hI04UI1Kn2qk0pSb62vflXZL53ep514m/ZN/Zs8VWr2mr/Bbw+ysMb4dOSJx7hkwR+dfNf7RH/BLm2sNNuPFf7PV/cySwoZH8M30vmGVQMkQSnnd6I+d3QMDwftkgdfelTbnBXof8mpp4mvTldSZ2Z7wJwrxDhZUsVhYJtaSilGcfNNL8Hdd0fizcxTWlxJa3ULxyxOUkjdSrKwOCCD0IPBFMtdR1HSNRt9a0i+ktry0mWa1uIW2vFIpyrKR0IIzX0b/wVA+Edr8O/j/B410e1WKx8X2Bu5UjGFW9iYJOcdtwMb+5ZjXzYQSCTjmvbpzVampdz+I+Ickr8NZ7Xy6o7ypSsntdaOMl2umn5H6yfsh/H2P9oz4F6X8QLgKupRFrLXIU6JeRY3kD0YFHHs/tXphzkn618E/8EhvH0uj/ABJ8VfC25nxba3pcepWsbN0ubdtj492ikGf+uIr73VO3P+cV4eIp+yrOKP7T8OuIJ8S8IYbF1Xeok4T/AMUHa/8A28rS+YZ+bnNOBBpMe56DvShRwRxWB9uJ/F+Ip3PX370gwDye/pThjGc9/agdmIeTzTfY1JtGRyaaVGcg0ANzzQrnkfzoIA6n9aFXqPp3zQAEnJNDsxo2jcR3odQRnPagB3GODSHbv5P0oDYGd3f0pMsGHzHrTeghdo4pQvHU03ceAGzz6U4M2Pvc/SkAnGeSTxSkDkEn7ooJJ53dvSgtgkbjnFAxuxQvAP50bRu78n0pVJA+/wBxS5O7hu9A2wCAfl6U5UGenak5HIbtSgkHOe3egQjKuM4xSEADr6UrFuOecelId397vQNJWPjv/gvoT/w6c+Ka7uM6P1P/AFFbWv5qHZi2Ca/pQ/4OAJZIv+CT3xN5+9c6Kp+h1S2/wr+a/HzZI4r+rPAr/klq/wD1+l/6RA/K+N/+RpD/AAL82fZn/BBHxR4V8B/8FH/D/wAQPHevW+l6LoHhfW9R1XUbt9sdvbxWTs7sfQLn69B1rnP+CqP/AAVB+Jv/AAUh+Nk+rNPd6T8OdDu5I/A/hRpCFSLJH2y4UHD3Uo5J58tSI14DFvmHTtV1LS1n/s2/ltxc27QXIikK+ZE2NyHHVTjkd6rkknJr9Glw1ganFDzqquaooRhC6+Gzk215u9r9EtN2fOrM68Mt+pw0i5cz89tPlYfCu38qJpoYl3Syqo9WbFeufsNfsa/FP9vT9onRv2dvhMIoLq/DXGq6vdITb6TYR4867lxyQoYBVBBd2RQRuyP6F/2Sv+CM37Av7InhO10vw/8ABLSfFmvrCo1Lxh4106HUL67kwNzKJVMduhPSOJVA4zuPNeLxl4kZLwZKNCrF1a0ldQjZWXRyb2v00bfa2p25Pw5jM4i6kWowWl319D+YVbu1nz5M6OR12uD/ACpO9f1N/tI/8Erf2CP2pvC03hr4n/s2eG4J3jZbXXvDunRaZqVmx/jiuLdVOQedrh0PQqRxX4A/8FTP+CZ3xB/4Jo/HK38D6trT694Q8RRS3XgzxQYBG13EjASW86DIS4i3JuA+VldXXAYqvLwd4oZJxfifqig6NfdRk01JLflkrXaWrTSdtr2Zvm3DWMyun7W/NDq109Tk/wBgX9uz40/8E/PjtZfGf4R6g01q7JB4n8NTzkWmuWQbLQSjorAZaOXG6NuRkFlb7f8A+DkH9o74Ufte/BL9mD9o34Lao11oXiDT/EzxrMoWa1lDaastvMg+7JHIrow6ZXIJBBP5ZqcDAOPetTVvHPi7VPBOm/Du/wBeuJtE0i/ub3TdOkfMdtPcLEs7oOxcQxZHT5Aeua97MuEsFiuJMJndJKNak5KT/ni4SjZ93FtWfa67W4MNmtejgKuDlrCdreTTT/Gxg7xj8K/bb/g0yJb4J/GYE/8AM46bx/26PX4kMMV+2v8AwaYqy/BT4yy54PjHTR+Vo9fL+Lrf+o1df3of+lo9XhRL+2Yej/Jn63hVB6mhlAx1/OlBOQCx/KjLHgMelfx2frQ1VHHXrzQPTJ6+tOAPHznrTWHX5u/pQVdBwORn86F24PPYdqAxH8RP4UKRydx+7QSDBSQeaMDnr19KUqSw2vx34pccH5j1oK6DAARnmlGM5B704ZxgMetJyD17+lAhVAAHPakxwSc9KUMwA+ft0xTd525z2oAcQM85/OggZx/Wm7ju+/8Ahj6UoJ4+Y9PSgFoIAvPtSFRkgE07LDPzfpRkhjk96AuMCqM4/SlAG7j+VOyR0b9KQlt3BP5UCFVVI6968g/b+Uf8MbePMc/8SyA/+TlvXr4LEAZPWvIv2/QR+xr4/wA8/wDEoix/4FwVrR0rR9UfP8W68K4//rxW/wDTcj8m9xyQf502dtsEjE/wN/I0AHfnFJcKTBIMf8s2/ka+nvof551Hoz9kP2bFX/hnjwLj/oUNN7f9O0ddk4UEjNcl+zpGIv2e/Ayg8DwjpvH/AG7R117HBPNfLVP4jP8ARzJ/+RVh7/yR/wDSURso9KWIYbj19aQkkgFj09qVS3QN+dQek9j82/8Agq5Dt/a4il4+bwVpx/8AIlyK+b9pH+NfTX/BV2Pd+1NZTMBz4LsR+U1zXzSEwCSa+gwv+7xP4M8RLLjnMP8Ar5I+tf8AgkGpHxi8XnP/ADKcX/pZHX36QDivgb/gkCufi94xI7eFYf8A0sWvvpieDv8A0ry8bpXZ/T/gyr8A4f8AxVP/AEtibVz16dKVQB+Ypd3JG/t7UA4A+fuO1cZ+p2FUAEf0qUA56VF5hU/KAeegHWvj39rz/gqLbeAtTvvhh+ztFaalq9rI8GoeJrhRJa2coJDRwJ0nkU5BY/u1IwN5BAulSqVp8sEeBxFxPk3CmXvF5jU5Y7JLWUn2iur/AAXVo+vNd1jRPDVg2p+I9atLC2X7015cLEg/FiBXk3jH9uv9lDwZM1tq3xo0uWRRzFp++6b/AMhK1fmF45+KPxD+Kmrv4g+JPjTUdbu5DlpNRumcL7KmdiD2UAVjkgKFjwB2A4r2aOVQa99/cfz3m30gsylUay3BxjHo6jcn90eVL736n6P6n/wVL/ZUs2K2N/4ivsdDa6A4B/GRlrNl/wCCtf7O8XEPgnxlNjoRYWy/+hT1+eOx+4NEg8tcscfU10/2VhUtb/f/AMA+VqeOXHtSV4ulH0p/5yZ+g0v/AAV++BNucR/Czxq/p+6sB/7c18JeNvEFv4s8aax4os4JYodS1S4uoopwu9VkkZgG2kjIBwcEiuekvrYPtNwgPu4q9bESRBl5GOuetKnhKNCTcEfJ8T8e8ScYUqdPMpxkqbbjyxUdXZPb0AKK/ST/AIJcsf8AhlO0GemuX2P+/tfm8VAr9H/+CWxJ/ZStTn/mPX45/wCutYZj/u/zPtvApv8A12l/16n/AOlQPefHniBvCPgLXfFkbANpWjXd4ufWOFnH6gV+NlxvkbfMxZyAXY9S2OT9Sc/nX64/tLzy2v7OXjyVPvDwnejP1iI/rX5FyyHcSfXpUZWkoSfmfRfSDrzlmOAo9FCb+bkl/wC2kEqH/Cv0v/4Ji+E08Ofsk6VqbRBZtd1a9v5G7somMMef+AxD86/M+ZsLmv1a/YXSGP8AZD+HwhI2t4fDHHqZZCf1zVZnK1FLzPI8BcPCpxXWqveFKVvnKC/K/wB56swHakdQCSPWlYnB+b8xR8w43Zrwz+txF5A59O9EkUNxA1tPGJI5EKOjjIYEYIP4UqhuOT0oBIA+b0ouO3c4mH9mP9nKHp8CvC5/3tLQ/wA6l/4Zq/ZykGG+AvhI5/vaJEf6V2LO2eW705JCCDuquefd/eecsmyfphqf/gEf8ip4Z8MeHfBmiW/hnwjoNppmnWoK2thZRCOKIFixCqOAMknj1r88f+CpXiO61b9puLQjLuh0nw7boi7vutIzu36ba/RtSc9a/MP/AIKNTyt+174kSVs7baxCZ/um3XH65rsy+PNiPkfk/jjU+r8FQpU1aMqsFZaKyjJpfgjxNcHg17r+xL+0n8K/2Yta17xN4+0PW7641K1ht7RdGtIZCiKxZtxkljxk7emc4rwZZSuKZcSF15Ga9irShVhyy2P5XybOMbkGZ08wwlvaU7tcyurtNbejPv1/+Cun7PinA+G3jnjv9isP/kumP/wV1/Z+UZX4aeOSfT7Lp4/nd1+fUobOdppg3E4ANYwwGFe6/E+8l43+IV9KlP8A8Fr/ADP0EX/gr78BQTu+E3jr/v3p3/yXSr/wV8+BP8Hwh8c/Upp//wAlV+fwhbOSp/KpBCwXofyq5ZfhUtF+LM/+I2eIb/5fQ/8ABcT6R/bf/bH+En7VHhHQ9P8ACnhLxHpmp6NqbzK+r21qI2hkjKuu6KdyDkIQNuPl69K+btu7gfhUUuQ3OfrilEuOOfyp06UaS5Y7HxGe8QZnxNmUsfjmnVkkm4pRTsrLRdbdT2n9gbxA3hv9q/wlcByouruS0k9xJE6/1r9StgBxX5L/ALHrPJ+1B4FSMHJ8Qw9u3NfrSGJfgnOeleXmKtVXof034B1Zy4ZxMHsqunzhG4m0A4B7dqXAHrzQSc43GhWIP3+1eefugKF3Zx3pQAR+NIpOc7j96lOeoc/e54oDYUgcAZ/Kmgc8fypQzcDd254oQt6nNA2wWMDHPajYgzx2p3phj070nPJ3UCG4XdjHakZV24H4/pT/AOLO7t6U1s4AyetUmgDOemOD60gJLZBApevH5UmPmqQAtwMtQDjof1pQDxg0Y9P5UAISR0FGSWJyOKRwRxjt3pQGJJoAcGJXqKDknGfSkAIGMilI54IpIYvIPTtSAkEf0pcDPUdPSjHzDnvTAaxB5pSScimkHinAEnFA+h8c/wDBfuISf8EmPikXH3H0Zhn21W2r+arcAc4r+tv9qn9mb4a/tifATxB+zn8XhqA8P+I0gF82lXfkXCGGdJ42R8HBDxr1ByMivjOH/g2O/wCCaSJtmm+I0rd2bxgAf0hr9y8MfELIeEskq4XHKfNKo5LlimrOMV3XVM+I4k4fxubY2NWi1ZRS1fW7f6n8+W4bevFNZ8Zr9Kv+C6P/AAR/+A3/AAT4+EPgP4t/s1WviWXT9V8T3OleJ5te1oXYhZrfzbXaPLXYCYpwT3+UV+aJJIr+g+H+Isv4oy2OPwV+RtrVWd07O6TfrvsfAZhltfLMS6Na199PM/aj/g01+HXh+L4dfGn40SW8b6vceJtN8PxTMvzw2sVqbplU9g8k6k+vlL6Cv173kkA1+Bv/AAbPftv+Bv2d/wBoLxR+zR8U9eg0vSviiLOTQNQvJgkMWtW4kRIWY8J58UpQE9XijXq4r98WUhtu0gjggjpX8qeLGExmF43xM66dqnLKD6OPLFaejTT80fqvDFajWyemqe8dH6/8HcazHOK/O7/g5q+H/hjxP/wTePjfV7eP+0PCvjfTLjSJz95XndreRR7MkhyP9kelfohIpAJY9K/Gz/g6W/bT0LUtO8I/sG+B9Xgub621KPxL48EMoY2YRGWxtHx0kfzHnKnkKsRI+cV5nh1gsZjeM8FHDJ3jNTbXSMdZN+TWnq0uptn9ajRymr7R7qy9WfjsrClkI25NMQYHNfcX/BDD/gnL8Kv+CiPx78Y6D8ebPV5PCXhTwit1KdHvzbSHUJ7hI7dTIAeNiXLbe+wdhz/Z+e5zg8gyqrmGKv7Omru2r1aSSV1q20tz8gweDq47FRoUvikfC8hHav26/wCDTYLH8BPjAxwCfG1h1/68zXuln/wbb/8ABL2yXE3gDxTcn/p48YXORx/skV9F/sYfsAfs1fsD+Htb8Mfs2+Fr7S7TxDfRXeqLe6tLdF5I0KKVMhO0YJ4Ffzvx74n8PcTcOVcBhI1FOTi05RSWkk+km/wPv8j4ZzDLsfGvVcbK+zd9V6Htakg8YpQc9+1Nxz1z+ApR9e1fgJ9yOGQBz+tNOT+frSrk4x6jqKQgn8DQAhzQM4JDdqUj3H4ChB1xjp6UAOXI4yelDE4PI60YOR0/KkAPp39KABfqOtDde350oyR0pOe/6CgBMtgc9qYu4Ak4p+0kDp0pApH/AOugAyQc5FO7cjtSfNnr3pevI9KAGgHnjnFK2Q3P86ULn8qQqQetABz1xSAHPfrS7fp+FAHzdqAFTOMH9a8k/b5Tf+x34+X/AKgycf8Ab1BXrRGR1HX1ryX9vJ8fsgeO0I66VCM/W7gFaUf4sfVHgcV/8ktj3/04q/8ApuR+TbR4c8d6WWL9y5/2G/katPGuSQB1qOVMwuAP4D/Kvpb6H+elVbn7D/s5SeZ+z34Gb18I6d/6TR11753H61xf7NJ3fs6eBGx18Iacf/JaOu2Od3HY18xP42f6M5O75Th3/ch/6ShmG4yKFzn3pwX6dPSlUHOOOKk9I/OX/gq5k/tP2XPTwfZ/+jZ6+Z36mvpn/gq5x+1Ha5I/5FCy/wDRk9fM7EZODX0OD1w8T+C/Eb/kucw/6+M+t/8AgkB/yVrxmcf8yvB/6VrX3wScde9fA/8AwSAUt8W/GZHQeFoP/Stf8K++GBHevKx6/wBofyP6g8FnfgCh/iqf+lyF3HPJ4x6U0se5B5pDnJ9vanLk4G7NcNz9WZ4L/wAFFfjtqnwQ/Z1vF8Mag1trfia4GlabcRtte3R1JmlX0ZYwQD2Lg9q/La2h8pBGowFGAO1fbv8AwWGvLuTxD4E0befs4s7+fYDwX3RLn8v518ZR2R/yK+gy6ko4fm7n8X+M+cYjH8a1MNN+5QUYxXrFSb9W39yRFbQXlzcR2dlbSTTTSLHDDChZpHY4CqBySTgACvtv9nf/AIJTNfaRa+Jv2h/EtxaTXEYf/hHdJZQ8IPRZpiD83qqDj+8e3jH/AATz8J6V4i/az8MxaxbpKlp9ou4UcZBljiJQ/UHke4r9RmzuPHfpWWPxdWnJU4Ox9N4O+H+ScQYOrmmZQ9ooz5IwbfLdJNydt90knpve/Tynwz+wr+yZ4ZgWO3+C2m3jAD97qsklyze53tj9K6ez/Zt/Z4skCWvwI8HKAOM+Hbdv/QlNdhk4HJ9qehb16V5ntq0t5P72f0dh+HOH8JFRo4SlFLtTiv0OXg+AvwKhmVk+CPg9TuAyPDFoD/6Lr8pfjqtnH8afFkWn28UMCeIbxIIYIwiIizMAFVQAAAOgr9Yfi58TPD/we+HGr/EvxTdrFZ6TZvKQTgyydI41HdmcqoHqa/HvWtZu/EOsXfiDUCPPvrqS5nAPAd2LEfma9DL+dtybPwLx7qZbhqOCwdGMY1Lym0klpZJXt3d7ejKhY7q/Rz/gliT/AMMqW4/6mHUP/R1fnF1Nfo//AMEtAF/ZVtuP+Y/f/wDo6tcwd6HzPkvAp/8AGby/69T/ADge1/GfQ5vE/wAGfF3hq1j3zX3hm+hhT+85gbaPzAr8dGkz82OvNftjG+CCwGM/MCOvtX5A/tHfDC9+Dnxx8T/D24tykNlq0r2BIwHtZD5kJH/AGA+oNZZXNXlH0f8AX4H2n0gstqypYHHxXurng/Ju0o/lI4aRjjpX6ef8E1fFEfiP9kLw7ZGUNLot1eadKoPI2ztImf8AgEq1+Ybg4619df8ABKD45WPhbxpq3wR1+9EUXiAreaOZGwDdRqVeMe7x4I9fLrpzCm6mHuump+d+DWd0cn43p06ztGvGVO77u0o/e4qK9T75Pc560vPJHr60nsM0uOevevn1c/tW1hUXoeffApW2opZjgAZJPH86VV+Xn9a8j/bm+N1l8DP2cde1sXUa6pqts2l6HCzfNJcTKV3AdwiFpD7L71cISqSUV1ODNMxw+VZdVxtd2hTi5P5K9vnsvM9LTxN4cnO2HxDp7c84voj/AOzU59d0SFfMk1uyVR1JvIwP/Qq/EWLTIYlCpEOBjOKeNIe9YWtvaCWWQ7YowuSzHgD8SQK9L+zHb4vw/wCCfzx/xMBW6Zen/wBxP/tD9wLK8tb6BbuxuIp4X5SWGVXRh6hgSDX5vf8ABUjQrrSP2qm1iSPEOreGrKaI+rRtLE38lr7w/Z0+Ho+EvwN8KfDfy1VtH0K3t5UUAASBMv0/2i1fOn/BW74XTax4E8OfF7TrYsdGvHsNRdR92GfDRsfYSLj/AIHXNg5ezxVumqPuPFfBYrOfDmVeUOWpD2dVx3t0kvPlUn9x8IFx6ivo3/gmj4a+Gvjn4x6r4Q+JHgvStZWfRGnsY9UtEmCPG43bQ3AOG/SvmxSxr0D9mP4ur8D/AI7eHfiPdM32K0vfK1ML3tZQY5T74Vt3/Aa9ivGU6TUdz+WeDsxwuV8UYTE4pJ04zXNzJNcr0bafZO/yP0uf9lH9mRyWb4AeEck/9ASL/Cm/8Mo/syMCh+AHhLrjH9iRf4V6BZ3dpqFpFf2Fwk0E0SyQzRNuV0YAhgR1BGCD707bjrXhe2rL7T+8/u1ZHkMlzLC0nf8AuQ/yPO/+GRP2X3OT8AfC34aUo/lSp+x5+yw4yfgB4WP10xf8a9FUZ6fnTwCF4PfuetJ1qv8AM/vZP+rvDzeuDpf+C4f5HnUX7HX7KZ6/s6+ECfVtFjP86e/7HP7K2cn9njwgPpo0dcd+2B+3X4Z/ZIv9G0BvBsniLV9XikuDYw6ktv8AZrdSFErsVbO58qoxztY9q8ah/wCCzsDf679m98e3ipf/AJHrSlTxtVc0b29f+CfD5vxP4V5Fj54HGxoxqQtdew5rXV94wavbzPqLQP2XP2cvB2t23inwt8D/AAzp+pWUols7610pUkgcZwysOh5/Wu3A45UflXz1+yj+3zcftU/Ee68C2HwcfRrey0mS9utRbXRceXh1RE2CFfvFuueNp619B9eBWdaNWE7VN/W59hw1mPDua5Z9ayVR9jJtXjBwTa0ejjF6bXsBJPHXigA46HpSEH9B3pQPU9vWsj3wBIbk96N3X5h16ZpQuW7daUgkHBHWgBASSOR07GgdMfzNLyDzj8qaPb07UAOB6DA69zQGIB+nekOc5B7/AOFKuTkgn6UAJk7s4pHLcEHjPSlwc5oYNzx0NACduvegDLYzzSBSO36UmDnpxQA7jjPp60uBjpSbSCDijB6hTQAFVz26etKFGTxzgd6YwPb0pcYbGOgFADgMr0pWA3dR19aaBhehoOd3TvSvqMfn+VJj5sY/GmYycgfpQOvC9vSmAp4wMZ4pVBPX09Kbjpxjj0pQBjp270Ah3sDTSw9aAo4IHvTXBB4H4076CPFf+Ci/7I1n+3J+xz4z/Z1e4jg1LU7Fbrw3dSfdg1S3YS2zMeys6+Wx7LI1fyu+KfC/iDwT4n1HwZ4u0W403VtIvpbLU9Ou02y2txE5SSNwejKykH6V/YXg7SAO3evza/4LP/8ABDW2/bKuLz9qH9lm3tNP+KCwA69oUsixW/ilUXCuHJCxXgUBQxwsoADEMA1fsnhNx3heHMVPLswly0KrupdIT21/uySV30aT2uz5HijIqmZQWIoK84qzXdf5o/AsfL04+lfcH7K3/BwV/wAFD/2YfCFp8Pb7xJofxE0OxiWLT4vHltNNeWsQGFjW7ikSV1A6CXzCOgIAAr40+IXw98f/AAi8a3/w5+KXgvU/D2vaXMYtQ0fWLN7e4gcf3kcA49COCOQSKyVcnv8ArX9OZlkuR8R4SMMbRhWhvFvX5xa1V+6ep+c0MVjstqt0pOEuv/BP0U+N/wDwcz/8FAvif4Wm8MfDrw34L+Hr3EZSbWNDs5ru9QHr5TXLskZ/2thI7EV+fPiTxL4h8Y6/e+K/Fuu3mqapqV1JdajqWoXLTT3U7tueWSRyWd2JJLEkmqRbAzmnWFteanexadp1pLPcTyBILeGMu8jngKqjJJJ4AFY5Pw7kPDkJLL6EaSe7W79ZO7svUMXj8dmEl7eblbYY7BBk9B6Cv6Sf+CCv7CWs/sU/sQWmpfEbRHsPG3xFu11/xFa3CbZbKEpts7Rx1DJES7KfuvM47V8af8EWP+CEHjOz8daJ+2B+3J4MOnWWkzR33gz4d6nD/pF1cgh4r2+jP+qjjOGSBvmdwGcKqhW/aOWd5VG4k5PU1/P3i/x9hc1SyXLpqVOLvUktnJbRT2aT1b2vZLZn3nCmRVMN/teIVpNe6u3mxkmCMqR7c0gyTj8qQr0GPxoC4A+X9K/BT7m9hSO/vR14HpTcHOMd6MEHIFAiRB0wDQQD69aaoPAx3pQD6HrQNDiuef60igFSAB0pNrDHFIBkdO1AWFYegNABAzjvSMuSOOntSMDg/wCFS2NIeowDx3oIyTx37GmAE5OP0pdpB+6evpVEseFyB16UgUkdO1IoAA+lJ24Hb0oAcV5wR6UDr0PSmkHPI/SjGPyoAeo9B2pWXJzimLxkDI4p2CCeKAEYEnPtTf4sE9KCP9n9KTBLYx3oAcMYPPevHv8AgoBN5X7Hfjg562dqv531uK9gCEjOK8c/4KDQn/hjfxvt/wCfayP5X9tWtD+LH1PneL7/AOqeYW/58Vf/AE3I/K/cT3pJCDG+f7pphfHekZzsYZ6g19E9j/PibumfsJ+zOAf2cvAZH/Qn6cf/ACWSu26knHeuC/Zbl839mnwA5zz4O04/+S6V3eDnp3/wr5up8bP9F8l1yjDf9e4f+kocwGBx2pVxkZI/PrTMd8fpS7cgDHQ1B6nQ/OL/AIKuuf8AhqmAHp/wiFhj/vuevmhmOTzX0p/wVdYf8NUWw4z/AMIdY8/9tJ6+aG4HNfRYJf7PE/gnxFd+Ocw/6+yPsD/gjwN3xZ8bZGceF7b/ANK6++JOuADXwN/wR3z/AMLV8atj/mWLb/0rr74YEnn0715eYK2I+4/p/wAF3/xgFD/FU/8AS5DSnPTqBTkG7AFNKktwOw7UqLtHI/8AHa4T9XufJP8AwVr+HN5q3gLwv8SrSFnXR9Sls7xgudsc6gqx9t8YH/Aq+EkjC9q/Y/4k/Drw18WfAWqfDnxdbGSw1azaGXb96M9VkU9mVgGHuor8qP2gvgN45/Z48fT+CvGdk2wsW03UUjIhvoc8SIfX+8vVT+Br6DLK0J0fZvdH8keOfCuMweerO6MW6NZRUmvsziuXXspRSs+6ZR+DHxS1P4K/FHRfibo9qJ5dKvBJJbltoniOVkjz23IWAPY4r9VfhV8Xvh98bPCdv4y+HPiSC/tZlBkiRx51q+OYpo87o3HQgj3GRg1+PrP6VNpWq6ro2oJqmiapdWVygwtzZXLwyAdcbkIOParxWChiHe9mfLeH3iTj+B1Uo+yVWhUd3G/K1La8XZ7qyaa1stV1/aHBzgHn6Vw/xa/aY+B/wPsZLj4ifEPT7e6RCY9Kt51mvJjj7qxISwz6nA96/LTUvjN8XNVtfseo/FTxJPEVKmOXXbhlI/775rj7pV8952JMkjFpHY5LH1J6muWnlet5S08j9GzH6QVaVFxwGCUZvrOd0v8At1JX/wDAke0/tg/tl+Mf2o9ei04QPpfhbT5i+naMsm4ySDIE8zDh3wTgfdXJxkkk+OKSV4Nejfsr/sr/ABA/ah8dxaJ4ft5rXQ7Wcf27rzx/u7SPuqE8PMRwqfi2FBNbH7dnw38MfB/9pDVfAng3SEsdLtLGxNlBGOiG2TLE/wATFtxLdySa6+ajTmqMOiPyDOcFxJnOWz4nzJtxnUUFKX2m1J+6tlCPLbTS7stnbyLkV+jv/BLU5/ZVtTn/AJj2of8Ao6vzhLArkV+jH/BK5zJ+ytCn9zxDqA/8iZ/rXJj7+xPs/AuS/wBd5L/p1P8A9KgfS6nAwT+dfJf/AAVF/Zxn8ZeELf48+FNPMt9oFv5OuRQplpLLORLx18tic/7LE/w19YAcfjSyRRXEbwXESyRSKVkR1BVlIwQQeCCCRivLpVZUaikuh/UvE/D2D4oySrl2I0U1o+sZLWMl6P71ddT8U1XcKs6bfX2i6pba7pF7La3lnOs1rcwtteKRTlWUjoQa+pv20v8AgnrrPw61G6+J3wM0WW98OTM017otqhebSj1OxRkvB3GMlOh+UA18qtGc4PBr6OjWp1qfNFn8IcScOZxwpmksHjoOMk/dkr2kukovqvxT0dmj70/Z2/4Ke+A9c0i28P8Ax7Emj6nEgR9at7ZpLW4x/E6IC0THvgFe/Fe/af8AtN/s36rZfb7H4+eDWi25Jk8S20ZHHdXcMPxFfkayYqGf5/vgN7kVw1MBSlK60P07JPHTinK8JGhiqUMRyqylK8ZfNrR+tk31dz9Pfip/wUX/AGY/hvpUs2jeNIvFF+inytP8PfvQ7f7UxAjUe+SfY1+f/wC0t+0p4/8A2ovHKeK/GLJbWVmrR6Po9sxMNmhPJ5+9I2Buc9cAAAcV5+Mk4OT6U0jBOBW+Hw1Og7rc+Z4u8T+JOMaKw+IcadG9+SF0m1tzNtt2+S62uIsYPJFfQ/8AwTp/Zzl+MHxli8a65YFtA8KSJdXLuvyT3XWGEeuCN7DsFH96vP8A9nP9l74mftLeLE0HwdYNb6dBIP7W1yeI/Z7NM85PR3x92McnvgZI/UH4O/CDwZ8Cvh9Y/DnwJY+VZ2i/vZX5kuZTy80hHV2PX04A4ArLG4lU4ckd2fS+FHAGL4gzOnmmMhbC0ndX/wCXkk9Eu8U9ZPZ25e9uozyfrxnqKw/if8OfD3xe+Hes/DTxTHusdasHtpSo+aMkfJIv+0rBWHuoraUc/wD1qkTjg/hXiJtO6P68r0KOKoSo1Y3jJNNPZpqzX3H43fE74X+Lvgz4+1P4b+N7MxahplwY3YDCzJ/BKnqrrhgffHY1gnr0+lfqd+2F+x74a/ai8Kpc2c0OneKdNiYaTqjp8si9fs82OTGT0PVCcjIyD+Z3xM+Gfjr4Q+L7nwP8Q/Dlxpuo2zfPDMvDr2dGHDoezKSDXv4XEwrwt1P4f4/4AzDgzMZWi5Yab9yf/tsu0l+O66pfQf7Gv/BQy/8Agrplv8L/AIt2V1qfhqE7dOv7UB7nTVJ+5tJHmxDqFzuXtkYWvuD4d/Hj4M/FezW9+HnxM0bUwR80EV8qTxnA4eGQrIh57qK/IAAkZHemyRJOw8+NXx90uucfTNTVwFOrK6dme5wr4ycQ8NYOGDxEFiKUVaPM3GaXRKSvdLpeLa720P2qu9T0nTbVrzU9WtLaFBl5ri5SNVHqSxArw749/wDBRL4CfCHSbi08Ia9b+LtdVSsFhpE2+3R8cebOMoAD1Clm9h1r8yDEpwroGAPAbnH51Izbl2n8KyjlkIv3nf8AA9rN/H7OcXh3Sy/Cxoyf2nLna9FyxV/W/oXvir8R/Gnxm+IWpfEvx9qhutU1KXdIwG1IoxwkSL/CijgD8eSSTz4j28jA9/SrZhLtwM19S/saf8E8vE3xH1uy+I3xp0WTT/DELLPb6XdIUm1XuoZTzHB0JJwXHAGCWHXOpTw8LvQ/KsnyXPOMs39jhoupVm7yk72V3rKcui/PZXeh7l/wS4+BN38MPgvcfEnxDZNBqfjKWOeCORcNHYR58gEHpvLPJ9GT0r6bA7c/nTIoYraNLe2jVERQFVVwABxgAcAe1OVeTxz7V8/WqyrVHN9T+6eG8iw3DWR0Mtoaxpxtfu95S+cm38xzDnr2FKoDDGaTbyDjtQAQ2cfpWZ7Yqrk/d+tKwwCMd/WkHB6UMpI6d/SgAI5Bx2pi89qXBJ5H6U0DnkdBQA4HI6frSrnke9R7SMfL0NKAcHA9KAJP4ugpWUHP581Go+bp0pxBxwO3WgAJ4xSZ5zinELyC3fimnZuGW7CgBc5A4P5UmfY0o2Z+9696Q7Acbv1oAQnB6UoJ3Z9hSNs6hu1Kqpk/N2HOaT2GhScjimtnd3609VTb16UMqbs57ipsUM3EcUqnB5Hb1oZU5+b9aNiA9ex71ZIde3anA5/L1puE4APb1p3yYPP60CFBPp+tNfGBTl2evb1pDsxnP60AMYnFODEDBFBVT/Ee3eghQBz37mgDyz9p/wDYq/ZU/bM0BNA/aU+CWjeJhbx7LPUbmExX1oCc/urmMrLGPYNj1FfDHxH/AODWH9inxPqT3/w9+OHxH8KxOcixWay1CJOuQDNCH/Nj0r9OTswfmHtzQNgP3u1fRZVxbxLkdP2eBxU6cf5VL3flF3X4HDicsy/GT5q1JSfdrU/Lrwd/waofscaRqS3Xjb9or4m67bry1pE2n2Ktz0LRwM2PoRX2X+yh/wAEwv2GP2K2i1b4DfADSLTWo1C/8JJqwa/1L3xcTlmjz/sbR7V74hQk89fWhlU4yarNOMOJ86p+zxuLnOPa9k/VKyfzJw+VZdhZ81Kkk/QazNkuWyxPJPU03e2fofSnFUIHPYd6aVUrw1fNnojt2RThnHApnHPzU9CuMbv88UCEKtnp3pOlPwmc7uppNqdm7UANTjGQacCMk7e9IoXjnv60DYFzu/WgY4tkcU0Hg4H8NKNmMbunvSDy8EE9qQ9AJI4NNzjODTm2Z+929aRgnPzUxAhwKXPfH60g2YJDd6cuznB7+tACBsAcfrSKxwRjtzSgKeN360iquCd3b1oEKT6CkJzmlYKSfm7+tAC55bv60AICckfkKduBzgfjSBUORn9aXCgn5u470AIcnoOlIPvD604BM9f1owm7qOtACK3AFcZ+0T8Krn44/BfXfhTZ63Fp0msQRRpezWxmWLZPHLkoGUtny8dR1z2rsflK53D2p6BQR83pRGUoyuYYzCYfH4Spha6vCpFxkr2upKz1Wq0fQ+FV/wCCO+uMMzftAWak9k8LyH+dzSS/8Ec9cKsF/aIteR/0Kb//ACTX3XhTnB7mmtgEjPcV2LH4hdfwR+b/APEHPD3rhH/4Mq//ACZz/wAJ/BTfDX4YeHfh1LqQvX0LRLWwe8WHyxOYY1QuEyducZxk4zjJrfD5Jye1ACnPPb1oAAPB/hrkk+Z3P0uhRp4ajGlT0jFJL0WiF3ggUoYmmHaCBu/WnpsI60jU+aP2tf8Agn1qP7TnxYj+JVt8WYdEVNGt7H7G+hNcHMbSMX3+cnXzOmONvXnjzUf8Eb7/ABl/2jY8+g8Kf/dNfcSMuQdw5PrT8qT1/i9a6YYzEU4qMZaLyR+fZj4XcE5tj6mNxWGcqlRuUnz1FdvyUkl8j5+/Y5/Ycn/ZP8Vaz4ll+KI14avpcdp5A0T7L5RWYSb93mvu9MYHrmvfm6U8sufvdvWo3IBGGHSsqtSdaXNLc+ryPI8s4ey+OBwEOSlFtpXb1bu9ZNvfzDIzj2FLu44/Omts7sPujvQhRl+9WZ69lYmDnua574pfCn4ffGXwrL4O+JHhuDUrGTlRKMPE/QOjj5kYeo/lW8GXf98dRSlkA69SKqMpRd0c+JwuHxlCVGvBShJWaaTTXZpnw38Uv+CSetDUJb34O/FG2a2ZiU07xDAwZB6CaIHP4rXmd/8A8Ey/2tbCUxweGtFvFHSS18QwgHk9pNp/Sv0tYLu69c0hUH8e2a7IZjiYdb+p+WY/wW4HxtRzhTnSv0hPT7pKVvkfm9pH/BMH9qLU5hFqlv4e0tO8l3rglI/4DCj1638Jf+CTXg7StQi1b4z+PJtbCsGOlaVE1tA3+y0hPmMPpt6V9ilVzu7k9KNi8DPbiipmGJmrXt6HRlXg7wNldZVXRlVktvaSuv8AwFKMX80yj4T8MeGPAPh628KeCtAtNL020QLb2VjCI40HfAHc9z1Pevn39qv/AIJ8D9pr4sf8LQi+KqaHu0q3tJbNtC+0lmi3/Pv85OoZRjHG3rzx9IFQCDnrT1AB+X09a5adSdOXNF6n2udcOZLxBlqwGOoqVFNNRTcUmlZW5WmrJ2sfFK/8EekVcD9oQZ/7FX/7qr6J/ZT/AGdZP2ZPhjJ8OJPGS62rarPeR3S6d9m2iTb8m3zHzggnOe/SvTC2BndSEqcfN39Kupia1WPLJ6eiPHyLw94R4bx/1zLsPyVLON+eo9Ha6tKTXTsIWOAAe9ODEGmfLgfP3p0e0kgn6Vgfa2HqxxuB6Hsa8Z+OX7CHwG+OF1Nrl1o76FrMwJk1TRgsfnN/ekjI2OffAJ9a9mGzaMHvTt0ZAGRVQnOnK8XY8zNcnyvO8K8Nj6MakO0le3muqfmrM+B/G/8AwSZ+KFhLJJ4A+I2i6nDk7I9RjktZMduQHX+VcJqH/BMz9rq3kKweDdKuR/fg8RWwB/76ZTX6aHYF4Pb1pkgTJPGd1dccfXjvZn5jjPA/gfFT5qaqU/KM9P8AyZSf4n5t+H/+CWX7UmrShdZ/4RvSIyfme71vzmH/AAGBGz+de1/B/wD4JQfDvw9cxan8X/Gs/iCRCGOn6fEbW2J9GYkyMPxWvrjCggjjjGM01QoGc/rSnjsRLrb0O/KvB3gXKqqqewdWS/5+Scl/4CrRfzTKPhrwt4Z8E6DB4W8G+H7TStOtVCwWNjCI4098DqfUnk96ugk44pxCZ4YdaFVeMn/Oa5G2z9Pp06dGmoQSSWiS0SXZIjVQuRjNODeoxShV6A/rQEUjr29aRY+OQqCM/rXN/FX4MfC344aAPDfxQ8H2uqQICbaWQbZrcn+KORcMh+hwe9dGgTH3u3rQWUEfMOgoTlF3TMMVhcLjaEqGIgpwlo4ySaa809D40+JX/BI3SLmeW8+FHxWe2QklLDXbTzFX2EsWD+amvIvEH/BL/wDat0qZ00vQ9E1VAflksddiXP8AwGYIRX6Ukpk4ak3Lj7/8XrXbTzDEQW9/U/Lcx8F+Bcwm5wpTpN/yTdvukpJfJI/MGP8A4Ju/tjyTeW3wsijH9+XX7IL+kxP6V3Hgr/gk18ctXeOTxz438P6JCSPMjt5JL2YD0woRM/8AAjX6DDaSeRwaVduBj+7jrVyzLESXRHHhfAzgjD1eap7Wou0ppL/ySMX+J4d8BP2AfgN8DZ4fEEmlv4i1qHDJqetKriJ/WOIDYh9Dgkete3s5JyRQ2317+tC7CTlvTvXDOc5u8nc/UcnyTKMhwn1fAUY04dorfzb3b83dgcn8OlIvB+opwCYznt60KFB69vWpPUbELc8A9KMn/PalIQd+g9aRdmfvDigQuTn8aXA7jvQgTP3u9KwTbncP8mgBhJ3ZHpTV6jHpTyqcc9qRRHnr2oATp2pNxGQQafhMde3rTGVMkhuw70ACt83elLkqcDj6U0Bck5705goXAP50ALt/zimleRk+lPIxTSMH8aAG429xSkZ6EUhPbuDSg8DGaABk9T2pVX5iP6UmBnp+lOCbWJoBDguFxn9KGUFs7vSg8rgD9KRjk9OmKB3ApzwRSbBngjpS8jNAPPXtQIaVwRyKNvGM05geMD9KUA4/CgBgXnGRQV/2hwKXBB6U3OR+FAD9owaQrkYA/WlxwcjtSAZP4UAMdCO3cUKDnoe9OKr6U1lAwfem9QHIvcA/nSgZUHHam7gOM9/WlGMDA7UhsUr2A7Uwx+1PXBOPakYAUCEZDzn14pyKexz9Ka4yc4705AOnuKBjgh7A9aQoSQf604YzjPekJBwfQUBqxqqMDBHWk2EcZFPXjoe/YU04DYyetAgVT3FIF+Xr/DS4x0FAVdpJHagBCoz3496QrnP19qXA70AKM59fSgAVMLn3pdmaFHFBxkc9qAAJ0NAT5eRRkYA9KVCO3bNACMoOTgdaTbycDvT2UZ+7TSuO3agY1Vzx7etO28nkfnQvBJA7Uo6/hQITb7ikIBfO4dRTsHbn+lNI+fp3FA0CqMdutOVeeM9KTgDHOM+tA28fSgBdnXPamlSTkYpwx60hIx1zxQIaFGSB6Um0biBTh0JPtSDg9O9ADSnpj86cF9COtLjpwfypAOP5cUAC54x6+tSY449ajHBySeKeDlfxoACvXj9KZIhBHFPAGMgUxlGcYpgNOfyApNuB8v8AOnFADz6Ck2gr1/HFIdxVBDfjTgpNN2jdnHf0p6jvjtQIXb047UmzkDFPUDPTtTXHPSgBAuQOKUxn2oJAHJHalBUAjPSgBjRkdqcEzxjtSNjoacoGPw9KdwG7CR0pGQ44GOaeAD2prLgUgIypGPr6+1CLzkfzpxVeu3HIpMKRxQA4ZwpANOUEgfTiowvAIH6U4Dp9KAHYOKGU880hwCaHPJ+tAC7eBx2pNp5xQMH/APVQFWgBCpzj+tKq8Zz/AJzQeB170Jxge9AAF9KUKc8YoGM9O9KfYfpQA0Jxjjp60EZGNp/KlHQ5Hb0oxwBt7elADSpycKetIoOOnfvTtq5Jx39KaAAKAAA+350qg4FIcDB9qUYwPpQAm0kcA9aUKc9O9Iq5GCKcFUE8DrQAbCAMCgITxjoKcQMdO1IPvHA7UABQdQQOKTbzxj86ccZwPSkwOtACKoz1HUd6cACOvQ9qaoBPPrTgAOAf4qAE28jH9Kbtz+VO68mmkDPAoAMZxj+dATOSCKMYHTvSp0INAAIxk/MKHjUDqOOlOH3utEnCHFAAVHXHemsBnFPJz3H0zTCeenWgBrAdPc0AeuetK3OMj9KF6Yx0oAAAD+FOwMnHpQMnFGTk/T1oAABjr3oZRu5PekzgZ469KCxB/wDr0AKAB+tKBk0mc/kaFPIxQAuFzx6UoAxwO1Jxjn0p2T+lADCBkHHakKjjA7U7OfX86aTjA56UALgYPHSgLQCSM+wzSrzyB260AIRxyaawBxz1NPbpimk9BnvQAzHJwf0oC4A5oAGcn9aUDgY/nQA5FAwc9qVgPShQcAe3Y0HJ5AoAR1Hp3oRRx/vChs+vf0pUB4Ge9ACqMdBSEDp7Uu360jDB4J6UAKuDimnBP40BugHrSDk496AH7V7HvSBF2k57c0uSef6UA4Gc9qAEZBxikKqFPPenN2BH6U3JOee9AAoHIBoPXihfzox/KgBcDAJFIoz19KOcAc/5NIp7YoAkwvT/AD2oYDGPb1pNxP6dqQnt7UAAA546U7AB4pqk8072z270ALtXaMUxlG45bvTt2OBTGJJ6d6BoTHBO7PNIuAeCelGSQeO/pSZOQOOlACg84yacDgcelMX3Ipw+7nHYUCHKOM03b8/GetKD/MdTQMlunegBdowMevrSBBjindhgCkGBQAm0Z6d6CMDA9aXPPBpDk/nQADBGKQgZpwU44NIQDg0AIEGSAf4aNoK5B70uTnr2pM4HJoAUqA3JzTlUAY9uxppOTSqxoAcg5/CmuAOgpQxB4Pb0pHPbHagBGHBwO1AHUY9aTrx7Uqrx39+aAA9fwpyAYzjtTSv1pwJ/DFACgCmuoA/GlBJ6UhPrQA0KMAgn73rTQAacD0A9fakAAoAUIABilCcYoBYqMClU8DigAZBzx0oYd8d6ViT0/HikYkngd6BiqBwRSgZA/DmkBwBQOnP8qBAVxx6GkVRxTmJxgetIpyRxQA3AyRSkAEnOKTHU0pY54/nQAq425yRxQVxgg9qRGwMAHgetBJJBI7UAIR15703A6805jnJx3poNABgZ/Cl2DA47UmTnjtTlzjk9qAEUcH6U4Lgkj1oVePw707HP40AIeDwe1NAFO69s00Z5AFAAeCMelKMYHNI3Xn09aUMRQAq4zkZ60pUYznJ3U1SR+dOBJ6+tABgAg5HSm7BmnAkHigHjGO1ADSgx19KFUDt0px4HSm8gk8UAHGeDSsM5pm4huRSsxwaAHAn1/Wmnrj+ZoIPYn8qQBs8k0DsKwPqPzox7ijaeOvTsaUA4xz+dACZ6YA/GlPfgZx60hDA4Genc0oByevT1oAacgc4oz83Sgg4/iprZJAye3egQ4Enn2pQTnjFNwe2elKqMGzg9PWgB2T7dBTucY44FRkNxjNOBbAGT0oAUgn0/GmsOnI6U4KT0zTdpGASaAFAO38u9KMgdRQF4xk9PWkCkHgn35oACOMH9TSYPHTrSkHHf86QA8cnrQA1VPt+VLyQMDt60vJ6E/nSEEgfMfzoAcvNIQR6dKBxjkjigg+poAGU/rTlHrj8aGXvk9fWhRgAhjQAoz7e/NNcYHanYIOMnrSOG9TwKAIwOnTrQAQce/elA6cmjbzkk9fWgA5x+FKpIPbp60bT05oRTgjnp1BoAXBz26UHoQQOvrSkDPU/nSHjPJ6+tACL06frQevbt3pATjgn/ADij5jySeg700AnzY5NIpOP/AK9O2nAGT055pgQgHk/nRpYBw6DgdB3peQOo/GgKc9+Peja3XJ6UgBST6fjTznOMjtUYB5wT05pxBLZJP50AOJJ6jtTMHd+PrTvoT+dNwQ+C3f1oAbjPQDrSjORnH50hH16+tLtOQeaBpir+tGD7dPShQ2eM04KcdT+dACKDzkjqO9KAd3brQFwDye1KFOcZPWgQhGABkfnRg4xx07mlYHA60hBXufxoAbyT/wDXpTkjPHX1pCCCOTShTjqetACgEjtQ2emO3c0YbHfrQQQep6UAB/DoO9NwQO3H1p5Az36DvSYJXgmgBrA5ycUvOT0oILHv1pQvuelAAPw6UjDp0/OlVTnqaNpOOT0oAAvfA6DvSgHHb8aNpPG49BmlxxjJ/OgBrL06dqACecjpSsDx17UgBA6n7vrQADPFIckc9aUAjpn86QqQMkmgBFBwBx19aUIc9BSgEYyT1pUB756+tADSDgcCjcQPw709gSBgn86ZtwMZPSgBCTnqPfilOehIP401gcnk+9Kc8ncetAxy5wPp60DJGePzpB2IY0LnplulAgYk8+9C5OOe/rQwY9z+NCjgfMetABt5JyOvc0HJ7DpSnrncevrSEZ6elAAoOMADp60ENxjHT1pFyAee3rTsdM56UANOec44PrSYOMY/OnFSM4J/SkCnHegAAI9KVQcDkdKQqe5PQUAEY5PSgByDvgU4D6dfWmJnHU0/nPU/nQAh9CBwPWmgdenHrSsD0BPT1pFDDJ5/OgBWHuOgox9OlDDpyeg70pBzk56etADVB3fjTsnHTvSKuOSx4680EfXr1oAXJ4/+KpBx27+tGMYHP1zSAHPU9fWgBw/Dp60nPPTt2oxk9T+dKoIU8np60ANIOeopHB2n6ZpxHzd/ekdeMEnp60AOI569/SkAIbilzxg+vXFAwW6/pSK1F7AUAcdfzo7DH8qXtjb39KYhrDBzntRjBNOPrn/x2kY8k57elAhjAFeKaeuM9MU/IK//AFqCMtznr6UAIoxn5h0pRgHrThgd/wBKDgEf0FADGGSAPSnADpnt3oI6fT0pwBHH+z/doAAO1IwOB0pRgc+3PFISMD6D+GgA5xg46DvSgE0mSOD6DtThgDj+VADSDjpTTnj61Iw4/L+Gm9xnPX0oAQU0g4FOHUjB/Kl2kgdf++aAGqTx9fWgE4x+tOwOMev92m8bc5H5UDHkkj8aUHjqO1NPOfr/AHaVTjH/AMTQIOvfv6UjZAGD27UuSPX8qQnjHPtxQA1c8Hd3oGN3WgbsjrS49QevpQMBz0FKoO0jjp3o59D79KVCcEAHp6UAgY+tMYnnnvT2+8OvT0prDgkZ/wC+aAtoRjIzk0vAOOOKUKO/4/LQVB6n9KaEGRgfSkUYzz2pVBIHGOP7tKi8HHp6UgDGTknv2pSuenp1zTiP84oPXqenpQMYAQT9KVutGCM4J6elB5Y849iKBAfb+VMO7d1708j09PSkwQ2T6/3aBiAcde9LjkY9PSjOB3+96U5QOMDqPSgBqjnp+NOxx1FAGD0/Sl64B/lQADgHn86TOGJBpc9fqP4aT5d2c/8AjtAgOBj6Ug6ZB+tLk4GM/wDfNH8P/wBagYwD5vw5p3OOvQ0Drj09qXJIAx39KADHy00kg9adkY49v4aQgfmPSgQ0k5654FKCdvApM465+7QuNufcdqYCkknGe9Lz/k0hznIznP8AdpwPqe3PFIAXOen40HJx/jQODwe3pR6Z9PSgB3IGPYUgyBwf1oLAd+gHagkAcfyoAa2OMfypR0xjtTSQcc9v7tL6j2FA7MBmkII+tKDgdf8Ax2gjjp6fw0CAc45HUU4HvmmhiMDPf0pFbJI/pQA8/dHNJyRigEkDn9KOOBn8hQAxhkn2pG59OtOYZGfb+7QwBJzk4P8AdoGICeBmlXOaAOAMHr/doTgZyfyoACOBigHGPmHWnMPT1/u0DOADnr6UCGg84z3oIPalCj0PHtShQRyO1ADADg8/w9xTgRx/hSgEdCenakOARwR06igBccE570mDjg04YweR1/u0gHy5/wDZaBiEc5B7DNJ2GPT0p3Xkn06rQqgqP93+7QA1M889qdkknnvQBx1PT0perYJ79xQIRge/8qaucnBp/U9O392mgev6YoAQnp9KO4+lBGTyD0HUUqj27elAApOeGHWlJ9+9IPvYyeD/AHaXPHOfvdMUAIQeDSLknrTwOQQD+ApFBPY8dcigAA45I60ucA8ilCnHTv6UmByP6UANP3utNY9QD+VPC5fofypGU4wAeKaAM57d6Qbt2dp60uQD0B59KaMbhxU2KuOIIwMUbuwHegEEDj9KDgjoPypkiM7Dt26ZoZmBPB+7QSv90dPSjCnIA7UAN3HGB/KlyxPTv6UhAA4FITzkAdfSgCTcTnjtQCc/hTRjnjtTkYZ6DgUABz0wenrTicDp/DTSwx0HT0pcrnOB044oAQsc4APQ00k8cHpS5B7D8qaSOMD0oAduYg5HanKzD8ulMBUg8Dj2py49O1AxSWwOvakyxxx3oLADoOKARwP60AKCeRj86DkAYGKBgc0bhgYFAhMkAHHbimnPOKfxjpUb4zgCgpDiW9D1pVJxjbTSV54HWnIVAHH6ULYT3FJPPB70053ZAp3y0hxkfSgNwXOBx3pDnnApUK9MUHBzwOvpQGgEknOM0ISFOF7UhxnGB+VAZcdB07CgQ4knHykUfMwOAetNLKSD70u8c8d+9A7gMhTxSNnOaAQQcil+XOcd/SgQgBxyD0oQkAj2pw27Rx2pqgDsOnpQA45z09KDnsvajKjsPypAVIHA6UAGG5wDQfvZxQGXByB09KU4z0FACDPp0FIA27PPalBA4HYUBgWz70AHOMbSDk96UZ4wD0prFSMj1PalDqMfSgBckHp3puTgDFAYE9vypTggcDt2oAaGbHFKpbd0oUDBGB19KFADZwPyoAd2BxSEnGAPxpcjjj9KTK44FADed3Q045I5H8VNJAPSnZUjHHWgABbB69RSEse3anYAHQU1tuenSgBMsDjH8PagE44Heg7M4/2fWhSFHHr2NACtuyeO9GWGfpQSCfx9aMLzx2oAVS2R1pGY9MdqMrkDAprFeDgUAOLHHQ9KQkk9DSYUdhTlC5wAOKAvYY27GMdqdkgHjtxSsFxwB2o+X0HQUDuIpI6Zoy2AAD0oUL6D8qCVxjH5UCDnAGDwaQE5zinjaQB7igIAc4oAZuYKBg/5xSgtxx2oIGPujp6UoCk9BQAnI4obIJ47075eeB+VD7SSMDrQO+gi54470KTikUqAOPShSuMY/SgBzFjzg9fShSQPoaMg9h17ChSvHT8qAFGe1KQfT9Kbnnp344pdwzkY6UCEDH0PSgk8YU9KTcuDn096TAyPpQA7c+DjPWmgtyMGlLKc8dKTcuOg6+lAxSTxwelKrfL93tTcrxwOgpQVwMAdKAYqk4+72pSSOcU1cYJx29KX5c9B19PpQICSe3RaRSeQRSnbn7o6elNAHPyjj2oAVj6DtSZ7Y7UNtwMAdKAQKAFBbd909aU5I6c7qRWG78aXIYcAdTzQAozxhe1C5zwvek3AHoOlKrDOMD8qAH/NgDFJzg/y5pQRgcDr6UblIOMdu1ADcnPT8aVumPekyA2RStJgHnvQAw4J6fpTQeRxS7TyaQA5xjoKAHZzgYo+opvII47mnDPHBoAQ4B4A6UoHP3R0pCPbtTsMM4/ujvQAjAAcL+lNPXkDrTjnb/8AXprAknjvQAmQOw/OlUgH+VIynuKMHIOKAHEj+72pS3fFN5xg+lDA8/WgBQSewNNY8cClAPTFIVOMe3rQMdnjIHanD1xTGBweO1OwegoED4x0pAeBlR1x0pWGV4Hakwf/AB6gBRgdqM8DgdKRVOSc9vWgggDjtQAo6D5aaxGCadgkYx6U1lO38qBrQRjk9O9KrDjA9KYynBzTkDE4/rQA/POQKCRwMUwbscfrS4Ocf1oEOU8AY/GjIOeO9AU8cd6TGRQAHGcYFNBwMY7etOCEgYpNjYwPSgAJyOFpN3X696Cp/T1oIOG4oAFIHYdacDn+Gm9unelXOeAetAEikYHFNUn+7SDIxx2oQkrQAp5/hHQUhOD07UuD+XvSFWzmgAUjJ47UpOTnFNUEAnHalwcng9RQApPtTQcN0H3qUhie9JhienegBVJOeO/rSZ56frSbSV6UYPFACrxzt7U5WyOlMAJPTtR8wOMGgZIpABGPSm8BuBSAnnHtQWOTz2oB6C5GBxQTkYxzjvSHPAB/WjaxoENJ5GAKerDA47+lM2k8Y/hp4Ug/8C9aB2HgYHSmsBn7opec5/rTXzxgdqBbASM/d/hFITxwvekILcewoCkDHv60AOyCeg/KlP0FJg5yR3Hehs549PWgBM4I4pvUg7e9OKkkUm0gAHsKAHD6U5TgDAphBHHofWnYJAxQANjGMfpSdc4HajB/SlKk/lQA0A9QKMnH3R2pSpPQHpSFTQA9SCANo+9TiRj7tMXcB17+vtRlieaAEOMfdHSlHUcDpTWJGOO1LzxweMUAKeOcUSYx0pDnbjFKQSDR1AZ7ADpSp9B0pdh/SkCtjI/PNADuOmKExgcUmD+lKi5wP50AGAeMUH6dqUKSKQr29qAG884Haj/gJoAbH4CkbPGP50DF3cHj9aQH27+tDDr9PWkx1oFcXr2pwIKjjtTAGyeO9OjDbVx2Hc0AKpwM4FKDnHA/KkAPT0o2kk4/nQApx6Ui4zwO/FBBxyOgpBnPTtQArHJ6dqTPI4/Okwf0FADHtQAq9c7Rwf8ACnDOOlNGc/iKVs4/GgBc/N0/WgHJ4A6UmTkcfrQuSOh6UAPB4xj9aARg8dqaAeuD0NLyM8dqAEOA3SkdwOQO/egk5IPoKbICeg70APGQThqbk5HP1qXB9M03ac/doAZyCBupyk9mHWgqBggfpQBx0oAUkgjBpNzev8NLtORwaNvP3T070AM3uF+8fyoYndj3pSg/u0Ffm6d6AAk9M0hzuBBp5Xrx270m0Z6dqAGEkgZPb0pQSep70MoBG2jb7UAKPqaazHoD2pwA7LSFfQduKAF57H0pSSMf0pdg28DtRtHTFACOxx94+1Cs3r3pGXjAFCjAA9z3oAVCwGd360HOAQe1AGegFAUlRhT0oAXJx15pj529entUm3/Z/SmOuRjFJjuMYnnmhSc8N/nilK5OAO9AXAH+FMQikkYz/ninZI/iPSk2gdAOvpS7OOnagBVY8fN3pcErjdRGp4wP0pwGD92gAGeKbk46noKkK9tv6UwJkHK9qAGHIx836U1icHmpWjHYdqY0ZOQB3oAaC2Tz3pyk+v6UKhxnFLt5PHf0oAQFsA7u1IrEjr3p4UYAIH5U1FyPu0DFDNn7wNGWz1H5UpUZ6fpQE4+72oENBbB5/h64pcnP3qFQZPy0pUA5xQA3ce7fpQGO7rzmn7cjpTdnzcjoaAG/MFzu796XJBBDU7aBwF70BQT93tQAwEgjntRzj7w60oUg4C0AcZxQCG/MM4fuKQscnmn7Ac8U3ZkkAUAKGIIw3TH9KUE44bnHpSFfSnBeBgdvSgBoznGe3PFOyc4z/FSbenHanBeM4PX0oGrgC2eM0jHOOT0p4UEDg0hTnhaQMZzk/N2oG4DO7uKkMYz93tTSo7DvTEN3HOB60ZYnP9KUqd2QD+NKEHcfnQAgyMc9vSkJPHPanBBkcdqRlGRgfWgBCx/velKC35UhXAz/AEpypxjb+lACc4/Cnbj6+lNZcDAFKBkYA7UAKM9QRQQcUKp6AUFcjgGgBuWH8R6/0pu47id1P2ggZHf19qYVyeB+lACFjjj0pcnsfSm446Y4pwT2oAXJ2nLGnZJzyabs46fpTmXA6UktQAE8c9qRScde3tS4/wA4pQo/u0wELN/e6ChcjHNG32+tCgCgBQTjOfzoyfXtR0OAKGBxjHagBgJxjPakJYkYb60Fc5+U9KNuOCtACbjzyaUFievegJkH5TTkXjOKAEUNjr3pVLEDntS7RnhaCgwOB0oAau7b1oy2Tlu9ATA59KUKPb8qAELEHhqapPr2p5TngUioAcbe1ACc54PagZ9f1pxQ8EL2o2k/w0AIpO7r39aCxxy3ejbh8YPWjaOm3v6UAJuOR8xoQk9+1Ls5Hy9qFTB4U9KAHA9sjpRuPODRg44FJjg5oAaS2c57ikbPUtj8acVy3A7+lI44PHNAEu7sR+tJxngH86QYGeT1pON3X8KYCsRkDbQOexpoIyBk/nSjBPU0gHADI+U0bQDwD09aaxGOvagkDPJoAUAFeQaCADjB600HcudxpcDdye/egBW+XPBpMnPT170h5IxnrSHAxgmgdgY89+nrQTjsenrTD65P3e5pwC44JoEKOP4T0oznHB6UgxkHJ/OkOMZ3H86BolDccg8Ad6UHHQdvWotxGcE9u9PJHYn86BAwGAdvcdqao6DHfsKU8gdeopAowOvWgBy+hU/iacMADC9u9NQLk9aXAGOtADiRwdv4GmPyM7ewoBGB8x/Om8bM5PA9aABxwcqevrQBnA2/rQyg5Oec0Ii8df8AOKBgM8/KevrSgkY+U9B3pqgEc5pcLxyelAhy9B8ppQ2CTjvTFIwPmPUUmeMZPWgasSg7hkKeR60oAIOFPSolIwMZpUYFTz2oESN1+6elNIBJO3nPrTdwyPmNLgc896AFVePu96Tv070qhQAeabx70AG4YHynp60iEYPynp600DIBJP50KBgnJ4B70ASEknhT1pcjrjt603jJ5P50mQOdx6UAOGOm3tTs/N/9eosjnBPSnoQWzuNADgARwpoIG7Ow0gwR1P5UEANw3cUDA4AOF70YGAdp6UwtgZ3nr60owSOe3rTEIFGeFNAGTjb0oUL2NAC5Gc9KQxyjrwe3egqucY/WkXHPzelOGMkk0CGsDjlT07GgZ5wp6nvTiFO3k/nTcDGeaBjQPmyAaep4+6fvVGAoPXt604KARz0agCVQMYwcig4B+6enrSKB7/lQ3J/DrQIUgDop6CkPI+76daDhe56etMBBXOT+dACsAG+73oJAz8v60xvvd/vetDMAeGPT1oAeGGec9PWkY8YwenrTAQCMN29aUEYB3H7tADj7Keg707GP4T+dNOPU9aUY45PSgBGPAwpoX1weB60hAIHXpSqo9T0HT8KAHrz/AA9vWg9M7TTQQBgE0h28HJoAXt909RTRknGDSDGARnrQNpYg5oGAAAACnpTuB0XtSBV2j6UYUAcnoO9Ag9wD2p34HrTDjbnJ4pcAjO6gY9eMDB/OlX1Cn86auBjBPX09qQbcZ5oEKenKn86FOCCAevrSMRnnNIvQEA9e1AxwYnJwaBjAwD09aZ0GMtwaVcHvQDFUDoUPSjZ0+Q9KaoGOvY0oABA56DnFAhwTr8nekAAX7lLhTn60mBjNACg88j2zQpyAPb1pgOT1PWnRnKjJJoGKBwcD9aUck8d/WkXBXqfrmjjceT19aBDvw7etNGAT8uOPWgEYwWPSk45wx/OmApOTjb2FKBnsfzphIB5J6DvQpB7nrSAdjn7p/OkI4OFPtzQQC2MnqKbjA5B688UD6DwvOdp6etIAB0Xt60mAMHNChfXt2oEOBI42mkB5Pymk+XPOelA284z270DQuBnO39KRwcEAenOKAAWPXrQVBx16dKA6EmeOD3pATu6/pSEnOCD160mTmgQ8Y45ozxnNM3EEYH6U7cSOPX0oAGOT1PSjPJwTSMSD+FIevSgBy5Vfvd/SlJG7Ge9Jk7Tx370hbnA9RQA7Pqf1puMkUb+Rz3pAxyP54oAQg/3u1OHXGaac8cdqUEj8hQAqg5ByeKac4wD2pw6Dg9KRgcYx2oAAeoLUqHHc03Jwcj0pQcAGgBe2MelKMAADHWmnAoBAxx3oAeDg5zQSCAcj8qaG9jSbicUAPBJwCe1JnGSD+VIG6YFKW4yB270AKc9ATSrnHJprHgk+tAYAY96AHEYPDVGxw33qXOe3pTXyTwPWgBVYcDcfypMZzzSLnjHtmlA45BoAD1wDQrAAnPb0oIyM47UKDyMc4oAdzxz+lOVuCcnr6Uwk8YBxS84xj9KAFB44J4pG68E0Bc9fWkPPOKAF6Ac9qF6Hn9KQZ2jAPT1oXpyO1ADyQSeaRhgcHtSbjzilyT1FADACScH9KUkg9aTGc8dqGJDEYNAD9xHBPb0pQSW+939KYCeTg0A89O9AAScHDd/Sl3cgbulNB+XoetLz0APWgAVuc7qdjpzTFyTwO1Kp46dqAHr9fSgH5vvU0NwaTdg8jvQBKO3NIeV60gc8ex9KQsduAKAGtjPXt6U4Hjg96jJOenYU4HoP9r0oAkXhep60pwep7U1WOMA96Mk0AKTzwewprEquAe/pSnrx6Uxjx3oARiC2SR19KXGeM01jlj9aUE5xjr7UACjkfN0FL+PaheT/APWoIx+VAD+T39O9KvTr0FNJwMfTtS5zxj9KAAj37UDj+Lt3oJJxj0pMnH4elACDg8H9aQn69PWjcR69aRumcdKAFzwAGPX1oBwT8xpB24/ioBPYUAPBOBz29BSjJPWmdBwO1OBxg4oACDk4PelbOSAaYxwM4oZuvNADwenzd6TJI5Pamhs4x6+lIGyvQ/WgBzHkn39aRcYx/WkYk8EHrQueMk9aAAgZODS5+bOaTGecHrRnIzj86AHIeD856UuckfN6UxTgEY7Uu7GOCKAH7hg5Pf0po5HB/SmluT1/KgHI6HrQAp69acBhRz2puCedppVzgfKelACjp15x6UE/MTn9KaCcdD0oJ68H86AHEnpnt6U36N29KCTnofzpFPtQArDGPpQDgdaNvPC9hRg46dBQAoOW+8eD6UEjGOfvU0tg9O/rSFh7/eoAcGwQM0A+jD8qaDkg4PFKoOOhoAd1Awwo9eT2pMnsDQpPIwe1ABjk9fzpWAwQB9OaME0MOcbupoGLgf3u/emkc9aUE46D86T+LJIoEIeoGfXvSjpwe/rTc5IwV6HvSgk8YHX1oAUD3H507ABPPFId2egpSW56dPWgBCeM5xSMCD+NC7ypzj86Gzu6DtQAZIzyOlCg5xjtQc5wAKMEMOB+JoAQgDjHanDGPwFIQf7vb1pV3Ht6d6AFAA456etIwGOB29KUbscAdPX6UjZwOB0HGaAADjk8/WhScdaRiQPu/wANIGYY47+tACkHAoB6c96DyB8o/GhSeOn5UAA69aMZA/nSJuyc4pdpwMhenpQALnA69KMjPGaAOAcDtTTwvbp60APYjHB780gPHXvTWY849fWnJuIHT86BjgAcnPf0pGXOMelCgkZAHX1pTk9s8UCERRxz39aXHXBpBnjIHX1oGeuB+dAClQelAAI4z0oUnHQfnQpyMEDp60AGMkUEYBP9aRsgjjt60pJwcjvQAD2NBAPQ0mWAPy0At/d/I0ALjAHPamgcfhSgkgDb+tNGdp4HT1oAcQCenp3pTjGM9qT5snj9aTk9u3rQAqgetDLk5Hp60i5GTjHHrThnPA/WgAKjsR0pABu696AWx939aOQ2do6+lAABjjPeggcYb86Ri2DwOvrQGbg8dPWmgAAZ4PajjANC5Pp09aDn0HT1pAKOmc46daT+LPvSDdngDqKTnd0HWgB/HAz3NHBGM035uBx+dOXJU8frQA3Az1pxHHB/i9aQg57fnQc5xx971oAcvK/jS9xTVz0wPzoOQQcCgBxb0Y9KQjK8Gglhxj+H1pis205AoAVhzSAkUH5j0FBBHOB/k0AOTrnHahl9B2pAWB6frSnPTA6etACnnv6UoHGM96Rs4+72HU0vPdf1oAG4A5PSkONuM9vWkYnHQdO5pSxGenT1oAQZz1pDyOCPzoy3YDt3oYEjtQAuOAc9xQOT1pATgdPzpEYhjwKAHDlcZ7Uo7f40wk7RkClBOBwOmaAEbnOKGznqOtIxbbkjtSHPoOvrQOw5egO7vQoGOvahc4HT86RQdvO386BCnBGM9qVCDj5u9NIJGTjp60q5AGMdaAHcdj3pGHOQaQEnqB+JpQSRkgfnQAKMgkHtQwxjntQuQDgDoKRmfI4H50AIRgn5qVQOxpCWOc4/KlXp2oAfgZ4pSq7QQe1NVmPb8jShiQPl7etACDpnPahlHNABwSAOnrRhsngdfWgBGA3HHpQo5JpfmIyQOnekUEE9OlADyo4IHakYDHB96XDZ6dh3pAW/u9vWmBGeG5bHPrR2PP8AFQ27PTvSNk5HGM+tIB6gbgQe1KoGePSmoTxnHSlUt6dqAsOwMZJoAGDk0Zb0/WkOeenJHegaHYGetBxjA9Kbk5PT86Cx2/hQHQMcfdpCBuztpxx19TTTtHGP1oEIeMcHrQApPKn8qCADjHenLgjoetABt9AaUrn+E9KQFR2PA9aCAGwM9PWgBQoKk7T+VIyjd0NBI2n/ABoJG7oTz60AG0A4KmlCgEHFAwM49PWlG3I/Q5oAayjsvb0pQueqntQcEj6etLwf/wBdAAACMbT0prAY+6elOXbwRTWxtzg9PWgBCnBO386AP9mnYXqOKABjOD+dAxpUYHynr6UmMkfKTTiFAxk/nQAOOvX1oEIqAHG0/nTtoAHymhQBzn8c0DbgHnp60AGwEDAPTvTCoI+729Kk4GODx70mB3z9aCkMZB129/SnIuMfL0pzBQCR60mFx+XegQoUdNnSgr/s9vSkBA796XK0AAUDBx3HakKjPA70qkY7/nScEkY7+tAhNo7L2oQcEFT92nYUnp+tNG0g89vWgB2Bn7p6UmwHPynr6Upx2zR8uCeevrQA0LgH5TSHGcYpRgjJB/OkO3P4etABgYGF7elIoyD8ueKcNuB9PWkQDGP60AKyjP3TSbQMDaelO4P1+tJ8vr29aAERAc/LninEDOQvb0pF2jP+NOIXPT9aAGlR02mmgYb7vennaT0/WmYTd36+v0oACMA4U9aUAEglT0owoB470oCkgY7UAIq842np6UuBt4U9O9Iu3sT09aX5cfh60AJtHJ29x2pAuG5Hf0p+AQev500BQ5PPWgpIAoAHye/SnAEjhc0ny8fT1pQQQaAYgUE/d7UpUD+H+L0pARuA/OngjHfr60CY0AE/d/SgqB0U9KcFBHT9aQ49DwPWgQjJznafujrTNoxyvcdqkIUHj0FMYDB56GgBCvPfr6Uu3nJBpnAbqetPGOfp3NADkTB+72oKgfw9vSlG307cc0HBwB2FAAwGPu9vSk24/g/SgkdOemOtHB5xQA0rj+Ht6UuBg/L/AA+lIwXGRn86cMY5HagBFAPG2kK8fdP5U4Aenb1pGwB3/wC+qAG7emFPWmKMNyh/On7RwOevrSBRnpQAbflGAfu0uMn7p6UqhdoAHb1pcLgDH60ANKDGNv40pQZ4H6UuOv8AjStjJ/DvQVcREGB8p4pQoHGw9KAQMfT1oG3GcfrQLcQqMZ2npSAcA4NObb156+tIuCOe3vQIbtA/hpQMDG09KMKfzpSFB6Hp60ACqCD8vQU0gcDaelKBkHqMD+9SHBx9PWgBNmAflNIBkH5T0pSBzjPWgAY6Hr1oAM88r3py4wBtPT1pAF3dDSqAAM+nrQA9VG05XtQVAJOw/lQgG3P9aUYJoAYwx0X+H0pqjk5U08hSMgdvWmgAMcg/XNADjjP3T0oAB/h7UnHTnp60oxn8OeaAE8sFuUPWhk68Hr60oxu4z1owMdD1NADVQHB2ngd6EUeh6UuAT36etCAZ6H/vqgYuBj7p+7S4xkBT2owOOOw70ZGDigLCY+bIB9+KGAwfp70hbDf/AF6R2BGPf1oH0DOT1pvUjnvT8rnGOntTcgtnFBIgPzD5u9KDngsPyoBXgYoyOmKAEbJz9KcDzwaRiB/D2o3An7vagA6jqaDndj3FKpXGAoNKQu7p3FACbiOSfqcUucEYP4UEqB0FAK7gQKAFwTj+eKOT+dJwcYWlBGPu9/8ACgAz6tTWJxjNLlTg7aT5TnA96Bi5I7/pQT6H9KX5cEbewo+XsKBCMOME9/SgDJ6/xGhiMcClV1wMr3NAABk9aTHA/ClUrnhf0oyAOF7UAIM9OelL1XJPalGM/dHSg7QvCjp6UADcDrSD0z39KVmXnjv6UIwznHekMb+NIWO7r29Kf8u37vamOVBximCFVumD3o565/Smqy8cdD/SlDKR0oCwvUc+npTRzkew/nSgrjhe1KpX+72HNABnJwX/AEppPB69e1P+Xj5e35U1tvIC96BDQc5z60Alj1/SlUAfw96FwT0oAVemQe1A57/pSqy4Hy9qFII4Xt2oAM8jn9KQn/a9O1KSu7hTQMf3aAEB6nP6UpJz17+lIrAE8DpSgqTkDtQAgyR9BR/EOfSnfKei00kA/d7igBMkDO49fSgckc0Erj7vel3JxhKAETBH4UvfAPb0oUgdF7Uu5fQdOtACZPOT+lBHzHnvRuj5HHbtQCoJGP0oKTE3HjDD8aAcjqKXcnHy96AV6bfxxQIQdevpTwff+KowV9KcGXrj+KgQ/PPWkb69qVSp6D9KCQDyKAEOeSTjj0ppG5ev44pxYE9O3rSAptztFADSuWwfWl6d6Usu77veglcngUAAPIyaGJOM9qQbQQcfpR8pHA7elACknnJ6e1Gc9/xxQzLjp+lGVzwtADTnH8qASD16DvQxB7fpSgr029hzQAoPv09qG+7SqVHGBSnbwCo/KgBmO3P3qQDJxk9e9OUrgfKOtA2g/dHX0oAQAYFL2FLlQB8vagHgcDpQAw5Iznt1oY5J57insF2/dHSmtgZ+Xv6UAG4gDBpN3HUHj0oyvBx2oBXbgDtQApb1bvQp4Hzevamkg846mnIwyBjoTQAoPGd1GMgEUAqBwMUoZewoAYvfPpQc/wB7t6Uu5ccr2pMrkfLQAhHX60KMfnS5XB+WlXbjp3oARRkdTx605BwMN2pV29l70ApgEDt6UACnjj+VLng8jj2poK7TxShlyTtPWgBD0/8ArU09+fpTzj+72poA54oAB7k4xSgZ6Hp7UfLnOO1LuX0/SgBB1GT/ABUHpw3ejcoP3f4qNwI+53oGlcQDPJOeKVeuPalUr0C9qBt/u9qBB04pCTzg9qUsoxwM03euTx6dqAGk4br34prsSevWn5UtkAdfSmsFwPl6UASBSOlN2nIGO1SAMTwaTYc9aAG7Oh29ulLt74pxVuMntQVO3gn6UDI36dO1IME49h3pWDZ6nNKFcHknoO1AhAFC9P1pSfm/EUoDFfvUMGzkE9aAAgZz/WlHUcUc80KDnjHSgAI6cUhHGMdBSndkf4UDd39O4oAbgnjbR7Y7U4A9yOnpSMpAznt6UAO4IzSMAR0oG4jr25pRux+FADG6YpMjOAO9PZWxTQrcc96AHKAOcfrQcbeB2oUORSjcRnPagBePTtSOOMAdqcNwx/hTSCce9ADXBOeO9C9B/jStuPp19KVQcDn0oAOM9P1pjjJ+7UgBzjcPyppUkfh6UDIsEkf73WlXr0704K5OSeT7fSkKEZ570AAx2BoByM4PSnENnIFIobB5PSgQ7aDyVpGAOcCl+YnH9KAHOcfpQPoIq5H40u0Z4FKA2OhowffpQIaB0+nrSL0p+1iBg/pTUDAHHpQAuwE9P1oKj07UvPXPp2+lBDdPagBig+nalA54Hb1oVW5+npTgjE8ntQAgAB4FNP3sZ70/DHp39qbhs5Ld/SgBmARjB/OgZ3Dg/nSlTyc9/SgI5IIJNAAv07U4rx07UiB+wNO+bA+lADdpOcik5DYxTsHkZ/SkCndn39KADaMDjtSYz0FPKtwAaQo23g9qBkYHTA7U/aeDjv60mxt3WnhWI69D6UAAHcDFDntTlB7U0q+R9PSgQ3OeCP4Rnmmt8oOB+tPKtnPfAppDEdaAG55wR3FOHJoKuTj+lKFPr29KAFC7mHGMe9GCB07c80qbt3X9KGDcc9u4oAaV46UAZ/hNP2N29u1ARs8/yoAYy5GcUBcgnHanspIwB2FGGI5Pb0oARcY6UHjHFKqmkZWx/wDWoAQdBwetAAHIH60o34Az39PakBc//qoAQngcdqcBkYx29aT5tv4elKNwx/hQApAI4XtSOp9KUg84ofd69/SgBm3I6dvWgLjoKeFbA5/SkAPUkdPSgBhUjt0NEeSRxTnRsH/CkRWBHP6UAIowMYp2ADwKFUnp1pSrHoR+dADccHg/doKZwdn504KcHt8tIcg4B7elAEZ74FOUgjp39aUhjkZ/SgB8d6AFXA6L+tAGAMDtRhs5x+dKN20HPagBBnpjtzS4HpQgbt6elADEkZHB9KAFIB4A/GkA5xj9acQf09KQA5JzQA0qB/CelAHouKcVY4ye1BQ4oAYBk4x3FLgMD8vf1pwB3cHHNIwbkg96ADjj6etIp549KcA2Qc0AMTxQA1g3UCkAPJxmnAMRzQqkKcn06igBu07jx2odflz704KdxzSlWI60AODE4GO9ICA3Hf3pATnHHWgN8w5HvzQNrQXPA4PT1pCcjBH60Bgccj86CR6Dr3NADWHPI/Wg45G3t608kZ/h6UjBcHBHSgQ3oOAfzoON3Tv60ZO3OB19aUgFuR39aB7icAcA9PWkB+bgH35px254I7Z+akOAwww/OgAJzyF7etKOvTtikyvHTp607jGeO3egQ3nB69PWhm54B6etOXHfH50m0HGSOg70AA6dD0HelB7Y/Wlx8pPHA9aOAOo/OgBpxwAKBkYwOh9aXK45x+dAI45FAAvGQBRkYGAe3egHDEkigsNoxjtxmgA3YIwD19abnvz+dLuU88Uh27ecZx0zQArHjp39aFPTHt3oYjvjr60sZ6EY6/3vpQFhwAz0P5012xxjt605MY5x+dMYg9x+dAxFIwOD19aQ9xg/nSrjg5H5/Sj5SDyPzpqwgPXgdvWheB07DvThtHBI6etIMAckdP71IBMcg4P50uevHT1NITggDHT1pdy4PI60AKh4xtH60p65ApFK+o6UbgTkkdu9AAcYBx+tMByDkfrTg4/yaRdpUnjp60DFPXA/nSggcY7etA65OKUADuOnrQIQd+O1B+9jFAwM9OnrSlgSQSPzoAQkYGAenrTT9/oevrTgQeuB+NIMZ6jr60AJgkdCPm9aAoGDt7U5sbcgj73rRkcZI/OgBi4/u/rTuABgfrSK3Pbp3NKNpxyv50AIOhBB/OkGd3T9aXI55H3h3pRtznI6+tAB1AGP1pcDGAD09aQcgcj6flTsjb26etADQACPl/WlPTp39aMqT1H50Eg9cfeoGKD2wfzoOOML29aAcdweOxpCVGMDt3NAgIyOn8PrSYABwO9O3D2+760isdvI70ABA3fd7+tHGOnQetKSM846+tBYdsdPWgYgxnp+tGRwNvajcucDHT1pQc4Bx09aAFAwMAdhSgDHA7etBPy8Y6Cjco7j35oENbnGF7UAcdO3rQSMAEj86XcDnp0HegYi49P1obpwopcjqCPzpWx0wPzoERdAPlHWmjIPAFS9h06+tMHDk8UDQg6ABf1pw6Dj0o4wOnToTQduOMcD1oEGSB0P50rHsB39abuGDkDr604kHk460AC9sD9aACRkZ6etKpUY/wAaRSCvI/WgYED0PX1pAAMcfjmnMQcnjr60g24GMdf730oENHXofzp4OOMfrTQABnj86euDjp09aAGjp07etNYHI4P50/I5xjp/eprEFhjFACeuR+tAxtxt/WnMRzyP1pAV29vzoARiAeF7etAPA47etHXkkfnSqQVGcdO5oARcgHjt6mlPJPHelXGD06ev0peMn69M0AJnHGP4fWkBwfu9vWl75Pp60gIx1H/fVADs4x9KMgADH60mQcdPuj+KgMGHOPzoAMjd07+tITnsevrRuG7PGM9AaGPBxjg96AFVsEdenrQpB520gccAkf8AfVCsuM5H/fVAC8YHHf1pVxg/QU1W5GcD8acGXk8du9A7MAPn+739aDyB8vQUoYbuoprMOOnSgQ4p296Yw+bGOlStjHB7+tRv1yPWgLjWP+c0oPAxSMM8Y6elKBxyDQAEjPT8aCSWOPSg8HgdqDgknA/OgBD0xjvSN97pSnBXJFBwWyR3oGnYM9eO3rQvXpTgR1FC8nOO1AMafT2pwGMcdqCMnp2pyjJ6UCEA45HbvSH6U/HGdv601uvTFABkEEYpM4oyOeO3rSZ9jQA0twAKUMeAOxpGK4zQpzj5fegY7cM55prE7RxQpyen40mcgcUCFBH6GjPPFIOe3ajv0oAcx4/GlU4H0NIeeQO9C5xigY7OTwKY55H0pc5PApCCcZHagQJ0H160oI3H8aRex29xQvXpQA4nHbt60AkKcjjFBxnpSLjByOcUAIx5/Cmh8Z4709hzwp6etMwBkn19aBrYVWJ7frSk9OO1Io46GlPXpQAoIwPpSLz0FAIx+HrQmM8D64oEOzkjA9O9BPfHajPHTtSEkjp7UDBSORjtTiRuxTAepwelOOc0AG4/pSK/zfjSse2KbnDdO/WgBd+QfqaXdggAdqaTgZ9/WlyARk0CEz3xRnC/j1o+XPSgnI4H60DFUgKcj0pA2W6HrQM4OBQD82cUCHZHAFIDgY9qM5AwKQnjIWgALc9KUn5enemng4xSn7pGOhoAUEbeO1ISDjAFAJweKTK5ztoAduJ4A6LQG4+hoPJ4HakHzJnAoGKzAk8frRnOeO1BwT070nuB2oAFY7gMUu4ccdqQHB6frQc8ZHagRJnj8KN2c8Umc9vSg4HQUAIzDg5zxSgjBx6UxvYdqcp46dqAFHA6UFieMdKQH2pCQen40DHA8AYPWkxySaByBx39aFJPUfrQAbsqOKTOR0o4wMD60m4YBHtQAMRk8UrH3ppPehznJ296AuPVsgUq4x9B1pqn8eKcnPbtQIRjgEY7mkUjjg8H1pz8jp39aau3II9aAFU/MeKcW6cdqYDzwO9KSc8LQAqnggDtSE5IJ60DkEAUncDaaAAt147+tIDkdO1KcEnoPrSDp0oGKT81LkYA9qQnngdBSZyvTtQIcjDn6U7qScVGhHXHanD0NACk8/hTQR/9fFOPLY9qQEdcUANftx2oLdOO1KwJ6DtSEYHOeBQAKTu6d/WhiCDgdzQOuMUNjaTjv60ANLcj2pFY55HNOIz1HY0gXnp0oAcOFwO1Lu68d6RV4HFJ1B49KB3FL/NxTXbnI/CjqTwaR1HJA/zmgCxuQ8D+VMYLupAxB69DQG+YHNAhSEIHSjauOO3TFNLZI5p4JI4NACMo7DtSHAJ4PSnGkI5wM8D1oAjIXacUvBYHHenYIB5P50h4bv8AnQA4BehH6Uqhc9O3pTcgdT096VTk/wD16AHYTjjtSgJgHHbsKbk+hpd3v2oAPkx909KYxXAwKdnjOT+dMbHAH60DDcpH5Uihccc0c+valH3etAIML1C03CnGFNOJ96FPTFADQF7D9KCowBingA/Nk0FeB+FAhmAOinpQQvBx09qcqkgE56UjIMdP1oGgO0jG3v6UALgUrd+vWhfTn86AABPT9KDs6Be3XFKOOmfzpT94Y9KAGhVAHFG1MnC/pT17c9/WkKryf60AJtAHC9vSgAbSCvb0p3HX+tNB4JHp60CBgmRkfpSbQc4B60HPGc0uc556UD6DAoxwDSlVPOP0pQRjr+tHQ/jQAAKVH+FCgdNv6UoAwMc8dqRaAFOzOQvp2oYKRwvb0p2M+v50hHv29aBjAo5IHb0pcLuzt/SjAGfxpSMnODQITCnt+lIAA2NtPwD/APrppwGzk9exoFcb8oHTv6UoxkEfpS43Dv1oxyMfzFACDGeAfyo2ggDH6Ui9elOGCMEfrQO4ihQMAdx2pAFGfl/SngDHXuO9Jxnj1oEIQvBx+lBVcZx29KXtilwcdf1oAZtGc4/GnELjp39KXHPFBHH/AAKgBuFx0/SmnGeF7U8gEcZppByPpQAh256dh2NHBGcfpSkHPIP3R3o+6vcfjQMUkE5waPlxjHWmk8596XPcn8aAFUKDwP0o2jIwKVWBbrQQMjigQYXHSnBV7Dt6Ug9M+nFOTjv+tADWVT0XtSYTHK9venEAjPtSEAdu1ADRjsKOMDAzx1pRzSkfpQAg27Rkd+OaMpnrQCVAye9APOc0AJxgDafypOBjilPIA9qAOmBQA0hAMgUrhTnA70pFKRycetAwQJxkdqeoTqB27CheF/ClAAGeenrQGo1tvp+QpoI469aVuecn86QEZHPegAAUnAH6UEA/dB6U0EdRThjp/WgQoUYICHpQQuRle3pQmCCc9BS+nUcd6AGnaARtPXim8Y+6enWnnkHk/nUZ46fzoGBKn+H9KBtwCB/D6UmOenelUcDGenrQAqbMZx2p4Ck5Azz6Uxe+PSpBnJ6/nQAFVPJB6elNAGScGnnHr2pmRzQIUhDzRtT+EfpSFuePQdqVScfhQA3ADdD1pcLjp39PelwS349qDnBznr3oAQhARgdqaAueBTx1HPb1pvfIoAVQoxx27igBMEYHbtSA/wAqVWwDzQMQomeFpSqnp6ULyevalYEjOaBCHpkn6Ugb5s5PX1pxQnv3poQk4oGOycYzSpggcUgXgDNKBgZ9qBDsA8cUrAEke1NweCeuKXGST7UAJgbev6UkmAetLjg4pHHzYHrQA0jrzzihOTjFAU5JBpVU8GgBTgDAHajjvzgU1gRjPpTgCePYUAJ6j+lJx3HYUoTPc8HrSFSeM9hQAu0YPFKBSAEA4J6U5RntQA0jihR0yf0pSoxSCM8HJ+tAxygZ60pA2jBFNAYmlAwBz29aAHKo4prL2IpwBwGpDyfpQA11x6Ui9MH1FOkXqeetNC9PqKBBgDil4HA44pFU9M0bSMAHtQA9DwKQjJ59aFHTnpQc5I96QxW7j+lNXG05PalZSSQKQLkHr0piEYAEYowOTnv9aCpBAzQU4J9/SgYIffvSkZNIqntRggj6CgQ4ABfwpBzz1pR079KRKBikDp9KDx+VLsJ/SgqAOvagBoBORjoKGX5vxpdpwaCcE8UALgGkxk9R19KUKSaQDLDk0CEKrgk+tJgEjnH1pSPQnrQEHAzQA1QM9B09KcRwBxQq0bSAMGgBQOPxFIBk/j2FKAQM59KTHPPrQApA4x/KjgL07UmDwOeho7cHtQAYGQAKUgY6Dr1pu05z704g45PegA2jGKawGf8A69KBkfSkK8jmgAKrnGP4fWmYABGehp5U+p+7TSmR1oAaw+bg9+lOAA7DpTWXn8acBgEg9qAFTk0HAxgfpSqpz1pCOQKAHAYGcdvSnAdsd+uKYAAM89KUD0NAAfYdqXAPYdKYwP6U9Rgck8CgAUc4xQcDHFAGTxSMDjJ7GgAGCByevY0ij5jj+VJgkDLHrQBk96AFAAAx6UoXgEjsKQDCj6UoXOD9KAGlQc5FOYAelIV9Pehgc9T1oGPQ9OaAOOAPypEHTFAUkY56UAJIMk4HemqBxnBpzqT69aaAcgg0AJjHb9KU+vT6imkc/jTtozkflQIVRkHI7elB7c9qRUzkY7UhU5FADyQQcHvTACR9KXb1Oe9GCRxQA1lBP4CgD5QPanlD69qTbwMelACKOCD6U/jPAHbtTVHX6UuD6mgBWxn8PSmDOcc0pBBOD9aRV570AKw6fSnLg9u1IV6DnpRjjr2oAUYzgnoaPk2nGOvamgEnr3p23rg9+aAEON33qaOTT9nIwfXtTAp9O1AxcD+VCjOT9KAvy/lSqh7+tACjlqVlAoAwf/rUrDk4Hb0oAaSv6+lNIGeG7+lAXOeR+dAUbhz096AFBHHOKCQO460YwQN1DJnHzdaBCEj+92o3DP3v4R1pHU5+929aTZ1w/Yd6AHFhtOGH5UHG4HPekAYoTv8A1pSh3dQOaAFBUdx+dG5AeGHfvSFCD97r7Uu05xu/SgBG25FGQe9BU5BHpShOCaAAFcdfxzSORjOfSl2HjmkZOev+eKAAMMdacpGBTSuM4btTguO9ACnGOSeo70cccnr3NLs+XJNIy479/wClA0N3DP3h+dICAAd3ajadxy3f1o2EqCGHT1pajH5XGPakZh/eoCHg57c0FTtzntTFcHZTk5+lIpUnGc/hSMuM/NSqgz97vigQDHr29KU7eMEUbCBjdSMuT979aABSABml+Uk4IPPrTUBwORjOKco9T3oADjrxTRjBOe1PVcgHcKRYyynntQMbxwM0AKMknvTjEQep6d6Qx7QTmgQi7dvbrRlT/EKNpwee9CqTn5u9ABkED5qIyCMGlWIEDLH8aFjyOD2oAeAP09KCRjr2xwKNpB4NJtP97tQAmRyOhpCV3Yz+tG0knnt0oMZyTn0oAAwz1/WgMA2Ax6+tJs5+99c0Kvz/AHh19aAFQjb97v3NKACeCPzppUheD3/rT1RjigAUJ/e7dqMjoM/lQqn1/hoKjdgtQAoIwc56+lNOMn/GlK4z83amsuCSGoARiAMZFAYdj260hU5Hz4/H6UBTj7360AKGGRz36U7Ixwe9MCkHOe/rTsHHJ/i7UDsG4f3v0pDtyBu7UbDnqfxpGQ5HzCgQo254YdBxSHaVIDCjHo3b1oXOw/N+tACELuwT3FLwM80bfm+93FDRns36UAKCAeD29KCQcZJoCZYfNShDxyenFACcA5B/M0AgDg9vWlaMgE56dKTysHg/jQAhKgYNOVlA+929aYycZz2pVQ+vY96AJAR2NNbGM5/WjacYB/OhkOOTQA3IGPm70AqTjdSiP0buaRY/n+9QAAgKOeKcCuBz29aQRHjB7UpjIAx6CgBGZcHBpHYY6jrQynGc9qRkxn5qAHAqAOR09KVSvqOnpTQpHJb9Kcq/Lnd2oADjp/SkXHGMflSsnPJzSKnAwaAE4PX+VHBPBoC5HBFKiE9x19aAEUjBIb+Gj5em4UoiJGQf4aRkII+b9aAABeckflSqAV60mw85Ydu9PVCB17+tACcZGCKByBz2pAv+13oC5xz2/rQAgIAwD29KDjOS34UKnHU9v6Uoj5OSetACAr60JtzwR270FD69qEQ5xnsKAHMB2PakbHrS7AOc9qNmSfm/WgBg6/e/ipQRjBbvR5Z3Y3D73rSlDjOe/r70ALkZGP5U3C+uMD0p4j5GTTTH3B7CgYgIHGe9OVgARn9DTCp3Yz3oC4z83YdqBEgIyfmFD4wTTQMMR/ntSshK9aAE25pNp3Z/KnFu2KQk7+f50BcCvIOKQrjtS784GKXcegFAEbLk4x2pQOSPpTjx2FBPJJHagY3G1eDSkc5x3FBbI6Uu7LDjvQHQQjPejbk5oLHPTr70BsHAHagQhXGPYUvt7UE5IO2lLY5/2aAGgHpjoKNpIpcnPT9KCe2KAF2g5JzxTgOOD2o3YHPoO9BbAzj9aABiQuM00+3rSuxx92kzkg4/ioGthAuDkUH7vI7Uoc9P60hY4/CgGKOfypSnGDSbiO1LvOMkUCEZcA+9AGT170rtnqOhoDYHFAxMdcZx9aRgc8Z6dadnnIFIxPQigBig56n71OVcdAetIp6YX9acDk5x3oGOUEDmgDIP+7QpyOlKrfKQP7tACFO4/nTWHWpN/OMfrTT0JxQIjALfnTgpHf8AWkyOSR39aUPmgBVG0D6UgyfypfMyBxTVYEdO1ACn+uaMY6fWgtk5oLYOCP1oAEHXI7UrLg9e/rTQwBJx2pfM3EnHb1oATBzQv3vxo3bu1AbDfiKAsKFLLSgHIzmmhiMfX1pytyDj9aBAp/lSDNKpzzjtQTzgjpQOwgGATmmFjk4NP3ZBBHpzTe/40B1EweCCfwpQDjGe1BY4FGTjH9aAEHBGPSnAd/8AapB1+nFODHABHf1oATbxx/OmuDkEdqeDxjH60jZ44+lAaDTz0/uik2lVODxTm6/gKQHIz/WgAAOencU7aTzSDJYD3p6t2x1oEIF+bkdPWgjHbFKGG4f40biSCRQMQr1xRtpd3oKUkA4xQAwrgfhSEd/9mnk5HSkJ6/SgEIozyBQTxzRk9gPegnIz70AAzxj+960KpzkGjcSBx39aEfPGP/HqAHY+UfSnEcY9qaW+UfSnFxycDigRGyDHTtSMuART2bvimu3J470ANx79vWlQfyo3cDilU5BOKBgcnkCkVfX0pTknpSqcY44+tADcdvSlUEHgd6N2eSKVTk9KAQoGVP0pGTvTkf5SMfw0jMAen60DsMA6/hThkjr3NIWxnj070A+3egVgxg8nvSBcAcdvSlyT+FAbIA9vWgQig4xil24PTv60BuDx2pd3zHjpQAhGDx6UijBxjtS7s9B2pFOSeO1ADgD1xjgdaNueRRu9uw70oY88dKAGgYb8RS7d3OO/X8aUMQ3A7ilyQM4zzQAmDkCmjJPHpT8kkHFNBz27UDGlPf8AWkCDnj0p5PPK/rSbxyMUCEVeTzSsDtOPSkDDJ47+tK7EqeO39aAH7Vz3ppHzU8keo49qbuGc5oGIFAI60oUHB5pCR/e7+lKD2B/SgQpAwDz0prL19wO1DHB4J6dhRxk9elA0NC4XqfyowM5yevpTuMd+MdqQ4Jzz1oBsCAe54pFX5hjPSlx6Ht6UDAOc0CFAAxwelO2KeOegpmRjr2p6sCOvYdqADYvv0pGUDHXtTsgDOf0prEY4NACEA9z0oAz3IxQCAp57dCaAQO9AAQMdTSbQMdfvZpW9c9/SkDdPmPX0oGgVRknJNG3gdenpQrL1yfypeOMEnj1oELtX0P6UmBjAz0pwwMH29KQlcZzQAMo9+tCoO2aGOedx60qEYHPNBWgBAfXn2pCoyOT0pxIHc/lTJGHTPQelAtwVVwOTS7QCevWmo+MHd36YpysD+PtQAqgAYGaABg9fu+lAIbpmhSNpw2OO4oEKRznJ/KkOMHnmlZgMZbt6U3OAeT19KB3DauCOevtTQBnOT1pVORk569xRyTxn8qAAIABjJ4oVVx36etLwQBnt0xQuMEZoANg7k/lSFQSfp6UvHB3dR/hSEgd+3pQAwAZPPb0oxhsc8+1AwfXp1obBbg0AKoA556c0BRvzk9R2oGDx6UoADZDd/SgAwMdT+VKvbr09KbkY6k8+lKMEggnp6UAKBg9+npS7QecnP0pFx2/lS8dPT2oAFRfX07UmwbvxpQRz83pQCAetAhpjHHWl2Ag9adxgcik7cH9KBjQgB5J96ULxgZ680cZBJ/SlyCv3uh9KAuARSM80FFHJz0pRgZ5prEAgg9uwoEIwAPGfuimhQowc4zSlhn7/AGFG4bc5PX0oGLhQ2RnrmlwBjGaTIJxk/lS5Azz29KAAKCRnPSjA4GD0pARnOccelIe2OaAHhRj8OacFyc8+9N34OM9O9KrA4x+PFAwKj36UpQEdD0pCy9jQHXB+btxQAmxc96Nqgd+1KDnvSMBgEE0CG4UYwe/NIqjOcmnBhgfP36UgI9f0oEGPlAyenpR949T+VHGByTSZAAPr7UDQvGMc0Mo5+vYUhPGR/KlZgQee9AAFHHWljQYwc0gIwOf0pyFRjmgBSgIyc0gVQBilJAHWhTjGW/SgBu1c45o2gdzQcZ/+tSHkd/yoAFGAeeg9KUgEg5P5U3dgHL9qN/vn8KBilUOSSe3agAEdxSBsk/MetKuMZyfxoELtUetG0ADk9KAc9KOMDntQIQKCO/Tv+FKEBYk8c+lIrccGlyDnnv6UAIFA9elCjk9enpSnnHJ6elIMAkg9KAHbRwMHoO1AUY7n8KCVzw3alLZ7/pQAgA3Z5pdo7E/lTASG+93pwYf3j19KAHBRgHnpTAAOBn8qcCOME8j0puA3r+VAxNoPRu3pRtHJ5pcDGM9BQpAzz6UCE2gHnPWhl4JGenGaXHPH8qGGT0/SgYrHt700/ezmnscgc96b34PfigQnfI9ewoz2yelKSMYzS8YHNADGPGM/jT/X6UnTmncAnB7etADdo2n/AApDgtx6ilJBXrmg7d3XvQAAZPWgZJ+9+PNLwCTuHT1oVhkD2oAa2O2Pyoz1+btSkA4wO1Lgen8NADQfQmkyT09PWnEA8j0ppC4GfTvQAoxjg9qMgdCfpSjGMZ7CgDuf50AIemM/mKTnjn8vrTscUqqrbRQAwYyckn8ad0XqPzpVAyeRxRwABTAM8dcfWkYnGKcMcc0wt3z+tIaFJ6jPfpinA8fe7U0nJOD3pykFcZ7UAwbOcAnmmP65qU43cHv61G+DjB7UCGL7E8GlU4HelUjPXvQAM596AHLkcCgNxj29aXGeh+lJxgjI6UABbp8360jfdJz+dBZRjDCkLAZwRyfSgBF6Yp2STwD1pFweaXAznNAxQM4/wpFPHXp7U7I2jDDp600EFelAhSc9+3pTScY57dqdnPB9qRscc9vWgBg9MdutB5PUUEKB17UpIJ659OaAFXnkenpQMZ696Bg9+gpVwDk460AJjjk9+uKMcg+3pSnAXggc0fLuBz2oAReDx6Up6ckfnRgZ4I6UbhtA3dvWgA47H0oU5bI/lQGH6igH5uDQAvO3736UZ4OWPWlIBwBSHGOOtAxvGfw9KUk/kaTgmlYjkZ70CE3Enr+hpHJ7HtThjoT2pvQ9e1ADScHGeMDjH0pFPGM9/SnFFJJHp6UKAQCSDyKADGT+NKcnp/KlIHbHWjIAOT29aAAfe69qQ9vpShgDRx1FAA2euaA3+1Tjgj8qCAv60FDC3bNAb19KRsHvSgdfpQGgqsSMZ/SlY5WhSAaDjHWgQmTgDNAJBPIP0pRsIBHYjmkUqc80CEPQYpMGnhRgYNJjA/CgBpAxjNKcDJz3oJBJ4pzYNAxF5Ax/KnKMjj+VICOOe/rTkKkdegoARupwfwoQ8Dn8aVj1578Ui4GMetAhvA7fkKPw7UvB6dAaU4bmgCMdSM9u1GcEfMf84p6qCD06U1wAw5xx60AJkc80q/dwDTcgE/MOtOTG3OaAF7/j70D7ox2FKcZyDn0peqjBz8tADBjkUE4JAPelGOfpSkAk47GgBmccZ7elAPU04qD37U3AHegB24g9/wAqUH+XpSHk4HpSjGMA9qADHPPrQenUdfSlGM8nvQSAPvD73rQAgOcZz07CjvjBp2FODu7HvSYGcj0oATHtRxnOcdKU9PyoGOeaAAYyfenMowB15pARnqKczDnBoAbjA6fpSdHzgce1IQM9859aaQC3WgBxAGBgf5xSgA8gD8qaAOMHinADigBxz6Dp2FGOuAOnpSHB79B60jAEnn+H1oATOASMflSEtv6DqO1BVQpwf1FIQobGe/tQA8lieQOPahQCR06elJs3Z47etKEBbp2oACQD90cDpilIyM47Y6UhUdgenrRgH9O9Aw2+ij8qaR3AHTjil2gjOD1PekIHHB6UCHAccgdu1KF9APbim49B6d6cFz0BoACDjA/lSAkEDA6+lKygr0P503aARk/xUAKOOePypcnAwB09KaqgHOf0owABz2oAXngkDp6Uxh/sjP0p4UHGQenemMq7cAdqBoUnAOAOvcU5COAAPyqNlXk8dfWnoBxz3oAcc8YA/Kmt9B+VKFH+QKQrk854FAhEAGB8p5pcYzgDr6UKo4B9fWgAY4HegB43DsPypuCwPA+76UEDp/WkAGDyPu+tAIHXB+6PyppJ56dfSlZAfy9aaVXn5h1oGKhwOg6+lKDg9B09KRVGDz3o2jPy+1AD1J2jAHT0pF6Hp+VNAXA+nrQuMfhQA9tue3X0o47AHj0pBzz83X1oIU9M9KBAoPbH5Ubefy7Ug2jPPanADOMfrQAKMHIA/KgZzwB970oAUe3FIMZHzAc+tA0BGc8A8+lKBggYHT0pNoPqfm9aAq4HPb1oEwUnPQcD0oI6cfpSDGeo4FLtX24FAAo5JwDyO1HGT0+mKUKpB9jSYGT/AI0DHMMkcD8qTGM8DpRtBx9aNoPAoATHOBilPrgZz6U3bg8UuMdv4vWgQ4Dv/SkYHpgdPSgKCelIwXsR0oACvJ+UHgUgUqMgD8qVwM43jp60gVQvX+VACkkMOB19KGHUlR+VIQN3XvQVBycdvWgBBywGB1oOeMY6elKFBPQ8D1ppQccdqAHjp0/Slx2AH5UmwEdPTFO8sdAP1oAjcDrgflS44OQOB6UhQHBI7c804KOSR2oGCjsAPypTwAMDj2pAoPQGkYDAwfzNAhegBwOvpSKMHjH5U0Y4+YdfWhQueW6e9AD8kqMAcD0oIwv3R0Ham4GP+A+tLgEDGTx60AKQBkY/Sh8DPQfhTcA0pABOD39aAFGRjgdfShTxjA/KjGQOKaEGOh6UDJC3PQflSKM4wB19Kay98d6EHTjpQA7AzwB19KXoeMflTABjpSgAZAP60CFGQpwo+76U11y2do6elKApGeOnqKayrnOR0oAQqeen5UqZx+PpS7VOeaAox+PrQA5eDjaPyoAwBwOnpRsUdB+tAUYBwelADQOD06elOAG4/KOvpTQox07UqquTx1NADzwBgDp6UwDknA/KlK8Y54WkVQSTigBSBnGB0HalAA6AdPSkwAeB2oAXse3pQAck9O/pQQT2/iPakAG7qOvrSsFweR19aAHAsCAAKbnHAUfd9KMDI+YU1R6DtQApAHOB19KAevA6dcUhVc9PTvRgAHIHagADncRx+VKzHHAHJ9KYV+bpxSYBAwO1AD8c/wD1qQDkeuafsyePxoCnI4oAaA2QacF5oK9PalAyBkUAB7jHakJJJB9KU55I9KCBknH6UAM5AwDSj7wH0pWAA4oI54B7dqAFKg9h+VBHzA4pwHrnpSAfNkCgBpOSDjtQQehx27UMPQdqX6Z6UAM/AfXFNbPTA6U8gY6H86awJ7UAOAz79KcBxxTUB5p68DjNACMPlpMnI5PWnEHHSjbnGAaAEAwcj1pOSACe1OxSbcgYBoATGQOKRkAXp2609R7HpSMAcmgaGlcA+uaEBHT+VK456d6FB447igYY9h+VJtGecU4DnIB96QjPQHpQSCr07cijbk9BSp1BHt2oIGSQO9Juw0hvIxSKSBgelPK8cD9KaFyCT6UxrQCM4wO3pRzz1pwHGOaNo5GDRcLaDVzjGO/ejaMduvYULmhutAhoGMdOKQDjkdqeRwKRFB/KgAI54H5mgZycY6U7ZnjB7UpX2PSgBijqPanYpApOTinEDOcUAN7nH6UmSG4H8VPIz0pFUlssO9AhBkjOO+acW6cUhUDoPWjHPfFAxoOfypcEtn+lJ+H6U5QOMUAKucHkdqCDnIoHT8qASSePxoARhkjjmlycbcUh6AEGlIx90UAMx7dvSnbR1wPvUhBzgDtTsHGSD19KAEx3xyPemNn9KkKnbTHU5+7QAbiTye1CnC4B/SjaM8DtRjjjigA5JzjuO1L1Ppz2oxzwO/YUvrx9DigQm3nkdutBByM0oBBoxnB9qAEA46enanYxxx+VBGP0peMc80ANKjAwO1G3Oc+lDg4/+tSgE5B7CgY3GBnHI96RicfSngYpjA+n6UCEBOcZ70i56g07GQDjvzRtA5xQA3+HB9KceaTHHAPIpQM0AIV45XtTjnnOfzpdnXmlcE5xQNjQMDAJ6djSbR6U4AkdO1AHGB0xQA1hk9B+FA7f1oOccUL0HHFACgccAUZORQBzwDQRzwDQA0Ej06elKAcg4FKBlTkHOKTGOOnFAhQeOacBkdB19KaueRjrT1U44BoATbz2/Kk24A4HA7U8g9gelBT5R9KAIwD6c+9IBycAHn0p+PakCgk8d6ADbz2/KgDqDzxTipPQdqbjOePpQApHPUdPWkzg5pWHIOO1IRkcDNADc89B27UpYkfj1x70AZODQwx+ZoAQsQQB296avPOO1L/EOKVR0wP0oAAD0HpRjkmnBQR0P5UoXg9aAGYAbpzSMp9zTwmTj+lEicc0APCrxx+lJtXOOaAxxgkcUmSWAoAGCDAzRwe3SjLHA+vajnpQAhxn8PWggc/4U1856dqCT79B3oAABjIOPwpwADYH8qYpYr/UU4A5696BsfgAfh6Uh2k4HYUuSO/XtTdxz+B70CEbbgewo+Xv6CkIORwelGCOnpQADZ0FIcHBz2pV3e/TvSENigBflA69h3py4x+FN5A/ClUtxj8aAH8cEE9qQLnAJ7+lBB2/lQAf1NAAAB0B/KggAAUAPyNp+pprZAHFADhtxye2aaxGDjPSkywxjNN3HbmgaHNt9e9Km319KYxbGfenLkc89KAHLt9D+VIxUEY9O1A3ehpGzkdenpQIVNoxz3pQRzg01SRjnv60BzjPvQNEgC9BTVHB/wB3uKAWYdO3WkGe3pUjFIUdP5U1scgHofSlJbsDTW3YJyaBq4bgR96jIyf8KaMgZ7ZoyapEskTbgfSkUjH4dqap47dPWlUnGPagCQbd3Q/lS4QflTAzZ4B60u8ng56UAKoX3GBQQPemhiCeucUjM244BoAeACePSkCjd/8AWpAW9OlCklsA0ADAAfj6UMF4we3ekYlRketNJJA69KBADg8HHNAYDBz29KaMk9+meaXnPB/M0AKpHOCRyO1KGAPWkUMc8d6AG3HjtQMdlBj/AA+lL8vIA7+lNG7IAzSjOM0AAwDxmnfKOMn71NCk/l3pTnGB/eoAcAvbNIyrkcfpSAtikbd29KBAUA/LpigAbc5I/CkyQeM9PWhWYjr+VAxxAzgevpRgZIx2ppzu70pyOmenrQIXA3d/agkYBBPSmgnI7fjQd3H0oGKSvPPp2pflz1PT3ppZu2e1A3dMn60AOJHYHp0zS/KAQc5xTDuIA2npTgTgkDtQAoxjgGkIGB1o56enrSEEgcmgQoC4AJ/MUgCk8GgFiBz3PehSxY4NA7CBQANpzx0FOVRxkdqTDYFKOAMD0oAdgYPB4ocLg4NN3HFDFiD1oHYX5cf/AFqT5cZ9vSglh0pFLYyPSgVgYKDgUiqBjmlOT+dAyMYP4ZoBAAMHr+VKfqaQFvTvQAT2oABjac+lBC8fT0pAG7f3aXLAgc0CFAAyT6+lOQjb3/KmckE5NGSB34oAkyvYdh2oBXaCD2pmW/ChXyB9KAFULj8qUBcke9MDNjqaXJ3Hj9aAHZXqB2poI5OT9KTLe/Sk+bJ47UAOfaTxnoKOM4AP5UhJz07etGST0oHsAHzf1xQVAGB6ntTRwwAHejcyrnnrQIcqrkEelCYB4zSckgnjihc/+O0DsSZUAfSgbME59O1MDHA/rSgtz1oHZWHDAbGaHAIwvNNDNuIyc0MWK8E9KCRd4/OgNlu34CmLg8ZPX0p643cseM0AOBBwOPzoJHQ/zoXaMHJ/OgYz1NADGIyP/iqAefwHenEDOBnr60FQM9eg70ANBAU5/wDQqcSu4dOvc0KBtzk0EYcDcevrQAbh14/KkBGeo6fSlI9SfzpABkDce/egA444/wDHqX8R+dIQOOT09adjjqfzoAb+AoYADn+dOwPfpSEAjOTQAnGOg6etOA+n5Cggep6CnDHqenrQAHae4/KgY46fepTwuSTQmOOT19aADgA8Dv60xypAGB/31T1wc8n86a5XAx/OgCNhgcY/OmEcdvyqYheCSe3emFVxgk/nQNCYGSMd6cpAAHt60u1cH/GlVV9e3tTC+ogPr/P2pGwDxjpjqaeB2z6UhGT1P1pARggEe59KFPHbr6CnBRx16jvQABnk9fWgQoI7evvSBhjHsO9OUE8ZNCAbTk9qB3EPvj8TSHaQd2KccbhzjNJgEHk9aQ9ho24OAOtJgdvXsaeFBByT+dLtXPJP0zTE9SIYAH0/vUq47enrTwilcZ7etIqgA89vQUCDjOMjrnrQQM9unrS4G7OT+lKBg/ePSgBi9fXj1oIG7Htz81PUDBye1KVHJyaAGcDn+tAYbu3X1p+0ep6UhAL9e/rQNDM8Z460m4Dt2PenMBt6nr60ALxknpQFhi47enrTgBgD39aFUE5zzj0FPQAY5/WgQigbTx37mgqM9B+dOUAA4J6jvQcAkZPSgBhXp069zSgcEHH50pA45PX1oUcdTQAgC54xQeB2607AzgntQwGOp+9QMYT24pDtyDx+dKFyev8AKgqMjBP6UCEyM8Y+6KAQF5/WnMoA4J+760i/dOWP50AIduecdR3pCeT06etPwN3U9fWkYKOh7etADQcsDig44yP1pdoJ6/rRjpk0DDAPpSgdsA8Uu0Y5J7d6dgdv50CGEAY/xoyOfp60rYGBk9utIFGOcngUAAPHagkD07d6VVHqaUgY5J/OgYikFRj1+tKDSqAAOT19aRR8x+Y9PWgBDtIHT9aM8AjHbtS7cAYJ6+tLtHH0oEMOCO35UrY5Gf1pSvynk9KHAA6n86AG5Axg/wDj1CnA5x+VAHv+opUAIPJ/OgYh2+g/OhWAxn09acVBB5P50m0AAZP50AICDn60vH+WpAODk96ev17e1AhowBgH+H+9QevXt/epwGe56UMBkdenrQAwlTnOPxoBGPx9TTig5Iz270KoI+agY3gHgjr60mQADjtz81SbVzjP60wqNo5PT1oENDDBHHT1pQeTjH8qRRnPPanooyc560AJ9MdPWk+U5H9afgYxk9KQAc5PrQA3jOPb1pQQO/60pUep6e1AUep6e1ADflzkkdaXgjr/ABf5707aN+N3FI4GDg9+1AxO4GfyakHHr09akUDjntTQB+ntQA3p/wDrpFI5zjtTjg4+tIAOfmPQd6B9AyC3B/WnMRjIA6etNXbk8mlcgAc0EkYBJ60o3A8UBT2NAByMHp70AKGJwT2p25qTDYA5/OnfPjqc0AId2Rz3o5yeT0pTuBGDRkg9T0GeaAEOQuVJoOdwzntRlguQTRlgfvd6Bijd1NJhieKcAfXtSAEkfSgQ054+lKCcfh3oIYY5PSg7sZ9qBiAnp7UEk0AH1Pvmj5uM54oAcCxBz2FKNw6HtSDdj73QUoJx940AIaTJGOe/rSktgDNJluPmPX1oELkjPNNZjxQN+Op/GgM2AQf1oAATx9KQ7sd/rTgCMcnpSN1GTQMG3DINOTPBprBsHB705ARjn60AOAYjn0pr7s//AF6d82eM5+lId2cH86AGrk4FJg5pyg8HPakOc5z39KBBkjgCgbsHHpRuI53dqFL4J54FAC8n8qTnnJ70uDu6npSfMM89D60DBTx170p3H0oBYAc96Du7E0AAzjj0pAGII9qMNgfT1oTd15/GgQEtnkY/Gjn9PWghuDz0oww7mgAG4gkenc0pyWNINwyP5ilYtnOaADJpoJ3Yz0NL8w6E++aQbt3U9fWgBAWIowSeBSlCR360BWBzk0AAB6j0pwzj6Uihs9frS88fT0oAFJwST0xSEsTwaUbsZ3elIu7JOaAEOcjilXOMHNBHuaFyMHmgABJxilIbGM96QbtwGaXJ25yetAw+bGKQhsjApy5A6mkwc0CGlWyOD0pCGAyAetSFW9f4RTCpC9fSgBCxLf8A16MnP4UfOD1py565PSgBqrk0o7U7kHrTTkEc0AOB4Gfaj5ux7Ug3Dqe9OGe56UAMYH07UKrY/D1pzA460DPXPagBF6cZ6UjE+9OGT0J6UjZ/vUAAZuOe9IC2fwpPnGOT1HelXd6n86AFGdoHPSnAHAFNBYAHPbvTgW4z6dxQAHOOD2ofIBz6+tKd3r0pJCcZB70AMwcA0qg9aQE9z2pyhuvPFACnOcH1pBnjBpcNjqaFLcHP1oAaN1Kfb065oJOevegMw5z0oAVCcEj070cnikDNg4PajLccmgBSeDnsfWkGcHFA3889DSqWA/H0oGIcnp6007iB9KcSScEmg528ntQIYATke3rSjIY5Hf1oAJHX86UZznNABzjp2pBnnHHFOwc8GkUNyM0ABLZ/CjLDjHahs5HJ/Ok5HIJ6d6AFGSfxFHzMOfWgF9/BPWhQxHU9aBhubI60mTnr2pwDZBB/Wm4brk/nQIQ5xmmjdz9KcQ3HXn1pAGwef0oGgGdx5pHLdu/anfNuIJNBUsM5/CgYAY5zSjrndSAsRkE03c245J60C0JdoyMmjHqaYHbgFj+dO38/eP0oEDAg9e1Bzk/MegoZ26hjTTI2CM9utAC5O3hj1pSCW+8evpTFcgH5j+dLvYkcnt3oGPH1PT0pQPmHzUwE+tKrMCCSenrQAMvIyaUDPO49KTc397FLuYdWPegBAD3b60hUAdaduYdGPSmFn7EnpQFh4H+12oH+9TC7YwT6UKxAGD+tAWHHJAO40YJx8x6+lALkfeP50BmwBn9aAE2nPWjJwBuNALBuSevrTCz9Mnj3oESLyB8x6UFff9KaGfAOT09aUFgudx/OgaFYerd6Vc8YY00scZ3HqKFZs/ePHvQCsPxzkk/lSEAcbu1AJ/vHp601nbJG4/nQA4DphjSHr949aarMBwf1oLHdnJoAU57E0qAkdT2poZiDlj+dKrMBjP60CHEHHWjaRn5j1oJPGGx+NBdhnmgY3JHO/vRgn+I/lSFmAxuP50gY/wB4/nQFiQDAHJ6UKAB1podgME9B1zSqxAPzH86BDtoB4J/KkIGME0ZOfvHgetGW7N+tACBQM80u3k5J60nzDuaCSWJLH25oAUg5yCaRV+fgnrSgn+9096Od2dx/OgACfXr3pdvuelHzYGCfzoJYAHJ6UAIBzyT70Y4xmk3NuPzHv3oDnuT+dAC44OHPNJj5jz37UA4B+amkkMTk9fWgY7pzntzRxjIJphc8HcfzpwZsZLH86AAjplj0zSjgfePWm5PqenrQSQcZPXsaA0Hp6g/pS4JIwT044piswH3j19aUOxPLGgQ7B9T0FNIIGcntS5Pr29aau4ZAJ/OgYEfNndThgA4NM3NkHcfzpd5yfmPQ0DskOxhshv0prA5GWoDkEfMelIzN13enU0CHY/2j+dKucdT0phc4+8e3WhXI4z0HrQIc/YbqQE4PzdqQs3dv1pC7DuaHoNWHrnk7qCCf4jTFdgfvGnFiRgsaSbG0hcEAcn86aFIPBP1pdzYAJpuSHPPr3pkjhnA+Y04ADHzUzc2AMn86UswA+Y9PWgBzHjG88UjknJ396Z5jLn5j19aXexHLGgbsKoI53dqcijGcmmqzAA7j09aVWYdz+dADiM5IakAPBDUhd/U9u9JucDOf1oAD1OWoyefm7elNLkNyTSbycgsfzoEOBbBIJ6UuDnqaYrsM/MenrRkkj5j09aAJBuGfmPaheRnPek3MMjcfzo3OF+8evrQA78TSYO0Dd2pu9gc7jx70gkYjG49KAFXj+KlAOclu/amB2HBalV2BJ3H60wJO/U0gAycNnihSSCSx/OkDHOdx/Gn0AVgc53dqQjnOTQWJ79qMsD97pSe4AB83U9acASOWPX1pFY7slj+dBZyASx496QDwDxyenrTdvzdfxpA7YGD2oVmyTuNACY/2qFTP8VBLDHzHj3oVmGefegAEfPJ+uKHXK4B+tKCx79qHZucn/OaBn//Z"
    return (
        f'<img src="data:image/jpeg;base64,/9j/4QhoRXhpZgAASUkqAAgAAAADAA4BAgD4AwAAMgAAADsBAgAlAAAAKgQAAGmHBAABAAAATwQAAAAAAABTaWduYXR1cmU6IDRqckdNZUlIVUN3Nm5LR2plSFY4TnQremFTNCs2OGt4N29GV1ljUWZKRDNyMGFJbVg3UE1mbEZpdElQaUg5VGkvTUlQbW1zZWJiNkxZVzZEOSt6cy9hRVdESVg1aXJkeUdtTzZuWkhrL0owVFgzcVo4TmVMNnhRWlJMNWUwQlQ3c1Nyak9LeVo4MlNqQStvOG5qcFpZdWJxZDEzT0h3VHUyZUIxWVBhbGZYdHhrQlJBR3FCeVViWDUrR0pnbk9HSmpMZFFuSmdSdCt6dStiUzF4UGV5d1h3d0dNNlQySC9TblA3bWxOR0xZcmRQT3FaZEtSS0RtTFpJbTNTcWdzWGYyU1FldnZwQ1NKUURLUTF5aWQyUjNYUVpvQ283WWhsckgwM2JheWV6N0dwbkVMNDEyNG9HcXdDUnRYMldkdkdWQzFHZWxCZ3hmV002aGN0TzEzSmFxM045aG9xdnFEVk81T1Rhc01Dci9nVXVWMzgwQ2JSMmkvMms1TFkxcElhWkd2b1RmeGsrRThTa25BRmVLeDkrampFZk5qT2xRSk05bUpTZzI5WDhyczJmOG45Q0xPb1JJU290OGc0eTMvNXJBaHpOTmt0Z0t3VDFqN3VsUndpeSswOTQ5SWVsMXF6N2QzVTV6L1FjWmRUVlRtZjkzVHZyVnJtK29zdlM4dUk5S281QXRZWGE0Ykx4cEQ2VTlIbXBsdDVKN0lkQjRHNjZSeGk0OTlQY1dnTTdDbXh4VVIrWFpPbW9TajNsQkNnc3puL29WUWRjKy82SG5GQmRTcysvejFIdy9pTStlQm9wYWpVbXFmRDlrUUIxWmxrZVFBNzJuMWlsWmduZHVsbS9DTUs4OWl6QkZZM0daWExaaks1eVZ3OGVWZk92eEpmbStySDFybWdvdks1bTdsNUZSM1ppTGtGZWhwaGJydFcrdmZ4WGZCSTR0aEF3UCt6dGJYMUQrelYzNDVoVDJQcjFLMEJzT1Y3eDAyRzNxSEt5U3dZQnpyMUxxYURyQytRcnlnVU1pNCtHcTdUTk5vSVZjeEhsNnZzYVNDSzVLbDhycWg5dXdhZTl1d3VpdjNCS29mWHQrOUVNSXBHMVJNaFRIRi9DOHNQcFNBYkRwWWs4emt0M1ZGZk9SUVpaRVR0M01mU28yV0RBOUJjM3IyT21WQ094NVRReGtnN3pLOFBkdUVlTjM0WVNncWQrZ29pK0F6QXp4Yy9yR1NUcVMramZobU1QMlVKUkt0VVo4OUFoNU81ZTJCRlh1UWRBMUdxS09DeVVJajgvdUVVWkJBRFg2SllNTUF6cmZwdm1YMnNIRU1sTDlQU2tuRE09AGQyYWVmMWMyLTk3NjMtNDE2Yi05MTc2LTAwMjk1NmUyZDhkMgABAIaSBwD/AwAAYQQAAAAAAABBU0NJSQAAAFNpZ25hdHVyZTogNGpyR01lSUhVQ3c2bktHamVIVjhOdCt6YVM0KzY4a3g3b0ZXWWNRZkpEM3IwYUltWDdQTWZsRml0SVBpSDlUaS9NSVBtbXNlYmI2TFlXNkQ5K3pzL2FFV0RJWDVpcmR5R21PNm5aSGsvSjBUWDNxWjhOZUw2eFFaUkw1ZTBCVDdzU3JqT0t5WjgyU2pBK284bmpwWll1YnFkMTNPSHdUdTJlQjFZUGFsZlh0eGtCUkFHcUJ5VWJYNStHSmduT0dKakxkUW5KZ1J0K3p1K2JTMXhQZXl3WHd3R002VDJIL1NuUDdtbE5HTFlyZFBPcVpkS1JLRG1MWkltM1NxZ3NYZjJTUWV2dnBDU0pRREtRMXlpZDJSM1hRWm9DbzdZaGxySDAzYmF5ZXo3R3BuRUw0MTI0b0dxd0NSdFgyV2R2R1ZDMUdlbEJneGZXTTZoY3RPMTNKYXEzTjlob3F2cURWTzVPVGFzTUNyL2dVdVYzODBDYlIyaS8yazVMWTFwSWFaR3ZvVGZ4aytFOFNrbkFGZUt4OStqakVmTmpPbFFKTTltSlNnMjlYOHJzMmY4bjlDTE9vUklTb3Q4ZzR5My81ckFoek5Oa3RnS3dUMWo3dWxSd2l5KzA5NDlJZWwxcXo3ZDNVNXovUWNaZFRWVG1mOTNUdnJWcm0rb3N2Uzh1STlLbzVBdFlYYTRiTHhwRDZVOUhtcGx0NUo3SWRCNEc2NlJ4aTQ5OVBjV2dNN0NteHhVUitYWk9tb1NqM2xCQ2dzem4vb1ZRZGMrLzZIbkZCZFNzKy96MUh3L2lNK2VCb3BhalVtcWZEOWtRQjFabGtlUUE3Mm4xaWxaZ25kdWxtL0NNSzg5aXpCRlkzR1pYTFpqSzV5Vnc4ZVZmT3Z4SmZtK3JIMXJtZ292SzVtN2w1RlIzWmlMa0ZlaHBoYnJ0Vyt2ZnhYZkJJNHRoQXdQK3p0YlgxRCt6VjM0NWhUMlByMUswQnNPVjd4MDJHM3FIS3lTd1lCenIxTHFhRHJDK1FyeWdVTWk0K0dxN1ROTm9JVmN4SGw2dnNhU0NLNUtsOHJxaDl1d2FlOXV3dWl2M0JLb2ZYdCs5RU1JcEcxUk1oVEhGL0M4c1BwU0FiRHBZazh6a3QzVkZmT1JRWlpFVHQzTWZTbzJXREE5QmMzcjJPbVZDT3g1VFF4a2c3eks4UGR1RWVOMzRZU2dxZCtnb2krQXpBenhjL3JHU1RxUytqZmhtTVAyVUpSS3RVWjg5QWg1TzVlMkJGWHVRZEExR3FLT0N5VUlqOC91RVVaQkFEWDZKWU1NQXpyZnB2bVgyc0hFTWxMOVBTa25ETT3/4AAQSkZJRgABAQAAAQABAAD/2wBDAAIBAQEBAQIBAQECAgICAgQDAgICAgUEBAMEBgUGBgYFBgYGBwkIBgcJBwYGCAsICQoKCgoKBggLDAsKDAkKCgr/2wBDAQICAgICAgUDAwUKBwYHCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgoKCgr/wAARCASQAxADASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD91grDgL+OaTa27oeafg5xg0AHIwP0rQzE8s8duacFOc7TS4bgYNKAe4PWgAAYY+U80u1uSBSEEdqXkE/SgBoVtpyO1BV+mO3rSnkfjRhsjrQAhXknFIFOeh6U8Z54pBnPAoAaVI4xShTjG30obOQcGnDOOR2oAaFJwNvSmsrEY21IFbpj2prAkfd7+lAEZQ4JwfalVTuwF/Gn7WII29qArdcdqAECnup605VY44NJgjHFClsjg8UAORGzkjvSgHAO09aATjoaT5sDg0DuLg91NNIIHI5//VSjJPTtSMD6UCGMrDt9KFVsZ2mnODzx3pADx8vegAwc52nrxRtOOFOMc0g3HsaX5uymgAAJHC96Apz0NKpPpSgkn7tACgHrtNCgnPFGDjgH8qTJA4zQANngY6UENycHrQTk/wD1qTJ5xnrQAqhsZINLg91/SmqW6gd6cevQ/lQAmxuOO1NCn0NPJbjjt3pozzhaAAK2fu9qXB2gbT0oOc8ikyQOnSgAwSTwfxpcENwKRd2cgGnZPpQAYbsKbtYv070pJI4FNG4NgCgBxDKuMd6UBiBx2pMkAjHTNJuJ7dqBikMSRg8UzDDqvU8U7JJOBSjJHTofSkA0Kw6A9ad82TwfagBjnjoPSgZDHg0wAqcAY7elG1uu31pWzjgdqbliMgUCEwemDRg+nf0pCTuGB2pwJxjH8VAxFVj/AA/nS7SMDHalUMRjHP0pGDA8A0CHAfoKTDEHA5oyw7fwikJbGAKAAht2dp5pQCScDtSHdxlTQGOeR2oAUKcj5TSEN02npTlbLcU07uwoANrEfd7ClCnHK9BSnPXBoJJzx24oAiZDx8poC89OgpWycEClAbHA/hoAAM8YP5UuDt+6etIM+lHzDsetAChW4oAYvyO/FL82Bwev9KRS3p+NABhgBkGjB9COPSja2B+lGGwOKAAjtg9uabtYdjS/N2FKQfTvxQAiIeMigKSOh604ZwOO1GSR06UANZSOAp69aaqnHSnNuPQUJuwMDp/9amAYOfu04ZHVTSDOTx39KMtjgUgFAY5IpHRi2MdqAXGcA9DQWJ/KgBpVuTjpSKjbehpeSTRzjgH8qAFAOfu96cuQBlT0poJ64P5U4EkDA7UAKAT/AA0bSSePpSDPQilyewNACMpzwvFNwSxytOO7H3T0603nJNABgjjBpyg5Hymm5OQdp6UqluOKAHLkkYH50rqxzj1poY54HelLnGMHqaADa2AMdqQBjnI796CxJG0Um5j0BoADk9qQBjnCnrQM46GlXODweDQAm1ixOMUrq3OOmKUbieAelI2SCCPegAJPQc0Atnp0FLzjr3pOd1AC5PQg9aWm5OBgHNOBJA5NABkg/h60biD+AoIYnIz9aMN1GenegBNxAyBRuYkcUh3BeppDknvQA8P6igE7hj1pmWB4J6Uqs2aAFJfI+lKCe+OlN5OB7Uo3dMHpQA4ZPb9aQ5xjr+NKCxPOaQ7iOp4oAcOnT9aTk9BQCwH5UDOaAGvuHT1FIGJx9aU57ZpMHg88GgByuc9DRucgDFIM8jmkw2B1/OgB4JwKQk46elC7gBjPSkJIGcmgAfPYd/WkBI60rFv1pFLY4oAXnsKDnOaBuznnr60fNx16UAChhg/1pc88jvQueDk9qQ55NAAS3YUis3XGKG3E9/zpAWwdoP1oAUsR0Wjcxz16+lIc8ZyOO9BzyOaAHKWC8g/iaXcc/dNMUlelKSSeKAFLNgYU9KRWJ5xSc4GPSkBYdzQA4sc8DpQS3XFNy2eM/nS5PoelACoWOTinZbOSO1MXcM9elOJO7NACjPXFGDu6d6M459qQFi3OetAAS2DwetJufPQ9KXnHOfvetBDEjg/hQAKXJ6Hp6UuWA6Hr6Uirz93tRzjGKABSTn8O9Lk5/wDr0gLYxkjmly3XJoGDE5GD+tMLMB0NOJPHU01hkYoENy2QMHj1qRN2OneowDu6VIuQPx60ASKDjgduOaawPp29aeCe3tTWLHgCgBhJB59B3pMt6dx1pTuB79BTQTjg9+aAFYtnoaNx60hznIz+NKN2e9ACqTu6dqMk44/Ohd4ORnpzR82RigBSTjpSbmxwKQlgOlAyRnBoADuHb9aUZHbt60hJxSjIyPagAG7H/wBelw2Oh60iljxz+NByQKAFG/HAP+RSKT6fnSknGAO9NUn0NADvmAGBnijLYGfam/NgA+lGTwAe1AAxIHUfnSsST0700s3qelKxY9STzQAoJx0oG707etALbcf57ULu64NACMWx0pFYj/65pSGxjn86Rc5BGetCAUZ5JFL8xGAKRc9aU5oARSwBI9PWly3YGkBcA7aCW4PPSgAJfkbf0oG4DOKQ7snk/nQue2aAD5uODSgnbwP4fWkOc0oJwO3FACgt1xS85/H1pAT6UZO78aAAliOnamgnOf5mlJOfwpF3ZJBoACDkYz0pQWPVTwKDk9B2oxk8UAHIOMHrRubHTuaAp3DGaUhsY5zmgAy2Rx2pAWHb86BuzkZwBQA2eBQAEkdaAxwTjvSZbGRnr3pAzDPJHNAEikk8A0MCVPH50wH5u9OyQO/TNADycccdfpTM/NyBSse3p703Pzfy5pALu6cH86VenTv600AZHApcADPFMB24ZB4/OlJGMD0phwD94fnSnnkelACBsLwPyozmTIHemsfl6Hp6Uo27ufagB3U8Dt6+9GRu6frS4HrSfLnGR+dACEjgEHp3pc8cjsO9IQox9P71L7YP4mgA3D0/M0FhjgdvWkGfQ9PWg/T9aAHFhk/h3oyQOn600kYOB2/vUbjjjP50APJBGNp7UA9BjvUZPHI/SlUggDjrQA4H5icUHoCF7d6aPf8AlS4GB0HSgBdwOMD9aaWAHToPWlwCP/r0jHgn+tAAX4696FYYHI/P6UhwTnHf0zSpQABs9v1pS3PA7etATIwR+lIVxxg0AODcDjvSZGCcfrSDgDCngilzkdO/rQAbgT+P96mq2BjHOOxp20H/APXTCq8genrQAMy5H8qUtkHANJg8AZ9sCkwBkkDr3FAxQTtzinAgnn19aRQPb8KFz/k0CFUggfT1pONuMfrSgDA/xoABHPp60ANJ5x9O9KCPTt60uCTwD+YpNpA5B6e1AAv07Uu7JOB+tN4yeD931pcZJ4oAXdz+HrSbvn6d6Nuex/Gk8sbug6+lADywxnr83rTsjjioxkKBjv605SMj/GgBynvt7UhbPbv600deB29aXGO1AADnt6Um7BOf50gPUkd/WjdknBpIp7ClgSABn8acAcfdPSmHOen6UqcDkCmSLt5+72pSwC4P971ppI7Y6CkZhjHv60APEnGAP1oL5IG01HkdyKMD/Z9qBj2b0B6U1Tjj39aMAHoOg/z1ozx+VAhxOWBx3pC3oP1puBnOO9KfqD+NAD8gEAD9aRiMgD09aaCNwAxz70u5e/p/eoAViMf/AF6M5PT170hPGcH86CM0AIxB7dh3pdwxye1Nbtx+tA74Hb1oAerA9B+tDONvT8zTAf8AOc0ue5P5GgB24cYbvTQ3J4/KgBcdT+ftRgZx60AGcgfKe1GRgHHT3pNowBgfnShc4HHT1oATcAvI/WlLA85HUfxU0jjkfrmnEgnigBykcYHb1pM/LjB/OlHYYP5U3BxkqfyoAGPJ4/WhXXjA7+tIc/3cULgY55z60APXBGQO/rS7gf4c/jTR9O/rSg8j6etAApx+XrSHlhwPzoABz0OR60HIxwaAAnOfwxzQD8v+BpD3OB19KBgjkj8aAHZBz8vf1pRjA+U/dpMjHUdfWgFdoPt2NACKRg+w9aMjcSR39aM8H6etAznoetACkg9fT1poIycDt60vbkHp600Dn/69ADtwLZA7etKG54H603jPT+H1ox14FAEgI3Zx3obBHHPzdaaOG6jIIpSxI6E8+tABkHgjtTQ3fb29aGySDt/lTOOw7e1ADiwPIHc96bvBPTuO9I2M9B+dA7jAP40AOVuef505myOmeKYMZ7dKV8kde9AAQ2Tn1pAG3fQcU/5QDwaTK78YNAAFOQRnpS4PYE80ZBA4PSlGOBt/SgBr59D0pDuJOOOKc5HYfrSMRznPT1oAjAfBzn86UZ3fiKMgAjBpwxvyQetADjuNJg5HJ/ClyD0B6elGQW5BoAacjjJ+7S4IzgUrdBwelLlT2PIHegBu0/3e1IVIHTtT8qB0PSkcgDp6c4oAaykg+1IQetOYqc5B4x2pMj1NADWGV4FA3dv71KSMdD2pVxxz39aBgobJ5PX1oLEKCB2pylcknP50jEYGAenrQIOTjA/Sm4YjoaeMYHFNJGOhxxQNAQcEk96VVbt/OlYqRnB+93FKjA4O0/lQABW64pCjdh0FOUr2X9DQxHUKelAhoQjt3pCD+tO46kHgjtSErk8HqaAEwwGPSkAYjgHpTtwxnB6etC4wQAT8tADCjEggUbWAP1qRsbuFNJnIPyk8+tAxoDEdO9ABz+Apy4xyOho4Jzg9B60CEAb07UKGIpQRtHB4HcUikFeR+lAxdrHvSFT6dqdlfQ9aDt7A9KBEeDkj260uw5Jx1pRtyevSl4LZwfxoAQKegHagZ3fjTgRjgdvSgfeyQevYGgBrKxGR60mW4HPSnMQRxn73HFNyvA56dxQAik5IGaXa3YdvShSM5/nS5GBx2oATD54FJgkkUqlRkAdx2oByeh60DuIyNkYFADAcA05tpAGDRgAEbTQAw7ieaCHPXP3qXA7DtQ2COFOd1AIFDdjQd3QUoxtxg0rEDoO3rQIT5j/3yKQbgp+b9aczA9j90dKRWG3nNA7CNknr3FB3ZOPSnZBbJU/lSHaSeO1AhNrHmgKQByelODLuHB7+tBK8HB4HcUABVjzjn6UFTnGKcSpXJz1oJHbNADGBxjHpQUOMhccelDsCBwenpRuBHKnoOooAaFNKc4xSjb1ANKVBAGDQAi78cE0qqc5/rSrjaMqevf6UKeSeaADBKjAo2EgcdqAw2gAHp604suBkUDIypxihlbkZ/I04leRg0rYIPynr3oENCnjApNrY4WnEr1x+lAKkdO3pQA0qeuO/pSAEgYGRT2wc4B6+lIGXAyD/AJxQAmG9O/pSjdjFINuc4704sOw/WgBEBIP+7SMpJBx+lOBBBG09KRiuQcGgBjZwcinLuxwO9BOcgA9aFIxyD1oAPmJ6mkG7HJPSnEgnlTRwVHyn7vrQMaoYjqaUBsnI70qgbeQenelymScd6BCbT6dvSmqCDjFSHb0A7elNBGCSP0oAMN6du9KA3NJ8oPQ9B2pwYYAx260AJg7vxGaRgzDp3pygE5wetBwecHrQMZhuOn50wZ6AVLu6DB/KkAB/h/M0CI9hyfp6UBDzz+lPwO4PSgbOeDQA0KdxFDA4xj3p3BOf6UHHA5PFUgHmmljnNKT70053cCpAUHovX8acDxx6UzccDkfnTgeeKAAkjof1pGJzxSsOfwppGOT6d6AGsRilz83GfzpMEryf0pdp3cnuKAHKx6U4HjPtTfu8gj86A3POOmKAFYnp7UoOex6U1iOCcdKASfTigB27jgdvWmMTgYHb1pckdT27mkIJHXtQAZOKQE+nalIIGf60nPX+tACEkf8A66crH5Rk9fWkI+Xr2oVeRz36UAODZ7H86XsMcUg75xijHAoGxwyMCmtxx7UE4GfakZsjqOlAgYnOPf1pUJAHfn1pGBIJzQoB43DrQA9TgdPxpC3H4UD1BHT1oIwO3SgA3e3frmgkZOB39aQZOMkdaXk5ORQAmQOnFAb5SSe3rQQMY9BSAccHt60AOLD0pN2c5HQ0h7cikIPPPegBwY9AO9KW9FPSmhSQfrS/lwfWgYpPyggHpSLnnjtRjI6jpSL0/rQA8/T0pC3GQO1GMnr096QnnGR0waAAEjNKSC3HpTc5J5oJyc5B9KBDieenakUjdyO/rScnpjj0oAO7qOtAC8EHA703qelKQQMAjrTSORkjp60AKpPYnpS7hjp2powPy9aXk9O3egBwPBHvSLnNIuTnkdqUdeoz7UAO7DilHI6U0nGBmlB7Z/WgBp4/KlPTH+1QRkdegoJweT/F60AAzj8KQsc4/XNLn6UjAjoO1ACMfm6dqTd2B70bTnH+e1IQSOo60APzzwKXJ/So+/XuKeDg8YoAFPI4oJPH0pM4YcUHJxj0oAcWOMc0FuP/AK9NJJ4xRk9z2oAGJK0q8flTSDjAPanKD37e9ACpmlJ45ApAdo6frQwOKAFB4Ax3pC4x0/WjB9f4vWm5JJoAUNwOKcGIIqPB25x29ad0GcdBQApPelYgnIprHg8UZyCcigBew+lC/j+dGcjqOlJzjt7UABOVx/WkDc8D8zQc9MiheQOR+dAB0PA70pY9qQcg59aMgd+nvQAoYFTuHalLDHQ/nTR04/u0pB4wO3c0AJnOSR09aVSAOKaQck8fnTlGQcEdfWgBc5P3aXIwOO1Jg+negEleT2oAVSMdO1Lnvg0wE9mHajqeo/OgBzHk4HT3phOSfr60pPvTcAk80AOLZxj0pwb0BphB/T1pR7kdKAHKwzzzz60uQR07+tNUHPBHX1o/hyGGM+tACjk4wPzpoJzn26A0uefvdvWkHP5UDF3cf/Xpp4BzS8kduKTnkcdOtACA8nHpSkkik79ulGOBQIkDZBGCfxppYBun0p+QRjj603+L7w6+tACbgcALTg3t39M00HoCf1pwIx94fXNABuBPCDpQTznb2/woY4OcjpQW25zj8KAEBXbkp3pQV3DCjqO1ICSOo69xTwDuB3DrQAm7/Z/SkyNw+XvTmODjcOfekBAYcj86AGscYyvb1oDccr6c0rDOMY6etBI/vCgAVwOi/rSFsD7v60oOOpHT0prEYzkdKAFLD+7+NGR1A/WlyCDkjoO9OXHcj86AGEDH3BQpBIyo+9TyoK8MvakxjByOtAArYz8tDMu0YXrQHHZhSHkAZWgBpZQBhetN3cZx+tPxgckUxgAPvCgBxYYOF70qt/snr/hSNznOODSoR03D8/pQA5WwOn60ZGchO1APGMj86Un0x+dACIwwOO/rQSMHilDYwSejetN3DqeOaAF3jPC/jTAwx90dPSnDk9R+VIOmNwzgUAN3jIwtLkc/L3HalJ5wewpQpAPzDr60AC4xnZ3pQwzwP1oX7uSR+dKCCeKAAHgcHp60gYYzjtTlIwOn500DjqOlAAW5xtPWmlueh9etKx5PIHPrTSVOBkdKAEDcnjt60u4Fj8tNH4dPWng5JOenvQAoI67aQEFvuDr6U4fN0I6dTR0b7w60ANYjH3R970oDDj5e1OBwMlh19aBk4IP50AMGM/c7UcZztpQpzwR0pePUdKAG5HOV9KA3JAH607rkAg803ox+YfnQA8sAR8tAOBnbmkJH98fnSgrj7w6UANyDj5aCec4780DBbG4dfWndRzj73rQAA8dB19KRiOPkHTjilB9DQeOhHTsaAGMRn7g6UDAGdvp2px4Odw6DpSA5XggUAIcbuEHX0o3YzhR+ApSCWHTrQcc8j86AEyN3TtSEgAfL29acuAOo6dKHUcYIoAaWA4we3OaAw9P1pSe26nAAdCOnrQAzjA+XtT8552ntSMBjjFOzkHp09aAEBGOh/OlJUDO0daBjGNw/Ohug+YUANDDjIHWmhhu+6KXccDDDr60iOM9R+dABlccIKdkDGFHI9aQAkDkU7G3AyKAGEgrkr+tO3cHC0hIx94UrENnpQAqsMDg9KA3H3TSBgAOnT1pQwIzn9al6DEJHPy0KQMcd6GYEcMPzpVPA5H50JiEyMZ20Z4+7S5GDyPzoIyM5HTvVAIpABGwdKQkZHydqVckfeHShmAONw6elADSeo205W4+739aQ85wRSoCFzkdaAFDegNIGwAQp6etLkHnigEFR06UAIp46dPelB5PHf1pB0pRjJ5/M0AHHXb29KQEZb5adkEZ4/E0Ic55FABnn7p6etGeThTSkgnqOnrSE+4/OgBA2G6d6QuOy4+agk5+8OvrSEkDqPvdjQA4Pz07elNGD/B2pd3TBFCexHT1oAQlRn5f1pSRzgelGQTjIz7mlzgHn0oAbkbskfTihj1ASkJyxOR19aUsMYDd/WgYcenfPSkB+bNIOc8d6By2aBD+RjH5U4dOB0pg6DH6U7HbH60AIxycUhIyduOnpSsPQCjYM8jt60AAwq9O/pQWGen0o25TOP1pG5bp6UAOznoO3FA6/hSdf/wBdGAp6frQAHgYx2oHXp27CkbBxxR7EZx70AG3AJCjpSd84pce360hHqP1oAevPNOXrkDtUYPHT0704HjP9aAHEgADHSk7gAd6CcgYpAcgDbQADjPA/KkzkAACjueOh9aQjOOPSgAHQDHpSHjIwOvpS8EdO3rQyjGcH86AGtxkKO/pT04xx3GKa4znjoeKVQOOO9MBy56Y6+1DegH8NIOOw/Og4I4HakADt9aQ5BOBSgAEcd6aQDyAevrQAu4gnFCtxjjp3pp5HAoTjjHpQBJwcfKOOnFL/AAkAd/SkXjoo6UHAB4HX0oATPB4zQT83Tv2pv4UYGen60ASA8DjtTRz2/SkBGAcdvWkUjBGOlADmGPuikyQfoPSgjJxj9fpSEd8dqAEHH4Up+9nHekx1yBQRlu/50APB9AOlCkZGemaQfQ9PWlX72cdBQAucrwAeeKTcAQAPypGAwTjv60hIOOOg9aAFVhnA/Sl5I4HpTAefu0o9NvTrQA/tuPrSY+bp3pAeo9xQByeKAHDGBgdKAe2KZ6AD9acBnigA+ooI9Bzu7UgUY+72oYn0/ioAcD7du9NJBGfakHB/+vSk56DtQA7dyQB+tNHAIGKCcnp+tIMbc8H8aAHEDOMDrS46gDtTTjdjA6+lO47gUAAzn8KD2HtRkZB/rSHHBAoACOMAdvSj/gPf0oY8cj9aTHYCgAJxxgdPSnDp07elMI9v1pw6cjtQA4cdP5Uxm7d6XHp/OmtkgA0AJnAGAOvpSIcEkj9KUDgDHf1poXk4A/OgB4I2DA/h/KlB9ulNA4GAOlOxkgYoAQ8ZFI3GQPWhgMZIPT1pGGe3egBy444HX0pU+n5CkXAGeenrQB22/rQNCsOwHelQdPY+lBGR05pEUcYpBcUd+O/pQxwcAdqTg8gHr60ZzTECH27UMehxSAcHjtSEHd90dPWgBd3B+vFKvAxjv6U0DknA/KlXp0oGOPJ6fpSAcDp0pCcn7p9+aUcgey0CEHAOPTtThgH7vWkAAz7D1o4yTgfnQAp47dvSkBzkUH6U0c5GKAHkn07Uv+FM6nkdvWnKeDx+tAAMbu3X0pDznAB57CjGSfl7+lBHB47+tAACAR0HHYUKwHK0YHH09aQAk0AOyew7U3A5J9u1J+HQUoI7e1A0hCvPPr6UNkEgc80mMnFK4GDgUDFwD270AfMDj9aTcMYKnrRuw2cHGaCR2AAPlFOGOO9N3DIHPXFKGGcY6+9AA3X7vb1pSASfl5x60EqBkL29aQsMHrwPWgBflC8D9aQgFs47jvSbwV78e9DMN2MHr60AO4PUdu9A+8OBSbwM4B6etAbDfh3NACEDsvalKj+72HekOD2P50u4c5B6DvQAgAxwv45pCOnHb1pwZRzg9PWmuw9D09aADGRkjoBSr9KTd7dh3oBwM4PvQApIAA2ijPQbRQSPQ/nSAjjr19aAHAAnO2kIyMbew6UisM4GfzpdwwMA/nQAo5H3e3ekIB6DtS7gDn+tN3jG7FACsAM/KOtKoAx8tIxBznPWlVhwDnqO9ACgD+6OtBA4+XtSBwOcH86QuoPTt60AOUA449KaR3xQkijHB6+tBcHOQevNA0IRgYC0ijqSufxpSwIzg/nQHXBGD09aAFwAfu/rS8YI296YXG76+9ODLg8d/WgQgA5G39aNvP3aVSuOR39aMgnv19aADA2529qavPbPFPyCBwfu+tNGDng9KBhwD0pcAfw9qMgnof8AOKN6+h796AAAc/LQQC33aAR6H86UlWJ47etAAMA8L+tCjn7p/OlBB5wenrSAgMOD19aBCNggkr1PpSfxD5e3pSuy+nf1pMrkcHp60AIuM/dpQuRkrSAgHp+tKWGCMHr60DFGMEbaToen40BxzkenQ0Bl3dD780ABUZxtHFOUf7P6Um8ccHn3o3LjofzoAXv0HSkcAjO0daUMuRjPHvQW4x/tetADcAdqMD+729aXIAxzSHGQQD+dAgOCeg4A70ZABIFDFcd+g700MMev40AO3DPTv6044/u0zcAcc/nSlwAQB+tACjhun50hA4+WhSN3T9aGZeOP1oAUgZHyjt2oA/2RSFhjIHp3pQcdB+tACYBA+UU4Dr8vb1puRjp29acGGDkHoO9AAq98UEcYx0pVIxwp/Og7cdD+dADdoAHy9/U0gUA52/jTwy4wQevrTQw3Ec/nQAYG37vb1pcDH3fxoyAo4PT1o3AYOD+dADWXttFDKOu3vQzLtJI96GYHt39aAFAA/h7cUqqN33aQMAAcHp60qsuOhoHcXbxwtC4wBtoLKBxmkDAAdfzoEJwP4aMAdB29aQMCOlKGB6j9aAFVRtJ2jp3NIyj+4PypyuoB47etMdwMZB6etACHHPy9h3oXAGMfrSMwyTg9PWlVlxnB69zQA7C7sBe/alAAAAXt60gKk55/OjcNo69PWmwBRkEle1GPmJ2ikVhtPHb1pcrvPB6+tIAwByF7UgXk/LmnEj+6enrQpBzx+tACbcdF7Cjgfw9uaUlc/hSbl9D065oAO+cd/WgkNzt7mk3DPOeo70MygZwevrQA4Hpx2PegDn7vUetIrjI4P50qkEZwfzoGIVyM47d6No5JUdqUlemO2DzSbhzwfrmgYgAz0pXA2HC9qAw3ZOaHYYPB6UDECd80hXBzSg8/d/Sk5J6UEC7SCCCetKAeKQc4GaUemTQArZJ60Yweo7UjkjselISQSR6UALtO3qPyoZcNye9Jklenel6np3oAXafX9KUKSwPH1xRz6UdD3oAQjABGKNpPf9KGPI49+lAJPVemOaADaeec/hSMOBz2pR/SmtwM47DrQABTilC5HUUmcA54pwJxj9aAEI4zkUgBBHPc05hxkj0pMHI/3qABR/tA/hSEALwe1KMntSMCAPpQAYOMA4ox2zRjFHIGTmgBzKQDz+lABIzupGJP4GgHj+tABgjoe1NPBHzdqd7/ANKbjGMZoAFHTmgjqAe/pQucAkHqKOSe/Wga0Db1yf0oCgqef4aUAnqKXHy4x/DQJibSOjD86ACQcmlOc4pNpAPHegByjPQ96XBHfvTVzj8aUHPb86ADbhe/Smhc9D2p2cgY9D2pF4U/SgAKnIOf0oKnPB9e1HGenSg9celAAo55x0p20g8VGOM/7tODck89KAHAGkwC3UdaGJIxj603d8/40ABB7MOtAHIGabuAWlzkigAGPWlxjuKavJHHanDg4/pQAbSQST+lAHJ5oHQjntSEkE59aAHYIxz/ACpMHGM9qaWPHNKDmgBcE4+gp+3jOf4ulRAk96kB7YP3qAF2k96aV5A/rTvbmmnOc4PSgAKkdx0FJt4+8PypcnOcHoKQklen40AIy4bqOtLgnuKRsk9D2peQelAChSWGf5UhXBAB/SlHDDjHBpDng4oAACRjP6U4Lng4pueOR0pykUABXI69qAvHXoKXtxQOBj2oAMdqRjgYyKQn0HamscrzQA8Djg/xelIAeu4fgKaCff71KGJ7UIegvG3IbtSkZHUUnOB/u0vI4HpzQIQqSCMj8qVlx3pOoyQegzSt64oATBwORQq/7X6UEnGOfwFIDxzQA45JxmkUdMevrSEk8g0L2A/SgBMY496UA5pAM9AaUcjkUAC8g4P8NIVP94flSc4P0pSxGATQA1l6kkUqg4OD3pOeeKcqnnjv6UAOUMPSlKYUH2oAJ7HgdxRngEA9KAGqDzz29KApznPX2pBn3pRz1BoAcVPQ/wAqQA5I/lRkdf0pBz19KAFYEdPT0pCO4/Gjr+Ao56EUAIF3NjPcUu3j7w60gB3f8CpwBA6HrQAKhyASKUDsD2HSjaQOnY0KDngdqADBBAzSEHkg9qU89u3cU3OCRj9KAEIO480j5w3XpS4JamyA7c0DuSqxA/GkLndnHanbVHGKaQu7GD+dAgUnI4/Wl3EkfL3pdowMr+dGFx93pQA1pD2HagyMSeB92kbb6HpSgKWxjt3NAAjlV+7Tg53dO4pCqgcLRlQwOD1oAdvOen60u45Hy9/WkAB6A/jSqBn8KAELE4+XsaUvxjbQQowAP1pG2EdDQAm85+6KazkjIA4HrSkA8YNI2MdD09aAAucElaXfxgJSbVPOD69aVQCMYP50APMny9KTccZ2j7xoJUDGDSbhgYHf1oAQMwJO2gyHaDt7UfL1P9aBtIGVNADt3H3e1DNhD8vb1pcKRgA0jbT2oAY8hAOR3pVkOchRSSAA4A70Ltx09qB6Artjhe9BJBHy/rTgFJ+6fypSF4+U0CGqTgcd/WjJxkLTlCnGQaTC5Py+vegAD7QBt7UCT5en8NHygfdNA27Twfu9zQAu45Hy0Fzg8UYXrtpCVOcA0AHmEKRt7+tAfnO3v603KnPB5NKAM9O/pQAoY4Hy9vWhXO3oOlKAMDjtSADHAPT1oAVmyeEzz60hLddo6UpC5zjvR8uOR29aAGBiCTjsKPMJY8frSjBJ+XtSlV3dP1oAaHI7Um87+B3p5UAcD9KaFG7JB6+lADdxK5C0ocgjj9aAgI6HrTgoBGAaAGo+D0H3aXd82cfrShQD909PSjC9hQMBL/s+nekLZJwvp3oG3k4PGKUYYnigQ0k5GF70qsR/DTtqkYC0gCAdOgoAYGIIwvapA5wPl70z5QRx6Zpy7RjA70AODHOcUjOeBjtS4THA+lI230NACGVs4C/w00OQD8v60pA3dOwoIUDkDrQAb/mA29/WlLE/w0EKG6Hr60ALgcHp60AKWII4FNZzwNopw2k8jtTWCkjjv60ABY4OF7+tKGJ7Cl2qT909aVVHQg/nQApbodo4pS+BwO3rSNtx0PT1peDnjtQBGScfdppc4AIHUU8hT2prBQPuntQAiyMQMjv/AEpUZgxJWjaMDjue9KFUfw/qaAF35UZA6UrPkfdHSkGCudv60uFPY/nQAxnIGCo6UNIcEkUrBSPu0jBRnA70AODnjjv600MSuNlKACAcelOAB/hNAxpJ7pSo3A+UcCghcZwaF28ADt60CEyRzspA5HOBTiFJ4FBCj+H9aAGCQ4PH8NBY7hx+tLtXacKfu96NoHVf0oAQPjOR2FOWQ4J2/wCcUmBz8tClcHIPWgByuf7vek3nA47etJuA4waUBcDA7UACvkD5fSlBIJwv60i7SOQfzpwC89eKAELHptFNDkHOBxinkLjp29aaACT8v60AN3Z7dvWlDYPSnEc8DtQAvYfpQAisd3T+L1pckrwB1pBgtwD970pVwRyOlADg54+XtQGHp2FGVz0PT1pAVz07elAAW5zt/WjJ5wvYd6CVPagFeePSgaEz82SO/rTZD8pwvX/61PBBY5zSPtwflNACknHT60nOelLu7YpCQDQIXnGSKOcf/XpC4AoDD9aAEY5OR6elAJye9DY3ZwOlL3OB29aADJ29Pyozzn3o6Dp+ZobBPTv60AKCfTtQpIP4UAjJ+nrQDg0AKSe1JuPcdqGOcEUA57dvWgBvX+GkY9qeDx+HrTSOMjr2oABkjGO1KuRzj86UcDGO3rQOOlACHPpRvwAPenHGOlNIHHHf1oAbuOTQOg/xpQeSaAflHH60AOUn0+tJn1FOHbAxx60jdc4oGNcn07+tIh5HHcdKdJ83Qd6Rf6igByHHQfrQWPBx2oGQeP50MRjGO1AApPGB0NBOTkUKwHfvQTyTQITJ7LQGwOnalOD0H60gIAPHb1oAC2McUhbg8c59aVsZ4FGQQeB19aAGj6UvO7p3oH3c4+nNBOTnHegBwY4Ax2pAcDp2pQRjHHT1poIH5UAKzHuPTvQCT27UHrzjrR9Ac4oARTknjtmnHOelIp68fmaUnJ4oAPw7UDGeB3pd3t075pNwDZOOvrQAg6Ehe9OBAxkHp60m7I49aUOMgcfnQA3jPSk5PboacCCc4H50jEEY68UANBOCcfrSA/N+NOB7D2pOM5460ALk4GB/nijcQPwpc8DAo5HAB6UANBORxS84wB3o79KUnPTs1ACgnGMCkJ7gUKRt/GkJHYCgBSxPOO1IGOM8D8KUsM9O3rTS3HT9aAHFju79aQEgfhRnLdBS8dAO3rQAgJzn2ozyM0v3T+BpGHTAxQAob27UofGePXvSduD2pcj/ACaAAseP8aAx547UhPYA04HqPb1oAQfSkJz2704cc0hwO3T3oAAxGB70KQOooUjA+XofWlVhnP8AWgBOoBxShj6UE5AGO3rR24HQetADD9McUMT6d6VupwKGJboOlAApx0FCkjnj60L09cClHA/DqTQAhJx07Uit04796Vscj8qVeD0oAarYPTvSk9cCgDnp3pcZ7DpQAinrwelGTkHFKpwD9PWgHJHHp0oATqDx3pvOMYqU8ggLTMDHTt60AMJOenYUqsccjoPWlYZPA7euaDjGPagAQn9OlOBPIx39aavfinZGf6ZoACT6DpTck5yKeT6DtTe5+X6UABOaBnA47UjDJ49KUdMAUACk7s470jNx07+tPX73I79qaxBGMZ59aAEDdOD35zSBj6dqXgEYH60gGTnH60AG44HHekDnofal6jgUDuMjqO9AwD89KVmOGwBxSf560rYNADiPqaaeoIX9KCyjgY6+lNLLu6fpQIDjjilyegFNyvGB29KcNp4x+lAC8+lGDnG00ny56dqUbc4x296ADqM7T19KUjkcfpSZQD5aU43cZ/KgAw3UigDLcijK9AO3pSZXd0/SgAI7DPT1ox6g9PWkO3jjt6U7KfpQAAH+6aCMYwvpQduOKQle47elADscHj86XHotJlQM47DtQpU8j+VACEcZApOeDtpTjAx/Kk44Oc80AAGf4eaADwCKPkyc/wAqTjA5oAfjvj9KRunNJlcflSMY+h7CgAPPAXvSqM44NIxXr7+lKpUAf4UBqL36GkYHPC9qBsz/APWpGKcfT0oGCk4H1o4PY9aRSuB7H0o+Uk8d/SgY/qMgGkGccjtSZHUUAptPPagkUgn+E0YPPBpG2dv5UnynODyD6UAOAJGcUhBJ6HrQpHQDv6Up2A4x27CgYAfKMDtTRk9jS/JgEY6elIu096BC85zg0vsQelBCbe3Sg7CeAOnpQAi+4peM529qbheeO3pQSueg7dqAHE5+6KQE7s470mQTkfjxSAjI+vpQOw4E45FLkkg0z5cYxnn0pVK5H+FAh4z1xR24HamgqDx+gpQy8dOlABg84HegDnOD1oBjxxjtSDbnoOvpQA4LjHy9qMcfdNICuAfalLJjGRxQAm3kECnEcZK00MoII6UpZdvbr6UD1EIyOAaawOenSjKsM4/SkJXPSgBTkH7tB6cCg7M4A7UnyleKBC85xjvTuemD+NMJXPBpcjPbp6UAOwB27Ug69D0oUrnp29KDt+nFADgM9F9KXBx9003KY5x0FLuQg47e1ACMP9n0pQeOnQUjbe1AK+vagBwz2U/lSPnrik3KeMfpSFh2Hp2oAMk4+XvQpz0U03KcfX0pQU5PtQA8ElR8valIJHQ03KBRigMpA+npQAMB120NnH3cUjFMYPahyvagBew+XNKvrtpilOD/AEpybcf/AFqAHE8cDt6UgHTikLKeoH5UisoA9j6UAOGOfl70enFN3Lnt19KUsvUUAABIOF/Sl/A0wFdpz/KnArxx29KAHeox+dIOn3T0pAV5+vpSBlxnPan0AXGcDafypCMgEL29KGdAeGH0zRlSoAx930pACjjpThkk4Hf1pqlCP/rU75N3Qcd8UAKR7U31+XpSnZj8PSmgpknH5CgBzAZ4Wj8KR2XPTt6UnygcDtQA4cHG3vQcY+6evrTcruz70ZTkcdfSgBcHrt7cUm3k4WlynRfT0pAVzwO9AC7f9mhB147+lHyjqO3pQCmCRj8qBigEHGD0pHVselIXUMf6Ch2XHP8AKgLCkHjPr1pu3LZOfxp5789/WkAG6gQgQADFOAHpRxxijoBgUAGAOnal4yeT0pCopRgk8dqAAAAdT19KGA3UvBXBPf0pHwCf8KAD1HJ4oVfmzmkwB37Uq8NmgBCBxj0p2OfwpCB2Hb0p3UcDtQA0+2elNIAxj0p3Hv0ppAJ6UAOUDBPPFOUZ6ZpqgAEU5OmcUABCkUh2jAGeDSsBikwDjr96gaEKjJOTTSOAf0qTHJOfWmkAAYoENHtQSCTmlwOOKaFzigYpGM8nrSr0pWXPPv1pMD9RQIXjrn9aY2Mgin5GcjimsADgnuc0ANXBxx3o4z170qgDHJ6+lGOfx5oADgd6AeCQT0oOM8fnQFG0jPagAyCcg0ZyD83elKDgY9KTbjP1oAFxilbHp9KRcYOO1K3XigA6jr29aRcE/eozwOe1Ccdu1ADhgDg+lLikJ9D6dqXt+FACKM5A9OtDAbuDQNvI9qVsFvpQAm0Z6UgUBsg96djgUBRux70AN425Hr3FHAOfanFRt4PekKjI5oAaCCeDSg9MGm4GcYpR6Z6UDHKODj+VJxkmlXG0j3o/i49aAAjnigjjqaMHijHBAoAbxkdelL24z1owMj60uOBz3oATHFIQM4NOGMdM00rzwv6UCDjO329aAML1pcDOc/w8ijA25z360AIQN2M8ilA5P0pCBn8fSlA4xz0oAFXn8KCRxSrjt6UhB7enPFAC++e4o4zR1P5UAcEe1AxG6cE/hSqBjOD931pGAI4HalAxn6UCBQPekYDGKcoA4pHA29O1ADcYAx/eoUYOc/XilCrgf73rSY5/rQAuQQOfwoUdPpTcDA/wpyAADjtQAEDBznihsdBnrSsvBx60MAenrQO2gijOME05UHoenrQgx360qjAyp7UCGkYGRTBzjmnuAOB+OKaFGR9aAGqRnvSgAD8KQLg/jTsAmgBVHB5J4oxyOD07mlVVKnI7UhXBFAAMYOCaT6NSkZyaQAYPFAARk96UAAdO3rQQM4pcAKB7etADVAweO1OAwT9aABgj+lGOT9aAFwOp9KFAJpTgHn0pEwSfpQAjD054pCMYIz71IyjgD0ppUYA/KgBi4B69DS4BGcnr6UoT5sdMGl2ADA9aAEB569qQEdqdtyw9hTQOc0AKAPf86Q9DzjBpTjA5pvAB+tDGhMjPBpJMcnNL3pH4BxQBKPc9+9J8oPag54Hv6U0KN3X9KBCjGR0+lOH4cH0poHA5/SlwenJ/CgBTjPAH5UpwQRjt1xTSpB6Hpzil285x29KADnZz+opGIzzjrSdV/wDrUjfezkUAOJByRijHIxTeucN29KcqnPFA7juwG0dPSl464HQdBTCvTH8qcBnt254oATv/APWpDjHQdBTtpxyD+VMYcD/CgQ7AHYdBTxgAYApnQH6elKpx0PagBT93p+dJkAjGOvpSEZHB/wDHaQKTjOfyoGOXkkcGlIGAABTV6kHP5UpXgY/lQIMA4wBTcqF5AzSkYxgfpTSfT+VADsgLjjrSDqMY6+n0pGyRyD19KEBABGevpQAozjPH5UHrjg/hQFzxg/lQwJOSP0oAQAYHA/L6UEcEYHWlQDHU/lQw5OfX0oAQDB7flQoGMgDp6UmPY/lQgypwSOOuKAFZcsOBQQMHp19KU8NgHt6UhXg/X0oGIvTt+VKTzzj8qQA8nB/KggkknP5UAAAwPlH5ULjGQB+VKAcDn9KRV4PH6UCHA88kdR2pQR6Dp6UFeeB6dqaRx+HpQAox0wOnYUoPzYOPyqMEZ/D0p65zg59uKAHKB1wD+FKoBbAx17CgDtjoPSk24bqOvpQAMPlwPX0ppwQMgdPSlYYHB/i9KaBnBB/SgBuAPT8qXHORjr6Um05/+saULwOPTtQMcpGeg6+lKNpY/dpo4zj8eKVOT1/SgB2AMcDp6UADHQdM9KCcAAenpSEnHQ9PSgQ0deg/KlOAOg+96U0jJ9ce1LgkdD170FaCjHt9BSNtyAAKAM//AKqNo7H8qBdQwSM/7PXFKqkrzigqRzg/dpOoGO3tQIUgZHTrSkg9McU05z+PoaUHqM/pQAbQSOB0pduMcDpQoyeT29KCo6Z7elAC7cDPHal7dB0o7Z9hzikJ6nPbnigBCAOqjt1FAGcnHQUjD/OKCMg89qAFGOwpCeOg7dqByen6UmDgdvwoAUZ/UdhQFy2ePypQp2g7h19KRQCeT3oAMLgHA/KlAGBjHbtTSOBz+opyjp9KAA8ryB+VDYwf8KQgEc+npSSdTn19KAFXsOPypVPHamZyBz+lKpJGBnp6UAPbB5IB59KQAHA460HJPIPX0oUcDg8ewoAbgBTwKABnt+VKVAzijaScgfpQAJ3x6elBGSDx05oVTjv09KAMEA56elADgowTtA59KABjoPypCpGf/iaAOO/5UAKMDg4/KmrwBn0pduTkg9PSgIcDHp6UAIvQ8D8BSjqenX0pFXI5B6UpUBuh4PpQAEDrx0/woTGD8o6elBHPPp3WkQcn/CgCQYzwB0oAHHA/KmnqMDt6UoJOMA9PSgBB97t970oIGOPWgA7sc8H0o7d+vegBccA4/SmcdwOv92lPXI9Kbgseh60DEJ54x9MUgUHJwO3al2ggDml2YBwOOO1AXG4Xd2pXxjORxRtyev6UrJkcEcLQIdk9v503J3cCjd6Z4FID83XpQA7J9Oh9aXJxTN44+anFgf4qABj6AUhJzkL2oc+56U0gZPJ6etADtxAP+NIxbdmmbgVOCfwpw57nr1zQMcM8/TjmnKTnp+tNBGcEnr605ThhjvQApB44NO7dO3c0wt0xS7yeD2oACTg4X9aYzH07UoPb9c0jcjGT+dAg3n9PWlUk9scUwsQOvQCnA4H3qAHHJHSjnjgUjH5cE/rRnB6nqe9ADlB9P1oycDj9aaGB4z0oLfKBk9PWgBcnHSmknnj8qCeMfSkY8d6BgxbP3T19aVScD5f1pG755o6DAPQUCFGf7vb1oYsei/rTSwJJBoLZOM0APQk44/WhgSScd/WmxtjGD37U4N3OfegBGDdcU1WODx0FOOCMYPSmY7c9BQBIC3Uj9aPmx07+tNVugzSrwOCetAAN2CNtHOc4/WgKMepzSFs/40DHDJA+nrQhOCAM8U1XyAM0BuOvb1oAcxI6D9aaxIHCnp60pYZ5zxSNjH4etAhqlsk4NSKT/k1Hxknn86epBfNAEq59KMHrg9fWmq3HfpSg/N170ANOcE47+tNJYMAFpS3HXv6UxuSCCaAAE5+7S84Ax29aao9AelO68c8e9AwBODx+tAZg1Ju6jntTSec570CJdzccZ49aMkrjFMzxilz2FABkhs7fzNKucdB1po6inZ6DnrQA5c4/H1obOO350gYYxmgk569qAF6HoM49aYxbBwO9Ozjuenemn5h680AIS2c4707J5PtTMYPA7ilzgdTQBIjH9KCx4GP1pgYgj6UpbIHJ4oAcSeuO9IGbnj9aazcYzSAj1oAcxJx9KUFgD9PWmElvwp6k4PPagAXP+TRkjjHelwAOppHAwDjvQAoLYHHfpmhdx7U0HA6nhvWlDgnG6gAOcDjt60oLYHHakJ+UAE9PWkJ96AFLHBwO1I5PoetIxyc89PWlJDHOTQALnrjtTlDdCKah7ZPSlBBORQAHOMY/WhckDgUE9vzoRiMHJoAXnJ+vrQSSc0gII5ozntigAHcBe1Jk8YXp70LyMn+760jEdM0AKHYg5B60Anafl/WmdMkE9aVSMdO9AEgJJzinc7eg6etMJ9M0ofgAE9KAFUH07UHd1x+tIHwMgkUhYHqKABs56frTQzAnjvQx/wDrU0EAn296AJSTxx2pST129qZnJzntQGU/l6UAKCQ33e9Lk46d/Woyefxpd+P4j1oAfye3b1pATngUgbnqelG4Z6UDsOAPHHajJ5OPTvTdwHSjdnI9KBCgtu/+vQxYqR7U3ec8E9KbI529eDTQDiFAySM0ny7vvd6MDOMUh69KQCkICMGlyM9e/HFJtyfu9PekIPUD9aBisR3YdKDjOdwprdenanDB4x2FAhoCY4x+Apy7Qevf0o2fLwuKAuG/EUDH4XGc55pPlJGCOlBGe3f1pDndwP1oAMrgc9qcNmM5HX1phBHRe1PA9h1oEKAhxg0jbMZz2pMHjj9aRsjOB2oAG2dc9hQCpA/rQwJH3e3FIVz0H60AOypX73aj5Tg7u9IRwcDtRz6fxUAL8uT8wpPlwMNj6UnfgfrSAZAx6UAOwCAcnPFISvenAA4+WkK/Ljb29aB3BtmCM0AJj7wpH4yQuPqaE47fr9KBBhOoNIAvY9qcFGMAfrQY/QUACBcDnv60o29jQoGAPQ0DuMfrQMPlxwecU0BTyW9KXHGAKQf0FACgJkAntSjYBw3Q00jodp/E0Hvx3oEOGOue9NwvUGgZ6YPU0BcHgDr60AIuzAII6UBl2/e7UgHTikHTgfrQA/KZOD+FHyf3utIF5PHf1pdvTj9aAEXZjrTlKbjg0wZ5GO1OA+Y8frQA9dvZh70q4LD5h1pgB6AfrTlJ3/iKAAgbc7qaVXI+btTyMr93ofWmkZIGKAGgKOQc8U4hM9unc0ijjHtQVySf60ANO3nkdu9N+XJz607bycCkA5oAX5OMH/PFO+UZwefam7QRwtOAPPH60DEUIOM0/CgcHnNNVe1OKjP3e9AhMA96RivZu1KVJ6LTWB7DtQApK5wG7UfLtzn0pAOeR2HU0u3C4H86AAhc/e70p25OT29aQg7/AMaNpPAFAB8gI+YdKacDoe1Lgkj5TSMp4GKABihyfelATPB7etIUyfujrSqvoBQAbRjg9qkwgzg9hTSMDgdqXHXC+lADvlOOaa4BH3qAOnH60j9OlACDZ13d6TKlvvd6AOf+BetIFwc46+9AC5GB8w6UcHv6Uh6Aj09aUA8Db+dAB8oXr+NObaASGpmBjkU4ng4FAxRsIGW7Uo24+9+GKaAOP8aFGB0H50AOO317UKEGMNSY9BzihcEAH+dAgBXHBo47NTccdO/rS8noOnvQMVdpBOe1IdpIwR0pF6cf3aCP9ntQIQhSDgilAXBOe9IwJzilUEjp3NTqUKNp5z355oUoQCD2pMEHp+tIOMcVQhwK44NINgJ570iDK4x3707aFJ4H50CEwoXO40ALjOf84o2jGAO3rSeooAUFAeG7Uvy9iPemgHPT8zTgp5wP1oATAz97vTW2j+LvTyvOPcU0rlc4/WgABXPDChSOzduaaVO4cGhVOMe1AxxKA4zSBlyct6d6CDn7vejb1AWgA+UkjdSPjgbh1oCnceKGQ4Py9DxTAl9Ov1pCo3Z56UoIA+8eDSAjdncaQhSBx1oIBHQ0oxkHcaQlR/EetACYGR8ppwUf3T0HemswHRj0oEg5+c9KAHbV29DSkAsOD1pu8bT8569qN4LdT1oGKcZ78etAwMdenrSEgfxGgEZBD/rQIGAyDg9KcAPQ0xiMjD9vWlXA53HtQA9QCOn60jKMdO1CsAOGPIoLYHDdhQAmAR0xxRxngUMy9Q/bmk3Yx81ACkDHQ0YBAIzyfekLADAY9e1AcHAJbrQAgHPK0h4Awp6UqlR0J6+tISCByenegBwIGDg9KDgrz7dqRTwPmNBI253HigAcgjOO/pQpGBwaQsMcMevrSpjH3v8APFAD1AxnBpSF/umkU46Mf84oZlB4J6UAIOMZB60DGeh600MOPmPXtSAg9z1oGh5Axggmm5GDwfu0oIGOT+dNyMY3Hp2oEB4P/wBalwOflJ5poYHBy1OUjB5brQAqj5c7TQ3XGDQuNvBNGATjdQAhAwDg00YxyD/hSjbx83akBGMbjQAvtjv3NLjpkdvWjIz94/5xRuAxhj0oARQOSFPSlx8xwO3pSK3UliPwpwxuOG/CgBwGMcE8UAANnB6ikDDsT+VAcBvvH71AC5HoevrRwSCFJ+lRl+OXPWnKwGME0AKvHJU9KCM9j0pqt6N2pcjoT260DW4uOvHpTcDceO9KCACQ3cUEgtncaAAAADrSqARnFN3AY+Y05WGOGNACgDjINK2MfdPWmgjOQ3p/SjeD0b+KgBcDGcHrSFRkcE8UKw7MaUsOoJ6UCE24/hPQdaXAIyQTzQWX1P3fWk3gDhz1FAAQCw+U9aCAcgA9KTcC3U8tSgg/xHpQAoH+z260Mo44pAwBHzfhQWBxhu1ACFRzxQAOOP1oJHQt6UAjONxNACtgD7p6UvGCMdqYT0wx6U7cOfmPSgBRgHpSHaR0PWgMB0JoJ4GCaABQMD5T19abxnOKcHGB8zdaZuAP3jQAYAH3T0peMZINNLDHU9PWnAjj5vSgBGAIzg8Ghs54B696GI2n5qRiOec4oKWw8Y4zmlUjGcU0MAB85oVhjhzQIecddppF6A4oLDsx9aRWXA+Y0CDHf3pRg4OD0pCV5IPelUj1PSgBFUYJIbpSEDPenLgDqelIxHHzGgBMDJ4NKoGw8GjI5wxoB+U/MaBgQM5wabkYGQelG4FuppARgcnkUAOQDH3TT+P7p/GmR428E9KeCOfmNMQhGex6U0AZ4U9PSnEjPDHpTAwAJLfrSAcQM9O1LgccGm55wG7UqsOMsTQAoHOdp6+tBAxjB60gf5up60pYYJBbr60AJtXqFJ4pAg/umnBgCME9KarAdGPSgBdoIHBpNo5GO9GR03Hr/hTlK4OWPagBoAzyDQwAGAD69KcNucg9qGC8YbtimrABzg/NSAnd1FKT2pM4akAc8YP4Yowe5H5Ubugx2pe2Tj8aAGtknqOlGG5we3pStweAOnpQT8x4H5UAIu7aef0pp4br/F6U7PHHrSMOenegBcc8Ffw+tKcg9f0oGTnI7UDGeg9qAGncOh/h9KUbume47Uhwew6elOAUdh0pgKN3H09KRicdQPwpVx6U1utIBDnB59O1LlvUdaTPbApVHGPagBWzj/61C5457+lNJ9hS56DjrQAihiSSw60YJA6cUo65x3oGSAMDigByglevakYHbjPb0p3J7Dp6UjDnOKAGsuOMjrSoCMcihm5IA70KwwPr1oAcN2Ovb0pr7vUflQCD0ApGOTwO1ACDJxg/pRk4PPXPalHI5HfvQeSfr2oAM+uKRQdvUdKUnmlXABOB0oATYcjkce1KRweR1pxHsKbuOCPegBVzjtQMjuPwpByOcUrHnpQAYOBz29KYAcHBHT0pxPyjI7Ug5z06d6AFwwPH8qCGHf8ASgjJwAO3QUrD2HegBgBOee3pSgnd/wDWoAGTwOlKVGeBQAKGIGSKPm3dR97mlxzjHQUmfmzjv60AMYEjgjr6Uo7cj8qOPQdaASCPYetAAmc8Ht6UoDZHP6ULz2HSlxwMgcUAADYyW7jtSE88n9KcCAOg7dab/F070DAA4GGH5U5d2Ov6UL0Bx0pc8dKBCE4P3v0pDu9e/pQTzwPypTgjpjnmgdhBk9xQxJI5FKvAz796Tn0H5UAHze3SgBtvUUu4k9B0ppI29B+dAgOd3UfepSWJ4I/D601jk9KcOc8Dp6UAKd2eT69qbzkHHb0pVPPSkOOwHA7UALzjj27UvzZxkflSH/dFKOmMUANbI79vSl5weR0pGxxx2pQcg5A6dcUAIC3TIpxPTpSL3oLY6UAA3EAA9/SmAMXOSDTi3A6daaDzn+lAAQRjBH+cU4AkAgjt2pOcfhTlx39KAAglcZ/SkbIBwe/TFOxxwKR8E8KKAG8nHPf0pUDFc5/SkUng47+tOUk9qAFOexH5UcgDHp2FDMSM496Bg4470ANO7HPr6U5c/wB4flTQBnp3/rTuP1oAdHnGB6U1wSfvAUqtgEYHA7mhiSM47UANOTn5hQNxX8KNx5BGeaQYx0/SgBMHOd3ehQdoOe3cCg8nHFKFOBx0FACx52np064p2CCeR1/u0injoOlL3OBQAjbscHt6UzDc/SnHH93tSADPQUAIQwPXt0xSrux17elLgHHHYUpHfb2oAaB8+AR1owSOSOvpSg8jjv60Y4xjv60AAB4ywpoLevanbiCMdvU03qen40AA3Z69z2pyk88+lNAAxwOtKuME4HGO1ADsnccN27Ujk4yG70AnOSKHJwQPrQAgXnp39KMfN/WnEjgkUg69KAAKRjrwPWlAx60Enj5aUcjBWgBjDJ/CkIz25xTj16dvWgqMk+3rQA0A465pWA/lQSNvt9aC3OAv60AL17H86TBJGePU5oBK5+XtQpAPT9aAGsOQfanbefw9aQ5Jxjt608AnqO1ACKvHfp60jA4xjtTwMfl2NNccdO1ACEZByOgGKXb36UZPINKp9B2oAaw46UEYwOeDTiBjGP1pOOOO/rQA1R3pwA469BSKSSePxpwxtHTp60ALgN2/WmvwOh6U8Dpx2prkgZxQBG/070qD+dJIT0C96WMk4z6igBQCTnn86QqffpT15PA/Wkbk49vWgBFXOPw70BcnJHenoOmP50Y5/GgBoXC4xSgHkj09aceTwv60g4BBU9PWgBhB/SjHXPrTmG3tSepx39aABelGznp37UDgdPyNLuyen60AN24X8KFGcjHanZGOfT1pB0PHagAwc4x3Hejbx+FKcZpSR09vWgBqr1GO1PKjdikU85I4+tKTk5x2oATGPl9qaQM56c04njgfmaaDlskd6AGMM569aAMnv0pTzk4HWlGMg4HfvQAIPr05p204yAelIuOw7etO7dO3rQA3BPP0pAMtz60uTjoO3WkB57daAFUDAGP1oxkCjcQAOP8AOKN+QRxQAhHQZ7UY54z96kHJHH61J1H3e/rQMaFPWggEjGfrTh0xt/WkI5HB/OgQ0g8f7tN5xye3rT257fwjvTcgr07+tAAQAfxpcegP50hHt39acOM8dvWgBMEnPPekI6ZHanLwe350j5HQdvWgAwT19vSgD2/WgnAIx270oOBn+tADW6AD0pQp64/h9aD/AEFLnAIx29aAAD0z+dI3TA9KUZzwKCBxwPzoAZyB36+tIox+frT+Co6df71Io5PH5mgAAGOnan4yBj86aMYHHb+9T16AgHp60AJjjkHp3pHUAf8A16XdwTSMQc9OvY0ANK+g70IpPY9O5pQcf/rpV9cdvWgYY4yB2oA6fWjOSQB39aUEcZ/nQIZj6/nS4x/Tmg8k8DrS/ePT9aABenXt60u3p2oU/Kcr29aMkY+U0AIRwaABzTuOeO/rSY4yB29aAE247d6AnH3f4aUnnH9acBwOO3PNADQCRjHagqcmnDgdO1DDn/69ADMEcj09aQLnPFPOMdO3rTV4JPr70ALjBH09aUA5zQx6fKenrSE9QVoAQD5gAe9BXIz+tKrHdkDofWnHG3IHegBgXB6dKbjB6VJwccdvWmHBPA/WgBNp7Z/OlUHk9KBwACPxzSqfb070ACr83WnFMjHoKAfm5H60rMMcY5oAjXsecUg4bIzT8DB5NIAu4HcaAEA6HaeppQBnOD05pQOB160dDnJoARj6A9KC3pnpQzD1PT0o3A55PQUDsNzhTwfzpCBuGQetPwNp+Y0MFLcsaAsMKDsp6UY+cHBp5xnIJoAGQQT70CGYxjKngetPUADAB6DvQQOM5/EU5SP7x6UDEx/snp601hkcA9PWng/7X6Ujc4y36UCGkccqeAOtAGcfKevrTsAjqelKABjBP5UANbgDAPHvTAc44qQ7SvJ7+lN4yMMetA0MQYJIWng4AIB6UDAOdxpVKlQMnP0oBodgY4B6U1/u/dPT1p/HYnkelIQMfeP5UCInUAE4/ioQYxhT+dSMoOck9elIqqACCfyoARRkdD270vvg0oBz97v6UMc85P5UAIpwB8p/OlxgEYPX1oXBAOTwfSgYBzk9aAFGCOh6etNGNp4/hp/GO/T0pFIAPJ6UAMYZb7ho98HrTiFOOTRlBkFj19KAGrnb0P50Zwc4b86UMu3gn8qOCchqAEXOBwelCDK4CnpSgd8np6UIRt6nP0oAUgZ6Hr60Yx2PT1pcg85NBOcAk/lQAwd/lPSnA8n5T+dIMc9enpQcE5yelACDgZAP50cF8kd6XcB3P1xSbhuzuPWgBCoI+7mlGRjAPSjII4Y9aA3TJPT0oARD9enr9KcvrtPT1pFwDnJ6elOA4HXp6UANxgE89u9IepOD+dSdc8nr2FMIBJJz+VADCOmFPalTjgA0uOByelOTGMljx7UANUDjANOIGMhT96gEZAyfypxGRznrQPoIoGfunpSMAegPSnKcdzQdpwQW6UCGOPRe1MXC8YPUd6lJDHqfu03OByx6igBAAD07igrnkA9PWnDaT95uvpTsc8Ht6UAMwARwfxPtSN7KelSBQDnJ/KmNjg8/lQA3bnnB7UY9QaftB/iPGKXaCMEmgCMnHOD0pQOvB/P6UOABjcenpShsgkk9KAEA6fKe3el7fdPHvSryNwJ+lB28cmgBo4A4PX1pqABs7T+dPyAB8x60g453GgBM8DIPT1+lL1HAPbFGRjgnp6Uoxxye3agBDgr900pJ5AzRt4wSevpSt0OSetBXQaM8H5unrQBxjafzp4AxnJ6elIABj5jQITAIzg0ozxgH86CevzH8qVDkDr+VA3YbtIH3T+dAwMYU9PWnYB5yevpQABwSeB6UEjFPBGD09aD1B2n86eMAH5j0pCAMHJoAaWwCMHt1NOU8dD6cUmRzye3anKRtyCevpQAqr1+VvzoUEAfKelKcZ6npRkYHJ6UAIuMdD09aABk8Hk+tC8Dgn8qcvcnPWgBhGF+6fu+tNXqeD0qQ7em4/lTQBk/e/KgBCMdj0FIAP7hp5Az36UmBx16elADT97vwfWkLE8DPX1p2QSeT16gUHB5y3WgAU8Dg/nTVGOx6VICBg5PT0pq4zkE9P7tADcDOcH060KOvB7d6cRnqT19KFxyQT27UD6AucnrQRgYANKOvU/lSnp0PSgQhbA696Td8w9qAPRaQhsnj1oAUuDjBo3cYBpuDuA28UpBHG39aABjz16D1pwOSRk9KYwPPHb1pwBz93P40DQ7oOG/Sh2O7Ge/pSjJX7v60j9encUAIT6GkDYPWkbIyNvcUZOcY7dzQIN3IC/hTg3+12xTcZ7du9OAPp+tAChuPvUMTg89KPT5f1prHPb0oAdu4xn6UZxzTTnH3fTqaCDxx+tACl+OD+lG4nHzfxelIQcY29qVQ3p3PegA991AJAGDQATn5aCDgYWgBwJ45pM8cEcelJzt+729KVixXO2gBGPH40obIGTSNkg/L+tAz/dP+cUDF3e9NJxQSey96Q56AdqBArYxnilDZFNXIxgd+5pVBAztHX1oAeCfXtQHwCN3YUg5xgdqQA8jae1ACls9/0phfGck/lSkEEYXt60jIcHC0ACucYz39KN2T1/KkCNj7vegA/wB3v60AODAgfShG7UiAkD5e3rTlU46D86ADcMfeFLv9KCp3fd/Wm4JxkdT60DDd1NDHng9qFDYJ29qUgkn5aAG7jnGaQMc/e/ipcMeNtADZ6d6BAG3DOe/pSqefvdqTacDC96dtIx8vagAVuCM9qdkdcimgN/dPQU75s9O3rQOwm7qP1ppY880pJ5BHcd6acknjv60ADEEdelKre/akAJI4/M/SlCk8YH50ACscghu3NPyMDDc01VbPT9fpTgTjoPvetACg8fepCxGOenelOc9BTXHt2oEIW+bG7tSFv9r8hQcg/dPShVYp92gBc5br39KduGD83akKktnB6ihgc8CgBN/zDmkJJI5oAbcOP1owTjK0APDehoLAdCPzpDkA8frSnccZX9aAI5GOOPwoDDB57UjBiBxS4Pp2FADlb3pS3HXvTeR0HfuaVgSORQAu7KgZ78cUmQD1o6jkdz3pADk/Jx9aAELHA+lOGeOO1IVbA+XtTgvAyKAAk4IJpGYEHB70MPl+7SMCSeKADf2z+lAbIxmkAbj5f1oAOOVHX1oAUnnrSo2AATSYPYDp60KpwOO9ACh85wRRuz37UmDjO2lUE4O2gBVfgj2oJGOopFUkdO39aVuoAX8aAG560qHjr34pCp54pyq3TFAAWAPBpA+QMelABz92gA4Hynp1zQAqPxS7gBjIpmDt+7S55Pynr60AOZs9Dn8aZv5OTRnodp/OkAPcUALvyRtbtS7gcYNMAbPT+GnAHn5f1oAVTzjPelyMZB7+lNGc/dPX0pecD5T97uaADeQRg0K3cHtQAcj5f1pAD/d7CgBe30oDYzg9KME/wfnSYPI20DsOB+Y4NKTleD+NIqtk8elOK8dMcUCGjJ4pOSc0oxmgDkAUANw2Rx60EMMYBp+OBzQBnoKAI2z79Kdk889qVl6fT0oCkHr+lJjQm5sZoLnOcntSkEjA/lTWHzZJ9O1JPUALHqPWm5IYYpwGe9CAZxVCBScjn607J2/gOaQgDpTwBnigBuWHemkkd+1PHIppX+VACbmAP0FO5wMUADHU9KUDtmgBG6UAnPXuaD060DPB3DrS6gALd89O9BJxnJ6UufQ03cSvUdKEMdk4ppJ2Z9u9OGT37dqCB0pghrMcGjd6c8f0odfcde1G3jg9aAGhnP4UEtnOTTivJGRSbenNAAN3B96VQ22lUDPWlAXOBQNCAscZoDHGc9hS4A6Uh4B5xxQSJk5FIQcHNPHTkmkA6gmgBoGPzpBnP48GnBRijaCeDz9aAERmwPoaVSduR+v0o2jAx6ULzxQApLkn/CkGc/zpTjPB+tAGetAxADjA9KUk80Kue9KQCcg9qAGYJJxQv3s+4pxA5ApoGGwT3oEALbR9adnGMelC8ZGT1ozyADQMFPHTtSZ5xigdcD0peoyDQAws2Tj2pAzbjj1pwXdn60BQWPNAAu7I4pwBAA9qAB0B/Wl4PQ0CG80uW9f4qOMilIAHB6GgeoZbNNcnt6U7kikK5O2gQ0k5yPSgZ2nilAw2M9hRjjKn9KAFyc9T1FGSDkUg69aevegBMEEEcUhzxj8akA5pjAAgAfpQA0lsflSb265pSox1oH1FADSTjJ9OKMtk/SlYLjI/lSgDBB9KAEycYFDE44z2pygnoaay5HX9KAAMQOD3P8qFZskkdqMcDnvSgk87qAFyQOPSlzjkGmZ4654pRnuaAAlttIS3OT6UpAPFK2ACAe9ADQWGMelKAccUoHA+nrSgZ70AN+bjBpAx4px6de9C4IHNACDcVyaVc5oGOg9e1JkDtQAKzY/ClLEkfWmjocHHFOBJ70AAyATmgHg89KMggjdmkGcZoAAzGkBbAOaGAzkHtzS7eAPagBoJxnPPFCs+SaFA6A9KcFGBj86AGgtQM459utOIz+ApFA5oAQbj1HQUuef5UEKCOe1GAelAxATuzz1FOBbb360ig7uvenKOMZ70kAgZgAc44pqk/oKceG49O1NAOeD25pghylvSlBOTxQBkfT1oByDzQAoY5Ofags34UgILHmhyACc9+KAAYHUjr3oAXdj3pFBH505QQetAhQF4xj3pwC7ccUnAA5pQTzyaAEZVzyKTauc4J4Halbd6mgZJOSeg70ANCrjkHqO1DKAen6U8D5c7h+dNcHPDdvWgCMhQT7ClG3OD6UpU54J6UKpzQAjAfp2FOAGeAfu+lIRnAGelKEbHOelAANo//VRgcfQdqUK3Ymm4OO/SgBRtx1pQFH5UAHHU8DigZIwD2oAQhcD8KAFIGPX0NKc44zSAdBk/nQA0Fd2P6UmFCjjt6U8Jyev500rgDB/WgBRtIH+FLkbc5pqgkAgUvzepoGKwB7Hr6UBRjofyoII65605QSAQT1oENIUHoevpTTt6Anp6U/GTwT601hz36UACbRjml+X9fSkUHAwSOlAznqevrQA7AAOPT0pMLg/T0o5xxSrnaeSOKAGnaGwD+lHy8nvn0pSORzik5IOG70DQgx1x39KUbSeB29KaFOMZNKBzn0xQDsOXaVBA7elIAAD9PShcgABu3SgA+p6UCA4z09O1HGeh4HpQw7ZNGD+lACrjBx6elBx6enamruycE9KUk54zQAEjv+opuVznNHpQuS33j1oAX5QM570fKccdvSg5x97vS4ORg549aAGqMH7vb0pQBjgd/ShVOaUA4wTQNAoX9aPlzwKcAcH5vpTcHdwaAYDbxg9qMjHA/SjB4AP+eKMMBwD+FAhMrke3tS5XHvn0pMHORS4OOT3oAUYI/wDrUYXg5owcZyevFGCCME9KAEwp59vSk4xyP0pfm9T92gDgkNj8aAE+UnnPX0+lOBXB57elIQc5z3pTuGcZ6UAKCueD1pG2YoBYsOaQhuMHtQANt/lQNuaUgjnJ7UmMjBNAxG2kf/WpQFwT7U0gjGPSlUHH3j93igQq47GggdR6+lCg54zQQ3GCetAANuB9aQBc8CnDOAcn71Iq85z3oATAwMeg7GgAA9P0pdpwMHPFAU0AIduDz+lKxB6H07UhyR3obPqetADl29jQMYH+FC8AcmlUEjOe1ADTtIOKBt4w3elcHse9NGePmoADtBz/AEpcKfXp6U0A5OCadtweP50AIFX9PSj5cj6elOUNtPXpQRyOT0oGhuVyf8KFAC5/HpQQckg0Lkjhj/nFAhcDOP6UBV2jgdPSggj160c7cZPSgBAAO3alwM4A6e1A9dx6UEYPXv3oACFz+HpTV2kn6elKRzwe1IoPXnpQA4heMDt6UqhSf/rUhByOvQU4Bhgc9KAEVRnn19KUhQOP73XFKoOQcnrSHJBO49aAGlQecfpSADOSP0pxxkEGmD6mgB2FxgevpSDbg49fSgDjqetG088n60AIp+Yj+lKTxjB6+lJjDUEcDHpQO48cc+lJuw2cH86cecDOeeaYcluKBC7iMdfzpwbn/wCvTDkgEevpSrkDFADy4PGPxzS9ScjoPWmZPp2pQxyR7UAPBIX6UjNz0/Wml8Dj19aRn+bp+tACn73I6e9IDhhgfjmgsSSfakUsGoAXPA+Xt607Ge3buaYSc4Apwb+VADhjjjt601sDHy9vWl3YzzTWY9cUAHUYPp604EEcDH40wHIOBSg+lACluOFH50Z4xgfnTSxHQfWgOflGR1oAcvyk8frSMQVB/TdQD3zSFiQDigAyP7v60nbBA/OlG707UhHOaBod+A6+tOUrtHy/jmmtnGScc0Kfl4xQIXIPBX/x6mscdv1pSwz179KYznofSgBykccdx3o3ZzxSRtngGkJOePWgB4YDjb+tC9DgY+X1phc5pVZgp4PSgBx5YHnp60hPB47/AN6lye47Um7r9fWgBARjkd/Wjoef/QqAeTz3obcTnFAwUkDHt60K3BG3t6mjJ29O1IuTxigQpOD/APXoyBjA7f3qMHt6UHpk+lAAuOeO3rRn5uR+tID1pSecA0AAGOQv60g+8PY9d1PPTGKbznkd6AA8j8aVSOOO396mbuPfNKH+YCgBwPJwPrzTgeQSPrzTFYntTi2VAoGKOQeO4703d82cd/WlB460gPOMd6AF4wMD9fpRgAfdFBJOKM/KfpQIQEZHHf1pckD8f71MLfN0708H5ce9AxRkDp+tIT04/M0oJAIz2pDnIwKBCMCeozhfWkXIGfQj+KnZJPPpTWJx170AOJG78em6jgjp29aQ5z0pQcnkdqAAfeyBj8aRiMDI7etKpO6kbJ4oAD9O470Z45H15oJz1pcnnigYxjjA9vWlGcE/7PrSMe+KASePagBwwedv60E8dP1pFY9sUjNkCgQ4E8c9/WheCeP1pA54Ge9Gff8AGgBSxwBj/wAepcjHA6e9N3HAGMcetKCemO1ACHGOn40E4zkd/WlbnJ/pSOSOnrQAqngYH/j3tT1IwML29ajRiB+FPGfXtQAjHHUd/WmjJxx39aexPIHrSJnGAKAGjABGO/rS7tvGP/HqCOfxpGyeRQAq9MD09aDyR1/OhTgHHp60bgep7UAGcZ+v96jPHT9c0m4nIA70A57d6AA8Z479jQDhRn0/vUNlqO34UAICMH5e3rTgcN07+v0pq5Ix7U7nnigA4wOO3rSL0Py9PenEn8hTVPJoAceDgjPApVYDoO3rSHJx9KXNAAG+bgdD60jHI5/vUDP6+tIcnkjnNACFzwMn86aGOeh/OnEsSMDpTec9O1ACqe5HcfxU7IOTj8M01c4xjvTlJIOTQAZAbgd/WmucA4OOaXOWpJCf1oAfyBxnrTed3KmnMW9j+NN3ENnFACnHB2/hSjJGMUgcjAzShvTFAAc54H60hJBPXp607cfXtTWLZ60DE3kLyD1HpSFznHPbvQdxHPr6UZbIG6gQEnkbT0oBOfun65pQc5A9KACCOB+dACZJx8p6U4E9Np4FNOcjIFLyPTpQAu4j1/OkJz2NAJ9RSE9On5UALzgkKe1Ck4xtNLk4wcdKBntjigdxCeOR37mgNgrx39aU5x1pvPGD3oEKGPv9aTdwMA/nSbmJ+lAZsDn9aAHDt8poJIH3T0oBIAP0pSTjt0oAGJx0PX1pFLYHB60rlucnvSKWHOR+f0oAQs2c4PWmszeh6U78R1pCCSOewoAahPHB6+tKCScBT1pygjGCOtHOenegBDkAnH5mhScfgP4qdyRwaQZ9e3tQAZ9jS7uDwevrTTnjpSjIBy3egBQSR0PWkOe6GhSQME8/WlLHPb86AEwcfdPShM/3TSgHHbpSKSPTgUAKc8fKelBJxkA9KU+2OAKbk4wcdKAEBJzweaUsc/dP1pATznsKU5LdvxoAXLeh6Ugzn7p+91FKCfypOQ3UdaB3sMJOO/X1pcnI6/nSMSB97vRg5zkdOgoEKjEnkGnAnA4PSmLu/SnqTwM/rQA5TxkqeooyQcbSaEzjnjkUAnceR1oATkAYU9PWjJ28KelBOMHjpSAnGeOlA7iEknoetLk46d/WgE5HFHOOSOtIaasOVjjAU9KMng4NIDx1FBbHRhTJHMMH7p6etICcZwfzoLNnqOnrSBiRmgAbOfx/vUuevyn86CTnqOtG45OMUAAJBzg/nQ2ePlNCkhs5FKTnHIoAQgj+E9qTJx909OtOJOMcfnRyOhH50ARtuPO00oyedp6UPnA4pRnHbpQAi5H8J/OhmPp39aVSf9mkJJGc+nagAVjgZB60gY5+6fypRuIAz3FIM9Dj86AAEhQCD0pV6DCnpTRkAYI/OlBxgcUAK2f7ppGJ7g9fWlJPPSlbPqOtACLnGdvb1p65/ummgtjt09aVCR1x0oAXnAyppV6Dg/nQScduPakDYAwRQAhzuxtPWkOcY2mlLc9R1pOcYwOlAAp4yAenrQT7H8TSKWGcHtS5bPTt6UABbAOB39abuOM8/lTiWGTnv6Ug3YoAD1+6adjKjg9KTnqcfnSjJA6dPWnoAiAkfdPTrSnrjaetAJx0HAoB+lIBDu6YPSkBOTwelOJ+nIpoJBPI6UAOLHjg9BSgk/wnpTCxyPoOpp6kk9KABc+hoI4xz19aUZ45zzSEkDqOvNACYwRwenrSAH+6acCcjmkVieaAEHToetIC2CMHqO9OyQKaS2DigBMnd900SEkdPek3MSRihi2D9KAJDnpTTknPp3pxPb3pON340AByMD39aUZ7kD6mjOcYpc4/OgA57HtQc5P8qRiAenagOCSMdqAAgkf40EHdnNBbg8UMcnOPrQADPr29aUA5/wDr0m7r9KVWwwNACEe1LyB+HrSMw4wO1Lu7f7NACEMB/wDXpDn06DrTsjqPSmEj0oAcM47fnSjkf400NweO1OUgdaAAg46U3qBx+tKzD0pAw4GO9AAAeoI6+tABIGMfnQG7gd6UHIHHagBVHAOaOcf40bvajcCc0ADA/wD6zSDpx7d6Hcdff0oVhwPcUAAHPHr60FT0x29aUEZ60FgR+FAAoOBx3FJzTlIyPamswySKAAjnHH50gBx17etBYZ6UIwAI4PHpQApUg0h6HkdaeTz0/SkL9cDvQA0BsdvzpefUfnSAjGcUFuelAC84B9u5oXJ4HpS7hgfSkVsk0ALtb1HbvSEH9PWnFh1+nNIx4/CgBgU88UuDn8PWgEcmnHrQAmD6dqQg7scdacSMZx2pu7DdO9A7CEEDj19aNp44PSlLZGcfxdKAV447e9AhiqQenanjOeB2pARnIXFKDkDA7UAOXJH4igZ3Z96AwxSK3zZNAAcgD6etGDj/AOvTs5AxSEgL+FADVHtTsED8fWmkjP4+lO3Ar+NAwwT6Uhz1HpS9vxprMM5x2oAOc9O1HOOg696QuM8D+Gk8w7elAEhzu7Uc9vSmF8tyO9ODA/lQAoznOP1owemO1IrYPPpQxB/KgB2TgED070vOOP1NN3jA57DvShxzg0CEZePwoCkDGR09aGYY60oYHP0oGIAf8mmnpinZANNZsjPtQIVQcDBHX0oHpSowAHB4NJuGdwoAbyRkUAEDIPak3ZA+WlVugoAMHnmlOTn60hOaViCaBirnAx/OlQH0/WkQjHJ70qlT9aAFOSORSDOB7e9IX4zQjDIx2oEGD1P6mjGe3QUZH60pPtQA0DOeR09aCOe3SnIwAI9qCc9u1ACbeD0696TBINO3ZyQvf1pARjmgBOc//XpRkKDjt60EgnNIWGB9KAFHIPQ0oJzk88+tMVuuc0/IJJPrQAhGevp60gB646U5iDz7U0HqcUAIc5GR2705cmmvyRx2p2RgcdB6UAKpII57+tBJxnPf1pgYK3Tv6UvmZ/h70AKc+v60gz0z+NL5mT+B70gI7AdPWgAz0/xo+Ygnjr60ZBGB7ZoB6n3oAAvPWkdTtwPxo3DdnFOZhgj0FAAQMdB1zTeAwyKfu7Yppb5v/rUALxgDFHbGKTdgAZHXvQWNACMRmjgk4GeKR2PXHalDHkkD8aAFOAvNDEEj8KTfkcfyo3Hd0/ioAcOe1GADnFIHGOn60bvQdvWgBCOcAdqcowenamljnp2p2T1x0oAAARkKelIRk9KUH6UjHjpQAoHHT86MY7Um/APH5ml3cdKAEfpnFNBHy/WnM3HFJk5AIH3vSgAAGM7f1owMDApVb6flQWIUHA6elAAD0OORQemQKN2OcdqCxA6D8qABwPTvSL9O9KzE5/CgHn8fSgAXGeBQfTHajcew/nSFyDQAo5HSmtgk/WnK3sKRyc5x3oGhCwz/APXoUjB+nrSAnP09qchO3H+z6UCHd+lJkc8frQSQe1G5sH/CgYgIPalbGfu0biQeO/pSFsnt+VAg7dO1KnPbnFNVyQOKVX+XIGePWgB/4elIR7dsUhY9x0pd/HSgAUdeKcw54FNDY6jtSls54H4mgAccYx2pnBbk96fv54HSmbm39O9A0xAQeff1pQQDgDHHrTdxxn39KUOQRj0oEAx1ApwA29D0pqOegHb0pwZumBQAAdcj0oXrQCQT06CguQT/AIUAOAAAytIQCOF7Uhc8c/nS7sdulADOrdOpp3bPvTSxzj0pQxxgetA0LkbelNYjOPalyaGJyBjt6UANJBIAHQDvQcAdP1pcknkdhSZGM4H5UCA4z0704EcgDtSZO4cD73pQW9qAHLjPTtTSeelG4g/hSEkkcUAOPTp2oB4PFISSOn6UufagAY98Ht3oGDk47UhJx92lz1H60AAbHb9aDjH/ANekDUpY4xigBQeBx3pOM5xil3kgcd6TOe1A0xMjA4pRjI4pNzBenalJ9FFAhCfakc5z/U0FuOlBfI6d6AFQjjg/nTk+n60wMfQdPSnq2BnrxQAEcYxSLzjjvTyewFMDYwKAAYB6d6dxjjH51GGb2/Cnq54oAVSACT6UFR6UBjjj0oLY4wOlABkd/X1pR0pA55I/lShvlPA6+lACMAecU0rgDjtTt1JuJGD6UAIg65Han4B429qYrcdqcrfNQApwT07UmAe344pxbjIpAxJwaAGsOny9hQ2MdO2KcWz+VIWoAjON2Se9Hykceuadkl+vekIwOR39KB9ADAHGO3rRnnilGR2HSkH0HT0oEHBFAXA5FAJB5A/KjdjPA6elAB36UrHg8dqbu5IwPypJGJU4PU+lAEoZccjnNISmc7e1BApOc8elACllBAwO/ak3LkHbSlRgH+tJtzxg0ADFcdO1BKZPy9qCp9PxpMZJ+goGKpQL9307Ubk3D5f4qQrheBRtOenf1oEKWTOAv6Ublz93saNhPalCjP8AKgBCV6he1O3IB0pNo4zRt+vSgBwKE/d7elNYrjhe1A68CgjjGO1AAxXqF6Yo3L2X9KCuR09KAuOi0ABKbPujp6UhZcglf4qRunHpSEZIH+1QMehQE8fpRuTbwvY0m0dqTGEB9qAHsVxwO1ISm0/L+lAA9O3WjaPT9aBCkoQcr+lKNgbOO/pSMOwFGMUAGVIxt/SkYpu4UUu3mmkdKABWUY+X0pcrjlaaOo49KAAScDv60DsOTYABt/SlRkx93+Gk28/KKAvHA/hoEOOwsCAPyoYoQTgflRj2pp6Hj9KBi5Qqfl/SkGzJ47+lAHBA9fSgD2oENXbx8v60Kybfu9qXZwCB2pFGc+w4oAdmPcTt70blz939aRhznHegg56dqAEDryQvanhlGfl9O1RgcnjtTwOSQPxoAcrR91/SkzHv4XvRt9vxpADu5HegBDs28Dv6UDYCOKcFBGSO9Jt+YYH6Ur6gIuz+7/D60AqD07UBc5+XtSlcDkUwDcnJAHQUhZMnAHX0oxnPNIB8xOKAHZXj5f0+lClcY2/pRtyB9KMccjtQAgZMj5c8UAr12/xUhGMYHagjgcfxd6BjgU67frSsycALTQAR0+lB5xQMVinoPuimKyAcr6dqUjJ/AU1lwOB+lAhwZd/3e4oJTsBTSuGxjuO9O2+v4UALlcj5aQleML2pwBLYxTdvIHtQIdlMZ2+nalAQ8hf0pNvNOA44A6UANOwj7v8ADSlk/u/54oZcjAHakC9cjoKAAbSfu0uUCj5fSlUDHTt60jDigBVKY+6OvP5U0bNxO0flS4/9CpMeg/SgAyhUYQdDSsUxwO1Nxhc47UoHPTtQA07dudg/Kg7ecrTmUYzj9KR1HIFACKy5Hy9qchXbgL29KQLx0HSnKozx6dqAHBkIzt9O1JlCAdtBGBwKQc0DWjG/L1K/pTkKDjZ344pNo5AH60FeflFAXHKUx93+GkbYSOB09KRc7Tx2pdo9D+VAgyhBwtLuTBwtNxweKACe3egLjhsznaOtNLIQOO1BHGNpoK9OO1ADVZcD5fSnIyAnj9KaF6+wox7UAPDIBwvb0oDLu6fpTcc5x2oHXgfrQA7cpH3f4aXcnOV700AE5x2pSAeMfjQMAVznb/FSsUPO3v6e9IFOeR/FQoJHQ0CFBTgbaF8v+7/DSEcjFNzk4A7UAKQmc4HGKPk+bAHT0o7EkdqPX9KAGjaCRj0prFcAlehp2CW6UkikgnHegBf89KBw2M9+tKPp+YpQRuyfXrigBMgEc/rRxx9adngYoz/tfnQA00hxjr2HenlvU9vSjOSR7CgBi4K5BpcfNTx8o6nrQTzkevpQAzgd+3PNHGQenFSDrn2pBwR9KAGZPB9qU8ryOw5oPbin+/tQBHz/APrpOcfhUmRwQO3rTWOOKADkDnPQUEgc4pQ3B47UBsdP5UAMbp19KQ54yT17/SnM3GM9x2pA/Tp+AoGIp9T+tGQMfSnA46E0gbABOelAhQcAc9qCeOfbrSqwHAPTrSkqo+YgfU4oC4jdOB3pB/hTj1xt7+lSKjEfcP8A3zQF0RYOeh603J469KmZGXO5Mc9ximOu1sYxgU7MCMZ4we4zQCR1Jp/PGB3pCcNgnvSHcRW460qnI69qC3oTQHwOp6UCBm5GKQtwTn9aUsCRyaN+c89/SgaDIwcnv60A9welCnj/AOxp3J70Axq9Og+76UgAx36U/IwOe1IuOcZ6UCGsOe/WkOevt1qQ49/zpDg9u3rQBGCcn6U4HBPWlGOfp60Hhu9AADzxRuw4we9Lnjihm569/SgBAeM/7VOHakByOp69qcGwQOaAGr1x7UdcEHFO3d89qUYOM0ARevtijHzEEd6k9/cc0ncfWgBoJwBz2pR04NL6fhSg4BOTQBHnn8O9BPHX+L1p+eRz+dHB7/xUAR+wP5UE8jmpR9T09KaSRigBgPqe3ekGCuQacW5wCelCnCnk9aAE/i696XOD17U7POcnrQST1zwKAGg88elISRjrwKcDz17UFumKADccZOe1Ln60px36jFAIPb9aAGk5A49KUAEHjt60E4A69PWnLzzk9KAEDEHr270jNgHHpThxyPSms2OB69QKAEBzwT39aaG+bGe/rS7sYye/pSg4PJNACZAAx6etKCBjntQGyo5PSnAnAwO1ADScrnNI3P8AF3p2DjH9KUnB6nr6UD6DRn1pVOQO9OU9DzQp+vSgQ3OAO3FAPAw3WnHGM5P501W6c/rQA3dxyeRRn0P604cmgjr/AIUANU5BGf4aCwBGD2pwIAOT29KCQMc9vSgBuQAef1pQQR1/Wjfgnr19KcjHHfrQAzrxilX7oz/dpx5PelB4H09aBkYBIJOeBRwSR/WnZO05z0oJwT160BYb3/ChMA9acW9CelIrHPX9KBADn8hSjGenQ0uemM9KUNgdeaAEX7w+tJwFyMU4N82Se9IzZXPPWgBmeRg0ikn8qduAIGaRCP09KAEyc4Gfz+lN3EE1KcY5x+VMIC5xigAU5NKwyPxpFYhuo/OnbvloAMnPUfTikBO7+tO5yOP0puDnJ/OgA3nA6fnShs9x2pMHAI/WlVSQOn5UABODxjp6UuSSen5UhH8qCCSeB+NADhnb0H5UvOccdfSkAwvAHWg9elADuT0A6U3JB5AoweoA/GnKDkHjpQBGc8Zx2pQzZ59PSlYDj5aaQT1H50AOzgcAdKa3A6ClGR2HHtTT7fyoACzc9OlJuwMnH50uMjg9qTBA60AIzHHUdfSkVskA4ps0kUMbyzzJGiKWd3YKqqBkkk8AD1Nfmr/wUY/4OMvgZ+zpdX/wo/Y903T/AIkeMYC8F34gknP9haXKMg4dCGvXB/hjKx+sh6V7GSZBm/EWMWGy+k5y69orvJvRL136XOTG4/CZfR9pXkkvxfoj9GvFfjfwf4A8PXfi/wAd+KdM0TSbCMyX2qavfx21vbqP4nkkKqo+pr4Z/aa/4OOP+Cf/AMEA+k/CG91X4r6vHJsdPDUJtdOTHUm9uFAkHoYkkB9RX4a/tPftoftQftn+Kv8AhL/2lfjDqviSRZN9npkkvlafZeggtY8RR46Z27vVj1rzMucc/pX9BcO+BmX0oRq5xWc5fyQ92Po5fE/lyn59mHG+InJxwsFFd3q/u2/M/Tr42f8AB0f+2B4nnntPgn8HPBfg+1cMIZbtZdSukBHBLuUQkc/wCvmD4gf8Fpv+ConxGdzqv7Xuv6fE3SDQbW2slX2Bjj3frXzA79sdOlMO4/wMfwr9PwHAfCGWRSoYKn6uKk/vld/ifO1s8zbEu860vk7fgj1rUP8AgoF+3dqs5ub/APbO+JzuTkkeMLhefoGAra8L/wDBT3/gon4QkWTQ/wBtD4gjb90XeutcD8pA2a8GLopw0gHsTU0Y3HcuSPavXlkGRzjyvDU2v8Ef8jkeOxy19pL72fbPww/4OF/+CpHw3u431D4y6Z4pt0zvtfEvh6CUOPdowjfrX1b8Ff8Ag6+v45obL9pH9keCaEuoudU8C655cqLnlhbXQKuQMnHnL+tfjzI+0bTx+NQs5YgDNfN5l4c8F5mmqmDhF94e4/8AyW343O/C8Q5zh3eNVtdnqvxuf1Dfsof8Fff+Cfv7ZjwaR8JPj3Y2Wvz4A8J+LIzpepBsZKrHMdk5HrC8g96+liXDEFCMdciv454VBZWPJUgqe4I6EHsa+2/2Dv8Agun+2X+xhPZ+EfEniOT4j+BYWCy+GPFV48lzaxd/sl6d0kJHZH8yPttXrX5PxH4G1qVJ1slrc9v+Xc7Jv0krK/k0v8R9Vl3G0ZVFDGQsv5l+q/yP6Pdx6nH4Um7AIyOleA/sM/8ABSn9lf8A4KBeE/7W+CHjYRa5bQCTWPBmrssOqad6logSJYwePNjLJ6kHivfQcDI/z0r8GxuAxuW4qWGxVNwqR3jJWa/rvs+h93Qr0cTSVSlJSi+qFLEHqtKpyp6daQgg8k0qjOR157muU1FHSnbsHGD+dIoyB/QUMuGyaAE3nHbpSqeOMdqbggYye9C89KQDyTxwPxoGepA4HekwAAQAeKPwHSgAGRnIHSg8np3oXJ4Awa5r4z/Fbw58C/hL4m+M3i+3uptK8LaHdarqMdjGHmeGCMyOEViAWwpwCRz3FXSpzrVI04K7bSS7t7EzlGEHKWyOk3Y4wPzpdxznivzPm/4Omv2CtgktPhB8WJgy5H/EjsF6/W8qBP8Ag6h/YZ3fP8DPiwBjr/Zenf8AyZX2v/EN+OLX+oT+5f5nj/6xZMnZ1l+P+R+nAPc4696d93jjgV+Zcf8AwdQ/sGhct8F/iwPb+x9P/wDkynj/AIOo/wBgpjj/AIUv8WP/AAUad/8AJlR/xDrjb/oBqfcv8w/1hyb/AJ/L8f8AI/TAHnPH0xSk+tfmTN/wdQfsMjJtvgb8WJPTOm6cv/t5UUH/AAdQ/sVvJtm+AXxSRSfvfZdOP6C6p/8AEOOOLX+oT+5f5ifEeSr/AJfI/Ttd3cUEHPT9a/PzwD/wcv8A/BM/xZMkHifUPHfhYsQDJrHhJpY0z3LWrynH4V9Yfs9ftwfsiftYWq3H7Ov7RHhXxVK0e9tP0/U1W9jH+3aybZl/FBXi5jwzxFlMXPGYSpCK6uErffa34nbh8zy/FO1KrFvtdX+49S7Dp24zTSxwc4p7IwOCpBHUEUwgha8M7hu7nmnKxPYfhTCDnnt7V5r+2n8XPGH7Pv7HfxU+Ovw/S0fXfB3w+1XWdHXUIPNga5trZ5Y/MQEbk3KMjIyM8itsPQnia8aMN5NJerdkE5RpwcnstT1EAAd/ypsnspOK/n7tf+Doj/gpIkKNN4V+FcmVBJHhO6Hb2vKjvv8Ag6E/4KSSoTF4d+F0XHbwjO3/AKFd1+qf8QW426xp/wDgf/APlnxhk17Xl93/AAT+gN856Dpjk0qE4ycflXmv7GPxT8afHb9kP4X/ABs+I4shr/i/wFpms6uNOtTDbi4uIFlcRpubYo3YAyfqetembcDoOvrX5VXozw1edGe8W0/VOzPp4TjUgprZq/3i9+g5PpQck9vejHIxijH04x1rIoQYz26UHBxnFOAPTPakYGgBCcDJx+VCk98UjAk5pQPYUAKSTwcdu1LlvQdPSk9KX147UAGT1GPfmmMeM8HGKdjngim0AIM9AB970pFLZ7flTtpIHA696RUx1A59qAFBOAOOlOBwB+HU0m3gcdvWnDnAx3oARjxzihiTknHWgjihu+fXtQO+gAnGOOlKCeuB2700dBwPzoHHVR09KBCkkjJoB5HPSkJPoM0qnOP60AID04FO5OeKacg9cUE46igBQSQenSkO7OQB09KBuwen50hU/wB0dPSgBM8HOKcnTt1pBk5GB+dOTpQAo9v50ZOBwOnrQeDwO3PNIegz6UDGqTyOOlISTwcfWnAZyMDp3pNvPIH5UCEBPPT86EJyc4pdn0PFCrycAfWgBdxPp0FLk57dOaCDntxilx7Dp6UANDHPUcGkZjjgDr0pcc8Y69aaRkduvpQAZJ5wKF/Dp6UgU5HA6U5RjsOlAB27dKTJOenFPwSAcfnTcHByf0oAaMgnH8qcWIB57UmOefzxSOOCB6UAS4J4Pp3NIFYkfL+tKGX15pMqGHP40ABVsj5f1oAbgY7+tG5eADRkdDn8qAEYEH7valbIzhTnHrSHGeAelBwCeO1AApYpyppxyW4U9fWmZ+XOD+VKTk556+tADiCOg7etKuQRx+tJuAyAaAw3UADK2Mle1IwOOnp3pGYcEEdPShnB4J6UDEyxPIH500lum3tTgRgHmmuAQOtACgkLkjrWP8RPiJ4E+EfgTVfid8T/ABZY6F4e0Sze61bVtSmEcNtCvVmJ79goyWJAAJIFXdd8QaF4S0G98U+J9XttP03TLOS61C/vZhHDbwRqWeR2OAFVQST7V/Op/wAFkP8Agrj4w/4KD/EWX4a/Di9udL+Efh7UGOh6aGKPrsyEgajdLnnIz5UR4jU5Pzkkfa8E8F47jPM/YUvdpQs6k/5V2XeT6L5vRHi51nNDJ8Nzy1k/hXf/AICOi/4K5/8ABcr4i/tu39/8Df2dr7UvCnwnilaK5ZZDDf8AigA433O05itz1FuDyDmTP3R+fEYCgKuAB0ApJBjq34V3H7OP7N3xu/a0+K9h8FP2ffAN54i8Q6gdy29soWK2iBw09xK3yQQrnl3IHYZJAP8AXuV5RknCOVexwyVKlBXlJvfvKUnu/N/LTQ/JcTisbm+J5ptyk9l+iRxyOoYBmxjkkmvqL9jf/gkH+3X+29bQ+IPhZ8K/7H8MSuAfGXi+VrDTiuRlosqZbrA/54oy9iwr9av+Ccn/AAb5/s1fsmW9l8R/2j7bT/if8QkCSj7daFtF0iUAHbbW8g/fsD/y2mBJIyqR1+hkQjgt0t4o1RI0CxxquFVQOAAOAB6CvyDijxvhQcsPkVNSa09rNaf9uw0b8nK3+Fo+tyzgnntUxsrf3V+r/wAvvPy0/Z4/4Nbf2a/CVpban+0v8a/EnjLUfLBurHQFXSrEP3Cn95Mw99yk+g6D6r8B/wDBFj/glz8O0jTSf2NvCuoyJg+f4le41NmI7kXErL+mK+oCw4xn6CkL8Z6V+LZjxvxbms3LE42o79FJxX/gMbL8D7PDZNleFX7ujH5q7+9nl2m/sJ/sO6TbLaad+xd8JIY1UAInw603p9TCawPG3/BMf/gnX8QLdoPE37EHwwbf1ksfCNvZyfg9sI2H4GvcGZeT+VAfOMZNeLTzXM6VTnhXmpd1OSf33OyWGw0lZwVvRHwX8av+DcT/AIJq/E+KebwX4S8S+Bbp1Ply+HfEks0SHHXyrvzRjPYEV8O/tK/8Gtv7SXguO51v9l34z6D43towzxaHr8f9lX7jnCpIS8Dt2+Zox71+6hKnsfyoGwcc19blXiZxnlMly4qVRdqnvp/N+990keTiuHMoxW9JRfeOn5afgfyJfGT4D/Gr9m/xnJ8O/j18LNd8I63EMnTte09oGkX+/GT8sqf7aFlPY1yxfcMfrX9bn7RX7MvwE/ay+HE/wr/aG+GOmeKNFm+aOC/i/eWsnaWCVcSQSDsyMD26cV+IH/BUP/g32+Kf7J9lqPxv/ZVuNS8c/D2DfcajpMkXmaxoMQJJZ1Qf6XAo6yoA6j764Bev33gvxdyjP5xwmYJUK70Tv7kn5N/C32l6KTeh8Nm/CWJwSdWg+eC+9f5/I/PfwF8RfHnwn8Zaf8Rvhj4w1HQNe0e5W40zWNJu2hnt5BzlWU59iDwRkEEHFfvN/wAEf/8Agun4P/bJGn/s6ftQ3en+HviqEEOlamqrBY+KyB/yzHSC8wMmH7snJj7xr/P8GDchgRjgg8GljubuyuY7+wu5beeCRZYJ4JCjxOp3K6spBVgQCCOQRkV9dxlwTlHGOB9niFy1Yr3Ki+KPr3j3T+VnqeZlGcYrKK94O8XvHo/8mf2MhjnAx70oLHOB2r44/wCCFX7T/wC03+1v+wlpnxB/ad8K3Kahp2pSaZoni65YK3iqxiAAvWTqJFcNE8n3ZWjLDndX2SIjnYASScDHc1/FGbZfXyjMq2CrNOVOTi2ndNrs/wCmtnqfseGrwxOHjVjtJX1I5bmC1ge5upkijjQtJJIwVUUDJJJ4AA5J7V8oeKP+C6f/AASh8J+PW+HOrftkaI96l19nmvdP0fUbvToZM4+a9ht2twoPVw5UdyBzX58f8HAP/BZyP4g3ep/sG/sneL92gW8rW/xI8W6ZccanKpw2mW8inmBSMTODiRh5Y+QNv/I+MhcYPQcAdq/XuD/CP+18sWNzWc6fPrCMbJ26SlzJ79Fa9tW9bL5HNeK/quI9lhoqVt29vRW/M/sc8NeJvDXjXw9ZeLfB/iCx1bStStluNO1PTbtJ7e6hcArJHIhKupB4IJFXVzzwevc1+S//AAai/HLx34o+EPxW/Z71/U7i60Lwbqemap4cWZywsjf/AGpbmBM/dQvbJKFHAZ5D/Ea/WjABr8t4myOpw5ntfLpy5vZtWfdNKSdujs1ddz6fLsZHMMFDERVubp+DFyxHA7etBDZOB2pMjP3jxShgR+FeJY7RVBORtPWvFP8AgpLv/wCHfnxpOOnwx1rGf+vOSva1xuNeKf8ABSg4/wCCfPxr2n/mmGtf+kklejktv7Yw3/XyH/pSObG/7lU/wv8AI/lTtUU2sOAP9Sv8qV0GKLM5tYuP+WS/yp7KSeK/0Iu5H4FLSTKroT0FIowwyeK+2P2Xv+CC37b/AO2D8CNB/aL+E194Li8P+I4ZZNOGseIXgnKxyvE25BC235o2xz0xXdv/AMGw3/BSWNgv9q/DkjPUeLXH84K+Jr8d8IYbESo1MbTUotppy2admvkz2aWR5vVpqcaMrPVaH55xgbQTTgCe1fokn/BsP/wUUU8+IPhyc9ceK34/8gVLf/8ABsT/AMFFrWyNzY638PLuQLkW8Xil1Y+2WgA/Wqj4hcFPfH0//AiZcP50n/Al9x+darip7DUtR0jUYNZ0i/uLS8tXD2t5aTtFNCwOQyOhDKR6g19G/tGf8Eif+Chf7Lei3Xij4mfs1a3Po1nzc614fCalbxJzl3NszsqgDlmUAd8V81o6ugdHDA9CDkGvp8vzTKs3oOpgq0Ksdm4yUl6OzZ5tfDYvB1LVoOD800foH+wf/wAHDv7XP7MWo2Pg/wDaFvLn4qeCUKxzR6rOBrNlF03QXbf64gfwTbs4xvXrX7qfs0ftNfBT9r74PaZ8c/gD40h1vw/qYKrKg2TWsy48y2uIj80MyEgMjeoIyCCf5JmOSRX1P/wSU/4KR+MP+Cdn7S1l4hvdUuZvh54luYbP4g6EpLI1vnat9GpOBcQZLAjl03xnhhj8e8RPCzLs1wk8dlNNU8RG7cY6Rqd1ZaKXZrd6Pe6+w4d4oxGHqxoYuXNB6Xe8fn2P6bdhJ4H1FeG/8FQMj/gmz8fPQ/CDXx+dlJXuGn6lpWtaVa67omow3djfW0dxZXdu4aOeF1DJIrDhlZSCCOxrw7/gqFx/wTa+PnP/ADSLXv8A0jkr+ZMnT/tfDp/8/If+lI/S8Y19UqP+6/yP5V4QRGoPZR3ps4+U8dRT0GFGB2FJIMofpX+getz8CW5/V1/wTvtmtP8Agn98Dbcj7vwj8P8AB/68Iq9kGSOR39a8l/YKZR+wl8EsdP8AhUvh7H/gvhr1jcu3A9a/z5zR3zSu/wC/P/0pn75hf92h6L8h2CTkD9aQ5HRe3rQT82efzoyOpJ6etcBuCgg8Dt60MCAPl7Uq4yD/AFprtggBjzTQDWBx07DqaXB67T7UhYcc9T2FKCD1JoYB82QcdhS8gHj+HuaaxBUc+lKCMHntSAAG4O39aQg4Bx2pQQf4aQ4wBj9aAFTgAY7/AN6gE7vu9vWkUgAAZ6+tIHGepoAfg7QAvYZ5pcHqF/Oo9/AG7tUiEED5qAEOcHg9KGDHPHf1pWxtPPamyEc896AAbsDj9aQbiOmePWgOMDrQpXjk9KAEYt1x1PrToy2Bx355pDggDNCYG3B702ApDcnFGGPRf1pMg8j19KDg9u1IATIGCp6etBycYX680LtAz/s0ErkDnpQApVucL39aUbtudtJ8vOM9R2oBAB/OgBSWJ6H86QbmUHHb1pu4Z4pVIIHJoGKoYg4XtS4JP3e/HNIpGD9KXKg4PrQAbW9O3rSBTz8vb1pdwPIz07UgI557etAhzZJzjt60KGPQdqaW54z93ilDDt6UAAU5Py9+xpHU44Xv60uRnknrQSoBPPJpoAAOR8p6etAUnnaelA2549PQ0Aj9KGAYJ7d/WkIODx2Hel3Ck3DkY7ikA35t3CmlYNg/KenpRxuJH8qQlSPwoAeH4xSbskHjNJg56UmMNx6UDsO3dPlHSng4/hPT1pgXHY9KUjjgfpQIUt/sjp60hPOcdv71NY9R7UinJ/KgBwYbc7e/rQWG7oOvc0mOMCkIwce9Ax5bHp+dIGO7OP1pMZIOKAATjFACMwJHyjntmlB/2e3rSFenHagKCMYzxQFxQwIwB29aCFzkgUYxxXiH/BRb9sjw9+wd+yB4u/aN1dY5tQ0+1Fn4X06Q/wDH9q9xlLWHH90PmR/SOJz2rowmFr47FQw1CN5zajFd23ZGdWtCjSlUnokrv5H5o/8AByL/AMFM9S1DxE3/AATy+C/iEw6dYrFc/FG+tZebq4OHh0vI/gRSsso/iZkQ/cYH8hWkJ+Ynr7Vd8VeLfEfjvxRqXjjxlrU2o6vrN/NfatqFw26S5uZXMkkjH1ZmJ/Gm+G/DmveMvEen+D/CekT6hqurX0Vnpun2qbpbm4lcJHEg7szMoH1r+5+E+HcFwjkVPBU7e6rzltzS+1J+Xbskl0PxHNcwr5tjnVl10S7Loj0r9jH9jL40ft4/HbTPgJ8EdJV7y6/f6tq9yjfZNHslYCS7uGHRFyAF+87EIvJ4/pZ/YR/YJ/Z//wCCe/wYt/hL8EPDy/aZkSTxJ4ou4l+367dhcGedx0XOdkQ+SNTgDOSeN/4JSf8ABOPwn/wTl/ZntPAjw2134619Yr/4ga7CufPvNvy2sbdfs8AYog/iO+QgF8D6dIPHFfzF4k8f1+K8e8JhZNYSm9F/z8a+3Ly/lXRavV6fpfDmRU8sw6q1FerLfy8l+oAqSTsHtmlLcDjtSAAGgKSuc84r8uufTDsnGMdvWmnheg6CnEU18Y/CkAxyOTtHWlU5wduf/wBVDAc4HegKAM0AA7HbzgdaGPzY29PelCjtjpSMoz+dACZBA4oyQxYZH0oUYxjsKCoHbvRcND8d/wDgul/wRH0mLRta/bc/Y28HJbSWqyX/AMQ/A2mQYjeMZaXUrKNfuleWmhUYK5kQAhlb4x/4I4/8Et/EX/BR345C/wDFsFzZ/C7wpcRTeMtYhJQ3jH5o9Ot3/wCesoGWYf6uPLfeZAf6U349D7EZBrnPgh8EPhP+zx4GT4a/BHwBpnhrQkvrm8XS9Lh2RCeeVpZZDkkks7HqeAFUYVQB+r5b4u5/l3C1TKn71XSNOq3rCL3v3a2g+l9fhR8vX4VwOIzJYnaO7j0b/RPqdH4U8K+GvAvhbT/Bfg/QrTS9H0mxis9M02xhEcNpbxoESJFHCqqgAfSvy5/4OB/+CwsHwM8Laj+xF+zF4sH/AAnGr2xh8da/p0+G8P2ci82cbr0u5VI3Y5ijPZnXb6r/AMFuv+CwWj/sF+An+BXwQ1i3u/i94ksN1uU2unhizkBAvZl6ee3PkxHuPMYbVAf+dzX9a1bxJq11r+vapcXt9fXL3F7e3czSS3EzsWeR3Yks7MSSx5JJJr3PCzw7lm1SOd5pH90neEX/AMvH/M7/AGU9v5n5LXh4m4gWGTweGfvbSfZdl5/l+WVsP/6q2vhz8PvGvxb8e6R8MPhr4autZ1/XtQisdI0uyj3S3M8jbVRR9eSTgAAkkAE1lRW9zd3SWdhbSzTTSLHDBBGXeR2OAqqOWJJAAHJzX9CP/BDD/gjlZfsR+DLf9pf9oLRYZ/i14i04G2sZkDDwpZSjm3Q/8/Ui4Erj7o/drxvLfsvGvF2C4Ryx16nvVJXVOH8z8+0V1fy3aPk8pyutm2I5I6RW77f8Hse9f8EnP+Cdfh7/AIJvfsuW3wwmu7fUPGOvTJqfj3WbflLi+2BVt4ieTBAvyITjcTI+AZCB9O5Xrjv60E8DjtTQSeK/jPMMfis0x1TF4mXNUm7t+b/RbJdFofruHw9LC0I0qatGOiFDc52npS7vVe1NABPTt1pSB6Vyamw5WGfuj24rxb/gpD83/BP/AONaKoOfhdrfH/bnJXs46nPpXjf/AAUUTd+wT8Z1H/RMNb/9IpK9DJ/+Rth/8cP/AEpHNjLfVKn+F/kfynWZ22cOf+eS/wAqkDc5xTLdCLWL2iX+Qpa/0Ii7H4BJ3kz+l3/ggteif/gk/wDCeIN/qrK+Q8dD9vnP9a+u5W56Cvi3/g36u2m/4JV/DpCf9VNqSDPtey19nlyxxzX8DcXQ5OK8cv8Ap9U/9LZ+85VLmyyi/wC7H8kNBGc8Uolx0PfFMzxg0ZO78a+fO4nSYqcoccdjivyk/wCC/v8AwSH+HviX4Uax+3P+zN4EttJ8UeHwb3x/omj2wjh1qxz+9vliXCrcw/fcqB5se8tllUn9V0B5zUWraFp3irRb3wprVss1nqlnLaXcTjIeOVCjAj0IY17vDnEGYcM5rTx+Ek04vVX0lHrFrqmvueq1Rw5jl9DMsLKjVV77eT6NH8d7MSNw7+9MIJUr2NdX8cvAMnwq+Mniv4ZyAA+H/Ed7p2Ac8QTvH/7LXK4OOK/vKjWjiaMasdpJNej1PwmcXTqNPof0U/8ABud+0zqPx7/4J16d4I8Sao91q3wz1yfw3M0rEv8AYwq3FkTnqBDL5YP/AExx2r3n/gqOwT/gmr8fSi8/8Kh1/Az/ANOUlfnH/wAGnHiK5h1T45eDCB5Elr4f1BeORIsl9CT+Ksv5V+jn/BUIf8a3fj0CeP8AhUevf+kclfxxxZltPK/EupQh8LrU5L/t/ln+DlY/Y8rxMsTw5GpLfka+66/Q/ldiQ+Wu7+6P5USKAhI9KWM/KAfQUrjKke1f2ly6n41d8x/V5+wPIG/YP+CJ9fhH4e/9N8Nesgkj7vf1ryH/AIJ/5H7B3wRz/wBEk8P/APpBFXrpwBxX+emaaZnX/wAc/wD0pn9AYb/d4ei/Idu5+7nn1p+QOSuKYM54FeGftyf8FGf2X/8Agn14Ih8VfHzxe/8AaWoxudA8J6RGs+p6sV4PlRZASMHhppCsanjJOBWWCwWLzHFRw2FpudSWijFXb+X9WHWr0cPTdSpJJLqz3gMM9Mcc5qObhd23j121+BX7V3/BzH+2n8Wr+40f9m/QdJ+FuhsSsM8MSalqzr2L3Ey+VGfaOLj+8etfE/xP/bZ/bJ+M92978Uv2q/iJrRc5aK68X3axD6Ro6oB9AK/Ycq8DuJsXTVTGVYUb9NZyXrb3fukz5LFcaZbSly0oufnsvx1/A/rFN7ZqQhuo8+hkGanjdSvycjgZHNfx5t408dPL57ePvEG/Od/9vXOfz312/wAMv2yf2vPg1qEeofDL9qX4g6Q0bArFB4runi49Y3dlP4ivWreAuMVNuljot+cGvx5n+Ryw44oN+9Sf3n9bLDnAWkI9u3rX4L/spf8ABz1+1d8Krmz0X9p7wfpnxK0SPCXN5bwpp2rov99JUHkykf3XjG7H3161+wf7Fn7f/wCy9/wUA+HcnxC/Zx8eC+azKprfh/UIxb6npEjDhLm3ySoODtkUtG+DtY4Nfl3EvAnEnCr5sZSvT6Tj70fm9Gv+3kr9Ln02W53l+aL9zK0uz0f/AAfkex5wOn60jcrnA7d6cR70jDIr489YjzgjKjr1zSh0DE5FfLH/AAWZ/at+Mf7Ff7CesfHn4Da7a6f4ks/EelWdtc3mnx3Uflzz7JAY5AVOVyM9R2r8fX/4OO/+CpcnT4reG1908E2Q/wDZa++4Y8Oc+4sy94zBSpqCk4+9KSd0k+kZd+54WZ8Q4HKq6pVlJtq+iX+aP6KgBgHA49qkUgcba/Db/gnN/wAF0/8AgoX+0J+3P8L/AIHfF74maPe+GvFHimOw1e2t/C1nbvJE0Uh2iREDL8wXkHPFfsd+0D+0V8Gv2VPhNqfxu+P/AI8s/DvhvSYx9pvbs5eWU52QQxj5ppnIwsaAsT7AkeZxHwZnPDOZU8DiUp1KiTioXle7aS1ine62SOjLs4weZYaVendRi7O9l+rO8ONvCZ4qO6YRjdJwPVuK/Bv9tf8A4OYf2mvi9ql54W/ZA0hPhr4XDtHb6xcRR3Ot3iZ4kZmDRWueuxAzDP8ArDXwN8Q/2rP2ovixqUmrfEn9o7x5rU8rEub/AMXXjLz6KJAo/AV+hZL4HcRY+gquNrQoX+zZzkvVKyX/AIE/M8PGcZ5fQqOFKLnbrsv8/wAD+thZ4WHyyK3A6HNSK/X5eB1r+QvQvjX8bfDFyl74Z+NXjTTZo/uS2Pi29jZfpiWvqL9ln/guz/wUQ/Zm1e3GpfGG4+IGhpIDdaB46/0zzU7qlzxNETnqGIz1B6V2Zj4D51h6LnhMVCo10acL+j95ffZeZlQ42wU5pVYOK77n9KW4cDb+tAOMcd/X6V80/wDBOn/gqL+zp/wUd8Cyap8M7yTRvFelwK3iXwPqk6m8sMnHmxkYFxbk8CVRwcBgjcV9LrhsV+J5jl+NyrGTwuLpuFSLs09/+G7NaNao+woYijiqSqUpXi+oi/7v15pQRjO39aAMDHNBB7fyriNgB4PA+760m4ZHA/OgL1wf4acFJHT6UAAIwRt7+tAbK4wOnrSAHJFAzjjNACHBP3R19aFYYGR29aUjrx3pMYA47UAKpGOR27mlDZboOvrTVHGBTgOef5UAGQBjb29aQMBk/wBaUqMcelNxySM9aAAnvgfd9aUMoOMfrS7R2z09aQgA8d6AFVstyO/rSEj07/3qTHzflQoyOlAChgMcdvWlVuc7R931puDkYFIP6UAPLDPQceppN2c49u9GT1oHc0AGQD+HrSMx6be9KOvXtTZBxwe/NADjj9aTOTjH50BwRzjr3NKD83bNAxwYYFBI9qCScY/lQCeBj9KBDHUHnjp60cZPBpzkk9OlJkc9OlAw4IzQcE89sd6QPtHBFO35bj1oCwYH6UL1/CnEn07UgOCPpQIRgvB+lIQvb09KcxzjjtQT1wR/k0AMOMZGPyr8M/8Ag6H/AGurzx18ePC37HHh3VN2keB7FdY1+CNxtk1a6TEQYDvFbHAz0Ny9fuVc3lrY28l7fSKkMMbSTSMeFRRlj9AAa/ks/a2+OF5+01+1D8QP2gL2+e4XxZ4uvr6ykkGD9kMpW2XHbEKxjHoK/ZvBLJIZhxLUx1RXjh43X+OeifyXN87Hx/GeNlh8ujRi9ZvX0Wv52PNySOf0r9R/+DY/9hi1+LPx41v9tj4haN5+kfDthYeEI548xy61NGS9wAept4G49HnVuqCvy4uSkMDzvnailmwOw5Nf1O/8EqP2XT+x9+wJ8N/g3qOmLba1/YSat4oULhjqd7/pE6t6lC6w59IhX6t4w8QTybhf6tSlapiXyefIleb/ACj/ANvHy/COAWMzH2sl7sNfn0/z+R9DnG0cd+4phIOP8aV5OOOOab5mcDPev5DP1gTv0NGeAKAwDZBAppc7RQMeTgZxSFieoIpFcnH06UZOOcUDSFbByffrigEDt+lDtnOPX1oDY/KgQZB6Dr7UEZ5Cn8RQG54A6+lKWJxgdqBCAcg4/Sgrkn6+lKG6c9x3pV5J6YoAjaPjAFfIv/BW/wD4Kq+B/wDgmx8GUTREtNY+Jnie3kTwb4elbKQgfK1/cgciCNui8GRxtHG4r6/+3R+2x8HP2Bv2ftU+O3xf1BX8lGg0DQopgtzrV+VJjtYQfU8s/REBY9AD/L5+1N+0z8Vv2v8A4567+0H8Z9dN7ruvXJdkQkQ2UA4itYFJ+SKNcKo+pOWYk/q3hl4fy4qxv1zGxawtN6/9PJfyryX2n8urt8txJnyy2l7Gi/3kv/JV39exzHxK+JXj34x/EHWPit8UPFN3rfiLX7+S91jVb2TdLczucsx9BwAFGAoAUAAAVgscgluAB36AUkrleQenav1M/wCCAX/BHi6+P2v6f+3D+0/4PRvAGl3Hm+CPDuow5HiK8RsC7kQ9bSJ1O0HiaRe6Id39M8QZ9lXCeUSxWI92EFaMVo5O2kYr+klrsj86wGBxOaYtU4at7vt5s9j/AOCAP/BGu28D6fpX7en7VvhDdr12i3Xw18K6nb/8g2EjKapPG3SdwcwowzGv7w/My7P1yDNuLNySeaXZj5m+uaT8Ohr+LuI+Icw4nzWeOxb1ey6Rj0ivJd+r1Z+w5fgKGXYZUaS9X3fdjlYFRk80LgjihWwB0/OhG4J4/KvCR2i8f3e3XFBIxwvb0oLDdkEUbhwcdqroAgwc143/AMFFJlg/YK+M0pAwPhjrfX/rzkFeyq27g+teKf8ABSIkfsAfGgAjP/CsdZ7/APTo9d+T65th/wDHD/0pHPjP9zqf4X+R/KrbsGtoweP3S/yFKOTTYP8Aj3jH/TNf5U8HnJr/AEEjrI/AJfEz+j7/AIN8oXg/4JWeAC38d5qjDPp9tl/wr7Tycj6elfGn/BADY3/BKj4buo6vqec/9f8APX2YVUHgDpX8G8Yu/FmP/wCv1T/0tn7rk/8AyKqH+GP5DBjP/wBalBGc0L39vSgZz2r5xanpEsRGMUXGqWOh2c+uanOsVtZwPPcSscBI0UszEnsADQgK5BAGfUV+bn/BfX/gqx4H/Z2+Cut/sf8Awb8Ww3nxL8YWBstZFhOH/wCEc0yVSJnmZfuXEqZSOL7wDmQ4AXd7GQZHjuIs1pYDCRblNpPtFdZPsktX/mcePxtHAYWVao7W/F9j8Pf2h/HkfxT+PPjP4lREbPEHirUNRjweNs1zJIv6MK46mMwIAHQDAFKuWO3I+tf37Rw9PDUY0obRSS9Fofg826k3J9T9kf8Ag038JzPF8c/HUkZEQOgabExHDPm9mcfgDH/30K/RL/gqNFv/AOCa/wAfFUcn4Ra8f/JOQ141/wAG/H7Muofs1f8ABObQNR8QWTQ6v8RdTn8W38cq4ZIZ1SK0Q/8AbtDE/wBZDXtv/BS1Bcf8E6PjxARjPwi8QcH/AK8ZTX8YcWZnSzXxKq4mm7x9tCK81Bxhf58tz9kyzDSwnDkact+Rv77v9T+VOINgfQVIy4HPpTkXag6dBQ/CnntX9sKyep+LuV2f1afsA4P7B3wSb/qkvh7/ANIIa9cJXHJ6147/AME8ZWk/YE+CDtnJ+Eugdf8Arxjr2RMFtuK/zvzbTNcQv78//Smf0DhtcND0X5Hj/wC3j+2L4I/YR/Zf8Q/tF+N7Zb19PjW20HRxLsbVNSlytvbg9lLZZ2/hRHbnGK/mI/aH/aI+LX7U3xg1r45/G3xZLrHiLXLnzLq4bIjhjH3LeFM4ihjHyog4A9SST+kX/B1J+0Pf6r8cfhz+y3p2qA6f4e8NP4k1a2jkGGvLuV4IN4z1SG3kIz2nJr8nGus8Z/Gv6r8GuFcLlPDcM2nFOtiLtPrGCdkl625n3uux+W8XZlWxWYPDRfuQ6d31b/IsOFYZGOe1ej/Ar9jD9rP9qCCa7/Z3/Zx8YeMbWGUxzajo2jO1nHIOqG5fbDuHdd+R6V7T/wAEWv2EdI/b+/bHs/Avj61kl8F+GLBtb8YpG5Q3FsjqkdqGBBXzpXVSQQQgkIIIFf0q+F/DHhfwR4csfBvgrw5YaRo+l2y2+maTplokFtaQqMLHHGgCooA6ACq8Q/FNcJYpYDB0VUrtKUnJvlgnsmlq297XVlZ63Fw9wx/atJ160nGF7K27/wCAfzI33/BFL/gqvptqby4/Yi8WOoXO23urGV/++UuS2fbFeC/Fn4HfGn4DeI/+ER+N/wAIvEvhDU2JEdn4k0aazaXHUxmRQJB7qSK/r2bawwFFcf8AHn9n34O/tQfC/Ufg58d/AVj4j8PanCUmsr6IFoWwQJYX+9BKucrIhDKeQa/Pss8eM2hiY/2hhYSp9eTmjJLuuZyT9NL90fQ4jgjCOk/Y1GpdL2a/Q/kQ2GvQf2Yf2nfjH+x18a9H+P8A8CvE76brujyjfGWP2fULYnMlpcoDiSCQDDKehwwwyg10f7eX7JfiT9iX9rHxj+zf4guJLqHQdQ3aLqcqBTqGnSgSW05A4DGMhWA4Dq4HSvHWXKnNf0Q5ZdnuVxnG1SjWinqtHGS6r0Z+e3r4DFNbTg/uaP6zv2Mf2qfAP7a37NXhT9pP4cgxWPiKw33WnySBpNOvI2MdxaSEfxRyqy5/iXa3RhXpzkYIJr8b/wDg1Q+PurRyfFD9mDUtULWSra+JtItXbPlSk/ZrkqM8BlW3J46rmv2Od8j5vwr+I+MMhXDfEeIwEfhi7x/wyScfmk7PzR+0ZPjv7Sy6nXe7WvqtGfAX/Byxc+V/wS81GDI/f+PNFT64kkb+lfzwrnOa/oU/4OYdz/8ABNIhTwPiFpBb8pq/ntAr+jfBSP8AxiEv+vsvyifnnGsv+FZL+6v1Pbf+CcPxD8M/CT9u74S/FHxnrNvp2kaD46srvVdQu5NkVvbKW8yRj2AUk/hXZf8ABWD/AIKN+Ov+Ci37SV54pkv7u0+H/h64ktfAHht5CEht84N5KmcG5mwGY9VXbGOFOfmi3fy/lxx719hf8ElP+CT3jH/go/8AE+613xHf3Wh/DTwzcxr4o1+FAJrqYgMthabgQZWXDO5yIkYEgllB+6zmjkGT4t8R5g0nRp8ib1sm2/dXWcr8qtr0W7PFy+rj8VD6hh/tO/8Aw/kj45s7O71G8XTtNtJbi4fPl28EZeRvoq5JrT1T4d/EPQ7I6lrngDXrK2UZa4u9GuI4wPUsyAY/Gv6tP2ff2Q/2Y/2V/Ddr4Y+APwR8O+G4rSERi8s9Nja9mHGWluXBllY45LMf6V6XJNNNE0FxK0iMuGjkO5WHoQeDX5HiPpAQp12sPgLw7yqWb+Sg0vvZ9dS4FvC9WvZ+SuvzR/HPE8cqeZFKrqf4lYEfpTjj0zX9IP8AwUu/4Is/syftsfDXVdd+Gvw90PwT8Ure3efQ/E2iWCWkWoTgZFvfxxBUmjkPy+aR5kZIYEgMrfzl+JfDfiHwZ4l1HwZ4u0ibTtW0i/mstT0+5XEltcROUkjYeoZSPwr9V4N48yrjbBzqYeLp1IW54SabV9mmt4vXWyd1qlpf5jOsjxGTVUpPmjLZ/o/M6H4C/Hz4qfsxfF7Q/jp8FvFE2j+I/D12J7G6jJ2SL0eGVekkMi5R0PDKT3wa/qa/Yo/am8Gftpfsv+EP2lPBkK28PiPTQ1/pwkDHT76M+Xc2xPU7JFYAnqu096/k0YnOa/Z7/g1I+PWqan4S+K/7M+q3zPBot9Y+ItHidydkd0HgnCjsPMhiJ9396/P/ABp4eoY7I1mkI2q0Wk33hJ2s/RtNdte57fBuYVKWMeFb92SuvJr/AIB+v/HXH6UEZ/8A1Um7t705WBOeK/ldH6cC8dj0peO360isMHjt60M4zkDHHrTACBzn+VIDxk9vajeSSaF6dB1oGB56A/lR1Axnp6U7OegFIH6D2oAap44BpQR6Hr3oBGOR+lBbk0CFJB4x29KaDnJ/pRuPbHA9aEJOT/WgB2QWH0owDxj9KXPPHp60qtz0/WgBgGGzjvzSkD079qXd82D69jSOeOB37mgBCRkcdqaAKXeMg4/U0Kec8UAHak3AAg0udw6Dp3pCeTzQNBu+f/61I5+X7vSjdlqHOQcY4HFACbWHIz19KUKd2TmnBOMAjOaQLg9RQIUJ0+lGwkjjPFAHQjFLjI5I/GgBrKewP5UhU7unb0pxUcdKChzyB0FADAjBeh4pdpLYwetG0BeAKXB3duvegBShOMA96NpyODS4HoOlIQMjAH50ABHTg9PSgg8gA9fSmntgUvXqM0AeO/8ABQr4k3vwZ/YQ+MXxN06fyrzSPhvq76fIeMXD2zxRH/v5ItfylC3S1ijtovuxIEX6AAV/TR/wXF1J9K/4JTfGGaJipn0rT7f8JNUtFP6Zr+ZqU5PJ71/UPgPhoRyPF1+sqij8oxTX/pTPzHjqq3jqVLoo3+9tfoeofsNfCGH4+/tl/C34N3lv5trr3jvTYb+PGd1qk6yzj6eXG4/Gv6xjIZHaRjyzE8+9fzWf8ECPCyeKP+Cqvw5aWHeulw6nqHH8JjspAD+bV/SiMgAZr47x1xkqvEWHw19IUr/OUnf8Io9ngiko5dOfeX5JA+MABhz6UmP8/jS/Njv+lNweBnv61+Hn2oANnv1pNpOMUAZbt+VKFOAcL+dACqpOCc9PSgoQOnb0pwUEAcdKTA29R09aB3YFeDxQFPZTSsCcnge5oA4GADx6UCECMTwD19KXYf8AIoHBzxS7MnFACLHkVw/7SX7SPwj/AGRvgvrnx9+N/idNL0DQ4N8z9ZbmY5EdvCnWSaRvlVB1PPABI6bxx438G/DHwZqnxF+IXiSz0fQ9FsZLzVtUv5hHDbQIMs7E9gO3UnAGSa/m6/4LCf8ABT7xJ/wUa+OuzwzdXlh8MPClxJF4L0SbKG5bo+o3CZ/10g4UH/Vx4UclyfuuA+CcXxlmns9Y0IWdSfZfyr+8+nZavs/DzzOqOUYa+83sv1fkebf8FFf+CgPxa/4KJ/H+8+L/AMQ5JLHSLTfbeEPC0dwXg0ax3ZVB2aZsBpJP4m4HyqoHz83ykk/lUzhurHOTwK+lP+CXv/BNL4n/APBSj49R+BdDe40rwbojR3HjnxaIsrp9uSSsERPyvcy4IROw3O3yqc/2BUqZPwpkd5WpYejH7kui7t/e2+7PyeKxmbY6y96c3/XyR6P/AMEWf+CReu/8FE/iofib8VLO4sfg/wCE79Brt0Mo+vXS4YabbsMcYIM0g+4h2j5nBX+jvRdH0DwxoNn4a8MaRa6dpmm2kdrYWFlCscNtBGoVI0RRhVVQAAOABWB8Evgj8L/2dfhNofwQ+DHhW30Xwz4dsVtdL0+3H3VHJd2PLyOxLu5+ZmYk8mvif/gtt/wV90X9hnwDN8A/gVrtvcfF7xFYZikjKyL4Ws5AR9tmByPtDD/URN3/AHrDaoD/AMmZ5nOeeKHE8KGFg+W9qcOkY9ZS8+sn00S6X/U8HhMFw1lrnUevV932X6H0dbft+fAjWP25o/2A/DOqHU/GEPhi81jXZrSVTBpLQ+SUs5D/ABXDpKZCg/1aoN3LgV7b5WRxX8zv/BGD4vav4P8A+Crnwt8W6zrNzcT+IvEN1p2q3l1cNJLdPewTIzyOxy7NIysWYkluTzX9McfKD+Wa5vEHg+lwZmdDCU5ualSjJyfWV5KVl0Wisu3fc0yDNJ5thp1ZK1pNJeVlYYUOOM9e1NCEA5z3qUKCBkDPvTCp28Y6etfAnuiAMT3/AC+lJsYc8/dpSPm5A/E0bT046UACpyfr3rxT/gpVlf8Agn38aijDj4Zazj/wEevbVX2H5V4l/wAFKkdv+CfHxrVTz/wq/Wuf+3SSvSyfTN8P/jh/6Ujmxv8AudT/AAv8j+Va3YG2iPrEv8qVnH4VBYsWsYCT/wAsV/kKkfp1xX+giZ+CNWmz9Jv2Af8Ag4Zg/YR/ZP8ADH7MEX7HbeKz4fa6Z9cb4g/YROZrmSbiH7DLswHC/fOcZ4zivaE/4O1bWQ/vP+Ce8i/7vxXB/wDcZX42OcHB7U3cB3r89xvhhwTj8XUxVfC3nOTlJ+0qK7k7t2U0lq9kfQUOJM4oUY04VLJKy0Wy+R+zsX/B2fohHz/8E/rkZ9Pikp/9xtVNZ/4OyneykHhv9gVIrojEUuo/FAvEp9WSPTwWHsGHTrX42iT0P6UplOOtcy8J+Av+gT/ypV/+TNf9ac9/5+f+Sr/I++P2mv8Ag4w/4KGfHrSLrwx4N8QaJ8ONMu4jHJH4KsmS7Kkcj7XO8kqnrynl18HXmo6jq2oXGr6vqE93eXczTXV1czNJLNIxyzu7ElmJOSSSSetVTMO5/WlWQEbjwB1JPAr7HJciybh+m6eX0I0097LV+r3fzbPJxWNxuOlzV5uXr+iLAOThTX1D/wAEk/8Agnh4q/4KI/tR2Pgu5sLiLwJ4dljv/iDrKAqsNmGytojdPOuCDGoHIXe/RK6P/gnN/wAEWP2rf2+NSsvFV1oF14H+HEkga68ca9ZtH9qizyNPgfDXbkEgSDEIIOXJG2v6D/2Uv2SfgT+xT8GNO+BP7PvhBNL0az/e3c8h33WpXRAD3VzLjMszYGScAABVCqAB+deJHidgMkwk8BltRTxUk03F3VPu29ubsuj1dtn9Hw9wzXxlWNfExtTWtnvL/gHoFjp+naTYW+j6PYRWlnZ26QWdrAgVIYkUKiKOyhQAB2Arxr/gpMwj/wCCdvx2ZgOPhD4g/wDSCWvazkNwfyrw/wD4KZs6f8E4vj247fB/xCf/ACQlr+WcpvLNsO3/AM/If+lI/S8ZH/Y6i/uv8j+WDOAMDsKjnbEbfSkViyg+wpJSdjE+lf6EqTbP5+5bSP6uv2B7eOH9hH4JRR8AfCXw9j/wXwmvWgpA615J+wLIz/sH/BGQ9T8I/Dp/8p0Net547/ga/wA8s0v/AGnX/wAcv/Smf0Hh1/s8PRfkfG/7YH/BDX9i79tb496x+0h8Yrzxp/wkWuQWkN5/ZniUw26Jb26QRrHHsIQbUBOOrMx715Pdf8Gwf/BOyZSItd+I8R7GPxUnH/fUJr9IHYnjjp1Jrxf9qT/goP8AsYfsZQeV+0d+0DoehaiyBofD0EpvNUlBGQRZwB5gpBHzsqr719FlPFfHElTwOX4ms7JKMIOTsl0UVfRHBicvyiPNVrwiurb/AMzj/wDgnZ/wSv8A2bv+CbN94s1P4F6p4mvrnxhBZw6lN4l1KO5MUds0rIkWyNNoJlYtnOdq9MV9ODjv+Nflf8X/APg6l/ZX8L3z2HwV/Z58Z+KVTIF9q93b6ZE5z1CAyvj6gH2rwrxj/wAHXfx+v5WXwB+yZ4S0+P8AgbVtdublvxCLGK+greHniNn+JeLxVBuc7XlUnBPRJK6cr6JW1R58c/yDA0/Z05qy6JNo/cfaNoyKcucYGa/Ay4/4Omf28pj/AKF8HfhbAOwfTb6Q/wDpUKrv/wAHR3/BRF/9R8OvhMv+94cvm/8Ab4V0x8FuOZL4Kf8A4MRhLjHJoveX3f8ABLv/AAdEaFZWH7dPhjV4YVWa/wDh5b/aGAwW2XMyrn14zX5p+UD17dK90/bt/b5+NP8AwUQ+KWmfF344eHvDen6ppmiJpkMfhm0nhheJZHkDMs00p3ZcjggYA4zXiRjOORX9M8HZTjcm4YwuBxaSqU42dndbvZ+h+YZvi6WLzOrWpfDJ3R+h3/BsPfy2n/BRLWtPTO27+Gl+H/4DcWzCv39cEjtX8/8A/wAGyqhf+CkF7wOfhrqeP+/1rX9ATAjoa/mfxpio8bSt/wA+4fqfp/BknLJV/if6HwB/wcpQhv8AgmLeynqnjzRiP++3FfzzhSDkdK/oW/4OXZWh/wCCYlzHjBk8f6KDz/tSH+lfz0Fj61+veCWnB0m/+fsvyifJca/8jZf4V+pb0nTNR1zVbTQNFtvOvb+6itrOH/npNI4RF/FmFf1a/sRfsr+Dv2Mv2VPBn7Ofgy2jCaDpCHVbxUAa/wBRk/eXV05HVnmZiM9FCL0UAfy4/s4n/jIv4ehl3A+OtIBH/b5FX9c2CY1Lc55618r4847EJYLBp/u3zza7yVkvuTdvVnrcCUafLWqte9ovkQshxx+lKw5yPWhwCCSO1BXNfzkfoQeW0iEIOqnkV/Nh/wAF9fhfp/wv/wCCpnxGXSrdYYdfNjrRRAAPMubWNpG+pcMT7k1/SlDkEc1/PH/wctQ+X/wU/wBVkA+94M0U5+kBH9K/YPBKpOPGEoJ6SpSv8pQZ8jxnCLylN9JL8mfAO3PQ1+mn/Bq/NND+3X4/tEk+S4+EUrOpPUpqliQf/Hz+dfmYhzxX6Xf8GsYdv2/fGoHK/wDCnr3P/gz06v3rxJgv9Rsd/gX/AKVE+E4clJZ1Rt3P3rCknnrmlwccUpXvgUbe2RX8Sn7SNVWweD0o2ndxnpzxShSc428j+lLt6Hjp6UAG08/WlCHb07UgBGRgdRSgnb26etACFSDxnr2pApKjOTx6U4gHnI60gAKjkdKAECnGBnilKHJ4PWkA4Odv40uBkkbaAG7Djp2oReuFNO2gjovT1pFAGTx+dAD9pzkA0BD0x+OKQfe/h/OlyD6cD1oAQqc4x0NIyEjg9+1Kc5zx+dByRzjOaAG7Dx1pFU5+lO2gkYx+FIBk9unrQAgRh0B6elBiY5JB/KnBcDkDp60oXCnp2oC5Hs5Iz+tK6kjj05pSPmPT86RzkD60Bck3DgY7+lNLAtx/Kj5en9KTA3dfzoAcGBxinqwOM1H8mBgjpTgVx16UAKzDgUEqSeO1Nbb6ikJUEnP6UAKGCr0oLLv4A6+lNOCCM/pSHaWzzQBLvXnjHApGcZB9KYGTsT0pwwTkg9KBiNg4OO3pSlgRjHamkKD07Uu0YGKBHx//AMF7ZRF/wSg+KQB++dHX6/8AE0tT/Sv5p5HO7OR1r+lH/gvyWH/BKH4mFP8An50YH6f2nb1/NZL3+tf1X4Fr/jGK7/6fP/0iB+V8ba5tD/AvzkfdX/BuM6j/AIKleHCe/hXWAPxgFf0Yqe2K/nG/4NzJEi/4Kl+Ft/8AF4c1ZR9fs9f0cKR2z9K/NPG1W4xi/wDp1D85H1HBf/Ioa/vP8kKduPuj64pMrj6GieW0tbeS7u50iihjZ5pZHCqiqCWYk8AADJJ9K/G39sz/AIOifFHh74mah4I/Ym+DPhvU9C0y7aBPGHjT7RN/au04MsFrBJF5UJ52M7szLhiqZAr4Lh3hXO+K8TKhltLncVeTbSir7Xb79Fq99ND3MwzTBZXBSxErX2W7fyP2RDDcevX1pSV2gV+cf/BK3/gv94D/AG1/HVt+z9+0d4L0vwF4/wBQbb4eutNvXbSdefGfIj84l7a5wPliZnEmMK+7CH9GyBjiuPPMhzXhzHvB5hTcJrXo00+qa0a9PR6m2Cx2Gx9FVaErodkEdO1JvHYfpQAOuT+VNO3HJP5V5B1j2YHPHf0oV1OARTSFGeaF28HP6UCHrtI49KZqN9p+k2M+rarew21rawvNdXNxIEjijUFmdmY4VQASSeABT41DNtXkngADrX4j/wDBej/gsmvxY1bVf2GP2XPEwbwnZXBtviD4q0+cFdanRvm0+3dTzbIwxK4/1rKVHyKS/wBNwnwtmHF2bxwWFVlvOXSEerfn2XV/NrzM2zShlWEdapv0Xd/1ueYf8Ftf+Cv19+294vm/Z7+A2szW/wAJtAv8vdxEofFN3GcC5ccEWyn/AFUZ+9/rGGdoX89JFUccHHtxU7FW6EEe1db8BPgB8U/2ofjDovwK+CvheTV/EWvXPlWdsnCRoOXnlfpHDGuWdzwAPXAP9r5XlOTcH5GsPh7Qo01eUn17yk+73b+S0SR+NYrF4zN8Y5z1lJ6L8kjoP2Iv2KvjH+3t+0DpnwF+D1jsef8A0jXtcuIybXRLBWAkupiPTOETOZHIUdSR/Td+yF+yh8Hf2KPgTo/7P3wO0MWuk6ZHvu7uVR9p1S7YDzby4Yfflcj6KAFXCqAOH/4Jyf8ABPf4U/8ABOz4A23wn8CiLUNev9l1408VNDtm1i9C4z6rBGCVij/hXJPzMxP0CgGMqe/QV/J3iPx9X4vx/sMO3HC037q/mf8APJf+krovNs/VeHsihlOH55q9SW/l5I4r9qn4k+LPg5+y78R/i54BtbafW/C/gXVdW0iG8GYmube1klj3juoZQSO4GK/k18e+PPGnxR8bar8SviL4qvdb1/Xb6S91fWNQlLzXc8h3NI59Sew4UAAAAAD+rX9t+5jsf2IvjNdzNhIvhR4hZif+wdPX8mqIRGueyj+Vfo3gHQovDY6ryrm5oK/W1m7X7X1PneOqklVowvpZ6fNHqP7CmrT6N+3F8GtThlZWi+Keg4Knsb+FT+hr+sdGA+Udj2r+Tz9g3RJtf/bn+DWkwIS0vxQ0MgAf3b2Jz+imv6v15OfevJ8eGv7Xwff2cv8A0o7OB7/VKvqvyJlccc9qaCCOgNINuBzSpgjr+lfgZ9yLkHoPTt7U7dzx6elNwvp2Hag7R09PSq6AOUg8V4v/AMFHoxL+wF8ao1A/5JfreP8AwDkr2UFQc147/wAFD3A/YN+Mu4jB+GOtjB/685K78o/5G2H/AMcP/Skc+M/3Sp/hf5H8oenn/QYf+uK/yFTgE/nUNoAlnCvpEo/QVJvANf6CJpPU/BJq82z9EP2Dv+DfDxJ+3v8Ass6D+0zoX7XFl4UXWp7uI6Jc+BnvfJME7xE+ct5HuzsB+4MZxz1r1k/8GlPxPBwP29PD592+G9x/S/r7R/4N2JCf+CVXgYE9NV1fr/1/SmvuAsM8tjmv5M4l8TeNsv4hxeGo4q0IVJxiuSm7JSaS1hfRd9T9ay7h7J6uX0pypXbim9Xvb1PxK/4hK/iqG5/bx8N49vhxc/8AydUsX/Bpb8RiQLn9u/QguefL+HM5P63wr9rQ2T96lODXif8AEWOPH/zF/wDklP8A+ROv/VrJf+fX4v8AzPyW+Gf/AAaf/ArTL2O5+Ln7Wvi/XY1T95a6FoVrpqls/wB52nbH5H3r7F/Zq/4Irf8ABNn9l+e21vwj+zrp+v61aurw6142lOrTRuOQ6JNmJGB5DKgIPSvqQeuRTs89a8PM+N+Lc2g4YnGTcXuk+VP1UbJnXQybKsNJShRV13V/zHYQABVACqFUAYAA6AegoZ+fwphIwOv5UuQR9OK+UPUTANgjFeJf8FLwz/8ABOb49hF5Pwe8RcH/ALB8te2DGRx3rxb/AIKTOsX/AATr+PLn/oj3iLP/AIL5q9DKP+Rrh/8AHD/0pGGMf+yVP8L/ACP5VEwqLxxikuMeUxB/hoDDYoz0FNnOY2Gf4a/0KXxH8/L4j+rz9glVX9g/4IKMcfCDw3/6bYK9aRd4HfNeTfsBgf8ADB/wR4H/ACSHw4Bn/sGwU39vT9oJ/wBlP9jX4kftAWTqt74c8K3M+lh+hvHAit//ACK6H8K/z6xWFrYvO54emrynUcV6uVl+J+/wrQo4JVJbKN/uR+eH/Bbv/guN4n+EXivU/wBjn9jDxOLLXrAtbeN/HdoQ0mnS9GsbI8gTL0kn6ofkTDBmX8WNWvL7XNWutf1vUbi9v76dpr6+vJ2lnuZWOWkkkclnYnksxJNO1PVdT1vUbjWdYvpLm8vJnnu7mZyzzSuSzuxPJYsSSfU1XZsDcx6V/bPCnCOVcJZZHDYWC57Lnnb3pvq2+3ZbJH4tmea4zNMS6lSWnRdEivchYU3s20DvnitPwd4A8f8AxBultPAPgTWtblY4WLSNLmuWP4Rqa/az/gjZ/wAEHPgCvwY8O/tWfto+B4fF/iTxLaRan4f8H6uC2m6NaP8ANC80PAubh1IciTMcYZVClgWr9UfCfg3wf4D06LR/A/hPS9FtIUCxWukafFbRoo6ALGqgCvzLiPxoy7KsfUwuAoOs4NpycuWN1o7aNtJ9dL9NNT6bLuD6+Jw8atafLzK9rXf6H8rGi/8ABOr/AIKCa1bpd6R+w58XLmJxlJIfh5qJU/j5Nbll/wAEvv8Ago/c8L+wj8WVz/z08C3if+hIK/qiFzKVAMzn6yGmqQG39/Uk18p/xHrPk/dwlL5uf+aPU/1GwDWtWX4H8nHx7/ZD/aW/ZWm0m1/aL+Dmr+D5dcgkm0qDWURJJ0RgrsEVyy4JA+YDrXnjIAOlfqH/AMHT/jK1v/2tfh54QhuFaTTPh3586K3Kma8l25HbIjr8uzKAMtX9A8IZ1iuIeG8NmOJiozqptqN7LVpWu29kup+dZzgaeAzKph6TbUXbX0R+g3/Bs5lP+Ckd0Sv/ADTXVMn/ALb2tf0Au+eAO/WvwF/4Nk4xP/wUV1SQL/q/hnqRz6ZubUV+/MgAr+ZPGr/kuJf9e4fkz9P4K0yRf4mfnr/wc0y7v+CaYTP/ADUHSD/6Nr+fGv6Cv+Dm0qn/AATVQdz8Q9IH6TV/Pop4r9e8Fv8AkjH/ANfZ/lE+T4zu82/7dR237NKq37SXw6Vuh8faPn/wNir+uVDtiT6Cv5Fv2c5RB+0Z8PJSfu+PdHPT/p9ir+uqPmFD/sj+VfC+PP8Av2C/wz/NHt8CaUa3qv1GMAF6fpQQOcAce1OYIaR1UE49a/n8++FQ4x9RX883/BzCyH/gpvqIGMjwXo4P/flv8a/oZTaBX87n/ByvMX/4Kf62v/PPwlow/wDJVT/Wv1/wSX/GZv8A69T/ADifJ8Z/8idf44/kz4EDAA81+nH/AAatQl/27vHtwB9z4QXQ/PVNP/wr8wd/r61+oP8AwamyB/24PiIDjj4Rzf8Ap0sK/fvEy/8AqJjv8C/9LifDcOQ/4W6Pr+h+8YcHj+lLvUcqOM9cVHuXP4+lLxjOe1fxEfso5HHTHb0oYrkEDt6UwFcE89KCy4FADmZeRx1FAkB4A6H0+lMJyT16+lC7cd6NQJdwzz60m4YH09ab8vr+lLgYGPT0oAFYYP8AhTgwBOB39KaoHJxSjbngHrQAFhj7v8PpSArySo/KghcfhSDbknP6UAOLAtkelKJBwcdvSmnBPHp6UHbjFADt43ZpHKkcep7U35S2M9KX5ewPWgBS4457elCPzn264pMLnv0pBjtQBIGBAwB1oLDnGO3QU0FegP6UqhcH/CgBGPzfj1xTWYEY9O9KQueD+lI4Hb0oATcPXvQME5zSkAAHI/Ok+XOSRQAuV4APb0oJXrn8xSZHcjrQCOgYUAOG3Iy1KdoPB7etJwCAGH50vGSQR0HQ0AMIXbwf0pGCk43elOIVh94daQou7lh24oHcMDnk9KVQAaXYuD0/OhQu7HH50CDCnGMdKftU8DHFM445FPBU9Me1AHx9/wAF77fzf+CT/wAUwByh0dvy1W1r+aScEHr3r+mH/gvMVH/BJ34sZ/556SP/ACq2tfzPXJ+Y896/qvwLduF6/wD1+f8A6RA/LON1/wAKsP8AAvzZ9s/8G7gZv+Cp3g5vTRNWz/4CtX9IKDKjiv5wP+DdVN//AAVR8Id8aDqx/wDJZq/o/TGOa/NfG6V+MIf9eo/+lSPpuDF/wky/xv8AJHD/ALVHgrxn8Rf2X/iN8P8A4ctjxBrngbVbDRcPtLXMtrIiKDxgknAPvX8kmt6NrPhnW7zw14h0uex1DTrp7a9sbqIpLbyxna8bqeVZWBBB7iv7GPNVVzu6V+Y3/Bcr/ginB+1Vp+oftc/speHY4viZZW3m+JfDVqiqviuFF/1kY4C3yqOO04AU4cKSvCPjTA8M5jUweN92lXcff/lktFzf3Xffo/K9jirJa2ZUY1qOsoX07ry8/wAz8HLS5uLS4jvbO5lgmhkWSCeCQo8TqQVdWXlWBAII5BAIr92f+CHv/BbC2/aO0/T/ANkn9r3xhBF8RbdVt/Cfii+cIvimMDCwStwBfKB7eeOR84YN+EkkU9pM9rdwSQzQyNHNDNGUeN1JDKykAqwIIIOCCCDzRb3lzZXMd9YXcsE0MiyQTwSFHidSCrKwOVYEAgjkEZFf0bxjwflfGeV+wxGk1rCot4v9Yvquvk7Nfn+U5tismxXPDVfaj3/yZ/Y0yADpgfSmMFA6/jX5m/8ABDn/AILXQ/tT6bp37Iv7VXiBI/idZW3l+GvEdywVfFcEa52OegvkVSSP+WyjePmDA/pixGBiv4rz7Icy4azOeBx0OWcdn0kukovqn/wHZpo/Y8DjsNmOGVai7p/g+z8wJHrSrjbnt9KZkcnrXwh/wWp/4K5aL+wV8O2+Dfwc1W2u/i54msC1kqlXXw3ZuCPt0y9PNPIhiPUgu3yqA2WTZNmGf5lTwOChzVJv5JdW30S3b/UMbjMPgMPKvWdor8fJeZ5D/wAHAH/BXwfBnw9qH7DX7Mfixo/GOq25i8f+ItNnw+h2brzYxOv3bmVT85BzFGccM42/hxG4TCpwBxwOlWtd1vV/EmsXev69qtxfX19cvcX17dzGSW4mdizyO7ElnZiSSeSTUGm2N9q+p22i6NYT3d5eXCQWlpawmSWeV2CpGiLyzMxAAHJJxX9qcI8KZfwZk6wtF3k9ak3pzS6vyS2S6Lzu3+NZtmmIznF+0ntsl2X+Z0Xwt+G3j/41fEPRvhN8KfCt1rfiPxBfpZaRpVmmZLiZzwPRVABZmOFVVZiQATX9If8AwSn/AOCWvw2/4Jx/B4Q3Is9Z+JHiC2jfxp4qRM5b732K2J5S2jP0MjDew+6q+Yf8ER/+CRVn+wf4Gb44/HHTbW5+LPiawCSx4Dr4ZsnwTZRt3nfA86QcZAjUlVJb79VlwMEfnX8+eKPiJLiHEPLMvn/s0Hq1/wAvJL/2xdO712sffcL8OrAU1icQv3j2X8q/zDYvGAPalVR06UfITyf1pBjBAP8An/Ir8bSufZNHzt/wV28dxfDr/gmJ8bNdFwscl14HuNLgJIG57xktQBnv+9P5V/Lw6jcdpGPav3o/4OePj9Y+Av2JvD3wPt7wDUPHfi6FjECdxtbJfOkb6eY0I/GvwTEu7qa/rLwOy6WF4UqYma/i1G16RSj+akflHG1f2uaRpr7EUvm9f8j6f/4Iw+CU8c/8FRPg7prwmSOz8SSajKAmdot7aWQE+g3Bea/p22AKMV+Av/BsN8J5PHH/AAUC1f4kz24Nr4L8BXk5lYcLNcyR28Yz6kGUj/dNfv6MdyOK/MPG/GxxHF0KMX/DpRT9W5S/Jo+p4LoSpZRzv7Un9ysiPA6g+tC4wfmpxC44x0pqkDuOvHNfjZ9aOGOMHtSNt7HtQSAcAjp60ErgfMOlACcZJrxf/go04T9gb4zEf9Ey1rv/ANOklez5XJO4cV4p/wAFIs/8MAfGjaR/yTHWj/5KPXo5Or5th/8AHD/0pHPjbfU6n+F/kfyn24LW0Rz/AMsl/kKVhg4oszmziP8A0yX+VOI+bNf6ApXPwWTtJn9H3/BvFEsX/BKfwBj+O+1Vj9ft01fbhAJ618V/8G+UKxf8Ep/h0QR81xqhP/gdNX2oVHXcOlfwlxl/yVmO/wCv1T/0pn7nlX/Iso/4Y/kIu3PWgYx1oXGcBh+dKCDzuH5180d4ZHOD37UYXPBFLxgnNNJXcfmHX1oAXjAyRQQuMg0fKDgEdKXC46igBMAtxzzXi3/BSZVb/gnZ8eFbGP8AhT3iLI/7h81e0naeQwrxj/gpMB/w7r+PIBB/4s94i/8ATfNXo5P/AMjbD/44f+lI58Z/ulT0f5H8pkRPlqfYUT8RMc/w0kH+pT/dFLMMxMPVTX+hl7s/BPtH9ZH7CkC2/wCwx8FI1HA+Enhzr/2DYDXgP/BwZHqM3/BKP4kRaepP77SjMB2jF/CW/kK+hP2HdjfsQfBZlxj/AIVJ4cI/8FlvUv7Zn7PFn+1b+yl4/wD2e5JESfxT4ZubSweT7sd3t3wMfQCVUz7Zr+CcDjaWX8WU8VV+GFZSfop3Z+416Mq+Uypw3cLL7j+SqOQkZPpzThNDE6zTrlEYM49QDk1Prmgaz4U16+8LeItOks9Q0y8ls9Qs51w8E8TlJI2HYqysp9xVYgMCCM+vFf3U5RqQvF3T6n4g1yz16H9ePwB17QfEvwK8GeIPC80cum3nhXT5rCSI/K0LW6FCPbGK6wntivw7/wCCMn/BePwZ+zL8LNN/ZJ/bKi1FPDOikxeDvHGn2zXR062ZiwsryFcyNEhJ8uaMMVUhGXChq/UPQf8AgqZ/wTw8S6QutaT+2L4Ba3dNwabxDFEwGO6OVYH2IzX8R8R8D8RZLmtWjPDzlHmfLOMXKMlfR3Sett09Uz9ny/O8txWEjP2iTsrptKz+Z9BAg/N7VneOvG/g74ZeCdV+I3xD8R2uj6FodhLe6vqt9KEitbeNSzyMT2AHTqeAMk18hfHj/gvl/wAEzPgVotxd2vx0HjXU4lP2fQvA1k99NO/ZfN+WCP8A3mkGPevyA/4Kdf8ABZr9oD/gozIvgSPTf+EK+GttcLNaeDLG8Msl9Ipyk9/MAPOccFY1AjQ4IDMN9epwv4X8TcQ4qPtaMqFC/vTmuXT+7F2cn20t3ZzZpxNl2AoNwkpz6JfqzyD/AIKFftc6p+27+2J42/aLumuItP1fUvI8N2VwcNaaXABFaxkD7reWodh/fkavGFk53E80z7JefZG1NbWQ20cyxNOIzsWQgkIW6BiFYgdSFJ7VEJs1/XuBwmGy7CU8Jh1aFOKil2SVkfkWIqVcTWlVnvJ3fzP0q/4Nfkjf9v3xJKVBZPhjeYP1vLYV+97gN3r8Cv8Ag15uFX9v/wASQEZL/DG8x+F3bV++zDI6Cv5R8ZL/AOu8/wDr3D8j9W4PVslXqz87f+DnJCf+CbVuQ3T4i6Tn8pq/n1U4PWv6Ef8Ag5ohEn/BNQNjO34haQQf+/or+fDaAeK/YvBb/kjX/wBfZ/lE+Q4xf/Ct/wBuo639n3DftBeAOf8AmetI/wDSyKv67ocCBP8AdFfyKfs5xmX9o34dwjnd4+0cY+t7FX9dmwKirjoBXwvjz/vuB/wz/OJ7nAv8Gt6r9RGPJwxpGAbJBzzQdoyeKDtz1HWv5/PvRYyq4Ga/nX/4OUB/xtH8Rj08L6L/AOkaV/RQNmMgiv53P+DlWLH/AAVF19wPv+FNFP8A5KqP6V+w+CGvGr/69T/OB8nxnpk6/wAa/Jn5+suORX6c/wDBqjIy/t0/EJM/e+EU/wCmqWFfmWIyRX6a/wDBq3CV/br8eOAP+SR3Gf8AwZ2Ff0D4mQ/4wTH/AOBf+lRPh+HZ/wDCzR9T95kIzljS4U9D9Kbwp6jrTxjsw/Ov4dP2QRQpHWk+Xg57ClCgj7w6UYXOCR9aBpilQSSD39KUBQOtJhcEBhilXaB94daBCkL3pQFIGB2o+XOdwpx2EA57UANCqQeaUBSxGeh9KaCuOCKUYJILDr60ABAIyD2puASeacSn94U0YJJyKAFJXOc9qU7ccHtTWZeMEdPWlBBGNw/E0AAC7slh1pSBjr3pq7S2Sw6047e7dzQAbVyMH6UgCk8U7aDg7h+dIqjOdw6etAAAMYOaUbQCAaBjgkighSDgjtQAmAW4NDIMZyOOlKCucbh+dDbSCQR7U7AN2c8GkK4OPpS78/nSMecmkAmzBBOetLtwRyaAxzjjr2pQ3GMcZoARh/KkOemewpWz+lIc5zjsKAEwQuAacB8wGfTvTckqcilBbdz6igCTGec0hHIIpAxHb3oLNnpQAnQAGlwduMd6YWPBxTlLMMY6f/WoA+QP+C9e4f8ABJ74qAHqdIGP+4pbV/NK43MTnvX9Lf8AwXoy3/BKD4pL6tpH/p0tq/mnlXDn61/Vfgav+MYr/wDX5/8ApED8s42a/tSH+Bfmz7k/4NyLPzf+CpHhufH+p8M6u3/kDH9a/owH071/O3/wbeDP/BT7RuM/8Ulq2OP+mQr+iPoM571+aeNqtxhH/r1D85H03BX/ACKJf43+SHlsDrTc5IH+zSE5HWgZOMelfj59dY/Kj/gu3/wRWX4zafq/7bH7I/hEnxpbo13478IabDz4hiUZe+t0H/L4oGXQf69QSP3q/vPw7R93Xg+lf2OAuDuBII6EGvyI/wCC7v8AwRMHiqLVv24v2OvBoGrIHvPiJ4K0u3/4/lGWfUrSNf8AlsBkzRKP3gzIo3hg39A+Fvia8J7PJc2n+72p1H9ntCT/AJf5X02elrfC8TcOqspYvDL3t5Lv5rz7n40+H9a1rwzrNp4l8N6vc6fqOnXUdzYX9lO0U1tMjBkljdcFXVgCCOQRX9AP/BFX/gsxp37cWjQfs6ftDajZ6d8W9Ks82d5lYofFlugy00a9FulAzJEOGGZEGNyr/PpA6soKnORwQetafhjxN4i8F+I7Dxh4Q1680rVtKvI7vTNS0+4aKe0nRgySxupBVlIBBBr9k404My7jPK3Rq+7VjrTn1i+z7xfVfNanxeT51iMmxXNHWL+Jd/8Ag+Z/Sl/wVm/4KfeBv+Cb/wAIEGmx22rfEnxLbSDwb4clbKRgfKb+5A5W3jY8LwZXGwdGK/zf/E/4kePfjF4/1j4qfE7xXda34h16/e81jVr2TdLczOckn0A4AUYCqAAAABWt8f8A9ob40ftSfFO/+NPx68e3fiLxHqMcUdxf3QVQscaBUjREAWNFA+6oAyWPViTxZkB6HpXH4f8AA2E4Ly60rTxM178//bY/3V+L1fRK8+zytnOJutKa+Ffq/MhmIRSxOBX7j/8ABBf/AII5W3wM8P6b+21+1H4Rz471SAXHgfw5qUPPh20dflu5Ub7t5KpyARmFCOjs23yf/ggN/wAEeoPGl1pn7en7VXhBZdIhdbn4ZeFdSgyt7IrArq1xG3WJSP3CMMOw80jaqZ/aR2ZmLOSSepPUmvyrxY8RliJTyPLJ+6tKs09+8Ivt/M+vw7Xv9Zwrw97OKxmJWr+FPp5v9AX5hggdeKA3agdBmk2sTwK/nxM+/TAckn3p8OXO1eSTx70wcD+tfKn/AAWB/b90b9gX9kXVPEum6rEvjjxYkuj+BbLdmT7Q6YlvMf3LdG3k9N5jX+KvQyzLsVm+PpYLDRvOo1FL16vyW7fRJswxWJpYShKtUdoxV2fjx/wcB/tf2n7Uv7feqeFvCOrrdeFvhjaHw3pMkMm6K4vFkL39wuOOZz5II6i1B6Gvhs5AyOfarN9cy3s7XU8zSPIxaR3YlmY8kk9yTyTXQ/BL4PeNf2gvi/4b+CHw301rvXPFOsQ6dpkKjrJI2Nx9FUZdj2VSe1f3fleXYLhnI6WDhK1OjCzk9NleUn6u8mfhuKxNbMsbKq1703t+SP3A/wCDXb9nS8+HH7GXiX9oPxBYiK9+Jfik/wBmsyYY6Xp4aCJs9cPcvdsOxCqa/TQ9MAfQGuK/Z6+DXhb9nT4JeFPgT4LUDTPCOg22lWjBceYIYwrSfV33Ofdq7MNn5sdfav4b4lzief5/icwe1Sba8o7RXyikj9ty7CrBYKnQ/lSXz6/iLtwAfQU0KcflT8kgcdqYCcZHpXhnYGDn+lNZf5UuSSevSkOegH1oECjrxXi//BR8f8a//jWxAwPhdrfU/wDTnJXs6tjNYvxK+H3hD4ufD3XPhX4/sXutE8RaXPp2r2sc7RNLbTIUkQOvK5ViMjkZrpwNeOFx9KtLaMot28mmZV6bq0JwW7TX3o/kA0/BsoR/0yX+QqbyyxAx3r+kW3/4N8/+CUFvbrDD+z5qWFXAJ8ZagT/6Npyf8G/H/BKYScfALU/ofGN/j/0ZX9Sx8cuE1vRrf+Aw/wDlh+Yy4IzZybU4fe//AJET/g32Vv8Ah1R8OwRjF1qoGf8Ar/mr7RC/MAPSuL/Z5/Z7+Ef7K/wl034IfA3w0+keGtJeZrCwkvZLgxmWQyOd8hLHLMTyeK7Q9Qa/mjP8wpZrneJxlJNRqVJySe9pSbV99de5+lYGhPDYKnRlvGKT+SEA/kKMdsUo5/L0oPHH9K8g6kNwPm4pQPm/GkyeeKATn8aBDtuMc9qUKe9IGJx9OaUMfT8aAEA6D2rxj/gpHx/wTw+PGcY/4U94jzk/9Q6Y/wBK9n6nhawvih8OvCPxi+GviH4R+P7KS50LxTol1pOs20U7RNLa3ETRSqHXlCUcjI5GciurBV44bG0q0toyi38mmZV4OrRlBdU0fyB24AiRT/dFOZCwbA7V/R5b/wDBvF/wSjgiCH4I65JgYzL43vyf/RlRXX/Bu7/wSplJ2fB3xDEMciPxxe/1Y1/VS8deEb/wq3/gMP8A5M/MZcE5s3dTh97/AMj6M/YQkY/sMfBTPUfCPw5n/wAFsFerI5RsisT4d+CPDfwx8AaF8M/Blk9to3hzRrXS9Jt5JjI0VrbxLFEpduWIRFG48nGTWxkkZxX8rYutDEYupVjtKTa9G7n6dQg6dGMHukkfkN/wXd/4Il+MfiZ4z1f9uL9j3wxLqmq6iPtPxA8D2EW6e7mAw2o2aD/WSMoBlhA3MRvXLFgfxkls7izuZLS8tnimhkaOeKRCrRuDhlYHlSDwQeRX9iqkhhivnz9q/wD4JT/sIftp38viP42fA6zTxDMuJPFXh6ZtP1GQ+skkWFnI7ear4r9n4I8X55Jg4YDNYOpShpGcbc8Utk02lJLo7ppaanx2d8JLG1XXwslGT3T2fnfofy5wxjp0qYWURO9o1PuRX7Z+Pv8Ag1g+Al9eST/Db9pzxXpUROY7fVdLt7vZ143L5ZP5VneGv+DVz4YW16j+Lf2ufEN1Bn54tP8ADsELH6M7tj8q/V4eMPAns+f28r9vZzv+VvxPkf8AVDPlK3Iv/Al/mfjB9lG3Crx7Cvdv2Ef+CaX7UX/BQnx1H4c+Cfg14PD9tcqmveN9ViaPStLTPzbpP+W8oHSCPLk9dgy4/a34Ff8ABvh/wTg+Dt1BrHirwNrPjy7gKsq+L9YL2xYetvAI0cf7Lbh6g19r+G9B8N+DPD9p4T8GeHtP0fSbCIRWOl6VZpb21vGOipFGAqKPQAV8dxH47YX6vKlktBub+3UskvNRTbb7XaXdPY9zLOBqvtFPGzVv5Y9fV/5H4+/8Fw/+CdXwY/YT/wCCQvw6+HPwRtZJx4e+LFrceI9du0UXes3d1p95FJdzbeMlkiVUHCIqoOASfxsjkI5P5V/W3+1V+yn8EP20/hLJ8D/2hPDdzqvhyXVLXUHtLTU5bRzPbsWjPmREMBknIB5Br5uX/g3k/wCCTyjDfs/6wfc+PdU/+PV4HBPirl+R5RKhmqq1Ksqk5uSUXfmaerco63v0sd+d8MYjHYpTw3LGKilZtrb0TPzM/wCDYXVBa/8ABSLU9Occ3fwu1Pb9VurNv5Zr+gx84r5s/Zd/4JKfsG/sa/Fpfjd+zv8ACe/0XxImlz6cLyfxVfXaG3mKmRDHNIyHJRecZGOK+kzyOa/PvEDiXAcVcQvH4SMowcYxtJJO6v2bX4nvZFl1fLMB7Cq03dvS9vxSPgD/AIOWoEk/4JkzsV5Xx9o5HP8AtSV/PKwIYjFf1oftT/so/BT9s34Wf8KY+Pmg3epeHzqUF+1rZ6i9qzTRZ2HenOBuPHevnCL/AIN6P+CVKn978Edaf/f8aX39HFfeeHfiXkXCmQSwOMhUlNzlL3Yxas1FdZLXTseDxBw1j80x6rUXFKyWrd/yZ/Pn+y8hk/ak+GSbTg/EXRAf/A6Gv66JfSvjrwf/AMEGv+CXngPxTpfjfwz8CtRh1PRtSgv9OuX8XXz+VcQyLJG2DJg4ZQcEYOK+wnlDHp1NfOeJvGmVcZYjDVMFGcVTUk+dJbtNWtKXbyPT4aybFZRTqRrNPmtazb790hjqCDzTTxmnscrkU1wT0Br8uPpxONnXtX88/wDwctCMf8FPNVKNy3g3Ri3PfyCP5Yr+hcdNvr1r5p/ag/4JC/sKftm/F+4+On7Qnw81fVfENzZwWss9r4mubWPyoUCIvlxMBwB171994ccUYDhHiL6/jIylDklG0Um7txfVrt3PC4iyyvmuXqhSaT5k9fJPyfc/mBjTK5r9Ov8Ag1ctyf22viLKV+58JZOT76pY/wCFfonZ/wDBvv8A8EnbQYP7Ot9LjvN4y1E/ymFeq/stf8Ezv2K/2J/HGpfEf9mX4Ry+HdY1fSP7L1G5bX7y6WW181JtmyeVlB3xodwGeMdK/U+MfGDhriDhnFZdhqVVTqRSTlGCXxJ62m307M+ZynhPMMDmNPEVJxcYu+jd/wAj3EKd3405VwKBjOTQDxjFfzcfoQKvByO1LsBwSelKoJB+lBBBxigBMYB+lC/1pSOtA6c+tACgc8DrzSBcAcdBSAnOKUEkDA7UACg9Aego2kE49aFJxnHOKAck5XoaADb2zTcc49qcST1Hak7GgAK44HpSBcGnYJ7dqT3x9aAADkZ/vUDkde/rR349aQZAzjvQA4EgjntSrngewpgJyM0qntntQA9RRtzk5xSBjShyM8dqAA4B60jdPwo3HJyO9IxOMD0pgAIA6Dr6UmVLZCj8qXAHUd80g254XoKQC7lIHH6UoII6Dr3pvy8YFLkY6UABYZyFHA9KM5Y/LQSP7tHynnHb/CgBNyheFz+FKSM52jrSYUL92lyN33aAFOORtpARuzinEIQcDtTflLcCgAIHGAKUEDoO1IWXpgduaUbTxgUAfIX/AAXoYf8ADqH4pjjppOOP+onbV/NNL1/Gv6WP+C8wB/4JS/FFTjGNK/8ATlb1/NNL161/V3gWr8MV/wDr8/8A0iB+VcbL/hWh/gX5s+7/APg2+IH/AAVB0TPfwlq4/wDIIr+ifAx0Ffzqf8G4ZH/D0HQCc/8AIsat/wCiK/orypGcdK/M/G//AJLGP/XqH5yPp+C/+RQ/8T/JCHB/hH5UqKDghRx7UmQR0pQy8DbX46fXj1AByAPyo345DYx0wabuUHgUhfdjI6UrMD8Uv+C9n/BGNfAE+r/t2/sl+FcaFPK138R/B+nQf8g2Rjl9TtY1/wCWDE5mjA/dkmQfKWC/kxE+7DA9ehFf2IXEFtd20trdW6SwyxtHLFKgZXRhgqQeCCCQQeMGvxb/AOCpX/Bun8Srf4h3nxt/4J3eF7TVNE1eZp9V+G32+K2uNLnY5ZrF5mVJLckk+SzK0ZOF3LgL/Rnhl4o0aVCOU53Vty6U6knpb+Wbe1ukn00fS/57xNwxKpJ4rBx1fxRX5r9UfkyFDDnr2r9Bv+CHX/BJJ/20/iB/w0P8etBcfCrwvf7IrOdSq+J9QQg/Zl9baM4MzD7xIiHJcrV/Y3/4N7f24fjH8VLC0/aR+HF58NPBVtcq+u6rql9bG9niBBMNnDFJITI4+USPtRMlvmICn98Phd8MPAPwX+Hmi/Cj4X+GbXRfDvh/T47LSNKs1xHbwIMAepJ6sx5ZmLEkkmvW8TfEvCYXA/2fk1ZTq1F704NNQj2TWnM/J+6td7HDwxwzVqV/rGMg1GOyfV+nZfibdjZ2ljbR2VjaRQQQII4IIYwqRooAVVUcKoAAAHAAAqVgpPCjp6UmVzjHenfKcYFfzA0fp4DAAGPTtSEIGzS/Lj6Vwn7Rn7TPwM/ZJ+Fl98Z/2g/iDZeHdBsgR5tw+ZrqXBKwW8Q+eeVugRAT9ByLo0K2JrRpUouUpOySV22+iS3InUhSg5Tdkupc+PHx1+F/7NHwh1745/GTxPFpHhzw5ZNdaheS8sccJFGvWSV2wiIOWZgBX8xX/BRP9vL4kf8ABQ39pPUvjn42jksNLiU2XhDw6ZtyaRpqsSkXHBkYkvI4+87Hsqgejf8ABVP/AIKv/FX/AIKSfERbKOC58O/DTRLtn8LeD/PyzsMqL28KnElwyk4HKxKSq5JZm+Rnj9/pX9aeGPh0+FsP/aGPSeKmtFv7OL6f4n9prbZdW/yriXiH+0qnsKD/AHa/8mff07ACG4/Sv2V/4Nlv+Ce1zpdpqX/BQr4oaDtkv4JtI+GkVzFgrCSUvNRXPTeQbdGH8KzY4YV8B/8ABK3/AIJv+Ov+CjP7Rdr4LjgubLwLoE0V38QPEMalRbWhYkWsTd7ifayIB90b5DwnP9OPhLwn4U+H/hTTPAvgjQbXS9F0WwhsdJ0uzj2RWltEgSOJFHRVUAD6V4fjLxrDC4N5FhJfvKi/eNfZh0j6y6r+Xf4kd3CGTOrV+u1V7q+Hzff5fn6F0fKoX0pBtB6d/WnEgjp9OKaxU9v0r+Yj9KuPDAoAFHT0pARj7o/KkUgAcfw05SMfd/SgBDgcBR0pCVz90flTyqjHHbsKawXHTtQAwbf7v04pGI3ZAyfXFKApyNopCELdAeBQNCZGegFOSTkUmF6YpUA3fdFAD8hlGVFGeRxnikG0jp3pflGOKABWA/hFBIxkikBGeFPWlUKRyO1A7DTjn5f0oXAb7o/KnYQj7vU0m1d3TnNA7DgQQAB/nil+UDhRSfIQMD0oJTbmgljdy56dqcGBHuDTflJwBS4G0cDtTYhckdDSMecYHXjilGPQUh2/3RSAQY9B0FKMBcBR1oIXj5e1Jxjgd6AHBuckD8qMqRjaPypMjPTFLw3btQAw43ZAppIOOKkwpONtNIH93tQAmccgD64pVYDoO3pRtU446YpQq7Rhaq4ChgR0H5U8NlfwqMhQMYpy7cZwOlDegDhj0H5UjPxjA/KgbTziglcDipAQHOMqOvpQH5z7+lC7SANnQ01sZzigAMoK4wPypFbOAQKbkEDK05dpPTtQA7gcYpZNuCMD8qT5QOnQUr7RxgUDG/L/AHR7cUqEAnAH4CkXbgfL+lKuz07+lAXH78jAH6Um4EdP0pCV6baF24HFAhCFySVHX0peAcbR+VKdpYnHekO0DIFAApAHKjgUrYODgflSDaB9ztSkoTnb+lACEjnAHX0pB/ujr6UpYHPy96Btx92gAGByAPypSVIGFHT0pDtHQdqPlwAFPSgAXGOg/Klyoz8o6+lINuCNvSlBUngd6AA7R0UdPSm4Uk/KPyp7Bf7vam4GT8tAAApI47UuQAML2pG29h2pSV7DoKpAJnLdB144oO0j7tKu3dkqOtBCEZ29/wCtJ7gNBAx8v0pARnhR09KViowMU0Fc8LSAUMP7o9elOVgAeB6dKZ8u0fL29Kcu3B47+lACgjcTgflQ2MYCjGPShSuTx+lK5U4OP0oAQ7PT9KQgZ6U7GRggde9GBnkjrQAwhSOlLwMcU4qrYxRtB49/WgBvHHH60fKT07c80pVQcCl2jpjsOaAGkLt+7+tBxnJHf1pwUMucUhVd3HrQAjEdh29aTjPIH504gcqOeKTYM4x29KAGHH93sKeT2xQVGRx1HpSlQeB1A/rQB8f/APBegkf8Eo/iiVHfSc59P7Tt6/mol+9+Nf0r/wDBewhP+CTvxUI7f2QB/wCDS1r+aeU5Y81/VvgS/wDjF6//AF+f/pED8s43/wCRtD/AvzZ92/8ABuCEP/BT7Qy/bwrquPr5Ir+ipeAOK/nN/wCDc2Yx/wDBUbwymf8AWeHNWH/kDNf0ZAjb1r8z8bv+Sxj/ANeo/nI+k4K/5FMv8T/JBkHqvT3pAQQMr39aVgu3qOnrQAvBHp61+PH2ALgdFpCMqPlHanKoJIH4U7Yu0HHpQIaF/wBkflSSYxjAxUmxSMZr5z/4KYf8FIfhD/wTR+C9j8UPiRot7rmqa7qLWHhfwvpsqxz6lOqb5GMj/LFFGhBeQg43KAGLAV14DA4vMsZDC4aDnUm7JLdv+vu3Mq1anh6TqVHaK3Z9ClY0+6gH0FLlQuAB+Vfjlbf8HZ9g3/H7+wfc+3lfEJP62dZ/iT/g7K8RSWbx+C/2EbGKcj5JdY+IUkiKfUpFZqT9Nw+tffrwq47k7fVP/KlL/wCTPD/1nyVa+1/B/wCR+zgUM2NnbNUvE3ibwx4K0O48TeMvENjpGm2sZe51DVLyO3hiUDks8hCgfjX8/nxV/wCDmP8A4KT/ABCtzZ+BE8B+AIjn974c8ONdXJB7GW/kmX15WNTXx18df2q/2kv2mtVOs/H343+JfFkxbKprOrSSxR/7kWfLT6KoFfT5V4H8QYiSlj60KUey9+X4Wj/5MzysZxrgqWmHg5Pu9F+rP27/AG5P+DkL9k/4DWN74I/ZWgT4p+Lo90Q1CBnh0KxcfxPccNdEHokI2n/notfip+1b+2P+0V+2z8Tn+Kn7RnxFudcv0DJp9kB5VlpsROfKtoFOyJfUjLN1ZmPNeXoAvf8ACnjB7dK/cOE+AOHuEo8+FhzVXvUlrL0XSK9EvO58XmmfY/NXy1HaPZbf8EcWJOSa9w/YI/YC+OP/AAUL+NEXwp+ENgttYWmybxP4pvYmNnotoWx5khH35DyI4QdzkdlDMvrn/BNL/giz+0X/AMFANSsvHetw3Hgr4XmYG68Y39r+91BAeY9PhbH2hjyPNOIk5yzEbK/oD/Zh/ZR+BH7HPwnsfgr+z74Ig0XRbM+ZMc77m+uCAGubmU/NNK2OWPAGAoVQFHy/iB4pYDh2nPBZdJVMVs2tY0/8T2cl/L0+1bZ+lkHC9fHyVfELlp/jL08vP7ij+x3+yH8Gf2IPgXpPwE+COiG307Tx5t9qFwoN1qt4wHm3lw4+9I5A9lUKigKoFepllIJCjr6UFRkYowuCB6+tfybicTXxmInXrycpybbb3be7Z+q0qVOjTUIKyWyGjHdaacHnFPCqRkUFAWO0Hr2rAsYpXHTtT48Y6dqaEB5x2p0YXHbpQA8BewpCqkcDt61JgZ4oKgngjpQBAUwTlapT+IPDdtdNY3XiKwjmj4eGS9jDqfQgtkVpFAT071+Sf7ZkTD9r34gSBzn/AISFxkEj+BK6cLh/rE3G9j4HxB42lwNltLFRoe1558tublto3e9n2P1eXWtAcfLrlkfcXaf405dZ0ENk63Zj/t7T/Gvxiju71Vwt7MB7TN/jQbu+J+a/nP8A22b/ABrr/s3+9+B+Ur6Qcrf8i7/yr/8AaH7PjV9Bbhdaszz2u0/+Kp39paMB/wAhe14/6eV/xr8XDe3a8/bJf+/rf40+PV79eBfzj/ts3+NNZan9r8A/4mDn/wBC7/yr/wDaH7P/ANp6Mc/8Te16/wDP0n+NIdY0JPva3Zjjvdp/jX4zLrOqZ+XU7gfSdv8AGkl1fVSOdSuD/wBt2/xo/sx/zfgP/iYWp/0Lv/Kv/wBofsymueHzkDXbL/wMT/Gmtr3h5c7vEFj1/wCf2P8A+Kr8YJNU1A5zezf9/W/xqH7dfMeLyX/v63+NNZU/5vwE/pCzX/Mu/wDKv/2h+0n/AAkPho8f8JFYcf8AT7H/APFUHX/Dh4HiGw/8DU/+Kr8XVvL4n/j9m/7+t/jTxd3vU3s3/f5v8aayp/zfgYy+kRNf8y7/AMq//aH7QW2q6PdzCCz1W1mc9EiuUYn8Aas/LgfKeD61+Y3/AATqu7mX9rvwpDNdSMpe5JVpCR/x7ye9fp3gbRz/AJxXBiaH1eryXufsHh7xo+OsmqY50fZcs3C3NzbRi73sv5u3QAFIzig4JBCj8aVdpGCfTvQVUHGK5z7uyGkqTnb2HekAUdFqQqOoHakUKVycfWgQhADY2jrS/KP4e1KQueD39aT5RkD09aBjeMj5R9aRgp5C07C54xSEDjp0oEBA7ijC8cYx70uQO47Upxk4HX2oAjbGM7fwoBHXb2pWAPQdvSlCjHTsKAEXA6L255pWwcYX9aVE4yB2pWQAZxigBqgYGV7460mF9P1qTA2g+4poCk4oAjIBAO0fiaUADkLTiq4BB/Wjao6HtQA07R/DStt9O9I23JORSsFI4PfigYDaQM4pBjOccZ9acqjjp1pVCle3SgegwgY4FKpHBxTyoOcUgRVx9fWgkQFSc7frSgA9v1poAB49fSpAAeQP0oAaoUA4Xt60h25ztp6qCOR2oZQCMelADQBkjaOtAAxytOwvOT+ZowhTIPYd6AEwp7frSELtGAOlOIBOR7UYG0H2oAjGOTilGMkY7+tAAYdO1KAMn6igBflxjb29abxzkfrT9q/pTflGckdPWgBG2k8Z6UgxwNv605lUnqOnrSqqnjjgUAAABzt/WhgvJx345pVUbsZ/Wkk2gYGOvagBjbcjj9aYOTnb196mKgkHHrTUjBPTtQAwYA6dvWlULzx6U4Kp6enrS7VGcAdqAE4DZxSORzx0NOKjJA/CkkXg8Z70AG0jkfypOQevH0p2SR06Um478HPWgYLgbcEfTFLn3FJuPGCfzpQze/WgQEZ6Y70YAPUdBQWOevagydfYetAxCBtzQc5HQ8+tG8svH86AxLd+ooAXAPJweKTGG6dvQ0pc5/DnrSbjn8PWgQ0jgcdqUY29e9BJyMZ6Uqs2MDn61SA+Pv8AgvYN3/BJ/wCKan/qEY/8GltX81EoO81/Sp/wXyZh/wAEovieV7y6QD9P7Tt6/mskOWP1r+qfA3/kl6//AF+f/pED8r43/wCRrD/AvzZ9wf8ABulEx/4Km+FCOg8Pavn/AMBjX9GuRwM81/Oj/wAG5jqn/BUfwzu6t4b1YL9fs5r+i/dkDHrX5r43f8lhD/r1H/0qR9LwV/yKJf43+SEJOeG7jvSZyACc07ccYyabuOB1/Ovx4+wuKuM1IoIAwvp2qJZOeP5mno/AP0oCxKin+7+lfjV/wduSMbj9n+33fKI/E7gZ77tLGa/ZVZCR1H1r8Z/+Dtpx/aPwBBP/AC7eJv8A0PTf8K+/8K9ePMH/ANv/APpuZ4PFGmR1vl/6Uj8cAvTp0p6detNQE9a+5/8AgkL/AMEefDX/AAU/8G+NvFfiD9oDUfBZ8IaxZ2SW9j4djvvtazQvJvLPNHsIKEYwfWv66zrOcv4ey+WOx0nGnFpNpOW7stEm9z8nweDxGYYhUaKvJ+dj4fGAKA4JA6Z9a/dX4ef8GrP7HXh+6+0fEv8AaC+I/igDGLa1az0uI+ufLikc59nFfV37Pn/BIj/gnX+zPJBqPw8/Zh0C51C2UbNU8Ro2p3OR33XJYA/QCvzTMPG7hbDRawtOpVfoor727/8Akp9FQ4KzOb/eSjFfefgD+yR/wS5/be/bS1FD8GfgjqC6NvC3PivX1NhpcAJGT58oHnEA52wh29q/Xj9hL/g3N/Zd/Z7Nl47/AGn7uP4peKIGSVNNuLcw6JaSA54tyd1zg/8APU7Tj7lfovHtjiSCKNUjiXbFFGgVUXsABwB7CpEZhwK/IuJfFnifP4yo0ZfV6T+zB+815z3+7lXkfWZdwpluAanNc8l1e33DLW1tLC1isrG2ihgt4lit4IYwiRIoAVFUcKoAAAAwBUnzHOR29KXJIGeePSkDHHGOBX5e9z6hAcccfpSZPOPX0pWYkjnt60mcA8n8c0hAuSvPNLjJ6D8qAxAOKMknPv1oAQDABHp2FLGxHTt+NAYkD/GhCcc9cUASA88fypxI6cfjTNxycfzNBc9vT1oAMgNnHftX5MftlyQL+1t49RnUN/b7kgn1RK/WXcST9etZ114M8GajeSahqfgzSLm4lOZJ59Mhd3PHVipJrpwuI+rzcrXufA+IXBU+OMspYWFZUnCfNdx5r+61bRrufjQmwjg5pWTOQFJ/Cv2WXwZ4ITiPwTo6/wC7pcI/9lo/4RLwhnjwfpX/AILYv/ia63mX938T8lX0fMS/+Zgv/BT/APkz8YLlhGcMcfXioUnRpQquCfQGv2hk8HeCpc+Z4K0duf4tLhOf/Ha8l/bq8G+DrD9kXxte6Z4P0q1uE0+22T22nRRuv+m2+cMqgjrVQzHmko8u/mefm3gVisqyuvjXjoyVKE529m1fki5Wvzu17WufmEj4A/xpzHcvFQO+D1o80gFvQHt7V6uyPwCUnEbKwRyCCfwp0alv4D+VfrV+y94R8K3H7NfgKW98K6bNJJ4P05pHmsI2ZibdCSSV55Ndt/wg/gYNkeCNHz/2C4f/AImvPeaRjJrl/E/oTCeAeIxmDp1/r6XPGMrezel0nb4z8Y9gA5U/lQeBzX7Pjwj4NX7vgzSOfTTIf/iaZJ4O8FSr+88FaOfXOlQ//E0v7Vj/ACfiaS+jtiZbZiv/AAU//kz8yf8AgnjdiP8AbH8HIG+9Pcg4/wCvaWv1F3gqCMVlWngnwTpuopqml+DNJtruI5iurbTYY5EyMHDKoI4NaSkxrk+1ediq6xNXmSsfsXhzwVV4FyepgaldVeebndR5d4xVrXf8pNkkdO9I3XPFNDnpjtSlmA5/OuY/QBQQO4HAoBwOvpTd7ZwfQUqNuGcDr6UCsBZs4JHWgsf0oJOR/iaQseBj8/woELxuGMdPSgg8Yx27Uqk5Ge/oKDwAc0CEJPqPypd3v29TSMxwefpQWPTJ6d6YDSRx8w/OnKRtJwOgppbI5NKpIJz6DvRcB+3PIHb0pecDH8qQOcd/fmlLnA+lIBMdOB19KbnHPH5UF3wMnvSBzuwaAAEkDnoPSlzlR/U0uTtGPzxRuIGc8e9ADGPynkdPWlPse4pWYkEZ6UuTg/WgYq9vp6U5Pu9TTVJAHsKVW4xzx15oC4rHP5UzpjBHXtSs3HB7UiseCTnmgQgzzjnn3pVBwD7UBie3elDEjkdqAHICF7cDmgqM5x06cUquQD/9ekd2GCaBjWY84b060IzY+8PzpHkIyMHoKFbjJJpJ3EODH179s0gxgYHalyRjPWk3ZGCOnvTAbj5TwOlCjB7dfSgM2OKMkHmgBQcc+1AI/wAiky1CM2c57UAOJOeP7voaNzZ/+tSbjn8KNxBBxQA5WwRyOvpSMflGD1P9aQM2cAnr60Fjj159aAE28ggfpSpn07UBzx249aRWPb0oAf74pB3xj8qNxz07e9KGbn6UDG9GIGPwFI+SPwp247sk0x2IUYpoRLjPUDr3puPnzgdfSnEDPGeRTBw3T9aQxcYAHBpQO20fWk4wOKDgcgUCBwAQAB+dJgknjt60jEE9P1pOM4x2HegYoxtIA/WlYZY9OvrTSQo/+vS5Bbp39aAHMBg8CgA7u1IcHn29aAFLc/zoEBHONvanBQeNvb/CmkL6duKcAuadwPj/AP4L3Irf8EnfinwOP7Ix/wCDS2r+aVxhiK/pW/4L7yGP/glB8TlX+KbRlP0/tS2r+ayReSTX9U+BavwxX/6/P/0iB+V8bv8A4Vof4F+cj7Z/4N4JTH/wVN8HY/i0bVFP/gM1f0dqQRyR7V/OH/wbxx7v+Cpng1j20bVSP/AVq/o6QDAyBX5z44K3F9P/AK9R/wDSpn0nBP8AyKZf43+SHntx+VMI6cf0p2Pb9aawGAP/AGavxs+xQgAyTx19aVQMDp0po659/WhRleB1HrQNMmDA8V+Mv/B2uxbXfgInUCw8SH/yLp/+FfswoyeVzzX43/8AB2nCn2v4Bz4G77P4lXPtv00/1r9E8J0v9fcIn/f/APTcz57ip/8ACHVt/d/9KR+OSdPxr9s/+DT8Rp8GfjG38R8XaZn6C0k/xr8TVA4Fftj/AMGoCn/hTnxjc9P+Es0wD/wEkr+hfF1L/UPEf4qf/pcT8/4Uf/C5T+f5M/WoDPIX9aGUHjAoGMc9veggZx7V/G5+vjQvTgde9OUD0HXtSqi5Hyjr6UoQDPHfigBwQH+EfnSBMKTgfd/vU4ADt+tJgYPHbsaQCFeeg/OjaCDkfrQ2PT9aQkc/X+9QMUKMcD9aNv8Asjt3pFYDt39aXjPH6U0ABeASvb1pFHHI/WlCgAHGePWkUDpjNIBxGScgUuBjG3t60nGcAHp60uAO3b1oECqM8r0NI3XkdKX5TnIpCoLHjvQMaeTyP1o2nd0pThTxTVIJ6dqYXsABI5A69fxryP8Ab7Ij/Y58dMO2n2/T/r9tq9eAU8EdzXkX7fip/wAMcePVH/QLgI/C7t60ofxo+qPn+Ln/AMYpmH/Xit/6bkflMTk0rcxtg/wn+VRNJhs0GU7Tg9jX1FlY/wA85vRn7EfszAD9nLwEF6DwfpuP/AZK7Rzzx61wn7MUhf8AZu8Akk5Pg/Tj/wCS6V3OcnHP518pP42f6NZPrlGH/wAEP/SUITkjIH4mjqOMfnSFQR+FKADwak9JOwgjJbPHT1rC+JfxK8B/B7wfc+O/iN4ig03TbbrLKctI56RxqOXc9lGTWd8dvj18OP2cvAUvj34iap5ceTHYWEBBuL+fBIiiUnk+pPyqMkkCvzD/AGjv2lviB+0141PivxlciC0gLLpGiwSEwWER7L/fcj70h5b2GBXZhcJLESu9I/1sfmfiD4kYDgzDexpWqYqS92HSP96fl2W78lqfWeq/8FdPhXaXbxaT8K9fuoVYhJZJoY9w9cZOKpP/AMFg/AAHHwY1s/8Ab9DXwsy5Gaice1eystwttvxP5yn4zcfuTaxEV/3Dh/kfdR/4LCeBtxx8E9ZwT1+3w08f8Fhvh/HGWl+DGsjHOBfQ18IEcdKjkAGRipeWYZ9PxIfjP4gL/mJj/wCC4f5H7NfCr4gWHxY+Gmg/EvSrGW1t9d0uG9htpmBeNZF3BWI4yK3eN2T0+teX/sTy+f8AskfDplbj/hE7Uc+wI/pVv9qD9o3wt+zJ8NpPGeuQi81C5kNvomkrLta8uMZwT/DGo+Z27DgckA+A6cvbOEe9j+vcJnWHo8L0c1x81GPsoTnLZaxTf3t6JeiOs8f/ABR+Hvwl8PN4p+JPiyy0ixVtqy3cmDI391EGWkb2UE180fFX/gq/4G0a6ex+FPw+utYC8LfanP8AZYmPqFAZyPrivjL4rfGP4h/G3xhN46+JHiGS+vJMrBGMrDaR5yIoUBwiD25PUknmuaaUt9416tLLqaV56v8AA/mnijxyz3HV5UsniqNLpJpSqPz1vGN+yTfmfS2t/wDBVv8AaJu5i2i+EvCNlHnhZLa4mP5+av8AKqUH/BVj9p+3bdcaH4NnA/hOmXCZ/ETmvm+d404ZsZ7E1Fw3Qg+wNdUcLh1pyn55LxG47nPmeYVL+qt9yVj678Jf8FgPHFtOkXj74J6Zcxg/PPo2qSRtj2SVWH619E/A39v39nb433MOg2viN9B1q4YJBpWvAQmZz/DHICY3PoMgn0r8u2iyOetQz26yIVYcHtRUy7Dzjpoz6bJfGjjXLK0XiaixEOsZxSdvKUUmn5u/oftwBjgrj1FDDjkH6V8P/wDBOT9t7Xr3WbL9nf4xazJeG4xD4W1q7kLS7wOLOVjy4IH7tjzkbDnK19wEJxgnp614NajOhPlkf1ZwnxVlnGGURx+DdltKL3jJbp/fdPZqzG7QQBjvSoo3ZIH4UAjgAHr60qj5sD8awbsz6fSwojyoGw5+tea/HH9rf4Cfs/BrDxz4vWXVVQMuiaYvn3XtuUcR/wDAyK8A/bW/4KH3Wk3d38Jv2e9bRZYmaHV/FFuQ2xgSGhtj0yOhl9eF6Zr4hvrq7v7mW9vrqSaeaRpJpppCzSMeSzMSSST3NephsvdRc09EfgPHXjTh8nxE8DksVVqR0lUesE+0Un7zXe9vU+vfiD/wV612WVofhh8HrWJAfluNd1BmJ56lIsflurg7r/grD+1NLIWt9G8GwrnhF0q4b9TPXzlJt68VBIuTXowwWHgvhPwvGeKPHuLqOTx04+UeWK/BI+pNC/4K6ftCWUynXvh94S1CMfeWFbm3Y/Q73H6V7P8ACX/grT8GfFc8WmfFPwlqXhaeQgG8jYXlqCe5KAOB77TX56pEe4p3lA/eFRUwOHmtreh2ZZ4vce5dVUnivapfZqRjJP5pKX3SR+03hvxT4Z8Z6Fb+J/CGvWmp6deJvtr6xuBJFKPZl7+o6jvV5TjB96/JL9nf9qD4nfs0+JxrHgvUTPp00qnVNCuZG+zXq98j+CTHSRfmHfIyp/UD4I/GnwT8fPh1ZfEbwLds9rc5S4tpcCW0nXG+GQDoy5HsQQwyCK8fEYadB90f014f+JWW8b0XRcfZYmCvKF7pr+aD6rut11utX1oxnjFOUDjApnHpThjPeuc/Sh6AMDxyfesH4lfFH4cfB7w23i34neL7PRrAHakl3J80z/3I0XLSN/sqCa84/a//AGw/DH7LfhNY7e3j1PxTqUROj6OZCFUdPPmI5WIHt1Y8DuR+aXxL+K/xG+M/i6bxz8T/ABVcapqM2QrStiOBM5EcSD5Y0HoB9cnk9mGwcq/vS0X5n5Lx94rZfwlN4LCRVbFdVf3YduZrVv8Aur5taX+zfih/wVy8F6Tcy2Xwp+GF5qyqSI73VrkWqP6EIu5sfXBryTWv+Ctf7TFy5OkeEvBtknZXsbmcj8TOv8q+bpMFdx/KqF48aNl2A+pr2aWAwsF8N/U/nXMvFfj3HVHL646a7QUYpfhf72z6Z07/AIK0ftRwzBr3w94KukB5T+y7mIkfUXB/lXp/w5/4K+2Nw0Vv8Ufg/Jb5AEt1oeoeYB6kJIAfwzXwpbyRuwCOCfY1oW+0kcdqupgsLKPw2OXB+KfH+CqKax0peU1GS/FX+5n68fBr9pH4LfHuzaT4a+M4Lm6jj33Gl3H7q7iHqYm5I/2lyPeu6C5ySP1r8XdG1/WvDWqW+veHNXubC+tJBJa3lnO0ckLjoysOQa+/P2Ff2+v+F03MHwi+MM8Fv4qEeNN1JVCR6soHKlRwk4AzgcOMkYPFeNisFKj70dUfvvh/4x4biLEQy/NoKlXlpGS+Cb7a35Zdk20+jvofUhQAHK00LknA6U+QDkc+/NR5yTkevOa4D9yHnrnaOlGM9FHSmnGen60oI9O3rQIAvzYx1PrRjA6Dr60mefx9aUAHqO/96gBQq8cDp60KAR2/Ogcnn065pAQDj+tA7D1APUDpSheD8o6daYOgPt60oIOT7etAARliMDPemuueRjilYjdn+tI+PU8e9AC85yP5U3+L6U/aByfX0ppA38H60CEJ6DPSjg9KOOMt29aUAHnPQ0DGtTfoe1SlQCOT70hTHOT09KAIjnHXt1pf4vxFLtAU84/ClKjd17+lAC96O/4etLgZ5NKEAOAO1AhpB6j0pwyfwpCuAMfyp2Ac4zQM+O/+C+kbS/8ABKD4nEfwS6Mxz7apbV/Nc45J96/pV/4L3kJ/wSb+Kh/7A4H/AINbWv5qGcZ61/VvgT/yS9f/AK/P/wBIgflPHC/4Vof4F+bPtn/g3pA/4em+CiRn/iUaqP8AyVav6OUAx0r+cT/g3qZf+Hp3gj30rVB/5KtX9DPxd+MHwr+AHw71L4s/Gjx3p3hvw5o8XmahqupzbI488BQOrux4VFBZjwATX5343UqlXjKjCmnKTpQSS1bfNOyS6n0fBUowyacpOyUnv6I6XAxjFII5JF/dRs2D/CpP8q/Fv9tn/g6B8c6xqd34L/YN+HdrpGnxsyJ458YWYnu5/wDbgss+XEO4Mxc88ovSvgP4nf8ABTj/AIKH/F+7a78eftnfEO43Nu+z2HiOWwgB9obQxRj/AL5rnyXwV4szKkquJcKCfSTbl/4DHRejkn3ReN4xyzDTcaac2u2i+9/5H9TsiPBxNE6+7IR/Onquen41/KV4F/4KJ/t+fC7VI9Z8Cftn/EuxmiYFUl8X3V1C3IOGhuHeNxx0ZSK+3/2Nf+Dnn9o/4e6ra+Gv2zPBWn+PNEZws/iDQrOLT9VgU9XMa4t58f3dsZP96qzfwT4oy+k6mFnCul0V4y+Sen/kwsJxnl1ZpVYuH4r+vkfuzswcV+Nn/B2srfbPgE2ePs/iYdf9vTK/VP8AZo/ao+Af7YPwvtvjB+zt8RrLxFolw3lzPbkrPZT4y0FxC2HglAPKuBxgjIINflV/wdqyEX3wDjYf8u3iUj/vvTa8fwvw9bDeIeFpVouMouommrNP2c9Gnqjt4lq06uQVJQd0+XVf4kfjqp4Br9t/+DUGID4C/F+f+941sF/KzP8AjX4iq+BX7d/8Gnsiv8A/i/FnlfG1gT+Nmf8ACv33xef/ABgeI/xU/wD0uJ8Jwon/AG3D5/kz9YRz270oB/SnBBnHPWjAGMHtX8bn68LGAQKdgdfehABtye/SgEY60AL1ppzt/wCA0owOmfxoGME57elACMPwphGAeakZeR+lNKg557+lAxq98+tKOenrQABk5NKoHqevrQIUYwOO1NHT8KkVRtB5zj1pgAxxnp60DDPGMUoIzx6daNozjmlIwSc9qAQmcA5oPJJ54o6knJprY3H/ABoACcngUgPP40nGeoFKoyeT+lAhVJI614//AMFBGZf2N/HYU9dPth+d9bCvYQMrjPevHv8AgoAhP7HHjodf+Jfbf+l1tWtD+NH1R89xd/ySmYf9eK3/AKbkflC7ndkGjedp+hpD1pGGEY542nn8K+p+yf56VFa5+xP7Mox+zh4BA/6E/Tsf+AyV2xJyf89q4v8AZnTH7OngNeePB+m8f9u0ddow5Ir5Ofxs/wBG8nVspw6/uQ/9JQcZ6UoORTWI/wAmlTaTgkdfWpPSsfm1/wAFWtV1O6/awi0m81GaW1svCFi9lbPISkJleYyFR0BYqpJ77R6CvnFX53A9K+hv+CrSlf2xG54PgvS8Y/37mvnUbh3r38I2qET+DPENylxtmF3/AMvZfmezfsf/ALLbftX+K9Y8Kr48TQDpOmJeee+nG580NL5e3AdNuMg55r3a4/4I6aipPk/tE2Z9N3hmT+k9c/8A8EfCf+FweMAf+hViPP8A1+JX36/Jx71z4jG4mnVcYy09F/kfs3hp4ccH8RcH0cbmGH56spTTfPOOik0tIyS2XY+G3/4I8eIvux/tAaYf97w7MP8A2rUM/wDwR08VEEp+0FpHsP8AhHZ//j1fdXRuB2pMDsec+tc/9pYz+b8F/kfef8Qa8PHvhH/4Mq//ACZyfwD+GFx8GPgt4Y+FN7rcWpz6BpMdnLfQwGJJ2Uk7gjElRz0JPTrX54/8FC/jZcfFn9o7U9JtrotpfhR20mwRW+Uujfv3+pkyv0jFfp5lo/njUFgMqGOASOgz9a/Ni7/4Jiftjavq13rWr2fhc3N7eTXNw6+IiwMkkjO3/LL1Y1WCnTVVzqPX/M+d8YMrz2rw5hMoyXDznSv7ygnK0aaShFvV9b678p89+ZubPTFfQf7Ef7Fg/aVvbvxb41v7iz8LaZOIZfsjbZr6fGTEjY+RQCCzDnkAdSRMn/BLX9qxBuNv4YPsddP/AMar7Z/ZM+EWtfAr9n3Qfhx4ngtk1W28+XU/sk/mxtLJM7ZD4G75So6dvauvF4uMaP7uSuz8z8OPDHMcdxFF57hJwoQi5WlFpSkmkovy1u11tbYd4M/ZS/Z18A2MeneG/hDoiLGuPNuLJZ5G92eTczH3Jq9r/wCzh8B/E9u1trfwj8Pzo4wc6XGpx9VAI/Ou03cY/WnoVJwD2rxVUq3vzM/q2GSZPToexjhqahty8kbfdax8N/tof8E7vD3gPwfefFz4HRXKWunRmbWPD8khm8uAfemgY5bC9WjbPy5IIxg/G7KpGevuK/ay+0611PTp9NvIFkhuYXimjcZDKwKkEehBIr8ZfGejJ4a8Y6x4ciHyWGq3Fsg/2UkZR+gr2sBip1IuE3qj+WfGngzK+Hcdh8dl8FThW5lKK0ipRs7pdE09lpdablDT7y80vU4NV0+5aC5tZ0lt5kOGjdW3KwPqGAP4V+wXwV+Ia/Fj4SeHfiOqqraxpMNxMiHhJSMSD8HDV+PGDkmv04/4JqatNqv7I+hRzSFjaXl5brk9FWdiB+tTmcb0lLzOrwDzGrR4hxOBv7tSnzW84SST+6bPdyMgV8w/8FL/ANpm9+E3gK2+EXgy/aHXPFMDm8uYXKyWmng7X2kcq0h+QHsofvivqAxuVCouSeg981+T/wC2b8SW+LH7SnirxNFdGW0tdQbTtNYHIEFsTECPYurt/wACrhwNBVa13sj9W8YeJ63D3C3scPLlq4h8ifVRtebXysvLmPM1uGIAB4AxigksoRVJLHCqBkk/SoTlelfUX/BL39n/AE74j/FC8+LfijT0uNP8KFDYRSoCj375KMQevlqCw9GKntXuVayoUnJ9D+SuHsgxXE2d0Mtw7tKo7X7Jayk/RJvz2Nv9mz/gltrPjTS7fxn+0Fq13olpcIHtvD1jhbx1PQzuwIgz/cAL+pU8V9EaX/wTs/ZD0uzW0PwnhusIAZby9mkc8dSS/WvbGC4ySeuc96GJHJPtXz9TF16sruVvQ/tHI/DTg3I8JGjDCQqSW86kVOTff3k7eisvI+aPid/wSx+AHirS5W+HM194Y1HaTA0dy1zbM3YPHISwHurAjrg9K+GPjj8CvH/7P3jWXwR8QdKEMwXzLS5hO6G7izgSRNj5h6jgg8ECv2BVsADmvGP28fgXp/xt+AGrPDYq2teH7eTUtGnC/OGjXdJF/uugII9Qp7VvhsZVjJRm7pnxfiJ4TZHmOU1MZlFBUcRTTlaCtGaWrXKtE7bNJa6O/T8r3Pavf/8AgnV8frn4PfHO18KatqnleH/FciWWopK3yRTnIt5+ehDnYT/ckPoK+fPPVhnnnkcUsF9LZXCXdvIUkjYOjjghhyD+detOmqkGn1P5XyDN8bw/nFHMMNpOnJP1XWL8pK6fkz9tdpGRjGOCKwvih8Q9A+Evw91f4k+KJdtho1i9zMF+9IR92Nf9pmKqPdhVP4EePF+Kvwa8M/EIHLato0E8x/6aFcP/AOPA186f8FafiRJo3w58P/Cexn2nWdRN7fqG5aG3HyKfbzHB+qCvCp0lOqoH918ScTUMo4Sq5xDX3FKHnKSXJ+LV/I+Kviv8TvFXxj8e6n8SPGl4Zb/U7guyBvkhTokKeiIuFH59Sa5mQndnpipdjfWvT/2NvghZ/Hf9oLSPCeuW/m6RZhtQ1iPtJBFgiI+zuUU+xavalJUad+iP4bwWGx3EWcwoRfNWrzSu+spPVt/iz0z9kz/gnPrnxk0i0+I3xc1O40bw7dKsljYWw23moRno+WH7mI9mwWYcgAENX134N/Yw/Zg8DWiWujfBzR5Cq4M9/bC5kb3LS7iTXpqRRRRLDDGqIihURFwFAGAAB0AHGKdgDAJ/WvHqYuvVd+ax/a3DPhzwvw1hI04UI1Kn2qk0pSb62vflXZL53ep514m/ZN/Zs8VWr2mr/Bbw+ysMb4dOSJx7hkwR+dfNf7RH/BLm2sNNuPFf7PV/cySwoZH8M30vmGVQMkQSnnd6I+d3QMDwftkgdfelTbnBXof8mpp4mvTldSZ2Z7wJwrxDhZUsVhYJtaSilGcfNNL8Hdd0fizcxTWlxJa3ULxyxOUkjdSrKwOCCD0IPBFMtdR1HSNRt9a0i+ktry0mWa1uIW2vFIpyrKR0IIzX0b/wVA+Edr8O/j/B410e1WKx8X2Bu5UjGFW9iYJOcdtwMb+5ZjXzYQSCTjmvbpzVampdz+I+Ickr8NZ7Xy6o7ypSsntdaOMl2umn5H6yfsh/H2P9oz4F6X8QLgKupRFrLXIU6JeRY3kD0YFHHs/tXphzkn618E/8EhvH0uj/ABJ8VfC25nxba3pcepWsbN0ubdtj492ikGf+uIr73VO3P+cV4eIp+yrOKP7T8OuIJ8S8IYbF1Xeok4T/AMUHa/8A28rS+YZ+bnNOBBpMe56DvShRwRxWB9uJ/F+Ip3PX370gwDye/pThjGc9/agdmIeTzTfY1JtGRyaaVGcg0ANzzQrnkfzoIA6n9aFXqPp3zQAEnJNDsxo2jcR3odQRnPagB3GODSHbv5P0oDYGd3f0pMsGHzHrTeghdo4pQvHU03ceAGzz6U4M2Pvc/SkAnGeSTxSkDkEn7ooJJ53dvSgtgkbjnFAxuxQvAP50bRu78n0pVJA+/wBxS5O7hu9A2wCAfl6U5UGenak5HIbtSgkHOe3egQjKuM4xSEADr6UrFuOecelId397vQNJWPjv/gvoT/w6c+Ka7uM6P1P/AFFbWv5qHZi2Ca/pQ/4OAJZIv+CT3xN5+9c6Kp+h1S2/wr+a/HzZI4r+rPAr/klq/wD1+l/6RA/K+N/+RpD/AAL82fZn/BBHxR4V8B/8FH/D/wAQPHevW+l6LoHhfW9R1XUbt9sdvbxWTs7sfQLn69B1rnP+CqP/AAVB+Jv/AAUh+Nk+rNPd6T8OdDu5I/A/hRpCFSLJH2y4UHD3Uo5J58tSI14DFvmHTtV1LS1n/s2/ltxc27QXIikK+ZE2NyHHVTjkd6rkknJr9Glw1ganFDzqquaooRhC6+Gzk215u9r9EtN2fOrM68Mt+pw0i5cz89tPlYfCu38qJpoYl3Syqo9WbFeufsNfsa/FP9vT9onRv2dvhMIoLq/DXGq6vdITb6TYR4867lxyQoYBVBBd2RQRuyP6F/2Sv+CM37Av7InhO10vw/8ABLSfFmvrCo1Lxh4106HUL67kwNzKJVMduhPSOJVA4zuPNeLxl4kZLwZKNCrF1a0ldQjZWXRyb2v00bfa2p25Pw5jM4i6kWowWl319D+YVbu1nz5M6OR12uD/ACpO9f1N/tI/8Erf2CP2pvC03hr4n/s2eG4J3jZbXXvDunRaZqVmx/jiuLdVOQedrh0PQqRxX4A/8FTP+CZ3xB/4Jo/HK38D6trT694Q8RRS3XgzxQYBG13EjASW86DIS4i3JuA+VldXXAYqvLwd4oZJxfifqig6NfdRk01JLflkrXaWrTSdtr2Zvm3DWMyun7W/NDq109Tk/wBgX9uz40/8E/PjtZfGf4R6g01q7JB4n8NTzkWmuWQbLQSjorAZaOXG6NuRkFlb7f8A+DkH9o74Ufte/BL9mD9o34Lao11oXiDT/EzxrMoWa1lDaastvMg+7JHIrow6ZXIJBBP5ZqcDAOPetTVvHPi7VPBOm/Du/wBeuJtE0i/ub3TdOkfMdtPcLEs7oOxcQxZHT5Aeua97MuEsFiuJMJndJKNak5KT/ni4SjZ93FtWfa67W4MNmtejgKuDlrCdreTTT/Gxg7xj8K/bb/g0yJb4J/GYE/8AM46bx/26PX4kMMV+2v8AwaYqy/BT4yy54PjHTR+Vo9fL+Lrf+o1df3of+lo9XhRL+2Yej/Jn63hVB6mhlAx1/OlBOQCx/KjLHgMelfx2frQ1VHHXrzQPTJ6+tOAPHznrTWHX5u/pQVdBwORn86F24PPYdqAxH8RP4UKRydx+7QSDBSQeaMDnr19KUqSw2vx34pccH5j1oK6DAARnmlGM5B704ZxgMetJyD17+lAhVAAHPakxwSc9KUMwA+ft0xTd525z2oAcQM85/OggZx/Wm7ju+/8Ahj6UoJ4+Y9PSgFoIAvPtSFRkgE07LDPzfpRkhjk96AuMCqM4/SlAG7j+VOyR0b9KQlt3BP5UCFVVI6968g/b+Uf8MbePMc/8SyA/+TlvXr4LEAZPWvIv2/QR+xr4/wA8/wDEoix/4FwVrR0rR9UfP8W68K4//rxW/wDTcj8m9xyQf502dtsEjE/wN/I0AHfnFJcKTBIMf8s2/ka+nvof551Hoz9kP2bFX/hnjwLj/oUNN7f9O0ddk4UEjNcl+zpGIv2e/Ayg8DwjpvH/AG7R117HBPNfLVP4jP8ARzJ/+RVh7/yR/wDSURso9KWIYbj19aQkkgFj09qVS3QN+dQek9j82/8Agq5Dt/a4il4+bwVpx/8AIlyK+b9pH+NfTX/BV2Pd+1NZTMBz4LsR+U1zXzSEwCSa+gwv+7xP4M8RLLjnMP8Ar5I+tf8AgkGpHxi8XnP/ADKcX/pZHX36QDivgb/gkCufi94xI7eFYf8A0sWvvpieDv8A0ry8bpXZ/T/gyr8A4f8AxVP/AEtibVz16dKVQB+Ypd3JG/t7UA4A+fuO1cZ+p2FUAEf0qUA56VF5hU/KAeegHWvj39rz/gqLbeAtTvvhh+ztFaalq9rI8GoeJrhRJa2coJDRwJ0nkU5BY/u1IwN5BAulSqVp8sEeBxFxPk3CmXvF5jU5Y7JLWUn2iur/AAXVo+vNd1jRPDVg2p+I9atLC2X7015cLEg/FiBXk3jH9uv9lDwZM1tq3xo0uWRRzFp++6b/AMhK1fmF45+KPxD+Kmrv4g+JPjTUdbu5DlpNRumcL7KmdiD2UAVjkgKFjwB2A4r2aOVQa99/cfz3m30gsylUay3BxjHo6jcn90eVL736n6P6n/wVL/ZUs2K2N/4ivsdDa6A4B/GRlrNl/wCCtf7O8XEPgnxlNjoRYWy/+hT1+eOx+4NEg8tcscfU10/2VhUtb/f/AMA+VqeOXHtSV4ulH0p/5yZ+g0v/AAV++BNucR/Czxq/p+6sB/7c18JeNvEFv4s8aax4os4JYodS1S4uoopwu9VkkZgG2kjIBwcEiuekvrYPtNwgPu4q9bESRBl5GOuetKnhKNCTcEfJ8T8e8ScYUqdPMpxkqbbjyxUdXZPb0AKK/ST/AIJcsf8AhlO0GemuX2P+/tfm8VAr9H/+CWxJ/ZStTn/mPX45/wCutYZj/u/zPtvApv8A12l/16n/AOlQPefHniBvCPgLXfFkbANpWjXd4ufWOFnH6gV+NlxvkbfMxZyAXY9S2OT9Sc/nX64/tLzy2v7OXjyVPvDwnejP1iI/rX5FyyHcSfXpUZWkoSfmfRfSDrzlmOAo9FCb+bkl/wC2kEqH/Cv0v/4Ji+E08Ofsk6VqbRBZtd1a9v5G7somMMef+AxD86/M+ZsLmv1a/YXSGP8AZD+HwhI2t4fDHHqZZCf1zVZnK1FLzPI8BcPCpxXWqveFKVvnKC/K/wB56swHakdQCSPWlYnB+b8xR8w43Zrwz+txF5A59O9EkUNxA1tPGJI5EKOjjIYEYIP4UqhuOT0oBIA+b0ouO3c4mH9mP9nKHp8CvC5/3tLQ/wA6l/4Zq/ZykGG+AvhI5/vaJEf6V2LO2eW705JCCDuquefd/eecsmyfphqf/gEf8ip4Z8MeHfBmiW/hnwjoNppmnWoK2thZRCOKIFixCqOAMknj1r88f+CpXiO61b9puLQjLuh0nw7boi7vutIzu36ba/RtSc9a/MP/AIKNTyt+174kSVs7baxCZ/um3XH65rsy+PNiPkfk/jjU+r8FQpU1aMqsFZaKyjJpfgjxNcHg17r+xL+0n8K/2Yta17xN4+0PW7641K1ht7RdGtIZCiKxZtxkljxk7emc4rwZZSuKZcSF15Ga9irShVhyy2P5XybOMbkGZ08wwlvaU7tcyurtNbejPv1/+Cun7PinA+G3jnjv9isP/kumP/wV1/Z+UZX4aeOSfT7Lp4/nd1+fUobOdppg3E4ANYwwGFe6/E+8l43+IV9KlP8A8Fr/ADP0EX/gr78BQTu+E3jr/v3p3/yXSr/wV8+BP8Hwh8c/Upp//wAlV+fwhbOSp/KpBCwXofyq5ZfhUtF+LM/+I2eIb/5fQ/8ABcT6R/bf/bH+En7VHhHQ9P8ACnhLxHpmp6NqbzK+r21qI2hkjKuu6KdyDkIQNuPl69K+btu7gfhUUuQ3OfrilEuOOfyp06UaS5Y7HxGe8QZnxNmUsfjmnVkkm4pRTsrLRdbdT2n9gbxA3hv9q/wlcByouruS0k9xJE6/1r9StgBxX5L/ALHrPJ+1B4FSMHJ8Qw9u3NfrSGJfgnOeleXmKtVXof034B1Zy4ZxMHsqunzhG4m0A4B7dqXAHrzQSc43GhWIP3+1eefugKF3Zx3pQAR+NIpOc7j96lOeoc/e54oDYUgcAZ/Kmgc8fypQzcDd254oQt6nNA2wWMDHPajYgzx2p3phj070nPJ3UCG4XdjHakZV24H4/pT/AOLO7t6U1s4AyetUmgDOemOD60gJLZBApevH5UmPmqQAtwMtQDjof1pQDxg0Y9P5UAISR0FGSWJyOKRwRxjt3pQGJJoAcGJXqKDknGfSkAIGMilI54IpIYvIPTtSAkEf0pcDPUdPSjHzDnvTAaxB5pSScimkHinAEnFA+h8c/wDBfuISf8EmPikXH3H0Zhn21W2r+arcAc4r+tv9qn9mb4a/tifATxB+zn8XhqA8P+I0gF82lXfkXCGGdJ42R8HBDxr1ByMivjOH/g2O/wCCaSJtmm+I0rd2bxgAf0hr9y8MfELIeEskq4XHKfNKo5LlimrOMV3XVM+I4k4fxubY2NWi1ZRS1fW7f6n8+W4bevFNZ8Zr9Kv+C6P/AAR/+A3/AAT4+EPgP4t/s1WviWXT9V8T3OleJ5te1oXYhZrfzbXaPLXYCYpwT3+UV+aJJIr+g+H+Isv4oy2OPwV+RtrVWd07O6TfrvsfAZhltfLMS6Na199PM/aj/g01+HXh+L4dfGn40SW8b6vceJtN8PxTMvzw2sVqbplU9g8k6k+vlL6Cv173kkA1+Bv/AAbPftv+Bv2d/wBoLxR+zR8U9eg0vSviiLOTQNQvJgkMWtW4kRIWY8J58UpQE9XijXq4r98WUhtu0gjggjpX8qeLGExmF43xM66dqnLKD6OPLFaejTT80fqvDFajWyemqe8dH6/8HcazHOK/O7/g5q+H/hjxP/wTePjfV7eP+0PCvjfTLjSJz95XndreRR7MkhyP9kelfohIpAJY9K/Gz/g6W/bT0LUtO8I/sG+B9Xgub621KPxL48EMoY2YRGWxtHx0kfzHnKnkKsRI+cV5nh1gsZjeM8FHDJ3jNTbXSMdZN+TWnq0uptn9ajRymr7R7qy9WfjsrClkI25NMQYHNfcX/BDD/gnL8Kv+CiPx78Y6D8ebPV5PCXhTwit1KdHvzbSHUJ7hI7dTIAeNiXLbe+wdhz/Z+e5zg8gyqrmGKv7Omru2r1aSSV1q20tz8gweDq47FRoUvikfC8hHav26/wCDTYLH8BPjAxwCfG1h1/68zXuln/wbb/8ABL2yXE3gDxTcn/p48YXORx/skV9F/sYfsAfs1fsD+Htb8Mfs2+Fr7S7TxDfRXeqLe6tLdF5I0KKVMhO0YJ4Ffzvx74n8PcTcOVcBhI1FOTi05RSWkk+km/wPv8j4ZzDLsfGvVcbK+zd9V6Htakg8YpQc9+1Nxz1z+ApR9e1fgJ9yOGQBz+tNOT+frSrk4x6jqKQgn8DQAhzQM4JDdqUj3H4ChB1xjp6UAOXI4yelDE4PI60YOR0/KkAPp39KABfqOtDde350oyR0pOe/6CgBMtgc9qYu4Ak4p+0kDp0pApH/AOugAyQc5FO7cjtSfNnr3pevI9KAGgHnjnFK2Q3P86ULn8qQqQetABz1xSAHPfrS7fp+FAHzdqAFTOMH9a8k/b5Tf+x34+X/AKgycf8Ab1BXrRGR1HX1ryX9vJ8fsgeO0I66VCM/W7gFaUf4sfVHgcV/8ktj3/04q/8ApuR+TbR4c8d6WWL9y5/2G/katPGuSQB1qOVMwuAP4D/Kvpb6H+elVbn7D/s5SeZ+z34Gb18I6d/6TR11753H61xf7NJ3fs6eBGx18Iacf/JaOu2Od3HY18xP42f6M5O75Th3/ch/6ShmG4yKFzn3pwX6dPSlUHOOOKk9I/OX/gq5k/tP2XPTwfZ/+jZ6+Z36mvpn/gq5x+1Ha5I/5FCy/wDRk9fM7EZODX0OD1w8T+C/Eb/kucw/6+M+t/8AgkB/yVrxmcf8yvB/6VrX3wScde9fA/8AwSAUt8W/GZHQeFoP/Stf8K++GBHevKx6/wBofyP6g8FnfgCh/iqf+lyF3HPJ4x6U0se5B5pDnJ9vanLk4G7NcNz9WZ4L/wAFFfjtqnwQ/Z1vF8Mag1trfia4GlabcRtte3R1JmlX0ZYwQD2Lg9q/La2h8pBGowFGAO1fbv8AwWGvLuTxD4E0befs4s7+fYDwX3RLn8v518ZR2R/yK+gy6ko4fm7n8X+M+cYjH8a1MNN+5QUYxXrFSb9W39yRFbQXlzcR2dlbSTTTSLHDDChZpHY4CqBySTgACvtv9nf/AIJTNfaRa+Jv2h/EtxaTXEYf/hHdJZQ8IPRZpiD83qqDj+8e3jH/AATz8J6V4i/az8MxaxbpKlp9ou4UcZBljiJQ/UHke4r9RmzuPHfpWWPxdWnJU4Ox9N4O+H+ScQYOrmmZQ9ooz5IwbfLdJNydt90knpve/Tynwz+wr+yZ4ZgWO3+C2m3jAD97qsklyze53tj9K6ez/Zt/Z4skCWvwI8HKAOM+Hbdv/QlNdhk4HJ9qehb16V5ntq0t5P72f0dh+HOH8JFRo4SlFLtTiv0OXg+AvwKhmVk+CPg9TuAyPDFoD/6Lr8pfjqtnH8afFkWn28UMCeIbxIIYIwiIizMAFVQAAAOgr9Yfi58TPD/we+HGr/EvxTdrFZ6TZvKQTgyydI41HdmcqoHqa/HvWtZu/EOsXfiDUCPPvrqS5nAPAd2LEfma9DL+dtybPwLx7qZbhqOCwdGMY1Lym0klpZJXt3d7ejKhY7q/Rz/gliT/AMMqW4/6mHUP/R1fnF1Nfo//AMEtAF/ZVtuP+Y/f/wDo6tcwd6HzPkvAp/8AGby/69T/ADge1/GfQ5vE/wAGfF3hq1j3zX3hm+hhT+85gbaPzAr8dGkz82OvNftjG+CCwGM/MCOvtX5A/tHfDC9+Dnxx8T/D24tykNlq0r2BIwHtZD5kJH/AGA+oNZZXNXlH0f8AX4H2n0gstqypYHHxXurng/Ju0o/lI4aRjjpX6ef8E1fFEfiP9kLw7ZGUNLot1eadKoPI2ztImf8AgEq1+Ybg4619df8ABKD45WPhbxpq3wR1+9EUXiAreaOZGwDdRqVeMe7x4I9fLrpzCm6mHuump+d+DWd0cn43p06ztGvGVO77u0o/e4qK9T75Pc560vPJHr60nsM0uOevevn1c/tW1hUXoeffApW2opZjgAZJPH86VV+Xn9a8j/bm+N1l8DP2cde1sXUa6pqts2l6HCzfNJcTKV3AdwiFpD7L71cISqSUV1ODNMxw+VZdVxtd2hTi5P5K9vnsvM9LTxN4cnO2HxDp7c84voj/AOzU59d0SFfMk1uyVR1JvIwP/Qq/EWLTIYlCpEOBjOKeNIe9YWtvaCWWQ7YowuSzHgD8SQK9L+zHb4vw/wCCfzx/xMBW6Zen/wBxP/tD9wLK8tb6BbuxuIp4X5SWGVXRh6hgSDX5vf8ABUjQrrSP2qm1iSPEOreGrKaI+rRtLE38lr7w/Z0+Ho+EvwN8KfDfy1VtH0K3t5UUAASBMv0/2i1fOn/BW74XTax4E8OfF7TrYsdGvHsNRdR92GfDRsfYSLj/AIHXNg5ezxVumqPuPFfBYrOfDmVeUOWpD2dVx3t0kvPlUn9x8IFx6ivo3/gmj4a+Gvjn4x6r4Q+JHgvStZWfRGnsY9UtEmCPG43bQ3AOG/SvmxSxr0D9mP4ur8D/AI7eHfiPdM32K0vfK1ML3tZQY5T74Vt3/Aa9ivGU6TUdz+WeDsxwuV8UYTE4pJ04zXNzJNcr0bafZO/yP0uf9lH9mRyWb4AeEck/9ASL/Cm/8Mo/syMCh+AHhLrjH9iRf4V6BZ3dpqFpFf2Fwk0E0SyQzRNuV0YAhgR1BGCD707bjrXhe2rL7T+8/u1ZHkMlzLC0nf8AuQ/yPO/+GRP2X3OT8AfC34aUo/lSp+x5+yw4yfgB4WP10xf8a9FUZ6fnTwCF4PfuetJ1qv8AM/vZP+rvDzeuDpf+C4f5HnUX7HX7KZ6/s6+ECfVtFjP86e/7HP7K2cn9njwgPpo0dcd+2B+3X4Z/ZIv9G0BvBsniLV9XikuDYw6ktv8AZrdSFErsVbO58qoxztY9q8ah/wCCzsDf679m98e3ipf/AJHrSlTxtVc0b29f+CfD5vxP4V5Fj54HGxoxqQtdew5rXV94wavbzPqLQP2XP2cvB2t23inwt8D/AAzp+pWUols7610pUkgcZwysOh5/Wu3A45UflXz1+yj+3zcftU/Ee68C2HwcfRrey0mS9utRbXRceXh1RE2CFfvFuueNp619B9eBWdaNWE7VN/W59hw1mPDua5Z9ayVR9jJtXjBwTa0ejjF6bXsBJPHXigA46HpSEH9B3pQPU9vWsj3wBIbk96N3X5h16ZpQuW7daUgkHBHWgBASSOR07GgdMfzNLyDzj8qaPb07UAOB6DA69zQGIB+nekOc5B7/AOFKuTkgn6UAJk7s4pHLcEHjPSlwc5oYNzx0NACduvegDLYzzSBSO36UmDnpxQA7jjPp60uBjpSbSCDijB6hTQAFVz26etKFGTxzgd6YwPb0pcYbGOgFADgMr0pWA3dR19aaBhehoOd3TvSvqMfn+VJj5sY/GmYycgfpQOvC9vSmAp4wMZ4pVBPX09Kbjpxjj0pQBjp270Ah3sDTSw9aAo4IHvTXBB4H4076CPFf+Ci/7I1n+3J+xz4z/Z1e4jg1LU7Fbrw3dSfdg1S3YS2zMeys6+Wx7LI1fyu+KfC/iDwT4n1HwZ4u0W403VtIvpbLU9Ou02y2txE5SSNwejKykH6V/YXg7SAO3evza/4LP/8ABDW2/bKuLz9qH9lm3tNP+KCwA69oUsixW/ilUXCuHJCxXgUBQxwsoADEMA1fsnhNx3heHMVPLswly0KrupdIT21/uySV30aT2uz5HijIqmZQWIoK84qzXdf5o/AsfL04+lfcH7K3/BwV/wAFD/2YfCFp8Pb7xJofxE0OxiWLT4vHltNNeWsQGFjW7ikSV1A6CXzCOgIAAr40+IXw98f/AAi8a3/w5+KXgvU/D2vaXMYtQ0fWLN7e4gcf3kcA49COCOQSKyVcnv8ArX9OZlkuR8R4SMMbRhWhvFvX5xa1V+6ep+c0MVjstqt0pOEuv/BP0U+N/wDwcz/8FAvif4Wm8MfDrw34L+Hr3EZSbWNDs5ru9QHr5TXLskZ/2thI7EV+fPiTxL4h8Y6/e+K/Fuu3mqapqV1JdajqWoXLTT3U7tueWSRyWd2JJLEkmqRbAzmnWFteanexadp1pLPcTyBILeGMu8jngKqjJJJ4AFY5Pw7kPDkJLL6EaSe7W79ZO7svUMXj8dmEl7eblbYY7BBk9B6Cv6Sf+CCv7CWs/sU/sQWmpfEbRHsPG3xFu11/xFa3CbZbKEpts7Rx1DJES7KfuvM47V8af8EWP+CEHjOz8daJ+2B+3J4MOnWWkzR33gz4d6nD/pF1cgh4r2+jP+qjjOGSBvmdwGcKqhW/aOWd5VG4k5PU1/P3i/x9hc1SyXLpqVOLvUktnJbRT2aT1b2vZLZn3nCmRVMN/teIVpNe6u3mxkmCMqR7c0gyTj8qQr0GPxoC4A+X9K/BT7m9hSO/vR14HpTcHOMd6MEHIFAiRB0wDQQD69aaoPAx3pQD6HrQNDiuef60igFSAB0pNrDHFIBkdO1AWFYegNABAzjvSMuSOOntSMDg/wCFS2NIeowDx3oIyTx37GmAE5OP0pdpB+6evpVEseFyB16UgUkdO1IoAA+lJ24Hb0oAcV5wR6UDr0PSmkHPI/SjGPyoAeo9B2pWXJzimLxkDI4p2CCeKAEYEnPtTf4sE9KCP9n9KTBLYx3oAcMYPPevHv8AgoBN5X7Hfjg562dqv531uK9gCEjOK8c/4KDQn/hjfxvt/wCfayP5X9tWtD+LH1PneL7/AOqeYW/58Vf/AE3I/K/cT3pJCDG+f7pphfHekZzsYZ6g19E9j/PibumfsJ+zOAf2cvAZH/Qn6cf/ACWSu26knHeuC/Zbl839mnwA5zz4O04/+S6V3eDnp3/wr5up8bP9F8l1yjDf9e4f+kocwGBx2pVxkZI/PrTMd8fpS7cgDHQ1B6nQ/OL/AIKuuf8AhqmAHp/wiFhj/vuevmhmOTzX0p/wVdYf8NUWw4z/AMIdY8/9tJ6+aG4HNfRYJf7PE/gnxFd+Ocw/6+yPsD/gjwN3xZ8bZGceF7b/ANK6++JOuADXwN/wR3z/AMLV8atj/mWLb/0rr74YEnn0715eYK2I+4/p/wAF3/xgFD/FU/8AS5DSnPTqBTkG7AFNKktwOw7UqLtHI/8AHa4T9XufJP8AwVr+HN5q3gLwv8SrSFnXR9Sls7xgudsc6gqx9t8YH/Aq+EkjC9q/Y/4k/Drw18WfAWqfDnxdbGSw1azaGXb96M9VkU9mVgGHuor8qP2gvgN45/Z48fT+CvGdk2wsW03UUjIhvoc8SIfX+8vVT+Br6DLK0J0fZvdH8keOfCuMweerO6MW6NZRUmvsziuXXspRSs+6ZR+DHxS1P4K/FHRfibo9qJ5dKvBJJbltoniOVkjz23IWAPY4r9VfhV8Xvh98bPCdv4y+HPiSC/tZlBkiRx51q+OYpo87o3HQgj3GRg1+PrP6VNpWq6ro2oJqmiapdWVygwtzZXLwyAdcbkIOParxWChiHe9mfLeH3iTj+B1Uo+yVWhUd3G/K1La8XZ7qyaa1stV1/aHBzgHn6Vw/xa/aY+B/wPsZLj4ifEPT7e6RCY9Kt51mvJjj7qxISwz6nA96/LTUvjN8XNVtfseo/FTxJPEVKmOXXbhlI/775rj7pV8952JMkjFpHY5LH1J6muWnlet5S08j9GzH6QVaVFxwGCUZvrOd0v8At1JX/wDAke0/tg/tl+Mf2o9ei04QPpfhbT5i+naMsm4ySDIE8zDh3wTgfdXJxkkk+OKSV4Nejfsr/sr/ABA/ah8dxaJ4ft5rXQ7Wcf27rzx/u7SPuqE8PMRwqfi2FBNbH7dnw38MfB/9pDVfAng3SEsdLtLGxNlBGOiG2TLE/wATFtxLdySa6+ajTmqMOiPyDOcFxJnOWz4nzJtxnUUFKX2m1J+6tlCPLbTS7stnbyLkV+jv/BLU5/ZVtTn/AJj2of8Ao6vzhLArkV+jH/BK5zJ+ytCn9zxDqA/8iZ/rXJj7+xPs/AuS/wBd5L/p1P8A9KgfS6nAwT+dfJf/AAVF/Zxn8ZeELf48+FNPMt9oFv5OuRQplpLLORLx18tic/7LE/w19YAcfjSyRRXEbwXESyRSKVkR1BVlIwQQeCCCRivLpVZUaikuh/UvE/D2D4oySrl2I0U1o+sZLWMl6P71ddT8U1XcKs6bfX2i6pba7pF7La3lnOs1rcwtteKRTlWUjoQa+pv20v8AgnrrPw61G6+J3wM0WW98OTM017otqhebSj1OxRkvB3GMlOh+UA18qtGc4PBr6OjWp1qfNFn8IcScOZxwpmksHjoOMk/dkr2kukovqvxT0dmj70/Z2/4Ke+A9c0i28P8Ax7Emj6nEgR9at7ZpLW4x/E6IC0THvgFe/Fe/af8AtN/s36rZfb7H4+eDWi25Jk8S20ZHHdXcMPxFfkayYqGf5/vgN7kVw1MBSlK60P07JPHTinK8JGhiqUMRyqylK8ZfNrR+tk31dz9Pfip/wUX/AGY/hvpUs2jeNIvFF+inytP8PfvQ7f7UxAjUe+SfY1+f/wC0t+0p4/8A2ovHKeK/GLJbWVmrR6Po9sxMNmhPJ5+9I2Buc9cAAAcV5+Mk4OT6U0jBOBW+Hw1Og7rc+Z4u8T+JOMaKw+IcadG9+SF0m1tzNtt2+S62uIsYPJFfQ/8AwTp/Zzl+MHxli8a65YFtA8KSJdXLuvyT3XWGEeuCN7DsFH96vP8A9nP9l74mftLeLE0HwdYNb6dBIP7W1yeI/Z7NM85PR3x92McnvgZI/UH4O/CDwZ8Cvh9Y/DnwJY+VZ2i/vZX5kuZTy80hHV2PX04A4ArLG4lU4ckd2fS+FHAGL4gzOnmmMhbC0ndX/wCXkk9Eu8U9ZPZ25e9uozyfrxnqKw/if8OfD3xe+Hes/DTxTHusdasHtpSo+aMkfJIv+0rBWHuoraUc/wD1qkTjg/hXiJtO6P68r0KOKoSo1Y3jJNNPZpqzX3H43fE74X+Lvgz4+1P4b+N7MxahplwY3YDCzJ/BKnqrrhgffHY1gnr0+lfqd+2F+x74a/ai8Kpc2c0OneKdNiYaTqjp8si9fs82OTGT0PVCcjIyD+Z3xM+Gfjr4Q+L7nwP8Q/Dlxpuo2zfPDMvDr2dGHDoezKSDXv4XEwrwt1P4f4/4AzDgzMZWi5Yab9yf/tsu0l+O66pfQf7Gv/BQy/8Agrplv8L/AIt2V1qfhqE7dOv7UB7nTVJ+5tJHmxDqFzuXtkYWvuD4d/Hj4M/FezW9+HnxM0bUwR80EV8qTxnA4eGQrIh57qK/IAAkZHemyRJOw8+NXx90uucfTNTVwFOrK6dme5wr4ycQ8NYOGDxEFiKUVaPM3GaXRKSvdLpeLa720P2qu9T0nTbVrzU9WtLaFBl5ri5SNVHqSxArw749/wDBRL4CfCHSbi08Ia9b+LtdVSsFhpE2+3R8cebOMoAD1Clm9h1r8yDEpwroGAPAbnH51Izbl2n8KyjlkIv3nf8AA9rN/H7OcXh3Sy/Cxoyf2nLna9FyxV/W/oXvir8R/Gnxm+IWpfEvx9qhutU1KXdIwG1IoxwkSL/CijgD8eSSTz4j28jA9/SrZhLtwM19S/saf8E8vE3xH1uy+I3xp0WTT/DELLPb6XdIUm1XuoZTzHB0JJwXHAGCWHXOpTw8LvQ/KsnyXPOMs39jhoupVm7yk72V3rKcui/PZXeh7l/wS4+BN38MPgvcfEnxDZNBqfjKWOeCORcNHYR58gEHpvLPJ9GT0r6bA7c/nTIoYraNLe2jVERQFVVwABxgAcAe1OVeTxz7V8/WqyrVHN9T+6eG8iw3DWR0Mtoaxpxtfu95S+cm38xzDnr2FKoDDGaTbyDjtQAQ2cfpWZ7Yqrk/d+tKwwCMd/WkHB6UMpI6d/SgAI5Bx2pi89qXBJ5H6U0DnkdBQA4HI6frSrnke9R7SMfL0NKAcHA9KAJP4ugpWUHP581Go+bp0pxBxwO3WgAJ4xSZ5zinELyC3fimnZuGW7CgBc5A4P5UmfY0o2Z+9696Q7Acbv1oAQnB6UoJ3Z9hSNs6hu1Kqpk/N2HOaT2GhScjimtnd3609VTb16UMqbs57ipsUM3EcUqnB5Hb1oZU5+b9aNiA9ex71ZIde3anA5/L1puE4APb1p3yYPP60CFBPp+tNfGBTl2evb1pDsxnP60AMYnFODEDBFBVT/Ee3eghQBz37mgDyz9p/wDYq/ZU/bM0BNA/aU+CWjeJhbx7LPUbmExX1oCc/urmMrLGPYNj1FfDHxH/AODWH9inxPqT3/w9+OHxH8KxOcixWay1CJOuQDNCH/Nj0r9OTswfmHtzQNgP3u1fRZVxbxLkdP2eBxU6cf5VL3flF3X4HDicsy/GT5q1JSfdrU/Lrwd/waofscaRqS3Xjb9or4m67bry1pE2n2Ktz0LRwM2PoRX2X+yh/wAEwv2GP2K2i1b4DfADSLTWo1C/8JJqwa/1L3xcTlmjz/sbR7V74hQk89fWhlU4yarNOMOJ86p+zxuLnOPa9k/VKyfzJw+VZdhZ81Kkk/QazNkuWyxPJPU03e2fofSnFUIHPYd6aVUrw1fNnojt2RThnHApnHPzU9CuMbv88UCEKtnp3pOlPwmc7uppNqdm7UANTjGQacCMk7e9IoXjnv60DYFzu/WgY4tkcU0Hg4H8NKNmMbunvSDy8EE9qQ9AJI4NNzjODTm2Z+929aRgnPzUxAhwKXPfH60g2YJDd6cuznB7+tACBsAcfrSKxwRjtzSgKeN360iquCd3b1oEKT6CkJzmlYKSfm7+tAC55bv60AICckfkKduBzgfjSBUORn9aXCgn5u470AIcnoOlIPvD604BM9f1owm7qOtACK3AFcZ+0T8Krn44/BfXfhTZ63Fp0msQRRpezWxmWLZPHLkoGUtny8dR1z2rsflK53D2p6BQR83pRGUoyuYYzCYfH4Spha6vCpFxkr2upKz1Wq0fQ+FV/wCCO+uMMzftAWak9k8LyH+dzSS/8Ec9cKsF/aIteR/0Kb//ACTX3XhTnB7mmtgEjPcV2LH4hdfwR+b/APEHPD3rhH/4Mq//ACZz/wAJ/BTfDX4YeHfh1LqQvX0LRLWwe8WHyxOYY1QuEyducZxk4zjJrfD5Jye1ACnPPb1oAAPB/hrkk+Z3P0uhRp4ajGlT0jFJL0WiF3ggUoYmmHaCBu/WnpsI60jU+aP2tf8Agn1qP7TnxYj+JVt8WYdEVNGt7H7G+hNcHMbSMX3+cnXzOmONvXnjzUf8Eb7/ABl/2jY8+g8Kf/dNfcSMuQdw5PrT8qT1/i9a6YYzEU4qMZaLyR+fZj4XcE5tj6mNxWGcqlRuUnz1FdvyUkl8j5+/Y5/Ycn/ZP8Vaz4ll+KI14avpcdp5A0T7L5RWYSb93mvu9MYHrmvfm6U8sufvdvWo3IBGGHSsqtSdaXNLc+ryPI8s4ey+OBwEOSlFtpXb1bu9ZNvfzDIzj2FLu44/Omts7sPujvQhRl+9WZ69lYmDnua574pfCn4ffGXwrL4O+JHhuDUrGTlRKMPE/QOjj5kYeo/lW8GXf98dRSlkA69SKqMpRd0c+JwuHxlCVGvBShJWaaTTXZpnw38Uv+CSetDUJb34O/FG2a2ZiU07xDAwZB6CaIHP4rXmd/8A8Ey/2tbCUxweGtFvFHSS18QwgHk9pNp/Sv0tYLu69c0hUH8e2a7IZjiYdb+p+WY/wW4HxtRzhTnSv0hPT7pKVvkfm9pH/BMH9qLU5hFqlv4e0tO8l3rglI/4DCj1638Jf+CTXg7StQi1b4z+PJtbCsGOlaVE1tA3+y0hPmMPpt6V9ilVzu7k9KNi8DPbiipmGJmrXt6HRlXg7wNldZVXRlVktvaSuv8AwFKMX80yj4T8MeGPAPh628KeCtAtNL020QLb2VjCI40HfAHc9z1Pevn39qv/AIJ8D9pr4sf8LQi+KqaHu0q3tJbNtC+0lmi3/Pv85OoZRjHG3rzx9IFQCDnrT1AB+X09a5adSdOXNF6n2udcOZLxBlqwGOoqVFNNRTcUmlZW5WmrJ2sfFK/8EekVcD9oQZ/7FX/7qr6J/ZT/AGdZP2ZPhjJ8OJPGS62rarPeR3S6d9m2iTb8m3zHzggnOe/SvTC2BndSEqcfN39Kupia1WPLJ6eiPHyLw94R4bx/1zLsPyVLON+eo9Ha6tKTXTsIWOAAe9ODEGmfLgfP3p0e0kgn6Vgfa2HqxxuB6Hsa8Z+OX7CHwG+OF1Nrl1o76FrMwJk1TRgsfnN/ekjI2OffAJ9a9mGzaMHvTt0ZAGRVQnOnK8XY8zNcnyvO8K8Nj6MakO0le3muqfmrM+B/G/8AwSZ+KFhLJJ4A+I2i6nDk7I9RjktZMduQHX+VcJqH/BMz9rq3kKweDdKuR/fg8RWwB/76ZTX6aHYF4Pb1pkgTJPGd1dccfXjvZn5jjPA/gfFT5qaqU/KM9P8AyZSf4n5t+H/+CWX7UmrShdZ/4RvSIyfme71vzmH/AAGBGz+de1/B/wD4JQfDvw9cxan8X/Gs/iCRCGOn6fEbW2J9GYkyMPxWvrjCggjjjGM01QoGc/rSnjsRLrb0O/KvB3gXKqqqewdWS/5+Scl/4CrRfzTKPhrwt4Z8E6DB4W8G+H7TStOtVCwWNjCI4098DqfUnk96ugk44pxCZ4YdaFVeMn/Oa5G2z9Pp06dGmoQSSWiS0SXZIjVQuRjNODeoxShV6A/rQEUjr29aRY+OQqCM/rXN/FX4MfC344aAPDfxQ8H2uqQICbaWQbZrcn+KORcMh+hwe9dGgTH3u3rQWUEfMOgoTlF3TMMVhcLjaEqGIgpwlo4ySaa809D40+JX/BI3SLmeW8+FHxWe2QklLDXbTzFX2EsWD+amvIvEH/BL/wDat0qZ00vQ9E1VAflksddiXP8AwGYIRX6Ukpk4ak3Lj7/8XrXbTzDEQW9/U/Lcx8F+Bcwm5wpTpN/yTdvukpJfJI/MGP8A4Ju/tjyTeW3wsijH9+XX7IL+kxP6V3Hgr/gk18ctXeOTxz438P6JCSPMjt5JL2YD0woRM/8AAjX6DDaSeRwaVduBj+7jrVyzLESXRHHhfAzgjD1eap7Wou0ppL/ySMX+J4d8BP2AfgN8DZ4fEEmlv4i1qHDJqetKriJ/WOIDYh9Dgkete3s5JyRQ2317+tC7CTlvTvXDOc5u8nc/UcnyTKMhwn1fAUY04dorfzb3b83dgcn8OlIvB+opwCYznt60KFB69vWpPUbELc8A9KMn/PalIQd+g9aRdmfvDigQuTn8aXA7jvQgTP3u9KwTbncP8mgBhJ3ZHpTV6jHpTyqcc9qRRHnr2oATp2pNxGQQafhMde3rTGVMkhuw70ACt83elLkqcDj6U0Bck5705goXAP50ALt/zimleRk+lPIxTSMH8aAG429xSkZ6EUhPbuDSg8DGaABk9T2pVX5iP6UmBnp+lOCbWJoBDguFxn9KGUFs7vSg8rgD9KRjk9OmKB3ApzwRSbBngjpS8jNAPPXtQIaVwRyKNvGM05geMD9KUA4/CgBgXnGRQV/2hwKXBB6U3OR+FAD9owaQrkYA/WlxwcjtSAZP4UAMdCO3cUKDnoe9OKr6U1lAwfem9QHIvcA/nSgZUHHam7gOM9/WlGMDA7UhsUr2A7Uwx+1PXBOPakYAUCEZDzn14pyKexz9Ka4yc4705AOnuKBjgh7A9aQoSQf604YzjPekJBwfQUBqxqqMDBHWk2EcZFPXjoe/YU04DYyetAgVT3FIF+Xr/DS4x0FAVdpJHagBCoz3496QrnP19qXA70AKM59fSgAVMLn3pdmaFHFBxkc9qAAJ0NAT5eRRkYA9KVCO3bNACMoOTgdaTbycDvT2UZ+7TSuO3agY1Vzx7etO28nkfnQvBJA7Uo6/hQITb7ikIBfO4dRTsHbn+lNI+fp3FA0CqMdutOVeeM9KTgDHOM+tA28fSgBdnXPamlSTkYpwx60hIx1zxQIaFGSB6Um0biBTh0JPtSDg9O9ADSnpj86cF9COtLjpwfypAOP5cUAC54x6+tSY449ajHBySeKeDlfxoACvXj9KZIhBHFPAGMgUxlGcYpgNOfyApNuB8v8AOnFADz6Ck2gr1/HFIdxVBDfjTgpNN2jdnHf0p6jvjtQIXb047UmzkDFPUDPTtTXHPSgBAuQOKUxn2oJAHJHalBUAjPSgBjRkdqcEzxjtSNjoacoGPw9KdwG7CR0pGQ44GOaeAD2prLgUgIypGPr6+1CLzkfzpxVeu3HIpMKRxQA4ZwpANOUEgfTiowvAIH6U4Dp9KAHYOKGU880hwCaHPJ+tAC7eBx2pNp5xQMH/APVQFWgBCpzj+tKq8Zz/AJzQeB170Jxge9AAF9KUKc8YoGM9O9KfYfpQA0Jxjjp60EZGNp/KlHQ5Hb0oxwBt7elADSpycKetIoOOnfvTtq5Jx39KaAAKAAA+350qg4FIcDB9qUYwPpQAm0kcA9aUKc9O9Iq5GCKcFUE8DrQAbCAMCgITxjoKcQMdO1IPvHA7UABQdQQOKTbzxj86ccZwPSkwOtACKoz1HUd6cACOvQ9qaoBPPrTgAOAf4qAE28jH9Kbtz+VO68mmkDPAoAMZxj+dATOSCKMYHTvSp0INAAIxk/MKHjUDqOOlOH3utEnCHFAAVHXHemsBnFPJz3H0zTCeenWgBrAdPc0AeuetK3OMj9KF6Yx0oAAAD+FOwMnHpQMnFGTk/T1oAABjr3oZRu5PekzgZ469KCxB/wDr0AKAB+tKBk0mc/kaFPIxQAuFzx6UoAxwO1Jxjn0p2T+lADCBkHHakKjjA7U7OfX86aTjA56UALgYPHSgLQCSM+wzSrzyB260AIRxyaawBxz1NPbpimk9BnvQAzHJwf0oC4A5oAGcn9aUDgY/nQA5FAwc9qVgPShQcAe3Y0HJ5AoAR1Hp3oRRx/vChs+vf0pUB4Ge9ACqMdBSEDp7Uu360jDB4J6UAKuDimnBP40BugHrSDk496AH7V7HvSBF2k57c0uSef6UA4Gc9qAEZBxikKqFPPenN2BH6U3JOee9AAoHIBoPXihfzox/KgBcDAJFIoz19KOcAc/5NIp7YoAkwvT/AD2oYDGPb1pNxP6dqQnt7UAAA546U7AB4pqk8072z270ALtXaMUxlG45bvTt2OBTGJJ6d6BoTHBO7PNIuAeCelGSQeO/pSZOQOOlACg84yacDgcelMX3Ipw+7nHYUCHKOM03b8/GetKD/MdTQMlunegBdowMevrSBBjindhgCkGBQAm0Z6d6CMDA9aXPPBpDk/nQADBGKQgZpwU44NIQDg0AIEGSAf4aNoK5B70uTnr2pM4HJoAUqA3JzTlUAY9uxppOTSqxoAcg5/CmuAOgpQxB4Pb0pHPbHagBGHBwO1AHUY9aTrx7Uqrx39+aAA9fwpyAYzjtTSv1pwJ/DFACgCmuoA/GlBJ6UhPrQA0KMAgn73rTQAacD0A9fakAAoAUIABilCcYoBYqMClU8DigAZBzx0oYd8d6ViT0/HikYkngd6BiqBwRSgZA/DmkBwBQOnP8qBAVxx6GkVRxTmJxgetIpyRxQA3AyRSkAEnOKTHU0pY54/nQAq425yRxQVxgg9qRGwMAHgetBJJBI7UAIR15703A6805jnJx3poNABgZ/Cl2DA47UmTnjtTlzjk9qAEUcH6U4Lgkj1oVePw707HP40AIeDwe1NAFO69s00Z5AFAAeCMelKMYHNI3Xn09aUMRQAq4zkZ60pUYznJ3U1SR+dOBJ6+tABgAg5HSm7BmnAkHigHjGO1ADSgx19KFUDt0px4HSm8gk8UAHGeDSsM5pm4huRSsxwaAHAn1/Wmnrj+ZoIPYn8qQBs8k0DsKwPqPzox7ijaeOvTsaUA4xz+dACZ6YA/GlPfgZx60hDA4Genc0oByevT1oAacgc4oz83Sgg4/iprZJAye3egQ4Enn2pQTnjFNwe2elKqMGzg9PWgB2T7dBTucY44FRkNxjNOBbAGT0oAUgn0/GmsOnI6U4KT0zTdpGASaAFAO38u9KMgdRQF4xk9PWkCkHgn35oACOMH9TSYPHTrSkHHf86QA8cnrQA1VPt+VLyQMDt60vJ6E/nSEEgfMfzoAcvNIQR6dKBxjkjigg+poAGU/rTlHrj8aGXvk9fWhRgAhjQAoz7e/NNcYHanYIOMnrSOG9TwKAIwOnTrQAQce/elA6cmjbzkk9fWgA5x+FKpIPbp60bT05oRTgjnp1BoAXBz26UHoQQOvrSkDPU/nSHjPJ6+tACL06frQevbt3pATjgn/ADij5jySeg700AnzY5NIpOP/AK9O2nAGT055pgQgHk/nRpYBw6DgdB3peQOo/GgKc9+Peja3XJ6UgBST6fjTznOMjtUYB5wT05pxBLZJP50AOJJ6jtTMHd+PrTvoT+dNwQ+C3f1oAbjPQDrSjORnH50hH16+tLtOQeaBpir+tGD7dPShQ2eM04KcdT+dACKDzkjqO9KAd3brQFwDye1KFOcZPWgQhGABkfnRg4xx07mlYHA60hBXufxoAbyT/wDXpTkjPHX1pCCCOTShTjqetACgEjtQ2emO3c0YbHfrQQQep6UAB/DoO9NwQO3H1p5Az36DvSYJXgmgBrA5ycUvOT0oILHv1pQvuelAAPw6UjDp0/OlVTnqaNpOOT0oAAvfA6DvSgHHb8aNpPG49BmlxxjJ/OgBrL06dqACecjpSsDx17UgBA6n7vrQADPFIckc9aUAjpn86QqQMkmgBFBwBx19aUIc9BSgEYyT1pUB756+tADSDgcCjcQPw709gSBgn86ZtwMZPSgBCTnqPfilOehIP401gcnk+9Kc8ncetAxy5wPp60DJGePzpB2IY0LnplulAgYk8+9C5OOe/rQwY9z+NCjgfMetABt5JyOvc0HJ7DpSnrncevrSEZ6elAAoOMADp60ENxjHT1pFyAee3rTsdM56UANOec44PrSYOMY/OnFSM4J/SkCnHegAAI9KVQcDkdKQqe5PQUAEY5PSgByDvgU4D6dfWmJnHU0/nPU/nQAh9CBwPWmgdenHrSsD0BPT1pFDDJ5/OgBWHuOgox9OlDDpyeg70pBzk56etADVB3fjTsnHTvSKuOSx4680EfXr1oAXJ4/+KpBx27+tGMYHP1zSAHPU9fWgBw/Dp60nPPTt2oxk9T+dKoIU8np60ANIOeopHB2n6ZpxHzd/ekdeMEnp60AOI569/SkAIbilzxg+vXFAwW6/pSK1F7AUAcdfzo7DH8qXtjb39KYhrDBzntRjBNOPrn/x2kY8k57elAhjAFeKaeuM9MU/IK//AFqCMtznr6UAIoxn5h0pRgHrThgd/wBKDgEf0FADGGSAPSnADpnt3oI6fT0pwBHH+z/doAAO1IwOB0pRgc+3PFISMD6D+GgA5xg46DvSgE0mSOD6DtThgDj+VADSDjpTTnj61Iw4/L+Gm9xnPX0oAQU0g4FOHUjB/Kl2kgdf++aAGqTx9fWgE4x+tOwOMev92m8bc5H5UDHkkj8aUHjqO1NPOfr/AHaVTjH/AMTQIOvfv6UjZAGD27UuSPX8qQnjHPtxQA1c8Hd3oGN3WgbsjrS49QevpQMBz0FKoO0jjp3o59D79KVCcEAHp6UAgY+tMYnnnvT2+8OvT0prDgkZ/wC+aAtoRjIzk0vAOOOKUKO/4/LQVB6n9KaEGRgfSkUYzz2pVBIHGOP7tKi8HHp6UgDGTknv2pSuenp1zTiP84oPXqenpQMYAQT9KVutGCM4J6elB5Y849iKBAfb+VMO7d1708j09PSkwQ2T6/3aBiAcde9LjkY9PSjOB3+96U5QOMDqPSgBqjnp+NOxx1FAGD0/Sl64B/lQADgHn86TOGJBpc9fqP4aT5d2c/8AjtAgOBj6Ug6ZB+tLk4GM/wDfNH8P/wBagYwD5vw5p3OOvQ0Drj09qXJIAx39KADHy00kg9adkY49v4aQgfmPSgQ0k5654FKCdvApM465+7QuNufcdqYCkknGe9Lz/k0hznIznP8AdpwPqe3PFIAXOen40HJx/jQODwe3pR6Z9PSgB3IGPYUgyBwf1oLAd+gHagkAcfyoAa2OMfypR0xjtTSQcc9v7tL6j2FA7MBmkII+tKDgdf8Ax2gjjp6fw0CAc45HUU4HvmmhiMDPf0pFbJI/pQA8/dHNJyRigEkDn9KOOBn8hQAxhkn2pG59OtOYZGfb+7QwBJzk4P8AdoGICeBmlXOaAOAMHr/doTgZyfyoACOBigHGPmHWnMPT1/u0DOADnr6UCGg84z3oIPalCj0PHtShQRyO1ADADg8/w9xTgRx/hSgEdCenakOARwR06igBccE570mDjg04YweR1/u0gHy5/wDZaBiEc5B7DNJ2GPT0p3Xkn06rQqgqP93+7QA1M889qdkknnvQBx1PT0perYJ79xQIRge/8qaucnBp/U9O392mgev6YoAQnp9KO4+lBGTyD0HUUqj27elAApOeGHWlJ9+9IPvYyeD/AHaXPHOfvdMUAIQeDSLknrTwOQQD+ApFBPY8dcigAA45I60ucA8ilCnHTv6UmByP6UANP3utNY9QD+VPC5fofypGU4wAeKaAM57d6Qbt2dp60uQD0B59KaMbhxU2KuOIIwMUbuwHegEEDj9KDgjoPypkiM7Dt26ZoZmBPB+7QSv90dPSjCnIA7UAN3HGB/KlyxPTv6UhAA4FITzkAdfSgCTcTnjtQCc/hTRjnjtTkYZ6DgUABz0wenrTicDp/DTSwx0HT0pcrnOB044oAQsc4APQ00k8cHpS5B7D8qaSOMD0oAduYg5HanKzD8ulMBUg8Dj2py49O1AxSWwOvakyxxx3oLADoOKARwP60AKCeRj86DkAYGKBgc0bhgYFAhMkAHHbimnPOKfxjpUb4zgCgpDiW9D1pVJxjbTSV54HWnIVAHH6ULYT3FJPPB70053ZAp3y0hxkfSgNwXOBx3pDnnApUK9MUHBzwOvpQGgEknOM0ISFOF7UhxnGB+VAZcdB07CgQ4knHykUfMwOAetNLKSD70u8c8d+9A7gMhTxSNnOaAQQcil+XOcd/SgQgBxyD0oQkAj2pw27Rx2pqgDsOnpQA45z09KDnsvajKjsPypAVIHA6UAGG5wDQfvZxQGXByB09KU4z0FACDPp0FIA27PPalBA4HYUBgWz70AHOMbSDk96UZ4wD0prFSMj1PalDqMfSgBckHp3puTgDFAYE9vypTggcDt2oAaGbHFKpbd0oUDBGB19KFADZwPyoAd2BxSEnGAPxpcjjj9KTK44FADed3Q045I5H8VNJAPSnZUjHHWgABbB69RSEse3anYAHQU1tuenSgBMsDjH8PagE44Heg7M4/2fWhSFHHr2NACtuyeO9GWGfpQSCfx9aMLzx2oAVS2R1pGY9MdqMrkDAprFeDgUAOLHHQ9KQkk9DSYUdhTlC5wAOKAvYY27GMdqdkgHjtxSsFxwB2o+X0HQUDuIpI6Zoy2AAD0oUL6D8qCVxjH5UCDnAGDwaQE5zinjaQB7igIAc4oAZuYKBg/5xSgtxx2oIGPujp6UoCk9BQAnI4obIJ47075eeB+VD7SSMDrQO+gi54470KTikUqAOPShSuMY/SgBzFjzg9fShSQPoaMg9h17ChSvHT8qAFGe1KQfT9Kbnnp344pdwzkY6UCEDH0PSgk8YU9KTcuDn096TAyPpQA7c+DjPWmgtyMGlLKc8dKTcuOg6+lAxSTxwelKrfL93tTcrxwOgpQVwMAdKAYqk4+72pSSOcU1cYJx29KX5c9B19PpQICSe3RaRSeQRSnbn7o6elNAHPyjj2oAVj6DtSZ7Y7UNtwMAdKAQKAFBbd909aU5I6c7qRWG78aXIYcAdTzQAozxhe1C5zwvek3AHoOlKrDOMD8qAH/NgDFJzg/y5pQRgcDr6UblIOMdu1ADcnPT8aVumPekyA2RStJgHnvQAw4J6fpTQeRxS7TyaQA5xjoKAHZzgYo+opvII47mnDPHBoAQ4B4A6UoHP3R0pCPbtTsMM4/ujvQAjAAcL+lNPXkDrTjnb/8AXprAknjvQAmQOw/OlUgH+VIynuKMHIOKAHEj+72pS3fFN5xg+lDA8/WgBQSewNNY8cClAPTFIVOMe3rQMdnjIHanD1xTGBweO1OwegoED4x0pAeBlR1x0pWGV4Hakwf/AB6gBRgdqM8DgdKRVOSc9vWgggDjtQAo6D5aaxGCadgkYx6U1lO38qBrQRjk9O9KrDjA9KYynBzTkDE4/rQA/POQKCRwMUwbscfrS4Ocf1oEOU8AY/GjIOeO9AU8cd6TGRQAHGcYFNBwMY7etOCEgYpNjYwPSgAJyOFpN3X696Cp/T1oIOG4oAFIHYdacDn+Gm9unelXOeAetAEikYHFNUn+7SDIxx2oQkrQAp5/hHQUhOD07UuD+XvSFWzmgAUjJ47UpOTnFNUEAnHalwcng9RQApPtTQcN0H3qUhie9JhienegBVJOeO/rSZ56frSbSV6UYPFACrxzt7U5WyOlMAJPTtR8wOMGgZIpABGPSm8BuBSAnnHtQWOTz2oB6C5GBxQTkYxzjvSHPAB/WjaxoENJ5GAKerDA47+lM2k8Y/hp4Ug/8C9aB2HgYHSmsBn7opec5/rTXzxgdqBbASM/d/hFITxwvekILcewoCkDHv60AOyCeg/KlP0FJg5yR3Hehs549PWgBM4I4pvUg7e9OKkkUm0gAHsKAHD6U5TgDAphBHHofWnYJAxQANjGMfpSdc4HajB/SlKk/lQA0A9QKMnH3R2pSpPQHpSFTQA9SCANo+9TiRj7tMXcB17+vtRlieaAEOMfdHSlHUcDpTWJGOO1LzxweMUAKeOcUSYx0pDnbjFKQSDR1AZ7ADpSp9B0pdh/SkCtjI/PNADuOmKExgcUmD+lKi5wP50AGAeMUH6dqUKSKQr29qAG884Haj/gJoAbH4CkbPGP50DF3cHj9aQH27+tDDr9PWkx1oFcXr2pwIKjjtTAGyeO9OjDbVx2Hc0AKpwM4FKDnHA/KkAPT0o2kk4/nQApx6Ui4zwO/FBBxyOgpBnPTtQArHJ6dqTPI4/Okwf0FADHtQAq9c7Rwf8ACnDOOlNGc/iKVs4/GgBc/N0/WgHJ4A6UmTkcfrQuSOh6UAPB4xj9aARg8dqaAeuD0NLyM8dqAEOA3SkdwOQO/egk5IPoKbICeg70APGQThqbk5HP1qXB9M03ac/doAZyCBupyk9mHWgqBggfpQBx0oAUkgjBpNzev8NLtORwaNvP3T070AM3uF+8fyoYndj3pSg/u0Ffm6d6AAk9M0hzuBBp5Xrx270m0Z6dqAGEkgZPb0pQSep70MoBG2jb7UAKPqaazHoD2pwA7LSFfQduKAF57H0pSSMf0pdg28DtRtHTFACOxx94+1Cs3r3pGXjAFCjAA9z3oAVCwGd360HOAQe1AGegFAUlRhT0oAXJx15pj529entUm3/Z/SmOuRjFJjuMYnnmhSc8N/nilK5OAO9AXAH+FMQikkYz/ninZI/iPSk2gdAOvpS7OOnagBVY8fN3pcErjdRGp4wP0pwGD92gAGeKbk46noKkK9tv6UwJkHK9qAGHIx836U1icHmpWjHYdqY0ZOQB3oAaC2Tz3pyk+v6UKhxnFLt5PHf0oAQFsA7u1IrEjr3p4UYAIH5U1FyPu0DFDNn7wNGWz1H5UpUZ6fpQE4+72oENBbB5/h64pcnP3qFQZPy0pUA5xQA3ce7fpQGO7rzmn7cjpTdnzcjoaAG/MFzu796XJBBDU7aBwF70BQT93tQAwEgjntRzj7w60oUg4C0AcZxQCG/MM4fuKQscnmn7Ac8U3ZkkAUAKGIIw3TH9KUE44bnHpSFfSnBeBgdvSgBoznGe3PFOyc4z/FSbenHanBeM4PX0oGrgC2eM0jHOOT0p4UEDg0hTnhaQMZzk/N2oG4DO7uKkMYz93tTSo7DvTEN3HOB60ZYnP9KUqd2QD+NKEHcfnQAgyMc9vSkJPHPanBBkcdqRlGRgfWgBCx/velKC35UhXAz/AEpypxjb+lACc4/Cnbj6+lNZcDAFKBkYA7UAKM9QRQQcUKp6AUFcjgGgBuWH8R6/0pu47id1P2ggZHf19qYVyeB+lACFjjj0pcnsfSm446Y4pwT2oAXJ2nLGnZJzyabs46fpTmXA6UktQAE8c9qRScde3tS4/wA4pQo/u0wELN/e6ChcjHNG32+tCgCgBQTjOfzoyfXtR0OAKGBxjHagBgJxjPakJYkYb60Fc5+U9KNuOCtACbjzyaUFievegJkH5TTkXjOKAEUNjr3pVLEDntS7RnhaCgwOB0oAau7b1oy2Tlu9ATA59KUKPb8qAELEHhqapPr2p5TngUioAcbe1ACc54PagZ9f1pxQ8EL2o2k/w0AIpO7r39aCxxy3ejbh8YPWjaOm3v6UAJuOR8xoQk9+1Ls5Hy9qFTB4U9KAHA9sjpRuPODRg44FJjg5oAaS2c57ikbPUtj8acVy3A7+lI44PHNAEu7sR+tJxngH86QYGeT1pON3X8KYCsRkDbQOexpoIyBk/nSjBPU0gHADI+U0bQDwD09aaxGOvagkDPJoAUAFeQaCADjB600HcudxpcDdye/egBW+XPBpMnPT170h5IxnrSHAxgmgdgY89+nrQTjsenrTD65P3e5pwC44JoEKOP4T0oznHB6UgxkHJ/OkOMZ3H86BolDccg8Ad6UHHQdvWotxGcE9u9PJHYn86BAwGAdvcdqao6DHfsKU8gdeopAowOvWgBy+hU/iacMADC9u9NQLk9aXAGOtADiRwdv4GmPyM7ewoBGB8x/Om8bM5PA9aABxwcqevrQBnA2/rQyg5Oec0Ii8df8AOKBgM8/KevrSgkY+U9B3pqgEc5pcLxyelAhy9B8ppQ2CTjvTFIwPmPUUmeMZPWgasSg7hkKeR60oAIOFPSolIwMZpUYFTz2oESN1+6elNIBJO3nPrTdwyPmNLgc896AFVePu96Tv070qhQAeabx70AG4YHynp60iEYPynp600DIBJP50KBgnJ4B70ASEknhT1pcjrjt603jJ5P50mQOdx6UAOGOm3tTs/N/9eosjnBPSnoQWzuNADgARwpoIG7Ow0gwR1P5UEANw3cUDA4AOF70YGAdp6UwtgZ3nr60owSOe3rTEIFGeFNAGTjb0oUL2NAC5Gc9KQxyjrwe3egqucY/WkXHPzelOGMkk0CGsDjlT07GgZ5wp6nvTiFO3k/nTcDGeaBjQPmyAaep4+6fvVGAoPXt604KARz0agCVQMYwcig4B+6enrSKB7/lQ3J/DrQIUgDop6CkPI+76daDhe56etMBBXOT+dACsAG+73oJAz8v60xvvd/vetDMAeGPT1oAeGGec9PWkY8YwenrTAQCMN29aUEYB3H7tADj7Keg707GP4T+dNOPU9aUY45PSgBGPAwpoX1weB60hAIHXpSqo9T0HT8KAHrz/AA9vWg9M7TTQQBgE0h28HJoAXt909RTRknGDSDGARnrQNpYg5oGAAAACnpTuB0XtSBV2j6UYUAcnoO9Ag9wD2p34HrTDjbnJ4pcAjO6gY9eMDB/OlX1Cn86auBjBPX09qQbcZ5oEKenKn86FOCCAevrSMRnnNIvQEA9e1AxwYnJwaBjAwD09aZ0GMtwaVcHvQDFUDoUPSjZ0+Q9KaoGOvY0oABA56DnFAhwTr8nekAAX7lLhTn60mBjNACg88j2zQpyAPb1pgOT1PWnRnKjJJoGKBwcD9aUck8d/WkXBXqfrmjjceT19aBDvw7etNGAT8uOPWgEYwWPSk45wx/OmApOTjb2FKBnsfzphIB5J6DvQpB7nrSAdjn7p/OkI4OFPtzQQC2MnqKbjA5B688UD6DwvOdp6etIAB0Xt60mAMHNChfXt2oEOBI42mkB5Pymk+XPOelA284z270DQuBnO39KRwcEAenOKAAWPXrQVBx16dKA6EmeOD3pATu6/pSEnOCD160mTmgQ8Y45ozxnNM3EEYH6U7cSOPX0oAGOT1PSjPJwTSMSD+FIevSgBy5Vfvd/SlJG7Ge9Jk7Tx370hbnA9RQA7Pqf1puMkUb+Rz3pAxyP54oAQg/3u1OHXGaac8cdqUEj8hQAqg5ByeKac4wD2pw6Dg9KRgcYx2oAAeoLUqHHc03Jwcj0pQcAGgBe2MelKMAADHWmnAoBAxx3oAeDg5zQSCAcj8qaG9jSbicUAPBJwCe1JnGSD+VIG6YFKW4yB270AKc9ATSrnHJprHgk+tAYAY96AHEYPDVGxw33qXOe3pTXyTwPWgBVYcDcfypMZzzSLnjHtmlA45BoAD1wDQrAAnPb0oIyM47UKDyMc4oAdzxz+lOVuCcnr6Uwk8YBxS84xj9KAFB44J4pG68E0Bc9fWkPPOKAF6Ac9qF6Hn9KQZ2jAPT1oXpyO1ADyQSeaRhgcHtSbjzilyT1FADACScH9KUkg9aTGc8dqGJDEYNAD9xHBPb0pQSW+939KYCeTg0A89O9AAScHDd/Sl3cgbulNB+XoetLz0APWgAVuc7qdjpzTFyTwO1Kp46dqAHr9fSgH5vvU0NwaTdg8jvQBKO3NIeV60gc8ex9KQsduAKAGtjPXt6U4Hjg96jJOenYU4HoP9r0oAkXhep60pwep7U1WOMA96Mk0AKTzwewprEquAe/pSnrx6Uxjx3oARiC2SR19KXGeM01jlj9aUE5xjr7UACjkfN0FL+PaheT/APWoIx+VAD+T39O9KvTr0FNJwMfTtS5zxj9KAAj37UDj+Lt3oJJxj0pMnH4elACDg8H9aQn69PWjcR69aRumcdKAFzwAGPX1oBwT8xpB24/ioBPYUAPBOBz29BSjJPWmdBwO1OBxg4oACDk4PelbOSAaYxwM4oZuvNADwenzd6TJI5Pamhs4x6+lIGyvQ/WgBzHkn39aRcYx/WkYk8EHrQueMk9aAAgZODS5+bOaTGecHrRnIzj86AHIeD856UuckfN6UxTgEY7Uu7GOCKAH7hg5Pf0po5HB/SmluT1/KgHI6HrQAp69acBhRz2puCedppVzgfKelACjp15x6UE/MTn9KaCcdD0oJ68H86AHEnpnt6U36N29KCTnofzpFPtQArDGPpQDgdaNvPC9hRg46dBQAoOW+8eD6UEjGOfvU0tg9O/rSFh7/eoAcGwQM0A+jD8qaDkg4PFKoOOhoAd1Awwo9eT2pMnsDQpPIwe1ABjk9fzpWAwQB9OaME0MOcbupoGLgf3u/emkc9aUE46D86T+LJIoEIeoGfXvSjpwe/rTc5IwV6HvSgk8YHX1oAUD3H507ABPPFId2egpSW56dPWgBCeM5xSMCD+NC7ypzj86Gzu6DtQAZIzyOlCg5xjtQc5wAKMEMOB+JoAQgDjHanDGPwFIQf7vb1pV3Ht6d6AFAA456etIwGOB29KUbscAdPX6UjZwOB0HGaAADjk8/WhScdaRiQPu/wANIGYY47+tACkHAoB6c96DyB8o/GhSeOn5UAA69aMZA/nSJuyc4pdpwMhenpQALnA69KMjPGaAOAcDtTTwvbp60APYjHB780gPHXvTWY849fWnJuIHT86BjgAcnPf0pGXOMelCgkZAHX1pTk9s8UCERRxz39aXHXBpBnjIHX1oGeuB+dAClQelAAI4z0oUnHQfnQpyMEDp60AGMkUEYBP9aRsgjjt60pJwcjvQAD2NBAPQ0mWAPy0At/d/I0ALjAHPamgcfhSgkgDb+tNGdp4HT1oAcQCenp3pTjGM9qT5snj9aTk9u3rQAqgetDLk5Hp60i5GTjHHrThnPA/WgAKjsR0pABu696AWx939aOQ2do6+lAABjjPeggcYb86Ri2DwOvrQGbg8dPWmgAAZ4PajjANC5Pp09aDn0HT1pAKOmc46daT+LPvSDdngDqKTnd0HWgB/HAz3NHBGM035uBx+dOXJU8frQA3Az1pxHHB/i9aQg57fnQc5xx971oAcvK/jS9xTVz0wPzoOQQcCgBxb0Y9KQjK8Gglhxj+H1pis205AoAVhzSAkUH5j0FBBHOB/k0AOTrnHahl9B2pAWB6frSnPTA6etACnnv6UoHGM96Rs4+72HU0vPdf1oAG4A5PSkONuM9vWkYnHQdO5pSxGenT1oAQZz1pDyOCPzoy3YDt3oYEjtQAuOAc9xQOT1pATgdPzpEYhjwKAHDlcZ7Uo7f40wk7RkClBOBwOmaAEbnOKGznqOtIxbbkjtSHPoOvrQOw5egO7vQoGOvahc4HT86RQdvO386BCnBGM9qVCDj5u9NIJGTjp60q5AGMdaAHcdj3pGHOQaQEnqB+JpQSRkgfnQAKMgkHtQwxjntQuQDgDoKRmfI4H50AIRgn5qVQOxpCWOc4/KlXp2oAfgZ4pSq7QQe1NVmPb8jShiQPl7etACDpnPahlHNABwSAOnrRhsngdfWgBGA3HHpQo5JpfmIyQOnekUEE9OlADyo4IHakYDHB96XDZ6dh3pAW/u9vWmBGeG5bHPrR2PP8AFQ27PTvSNk5HGM+tIB6gbgQe1KoGePSmoTxnHSlUt6dqAsOwMZJoAGDk0Zb0/WkOeenJHegaHYGetBxjA9Kbk5PT86Cx2/hQHQMcfdpCBuztpxx19TTTtHGP1oEIeMcHrQApPKn8qCADjHenLgjoetABt9AaUrn+E9KQFR2PA9aCAGwM9PWgBQoKk7T+VIyjd0NBI2n/ABoJG7oTz60AG0A4KmlCgEHFAwM49PWlG3I/Q5oAayjsvb0pQueqntQcEj6etLwf/wBdAAACMbT0prAY+6elOXbwRTWxtzg9PWgBCnBO386AP9mnYXqOKABjOD+dAxpUYHynr6UmMkfKTTiFAxk/nQAOOvX1oEIqAHG0/nTtoAHymhQBzn8c0DbgHnp60AGwEDAPTvTCoI+729Kk4GODx70mB3z9aCkMZB129/SnIuMfL0pzBQCR60mFx+XegQoUdNnSgr/s9vSkBA796XK0AAUDBx3HakKjPA70qkY7/nScEkY7+tAhNo7L2oQcEFT92nYUnp+tNG0g89vWgB2Bn7p6UmwHPynr6Upx2zR8uCeevrQA0LgH5TSHGcYpRgjJB/OkO3P4etABgYGF7elIoyD8ueKcNuB9PWkQDGP60AKyjP3TSbQMDaelO4P1+tJ8vr29aAERAc/LninEDOQvb0pF2jP+NOIXPT9aAGlR02mmgYb7vennaT0/WmYTd36+v0oACMA4U9aUAEglT0owoB470oCkgY7UAIq842np6UuBt4U9O9Iu3sT09aX5cfh60AJtHJ29x2pAuG5Hf0p+AQev500BQ5PPWgpIAoAHye/SnAEjhc0ny8fT1pQQQaAYgUE/d7UpUD+H+L0pARuA/OngjHfr60CY0AE/d/SgqB0U9KcFBHT9aQ49DwPWgQjJznafujrTNoxyvcdqkIUHj0FMYDB56GgBCvPfr6Uu3nJBpnAbqetPGOfp3NADkTB+72oKgfw9vSlG307cc0HBwB2FAAwGPu9vSk24/g/SgkdOemOtHB5xQA0rj+Ht6UuBg/L/AA+lIwXGRn86cMY5HagBFAPG2kK8fdP5U4Aenb1pGwB3/wC+qAG7emFPWmKMNyh/On7RwOevrSBRnpQAbflGAfu0uMn7p6UqhdoAHb1pcLgDH60ANKDGNv40pQZ4H6UuOv8AjStjJ/DvQVcREGB8p4pQoHGw9KAQMfT1oG3GcfrQLcQqMZ2npSAcA4NObb156+tIuCOe3vQIbtA/hpQMDG09KMKfzpSFB6Hp60ACqCD8vQU0gcDaelKBkHqMD+9SHBx9PWgBNmAflNIBkH5T0pSBzjPWgAY6Hr1oAM88r3py4wBtPT1pAF3dDSqAAM+nrQA9VG05XtQVAJOw/lQgG3P9aUYJoAYwx0X+H0pqjk5U08hSMgdvWmgAMcg/XNADjjP3T0oAB/h7UnHTnp60oxn8OeaAE8sFuUPWhk68Hr60oxu4z1owMdD1NADVQHB2ngd6EUeh6UuAT36etCAZ6H/vqgYuBj7p+7S4xkBT2owOOOw70ZGDigLCY+bIB9+KGAwfp70hbDf/AF6R2BGPf1oH0DOT1pvUjnvT8rnGOntTcgtnFBIgPzD5u9KDngsPyoBXgYoyOmKAEbJz9KcDzwaRiB/D2o3An7vagA6jqaDndj3FKpXGAoNKQu7p3FACbiOSfqcUucEYP4UEqB0FAK7gQKAFwTj+eKOT+dJwcYWlBGPu9/8ACgAz6tTWJxjNLlTg7aT5TnA96Bi5I7/pQT6H9KX5cEbewo+XsKBCMOME9/SgDJ6/xGhiMcClV1wMr3NAABk9aTHA/ClUrnhf0oyAOF7UAIM9OelL1XJPalGM/dHSg7QvCjp6UADcDrSD0z39KVmXnjv6UIwznHekMb+NIWO7r29Kf8u37vamOVBximCFVumD3o565/Smqy8cdD/SlDKR0oCwvUc+npTRzkew/nSgrjhe1KpX+72HNABnJwX/AEppPB69e1P+Xj5e35U1tvIC96BDQc5z60Alj1/SlUAfw96FwT0oAVemQe1A57/pSqy4Hy9qFII4Xt2oAM8jn9KQn/a9O1KSu7hTQMf3aAEB6nP6UpJz17+lIrAE8DpSgqTkDtQAgyR9BR/EOfSnfKei00kA/d7igBMkDO49fSgckc0Erj7vel3JxhKAETBH4UvfAPb0oUgdF7Uu5fQdOtACZPOT+lBHzHnvRuj5HHbtQCoJGP0oKTE3HjDD8aAcjqKXcnHy96AV6bfxxQIQdevpTwff+KowV9KcGXrj+KgQ/PPWkb69qVSp6D9KCQDyKAEOeSTjj0ppG5ev44pxYE9O3rSAptztFADSuWwfWl6d6Usu77veglcngUAAPIyaGJOM9qQbQQcfpR8pHA7elACknnJ6e1Gc9/xxQzLjp+lGVzwtADTnH8qASD16DvQxB7fpSgr029hzQAoPv09qG+7SqVHGBSnbwCo/KgBmO3P3qQDJxk9e9OUrgfKOtA2g/dHX0oAQAYFL2FLlQB8vagHgcDpQAw5Iznt1oY5J57insF2/dHSmtgZ+Xv6UAG4gDBpN3HUHj0oyvBx2oBXbgDtQApb1bvQp4Hzevamkg846mnIwyBjoTQAoPGd1GMgEUAqBwMUoZewoAYvfPpQc/wB7t6Uu5ccr2pMrkfLQAhHX60KMfnS5XB+WlXbjp3oARRkdTx605BwMN2pV29l70ApgEDt6UACnjj+VLng8jj2poK7TxShlyTtPWgBD0/8ArU09+fpTzj+72poA54oAB7k4xSgZ6Hp7UfLnOO1LuX0/SgBB1GT/ABUHpw3ejcoP3f4qNwI+53oGlcQDPJOeKVeuPalUr0C9qBt/u9qBB04pCTzg9qUsoxwM03euTx6dqAGk4br34prsSevWn5UtkAdfSmsFwPl6UASBSOlN2nIGO1SAMTwaTYc9aAG7Oh29ulLt74pxVuMntQVO3gn6UDI36dO1IME49h3pWDZ6nNKFcHknoO1AhAFC9P1pSfm/EUoDFfvUMGzkE9aAAgZz/WlHUcUc80KDnjHSgAI6cUhHGMdBSndkf4UDd39O4oAbgnjbR7Y7U4A9yOnpSMpAznt6UAO4IzSMAR0oG4jr25pRux+FADG6YpMjOAO9PZWxTQrcc96AHKAOcfrQcbeB2oUORSjcRnPagBePTtSOOMAdqcNwx/hTSCce9ADXBOeO9C9B/jStuPp19KVQcDn0oAOM9P1pjjJ+7UgBzjcPyppUkfh6UDIsEkf73WlXr0704K5OSeT7fSkKEZ570AAx2BoByM4PSnENnIFIobB5PSgQ7aDyVpGAOcCl+YnH9KAHOcfpQPoIq5H40u0Z4FKA2OhowffpQIaB0+nrSL0p+1iBg/pTUDAHHpQAuwE9P1oKj07UvPXPp2+lBDdPagBig+nalA54Hb1oVW5+npTgjE8ntQAgAB4FNP3sZ70/DHp39qbhs5Ld/SgBmARjB/OgZ3Dg/nSlTyc9/SgI5IIJNAAv07U4rx07UiB+wNO+bA+lADdpOcik5DYxTsHkZ/SkCndn39KADaMDjtSYz0FPKtwAaQo23g9qBkYHTA7U/aeDjv60mxt3WnhWI69D6UAAHcDFDntTlB7U0q+R9PSgQ3OeCP4Rnmmt8oOB+tPKtnPfAppDEdaAG55wR3FOHJoKuTj+lKFPr29KAFC7mHGMe9GCB07c80qbt3X9KGDcc9u4oAaV46UAZ/hNP2N29u1ARs8/yoAYy5GcUBcgnHanspIwB2FGGI5Pb0oARcY6UHjHFKqmkZWx/wDWoAQdBwetAAHIH60o34Az39PakBc//qoAQngcdqcBkYx29aT5tv4elKNwx/hQApAI4XtSOp9KUg84ofd69/SgBm3I6dvWgLjoKeFbA5/SkAPUkdPSgBhUjt0NEeSRxTnRsH/CkRWBHP6UAIowMYp2ADwKFUnp1pSrHoR+dADccHg/doKZwdn504KcHt8tIcg4B7elAEZ74FOUgjp39aUhjkZ/SgB8d6AFXA6L+tAGAMDtRhs5x+dKN20HPagBBnpjtzS4HpQgbt6elADEkZHB9KAFIB4A/GkA5xj9acQf09KQA5JzQA0qB/CelAHouKcVY4ye1BQ4oAYBk4x3FLgMD8vf1pwB3cHHNIwbkg96ADjj6etIp549KcA2Qc0AMTxQA1g3UCkAPJxmnAMRzQqkKcn06igBu07jx2odflz704KdxzSlWI60AODE4GO9ICA3Hf3pATnHHWgN8w5HvzQNrQXPA4PT1pCcjBH60Bgccj86CR6Dr3NADWHPI/Wg45G3t608kZ/h6UjBcHBHSgQ3oOAfzoON3Tv60ZO3OB19aUgFuR39aB7icAcA9PWkB+bgH35px254I7Z+akOAwww/OgAJzyF7etKOvTtikyvHTp607jGeO3egQ3nB69PWhm54B6etOXHfH50m0HGSOg70AA6dD0HelB7Y/Wlx8pPHA9aOAOo/OgBpxwAKBkYwOh9aXK45x+dAI45FAAvGQBRkYGAe3egHDEkigsNoxjtxmgA3YIwD19abnvz+dLuU88Uh27ecZx0zQArHjp39aFPTHt3oYjvjr60sZ6EY6/3vpQFhwAz0P5012xxjt605MY5x+dMYg9x+dAxFIwOD19aQ9xg/nSrjg5H5/Sj5SDyPzpqwgPXgdvWheB07DvThtHBI6etIMAckdP71IBMcg4P50uevHT1NITggDHT1pdy4PI60AKh4xtH60p65ApFK+o6UbgTkkdu9AAcYBx+tMByDkfrTg4/yaRdpUnjp60DFPXA/nSggcY7etA65OKUADuOnrQIQd+O1B+9jFAwM9OnrSlgSQSPzoAQkYGAenrTT9/oevrTgQeuB+NIMZ6jr60AJgkdCPm9aAoGDt7U5sbcgj73rRkcZI/OgBi4/u/rTuABgfrSK3Pbp3NKNpxyv50AIOhBB/OkGd3T9aXI55H3h3pRtznI6+tAB1AGP1pcDGAD09aQcgcj6flTsjb26etADQACPl/WlPTp39aMqT1H50Eg9cfeoGKD2wfzoOOML29aAcdweOxpCVGMDt3NAgIyOn8PrSYABwO9O3D2+760isdvI70ABA3fd7+tHGOnQetKSM846+tBYdsdPWgYgxnp+tGRwNvajcucDHT1pQc4Bx09aAFAwMAdhSgDHA7etBPy8Y6Cjco7j35oENbnGF7UAcdO3rQSMAEj86XcDnp0HegYi49P1obpwopcjqCPzpWx0wPzoERdAPlHWmjIPAFS9h06+tMHDk8UDQg6ABf1pw6Dj0o4wOnToTQduOMcD1oEGSB0P50rHsB39abuGDkDr604kHk460AC9sD9aACRkZ6etKpUY/wAaRSCvI/WgYED0PX1pAAMcfjmnMQcnjr60g24GMdf730oENHXofzp4OOMfrTQABnj86euDjp09aAGjp07etNYHI4P50/I5xjp/eprEFhjFACeuR+tAxtxt/WnMRzyP1pAV29vzoARiAeF7etAPA47etHXkkfnSqQVGcdO5oARcgHjt6mlPJPHelXGD06ev0peMn69M0AJnHGP4fWkBwfu9vWl75Pp60gIx1H/fVADs4x9KMgADH60mQcdPuj+KgMGHOPzoAMjd07+tITnsevrRuG7PGM9AaGPBxjg96AFVsEdenrQpB520gccAkf8AfVCsuM5H/fVAC8YHHf1pVxg/QU1W5GcD8acGXk8du9A7MAPn+739aDyB8vQUoYbuoprMOOnSgQ4p296Yw+bGOlStjHB7+tRv1yPWgLjWP+c0oPAxSMM8Y6elKBxyDQAEjPT8aCSWOPSg8HgdqDgknA/OgBD0xjvSN97pSnBXJFBwWyR3oGnYM9eO3rQvXpTgR1FC8nOO1AMafT2pwGMcdqCMnp2pyjJ6UCEA45HbvSH6U/HGdv601uvTFABkEEYpM4oyOeO3rSZ9jQA0twAKUMeAOxpGK4zQpzj5fegY7cM55prE7RxQpyen40mcgcUCFBH6GjPPFIOe3ajv0oAcx4/GlU4H0NIeeQO9C5xigY7OTwKY55H0pc5PApCCcZHagQJ0H160oI3H8aRex29xQvXpQA4nHbt60AkKcjjFBxnpSLjByOcUAIx5/Cmh8Z4709hzwp6etMwBkn19aBrYVWJ7frSk9OO1Io46GlPXpQAoIwPpSLz0FAIx+HrQmM8D64oEOzkjA9O9BPfHajPHTtSEkjp7UDBSORjtTiRuxTAepwelOOc0AG4/pSK/zfjSse2KbnDdO/WgBd+QfqaXdggAdqaTgZ9/WlyARk0CEz3xRnC/j1o+XPSgnI4H60DFUgKcj0pA2W6HrQM4OBQD82cUCHZHAFIDgY9qM5AwKQnjIWgALc9KUn5enemng4xSn7pGOhoAUEbeO1ISDjAFAJweKTK5ztoAduJ4A6LQG4+hoPJ4HakHzJnAoGKzAk8frRnOeO1BwT070nuB2oAFY7gMUu4ccdqQHB6frQc8ZHagRJnj8KN2c8Umc9vSg4HQUAIzDg5zxSgjBx6UxvYdqcp46dqAFHA6UFieMdKQH2pCQen40DHA8AYPWkxySaByBx39aFJPUfrQAbsqOKTOR0o4wMD60m4YBHtQAMRk8UrH3ppPehznJ296AuPVsgUq4x9B1pqn8eKcnPbtQIRjgEY7mkUjjg8H1pz8jp39aau3II9aAFU/MeKcW6cdqYDzwO9KSc8LQAqnggDtSE5IJ60DkEAUncDaaAAt147+tIDkdO1KcEnoPrSDp0oGKT81LkYA9qQnngdBSZyvTtQIcjDn6U7qScVGhHXHanD0NACk8/hTQR/9fFOPLY9qQEdcUANftx2oLdOO1KwJ6DtSEYHOeBQAKTu6d/WhiCDgdzQOuMUNjaTjv60ANLcj2pFY55HNOIz1HY0gXnp0oAcOFwO1Lu68d6RV4HFJ1B49KB3FL/NxTXbnI/CjqTwaR1HJA/zmgCxuQ8D+VMYLupAxB69DQG+YHNAhSEIHSjauOO3TFNLZI5p4JI4NACMo7DtSHAJ4PSnGkI5wM8D1oAjIXacUvBYHHenYIB5P50h4bv8AnQA4BehH6Uqhc9O3pTcgdT096VTk/wD16AHYTjjtSgJgHHbsKbk+hpd3v2oAPkx909KYxXAwKdnjOT+dMbHAH60DDcpH5Uihccc0c+valH3etAIML1C03CnGFNOJ96FPTFADQF7D9KCowBingA/Nk0FeB+FAhmAOinpQQvBx09qcqkgE56UjIMdP1oGgO0jG3v6UALgUrd+vWhfTn86AABPT9KDs6Be3XFKOOmfzpT94Y9KAGhVAHFG1MnC/pT17c9/WkKryf60AJtAHC9vSgAbSCvb0p3HX+tNB4JHp60CBgmRkfpSbQc4B60HPGc0uc556UD6DAoxwDSlVPOP0pQRjr+tHQ/jQAAKVH+FCgdNv6UoAwMc8dqRaAFOzOQvp2oYKRwvb0p2M+v50hHv29aBjAo5IHb0pcLuzt/SjAGfxpSMnODQITCnt+lIAA2NtPwD/APrppwGzk9exoFcb8oHTv6UoxkEfpS43Dv1oxyMfzFACDGeAfyo2ggDH6Ui9elOGCMEfrQO4ihQMAdx2pAFGfl/SngDHXuO9Jxnj1oEIQvBx+lBVcZx29KXtilwcdf1oAZtGc4/GnELjp39KXHPFBHH/AAKgBuFx0/SmnGeF7U8gEcZppByPpQAh256dh2NHBGcfpSkHPIP3R3o+6vcfjQMUkE5waPlxjHWmk8596XPcn8aAFUKDwP0o2jIwKVWBbrQQMjigQYXHSnBV7Dt6Ug9M+nFOTjv+tADWVT0XtSYTHK9venEAjPtSEAdu1ADRjsKOMDAzx1pRzSkfpQAg27Rkd+OaMpnrQCVAye9APOc0AJxgDafypOBjilPIA9qAOmBQA0hAMgUrhTnA70pFKRycetAwQJxkdqeoTqB27CheF/ClAAGeenrQGo1tvp+QpoI469aVuecn86QEZHPegAAUnAH6UEA/dB6U0EdRThjp/WgQoUYICHpQQuRle3pQmCCc9BS+nUcd6AGnaARtPXim8Y+6enWnnkHk/nUZ46fzoGBKn+H9KBtwCB/D6UmOenelUcDGenrQAqbMZx2p4Ck5Azz6Uxe+PSpBnJ6/nQAFVPJB6elNAGScGnnHr2pmRzQIUhDzRtT+EfpSFuePQdqVScfhQA3ADdD1pcLjp39PelwS349qDnBznr3oAQhARgdqaAueBTx1HPb1pvfIoAVQoxx27igBMEYHbtSA/wAqVWwDzQMQomeFpSqnp6ULyevalYEjOaBCHpkn6Ugb5s5PX1pxQnv3poQk4oGOycYzSpggcUgXgDNKBgZ9qBDsA8cUrAEke1NweCeuKXGST7UAJgbev6UkmAetLjg4pHHzYHrQA0jrzzihOTjFAU5JBpVU8GgBTgDAHajjvzgU1gRjPpTgCePYUAJ6j+lJx3HYUoTPc8HrSFSeM9hQAu0YPFKBSAEA4J6U5RntQA0jihR0yf0pSoxSCM8HJ+tAxygZ60pA2jBFNAYmlAwBz29aAHKo4prL2IpwBwGpDyfpQA11x6Ui9MH1FOkXqeetNC9PqKBBgDil4HA44pFU9M0bSMAHtQA9DwKQjJ59aFHTnpQc5I96QxW7j+lNXG05PalZSSQKQLkHr0piEYAEYowOTnv9aCpBAzQU4J9/SgYIffvSkZNIqntRggj6CgQ4ABfwpBzz1pR079KRKBikDp9KDx+VLsJ/SgqAOvagBoBORjoKGX5vxpdpwaCcE8UALgGkxk9R19KUKSaQDLDk0CEKrgk+tJgEjnH1pSPQnrQEHAzQA1QM9B09KcRwBxQq0bSAMGgBQOPxFIBk/j2FKAQM59KTHPPrQApA4x/KjgL07UmDwOeho7cHtQAYGQAKUgY6Dr1pu05z704g45PegA2jGKawGf8A69KBkfSkK8jmgAKrnGP4fWmYABGehp5U+p+7TSmR1oAaw+bg9+lOAA7DpTWXn8acBgEg9qAFTk0HAxgfpSqpz1pCOQKAHAYGcdvSnAdsd+uKYAAM89KUD0NAAfYdqXAPYdKYwP6U9Rgck8CgAUc4xQcDHFAGTxSMDjJ7GgAGCByevY0ij5jj+VJgkDLHrQBk96AFAAAx6UoXgEjsKQDCj6UoXOD9KAGlQc5FOYAelIV9Pehgc9T1oGPQ9OaAOOAPypEHTFAUkY56UAJIMk4HemqBxnBpzqT69aaAcgg0AJjHb9KU+vT6imkc/jTtozkflQIVRkHI7elB7c9qRUzkY7UhU5FADyQQcHvTACR9KXb1Oe9GCRxQA1lBP4CgD5QPanlD69qTbwMelACKOCD6U/jPAHbtTVHX6UuD6mgBWxn8PSmDOcc0pBBOD9aRV570AKw6fSnLg9u1IV6DnpRjjr2oAUYzgnoaPk2nGOvamgEnr3p23rg9+aAEON33qaOTT9nIwfXtTAp9O1AxcD+VCjOT9KAvy/lSqh7+tACjlqVlAoAwf/rUrDk4Hb0oAaSv6+lNIGeG7+lAXOeR+dAUbhz096AFBHHOKCQO460YwQN1DJnHzdaBCEj+92o3DP3v4R1pHU5+929aTZ1w/Yd6AHFhtOGH5UHG4HPekAYoTv8A1pSh3dQOaAFBUdx+dG5AeGHfvSFCD97r7Uu05xu/SgBG25FGQe9BU5BHpShOCaAAFcdfxzSORjOfSl2HjmkZOev+eKAAMMdacpGBTSuM4btTguO9ACnGOSeo70cccnr3NLs+XJNIy479/wClA0N3DP3h+dICAAd3ajadxy3f1o2EqCGHT1pajH5XGPakZh/eoCHg57c0FTtzntTFcHZTk5+lIpUnGc/hSMuM/NSqgz97vigQDHr29KU7eMEUbCBjdSMuT979aABSABml+Uk4IPPrTUBwORjOKco9T3oADjrxTRjBOe1PVcgHcKRYyynntQMbxwM0AKMknvTjEQep6d6Qx7QTmgQi7dvbrRlT/EKNpwee9CqTn5u9ABkED5qIyCMGlWIEDLH8aFjyOD2oAeAP09KCRjr2xwKNpB4NJtP97tQAmRyOhpCV3Yz+tG0knnt0oMZyTn0oAAwz1/WgMA2Ax6+tJs5+99c0Kvz/AHh19aAFQjb97v3NKACeCPzppUheD3/rT1RjigAUJ/e7dqMjoM/lQqn1/hoKjdgtQAoIwc56+lNOMn/GlK4z83amsuCSGoARiAMZFAYdj260hU5Hz4/H6UBTj7360AKGGRz36U7Ixwe9MCkHOe/rTsHHJ/i7UDsG4f3v0pDtyBu7UbDnqfxpGQ5HzCgQo254YdBxSHaVIDCjHo3b1oXOw/N+tACELuwT3FLwM80bfm+93FDRns36UAKCAeD29KCQcZJoCZYfNShDxyenFACcA5B/M0AgDg9vWlaMgE56dKTysHg/jQAhKgYNOVlA+929aYycZz2pVQ+vY96AJAR2NNbGM5/WjacYB/OhkOOTQA3IGPm70AqTjdSiP0buaRY/n+9QAAgKOeKcCuBz29aQRHjB7UpjIAx6CgBGZcHBpHYY6jrQynGc9qRkxn5qAHAqAOR09KVSvqOnpTQpHJb9Kcq/Lnd2oADjp/SkXHGMflSsnPJzSKnAwaAE4PX+VHBPBoC5HBFKiE9x19aAEUjBIb+Gj5em4UoiJGQf4aRkII+b9aAABeckflSqAV60mw85Ydu9PVCB17+tACcZGCKByBz2pAv+13oC5xz2/rQAgIAwD29KDjOS34UKnHU9v6Uoj5OSetACAr60JtzwR270FD69qEQ5xnsKAHMB2PakbHrS7AOc9qNmSfm/WgBg6/e/ipQRjBbvR5Z3Y3D73rSlDjOe/r70ALkZGP5U3C+uMD0p4j5GTTTH3B7CgYgIHGe9OVgARn9DTCp3Yz3oC4z83YdqBEgIyfmFD4wTTQMMR/ntSshK9aAE25pNp3Z/KnFu2KQk7+f50BcCvIOKQrjtS784GKXcegFAEbLk4x2pQOSPpTjx2FBPJJHagY3G1eDSkc5x3FBbI6Uu7LDjvQHQQjPejbk5oLHPTr70BsHAHagQhXGPYUvt7UE5IO2lLY5/2aAGgHpjoKNpIpcnPT9KCe2KAF2g5JzxTgOOD2o3YHPoO9BbAzj9aABiQuM00+3rSuxx92kzkg4/ioGthAuDkUH7vI7Uoc9P60hY4/CgGKOfypSnGDSbiO1LvOMkUCEZcA+9AGT170rtnqOhoDYHFAxMdcZx9aRgc8Z6dadnnIFIxPQigBig56n71OVcdAetIp6YX9acDk5x3oGOUEDmgDIP+7QpyOlKrfKQP7tACFO4/nTWHWpN/OMfrTT0JxQIjALfnTgpHf8AWkyOSR39aUPmgBVG0D6UgyfypfMyBxTVYEdO1ACn+uaMY6fWgtk5oLYOCP1oAEHXI7UrLg9e/rTQwBJx2pfM3EnHb1oATBzQv3vxo3bu1AbDfiKAsKFLLSgHIzmmhiMfX1pytyDj9aBAp/lSDNKpzzjtQTzgjpQOwgGATmmFjk4NP3ZBBHpzTe/40B1EweCCfwpQDjGe1BY4FGTjH9aAEHBGPSnAd/8AapB1+nFODHABHf1oATbxx/OmuDkEdqeDxjH60jZ44+lAaDTz0/uik2lVODxTm6/gKQHIz/WgAAOencU7aTzSDJYD3p6t2x1oEIF+bkdPWgjHbFKGG4f40biSCRQMQr1xRtpd3oKUkA4xQAwrgfhSEd/9mnk5HSkJ6/SgEIozyBQTxzRk9gPegnIz70AAzxj+960KpzkGjcSBx39aEfPGP/HqAHY+UfSnEcY9qaW+UfSnFxycDigRGyDHTtSMuART2bvimu3J470ANx79vWlQfyo3cDilU5BOKBgcnkCkVfX0pTknpSqcY44+tADcdvSlUEHgd6N2eSKVTk9KAQoGVP0pGTvTkf5SMfw0jMAen60DsMA6/hThkjr3NIWxnj070A+3egVgxg8nvSBcAcdvSlyT+FAbIA9vWgQig4xil24PTv60BuDx2pd3zHjpQAhGDx6UijBxjtS7s9B2pFOSeO1ADgD1xjgdaNueRRu9uw70oY88dKAGgYb8RS7d3OO/X8aUMQ3A7ilyQM4zzQAmDkCmjJPHpT8kkHFNBz27UDGlPf8AWkCDnj0p5PPK/rSbxyMUCEVeTzSsDtOPSkDDJ47+tK7EqeO39aAH7Vz3ppHzU8keo49qbuGc5oGIFAI60oUHB5pCR/e7+lKD2B/SgQpAwDz0prL19wO1DHB4J6dhRxk9elA0NC4XqfyowM5yevpTuMd+MdqQ4Jzz1oBsCAe54pFX5hjPSlx6Ht6UDAOc0CFAAxwelO2KeOegpmRjr2p6sCOvYdqADYvv0pGUDHXtTsgDOf0prEY4NACEA9z0oAz3IxQCAp57dCaAQO9AAQMdTSbQMdfvZpW9c9/SkDdPmPX0oGgVRknJNG3gdenpQrL1yfypeOMEnj1oELtX0P6UmBjAz0pwwMH29KQlcZzQAMo9+tCoO2aGOedx60qEYHPNBWgBAfXn2pCoyOT0pxIHc/lTJGHTPQelAtwVVwOTS7QCevWmo+MHd36YpysD+PtQAqgAYGaABg9fu+lAIbpmhSNpw2OO4oEKRznJ/KkOMHnmlZgMZbt6U3OAeT19KB3DauCOevtTQBnOT1pVORk569xRyTxn8qAAIABjJ4oVVx36etLwQBnt0xQuMEZoANg7k/lSFQSfp6UvHB3dR/hSEgd+3pQAwAZPPb0oxhsc8+1AwfXp1obBbg0AKoA556c0BRvzk9R2oGDx6UoADZDd/SgAwMdT+VKvbr09KbkY6k8+lKMEggnp6UAKBg9+npS7QecnP0pFx2/lS8dPT2oAFRfX07UmwbvxpQRz83pQCAetAhpjHHWl2Ag9adxgcik7cH9KBjQgB5J96ULxgZ680cZBJ/SlyCv3uh9KAuARSM80FFHJz0pRgZ5prEAgg9uwoEIwAPGfuimhQowc4zSlhn7/AGFG4bc5PX0oGLhQ2RnrmlwBjGaTIJxk/lS5Azz29KAAKCRnPSjA4GD0pARnOccelIe2OaAHhRj8OacFyc8+9N34OM9O9KrA4x+PFAwKj36UpQEdD0pCy9jQHXB+btxQAmxc96Nqgd+1KDnvSMBgEE0CG4UYwe/NIqjOcmnBhgfP36UgI9f0oEGPlAyenpR949T+VHGByTSZAAPr7UDQvGMc0Mo5+vYUhPGR/KlZgQee9AAFHHWljQYwc0gIwOf0pyFRjmgBSgIyc0gVQBilJAHWhTjGW/SgBu1c45o2gdzQcZ/+tSHkd/yoAFGAeeg9KUgEg5P5U3dgHL9qN/vn8KBilUOSSe3agAEdxSBsk/MetKuMZyfxoELtUetG0ADk9KAc9KOMDntQIQKCO/Tv+FKEBYk8c+lIrccGlyDnnv6UAIFA9elCjk9enpSnnHJ6elIMAkg9KAHbRwMHoO1AUY7n8KCVzw3alLZ7/pQAgA3Z5pdo7E/lTASG+93pwYf3j19KAHBRgHnpTAAOBn8qcCOME8j0puA3r+VAxNoPRu3pRtHJ5pcDGM9BQpAzz6UCE2gHnPWhl4JGenGaXHPH8qGGT0/SgYrHt700/ezmnscgc96b34PfigQnfI9ewoz2yelKSMYzS8YHNADGPGM/jT/X6UnTmncAnB7etADdo2n/AApDgtx6ilJBXrmg7d3XvQAAZPWgZJ+9+PNLwCTuHT1oVhkD2oAa2O2Pyoz1+btSkA4wO1Lgen8NADQfQmkyT09PWnEA8j0ppC4GfTvQAoxjg9qMgdCfpSjGMZ7CgDuf50AIemM/mKTnjn8vrTscUqqrbRQAwYyckn8ad0XqPzpVAyeRxRwABTAM8dcfWkYnGKcMcc0wt3z+tIaFJ6jPfpinA8fe7U0nJOD3pykFcZ7UAwbOcAnmmP65qU43cHv61G+DjB7UCGL7E8GlU4HelUjPXvQAM596AHLkcCgNxj29aXGeh+lJxgjI6UABbp8360jfdJz+dBZRjDCkLAZwRyfSgBF6Yp2STwD1pFweaXAznNAxQM4/wpFPHXp7U7I2jDDp600EFelAhSc9+3pTScY57dqdnPB9qRscc9vWgBg9MdutB5PUUEKB17UpIJ659OaAFXnkenpQMZ696Bg9+gpVwDk460AJjjk9+uKMcg+3pSnAXggc0fLuBz2oAReDx6Up6ckfnRgZ4I6UbhtA3dvWgA47H0oU5bI/lQGH6igH5uDQAvO3736UZ4OWPWlIBwBSHGOOtAxvGfw9KUk/kaTgmlYjkZ70CE3Enr+hpHJ7HtThjoT2pvQ9e1ADScHGeMDjH0pFPGM9/SnFFJJHp6UKAQCSDyKADGT+NKcnp/KlIHbHWjIAOT29aAAfe69qQ9vpShgDRx1FAA2euaA3+1Tjgj8qCAv60FDC3bNAb19KRsHvSgdfpQGgqsSMZ/SlY5WhSAaDjHWgQmTgDNAJBPIP0pRsIBHYjmkUqc80CEPQYpMGnhRgYNJjA/CgBpAxjNKcDJz3oJBJ4pzYNAxF5Ax/KnKMjj+VICOOe/rTkKkdegoARupwfwoQ8Dn8aVj1578Ui4GMetAhvA7fkKPw7UvB6dAaU4bmgCMdSM9u1GcEfMf84p6qCD06U1wAw5xx60AJkc80q/dwDTcgE/MOtOTG3OaAF7/j70D7ox2FKcZyDn0peqjBz8tADBjkUE4JAPelGOfpSkAk47GgBmccZ7elAPU04qD37U3AHegB24g9/wAqUH+XpSHk4HpSjGMA9qADHPPrQenUdfSlGM8nvQSAPvD73rQAgOcZz07CjvjBp2FODu7HvSYGcj0oATHtRxnOcdKU9PyoGOeaAAYyfenMowB15pARnqKczDnBoAbjA6fpSdHzgce1IQM9859aaQC3WgBxAGBgf5xSgA8gD8qaAOMHinADigBxz6Dp2FGOuAOnpSHB79B60jAEnn+H1oATOASMflSEtv6DqO1BVQpwf1FIQobGe/tQA8lieQOPahQCR06elJs3Z47etKEBbp2oACQD90cDpilIyM47Y6UhUdgenrRgH9O9Aw2+ij8qaR3AHTjil2gjOD1PekIHHB6UCHAccgdu1KF9APbim49B6d6cFz0BoACDjA/lSAkEDA6+lKygr0P503aARk/xUAKOOePypcnAwB09KaqgHOf0owABz2oAXngkDp6Uxh/sjP0p4UHGQenemMq7cAdqBoUnAOAOvcU5COAAPyqNlXk8dfWnoBxz3oAcc8YA/Kmt9B+VKFH+QKQrk854FAhEAGB8p5pcYzgDr6UKo4B9fWgAY4HegB43DsPypuCwPA+76UEDp/WkAGDyPu+tAIHXB+6PyppJ56dfSlZAfy9aaVXn5h1oGKhwOg6+lKDg9B09KRVGDz3o2jPy+1AD1J2jAHT0pF6Hp+VNAXA+nrQuMfhQA9tue3X0o47AHj0pBzz83X1oIU9M9KBAoPbH5Ubefy7Ug2jPPanADOMfrQAKMHIA/KgZzwB970oAUe3FIMZHzAc+tA0BGc8A8+lKBggYHT0pNoPqfm9aAq4HPb1oEwUnPQcD0oI6cfpSDGeo4FLtX24FAAo5JwDyO1HGT0+mKUKpB9jSYGT/AI0DHMMkcD8qTGM8DpRtBx9aNoPAoATHOBilPrgZz6U3bg8UuMdv4vWgQ4Dv/SkYHpgdPSgKCelIwXsR0oACvJ+UHgUgUqMgD8qVwM43jp60gVQvX+VACkkMOB19KGHUlR+VIQN3XvQVBycdvWgBBywGB1oOeMY6elKFBPQ8D1ppQccdqAHjp0/Slx2AH5UmwEdPTFO8sdAP1oAjcDrgflS44OQOB6UhQHBI7c804KOSR2oGCjsAPypTwAMDj2pAoPQGkYDAwfzNAhegBwOvpSKMHjH5U0Y4+YdfWhQueW6e9AD8kqMAcD0oIwv3R0Ham4GP+A+tLgEDGTx60AKQBkY/Sh8DPQfhTcA0pABOD39aAFGRjgdfShTxjA/KjGQOKaEGOh6UDJC3PQflSKM4wB19Kay98d6EHTjpQA7AzwB19KXoeMflTABjpSgAZAP60CFGQpwo+76U11y2do6elKApGeOnqKayrnOR0oAQqeen5UqZx+PpS7VOeaAox+PrQA5eDjaPyoAwBwOnpRsUdB+tAUYBwelADQOD06elOAG4/KOvpTQox07UqquTx1NADzwBgDp6UwDknA/KlK8Y54WkVQSTigBSBnGB0HalAA6AdPSkwAeB2oAXse3pQAck9O/pQQT2/iPakAG7qOvrSsFweR19aAHAsCAAKbnHAUfd9KMDI+YU1R6DtQApAHOB19KAevA6dcUhVc9PTvRgAHIHagADncRx+VKzHHAHJ9KYV+bpxSYBAwO1AD8c/wD1qQDkeuafsyePxoCnI4oAaA2QacF5oK9PalAyBkUAB7jHakJJJB9KU55I9KCBknH6UAM5AwDSj7wH0pWAA4oI54B7dqAFKg9h+VBHzA4pwHrnpSAfNkCgBpOSDjtQQehx27UMPQdqX6Z6UAM/AfXFNbPTA6U8gY6H86awJ7UAOAz79KcBxxTUB5p68DjNACMPlpMnI5PWnEHHSjbnGAaAEAwcj1pOSACe1OxSbcgYBoATGQOKRkAXp2609R7HpSMAcmgaGlcA+uaEBHT+VK456d6FB447igYY9h+VJtGecU4DnIB96QjPQHpQSCr07cijbk9BSp1BHt2oIGSQO9Juw0hvIxSKSBgelPK8cD9KaFyCT6UxrQCM4wO3pRzz1pwHGOaNo5GDRcLaDVzjGO/ejaMduvYULmhutAhoGMdOKQDjkdqeRwKRFB/KgAI54H5mgZycY6U7ZnjB7UpX2PSgBijqPanYpApOTinEDOcUAN7nH6UmSG4H8VPIz0pFUlssO9AhBkjOO+acW6cUhUDoPWjHPfFAxoOfypcEtn+lJ+H6U5QOMUAKucHkdqCDnIoHT8qASSePxoARhkjjmlycbcUh6AEGlIx90UAMx7dvSnbR1wPvUhBzgDtTsHGSD19KAEx3xyPemNn9KkKnbTHU5+7QAbiTye1CnC4B/SjaM8DtRjjjigA5JzjuO1L1Ppz2oxzwO/YUvrx9DigQm3nkdutBByM0oBBoxnB9qAEA46enanYxxx+VBGP0peMc80ANKjAwO1G3Oc+lDg4/+tSgE5B7CgY3GBnHI96RicfSngYpjA+n6UCEBOcZ70i56g07GQDjvzRtA5xQA3+HB9KceaTHHAPIpQM0AIV45XtTjnnOfzpdnXmlcE5xQNjQMDAJ6djSbR6U4AkdO1AHGB0xQA1hk9B+FA7f1oOccUL0HHFACgccAUZORQBzwDQRzwDQA0Ej06elKAcg4FKBlTkHOKTGOOnFAhQeOacBkdB19KaueRjrT1U44BoATbz2/Kk24A4HA7U8g9gelBT5R9KAIwD6c+9IBycAHn0p+PakCgk8d6ADbz2/KgDqDzxTipPQdqbjOePpQApHPUdPWkzg5pWHIOO1IRkcDNADc89B27UpYkfj1x70AZODQwx+ZoAQsQQB296avPOO1L/EOKVR0wP0oAAD0HpRjkmnBQR0P5UoXg9aAGYAbpzSMp9zTwmTj+lEicc0APCrxx+lJtXOOaAxxgkcUmSWAoAGCDAzRwe3SjLHA+vajnpQAhxn8PWggc/4U1856dqCT79B3oAABjIOPwpwADYH8qYpYr/UU4A5696BsfgAfh6Uh2k4HYUuSO/XtTdxz+B70CEbbgewo+Xv6CkIORwelGCOnpQADZ0FIcHBz2pV3e/TvSENigBflA69h3py4x+FN5A/ClUtxj8aAH8cEE9qQLnAJ7+lBB2/lQAf1NAAAB0B/KggAAUAPyNp+pprZAHFADhtxye2aaxGDjPSkywxjNN3HbmgaHNt9e9Km319KYxbGfenLkc89KAHLt9D+VIxUEY9O1A3ehpGzkdenpQIVNoxz3pQRzg01SRjnv60BzjPvQNEgC9BTVHB/wB3uKAWYdO3WkGe3pUjFIUdP5U1scgHofSlJbsDTW3YJyaBq4bgR96jIyf8KaMgZ7ZoyapEskTbgfSkUjH4dqap47dPWlUnGPagCQbd3Q/lS4QflTAzZ4B60u8ng56UAKoX3GBQQPemhiCeucUjM244BoAeACePSkCjd/8AWpAW9OlCklsA0ADAAfj6UMF4we3ekYlRketNJJA69KBADg8HHNAYDBz29KaMk9+meaXnPB/M0AKpHOCRyO1KGAPWkUMc8d6AG3HjtQMdlBj/AA+lL8vIA7+lNG7IAzSjOM0AAwDxmnfKOMn71NCk/l3pTnGB/eoAcAvbNIyrkcfpSAtikbd29KBAUA/LpigAbc5I/CkyQeM9PWhWYjr+VAxxAzgevpRgZIx2ppzu70pyOmenrQIXA3d/agkYBBPSmgnI7fjQd3H0oGKSvPPp2pflz1PT3ppZu2e1A3dMn60AOJHYHp0zS/KAQc5xTDuIA2npTgTgkDtQAoxjgGkIGB1o56enrSEEgcmgQoC4AJ/MUgCk8GgFiBz3PehSxY4NA7CBQANpzx0FOVRxkdqTDYFKOAMD0oAdgYPB4ocLg4NN3HFDFiD1oHYX5cf/AFqT5cZ9vSglh0pFLYyPSgVgYKDgUiqBjmlOT+dAyMYP4ZoBAAMHr+VKfqaQFvTvQAT2oABjac+lBC8fT0pAG7f3aXLAgc0CFAAyT6+lOQjb3/KmckE5NGSB34oAkyvYdh2oBXaCD2pmW/ChXyB9KAFULj8qUBcke9MDNjqaXJ3Hj9aAHZXqB2poI5OT9KTLe/Sk+bJ47UAOfaTxnoKOM4AP5UhJz07etGST0oHsAHzf1xQVAGB6ntTRwwAHejcyrnnrQIcqrkEelCYB4zSckgnjihc/+O0DsSZUAfSgbME59O1MDHA/rSgtz1oHZWHDAbGaHAIwvNNDNuIyc0MWK8E9KCRd4/OgNlu34CmLg8ZPX0p643cseM0AOBBwOPzoJHQ/zoXaMHJ/OgYz1NADGIyP/iqAefwHenEDOBnr60FQM9eg70ANBAU5/wDQqcSu4dOvc0KBtzk0EYcDcevrQAbh14/KkBGeo6fSlI9SfzpABkDce/egA444/wDHqX8R+dIQOOT09adjjqfzoAb+AoYADn+dOwPfpSEAjOTQAnGOg6etOA+n5Cggep6CnDHqenrQAHae4/KgY46fepTwuSTQmOOT19aADgA8Dv60xypAGB/31T1wc8n86a5XAx/OgCNhgcY/OmEcdvyqYheCSe3emFVxgk/nQNCYGSMd6cpAAHt60u1cH/GlVV9e3tTC+ogPr/P2pGwDxjpjqaeB2z6UhGT1P1pARggEe59KFPHbr6CnBRx16jvQABnk9fWgQoI7evvSBhjHsO9OUE8ZNCAbTk9qB3EPvj8TSHaQd2KccbhzjNJgEHk9aQ9ho24OAOtJgdvXsaeFBByT+dLtXPJP0zTE9SIYAH0/vUq47enrTwilcZ7etIqgA89vQUCDjOMjrnrQQM9unrS4G7OT+lKBg/ePSgBi9fXj1oIG7Htz81PUDBye1KVHJyaAGcDn+tAYbu3X1p+0ep6UhAL9e/rQNDM8Z460m4Dt2PenMBt6nr60ALxknpQFhi47enrTgBgD39aFUE5zzj0FPQAY5/WgQigbTx37mgqM9B+dOUAA4J6jvQcAkZPSgBhXp069zSgcEHH50pA45PX1oUcdTQAgC54xQeB2607AzgntQwGOp+9QMYT24pDtyDx+dKFyev8AKgqMjBP6UCEyM8Y+6KAQF5/WnMoA4J+760i/dOWP50AIduecdR3pCeT06etPwN3U9fWkYKOh7etADQcsDig44yP1pdoJ6/rRjpk0DDAPpSgdsA8Uu0Y5J7d6dgdv50CGEAY/xoyOfp60rYGBk9utIFGOcngUAAPHagkD07d6VVHqaUgY5J/OgYikFRj1+tKDSqAAOT19aRR8x+Y9PWgBDtIHT9aM8AjHbtS7cAYJ6+tLtHH0oEMOCO35UrY5Gf1pSvynk9KHAA6n86AG5Axg/wDj1CnA5x+VAHv+opUAIPJ/OgYh2+g/OhWAxn09acVBB5P50m0AAZP50AICDn60vH+WpAODk96ev17e1AhowBgH+H+9QevXt/epwGe56UMBkdenrQAwlTnOPxoBGPx9TTig5Iz270KoI+agY3gHgjr60mQADjtz81SbVzjP60wqNo5PT1oENDDBHHT1pQeTjH8qRRnPPanooyc560AJ9MdPWk+U5H9afgYxk9KQAc5PrQA3jOPb1pQQO/60pUep6e1AUep6e1ADflzkkdaXgjr/ABf5707aN+N3FI4GDg9+1AxO4GfyakHHr09akUDjntTQB+ntQA3p/wDrpFI5zjtTjg4+tIAOfmPQd6B9AyC3B/WnMRjIA6etNXbk8mlcgAc0EkYBJ60o3A8UBT2NAByMHp70AKGJwT2p25qTDYA5/OnfPjqc0AId2Rz3o5yeT0pTuBGDRkg9T0GeaAEOQuVJoOdwzntRlguQTRlgfvd6Bijd1NJhieKcAfXtSAEkfSgQ054+lKCcfh3oIYY5PSg7sZ9qBiAnp7UEk0AH1Pvmj5uM54oAcCxBz2FKNw6HtSDdj73QUoJx940AIaTJGOe/rSktgDNJluPmPX1oELkjPNNZjxQN+Op/GgM2AQf1oAATx9KQ7sd/rTgCMcnpSN1GTQMG3DINOTPBprBsHB705ARjn60AOAYjn0pr7s//AF6d82eM5+lId2cH86AGrk4FJg5pyg8HPakOc5z39KBBkjgCgbsHHpRuI53dqFL4J54FAC8n8qTnnJ70uDu6npSfMM89D60DBTx170p3H0oBYAc96Du7E0AAzjj0pAGII9qMNgfT1oTd15/GgQEtnkY/Gjn9PWghuDz0oww7mgAG4gkenc0pyWNINwyP5ilYtnOaADJpoJ3Yz0NL8w6E++aQbt3U9fWgBAWIowSeBSlCR360BWBzk0AAB6j0pwzj6Uihs9frS88fT0oAFJwST0xSEsTwaUbsZ3elIu7JOaAEOcjilXOMHNBHuaFyMHmgABJxilIbGM96QbtwGaXJ25yetAw+bGKQhsjApy5A6mkwc0CGlWyOD0pCGAyAetSFW9f4RTCpC9fSgBCxLf8A16MnP4UfOD1py565PSgBqrk0o7U7kHrTTkEc0AOB4Gfaj5ux7Ug3Dqe9OGe56UAMYH07UKrY/D1pzA460DPXPagBF6cZ6UjE+9OGT0J6UjZ/vUAAZuOe9IC2fwpPnGOT1HelXd6n86AFGdoHPSnAHAFNBYAHPbvTgW4z6dxQAHOOD2ofIBz6+tKd3r0pJCcZB70AMwcA0qg9aQE9z2pyhuvPFACnOcH1pBnjBpcNjqaFLcHP1oAaN1Kfb065oJOevegMw5z0oAVCcEj070cnikDNg4PajLccmgBSeDnsfWkGcHFA3889DSqWA/H0oGIcnp6007iB9KcSScEmg528ntQIYATke3rSjIY5Hf1oAJHX86UZznNABzjp2pBnnHHFOwc8GkUNyM0ABLZ/CjLDjHahs5HJ/Ok5HIJ6d6AFGSfxFHzMOfWgF9/BPWhQxHU9aBhubI60mTnr2pwDZBB/Wm4brk/nQIQ5xmmjdz9KcQ3HXn1pAGwef0oGgGdx5pHLdu/anfNuIJNBUsM5/CgYAY5zSjrndSAsRkE03c245J60C0JdoyMmjHqaYHbgFj+dO38/eP0oEDAg9e1Bzk/MegoZ26hjTTI2CM9utAC5O3hj1pSCW+8evpTFcgH5j+dLvYkcnt3oGPH1PT0pQPmHzUwE+tKrMCCSenrQAMvIyaUDPO49KTc397FLuYdWPegBAD3b60hUAdaduYdGPSmFn7EnpQFh4H+12oH+9TC7YwT6UKxAGD+tAWHHJAO40YJx8x6+lALkfeP50BmwBn9aAE2nPWjJwBuNALBuSevrTCz9Mnj3oESLyB8x6UFff9KaGfAOT09aUFgudx/OgaFYerd6Vc8YY00scZ3HqKFZs/ePHvQCsPxzkk/lSEAcbu1AJ/vHp601nbJG4/nQA4DphjSHr949aarMBwf1oLHdnJoAU57E0qAkdT2poZiDlj+dKrMBjP60CHEHHWjaRn5j1oJPGGx+NBdhnmgY3JHO/vRgn+I/lSFmAxuP50gY/wB4/nQFiQDAHJ6UKAB1podgME9B1zSqxAPzH86BDtoB4J/KkIGME0ZOfvHgetGW7N+tACBQM80u3k5J60nzDuaCSWJLH25oAUg5yCaRV+fgnrSgn+9096Od2dx/OgACfXr3pdvuelHzYGCfzoJYAHJ6UAIBzyT70Y4xmk3NuPzHv3oDnuT+dAC44OHPNJj5jz37UA4B+amkkMTk9fWgY7pzntzRxjIJphc8HcfzpwZsZLH86AAjplj0zSjgfePWm5PqenrQSQcZPXsaA0Hp6g/pS4JIwT044piswH3j19aUOxPLGgQ7B9T0FNIIGcntS5Pr29aau4ZAJ/OgYEfNndThgA4NM3NkHcfzpd5yfmPQ0DskOxhshv0prA5GWoDkEfMelIzN13enU0CHY/2j+dKucdT0phc4+8e3WhXI4z0HrQIc/YbqQE4PzdqQs3dv1pC7DuaHoNWHrnk7qCCf4jTFdgfvGnFiRgsaSbG0hcEAcn86aFIPBP1pdzYAJpuSHPPr3pkjhnA+Y04ADHzUzc2AMn86UswA+Y9PWgBzHjG88UjknJ396Z5jLn5j19aXexHLGgbsKoI53dqcijGcmmqzAA7j09aVWYdz+dADiM5IakAPBDUhd/U9u9JucDOf1oAD1OWoyefm7elNLkNyTSbycgsfzoEOBbBIJ6UuDnqaYrsM/MenrRkkj5j09aAJBuGfmPaheRnPek3MMjcfzo3OF+8evrQA78TSYO0Dd2pu9gc7jx70gkYjG49KAFXj+KlAOclu/amB2HBalV2BJ3H60wJO/U0gAycNnihSSCSx/OkDHOdx/Gn0AVgc53dqQjnOTQWJ79qMsD97pSe4AB83U9acASOWPX1pFY7slj+dBZyASx496QDwDxyenrTdvzdfxpA7YGD2oVmyTuNACY/2qFTP8VBLDHzHj3oVmGefegAEfPJ+uKHXK4B+tKCx79qHZucn/OaBn//Z" '
        f'width="{width}" style="display:block;object-fit:contain;" alt="PlanTrace">'
    )


# -----------------------------------------------------------------------------
# HELPER: KPI card
# -----------------------------------------------------------------------------

def kpi(label: str, value, sub: str = "",
        colour: str = PT_NAVY, accent: str = PT_AMBER) -> str:
    """Premium KPI card HTML."""
    sub_html = (
        f'<div class="kpi-sub">{sub}</div>' if sub else ""
    )
    return (
        f'<div class="kpi-card" style="border-top:3px solid {accent};">'
        f'<div class="kpi-label">{label}</div>'
        f'<div class="kpi-value" style="color:{colour};">{value}</div>'
        f'{sub_html}</div>'
    )


def kpi_row(cards: list):
    """Render a list of (label, value, sub, colour, accent) tuples as equal columns."""
    cols = st.columns(len(cards))
    for col, card in zip(cols, cards):
        args = card if isinstance(card, tuple) else (card,)
        col.markdown(kpi(*args), unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# HELPER: Status chips
# -----------------------------------------------------------------------------

def chip(label: str, style: str = "grey") -> str:
    """Inline HTML status chip. style: red|amber|green|blue|grey|teal"""
    icons = {
        "red": "●", "amber": "●", "green": "●",
        "blue": "●", "grey": "○", "teal": "●",
    }
    icon = icons.get(style, "●")
    return f'<span class="chip chip-{style}">{icon} {label}</span>'


def float_chip(tf) -> str:
    """Colour-coded float value chip."""
    if tf is None:
        return chip("-", "grey")
    try:
        f = float(tf)
    except Exception:
        return chip(str(tf), "grey")
    if f < 0:
        return chip(f"{f}d", "red")
    if f == 0:
        return chip("Critical", "red")
    if f <= 10:
        return chip(f"{f}d", "amber")
    return chip(f"{f}d", "green")


def status_chip(status_str: str) -> str:
    """Status badge for activity status."""
    s = str(status_str).strip()
    mapping = {
        "TK_NotStart": ("Not Started", "grey"),
        "Not Started": ("Not Started", "grey"),
        "TK_Active":   ("In Progress", "blue"),
        "In Progress": ("In Progress", "blue"),
        "TK_Complete": ("Complete",    "green"),
        "Complete":    ("Complete",    "green"),
    }
    label, style = mapping.get(s, (s or "-", "grey"))
    return chip(label, style)


# -----------------------------------------------------------------------------
# HELPER: Page header
# -----------------------------------------------------------------------------

def page_header(title: str, description: str = "", icon: str = ""):
    """Render the dark-navy per-page header with programme status strip."""
    prog_loaded = "programme" in st.session_state

    # Programme status strip
    if prog_loaded:
        prog   = st.session_state["programme"]
        pname  = prog.get("project_info", {}).get("name", "")
        ddate  = prog.get("project_info", {}).get("data_date")
        ntasks = len(prog.get("tasks_df", []))
        nrels  = len(prog.get("relationships_df", []))
        has_res = not prog.get("task_resources_df", pd.DataFrame()).empty
        has_notes = bool(st.session_state.get("_notes_text", ""))
        has_comp  = "_mi_prev" in st.session_state
        dd_str = format_date(ddate) if ddate else "N/A"

        chips_html = (
            chip("Loaded", "green") + "&nbsp;" +
            chip(f"{ntasks:,} activities", "grey") + "&nbsp;" +
            chip(f"{nrels:,} relationships", "grey") + "&nbsp;" +
            (chip("Resources", "green") if has_res else chip("No Resources", "grey")) + "&nbsp;" +
            (chip("Notes", "green") if has_notes else chip("No Notes", "grey")) + "&nbsp;" +
            (chip("Comparison", "amber") if has_comp else chip("No Comparison", "grey"))
        )
        prog_strip = (
            f'<div style="margin-top:10px;display:flex;align-items:center;'
            f'gap:8px;flex-wrap:wrap;">'
            f'<span style="font-size:11px;color:#8FA3B1;font-weight:700;">'
            f'{pname or "Programme"}</span>'
            f'<span style="color:#1E3A4A;">|</span>'
            f'<span style="font-size:11px;color:#4B6275;">Data date: '
            f'<strong style="color:#8FA3B1;">{dd_str}</strong></span>'
            f'<span style="color:#1E3A4A;">|</span>'
            f'{chips_html}'
            f'</div>'
        )
    else:
        prog_strip = (
            f'<div style="margin-top:8px;">'
            f'{chip("No programme loaded", "grey")}'
            f'</div>'
        )

    icon_html = ""
    desc_html = (
        f'<div style="font-size:11px;color:#304a5e;margin-top:3px;line-height:1.5;">'
        f'{description}</div>'
    ) if description else ""

    st.markdown(
        f"""
        <div class="pt-page-header">
            <div style="display:flex;justify-content:space-between;
                        align-items:flex-start;flex-wrap:wrap;gap:8px;">
                <div style="flex:1;min-width:0;">
                    <div style="font-size:9px;font-weight:700;color:#1e3040;
                                text-transform:uppercase;letter-spacing:2px;margin-bottom:4px;">
                        PlanTrace
                    </div>
                    <div style="font-size:20px;font-weight:800;color:#FFFFFF;
                                letter-spacing:-0.3px;line-height:1.1;">
                        {title}
                    </div>
                    {desc_html}
                    {prog_strip}
                </div>
            </div>
        </div>
        <div style="padding:20px 24px 0 24px;">
        """,
        unsafe_allow_html=True,
    )
    st.markdown("</div>", unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# HELPER: Data quality card
# -----------------------------------------------------------------------------

def data_quality_card(data: dict):
    """Compact data quality summary card."""
    tasks  = data.get("tasks_df", pd.DataFrame())
    rels   = data.get("relationships_df", pd.DataFrame())
    res    = data.get("task_resources_df", pd.DataFrame())
    cals   = data.get("calendars_df", pd.DataFrame())
    notes  = bool(st.session_state.get("_notes_text", ""))
    comp   = "_mi_prev" in st.session_state

    items = [
        ("Activities",    len(tasks),         bool(len(tasks))),
        ("Relationships", len(rels),           bool(len(rels))),
        ("Resources",     len(res) if not res.empty else 0, not res.empty),
        ("Calendars",     len(cals) if not cals.empty else 0, not cals.empty),
        ("Notes",         "Yes" if notes else "No", notes),
        ("Comparison",    "Yes" if comp else "No",  comp),
    ]

    good_count = sum(1 for _, _, ok in items if ok)
    if good_count >= 5:
        overall_label, overall_style = "Good", "green"
    elif good_count >= 3:
        overall_label, overall_style = "Partial", "amber"
    else:
        overall_label, overall_style = "Limited", "grey"

    rows_html = ""
    for label, val, ok in items:
        dot_col = PT_GREEN if ok else "#9CA3AF"
        rows_html += (
            f'<div class="dq-row">'
            f'<span style="font-size:12px;color:#374151;">{label}</span>'
            f'<span style="display:flex;align-items:center;gap:5px;">'
            f'<span style="font-size:12px;font-weight:600;color:#0A1628;">{val}</span>'
            f'<span style="color:{dot_col};font-size:10px;">{"●" if ok else "○"}</span>'
            f'</span></div>'
        )

    st.markdown(
        f'<div style="background:#FFFFFF;border-radius:10px;padding:16px 18px;'
        f'border:1px solid #E5E7EB;box-shadow:0 1px 4px rgba(10,22,40,0.07);">'
        f'<div style="display:flex;justify-content:space-between;align-items:center;'
        f'margin-bottom:12px;">'
        f'<span style="font-size:11px;font-weight:700;color:#9CA3AF;text-transform:uppercase;'
        f'letter-spacing:1px;">Data Quality</span>'
        f'{chip(overall_label, overall_style)}'
        f'</div>'
        f'{rows_html}'
        f'</div>',
        unsafe_allow_html=True,
    )


# -----------------------------------------------------------------------------
# HELPER: PM Attention panel (auto-generates top issues)
# -----------------------------------------------------------------------------

def pm_attention_panel(data: dict, near_crit_days: float):
    """Auto-generate top 5 PM attention items from programme data."""
    tasks = data.get("tasks_df", pd.DataFrame())
    rels  = data.get("relationships_df", pd.DataFrame())
    notes = st.session_state.get("_notes_text", "")

    if tasks.empty:
        return

    tasks = get_critical_threshold(tasks, near_crit_days)
    items = []

    # 1. Negative float
    if "total_float_days" in tasks.columns:
        neg = tasks[tasks["total_float_days"].apply(lambda f: safe_float(f, 0) < 0)]
        if not neg.empty:
            items.append(("red", "Negative Float",
                f"{len(neg)} {'activity' if len(neg)==1 else 'activities'} cannot meet target dates. "
                "Immediate recovery action required."))

    # 2. Critical not started
    if "is_critical" in tasks.columns and "status" in tasks.columns:
        cns = tasks[tasks["is_critical"] & tasks["status"].apply(
            lambda s: str(s) in ("TK_NotStart","Not Started"))]
        if not cns.empty:
            items.append(("red", "Critical Activities Not Started",
                f"{len(cns)} critical {'activity' if len(cns)==1 else 'activities'} not yet started. "
                "Any delay will push out the project finish date."))

    # 3. Open-ended activities
    if not rels.empty and "task_id" in tasks.columns:
        with_pred = set(rels["succ_task_id"].dropna()) if "succ_task_id" in rels.columns else set()
        with_succ = set(rels["pred_task_id"].dropna()) if "pred_task_id" in rels.columns else set()
        open_starts = tasks[~tasks["task_id"].isin(with_pred)]
        open_ends   = tasks[~tasks["task_id"].isin(with_succ)]
        n_open = len(open_starts) + len(open_ends)
        if n_open > 0:
            items.append(("amber", "Open Logic",
                f"{n_open} activities have missing predecessor or successor logic. "
                "Float calculations may be unreliable."))

    # 4. Near-critical due soon
    from datetime import timedelta
    now = datetime.now()
    four_wks = now + timedelta(weeks=4)
    if "is_near_critical" in tasks.columns and "eff_finish" in tasks.columns:
        nc_soon = tasks[
            tasks["is_near_critical"] &
            tasks["eff_finish"].apply(
                lambda d: d is not None and hasattr(d,"date") and d <= four_wks)
        ]
        if not nc_soon.empty:
            items.append(("amber", "Near-Critical Due in 4 Weeks",
                f"{len(nc_soon)} near-critical {'activity is' if len(nc_soon)==1 else 'activities are'} "
                "finishing within 4 weeks with limited float buffer."))

    # 5. Planning notes risk words
    if notes and "task_code" in tasks.columns:
        import re as _re
        risk_acts = []
        for _, t in tasks.head(200).iterrows():
            code = str(t.get("task_code",""))
            if not code or code not in notes:
                continue
            idx = notes.find(code)
            snippet = notes[max(0,idx-200):idx+200]
            for word in _RISK_WORDS:
                if _re.search(r'\b'+_re.escape(word)+r'\b', snippet, _re.IGNORECASE):
                    risk_acts.append(code)
                    break
        if risk_acts:
            items.append(("red", "Risk Keywords in Planning Notes",
                f"{len(risk_acts)} {'activity' if len(risk_acts)==1 else 'activities'} "
                f"mentioned alongside risk keywords (delay, blocked, CE, EWN etc.) in planning notes."))

    # 6. Comparison: major slips
    if "_mi_prev" in st.session_state and "_mi_curr" in st.session_state:
        try:
            prev_t = st.session_state["_mi_prev"]["tasks_df"]
            curr_t = st.session_state["_mi_curr"]["tasks_df"]
            if not prev_t.empty and not curr_t.empty:
                merged = prev_t[["task_code","eff_finish"]].merge(
                    curr_t[["task_code","eff_finish"]], on="task_code",
                    suffixes=("_p","_c"), how="inner")
                slips = 0
                for _, r in merged.iterrows():
                    try:
                        d = int((pd.Timestamp(r["eff_finish_c"]) - pd.Timestamp(r["eff_finish_p"])).days)
                        if d > 14:
                            slips += 1
                    except Exception:
                        pass
                if slips:
                    items.append(("amber", "Major Date Slippage Detected",
                        f"{slips} activities have slipped more than 14 days compared to the previous revision."))
        except Exception:
            pass

    if not items:
        return

    # Show top 5
    items = items[:5]
    st.markdown(
        '<div class="section-label" style="margin-bottom:12px;">PM Attention Required</div>',
        unsafe_allow_html=True,
    )
    for style, title, body in items:
        cls = "attention-item" if style == "red" else (
              "attention-item attention-item-amber" if style == "amber"
              else "attention-item attention-item-blue")
        st.markdown(
            f'<div class="{cls}">'
            f'<div style="font-weight:700;color:#0A1628;font-size:13px;margin-bottom:3px;">'
            f'{title}</div>'
            f'<div style="font-size:12px;color:#374151;line-height:1.5;">{body}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
    st.markdown("<br>", unsafe_allow_html=True)


# -----------------------------------------------------------------------------
# HELPER: Empty state card
# -----------------------------------------------------------------------------

def empty_state(icon: str, title: str, body: str, action: str = ""):
    """Premium empty-state card."""
    action_html = (
        f'<div style="margin-top:14px;display:inline-block;background:#071827;'
        f'color:white;border-radius:6px;padding:8px 18px;font-size:12px;font-weight:700;'
        f'letter-spacing:0.5px;border-left:3px solid #F5A623;">'
        f'{action}</div>'
    ) if action else ""
    icon_html = f'<div style="font-size:32px;margin-bottom:10px;">{icon}</div>' if icon else ""
    st.markdown(
        f'<div class="empty-state">'
        f'{icon_html}'
        f'<div style="font-size:16px;font-weight:700;color:#071827;margin-bottom:8px;">'
        f'{title}</div>'
        f'<div style="font-size:13px;color:#6B7280;max-width:400px;margin:0 auto;'
        f'line-height:1.6;">{body}</div>'
        f'{action_html}'
        f'</div>',
        unsafe_allow_html=True,
    )


# -----------------------------------------------------------------------------
# LANDING PAGE  (shown when no XER uploaded)
# -----------------------------------------------------------------------------

def _landing_page():
    """Professional control centre landing page shown when no programme is loaded."""

    # Control bar header
    st.markdown(
        f"""
        <div style="background:linear-gradient(180deg,#071827 0%,#0B2438 100%);
                    margin:-0px -28px 0 -28px;padding:18px 28px 16px 28px;
                    border-bottom:2px solid #F5A623;">
            <div style="display:flex;align-items:flex-end;justify-content:space-between;">
                <div>
                    <div style="font-size:11px;color:#2D4557;font-weight:700;
                                text-transform:uppercase;letter-spacing:2px;margin-bottom:4px;">
                        PlanTrace
                    </div>
                    <div style="font-size:26px;font-weight:800;color:#FFFFFF;
                                letter-spacing:-0.5px;line-height:1;">
                        Control Centre
                    </div>
                    <div style="font-size:13px;color:#4B6275;margin-top:5px;">
                        Planning intelligence for project delivery teams.
                    </div>
                </div>
                <div style="text-align:right;">
                    <div style="font-size:10px;color:#2D4557;font-weight:700;
                                text-transform:uppercase;letter-spacing:1px;margin-bottom:4px;">
                        System Status
                    </div>
                    <div style="display:flex;align-items:center;gap:6px;">
                        <div style="width:6px;height:6px;border-radius:50%;background:#16A34A;"></div>
                        <span style="font-size:11px;color:#4B6275;font-weight:600;">Online</span>
                    </div>
                </div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Status strip
    st.markdown(
        f"""
        <div style="background:#0B2438;margin:-0px -28px 0 -28px;
                    padding:8px 28px;border-bottom:1px solid #102A43;
                    display:flex;align-items:center;gap:0;flex-wrap:wrap;">
            <div style="display:flex;align-items:center;gap:8px;padding-right:20px;
                        border-right:1px solid #102A43;margin-right:20px;">
                <div style="width:5px;height:5px;border-radius:50%;background:#374151;"></div>
                <span style="font-size:9px;color:#2D4557;font-weight:700;
                             text-transform:uppercase;letter-spacing:0.8px;">Programme</span>
                <span style="font-size:11px;color:#3A5060;font-weight:600;margin-left:4px;">Not Loaded</span>
            </div>
            <div style="display:flex;align-items:center;gap:8px;padding-right:20px;
                        border-right:1px solid #102A43;margin-right:20px;">
                <span style="font-size:9px;color:#2D4557;font-weight:700;
                             text-transform:uppercase;letter-spacing:0.8px;">Relationships</span>
                <span style="font-size:11px;color:#3A5060;font-weight:600;margin-left:4px;">Waiting</span>
            </div>
            <div style="display:flex;align-items:center;gap:8px;padding-right:20px;
                        border-right:1px solid #102A43;margin-right:20px;">
                <span style="font-size:9px;color:#2D4557;font-weight:700;
                             text-transform:uppercase;letter-spacing:0.8px;">Resources</span>
                <span style="font-size:11px;color:#3A5060;font-weight:600;margin-left:4px;">Waiting</span>
            </div>
            <div style="display:flex;align-items:center;gap:8px;padding-right:20px;
                        border-right:1px solid #102A43;margin-right:20px;">
                <span style="font-size:9px;color:#2D4557;font-weight:700;
                             text-transform:uppercase;letter-spacing:0.8px;">Notes</span>
                <span style="font-size:11px;color:#3A5060;font-weight:600;margin-left:4px;">Not Loaded</span>
            </div>
            <div style="display:flex;align-items:center;gap:8px;">
                <span style="font-size:9px;color:#2D4557;font-weight:700;
                             text-transform:uppercase;letter-spacing:0.8px;">Comparison</span>
                <span style="font-size:11px;color:#3A5060;font-weight:600;margin-left:4px;">Not Loaded</span>
            </div>
        </div>
        <div style="margin-bottom:24px;"></div>
        """,
        unsafe_allow_html=True,
    )

    # Upload CTA + description
    cta_col, desc_col = st.columns([1, 2], gap="large")
    with cta_col:
        st.markdown(
            f"""
            <div style="background:#071827;border-radius:8px;padding:22px;
                        border:1px solid #0B2438;border-left:3px solid #F5A623;">
                <div style="font-size:10px;color:#F5A623;font-weight:700;
                            text-transform:uppercase;letter-spacing:1.5px;margin-bottom:8px;">
                    Upload Programme
                </div>
                <div style="font-size:13px;color:#8FA3B1;line-height:1.6;margin-bottom:14px;">
                    Upload a Primavera P6 XER file using the <strong style="color:#CBD5E1;">
                    sidebar on the left</strong> to activate the control centre.
                </div>
                <div style="font-size:11px;color:#2D4557;line-height:1.7;">
                    Export from P6:<br>
                    <span style="color:#4B6275;">File &rarr; Export &rarr; Primavera P6 XER</span>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with desc_col:
        st.markdown(
            f"""
            <div style="padding:4px 0;">
                <div style="font-size:14px;color:#111827;font-weight:600;margin-bottom:8px;">
                    Planning intelligence for project delivery teams
                </div>
                <div style="font-size:13px;color:#6B7280;line-height:1.7;margin-bottom:16px;">
                    Upload a Primavera P6 XER programme to interrogate logic, critical paths,
                    labour demand, schedule quality and programme movement — without opening P6.
                </div>
                <div style="display:flex;gap:20px;flex-wrap:wrap;">
                    <div style="text-align:center;">
                        <div style="font-size:20px;font-weight:800;color:#071827;">11</div>
                        <div style="font-size:10px;color:#6B7280;text-transform:uppercase;
                                    letter-spacing:0.8px;font-weight:600;">Health Checks</div>
                    </div>
                    <div style="text-align:center;">
                        <div style="font-size:20px;font-weight:800;color:#071827;">9</div>
                        <div style="font-size:10px;color:#6B7280;text-transform:uppercase;
                                    letter-spacing:0.8px;font-weight:600;">Analysis Modules</div>
                    </div>
                    <div style="text-align:center;">
                        <div style="font-size:20px;font-weight:800;color:#071827;">0</div>
                        <div style="font-size:10px;color:#6B7280;text-transform:uppercase;
                                    letter-spacing:0.8px;font-weight:600;">P6 Licence Required</div>
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.markdown("<div style='margin-bottom:24px;'></div>", unsafe_allow_html=True)

    # Section label
    st.markdown(
        '<div class="section-label">Analysis Modules</div>',
        unsafe_allow_html=True,
    )

    # Module cards - row 1
    c1, c2, c3, c4 = st.columns(4, gap="small")
    modules_row1 = [
        (c1, "Logic Trace",
         "Trace predecessor and successor chains through the programme network.",
         ["Predecessors", "Successors", "Full chain", "Export"]),
        (c2, "Critical Path",
         "Identify critical work, near-critical activities and negative float conditions.",
         ["Critical work", "Near-critical", "Negative float", "Driving path"]),
        (c3, "Labour Demand",
         "Labour histograms by week, month, WBS and resource type.",
         ["Weekly histogram", "Monthly histogram", "WBS view", "Resource view"]),
        (c4, "Programme Health",
         "Eleven automated schedule quality checks against DCMA and industry standards.",
         ["Open logic", "Constraints", "Long durations", "Float issues"]),
    ]
    for col, title, desc, outputs in modules_row1:
        outputs_html = "".join(f"<li>{o}</li>" for o in outputs)
        col.markdown(
            f'<div class="module-card">'
            f'<div class="module-label">Module</div>'
            f'<div class="module-title">{title}</div>'
            f'<div class="module-desc">{desc}</div>'
            f'<div style="font-size:9px;color:#9CA3AF;text-transform:uppercase;'
            f'letter-spacing:1px;font-weight:700;margin-bottom:5px;">Outputs</div>'
            f'<ul class="module-outputs">{outputs_html}</ul>'
            f'</div>',
            unsafe_allow_html=True,
        )

    st.markdown("<div style='margin-bottom:12px;'></div>", unsafe_allow_html=True)

    # Row 2
    st.markdown('<div class="section-label">Additional Modules</div>', unsafe_allow_html=True)
    c1, c2, c3 = st.columns(3, gap="small")
    modules_row2 = [
        (c1, "Programme Comparison",
         "Intelligence on movement between two XER revisions. Identify slippage, gains and changes.",
         ["Date movement", "Slip analysis", "Gained activities", "Risk & opportunity"]),
        (c2, "PM Actions",
         "Auto-generated prioritised action list based on programme analysis and risk indicators.",
         ["Prioritised actions", "Risk items", "Float warnings", "Export"]),
        (c3, "Reports",
         "Export all programme data, analysis results and health checks to formatted Excel workbooks.",
         ["Activity export", "Critical path", "Health check", "Full workbook"]),
    ]
    for col, title, desc, outputs in modules_row2:
        outputs_html = "".join(f"<li>{o}</li>" for o in outputs)
        col.markdown(
            f'<div class="module-card">'
            f'<div class="module-label">Module</div>'
            f'<div class="module-title">{title}</div>'
            f'<div class="module-desc">{desc}</div>'
            f'<div style="font-size:9px;color:#9CA3AF;text-transform:uppercase;'
            f'letter-spacing:1px;font-weight:700;margin-bottom:5px;">Outputs</div>'
            f'<ul class="module-outputs">{outputs_html}</ul>'
            f'</div>',
            unsafe_allow_html=True,
        )

    # Footer
    st.markdown(
        f'<div style="margin-top:32px;padding:16px 0;border-top:1px solid #E5E7EB;">'
        f'<span style="font-size:11px;color:#9CA3AF;">'
        f'<strong style="color:#071827;">PlanTrace</strong>'
        f' &nbsp;&mdash;&nbsp; Planning intelligence for project delivery'
        f' &nbsp;&mdash;&nbsp; No P6 licence required'
        f'</span></div>',
        unsafe_allow_html=True,
    )


# -----------------------------------------------------------------------------
# SIDEBAR
# -----------------------------------------------------------------------------

def _sidebar() -> tuple:
    """
    Render the PlanTrace branded sidebar.
    Returns (active_page_key: str, near_crit_days: int).
    """
    with st.sidebar:

        # -- Brand header ------------------------------------------------------
        logo_html = _logo_tag(width=34)
        st.markdown(
            f"""<div style="padding:11px 12px 9px 12px;border-bottom:1px solid #0c2030;background:#071827;">
<div style="display:flex;align-items:center;gap:8px;">
<div style="flex-shrink:0;width:34px;height:34px;overflow:hidden;border-radius:5px;background:#0C2132;display:flex;align-items:center;justify-content:center;">{logo_html}</div>
<div>
<div style="font-size:14px;font-weight:800;line-height:1;letter-spacing:-0.3px;margin-bottom:2px;"><span style="color:#FFFFFF;">Plan</span><span style="color:#F5A623;">Trace</span></div>
<div style="font-size:8px;color:#1a2f40;text-transform:uppercase;letter-spacing:1.8px;font-weight:700;">Planning Intelligence</div>
</div></div></div>""",
            unsafe_allow_html=True,
        )

        # -- XER Upload --------------------------------------------------------
        st.markdown(
            '<div style="padding:8px 10px 4px 10px;">'
            '<div style="font-size:8px;color:#1a2e3a;font-weight:700;'
            'letter-spacing:1.5px;text-transform:uppercase;margin-bottom:4px;">'
            'Programme File</div>',
            unsafe_allow_html=True,
        )
        xer_file = st.file_uploader(
            "Upload XER", type=["xer"],
            label_visibility="collapsed",
            help="Export from P6: File > Export > Primavera P6 XER",
            key="sidebar_xer_upload",
        )
        st.markdown("</div>", unsafe_allow_html=True)

        # Parse on new upload
        if xer_file is not None:
            ck = f"xer_{xer_file.name}_{xer_file.size}"
            if st.session_state.get("_xer_cache_key") != ck:
                with st.spinner("Parsing programme..."):
                    try:
                        parsed = parse_xer(xer_file.read())
                        st.session_state["programme"]      = parsed
                        st.session_state["_xer_cache_key"] = ck
                        st.session_state["_xer_filename"]  = xer_file.name
                    except Exception as e:
                        st.error(f"Parse error: {e}")
                        st.session_state.pop("programme", None)

        # -- Programme status card ---------------------------------------------
        if "programme" in st.session_state:
            prog   = st.session_state["programme"]
            fname  = st.session_state.get("_xer_filename", "")
            ntasks = len(prog.get("tasks_df", []))
            nrels  = len(prog.get("relationships_df", []))
            ddate  = prog.get("project_info", {}).get("data_date")
            pname  = prog.get("project_info", {}).get("name", "")
            dd_s   = format_date(ddate) if ddate else "N/A"
            pname_line = (
                f'<div style="font-size:10px;color:#4B5563;margin-bottom:3px;'
                f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis;"'
                f' title="{pname}">{pname}</div>'
                if pname else ""
            )
            st.markdown(
                f'<div style="margin:6px 8px;background:#0B2438;'
                f'border:1px solid #102A43;border-radius:4px;padding:8px 10px;">'
                f'<div style="display:flex;align-items:center;gap:4px;margin-bottom:4px;">'
                f'<div style="width:4px;height:4px;border-radius:50%;background:{PT_GREEN};flex-shrink:0;"></div>'
                f'<div style="font-size:8px;font-weight:700;color:{PT_GREEN};letter-spacing:1px;text-transform:uppercase;">Loaded</div>'
                f'</div>'
                f'<div style="font-size:10px;font-weight:600;color:#8fa3b5;margin-bottom:1px;'
                f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis;" title="{fname}">{fname}</div>'
                f'{pname_line}'
                f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:3px;margin-top:5px;">'
                f'<div style="background:#071827;border-radius:3px;padding:3px 7px;">'
                f'<div style="font-size:7px;color:#1e3040;text-transform:uppercase;letter-spacing:0.8px;font-weight:700;">Activities</div>'
                f'<div style="font-size:13px;font-weight:800;color:#F5A623;">{ntasks:,}</div>'
                f'</div>'
                f'<div style="background:#071827;border-radius:3px;padding:3px 7px;">'
                f'<div style="font-size:7px;color:#1e3040;text-transform:uppercase;letter-spacing:0.8px;font-weight:700;">Rels</div>'
                f'<div style="font-size:13px;font-weight:800;color:#F5A623;">{nrels:,}</div>'
                f'</div></div>'
                f'<div style="font-size:9px;color:#1e3040;margin-top:4px;">'
                f'Data date: <span style="color:#304050;">{dd_s}</span></div>'
                f'</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                f'<div style="margin:6px 8px;background:#0B2438;'
                f'border:1px solid #0f2236;border-radius:4px;padding:8px 10px;">'
                f'<div style="display:flex;align-items:center;gap:4px;margin-bottom:3px;">'
                f'<div style="width:4px;height:4px;border-radius:50%;background:#1e3040;flex-shrink:0;"></div>'
                f'<div style="font-size:8px;font-weight:700;color:#1e3040;text-transform:uppercase;letter-spacing:1px;">No Programme</div>'
                f'</div>'
                f'<div style="font-size:10px;color:#182535;line-height:1.4;">'
                f'Upload a .xer file above.</div>'
                f'</div>',
                unsafe_allow_html=True,
            )

        # -- Navigation --------------------------------------------------------
        st.markdown(
            '<div style="margin:8px 0 2px 0;padding:6px 10px 4px 10px;'
            'border-top:1px solid #0a1f33;">'
            '<div style="font-size:8px;color:#182535;font-weight:700;'
            'letter-spacing:1.5px;text-transform:uppercase;">Navigation</div>'
            '</div>',
            unsafe_allow_html=True,
        )

        prog_loaded = "programme" in st.session_state

        if "nav_page" not in st.session_state:
            st.session_state["nav_page"] = "overview"
        current = st.session_state["nav_page"]

        for key, label, icon in _NAV:
            is_active   = (current == key)
            is_disabled = (not prog_loaded) and (key in _NEEDS_PROG)

            if is_active:
                bg_style  = f"background:#0B2438!important;"
                txt_style = f"color:#F5A623!important;"
                border    = f"border-left:3px solid #F5A623!important;"
            elif is_disabled:
                bg_style  = "background:transparent!important;"
                txt_style = "color:#1A2E3A!important;"
                border    = "border-left:3px solid transparent!important;"
            else:
                bg_style  = "background:transparent!important;"
                txt_style = "color:#4B6275!important;"
                border    = "border-left:3px solid transparent!important;"

            opacity = "opacity:0.3;" if is_disabled else ""

            st.markdown(
                f'<div style="margin:1px 0;{bg_style}{border}{opacity}">',
                unsafe_allow_html=True,
            )
            if st.button(f"{label}", key=f"nav_{key}",
                         use_container_width=True, disabled=is_disabled):
                st.session_state["nav_page"] = key
                st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)

        # -- Settings ----------------------------------------------------------
        st.markdown(
            '<div style="margin:4px 0 2px 0;padding:6px 10px 4px 10px;'
            'border-top:1px solid #0a1f33;">'
            '<div style="font-size:8px;color:#182535;font-weight:700;'
            'letter-spacing:1.5px;text-transform:uppercase;">Settings</div>'
            '</div>',
            unsafe_allow_html=True,
        )
        near_crit_days = st.slider(
            "Near-Critical Float (days)",
            min_value=1, max_value=30, value=10, step=1,
        )

        # -- Footer ------------------------------------------------------------
        st.markdown(
            '<div style="padding:8px 10px;border-top:1px solid #0a1f33;margin-top:8px;">'
            '<div style="font-size:8px;color:#182535;line-height:1.6;font-weight:600;">'
            'P6: File &rarr; Export &rarr; P6 XER</div>'
            '<div style="font-size:8px;color:#111e28;margin-top:1px;">v4.0 &nbsp;&mdash;&nbsp; PlanTrace</div>'
            '</div>',
            unsafe_allow_html=True,
        )

    return st.session_state["nav_page"], near_crit_days


# -----------------------------------------------------------------------------
# GUARD
# -----------------------------------------------------------------------------

def _guard() -> bool:
    """Show empty state and return True if no programme loaded."""
    if "programme" not in st.session_state:
        empty_state(
            "",
            "No Programme Loaded",
            "Upload a .xer file using the sidebar on the left to unlock this page. "
            "Export from Primavera P6 via File &rarr; Export &rarr; Primavera P6 XER.",
            "Upload in Sidebar",
        )
        return True
    return False


# -----------------------------------------------------------------------------
# MAIN
# -----------------------------------------------------------------------------

def main():
    active, near_crit_days = _sidebar()
    prog_loaded = "programme" in st.session_state
    data = st.session_state.get("programme", {})

    # -- OVERVIEW ------------------------------------------------------------
    if active == "overview":
        if not prog_loaded:
            _landing_page()
        else:
            page_header(
                "Overview",
                "Programme summary, KPI indicators and top issues requiring your attention.",
                "",
            )
            # Data quality card on the right, KPI row on the left
            tasks = data.get("tasks_df", pd.DataFrame())
            if not tasks.empty:
                tasks_kpi = get_critical_threshold(tasks, near_crit_days)
                n_crit  = int(tasks_kpi["is_critical"].sum()) if "is_critical" in tasks_kpi.columns else 0
                n_nc    = int(tasks_kpi["is_near_critical"].sum()) if "is_near_critical" in tasks_kpi.columns else 0
                n_neg   = int(tasks_kpi["total_float_days"].apply(lambda f: safe_float(f,0) < 0).sum()) if "total_float_days" in tasks_kpi.columns else 0
                n_rels  = len(data.get("relationships_df", pd.DataFrame()))

                kpi_row([
                    ("Total Activities",    len(tasks),       "",                   PT_NAVY,  PT_AMBER),
                    ("Critical",            n_crit,           "float ≤ 0",          PT_RED,   PT_RED),
                    ("Near-Critical",       n_nc,             f"float ≤ {near_crit_days}d", "#92400E", PT_AMBER),
                    ("Negative Float",      n_neg,            "beyond target date", PT_RED,   PT_RED) if n_neg > 0
                        else ("Negative Float", n_neg,        "all on track",       PT_GREEN, PT_GREEN),
                    ("Relationships",       n_rels,           "",                   PT_NAVY,  PT_AMBER),
                ])
                st.markdown("<br>", unsafe_allow_html=True)

            # Attention panel + data quality side by side
            left_col, right_col = st.columns([3, 1], gap="large")
            with left_col:
                pm_attention_panel(data, near_crit_days)
            with right_col:
                data_quality_card(data)

            # Existing overview page content below
            page_project_summary(data, near_crit_days)

    # -- PROGRAMME ------------------------------------------------------------
    elif active == "programme":
        page_header("Programme", "Search activities, plan lookahead, track milestones and review planning notes.", "")
        if _guard():
            return
        p_tabs = st.tabs(["Activity Search", "Lookahead", "Milestones", "Planning Notes"])
        with p_tabs[0]: page_activity_search(data, near_crit_days)
        with p_tabs[1]: page_lookahead(data, near_crit_days)
        with p_tabs[2]: page_milestone_tracker(data, near_crit_days)
        with p_tabs[3]: page_planning_notes(data)

    # -- LOGIC ----------------------------------------------------------------
    elif active == "logic":
        page_header("Logic", "Trace predecessor and successor chains through the programme network.", "")
        if _guard():
            return
        l_tabs = st.tabs(["Logic Trace", "Path to Selected Activity"])
        with l_tabs[0]: page_logic_trace(data, near_crit_days)
        with l_tabs[1]: page_critical_path_to_activity(data, near_crit_days)

    # -- CRITICAL PATH --------------------------------------------------------
    elif active == "critical":
        page_header("Critical Path", "Full critical path, near-critical activities and negative float analysis.", "")
        if _guard():
            return
        page_critical_path(data, near_crit_days)

    # -- LABOUR ---------------------------------------------------------------
    elif active == "labour":
        page_header("Labour", "Labour demand histograms by week, month, resource and WBS.", "")
        if _guard():
            return
        task_res = data.get("task_resources_df", pd.DataFrame())
        if task_res.empty:
            empty_state(
                "",
                "No Resource Data Found",
                "This XER file does not contain resource loading. "
                "Resources must be assigned in P6 and included in the export. "
                "You can also upload a separate CSV file with columns: "
                "task_code, rsrc_name, target_qty, target_start, target_finish.",
            )
        page_labour_histogram(data)

    # -- HEALTH CHECK ---------------------------------------------------------
    elif active == "health":
        page_header("Health Check", "Automated schedule quality checks covering logic, float, constraints and durations.", "")
        if _guard():
            return
        page_health_check(data, near_crit_days)

    # -- COMPARISON -----------------------------------------------------------
    elif active == "comparison":
        page_header("Comparison", "Programme movement intelligence between two XER revisions.", "")
        c_tabs = st.tabs(["Programme Movement", "Risk & Opportunity"])
        with c_tabs[0]:
            page_programme_comparison()
        with c_tabs[1]:
            if _guard():
                pass
            else:
                page_risk_register(data, near_crit_days)

    # -- PM ACTIONS -----------------------------------------------------------
    elif active == "pm_actions":
        page_header("PM Actions", "Auto-generated prioritised action list based on programme analysis.", "")
        if _guard():
            return
        page_pm_actions(data, near_crit_days)

    # -- RISK REGISTER --------------------------------------------------------
    elif active == "risk":
        page_header("Risk Register", "Auto-generated risk and opportunity register from programme data.", "")
        if _guard():
            return
        page_risk_register(data, near_crit_days)

    # -- REPORTS --------------------------------------------------------------
    elif active == "reports":
        page_header("Reports", "Export all programme data and analysis to formatted Excel workbooks.", "")
        if _guard():
            return
        page_export_reports(data, near_crit_days)

    # -- SETTINGS -------------------------------------------------------------
    elif active == "settings":
        page_header("Settings", "App configuration and loaded programme details.", "")
        st.markdown("### Configuration")
        st.info(
            f"Near-critical float threshold: **{near_crit_days} days**. "
            "Adjust using the slider in the sidebar."
        )
        if prog_loaded:
            st.markdown("### Loaded Programme")
            proj = data.get("project_info", {})
            st.json({
                "name":          proj.get("name", ""),
                "data_date":     format_date(proj.get("data_date")),
                "activities":    len(data.get("tasks_df", [])),
                "relationships": len(data.get("relationships_df", [])),
                "parse_method":  data.get("parse_method", ""),
            })
            st.markdown("### Data Quality")
            data_quality_card(data)


if __name__ == "__main__":
    main()
